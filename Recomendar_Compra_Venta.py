# Importaciones livianas primero (para mostrar interfaz rápido)
import tkinter as tk
from tkinter import filedialog, ttk, messagebox, simpledialog
import os
import sys
import gc
import json
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
import threading
import numpy as np

# =====================================================
# DETECCIÓN DE MODO EJECUTABLE vs SCRIPT
# =====================================================
def es_ejecutable():
    """Detecta si el script corre como ejecutable (.exe) compilado con PyInstaller"""
    return getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS')

def obtener_ruta_base():
    """Obtiene la ruta base del ejecutable o del script"""
    if es_ejecutable():
        return Path(sys.executable).parent
    else:
        return Path(__file__).parent

def obtener_carpeta_datos():
    """Obtiene la carpeta de datos (data/) - SIEMPRE portable"""
    ruta_base = obtener_ruta_base()
    # Buscar data/ en el directorio actual
    carpeta_data = ruta_base / "data"
    if carpeta_data.exists():
        return carpeta_data
    # Si no existe, buscar en el directorio padre (estructura compartida)
    carpeta_data_padre = ruta_base.parent / "data"
    if carpeta_data_padre.exists():
        return carpeta_data_padre
    # Si no existe en ningun lugar, crear en el directorio actual
    carpeta_data.mkdir(parents=True, exist_ok=True)
    return carpeta_data

# Variables globales para modo ejecutable
MODO_EJECUTABLE = es_ejecutable()
CARPETA_DATOS_PORTABLE = obtener_carpeta_datos()

# Variables globales para bibliotecas cargadas en segundo plano
yf = None
pd = None
plt = None
FigureCanvasTkAgg = None
mdates = None

# Estado de carga
carga_completa = False
progreso_carga = 0
libs_cargadas = {"yfinance": False, "pandas": False, "matplotlib": False}

def cargar_bibliotecas_async(root, label_progreso):
    """Carga las bibliotecas pesadas en segundo plano"""
    global yf, pd, plt, FigureCanvasTkAgg, mdates, carga_completa, progreso_carga, libs_cargadas

    def actualizar_progreso(texto, porcentaje):
        global progreso_carga
        progreso_carga = porcentaje
        if label_progreso.winfo_exists():
            label_progreso.config(text=f"{texto} ({porcentaje}%)")
            root.update_idletasks()

    try:
        # Cargar pandas (40%)
        actualizar_progreso("Cargando pandas...", 20)
        import pandas
        pd = pandas
        libs_cargadas["pandas"] = True
        actualizar_progreso("pandas cargado", 40)

        # Cargar yfinance (60%)
        actualizar_progreso("Cargando yfinance...", 50)
        import yfinance
        yf = yfinance
        libs_cargadas["yfinance"] = True
        actualizar_progreso("yfinance cargado", 60)

        # Cargar matplotlib (100%)
        actualizar_progreso("Cargando matplotlib...", 70)
        import matplotlib.pyplot
        import matplotlib.dates
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg as FCA
        plt = matplotlib.pyplot
        mdates = matplotlib.dates
        FigureCanvasTkAgg = FCA
        libs_cargadas["matplotlib"] = True
        actualizar_progreso("Listo", 100)

        carga_completa = True
        if label_progreso.winfo_exists():
            label_progreso.config(text="")
    except Exception as e:
        if label_progreso.winfo_exists():
            label_progreso.config(text=f"Error: {e}")

def verificar_libs_cargadas(libs_requeridas):
    """Verifica si las bibliotecas requeridas están cargadas"""
    for lib in libs_requeridas:
        if not libs_cargadas.get(lib, False):
            return False
    return True

def requiere_libs(libs_requeridas):
    """Decorador para verificar si las bibliotecas están cargadas antes de ejecutar"""
    def decorador(func):
        def wrapper(*args, **kwargs):
            if not verificar_libs_cargadas(libs_requeridas):
                messagebox.showwarning("Esperar", "Esperar que se carguen los recursos del sistema.")
                return
            return func(*args, **kwargs)
        return wrapper
    return decorador

# Lista de tickers por defecto
TICKERS_DEFAULT = ["AAPL","AMZN","AVGO","BRK-B","GLD","META","MSFT","NVDA","PLTR","QQQ","SPY","TSLA"]

# Configuracion PORTABLE - siempre usa carpeta data/ relativa al script/exe
TICKERS_CONFIG_FILE = CARPETA_DATOS_PORTABLE / "tickers_descarga.json"


def cargar_tickers_config():
    """Carga la configuracion de tickers por plataforma.
    Si no existe o tiene formato antiguo, migra automaticamente."""
    if TICKERS_CONFIG_FILE.exists():
        try:
            with open(TICKERS_CONFIG_FILE, 'r', encoding='utf-8') as f:
                datos = json.load(f)

                # Verificar si es formato nuevo (con plataformas)
                if "plataformas" in datos:
                    return datos

                # Formato antiguo: migrar automaticamente
                tickers_antiguos = datos.get("tickers", TICKERS_DEFAULT.copy())
                datos_nuevos = {
                    "plataformas": {
                        "TYBA": {
                            "tickers": tickers_antiguos,
                            "mercado": "NYSE",
                            "moneda": "USD"
                        }
                    }
                }
                # Guardar formato nuevo
                guardar_tickers_config(datos_nuevos)
                return datos_nuevos

        except Exception as e:
            print(f"[WARN] Error cargando tickers config: {e}")

    # Retornar estructura por defecto
    return {
        "plataformas": {
            "TYBA": {
                "tickers": TICKERS_DEFAULT.copy(),
                "mercado": "NYSE",
                "moneda": "USD"
            }
        }
    }


def obtener_tickers_plataforma(plataforma, modo=None):
    """Retorna la lista de tickers para una plataforma y modo especificos.

    Args:
        plataforma: Nombre de la plataforma (ej: TYBA, IBKR-UK)
        modo: Modo de operacion (Paper/Real). Si es None, retorna todos los tickers de la plataforma.
    """
    config = cargar_tickers_config()
    plat_info = config.get("plataformas", {}).get(plataforma, {})

    # Nueva estructura con modos
    if "modos" in plat_info:
        if modo:
            return plat_info.get("modos", {}).get(modo, {}).get("tickers", [])
        else:
            # Sin modo especificado: retornar todos los tickers de todos los modos
            todos = set()
            for modo_info in plat_info.get("modos", {}).values():
                todos.update(modo_info.get("tickers", []))
            return sorted(list(todos))
    else:
        # Estructura antigua (compatibilidad)
        return plat_info.get("tickers", [])


def obtener_tickers_unicos():
    """Retorna un set de todos los tickers unicos de todas las plataformas y modos."""
    config = cargar_tickers_config()
    tickers_unicos = set()
    for plat_info in config.get("plataformas", {}).values():
        if "modos" in plat_info:
            for modo_info in plat_info.get("modos", {}).values():
                tickers_unicos.update(modo_info.get("tickers", []))
        else:
            tickers_unicos.update(plat_info.get("tickers", []))
    return sorted(list(tickers_unicos))


def obtener_plataformas():
    """Retorna lista de nombres de plataformas disponibles."""
    config = cargar_tickers_config()
    return list(config.get("plataformas", {}).keys())


def guardar_tickers_config(datos_config):
    """Guarda la configuracion de tickers por plataforma y sincroniza con GitHub.

    Args:
        datos_config: dict con estructura {"plataformas": {...}}
    """
    import subprocess
    try:
        # Asegurar que la carpeta existe
        TICKERS_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(TICKERS_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(datos_config, f, indent=2, ensure_ascii=False)

        # Intentar sincronizar con GitHub (si es un repo git)
        repo_path = str(obtener_ruta_base())
        try:
            # Verificar si es un repositorio git
            check_git = subprocess.run(
                ["git", "rev-parse", "--is-inside-work-tree"],
                cwd=repo_path, capture_output=True, text=True, timeout=10
            )
            if check_git.returncode == 0:
                # Es un repo git, hacer commit y push
                archivo_rel = "data/tickers_descarga.json"
                subprocess.run(["git", "add", archivo_rel], cwd=repo_path, capture_output=True, timeout=10)
                subprocess.run(
                    ["git", "commit", "-m", "Actualizar configuracion de tickers/plataformas"],
                    cwd=repo_path, capture_output=True, timeout=10
                )
                resultado_push = subprocess.run(
                    ["git", "push", "origin", "main"],
                    cwd=repo_path, capture_output=True, text=True, timeout=30
                )
                if resultado_push.returncode == 0:
                    print("[INFO] Tickers sincronizados con GitHub")
                else:
                    print(f"[WARN] No se pudo hacer push: {resultado_push.stderr}")
        except Exception as e:
            print(f"[WARN] No se pudo sincronizar con GitHub: {e}")

        return True
    except Exception as e:
        print(f"[ERROR] Error guardando tickers config: {e}")
        return False


def agregar_ticker_plataforma(plataforma, ticker, modo="Real"):
    """Agrega un ticker a una plataforma y modo especificos."""
    config = cargar_tickers_config()
    if plataforma not in config.get("plataformas", {}):
        return False, f"Plataforma '{plataforma}' no existe"

    plat_info = config["plataformas"][plataforma]

    # Nueva estructura con modos
    if "modos" in plat_info:
        if modo not in plat_info["modos"]:
            plat_info["modos"][modo] = {"tickers": []}
        tickers = plat_info["modos"][modo].get("tickers", [])
        if ticker in tickers:
            return False, f"Ticker '{ticker}' ya existe en {plataforma} ({modo})"
        tickers.append(ticker)
        tickers.sort()
        plat_info["modos"][modo]["tickers"] = tickers
    else:
        # Estructura antigua
        tickers = plat_info.get("tickers", [])
        if ticker in tickers:
            return False, f"Ticker '{ticker}' ya existe en {plataforma}"
        tickers.append(ticker)
        tickers.sort()
        plat_info["tickers"] = tickers

    guardar_tickers_config(config)
    return True, f"Ticker '{ticker}' agregado a {plataforma} ({modo})"


def quitar_ticker_plataforma(plataforma, ticker, modo="Real"):
    """Quita un ticker de una plataforma y modo especificos."""
    config = cargar_tickers_config()
    if plataforma not in config.get("plataformas", {}):
        return False, f"Plataforma '{plataforma}' no existe"

    plat_info = config["plataformas"][plataforma]

    # Nueva estructura con modos
    if "modos" in plat_info:
        if modo not in plat_info.get("modos", {}):
            return False, f"Modo '{modo}' no existe en {plataforma}"
        tickers = plat_info["modos"][modo].get("tickers", [])
        if ticker not in tickers:
            return False, f"Ticker '{ticker}' no existe en {plataforma} ({modo})"
        tickers.remove(ticker)
        plat_info["modos"][modo]["tickers"] = tickers
    else:
        # Estructura antigua
        tickers = plat_info.get("tickers", [])
        if ticker not in tickers:
            return False, f"Ticker '{ticker}' no existe en {plataforma}"
        tickers.remove(ticker)
        plat_info["tickers"] = tickers

    guardar_tickers_config(config)
    return True, f"Ticker '{ticker}' eliminado de {plataforma} ({modo})"


# ============================================================================
# FUNCIONES PARA LISTA GLOBAL DE TICKERS
# ============================================================================

def obtener_tickers_globales():
    """Retorna la lista global de tickers (con parámetros calculados)."""
    config = cargar_tickers_config()
    return config.get("tickers_globales", [])


def agregar_ticker_global(ticker):
    """Agrega un ticker a la lista global.

    Args:
        ticker: Símbolo del ticker a agregar

    Returns:
        tuple: (exito: bool, mensaje: str)
    """
    config = cargar_tickers_config()
    tickers_globales = config.get("tickers_globales", [])

    if ticker in tickers_globales:
        return False, f"Ticker '{ticker}' ya existe en la lista global"

    tickers_globales.append(ticker)
    tickers_globales.sort()
    config["tickers_globales"] = tickers_globales

    guardar_tickers_config(config)
    return True, f"Ticker '{ticker}' agregado a la lista global"


def quitar_ticker_global(ticker):
    """Quita un ticker de la lista global (solo si no está en ninguna plataforma).

    La información del ticker (parámetros, histórico) se mantiene.
    Solo se puede quitar si no está asignado a ninguna plataforma/modo.

    Args:
        ticker: Símbolo del ticker a quitar

    Returns:
        tuple: (exito: bool, mensaje: str)
    """
    config = cargar_tickers_config()
    tickers_globales = config.get("tickers_globales", [])

    if ticker not in tickers_globales:
        return False, f"Ticker '{ticker}' no existe en la lista global"

    # Verificar si está en alguna plataforma/modo
    plataformas_con_ticker = []
    for plat_nombre, plat_info in config.get("plataformas", {}).items():
        if "modos" in plat_info:
            for modo, modo_info in plat_info["modos"].items():
                tickers_modo = modo_info.get("tickers", [])
                if ticker in tickers_modo:
                    plataformas_con_ticker.append(f"{plat_nombre} ({modo})")

    # Si está en alguna plataforma, no permitir quitar
    if plataformas_con_ticker:
        return False, f"No se puede quitar '{ticker}'. Está en: {', '.join(plataformas_con_ticker)}"

    # Quitar solo de la lista global (la información/parámetros se mantiene)
    tickers_globales.remove(ticker)
    config["tickers_globales"] = tickers_globales
    guardar_tickers_config(config)

    return True, f"Ticker '{ticker}' quitado de la lista global (información conservada)"


def ticker_existe_en_global(ticker):
    """Verifica si un ticker existe en la lista global."""
    return ticker in obtener_tickers_globales()


def agregar_plataforma_tickers(nombre, mercado="NYSE", moneda="USD"):
    """Agrega una nueva plataforma a la configuracion de tickers."""
    config = cargar_tickers_config()
    if nombre in config.get("plataformas", {}):
        return False, f"Plataforma '{nombre}' ya existe"

    config["plataformas"][nombre] = {
        "mercado": mercado,
        "moneda": moneda,
        "modos": {
            "Real": {"tickers": []},
            "Paper": {"tickers": []}
        }
    }
    guardar_tickers_config(config)

    # Tambien agregar a historial_operaciones.json para mantener sincronizacion
    try:
        hist_data = cargar_historial_operaciones_completo()
        if nombre not in hist_data.get("config_plataformas", {}):
            hist_data["config_plataformas"][nombre] = {
                "moneda": moneda,
                "descripcion": f"{nombre} - Inversiones {moneda}"
            }
            guardar_historial_operaciones(
                hist_data.get("operaciones", []),
                hist_data.get("config_plataformas")
            )
    except Exception as e:
        print(f"[WARN] No se pudo sincronizar plataforma con historial: {e}")

    return True, f"Plataforma '{nombre}' creada"


def eliminar_plataforma_tickers(nombre):
    """Elimina una plataforma de la configuracion de tickers."""
    config = cargar_tickers_config()
    if nombre not in config.get("plataformas", {}):
        return False, f"Plataforma '{nombre}' no existe"

    if len(config["plataformas"]) <= 1:
        return False, "No se puede eliminar la ultima plataforma"

    del config["plataformas"][nombre]
    guardar_tickers_config(config)
    return True, f"Plataforma '{nombre}' eliminada"


# Cargar tickers desde archivo (todos los unicos de todas las plataformas)
tickers_config = cargar_tickers_config()  # Estructura completa con plataformas
tickers = obtener_tickers_unicos()  # Lista de tickers unicos para descarga

# Configuracion PORTABLE
CONFIG_FILE = None  # No se usa archivo de configuracion externo
UBICACION_JSON_PORTABLE = CARPETA_DATOS_PORTABLE
DATOS_CSV_PORTABLE = CARPETA_DATOS_PORTABLE / "datos_1dia_crudos.csv"  # Archivo principal
AUTO_UPDATE_LOG_PORTABLE = CARPETA_DATOS_PORTABLE / "auto_update_log.csv"  # Log historico
BACKUPS_FOLDER = CARPETA_DATOS_PORTABLE / "backups"  # Carpeta de respaldos


def crear_backup_datos(motivo="manual"):
    """
    Crea un backup de todos los archivos críticos en data/backups/
    Se debe llamar ANTES de cualquier operación que modifique datos.

    Args:
        motivo: Descripción del motivo del backup (ej: "antes_sync", "antes_actualizar")

    Returns:
        str: Ruta de la carpeta de backup creada, o None si falla
    """
    import shutil

    try:
        # Crear carpeta de backups si no existe
        BACKUPS_FOLDER.mkdir(parents=True, exist_ok=True)

        # Crear subcarpeta con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_folder = BACKUPS_FOLDER / f"{timestamp}_{motivo}"
        backup_folder.mkdir(parents=True, exist_ok=True)

        # Lista de archivos críticos a respaldar
        archivos_criticos = [
            "auto_update_log.csv",
            "datos_1dia_crudos.csv",
            "parametros_activos.json",
            "historial_senales.json",
            "Resultado_de_Analisis.json",
            "tickers_descarga.json",
            "historial_operaciones.json"
        ]

        archivos_respaldados = 0
        for archivo in archivos_criticos:
            origen = CARPETA_DATOS_PORTABLE / archivo
            if origen.exists():
                destino = backup_folder / archivo
                shutil.copy2(origen, destino)
                archivos_respaldados += 1
                print(f"[Backup] {archivo} -> {backup_folder.name}/")

        print(f"[Backup] Completado: {archivos_respaldados} archivos en {backup_folder}")

        # Limpiar backups antiguos (mantener solo los últimos 10)
        limpiar_backups_antiguos(max_backups=10)

        return str(backup_folder)

    except Exception as e:
        print(f"[Backup] ERROR: {e}")
        return None


def limpiar_backups_antiguos(max_backups=10):
    """Elimina backups antiguos, manteniendo solo los más recientes"""
    import shutil

    try:
        if not BACKUPS_FOLDER.exists():
            return

        # Listar carpetas de backup ordenadas por fecha (más antiguas primero)
        backups = sorted([d for d in BACKUPS_FOLDER.iterdir() if d.is_dir()])

        # Eliminar los más antiguos si hay más del límite
        while len(backups) > max_backups:
            backup_antiguo = backups.pop(0)
            shutil.rmtree(backup_antiguo)
            print(f"[Backup] Eliminado backup antiguo: {backup_antiguo.name}")

    except Exception as e:
        print(f"[Backup] Error limpiando backups antiguos: {e}")


def restaurar_backup(backup_folder):
    """
    Restaura archivos desde una carpeta de backup.

    Args:
        backup_folder: Ruta a la carpeta de backup a restaurar
    """
    import shutil

    backup_path = Path(backup_folder)
    if not backup_path.exists():
        print(f"[Backup] ERROR: No existe la carpeta {backup_folder}")
        return False

    try:
        for archivo in backup_path.iterdir():
            if archivo.is_file():
                destino = CARPETA_DATOS_PORTABLE / archivo.name
                shutil.copy2(archivo, destino)
                print(f"[Backup] Restaurado: {archivo.name}")

        print(f"[Backup] Restauración completada desde {backup_path.name}")
        return True

    except Exception as e:
        print(f"[Backup] ERROR en restauración: {e}")
        return False


def siguiente_dia_trading(fecha, retornar_feriados=False):
    """
    Calcula el siguiente día de trading después de la fecha dada.
    Salta fines de semana y feriados principales de USA.

    Args:
        fecha: datetime o date object
        retornar_feriados: Si True, retorna también lista de feriados saltados

    Returns:
        datetime.date del siguiente día de trading
        Si retornar_feriados=True: (datetime.date, list de strings con feriados)
    """
    from datetime import timedelta

    # Convertir a date si es datetime
    if hasattr(fecha, 'date'):
        fecha = fecha.date()

    # Feriados principales de USA 2025-2026 (mercado cerrado) con nombres
    feriados_usa = {
        # 2025
        datetime(2025, 1, 1).date(): "New Year's Day",
        datetime(2025, 1, 20).date(): "MLK Day",
        datetime(2025, 2, 17).date(): "Presidents Day",
        datetime(2025, 4, 18).date(): "Good Friday",
        datetime(2025, 5, 26).date(): "Memorial Day",
        datetime(2025, 6, 19).date(): "Juneteenth",
        datetime(2025, 7, 4).date(): "Independence Day",
        datetime(2025, 9, 1).date(): "Labor Day",
        datetime(2025, 11, 27).date(): "Thanksgiving",
        datetime(2025, 12, 25).date(): "Christmas",
        # 2026
        datetime(2026, 1, 1).date(): "New Year's Day",
        datetime(2026, 1, 19).date(): "MLK Day",
        datetime(2026, 2, 16).date(): "Presidents Day",
        datetime(2026, 4, 3).date(): "Good Friday",
        datetime(2026, 5, 25).date(): "Memorial Day",
        datetime(2026, 6, 19).date(): "Juneteenth",
        datetime(2026, 7, 3).date(): "Independence Day (observed)",
        datetime(2026, 9, 7).date(): "Labor Day",
        datetime(2026, 11, 26).date(): "Thanksgiving",
        datetime(2026, 12, 25).date(): "Christmas",
    }

    siguiente = fecha + timedelta(days=1)
    feriados_saltados = []

    # Avanzar hasta encontrar un día de trading válido
    while siguiente.weekday() >= 5 or siguiente in feriados_usa:  # 5=sábado, 6=domingo
        # Si es feriado (no fin de semana), registrarlo
        if siguiente.weekday() < 5 and siguiente in feriados_usa:
            feriados_saltados.append(f"{siguiente.strftime('%d-%m-%Y')} {feriados_usa[siguiente]}")
        siguiente += timedelta(days=1)

    if retornar_feriados:
        return siguiente, feriados_saltados
    return siguiente


def crear_estructura_slots_vacia():
    """Crea una estructura de slots vacía con valores por defecto"""
    return {
        "version": "2.0",
        "slots": {
            "1": {"nombre": "1", "parametros_activos": []},
            "2": {"nombre": "2", "parametros_activos": []},
            "3": {"nombre": "3", "parametros_activos": []},
            "4": {"nombre": "4", "parametros_activos": []},
            "5": {"nombre": "5", "parametros_activos": []}
        }
    }


def migrar_parametros_v1_a_v2(datos_v1):
    """Migra formato antiguo (lista) a nuevo formato (slots)"""
    parametros_lista = datos_v1.get("parametros_activos", [])
    estructura = crear_estructura_slots_vacia()
    estructura["slots"]["1"]["parametros_activos"] = parametros_lista
    return estructura


def obtener_parametros_slot(datos_slots, slot_id):
    """Obtiene la lista de parámetros de un slot específico"""
    return datos_slots.get("slots", {}).get(slot_id, {}).get("parametros_activos", [])


def obtener_nombre_slot(datos_slots, slot_id):
    """Obtiene el nombre de un slot"""
    return datos_slots.get("slots", {}).get(slot_id, {}).get("nombre", slot_id)


def cargar_parametros_activos():
    """Carga los parametros activos desde carpeta data (portable).
    Retorna: (datos_slots, error) - datos_slots es la estructura completa con todos los slots"""
    archivo_params = UBICACION_JSON_PORTABLE / "parametros_activos.json"
    if not archivo_params.exists():
        return None, f"No existe el archivo:\n{archivo_params}\n\nCopia los archivos de datos a la carpeta 'data'."
    try:
        with open(archivo_params, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Detectar versión del formato
        if "version" in datos and datos.get("version") == "2.0":
            # Ya es formato nuevo con slots
            # Verificar si hay al menos un slot con parámetros
            tiene_parametros = any(
                obtener_parametros_slot(datos, s) for s in ["1", "2", "3", "4", "5"]
            )
            if not tiene_parametros:
                return None, "No hay parametros activos configurados en ningún slot"
            return datos, None
        else:
            # Formato antiguo - migrar a v2
            datos_migrados = migrar_parametros_v1_a_v2(datos)
            # Guardar en formato nuevo
            with open(archivo_params, 'w', encoding='utf-8') as f:
                json.dump(datos_migrados, f, indent=2, ensure_ascii=False)
            if not datos_migrados["slots"]["1"]["parametros_activos"]:
                return None, "No hay parametros activos configurados"
            return datos_migrados, None
    except Exception as e:
        return None, f"Error cargando parametros: {e}"


def obtener_ruta_historial():
    """Obtiene la ruta del archivo de historial de operaciones (portable)"""
    return UBICACION_JSON_PORTABLE / "historial_operaciones.json"


def cargar_historial_operaciones():
    """Carga el historial de operaciones confirmadas (solo lista de operaciones)"""
    ruta = obtener_ruta_historial()
    if ruta is None or not ruta.exists():
        return []

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)
            return datos.get("operaciones", [])
    except Exception as e:
        print(f"[ERROR] Error cargando historial: {e}")
        return []


def cargar_historial_operaciones_completo():
    """Carga el historial completo incluyendo config_plataformas"""
    ruta = obtener_ruta_historial()
    if ruta is None or not ruta.exists():
        return {"operaciones": [], "config_plataformas": {"TYBA": {"moneda": "USD", "descripcion": "Tyba - Inversiones USD"}}}

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)
            # Asegurar que existe config_plataformas
            if "config_plataformas" not in datos:
                datos["config_plataformas"] = {"TYBA": {"moneda": "USD", "descripcion": "Tyba - Inversiones USD"}}
            return datos
    except Exception as e:
        print(f"[ERROR] Error cargando historial: {e}")
        return {"operaciones": [], "config_plataformas": {"TYBA": {"moneda": "USD", "descripcion": "Tyba - Inversiones USD"}}}


def agregar_operacion_sin_duplicado(operacion):
    """
    Agrega una operación al historial solo si no existe un duplicado.
    Duplicado = misma fecha, ticker, tipo, precio, cantidad, plataforma, modo.

    Returns:
        bool: True si se agregó, False si ya existía
    """
    datos = cargar_historial_operaciones_completo()
    operaciones = datos.get("operaciones", [])

    # Clave única para detectar duplicados
    ticker = operacion.get('ticker_symbol') or operacion.get('symbol', '')
    clave_nueva = (
        operacion.get('fecha', ''),
        ticker,
        operacion.get('tipo', '').lower(),
        operacion.get('precio', 0),
        operacion.get('cantidad', 0),
        operacion.get('plataforma', ''),
        operacion.get('modo', '')
    )

    # Verificar si ya existe
    for op in operaciones:
        ticker_existente = op.get('ticker_symbol') or op.get('symbol', '')
        clave_existente = (
            op.get('fecha', ''),
            ticker_existente,
            op.get('tipo', '').lower(),
            op.get('precio', 0),
            op.get('cantidad', 0),
            op.get('plataforma', ''),
            op.get('modo', '')
        )
        if clave_nueva == clave_existente:
            print(f"[INFO] Operación duplicada ignorada: {ticker} {operacion.get('tipo')} {operacion.get('fecha')}")
            return False

    # Agregar y guardar
    operaciones.append(operacion)
    guardar_historial_operaciones(operaciones, datos.get("config_plataformas"))
    return True


def guardar_historial_operaciones(operaciones, config_plataformas=None):
    """Guarda el historial de operaciones preservando config_plataformas"""
    ruta = obtener_ruta_historial()
    if ruta is None:
        messagebox.showerror("Error", "No hay ubicación configurada para guardar el historial.")
        return False

    try:
        # Cargar config existente si no se proporciona
        if config_plataformas is None:
            datos_existentes = cargar_historial_operaciones_completo()
            config_plataformas = datos_existentes.get("config_plataformas", {})

        datos = {"config_plataformas": config_plataformas, "operaciones": operaciones}
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Error guardando historial:\n{e}")
        return False


def calcular_posiciones_ibkr(modo):
    """Calcula las posiciones actuales de IBKR-UK desde el historial de operaciones."""
    try:
        datos = cargar_historial_operaciones_completo()
        operaciones = datos.get("operaciones", [])

        # Filtrar por plataforma IBKR-UK y modo
        ops_filtradas = [
            op for op in operaciones
            if op.get("plataforma") == "IBKR-UK"
            and op.get("modo", "Real").lower() == modo.lower()
        ]

        # Calcular posiciones
        posiciones = {}
        for op in ops_filtradas:
            ticker = op.get("ticker_symbol") or op.get("symbol", "")
            if not ticker:
                continue

            tipo = op.get("tipo", "").lower()
            cantidad = op.get("cantidad", 0)

            if ticker not in posiciones:
                posiciones[ticker] = 0

            if tipo == "compra":
                posiciones[ticker] += cantidad
            elif tipo == "venta":
                posiciones[ticker] -= cantidad

        # Filtrar solo posiciones > 0
        return {k: v for k, v in posiciones.items() if v > 0}

    except Exception as e:
        print(f"[WARN] Error calculando posiciones IBKR: {e}")
        return {}


def guardar_sync_ibkr(modo, capital, posiciones, fecha_sync=None, balances_por_moneda=None):
    """Guarda los datos de sincronización de IBKR (Paper o Live) en historial_operaciones.json.
    Fuente única de datos para sync IBKR."""
    from datetime import datetime

    fecha_actual = fecha_sync or datetime.now().strftime("%Y-%m-%d %H:%M")

    # Guardar en historial_operaciones.json (fuente única)
    datos = cargar_historial_operaciones_completo()
    config = datos.get("config_plataformas", {})

    # Asegurar que existe IBKR-UK en config
    if "IBKR-UK" not in config:
        config["IBKR-UK"] = {"moneda": "USD", "descripcion": "Interactive Brokers UK"}

    # Clave según modo
    clave = f"ultimo_sync_{modo.lower()}"

    sync_data = {
        "fecha": fecha_actual,
        "capital": capital,
        "posiciones": posiciones  # dict con detalle {ticker: cantidad}
    }

    # Agregar datos de monedas si están disponibles (nuevo formato con cash y stocks)
    if balances_por_moneda:
        if isinstance(balances_por_moneda, dict) and "cash" in balances_por_moneda:
            sync_data["balances_por_moneda"] = balances_por_moneda.get("cash", {})
            sync_data["stocks_por_moneda"] = balances_por_moneda.get("stocks", {})
        else:
            # Compatibilidad con formato antiguo
            sync_data["balances_por_moneda"] = balances_por_moneda

    config["IBKR-UK"][clave] = sync_data

    # Guardar historial
    guardar_historial_operaciones(datos.get("operaciones", []), config)
    print(f"[Sync] IBKR-UK {modo} guardado en historial_operaciones.json")


def subir_estado_ibkr_a_github(modo):
    """Sube el estado de IBKR a GitHub automáticamente después de sincronizar."""
    import subprocess

    repo_path = str(obtener_ruta_base())
    sync_file = "data/estado_ibkr_sync.json"

    try:
        # Verificar si es un repositorio git
        check_git = subprocess.run(
            ["git", "rev-parse", "--is-inside-work-tree"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=10
        )
        if check_git.returncode != 0:
            print("[Sync GitHub] No es un repositorio git, omitiendo push")
            return False

        # git add
        result = subprocess.run(
            ["git", "add", sync_file],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode != 0:
            print(f"[Sync GitHub] Error en git add: {result.stderr}")
            return False

        # Verificar si hay cambios para commitear
        status = subprocess.run(
            ["git", "status", "--porcelain", sync_file],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=10
        )
        if not status.stdout.strip():
            print("[Sync GitHub] Sin cambios nuevos para subir")
            return True

        # git commit
        from datetime import datetime
        fecha_commit = datetime.now().strftime("%Y-%m-%d %H:%M")
        mensaje = f"Sync IBKR-UK {modo} - {fecha_commit}"

        result = subprocess.run(
            ["git", "commit", "-m", mensaje],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode != 0 and "nothing to commit" not in result.stdout:
            print(f"[Sync GitHub] Error en git commit: {result.stderr}")
            return False

        # git push
        print("[Sync GitHub] Subiendo estado a GitHub...")
        result = subprocess.run(
            ["git", "push"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=60
        )
        if result.returncode != 0:
            print(f"[Sync GitHub] Error en git push: {result.stderr}")
            # Intentar con --set-upstream si es necesario
            if "no upstream branch" in result.stderr:
                result = subprocess.run(
                    ["git", "push", "--set-upstream", "origin", "main"],
                    cwd=repo_path,
                    capture_output=True,
                    text=True,
                    timeout=60
                )
                if result.returncode != 0:
                    print(f"[Sync GitHub] Error en push con upstream: {result.stderr}")
                    return False

        print(f"[Sync GitHub] ✓ Estado IBKR-UK {modo} subido a GitHub")
        return True

    except subprocess.TimeoutExpired:
        print("[Sync GitHub] Timeout - no se pudo subir a GitHub")
        return False
    except Exception as e:
        print(f"[Sync GitHub] Error: {e}")
        return False


def cargar_sync_ibkr(modo):
    """Carga los datos de última sincronización de IBKR (Paper o Real)"""
    return cargar_sync_plataforma("IBKR-UK", modo)


def cargar_sync_plataforma(plataforma, modo):
    """Carga los datos de última sincronización de cualquier plataforma.

    Args:
        plataforma: Nombre de la plataforma (IBKR-UK, TYBA, etc.)
        modo: "Paper" o "Real"

    Returns:
        dict con {fecha, capital, posiciones, balances_por_moneda} o None
    """
    datos = cargar_historial_operaciones_completo()
    config = datos.get("config_plataformas", {})

    plat_config = config.get(plataforma, {})
    clave = f"ultimo_sync_{modo.lower()}"

    # Buscar con clave actual, si no existe buscar con "live" por compatibilidad
    result = plat_config.get(clave, None)
    if result is None and modo.lower() == "real":
        result = plat_config.get("ultimo_sync_live", None)

    return result


def guardar_sync_plataforma(plataforma, modo, capital, posiciones, fecha_sync=None, balances_por_moneda=None):
    """Guarda los datos de sincronización de cualquier plataforma.

    Args:
        plataforma: Nombre de la plataforma (IBKR-UK, TYBA, etc.)
        modo: "Paper" o "Real"
        capital: String con el capital (ej: "$10,000.00")
        posiciones: dict {ticker: cantidad} o número de posiciones
        fecha_sync: Fecha/hora de sincronización (opcional)
        balances_por_moneda: dict {moneda: valor} o {"cash": {...}, "stocks": {...}}
    """
    from datetime import datetime

    fecha_actual = fecha_sync or datetime.now().strftime("%Y-%m-%d %H:%M")

    datos = cargar_historial_operaciones_completo()
    config = datos.get("config_plataformas", {})

    # Asegurar que existe la plataforma en config
    if plataforma not in config:
        config[plataforma] = {"moneda": "USD", "descripcion": plataforma}

    # Clave según modo
    clave = f"ultimo_sync_{modo.lower()}"

    sync_data = {
        "fecha": fecha_actual,
        "capital": capital,
        "posiciones": posiciones
    }

    # Agregar datos de monedas si están disponibles
    if balances_por_moneda:
        if isinstance(balances_por_moneda, dict) and "cash" in balances_por_moneda:
            sync_data["balances_por_moneda"] = balances_por_moneda.get("cash", {})
            sync_data["stocks_por_moneda"] = balances_por_moneda.get("stocks", {})
        else:
            sync_data["balances_por_moneda"] = balances_por_moneda

    config[plataforma][clave] = sync_data

    guardar_historial_operaciones(datos.get("operaciones", []), config)
    print(f"[Sync] {plataforma} {modo} guardado en historial_operaciones.json")


# ============================================================================
# FUNCIONES PARA TRANSFERENCIAS (DEPÓSITOS/RETIROS)
# ============================================================================

MONEDAS_DISPONIBLES = ["USD", "EUR", "GBP", "JPY", "CHF", "PEN"]


def guardar_transferencia(plataforma, monto, moneda, fecha, descripcion=""):
    """Guarda una transferencia (depósito/retiro) para una plataforma en modo Real.

    Args:
        plataforma: Nombre de la plataforma (ej: "IBKR-UK", "TYBA")
        monto: Cantidad (positivo=depósito, negativo=retiro)
        moneda: Código de moneda (USD, EUR, GBP, etc.)
        fecha: Fecha de la transferencia (YYYY-MM-DD)
        descripcion: Descripción opcional

    Returns:
        tuple: (exito: bool, mensaje: str)
    """
    from datetime import datetime

    try:
        datos = cargar_historial_operaciones_completo()
        config = datos.get("config_plataformas", {})

        # Asegurar que existe la plataforma en config
        if plataforma not in config:
            config[plataforma] = {"moneda": "USD", "descripcion": plataforma}

        # Asegurar que existe la lista de transferencias
        if "transferencias" not in config[plataforma]:
            config[plataforma]["transferencias"] = []

        # Crear registro de transferencia
        transferencia = {
            "fecha": fecha,
            "monto": float(monto),
            "moneda": moneda,
            "tipo": "deposito" if float(monto) >= 0 else "retiro",
            "descripcion": descripcion,
            "fecha_registro": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        config[plataforma]["transferencias"].append(transferencia)

        # Ordenar por fecha
        config[plataforma]["transferencias"].sort(key=lambda x: x["fecha"])

        # Guardar
        guardar_historial_operaciones(datos.get("operaciones", []), config)

        tipo_texto = "Depósito" if float(monto) >= 0 else "Retiro"
        return True, f"{tipo_texto} de {moneda} {abs(float(monto)):,.2f} registrado"

    except Exception as e:
        return False, f"Error guardando transferencia: {e}"


def cargar_transferencias(plataforma):
    """Carga las transferencias de una plataforma.

    Returns:
        list: Lista de transferencias [{fecha, monto, moneda, tipo, descripcion}]
    """
    datos = cargar_historial_operaciones_completo()
    config = datos.get("config_plataformas", {})
    plat_config = config.get(plataforma, {})
    return plat_config.get("transferencias", [])


def calcular_capital_invertido(plataforma, moneda_filtro=None):
    """Calcula el capital invertido (depósitos - retiros) para una plataforma.

    Args:
        plataforma: Nombre de la plataforma
        moneda_filtro: Si se especifica, solo suma esa moneda

    Returns:
        dict: {moneda: total} o float si se especifica moneda_filtro
    """
    transferencias = cargar_transferencias(plataforma)

    if moneda_filtro:
        total = sum(t["monto"] for t in transferencias if t["moneda"] == moneda_filtro)
        return total

    # Agrupar por moneda
    totales = {}
    for t in transferencias:
        moneda = t["moneda"]
        if moneda not in totales:
            totales[moneda] = 0
        totales[moneda] += t["monto"]

    return totales


def eliminar_transferencia(plataforma, indice):
    """Elimina una transferencia por índice.

    Args:
        plataforma: Nombre de la plataforma
        indice: Índice de la transferencia a eliminar

    Returns:
        tuple: (exito: bool, mensaje: str)
    """
    try:
        datos = cargar_historial_operaciones_completo()
        config = datos.get("config_plataformas", {})

        if plataforma not in config:
            return False, f"Plataforma '{plataforma}' no existe"

        transferencias = config[plataforma].get("transferencias", [])

        if indice < 0 or indice >= len(transferencias):
            return False, "Índice inválido"

        eliminada = transferencias.pop(indice)
        guardar_historial_operaciones(datos.get("operaciones", []), config)

        return True, f"Transferencia del {eliminada['fecha']} eliminada"

    except Exception as e:
        return False, f"Error eliminando transferencia: {e}"


def obtener_ruta_senales():
    """Obtiene la ruta del archivo de historial de senales (portable)"""
    return UBICACION_JSON_PORTABLE / "historial_senales.json"


def guardar_ruta_csv(ruta_csv):
    """En modo portable no se guarda ruta CSV externa"""
    pass  # Siempre usa auto_update_log.csv en data/


def cargar_ruta_csv():
    """Retorna la ruta del CSV principal en carpeta data (portable)"""
    return str(DATOS_CSV_PORTABLE)


def sincronizar_desde_github():
    """
    Sincroniza datos desde GitHub siguiendo el flujo normal:
    0. BACKUP automático antes de cualquier cambio
    1. Descarga datos de GitHub
    2. Filtra solo los que NO existen en auto_update_log.csv local
    3. Guarda los nuevos en datos_1dia_crudos.csv
    4. Muestra en tabla
    5. Merge a auto_update_log.csv
    """
    import subprocess
    import io

    # *** BACKUP AUTOMÁTICO ANTES DE SINCRONIZAR ***
    backup_path = crear_backup_datos("antes_sync_github")
    if backup_path:
        print(f"[Sync] Backup creado: {backup_path}")

    # Rutas (portable)
    repo_path = str(obtener_ruta_base())
    log_file = str(AUTO_UPDATE_LOG_PORTABLE)
    csv_file = str(DATOS_CSV_PORTABLE)

    try:
        # 0. Verificar si es un repositorio git
        check_git = subprocess.run(
            ["git", "rev-parse", "--is-inside-work-tree"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=10
        )
        if check_git.returncode != 0:
            messagebox.showinfo(
                "Sync GitHub",
                "Esta funcion no esta disponible en la version portable.\n\n"
                "Para sincronizar datos desde GitHub, usa la carpeta TRADING\n"
                "que contiene el repositorio git."
            )
            return False

        # 1. Leer datos locales del log histórico
        df_local = None
        local_keys = set()
        if os.path.exists(log_file):
            df_local = pd.read_csv(log_file, parse_dates=['Date'])
            df_local = df_local.loc[:, ~df_local.columns.duplicated()]
            df_local['Date'] = pd.to_datetime(df_local['Date']).dt.normalize()
            local_keys = set(zip(
                df_local['Date'].dt.strftime('%Y-%m-%d'),
                df_local['Ticker']
            ))
            print(f"[Sync] Datos locales en log: {len(df_local)} registros")

        # 2. Hacer git fetch para actualizar referencias
        print("[Sync] Conectando a GitHub...")
        result = subprocess.run(
            ["git", "fetch", "origin", "main"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0:
            messagebox.showerror("Error", f"Error en fetch:\n{result.stderr}")
            return False

        # 3. Obtener el archivo desde GitHub
        result = subprocess.run(
            ["git", "show", "origin/main:data/auto_update_log.csv"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0 or not result.stdout.strip():
            messagebox.showerror("Error", f"No se pudo obtener datos de GitHub:\n{result.stderr}")
            return False

        df_github = pd.read_csv(io.StringIO(result.stdout), parse_dates=['Date'])
        df_github = df_github.loc[:, ~df_github.columns.duplicated()]
        df_github['Date'] = pd.to_datetime(df_github['Date']).dt.normalize()
        print(f"[Sync] Datos en GitHub: {len(df_github)} registros")

        # 4. Filtrar solo registros que NO existen localmente
        github_keys = df_github[['Date', 'Ticker']].apply(
            lambda r: (r['Date'].strftime('%Y-%m-%d'), r['Ticker']), axis=1
        )
        mask_nuevos = ~github_keys.isin(local_keys)
        df_nuevos = df_github.loc[mask_nuevos].copy()

        if df_nuevos.empty:
            # Aunque no hay datos nuevos, mostrar el último día disponible en tabla
            if df_local is not None and not df_local.empty:
                ultima_fecha = df_local['Date'].max()
                df_ultimo_dia = df_local[df_local['Date'] == ultima_fecha].copy()
                # Guardar en csv temporal para mostrar
                temp_csv = str(DATOS_CSV_PORTABLE)
                df_ultimo_dia.to_csv(temp_csv, index=False, float_format="%.2f")
                mostrar_datos_en_tabla(temp_csv)

                # Verificar tickers faltantes aunque no haya datos nuevos
                tickers_configurados = obtener_tickers_unicos()
                tickers_en_csv = set(df_local['Ticker'].unique())
                tickers_faltantes = [t for t in tickers_configurados if t not in tickers_en_csv]

                if tickers_faltantes:
                    messagebox.showinfo("Sincronización",
                        f"Ya tienes los datos más recientes.\n\n"
                        f"Última fecha: {ultima_fecha.strftime('%Y-%m-%d')}\n"
                        f"Registros: {len(df_ultimo_dia)}\n\n"
                        f"⚠️ Tickers sin precios: {', '.join(tickers_faltantes)}")

                    respuesta = messagebox.askyesno("Tickers sin precios",
                        f"¿Desea descargar precios históricos para:\n{', '.join(tickers_faltantes)}?")

                    if respuesta:
                        import yfinance as yf
                        from datetime import datetime, timedelta

                        fecha_inicio = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')
                        fecha_fin = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
                        registros_agregados = 0
                        df_combined = df_local.copy()

                        for ticker in tickers_faltantes:
                            try:
                                print(f"[Sync] Descargando {ticker}...")
                                df_ticker = yf.download(ticker, start=fecha_inicio, end=fecha_fin, progress=False)
                                if df_ticker.empty:
                                    continue
                                if isinstance(df_ticker.columns, pd.MultiIndex):
                                    df_ticker.columns = df_ticker.columns.get_level_values(0)
                                df_ticker = df_ticker.reset_index()
                                df_ticker['Ticker'] = ticker
                                df_ticker = df_ticker[['Date', 'Ticker', 'Open', 'High', 'Low', 'Close']]
                                df_ticker['Date'] = pd.to_datetime(df_ticker['Date']).dt.normalize()
                                df_combined = pd.concat([df_combined, df_ticker], ignore_index=True)
                                registros_agregados += len(df_ticker)
                            except Exception as e:
                                print(f"[Sync] Error descargando {ticker}: {e}")

                        if registros_agregados > 0:
                            df_combined = df_combined.sort_values(['Date', 'Ticker']).reset_index(drop=True)
                            df_combined.to_csv(log_file, index=False, float_format="%.2f")
                            messagebox.showinfo("Descarga completada",
                                f"Se agregaron {registros_agregados} registros de precios.")
                else:
                    messagebox.showinfo("Sincronización",
                        f"Ya tienes los datos más recientes.\n\n"
                        f"Última fecha: {ultima_fecha.strftime('%Y-%m-%d')}\n"
                        f"Registros: {len(df_ultimo_dia)}")
            else:
                messagebox.showinfo("Sincronización", "Ya tienes los datos más recientes.")
            return True

        print(f"[Sync] Registros nuevos encontrados: {len(df_nuevos)}")

        # 5. Guardar nuevos en datos_1dia_crudos.csv (como el flujo manual)
        df_nuevos.to_csv(csv_file, index=False, float_format="%.2f")
        print(f"[Sync] Guardado en {csv_file}")

        # 6. Mostrar en tabla (no cambiar entry_ruta)
        mostrar_datos_en_tabla(csv_file)

        # 7. Merge a auto_update_log.csv (flujo normal)
        if df_local is not None:
            df_combined = pd.concat([df_local, df_nuevos], ignore_index=True)
            df_combined = df_combined.sort_values(['Date', 'Ticker']).reset_index(drop=True)
        else:
            df_combined = df_nuevos

        df_combined.to_csv(log_file, index=False, float_format="%.2f")
        print(f"[Sync] Log actualizado: {len(df_combined)} registros totales")

        # 8. Verificar si hay valores NaN en los datos
        df_con_nan = df_combined[df_combined['Close'].isna()]
        if not df_con_nan.empty:
            # Construir lista de tickers/fechas con NaN
            nan_info = []
            nan_tickers_fechas = []  # Lista de (ticker, fecha) para re-descarga
            for _, row in df_con_nan.iterrows():
                fecha_str = row['Date'].strftime('%Y-%m-%d') if pd.notna(row['Date']) else 'N/A'
                nan_info.append(f"  - {row['Ticker']} ({fecha_str})")
                if pd.notna(row['Date']):
                    nan_tickers_fechas.append((row['Ticker'], row['Date']))
            nan_list = "\n".join(nan_info[:10])  # Mostrar máximo 10
            if len(nan_info) > 10:
                nan_list += f"\n  ... y {len(nan_info) - 10} más"

            # Preguntar si quiere re-descargar automáticamente
            respuesta = messagebox.askyesno("Datos incompletos",
                f"Se detectaron valores vacíos (NaN) en los siguientes tickers:\n\n"
                f"{nan_list}\n\n"
                f"¿Desea re-descargar automáticamente los datos faltantes desde Yahoo Finance?")

            if respuesta and nan_tickers_fechas:
                # Re-descargar datos faltantes
                import yfinance as yf
                corregidos = 0
                errores = []

                for ticker, fecha in nan_tickers_fechas:
                    try:
                        fecha_inicio = fecha.strftime('%Y-%m-%d')
                        fecha_fin = (fecha + pd.Timedelta(days=1)).strftime('%Y-%m-%d')

                        yf_ticker = yf.Ticker(ticker)
                        hist = yf_ticker.history(start=fecha_inicio, end=fecha_fin)

                        if not hist.empty:
                            # Actualizar valores en df_combined
                            mask = (df_combined['Date'] == fecha) & (df_combined['Ticker'] == ticker)
                            df_combined.loc[mask, 'Open'] = hist.iloc[0]['Open']
                            df_combined.loc[mask, 'High'] = hist.iloc[0]['High']
                            df_combined.loc[mask, 'Low'] = hist.iloc[0]['Low']
                            df_combined.loc[mask, 'Close'] = hist.iloc[0]['Close']
                            corregidos += 1
                            print(f"[Sync] Corregido: {ticker} ({fecha_inicio}) - Close: {hist.iloc[0]['Close']:.2f}")
                        else:
                            errores.append(f"{ticker} ({fecha_inicio})")
                    except Exception as e:
                        errores.append(f"{ticker} ({fecha_inicio}): {str(e)[:30]}")

                # Guardar CSV actualizado
                df_combined.to_csv(log_file, index=False, float_format="%.2f")

                # Mostrar resultado
                if corregidos > 0 and not errores:
                    messagebox.showinfo("Re-descarga completada",
                        f"Se corrigieron {corregidos} registros exitosamente.\n\n"
                        f"Total en log: {len(df_combined)}")
                elif corregidos > 0 and errores:
                    messagebox.showwarning("Re-descarga parcial",
                        f"Corregidos: {corregidos}\n"
                        f"Errores: {len(errores)}\n\n"
                        f"No se pudieron obtener:\n" + "\n".join(errores[:5]))
                else:
                    messagebox.showerror("Error en re-descarga",
                        f"No se pudieron obtener los datos.\n\n"
                        f"Prueba con 'Descargar Precios' manualmente.")
            elif not respuesta:
                messagebox.showinfo("Sincronización",
                    f"Sincronización completada con advertencias.\n\n"
                    f"Registros nuevos: {len(df_nuevos)}\n"
                    f"Registros con NaN: {len(nan_info)}\n\n"
                    f"Usa 'Descargar Precios' para corregir manualmente.")
        else:
            messagebox.showinfo("Sincronización",
                f"Sincronización completada.\n\n"
                f"Registros nuevos: {len(df_nuevos)}\n"
                f"Total en log: {len(df_combined)}")

        # 9. Verificar si hay tickers configurados que no tienen precios en el CSV
        tickers_configurados = obtener_tickers_unicos()
        tickers_en_csv = set(df_combined['Ticker'].unique()) if not df_combined.empty else set()
        tickers_faltantes = [t for t in tickers_configurados if t not in tickers_en_csv]

        if tickers_faltantes:
            respuesta = messagebox.askyesno("Tickers sin precios",
                f"Los siguientes tickers están configurados pero no tienen precios en el CSV:\n\n"
                f"{', '.join(tickers_faltantes)}\n\n"
                f"¿Desea descargar sus precios históricos ahora?")

            if respuesta:
                import yfinance as yf
                from datetime import datetime, timedelta

                # Descargar últimos 60 días de cada ticker faltante
                fecha_inicio = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')
                fecha_fin = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

                registros_agregados = 0
                errores_descarga = []

                for ticker in tickers_faltantes:
                    try:
                        print(f"[Sync] Descargando {ticker}...")
                        df_ticker = yf.download(ticker, start=fecha_inicio, end=fecha_fin, progress=False)

                        if df_ticker.empty:
                            errores_descarga.append(ticker)
                            continue

                        # Manejar MultiIndex de columnas
                        if isinstance(df_ticker.columns, pd.MultiIndex):
                            df_ticker.columns = df_ticker.columns.get_level_values(0)

                        df_ticker = df_ticker.reset_index()
                        df_ticker['Ticker'] = ticker
                        df_ticker = df_ticker[['Date', 'Ticker', 'Open', 'High', 'Low', 'Close']]
                        df_ticker['Date'] = pd.to_datetime(df_ticker['Date']).dt.normalize()

                        # Agregar al CSV
                        df_combined = pd.concat([df_combined, df_ticker], ignore_index=True)
                        registros_agregados += len(df_ticker)
                        print(f"[Sync] {ticker}: {len(df_ticker)} registros agregados")

                    except Exception as e:
                        errores_descarga.append(f"{ticker}: {str(e)[:30]}")

                # Guardar CSV actualizado
                df_combined = df_combined.sort_values(['Date', 'Ticker']).reset_index(drop=True)
                df_combined.to_csv(log_file, index=False, float_format="%.2f")

                if registros_agregados > 0:
                    msg = f"Se agregaron {registros_agregados} registros de precios."
                    if errores_descarga:
                        msg += f"\n\nNo se pudieron descargar:\n{', '.join(errores_descarga)}"
                    messagebox.showinfo("Descarga completada", msg)
                elif errores_descarga:
                    messagebox.showerror("Error", f"No se pudieron descargar:\n{', '.join(errores_descarga)}")

        return True

    except subprocess.TimeoutExpired:
        messagebox.showerror("Error", "Timeout: La sincronización tardó demasiado.")
        return False
    except FileNotFoundError:
        messagebox.showerror("Error", "Git no está instalado o no se encuentra en el PATH.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error inesperado:\n{e}")
        return False


def crear_estructura_senales_vacia():
    """Crea una estructura de señales vacía con slots"""
    return {
        "version": "2.0",
        "senales_por_slot": {
            "1": [],
            "2": [],
            "3": [],
            "4": [],
            "5": [],
            "6": []
        }
    }


def cargar_historial_senales():
    """Carga el historial de señales generadas (estructura con slots v2.0)"""
    ruta = obtener_ruta_senales()
    if ruta is None or not ruta.exists():
        return crear_estructura_senales_vacia()

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Detectar versión del formato
        if "version" in datos and datos.get("version") == "2.0":
            return datos
        else:
            # Formato antiguo - migrar a v2
            senales_antiguas = datos.get("senales", [])
            estructura = crear_estructura_senales_vacia()
            estructura["senales_por_slot"]["1"] = senales_antiguas
            # Guardar migrado
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(estructura, f, indent=2, ensure_ascii=False)
            return estructura
    except Exception as e:
        print(f"[ERROR] Error cargando historial de señales: {e}")
        return crear_estructura_senales_vacia()


def cargar_senales_slot(slot_id):
    """Carga las señales de un slot específico"""
    datos = cargar_historial_senales()
    return datos.get("senales_por_slot", {}).get(slot_id, [])


def guardar_historial_senales(senales_nuevas, slot_id="1", slot_nombre="1", fecha_override=None, plataforma=None, modo=None, fecha_cierre_usado=None):
    """Guarda las senales generadas en el historial para un slot especifico (evita duplicados por fecha y simbolo)

    Args:
        senales_nuevas: Lista de senales a guardar
        slot_id: ID del slot
        slot_nombre: Nombre del slot
        fecha_override: Fecha opcional para senales historicas (formato YYYY-MM-DD HH:MM:SS)
        plataforma: Plataforma de inversion (ej: TYBA, IBKR-UK)
        modo: Modo de operacion (Paper/Real)
        fecha_cierre_usado: Fecha del precio de cierre usado para calculos (del CSV)
    """
    ruta = obtener_ruta_senales()
    if ruta is None:
        print("[WARN] No hay ubicación configurada para guardar señales.")
        return False

    try:
        # Cargar estructura completa de señales
        datos_senales = cargar_historial_senales()

        # Obtener señales existentes del slot
        senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])

        # Usar fecha override si se proporciona, sino usar ahora
        if fecha_override:
            fecha_generacion = fecha_override
        else:
            fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        fecha_hoy = fecha_generacion[:10]  # Solo la fecha (YYYY-MM-DD)

        # Para señales históricas, eliminar señales existentes de esa fecha, plataforma y modo en este slot
        if fecha_override:
            plat_actual = plataforma or 'TYBA'
            modo_actual = (modo or 'real').lower()
            senales_slot = [sen for sen in senales_slot
                          if not (sen.get("fecha_generacion", "")[:10] == fecha_hoy
                                  and sen.get("plataforma", "TYBA") == plat_actual
                                  and sen.get("modo", "real").lower() == modo_actual)]

        # Crear conjunto de señales existentes para verificar duplicados (fecha + symbol + plataforma + modo)
        senales_existentes_keys = set()
        for sen in senales_slot:
            fecha_sen = sen.get("fecha_generacion", "")[:10]
            symbol_sen = sen.get("symbol", "")
            plat_sen = sen.get("plataforma", "TYBA")
            modo_sen = sen.get("modo", "real").lower()
            senales_existentes_keys.add((fecha_sen, symbol_sen, plat_sen, modo_sen))

        # Contador de señales nuevas agregadas
        senales_agregadas = 0

        for senal in senales_nuevas:
            if senal.get('estado') == 'OK':
                symbol = senal.get('symbol')
                plat_senal = plataforma or senal.get('plataforma', 'TYBA')
                modo_senal = (modo or senal.get('modo', 'real')).lower()

                # Verificar si ya existe una señal para esta fecha, símbolo, plataforma y modo en este slot
                if (fecha_hoy, symbol, plat_senal, modo_senal) in senales_existentes_keys:
                    print(f"[INFO] Señal duplicada ignorada: {symbol} ({fecha_hoy}) {plat_senal}/{modo_senal} en slot {slot_id}")
                    continue

                # Calcular las 3 fechas de referencia
                fecha_analisis_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                fecha_cierre_str = fecha_cierre_usado.strftime("%Y-%m-%d") if fecha_cierre_usado else None

                nueva_senal = {
                    "fecha_generacion": fecha_generacion,
                    "fecha_senal": fecha_hoy,
                    "symbol": symbol,
                    "plataforma": plataforma or senal.get('plataforma', 'TYBA'),
                    "modo": modo or senal.get('modo', 'Paper'),
                    "precio_cierre": senal.get('cierre'),
                    "precio_compra_sugerido": senal.get('precio_compra'),
                    "cant_compra": senal.get('cant_compra'),
                    "opc_compra": senal.get('opc_compra'),
                    "precio_venta_sugerido": senal.get('precio_venta'),
                    "cant_venta": senal.get('cant_venta'),
                    "opc_venta": senal.get('opc_venta'),
                    "acciones_cartera": senal.get('acciones_cartera'),
                    "precio_compra_minimo": senal.get('precio_compra_minimo', 0),
                    "ganancia_min_pct": senal.get('ganancia_min_pct', 0),
                    "limite_tipo": senal.get('limite_tipo', 'acciones'),
                    "limite_valor": senal.get('limite_valor', 10),
                    "slot_id": slot_id,
                    "slot_nombre": slot_nombre,
                    "tendencia": senal.get('tendencia', 'N/A'),
                    "tendencia_larga": senal.get('tendencia_larga', 'N/A'),
                    "slot_origen_compra": senal.get('slot_origen_compra', ''),
                    "slot_origen_venta": senal.get('slot_origen_venta', ''),
                    # 3 fechas de referencia (v3.5.0)
                    "fecha_cierre_usado": fecha_cierre_str,  # Fecha del precio de cierre del CSV
                    "fecha_analisis": fecha_analisis_actual,  # Cuando se ejecutó generar_senales
                    "fecha_trading": fecha_hoy  # Día de trading para el cual aplican
                }
                senales_slot.append(nueva_senal)
                senales_existentes_keys.add((fecha_hoy, symbol))
                senales_agregadas += 1

        # Actualizar slot en la estructura
        datos_senales["senales_por_slot"][slot_id] = senales_slot

        # Guardar todo
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(datos_senales, f, indent=2, ensure_ascii=False)

        print(f"[INFO] Slot {slot_id}: {senales_agregadas} señales nuevas guardadas")
        return True

    except Exception as e:
        print(f"[ERROR] Error guardando señales: {e}")
        return False


def calcular_cartera(operaciones_param=None, plataforma=None, modo=None):
    """Calcula el estado actual de la cartera basandose en el historial de operaciones.

    Args:
        operaciones_param: Lista de operaciones (si None, carga del archivo)
        plataforma: Si se especifica, filtra operaciones por esta plataforma
        modo: Si se especifica, filtra operaciones por este modo (paper/real)
    """
    operaciones = operaciones_param if operaciones_param is not None else cargar_historial_operaciones()

    # Filtrar por plataforma si se especifica
    if plataforma:
        operaciones = [op for op in operaciones if op.get("plataforma", "TYBA") == plataforma]

    # Filtrar por modo si se especifica
    if modo:
        modo_lower = modo.lower()
        def get_modo_op(op):
            if "modo" in op:
                return op["modo"].lower()
            # Default: TYBA=real, resto=paper
            return "real" if op.get("plataforma", "TYBA") == "TYBA" else "paper"
        operaciones = [op for op in operaciones if get_modo_op(op) == modo_lower]

    cartera = {}

    # Primero, construir lista de compras por ticker para rastrear precios individuales
    compras_por_ticker = {}

    for op in operaciones:
        symbol = op.get("ticker_symbol")
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)

        if symbol not in cartera:
            cartera[symbol] = {
                "acciones": 0,
                "total_comprado": 0,
                "total_vendido": 0,
                "precio_promedio_compra": 0,
                "capital_invertido": 0,
                "precio_compra_minimo": 0
            }
            compras_por_ticker[symbol] = []  # Lista de (precio, cantidad_restante)

        if tipo == "compra":
            # Actualizar precio promedio de compra
            total_acciones_previas = cartera[symbol]["acciones"]
            capital_previo = cartera[symbol]["capital_invertido"]
            nuevo_capital = capital_previo + (precio * cantidad)
            nuevas_acciones = total_acciones_previas + cantidad

            cartera[symbol]["acciones"] = nuevas_acciones
            cartera[symbol]["total_comprado"] += cantidad
            cartera[symbol]["capital_invertido"] = nuevo_capital
            if nuevas_acciones > 0:
                cartera[symbol]["precio_promedio_compra"] = nuevo_capital / nuevas_acciones

            # Agregar compra a la lista (ordenada por precio ascendente para FIFO por precio más bajo)
            compras_por_ticker[symbol].append([precio, cantidad])
            compras_por_ticker[symbol].sort(key=lambda x: x[0])  # Ordenar por precio

        elif tipo == "venta":
            cartera[symbol]["acciones"] -= cantidad
            cartera[symbol]["total_vendido"] += cantidad

            # Descontar de las compras usando FIFO (primero las de precio más bajo)
            # y calcular el capital invertido que se libera
            cantidad_a_descontar = cantidad
            capital_liberado = 0
            for compra in compras_por_ticker[symbol]:
                if cantidad_a_descontar <= 0:
                    break
                if compra[1] > 0:
                    descontar = min(compra[1], cantidad_a_descontar)
                    capital_liberado += descontar * compra[0]  # Costo real de las acciones vendidas
                    compra[1] -= descontar
                    cantidad_a_descontar -= descontar

            # Reducir capital invertido por el costo real de las acciones vendidas
            cartera[symbol]["capital_invertido"] -= capital_liberado

            # Recalcular precio promedio de compra con las acciones restantes
            if cartera[symbol]["acciones"] > 0:
                cartera[symbol]["precio_promedio_compra"] = cartera[symbol]["capital_invertido"] / cartera[symbol]["acciones"]
            else:
                cartera[symbol]["precio_promedio_compra"] = 0
                cartera[symbol]["capital_invertido"] = 0  # Asegurar que sea 0 si no hay acciones

            # Limpiar compras agotadas
            compras_por_ticker[symbol] = [c for c in compras_por_ticker[symbol] if c[1] > 0]

    # Calcular precio de compra mínimo y guardar lista FIFO para cada ticker
    for symbol in cartera:
        if compras_por_ticker.get(symbol) and cartera[symbol]["acciones"] > 0:
            # El precio mínimo es el primero de la lista ordenada
            cartera[symbol]["precio_compra_minimo"] = compras_por_ticker[symbol][0][0]
            # Guardar lista completa de precios FIFO: [(precio, cantidad), ...]
            cartera[symbol]["precios_fifo"] = [(c[0], c[1]) for c in compras_por_ticker[symbol] if c[1] > 0]
        else:
            cartera[symbol]["precio_compra_minimo"] = 0
            cartera[symbol]["precios_fifo"] = []

    return cartera


def calcular_cartera_historica(fecha_limite, plataforma=None, modo=None):
    """
    Calcula el estado de la cartera hasta una fecha especifica.
    Util para regenerar senales historicas con la cartera que existia en esa fecha.

    Args:
        fecha_limite: Fecha limite (str YYYY-MM-DD o date). Las operaciones de esta fecha
                      en adelante NO se incluyen.
        plataforma: Si se especifica, filtra operaciones por esta plataforma.
        modo: Si se especifica, filtra operaciones por este modo (paper/real).

    Returns:
        dict: Cartera con acciones y precio_compra_minimo por ticker
    """
    operaciones = cargar_historial_operaciones()

    # Filtrar por plataforma si se especifica
    if plataforma:
        operaciones = [op for op in operaciones if op.get("plataforma", "TYBA") == plataforma]

    # Filtrar por modo si se especifica
    if modo:
        modo_lower = modo.lower()
        def get_modo_op(op):
            if "modo" in op:
                return op["modo"].lower()
            return "real" if op.get("plataforma", "TYBA") == "TYBA" else "paper"
        operaciones = [op for op in operaciones if get_modo_op(op) == modo_lower]

    cartera = {}
    compras_por_ticker = {}

    # Convertir fecha_limite a string si es necesario
    if hasattr(fecha_limite, 'strftime'):
        fecha_limite_str = fecha_limite.strftime("%Y-%m-%d")
    else:
        fecha_limite_str = str(fecha_limite)

    for op in operaciones:
        fecha_op = op.get("fecha", "")
        # Solo procesar operaciones ANTERIORES a la fecha límite
        if fecha_op >= fecha_limite_str:
            continue

        symbol = op.get("ticker_symbol")
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)

        if symbol not in cartera:
            cartera[symbol] = {
                "acciones": 0,
                "capital_invertido": 0,
                "precio_compra_minimo": 0
            }
            compras_por_ticker[symbol] = []

        if tipo == "compra":
            cartera[symbol]["acciones"] += cantidad
            cartera[symbol]["capital_invertido"] += precio * cantidad
            compras_por_ticker[symbol].append([precio, cantidad])
            compras_por_ticker[symbol].sort(key=lambda x: x[0])

        elif tipo == "venta":
            cartera[symbol]["acciones"] -= cantidad
            # Descontar de las compras más baratas primero (FIFO por precio)
            restante = cantidad
            nuevas_compras = []
            for compra in compras_por_ticker[symbol]:
                if restante <= 0:
                    nuevas_compras.append(compra)
                elif compra[1] <= restante:
                    restante -= compra[1]
                else:
                    compra[1] -= restante
                    restante = 0
                    nuevas_compras.append(compra)
            compras_por_ticker[symbol] = nuevas_compras

    # Calcular precio_compra_minimo y guardar lista FIFO para cada ticker
    for symbol in cartera:
        if compras_por_ticker.get(symbol) and cartera[symbol]["acciones"] > 0:
            cartera[symbol]["precio_compra_minimo"] = compras_por_ticker[symbol][0][0]
            # Guardar lista completa de precios FIFO: [(precio, cantidad), ...]
            cartera[symbol]["precios_fifo"] = [(c[0], c[1]) for c in compras_por_ticker[symbol] if c[1] > 0]
        else:
            cartera[symbol]["precio_compra_minimo"] = 0
            cartera[symbol]["precios_fifo"] = []

    return cartera


def calcular_cant_venta_valida_fifo(precios_fifo, precio_venta, cant_deseada, ganancia_min_pct):
    """
    Calcula la cantidad máxima de acciones que se pueden vender cumpliendo:
    1. Ninguna acción se vende a pérdida (precio_venta > precio_compra de cada acción)
    2. La ganancia total % sobre el costo FIFO >= ganancia_min_pct

    Args:
        precios_fifo: Lista de tuplas [(precio_compra, cantidad), ...] ordenada por precio ascendente
        precio_venta: Precio de venta sugerido
        cant_deseada: Cantidad de acciones que se desea vender
        ganancia_min_pct: Porcentaje mínimo de ganancia requerido

    Returns:
        tuple: (cantidad_valida, motivo)
            - cantidad_valida: Número de acciones que se pueden vender (0 si ninguna cumple)
            - motivo: "OK", "ESPERAR (pérdida individual)" o "ESPERAR (ganancia insuficiente)"
    """
    if not precios_fifo or cant_deseada <= 0 or precio_venta <= 0:
        return (0, "N/A")

    # Expandir lista FIFO a acciones individuales (para procesar una por una)
    acciones_fifo = []
    for precio, cantidad in precios_fifo:
        acciones_fifo.extend([precio] * int(cantidad))

    if not acciones_fifo:
        return (0, "N/A")

    # Limitar a la cantidad deseada o disponible
    max_acciones = min(cant_deseada, len(acciones_fifo))

    # Probar desde la cantidad máxima hacia abajo
    for cant in range(max_acciones, 0, -1):
        # Tomar las primeras 'cant' acciones (FIFO)
        acciones_a_vender = acciones_fifo[:cant]

        # Verificar que ninguna acción se venda a pérdida
        precio_max_compra = max(acciones_a_vender)
        if precio_max_compra > precio_venta:
            # Esta cantidad incluye acciones que se venderían a pérdida
            # Probar con menos acciones
            continue

        # Calcular ganancia total
        costo_total = sum(acciones_a_vender)
        ingreso_total = precio_venta * cant
        ganancia_pct = ((ingreso_total - costo_total) / costo_total) * 100 if costo_total > 0 else 0

        if ganancia_pct >= ganancia_min_pct:
            return (cant, "OK")

    # Si llegamos aquí, ninguna cantidad cumple ambas condiciones
    # Determinar el motivo principal
    if acciones_fifo and acciones_fifo[0] > precio_venta:
        return (0, "ESPERAR (pérdida individual)")
    else:
        return (0, "ESPERAR (ganancia insuficiente)")


def calcular_ganancia_perdida(operaciones_param=None):
    """Calcula el total de ganancia o pérdida efectiva de las operaciones

    Fórmula: Ganancia/Pérdida = (Ventas + Valor actual cartera) - Compras

    Las acciones no vendidas se valoran al último precio de cierre disponible.
    """
    operaciones = operaciones_param if operaciones_param is not None else cargar_historial_operaciones()
    total_compras = 0
    total_ventas = 0

    for op in operaciones:
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)
        monto = precio * cantidad

        if tipo == "compra":
            total_compras += monto
        elif tipo == "venta":
            total_ventas += monto

    # Calcular valor actual de la cartera (acciones no vendidas)
    cartera = calcular_cartera(operaciones)
    valor_cartera = 0

    # Obtener últimos precios de cierre
    ultimos_precios = {}
    if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
        try:
            df_log = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
            # Para cada ticker, obtener el último precio de cierre
            for ticker in df_log['Ticker'].unique():
                df_ticker = df_log[df_log['Ticker'] == ticker].sort_values('Date')
                if not df_ticker.empty:
                    ultimos_precios[ticker] = df_ticker.iloc[-1]['Close']
        except Exception as e:
            print(f"[WARN] Error leyendo precios: {e}")

    # Calcular valor de la cartera
    for symbol, datos in cartera.items():
        acciones = datos.get("acciones", 0)
        if acciones > 0 and symbol in ultimos_precios:
            precio = ultimos_precios[symbol]
            # Ignorar precios NaN
            if pd.notna(precio):
                valor_cartera += acciones * precio

    # Ganancia/Pérdida = (Ventas + Valor cartera) - Compras
    ganancia_perdida = (total_ventas + valor_cartera) - total_compras

    # Calcular ganancia realizada (solo de acciones vendidas)
    ganancia_realizada = calcular_ganancia_realizada(operaciones)

    return {
        "total_compras": total_compras,
        "total_ventas": total_ventas,
        "valor_cartera": valor_cartera,
        "ganancia_perdida": ganancia_perdida,
        "ganancia_realizada": ganancia_realizada
    }


def calcular_ganancia_realizada(operaciones_param=None):
    """Calcula la ganancia/pérdida realizada solo de acciones que se vendieron.

    Usa FIFO por precio más bajo: las ventas se asignan primero a las compras
    de menor precio. La ganancia realizada es la diferencia entre el precio
    de venta y el precio de compra de cada acción vendida.
    """
    operaciones = operaciones_param if operaciones_param is not None else cargar_historial_operaciones()

    # Ordenar por fecha para procesar en orden cronológico
    operaciones_ordenadas = sorted(operaciones, key=lambda x: x.get("fecha", ""))

    # Diccionario de compras disponibles por ticker
    # Cada entrada es una lista de [precio, cantidad_disponible]
    compras_por_ticker = {}

    ganancia_total = 0

    for op in operaciones_ordenadas:
        symbol = op.get("ticker_symbol")
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)

        if symbol not in compras_por_ticker:
            compras_por_ticker[symbol] = []

        if tipo == "compra":
            # Agregar compra a la lista (ordenada por precio ascendente)
            compras_por_ticker[symbol].append([precio, cantidad])
            compras_por_ticker[symbol].sort(key=lambda x: x[0])

        elif tipo == "venta":
            # Consumir de las compras de menor precio primero
            cantidad_a_vender = cantidad
            precio_venta = precio

            for compra in compras_por_ticker[symbol]:
                if cantidad_a_vender <= 0:
                    break
                if compra[1] > 0:
                    # Cantidad a consumir de esta compra
                    consumir = min(compra[1], cantidad_a_vender)
                    precio_compra = compra[0]

                    # Calcular ganancia de esta porción
                    ganancia_porcion = (precio_venta - precio_compra) * consumir
                    ganancia_total += ganancia_porcion

                    # Reducir cantidad disponible de esta compra
                    compra[1] -= consumir
                    cantidad_a_vender -= consumir

            # Limpiar compras agotadas
            compras_por_ticker[symbol] = [c for c in compras_por_ticker[symbol] if c[1] > 0]

    return ganancia_total


def administrar_historial():
    """Abre ventana para gestionar el historial de operaciones"""
    ruta = obtener_ruta_historial()
    if ruta is None:
        messagebox.showerror("Error", "No hay ubicacion configurada.\nVerifica que exista la carpeta data/")
        return

    # Cargar datos completos (incluyendo config_plataformas)
    datos_historial = cargar_historial_operaciones_completo()
    operaciones = datos_historial.get("operaciones", [])
    config_plataformas = datos_historial.get("config_plataformas", {"TYBA": {"moneda": "USD"}})

    # Crear ventana
    ventana_hist = tk.Toplevel(root)
    ventana_hist.title("Historial de Operaciones")
    ventana_hist.geometry("900x750")

    # Frame selector de plataforma y modo
    frame_plataforma = tk.Frame(ventana_hist, pady=5)
    frame_plataforma.pack(fill="x", padx=10)

    tk.Label(frame_plataforma, text="Plataforma:", font=("Arial", 10, "bold")).pack(side="left", padx=(0, 5))
    plataforma_var = tk.StringVar(value=list(config_plataformas.keys())[0] if config_plataformas else "TYBA")
    combo_plataforma = ttk.Combobox(frame_plataforma, textvariable=plataforma_var,
                                     values=list(config_plataformas.keys()),
                                     state="readonly", width=15)
    combo_plataforma.pack(side="left", padx=(0, 10))

    # Selector de modo (Paper/Real/Todos)
    tk.Label(frame_plataforma, text="Modo:", font=("Arial", 10, "bold")).pack(side="left", padx=(10, 5))
    modo_var = tk.StringVar(value="Real")  # Default: Real (no mezclar Paper y Real)
    combo_modo = ttk.Combobox(frame_plataforma, textvariable=modo_var,
                               values=["Real", "Paper", "Todos"],
                               state="readonly", width=8)
    combo_modo.pack(side="left", padx=(0, 10))

    # Mostrar moneda de la plataforma seleccionada
    moneda_plat = config_plataformas.get(plataforma_var.get(), {}).get("moneda", "USD")
    lbl_moneda = tk.Label(frame_plataforma, text=f"Moneda: {moneda_plat}", font=("Arial", 9), fg="gray")
    lbl_moneda.pack(side="left", padx=10)

    # Descripción de la plataforma
    desc_plat = config_plataformas.get(plataforma_var.get(), {}).get("descripcion", "")
    lbl_desc = tk.Label(frame_plataforma, text=desc_plat, font=("Arial", 9), fg="gray")
    lbl_desc.pack(side="left", padx=10)

    def obtener_operaciones_plataforma():
        """Retorna las operaciones filtradas por plataforma y modo"""
        plat = plataforma_var.get()
        modo = modo_var.get()

        # Filtrar por plataforma
        ops_filtradas = [op for op in operaciones if op.get("plataforma", "TYBA") == plat]

        # Filtrar por modo si no es "Todos"
        # Default depende de la plataforma: TYBA=Real, IBKR-UK=Paper
        if modo != "Todos":
            modo_lower = modo.lower()  # "paper" o "real"
            def get_modo_default(op):
                if "modo" in op:
                    return op["modo"].lower()
                # Sin campo modo: TYBA es Real, resto es Paper
                return "real" if op.get("plataforma", "TYBA") == "TYBA" else "paper"
            ops_filtradas = [op for op in ops_filtradas if get_modo_default(op) == modo_lower]

        return ops_filtradas

    # Frame superior - Estado de cartera
    frame_cartera = tk.LabelFrame(ventana_hist, text="Estado Actual de Cartera", pady=5, padx=5)
    frame_cartera.pack(fill="x", padx=10, pady=5)

    # Treeview para cartera
    cols_cartera = ("Symbol", "Acciones", "P. Prom. Compra", "Capital Invertido")
    tree_cartera = ttk.Treeview(frame_cartera, columns=cols_cartera, show="headings", height=4)

    for col in cols_cartera:
        tree_cartera.heading(col, text=col)
        tree_cartera.column(col, width=120, anchor="center")

    tree_cartera.pack(fill="x", pady=5)

    def actualizar_cartera():
        """Actualiza la vista de cartera con operaciones de la plataforma seleccionada"""
        for item in tree_cartera.get_children():
            tree_cartera.delete(item)

        ops_plataforma = obtener_operaciones_plataforma()
        cartera = calcular_cartera(ops_plataforma)
        # Ordenar alfabéticamente por symbol
        for symbol, datos in sorted(cartera.items(), key=lambda x: x[0].upper()):
            if datos["acciones"] > 0 or datos["total_comprado"] > 0:
                tree_cartera.insert("", "end", values=(
                    symbol,
                    datos["acciones"],
                    f"${datos['precio_promedio_compra']:.2f}" if datos['precio_promedio_compra'] > 0 else "-",
                    f"${datos['capital_invertido']:.2f}" if datos['capital_invertido'] > 0 else "-"
                ))

    actualizar_cartera()

    # Frame resumen - Ganancia/Pérdida
    frame_resumen = tk.LabelFrame(ventana_hist, text="Resumen de Operaciones (Calculado con el precio de cierre)", pady=5, padx=5)
    frame_resumen.pack(fill="x", padx=10, pady=5)

    # Labels para mostrar resumen
    frame_resumen_inner = tk.Frame(frame_resumen)
    frame_resumen_inner.pack(fill="x", pady=5)

    lbl_compras = tk.Label(frame_resumen_inner, text="Compras: $0.00", font=("Arial", 9))
    lbl_compras.pack(side="left", padx=10)

    lbl_ventas = tk.Label(frame_resumen_inner, text="Ventas: $0.00", font=("Arial", 9))
    lbl_ventas.pack(side="left", padx=10)

    lbl_cartera = tk.Label(frame_resumen_inner, text="Cartera: $0.00", font=("Arial", 9), fg="#0066cc")
    lbl_cartera.pack(side="left", padx=10)

    lbl_realizada = tk.Label(frame_resumen_inner, text="Realizada: $0.00", font=("Arial", 9, "bold"))
    lbl_realizada.pack(side="left", padx=10)

    lbl_global = tk.Label(frame_resumen_inner, text="Global: $0.00", font=("Arial", 9, "bold"))
    lbl_global.pack(side="left", padx=10)

    # Separador y botón Total Real
    tk.Label(frame_resumen_inner, text="|", font=("Arial", 10), fg="gray").pack(side="left", padx=(10, 5))

    def mostrar_total_real():
        """Muestra el total de todas las plataformas en modo Real"""
        ops_real_total = []
        for plat in config_plataformas.keys():
            ops_plat = [op for op in operaciones if op.get("plataforma", "TYBA") == plat]
            def get_modo(op):
                if "modo" in op:
                    return op["modo"].lower()
                return "real" if op.get("plataforma", "TYBA") == "TYBA" else "paper"
            ops_real = [op for op in ops_plat if get_modo(op) == "real"]
            ops_real_total.extend(ops_real)

        if not ops_real_total:
            messagebox.showinfo("Total Real", "No hay operaciones en modo Real.", parent=ventana_hist)
            return

        resultado = calcular_ganancia_perdida(ops_real_total)
        realizada = resultado['ganancia_realizada']
        global_val = resultado['ganancia_perdida']

        color_r = "ganancia" if realizada >= 0 else "pérdida"
        color_g = "ganancia" if global_val >= 0 else "pérdida"

        msg = "TOTAL REAL (todas las plataformas)" + chr(10) + chr(10)
        msg += f"Compras: ${resultado['total_compras']:,.2f}" + chr(10)
        msg += f"Ventas: ${resultado['total_ventas']:,.2f}" + chr(10)
        msg += f"Cartera: ${resultado['valor_cartera']:,.2f}" + chr(10) + chr(10)
        msg += f"Realizada: ${realizada:,.2f} ({color_r})" + chr(10)
        msg += f"Global: ${global_val:,.2f} ({color_g})"

        messagebox.showinfo("Total Real", msg, parent=ventana_hist)

    btn_total_real = tk.Button(frame_resumen_inner, text="Total Real", font=("Arial", 8),
                                command=mostrar_total_real, bg="#e0e0ff")
    btn_total_real.pack(side="left", padx=5)

    def actualizar_resumen():
        """Actualiza el resumen de ganancia/pérdida de la plataforma seleccionada"""
        ops_plataforma = obtener_operaciones_plataforma()
        resultado = calcular_ganancia_perdida(ops_plataforma)
        lbl_compras.config(text=f"Compras: ${resultado['total_compras']:,.2f}")
        lbl_ventas.config(text=f"Ventas: ${resultado['total_ventas']:,.2f}")
        lbl_cartera.config(text=f"Cartera: ${resultado['valor_cartera']:,.2f}")

        # Ganancia realizada (solo de acciones vendidas)
        gr = resultado['ganancia_realizada']
        if gr >= 0:
            lbl_realizada.config(text=f"Realizada: ${gr:,.2f}", fg="green")
        else:
            lbl_realizada.config(text=f"Realizada: -${abs(gr):,.2f}", fg="red")

        # Ganancia global (ventas + cartera - compras)
        gp = resultado['ganancia_perdida']
        if gp >= 0:
            lbl_global.config(text=f"Global: ${gp:,.2f}", fg="green")
        else:
            lbl_global.config(text=f"Global: -${abs(gp):,.2f}", fg="red")

    actualizar_resumen()

    # Frame Capital y Posiciones (visible para todas las plataformas)
    frame_capital = tk.LabelFrame(ventana_hist, text="Capital y Posiciones (según última sincronización o entrada manual)", pady=5, padx=5)

    # Variables para Paper
    ibkr_paper_capital_var = tk.StringVar(value="-")
    ibkr_paper_pos_var = tk.StringVar(value="-")

    # Variables para Live
    ibkr_live_capital_var = tk.StringVar(value="-")
    ibkr_live_pos_var = tk.StringVar(value="-")
    ibkr_live_balances_var = tk.StringVar(value="")  # Balances por moneda (GBP, USD)
    ibkr_live_invertido_var = tk.StringVar(value="")  # Capital invertido

    # Frame con dos columnas: Paper y Live
    frame_capital_datos = tk.Frame(frame_capital)
    frame_capital_datos.pack(fill="x", pady=5)

    # Variables para fechas de sync
    ibkr_paper_fecha_var = tk.StringVar(value="Sin datos")
    ibkr_live_fecha_var = tk.StringVar(value="Sin datos")

    # Cargar datos guardados
    sync_paper = cargar_sync_ibkr("paper")
    if sync_paper:
        ibkr_paper_capital_var.set(sync_paper.get("capital", "-"))
        pos_paper = sync_paper.get("posiciones", "-")
        # Si es dict, mostrar número de posiciones; si es string, mostrar directo
        if isinstance(pos_paper, dict):
            ibkr_paper_pos_var.set(str(len(pos_paper)))
        else:
            ibkr_paper_pos_var.set(pos_paper)
        ibkr_paper_fecha_var.set(f"Sync: {sync_paper.get('fecha', '-')}")

    sync_real = cargar_sync_ibkr("real")
    if sync_real:
        ibkr_live_capital_var.set(sync_real.get("capital", "-"))
        pos_real = sync_real.get("posiciones", "-")
        # Si es dict, mostrar número de posiciones; si es string, mostrar directo
        if isinstance(pos_real, dict):
            ibkr_live_pos_var.set(str(len(pos_real)))
        else:
            ibkr_live_pos_var.set(pos_real)
        ibkr_live_fecha_var.set(f"Sync: {sync_real.get('fecha', '-')}")
        # Balances por moneda (Cash)
        balances = sync_real.get("balances_por_moneda", {})
        if balances:
            bal_str = " / ".join([f"{m}: {v:,.2f}" for m, v in balances.items() if abs(v) > 0.01])
            ibkr_live_balances_var.set(f"Cash: {bal_str}")

    # Cargar capital invertido para IBKR-UK
    capital_inv = calcular_capital_invertido("IBKR-UK")
    if capital_inv:
        inv_str = " | ".join([f"{m}: {v:,.2f}" for m, v in capital_inv.items()])
        ibkr_live_invertido_var.set(f"Invertido: {inv_str}")

    # Columna Paper
    frame_paper = tk.Frame(frame_capital_datos)
    frame_paper.pack(side="left", padx=20)

    # Encabezado: PAPER (Simulador) + Fecha sync
    frame_paper_header = tk.Frame(frame_paper)
    frame_paper_header.pack(anchor="w")
    tk.Label(frame_paper_header, text="PAPER (Simulador)", font=("Arial", 9, "bold"),
             fg="#6f42c1").pack(side="left")
    tk.Label(frame_paper_header, textvariable=ibkr_paper_fecha_var,
             font=("Arial", 8), fg="gray").pack(side="left", padx=(10, 0))

    frame_paper_data = tk.Frame(frame_paper)
    frame_paper_data.pack(anchor="w")
    tk.Label(frame_paper_data, text="Capital:", font=("Arial", 9)).pack(side="left")
    tk.Label(frame_paper_data, textvariable=ibkr_paper_capital_var,
             font=("Arial", 10, "bold"), fg="#0066cc").pack(side="left", padx=5)
    tk.Label(frame_paper_data, text="Pos:", font=("Arial", 9)).pack(side="left", padx=(10, 0))
    tk.Label(frame_paper_data, textvariable=ibkr_paper_pos_var,
             font=("Arial", 9)).pack(side="left", padx=5)

    # Separador entre Paper y Live
    separador_paper_live = tk.Label(frame_capital_datos, text="|", font=("Arial", 12), fg="gray")
    separador_paper_live.pack(side="left", padx=10)

    # Columna Live
    frame_live = tk.Frame(frame_capital_datos)
    frame_live.pack(side="left", padx=20)

    # Encabezado: LIVE (Real) + Fecha sync
    frame_live_header = tk.Frame(frame_live)
    frame_live_header.pack(anchor="w")
    tk.Label(frame_live_header, text="LIVE (Real)", font=("Arial", 9, "bold"),
             fg="#dc3545").pack(side="left")
    tk.Label(frame_live_header, textvariable=ibkr_live_fecha_var,
             font=("Arial", 8), fg="gray").pack(side="left", padx=(10, 0))

    frame_live_data = tk.Frame(frame_live)
    frame_live_data.pack(anchor="w")
    tk.Label(frame_live_data, text="Capital:", font=("Arial", 9)).pack(side="left")
    tk.Label(frame_live_data, textvariable=ibkr_live_capital_var,
             font=("Arial", 10, "bold"), fg="#0066cc").pack(side="left", padx=5)
    tk.Label(frame_live_data, text="Pos:", font=("Arial", 9)).pack(side="left", padx=(10, 0))
    tk.Label(frame_live_data, textvariable=ibkr_live_pos_var,
             font=("Arial", 9)).pack(side="left", padx=5)
    # Balances por moneda (GBP, USD)
    tk.Label(frame_live, textvariable=ibkr_live_balances_var,
             font=("Arial", 8), fg="#006600").pack(anchor="w")
    # Capital invertido
    tk.Label(frame_live, textvariable=ibkr_live_invertido_var,
             font=("Arial", 8), fg="#cc6600").pack(anchor="w")

    def consultar_ibkr_datos(puerto, modo_texto):
        """Consulta datos de IBKR para un puerto específico.

        Returns:
            tuple: (capital_str, num_posiciones, pos_activas, balances_por_moneda)
            balances_por_moneda es dict {moneda: valor} para mostrar GBP/USD separados
        """
        try:
            from ib_insync import IB
            ib = IB()
            ib.connect('127.0.0.1', puerto, clientId=3, timeout=5)

            if not ib.isConnected():
                return None, None, "No conectado", {}

            acc_values = ib.accountValues()

            # Obtener balances por moneda
            balances_por_moneda = {}
            net_liq = 0
            moneda_base = "USD"

            # Detectar moneda base
            for av in acc_values:
                if av.tag == "NetLiquidation" and av.currency and av.currency != "BASE":
                    moneda_base = av.currency
                    break

            # Obtener CashBalance por cada moneda
            for av in acc_values:
                if av.tag == "CashBalance" and av.currency and av.currency != "BASE":
                    try:
                        balances_por_moneda[av.currency] = float(av.value)
                    except:
                        pass
                elif av.tag == "NetLiquidation" and (av.currency == moneda_base or av.currency == "BASE"):
                    try:
                        net_liq = float(av.value)
                    except:
                        pass

            # Construir string de capital (total en moneda base)
            simbolo = {"USD": "$", "GBP": "£", "EUR": "€", "JPY": "¥", "CHF": "Fr"}.get(moneda_base, moneda_base + " ")
            total_cash = sum(balances_por_moneda.values()) if balances_por_moneda else 0

            posiciones = ib.positions()
            pos_activas = [p for p in posiciones if int(p.position) != 0]

            ib.disconnect()

            # Formato: mostrar Net Liquidation (total cuenta)
            capital_str = f"{simbolo}{net_liq:,.2f}"

            return capital_str, f"{len(pos_activas)}", pos_activas, balances_por_moneda

        except Exception as e:
            return None, None, str(e), {}

    def consultar_datos_guardados():
        """Actualiza los labels con datos guardados de la plataforma actual"""
        plat_actual = plataforma_var.get()
        actualizar_datos_capital_plataforma(plat_actual)

    def sincronizar_historial_ibkr():
        """Sincroniza historial de ejecuciones Y capital/posiciones desde IBKR"""
        from datetime import datetime
        try:
            from ib_insync import IB, ExecutionFilter

            # Usar el modo seleccionado en la interfaz
            modo_seleccionado = modo_var.get()  # "Todos", "Paper" o "Real"
            fecha_sync = datetime.now().strftime("%Y-%m-%d %H:%M")

            operaciones_nuevas = []
            resumen = f"Sincronizacion: {fecha_sync}\n"
            resumen += f"Modo: {modo_seleccionado}\n\n"

            def sincronizar_modo_ibkr(puerto, modo_texto):
                """Sincroniza ejecuciones y capital/posiciones de un modo IBKR"""
                try:
                    ib = IB()
                    ib.connect('127.0.0.1', puerto, clientId=4, timeout=10)

                    if not ib.isConnected():
                        return [], None, None, f"{modo_texto}: No conectado", {}

                    # 1. Obtener capital y posiciones
                    acc_values = ib.accountValues()
                    cash = 0
                    net_liq = 0
                    moneda_base = "USD"
                    balances_por_moneda = {}

                    for av in acc_values:
                        if av.tag == "NetLiquidation" and av.currency and av.currency != "BASE":
                            moneda_base = av.currency
                            break

                    # Obtener CashBalance por moneda, NetLiquidation y StockMarketValue
                    stock_value_by_currency = {}
                    for av in acc_values:
                        currency = av.currency or ""
                        if av.tag == "CashBalance" and currency and currency != "BASE":
                            try:
                                balances_por_moneda[currency] = float(av.value)
                            except:
                                pass
                        elif av.tag == "StockMarketValue" and currency and currency != "BASE":
                            try:
                                stock_value_by_currency[currency] = float(av.value)
                            except:
                                pass
                        elif av.tag == "NetLiquidation" and (currency == moneda_base or currency == "BASE"):
                            try:
                                net_liq = float(av.value)
                            except:
                                pass
                        elif currency == moneda_base or currency == "" or currency == "BASE":
                            if av.tag == "AvailableFunds":
                                cash = float(av.value)
                            elif av.tag == "CashBalance" and cash == 0:
                                cash = float(av.value)

                    simbolos = {"USD": "$", "GBP": "£", "EUR": "€", "JPY": "¥", "CHF": "Fr"}
                    simbolo_base = simbolos.get(moneda_base, moneda_base + " ")

                    # Construir desglose del capital
                    componentes = []
                    # Agregar valor de acciones por moneda
                    for curr, val in stock_value_by_currency.items():
                        if abs(val) > 0.01:
                            simb = simbolos.get(curr, curr + " ")
                            componentes.append(f"{simb}{val:,.2f}")
                    # Agregar efectivo por moneda
                    for curr, val in balances_por_moneda.items():
                        if abs(val) > 0.01:
                            simb = simbolos.get(curr, curr + " ")
                            componentes.append(f"{simb}{val:,.2f}")

                    # Formato: "£4121.87 = $4400 + £779.68"
                    if net_liq > 0:
                        if componentes:
                            capital_str = f"{simbolo_base}{net_liq:,.2f} = {' + '.join(componentes)}"
                        else:
                            capital_str = f"{simbolo_base}{net_liq:,.2f}"
                    else:
                        capital_str = f"{simbolo_base}{cash:,.2f}"

                    posiciones = ib.positions()
                    pos_activas = [p for p in posiciones if int(p.position) != 0]
                    # Crear dict con detalle de posiciones {ticker: cantidad}
                    pos_dict = {p.contract.symbol: int(p.position) for p in pos_activas}

                    # 2. Obtener ejecuciones
                    exec_filter = ExecutionFilter()
                    executions = ib.reqExecutions(exec_filter)
                    ib.sleep(1)
                    fills = ib.fills()

                    ops = []
                    ops_procesadas = set()  # Clave única: ticker+fecha+hora+tipo+cantidad

                    for fill in fills:
                        exec_info = fill.execution
                        contract = fill.contract

                        # Ignorar conversiones de moneda
                        if contract.symbol in ['GBP', 'USD', 'EUR']:
                            continue

                        try:
                            exec_time = datetime.strptime(exec_info.time, "%Y%m%d  %H:%M:%S")
                        except:
                            try:
                                exec_time = datetime.fromisoformat(str(exec_info.time).replace('+00:00', ''))
                            except:
                                exec_time = datetime.now()

                        # Clave única para evitar duplicados
                        clave = f"{contract.symbol}_{exec_time.strftime('%Y%m%d%H%M%S')}_{exec_info.side}_{int(abs(fill.execution.shares))}"
                        if clave in ops_procesadas:
                            continue
                        ops_procesadas.add(clave)

                        op = {
                            "fecha": exec_time.strftime("%Y-%m-%d"),
                            "ticker_symbol": contract.symbol,
                            "tipo": "compra" if exec_info.side == "BOT" else "venta",
                            "precio": round(fill.execution.avgPrice, 2),
                            "cantidad": int(abs(fill.execution.shares)),
                            "plataforma": "IBKR-UK",
                            "modo": modo_texto,
                            "fuente": "sync_ibkr",
                            "hora": exec_time.strftime("%H:%M:%S"),
                            "comision": round(fill.commissionReport.commission, 2) if fill.commissionReport else 0,
                            "exec_id": clave
                        }
                        ops.append(op)

                    for exec_trade in executions:
                        exec_info = exec_trade.execution
                        contract = exec_trade.contract

                        # Ignorar conversiones de moneda
                        if contract.symbol in ['GBP', 'USD', 'EUR']:
                            continue

                        try:
                            exec_time = datetime.strptime(exec_info.time, "%Y%m%d  %H:%M:%S")
                        except:
                            try:
                                exec_time = datetime.fromisoformat(str(exec_info.time).replace('+00:00', ''))
                            except:
                                exec_time = datetime.now()

                        clave = f"{contract.symbol}_{exec_time.strftime('%Y%m%d%H%M%S')}_{exec_info.side}_{int(abs(exec_info.shares))}"
                        if clave in ops_procesadas:
                            continue
                        ops_procesadas.add(clave)

                        op = {
                            "fecha": exec_time.strftime("%Y-%m-%d"),
                            "ticker_symbol": contract.symbol,
                            "tipo": "compra" if exec_info.side == "BOT" else "venta",
                            "precio": round(exec_info.avgPrice, 2),
                            "cantidad": int(abs(exec_info.shares)),
                            "plataforma": "IBKR-UK",
                            "modo": modo_texto,
                            "fuente": "sync_ibkr",
                            "hora": exec_time.strftime("%H:%M:%S"),
                            "comision": 0,
                            "exec_id": clave
                        }
                        ops.append(op)

                    ib.disconnect()

                    msg = f"=== {modo_texto.upper()} ===\n"
                    msg += f"Capital: {capital_str}\n"
                    # Mostrar balances por moneda
                    if balances_por_moneda:
                        bal_str = " | ".join([f"{m}: {v:,.2f}" for m, v in balances_por_moneda.items() if abs(v) > 0.01])
                        if bal_str:
                            msg += f"Balances: {bal_str}\n"
                    msg += f"Posiciones: {len(pos_dict)}\n"
                    # Mostrar detalle de posiciones
                    if pos_activas:
                        msg += "Detalle:\n"
                        for p in pos_activas:
                            msg += f"  - {p.contract.symbol}: {int(p.position)} acciones\n"
                    msg += f"Ejecuciones: {len(ops)}\n"

                    # Combinar balances y stock values para guardar
                    datos_monedas = {
                        "cash": balances_por_moneda,
                        "stocks": stock_value_by_currency
                    }
                    return ops, capital_str, pos_dict, msg, datos_monedas

                except Exception as e:
                    return [], None, None, f"=== {modo_texto.upper()} ===\nError: {str(e)}\n", {"cash": {}, "stocks": {}}

            # Sincronizar según modo seleccionado
            if modo_seleccionado in ["Paper", "Todos"]:
                ops_paper, capital_p, pos_p, msg_paper, balances_p = sincronizar_modo_ibkr(7497, "Paper")
                operaciones_nuevas.extend(ops_paper)
                resumen += msg_paper + "\n"
                if capital_p:
                    guardar_sync_ibkr("paper", capital_p, pos_p, fecha_sync, balances_p)
                    ibkr_paper_capital_var.set(capital_p)
                    # pos_p ahora es dict, mostrar número de posiciones
                    ibkr_paper_pos_var.set(str(len(pos_p)) if isinstance(pos_p, dict) else pos_p)
                    ibkr_paper_fecha_var.set(f"Sync: {fecha_sync}")

            if modo_seleccionado in ["Real", "Todos"]:
                ops_real, capital_r, pos_r, msg_real, balances_r = sincronizar_modo_ibkr(7496, "Real")
                operaciones_nuevas.extend(ops_real)
                resumen += msg_real + "\n"
                if capital_r:
                    guardar_sync_ibkr("real", capital_r, pos_r, fecha_sync, balances_r)
                    ibkr_live_capital_var.set(capital_r)
                    # pos_r ahora es dict, mostrar número de posiciones
                    ibkr_live_pos_var.set(str(len(pos_r)) if isinstance(pos_r, dict) else pos_r)
                    ibkr_live_fecha_var.set(f"Sync: {fecha_sync}")
                    # Mostrar balances por moneda (Cash)
                    if balances_r and balances_r.get("cash"):
                        cash_dict = balances_r["cash"]
                        bal_str = " / ".join([f"{m}: {v:,.2f}" for m, v in cash_dict.items() if abs(v) > 0.01])
                        ibkr_live_balances_var.set(f"Cash: {bal_str}")

            # Procesar operaciones nuevas
            if operaciones_nuevas:
                todas_ops = cargar_historial_operaciones()

                # Crear claves únicas de operaciones existentes (fecha+ticker+tipo+precio+cantidad+plataforma+modo)
                def clave_operacion(op):
                    ticker = op.get('ticker_symbol') or op.get('symbol', '')
                    return (
                        op.get('fecha', ''),
                        ticker,
                        op.get('tipo', '').lower(),
                        op.get('precio', 0),
                        op.get('cantidad', 0),
                        op.get('plataforma', ''),
                        op.get('modo', '')
                    )

                claves_existentes = set(clave_operacion(op) for op in todas_ops)

                # Filtrar solo operaciones que no existan
                ops_filtradas = [op for op in operaciones_nuevas
                                if clave_operacion(op) not in claves_existentes]

                if ops_filtradas:
                    todas_ops.extend(ops_filtradas)
                    guardar_historial_operaciones(todas_ops)
                    resumen += f"{len(ops_filtradas)} operaciones nuevas agregadas."
                    actualizar_historial()
                    actualizar_cartera()
                    actualizar_resumen()
                else:
                    resumen += "No hay operaciones nuevas."

            messagebox.showinfo("Sync IBKR", resumen, parent=ventana_hist)

        except ImportError:
            messagebox.showerror("Error", "Librería ib_insync no instalada.\n\nEjecuta: pip install ib_insync", parent=ventana_hist)
        except Exception as e:
            messagebox.showerror("Error", f"Error sincronizando:\n{str(e)}", parent=ventana_hist)

    # Frame para botones IBKR (2 filas)
    frame_capital_botones = tk.Frame(frame_capital_datos)
    frame_capital_botones.pack(side="right", padx=10)

    # Fila 1: Ver Guardado y Sync IBKR
    frame_capital_fila1 = tk.Frame(frame_capital_botones)
    frame_capital_fila1.pack(anchor="e")

    btn_consultar_ibkr = tk.Button(frame_capital_fila1, text="Ver Guardado",
                                   command=consultar_datos_guardados,
                                   bg="#6c757d", fg="white", font=("Arial", 9, "bold"))
    btn_consultar_ibkr.pack(side="left", padx=2)

    btn_sync_ibkr = tk.Button(frame_capital_fila1, text="Sync IBKR",
                              command=sincronizar_historial_ibkr,
                              bg="#17a2b8", fg="white", font=("Arial", 9, "bold"))
    btn_sync_ibkr.pack(side="left", padx=2)

    btn_editar_manual = tk.Button(frame_capital_fila1, text="Editar Manual",
                                  command=lambda: abrir_ventana_editar_capital(),
                                  bg="#28a745", fg="white", font=("Arial", 9, "bold"))
    btn_editar_manual.pack(side="left", padx=2)

    # Fila 2: Transferencias
    frame_capital_fila2 = tk.Frame(frame_capital_botones)
    frame_capital_fila2.pack(anchor="e", pady=(2, 0))

    def abrir_ventana_transferencias():
        """Abre ventana para registrar depósitos/retiros."""
        from datetime import datetime

        # Intentar importar tkcalendar (opcional)
        try:
            from tkcalendar import DateEntry
            tiene_calendar = True
        except ImportError:
            tiene_calendar = False

        vent_trans = tk.Toplevel(ventana_hist)
        vent_trans.title("Transferencias - Capital Invertido")
        vent_trans.geometry("600x500")
        vent_trans.transient(ventana_hist)
        vent_trans.grab_set()

        # Frame superior - Formulario
        frame_form = tk.LabelFrame(vent_trans, text="Registrar Transferencia", padx=10, pady=10)
        frame_form.pack(fill="x", padx=10, pady=10)

        # Plataforma (solo REAL)
        frame_plat = tk.Frame(frame_form)
        frame_plat.pack(fill="x", pady=2)
        tk.Label(frame_plat, text="Plataforma:", width=12, anchor="w").pack(side="left")
        plat_trans_var = tk.StringVar(value="IBKR-UK")
        plataformas_real = [p for p in obtener_plataformas()]
        combo_plat = ttk.Combobox(frame_plat, textvariable=plat_trans_var,
                                   values=plataformas_real, state="readonly", width=15)
        combo_plat.pack(side="left", padx=5)

        # Fecha
        frame_fecha = tk.Frame(frame_form)
        frame_fecha.pack(fill="x", pady=2)
        tk.Label(frame_fecha, text="Fecha:", width=12, anchor="w").pack(side="left")
        if tiene_calendar:
            date_entry = DateEntry(frame_fecha, width=12, date_pattern="yyyy-mm-dd")
            date_entry.pack(side="left", padx=5)
        else:
            # Si no tiene tkcalendar, usar Entry simple
            date_entry = tk.Entry(frame_fecha, width=12)
            date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
            date_entry.pack(side="left", padx=5)
            tk.Label(frame_fecha, text="(YYYY-MM-DD)", font=("Arial", 8), fg="gray").pack(side="left")

        # Monto
        frame_monto = tk.Frame(frame_form)
        frame_monto.pack(fill="x", pady=2)
        tk.Label(frame_monto, text="Monto:", width=12, anchor="w").pack(side="left")
        entry_monto = tk.Entry(frame_monto, width=15)
        entry_monto.pack(side="left", padx=5)
        tk.Label(frame_monto, text="(positivo=depósito, negativo=retiro)",
                 font=("Arial", 8), fg="gray").pack(side="left")

        # Moneda
        frame_moneda = tk.Frame(frame_form)
        frame_moneda.pack(fill="x", pady=2)
        tk.Label(frame_moneda, text="Moneda:", width=12, anchor="w").pack(side="left")
        moneda_var = tk.StringVar(value="USD")
        combo_moneda = ttk.Combobox(frame_moneda, textvariable=moneda_var,
                                     values=MONEDAS_DISPONIBLES, state="readonly", width=8)
        combo_moneda.pack(side="left", padx=5)

        # Descripción
        frame_desc = tk.Frame(frame_form)
        frame_desc.pack(fill="x", pady=2)
        tk.Label(frame_desc, text="Descripción:", width=12, anchor="w").pack(side="left")
        entry_desc = tk.Entry(frame_desc, width=30)
        entry_desc.pack(side="left", padx=5)

        # Label de resultado
        lbl_resultado = tk.Label(frame_form, text="", fg="blue")
        lbl_resultado.pack(pady=5)

        def registrar_transferencia():
            plat = plat_trans_var.get()
            try:
                if tiene_calendar:
                    fecha = date_entry.get_date().strftime("%Y-%m-%d")
                else:
                    fecha = date_entry.get()
            except:
                fecha = date_entry.get()

            try:
                monto = float(entry_monto.get().replace(",", ""))
            except ValueError:
                lbl_resultado.config(text="Monto inválido", fg="red")
                return

            moneda = moneda_var.get()
            desc = entry_desc.get().strip()

            exito, mensaje = guardar_transferencia(plat, monto, moneda, fecha, desc)

            if exito:
                lbl_resultado.config(text=mensaje, fg="green")
                entry_monto.delete(0, tk.END)
                entry_desc.delete(0, tk.END)
                actualizar_lista_transferencias()
                actualizar_capital_invertido()
            else:
                lbl_resultado.config(text=mensaje, fg="red")

        tk.Button(frame_form, text="Registrar", command=registrar_transferencia,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=5)

        # Frame inferior - Lista de transferencias
        frame_lista = tk.LabelFrame(vent_trans, text="Historial de Transferencias", padx=10, pady=10)
        frame_lista.pack(fill="both", expand=True, padx=10, pady=5)

        # Treeview para mostrar transferencias
        cols_trans = ("Fecha", "Tipo", "Monto", "Moneda", "Descripción")
        tree_trans = ttk.Treeview(frame_lista, columns=cols_trans, show="headings", height=10)
        for col in cols_trans:
            tree_trans.heading(col, text=col)
            tree_trans.column(col, width=100 if col != "Descripción" else 150, anchor="center")
        tree_trans.pack(fill="both", expand=True)

        # Scrollbar
        scroll_trans = ttk.Scrollbar(frame_lista, orient="vertical", command=tree_trans.yview)
        tree_trans.configure(yscrollcommand=scroll_trans.set)

        def actualizar_lista_transferencias():
            tree_trans.delete(*tree_trans.get_children())
            plat = plat_trans_var.get()
            transferencias = cargar_transferencias(plat)
            for t in transferencias:
                tipo = "Depósito" if t["monto"] >= 0 else "Retiro"
                tree_trans.insert("", "end", values=(
                    t["fecha"],
                    tipo,
                    f"{abs(t['monto']):,.2f}",
                    t["moneda"],
                    t.get("descripcion", "")
                ))

        def eliminar_seleccionada():
            seleccion = tree_trans.selection()
            if not seleccion:
                messagebox.showwarning("Aviso", "Selecciona una transferencia", parent=vent_trans)
                return
            if not messagebox.askyesno("Confirmar", "¿Eliminar transferencia seleccionada?", parent=vent_trans):
                return

            idx = tree_trans.index(seleccion[0])
            plat = plat_trans_var.get()
            exito, mensaje = eliminar_transferencia(plat, idx)
            if exito:
                actualizar_lista_transferencias()
                actualizar_capital_invertido()
                lbl_resultado.config(text=mensaje, fg="blue")
            else:
                lbl_resultado.config(text=mensaje, fg="red")

        # Botón eliminar
        tk.Button(frame_lista, text="Eliminar Seleccionada", command=eliminar_seleccionada,
                  bg="#dc3545", fg="white").pack(pady=5)

        # Label de totales
        lbl_totales = tk.Label(vent_trans, text="", font=("Arial", 10, "bold"), fg="#006600")
        lbl_totales.pack(pady=5)

        def actualizar_capital_invertido():
            plat = plat_trans_var.get()
            totales = calcular_capital_invertido(plat)
            if totales:
                total_str = " | ".join([f"{m}: {v:,.2f}" for m, v in totales.items()])
                lbl_totales.config(text=f"Capital Invertido ({plat}): {total_str}")
                # Actualizar también la variable de la GUI principal
                ibkr_live_invertido_var.set(f"Invertido: {total_str}")
            else:
                lbl_totales.config(text=f"Capital Invertido ({plat}): Sin transferencias")
                ibkr_live_invertido_var.set("")

        # Actualizar al cambiar plataforma
        combo_plat.bind("<<ComboboxSelected>>", lambda e: (actualizar_lista_transferencias(), actualizar_capital_invertido()))

        # Cargar datos iniciales
        actualizar_lista_transferencias()
        actualizar_capital_invertido()

    btn_transferencias = tk.Button(frame_capital_fila2, text="Transferencias",
                                   command=abrir_ventana_transferencias,
                                   bg="#fd7e14", fg="white", font=("Arial", 9, "bold"))
    btn_transferencias.pack(side="right", padx=2)

    def actualizar_datos_capital_plataforma(plataforma):
        """Carga y muestra los datos de capital/posiciones para la plataforma actual."""
        modo_actual = modo_var.get()  # "Todos", "Paper" o "Real"

        # Para IBKR, cargar ambos modos
        if plataforma.startswith("IBKR"):
            # Cargar Paper
            sync_paper = cargar_sync_plataforma(plataforma, "Paper")
            if sync_paper:
                ibkr_paper_capital_var.set(sync_paper.get("capital", "-"))
                pos_paper = sync_paper.get("posiciones", "-")
                if isinstance(pos_paper, dict):
                    ibkr_paper_pos_var.set(str(len(pos_paper)))
                else:
                    ibkr_paper_pos_var.set(str(pos_paper) if pos_paper else "-")
                ibkr_paper_fecha_var.set(f"Sync: {sync_paper.get('fecha', '-')}")
            else:
                ibkr_paper_capital_var.set("-")
                ibkr_paper_pos_var.set("-")
                ibkr_paper_fecha_var.set("Sin datos")

        # Cargar Real (para todas las plataformas)
        sync_real = cargar_sync_plataforma(plataforma, "Real")
        if sync_real:
            ibkr_live_capital_var.set(sync_real.get("capital", "-"))
            pos_real = sync_real.get("posiciones", "-")
            if isinstance(pos_real, dict):
                ibkr_live_pos_var.set(str(len(pos_real)))
            else:
                ibkr_live_pos_var.set(str(pos_real) if pos_real else "-")
            ibkr_live_fecha_var.set(f"Sync: {sync_real.get('fecha', '-')}")

            # Balances por moneda (Cash)
            balances = sync_real.get("balances_por_moneda", {})
            if balances:
                bal_str = " / ".join([f"{m}: {v:,.2f}" for m, v in balances.items() if abs(v) > 0.01])
                ibkr_live_balances_var.set(f"Cash: {bal_str}")
            else:
                ibkr_live_balances_var.set("")
        else:
            ibkr_live_capital_var.set("-")
            ibkr_live_pos_var.set("-")
            ibkr_live_fecha_var.set("Sin datos")
            ibkr_live_balances_var.set("")

        # Cargar capital invertido
        capital_inv = calcular_capital_invertido(plataforma)
        if capital_inv:
            inv_str = " | ".join([f"{m}: {v:,.2f}" for m, v in capital_inv.items()])
            ibkr_live_invertido_var.set(f"Invertido: {inv_str}")
        else:
            ibkr_live_invertido_var.set("")

    def abrir_ventana_editar_capital():
        """Abre ventana para editar manualmente capital y posiciones."""
        from datetime import datetime

        plat_actual = plataforma_var.get()
        modo_edicion = "Real"  # Por defecto editar Real

        vent_edit = tk.Toplevel(ventana_hist)
        vent_edit.title(f"Editar Capital - {plat_actual}")
        vent_edit.geometry("400x350")
        vent_edit.resizable(False, False)
        vent_edit.transient(ventana_hist)
        vent_edit.grab_set()

        # Cargar datos actuales
        sync_actual = cargar_sync_plataforma(plat_actual, modo_edicion)

        # Frame principal
        frame_form = tk.Frame(vent_edit, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        tk.Label(frame_form, text=f"Plataforma: {plat_actual} ({modo_edicion})",
                 font=("Arial", 11, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15))

        # Capital total
        tk.Label(frame_form, text="Capital Total:", font=("Arial", 10)).grid(row=1, column=0, sticky="e", pady=5)
        entry_capital = tk.Entry(frame_form, width=20, font=("Arial", 10))
        entry_capital.grid(row=1, column=1, sticky="w", pady=5, padx=5)
        if sync_actual and sync_actual.get("capital"):
            entry_capital.insert(0, sync_actual.get("capital", ""))

        # Número de posiciones
        tk.Label(frame_form, text="Nº Posiciones:", font=("Arial", 10)).grid(row=2, column=0, sticky="e", pady=5)
        entry_posiciones = tk.Entry(frame_form, width=20, font=("Arial", 10))
        entry_posiciones.grid(row=2, column=1, sticky="w", pady=5, padx=5)
        if sync_actual:
            pos = sync_actual.get("posiciones", "")
            if isinstance(pos, dict):
                entry_posiciones.insert(0, str(len(pos)))
            else:
                entry_posiciones.insert(0, str(pos) if pos else "")

        # Separador
        tk.Label(frame_form, text="─" * 40, fg="gray").grid(row=3, column=0, columnspan=2, pady=10)

        # Cash por moneda
        tk.Label(frame_form, text="Cash por Moneda:", font=("Arial", 10, "bold")).grid(row=4, column=0, columnspan=2, pady=(0, 5))

        balances_actual = {}
        if sync_actual:
            balances_actual = sync_actual.get("balances_por_moneda", {})

        entries_moneda = {}
        monedas_principales = ["USD", "GBP", "EUR"]
        for i, moneda in enumerate(monedas_principales):
            tk.Label(frame_form, text=f"{moneda}:", font=("Arial", 10)).grid(row=5+i, column=0, sticky="e", pady=2)
            entry_mon = tk.Entry(frame_form, width=15, font=("Arial", 10))
            entry_mon.grid(row=5+i, column=1, sticky="w", pady=2, padx=5)
            if moneda in balances_actual:
                entry_mon.insert(0, f"{balances_actual[moneda]:.2f}")
            entries_moneda[moneda] = entry_mon

        def guardar_cambios():
            try:
                capital_raw = entry_capital.get().strip()
                posiciones = entry_posiciones.get().strip()

                # Parsear capital como número y formatear con coma de miles
                try:
                    # Remover símbolos de moneda y comas existentes
                    capital_limpio = capital_raw.replace("$", "").replace("£", "").replace("€", "").replace(",", "").strip()
                    capital_num = float(capital_limpio)
                    # Detectar símbolo de moneda original o usar $
                    if "£" in capital_raw:
                        simbolo = "£"
                    elif "€" in capital_raw:
                        simbolo = "€"
                    else:
                        simbolo = "$"
                    capital = f"{simbolo}{capital_num:,.2f}"
                except ValueError:
                    capital = capital_raw  # Si no es número, guardar como está

                # Validar posiciones como número
                try:
                    pos_num = int(posiciones) if posiciones else 0
                except ValueError:
                    pos_num = posiciones

                # Recopilar balances
                balances = {}
                for moneda, entry in entries_moneda.items():
                    val = entry.get().strip()
                    if val:
                        try:
                            balances[moneda] = float(val.replace(",", ""))
                        except ValueError:
                            pass

                fecha_sync = datetime.now().strftime("%Y-%m-%d %H:%M")

                # Guardar usando la función genérica
                guardar_sync_plataforma(plat_actual, modo_edicion, capital, pos_num, fecha_sync, balances)

                # Actualizar la visualización
                actualizar_datos_capital_plataforma(plat_actual)

                messagebox.showinfo("Guardado", f"Datos de {plat_actual} guardados correctamente.", parent=vent_edit)
                vent_edit.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}", parent=vent_edit)

        # Botones
        frame_btns = tk.Frame(frame_form)
        frame_btns.grid(row=8, column=0, columnspan=2, pady=20)

        tk.Button(frame_btns, text="Guardar", command=guardar_cambios,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold"), width=10).pack(side="left", padx=5)
        tk.Button(frame_btns, text="Cancelar", command=vent_edit.destroy,
                  bg="#dc3545", fg="white", font=("Arial", 10, "bold"), width=10).pack(side="left", padx=5)

    def mostrar_ocultar_frame_capital(*args):
        """Muestra el frame de Capital y Posiciones para todas las plataformas.
        Adapta la visualización según la plataforma (IBKR muestra Paper/Real, otras solo Real)."""
        plat = plataforma_var.get()

        # Siempre mostrar el frame
        frame_capital.pack(fill="x", padx=10, pady=5, after=frame_resumen)

        # Para IBKR: mostrar Paper y Real, botón Sync IBKR
        if plat.startswith("IBKR"):
            frame_paper.pack(side="left", padx=20)
            separador_paper_live.pack(side="left", padx=10)
            btn_sync_ibkr.pack(side="left", padx=2)
            ventana_hist.geometry("900x800")
        else:
            # Para otras plataformas: ocultar Paper, mostrar solo Real
            frame_paper.pack_forget()
            separador_paper_live.pack_forget()
            btn_sync_ibkr.pack_forget()
            ventana_hist.geometry("900x750")

        # Cargar datos de la plataforma actual
        actualizar_datos_capital_plataforma(plat)

    # Vincular al cambio de plataforma
    plataforma_var.trace_add("write", mostrar_ocultar_frame_capital)

    # Mostrar/ocultar inicialmente
    mostrar_ocultar_frame_capital()

    # Frame inferior - Botones (crear ANTES del historial y empaquetar con side="bottom")
    frame_botones = tk.Frame(ventana_hist, pady=10)
    frame_botones.pack(fill="x", padx=10, side="bottom")

    # Frame medio - Historial de operaciones
    frame_historial = tk.LabelFrame(ventana_hist, text="Historial de Operaciones", pady=5, padx=5)
    frame_historial.pack(fill="both", expand=True, padx=10, pady=5)

    # Frame de filtros
    frame_filtros_hist = tk.Frame(frame_historial)
    frame_filtros_hist.pack(fill="x", pady=(0, 5))

    tk.Label(frame_filtros_hist, text="Filtrar:", font=("Arial", 9)).pack(side="left", padx=(0, 5))

    tk.Label(frame_filtros_hist, text="Ticker:", font=("Arial", 9)).pack(side="left")
    filtro_ticker_var = tk.StringVar(value="Todos")
    combo_filtro_ticker = ttk.Combobox(frame_filtros_hist, textvariable=filtro_ticker_var,
                                        state="readonly", width=10)
    combo_filtro_ticker.pack(side="left", padx=(2, 10))

    tk.Label(frame_filtros_hist, text="Fecha:", font=("Arial", 9)).pack(side="left")
    filtro_fecha_var = tk.StringVar(value="Todos")
    combo_filtro_fecha = ttk.Combobox(frame_filtros_hist, textvariable=filtro_fecha_var,
                                       state="readonly", width=12)
    combo_filtro_fecha.pack(side="left", padx=2)

    # Labels para mostrar ganancia del ticker seleccionado
    tk.Label(frame_filtros_hist, text="|", font=("Arial", 10), fg="gray").pack(side="left", padx=(10, 5))

    # Checkbox para ocultar vendidas
    ocultar_vendidas_var = tk.BooleanVar(value=False)
    chk_ocultar_vendidas = tk.Checkbutton(frame_filtros_hist, text="Ocultar vendidas",
                                           variable=ocultar_vendidas_var, font=("Arial", 9))
    chk_ocultar_vendidas.pack(side="left", padx=(0, 10))

    tk.Label(frame_filtros_hist, text="|", font=("Arial", 10), fg="gray").pack(side="left", padx=(5, 5))
    lbl_realizada_filtro = tk.Label(frame_filtros_hist, text="Realizada: -", font=("Arial", 9))
    lbl_realizada_filtro.pack(side="left", padx=(0, 10))
    lbl_global_filtro = tk.Label(frame_filtros_hist, text="Global: -", font=("Arial", 9))
    lbl_global_filtro.pack(side="left")

    def calcular_ganancia_realizada_fecha(ops_todas, ticker_filtro, fecha_filtro):
        """
        Calcula ganancia realizada de ventas en una fecha específica.
        Usa TODAS las operaciones para conocer el costo base, pero solo cuenta
        la ganancia de ventas que coinciden con los filtros.
        """
        # Ordenar por fecha
        operaciones_ordenadas = sorted(ops_todas, key=lambda x: x.get("fecha", ""))

        # Diccionario de compras disponibles por ticker
        compras_por_ticker = {}
        ganancia_total = 0

        for op in operaciones_ordenadas:
            symbol = op.get("ticker_symbol")
            tipo = op.get("tipo", "").lower()
            cantidad = op.get("cantidad", 0)
            precio = op.get("precio", 0)
            fecha_op = op.get("fecha", "")

            # Aplicar filtro de ticker si existe
            if ticker_filtro != "Todos" and symbol != ticker_filtro:
                continue

            if symbol not in compras_por_ticker:
                compras_por_ticker[symbol] = []

            if tipo == "compra":
                compras_por_ticker[symbol].append([precio, cantidad])
                compras_por_ticker[symbol].sort(key=lambda x: x[0])

            elif tipo == "venta":
                cantidad_a_vender = cantidad
                precio_venta = precio

                for compra in compras_por_ticker.get(symbol, []):
                    if cantidad_a_vender <= 0:
                        break
                    if compra[1] > 0:
                        consumir = min(compra[1], cantidad_a_vender)
                        precio_compra = compra[0]

                        # Solo contar ganancia si la venta es de la fecha filtrada
                        if fecha_filtro == "Todos" or fecha_op == fecha_filtro:
                            ganancia_porcion = (precio_venta - precio_compra) * consumir
                            ganancia_total += ganancia_porcion

                        compra[1] -= consumir
                        cantidad_a_vender -= consumir

                compras_por_ticker[symbol] = [c for c in compras_por_ticker.get(symbol, []) if c[1] > 0]

        return ganancia_total

    def calcular_metricas_filtradas(ticker, fecha):
        """
        Calcula ganancia realizada y global según los filtros aplicados.
        - ticker="Todos" y fecha="Todos" → None, None
        - ticker específico y fecha="Todos" → métricas del ticker (todas las fechas)
        - ticker="Todos" y fecha específica → métricas de la fecha (todos los tickers)
        - ticker específico y fecha específica → métricas del ticker en esa fecha
        """
        if ticker == "Todos" and fecha == "Todos":
            return None, None

        ops_plataforma = obtener_operaciones_plataforma()

        # Operaciones filtradas (para mostrar)
        ops_filtradas = ops_plataforma
        if ticker != "Todos":
            ops_filtradas = [op for op in ops_filtradas if op.get("ticker_symbol") == ticker]
        if fecha != "Todos":
            ops_filtradas = [op for op in ops_filtradas if op.get("fecha") == fecha]

        if not ops_filtradas:
            return 0, 0

        # Calcular ganancia realizada usando todas las operaciones para el costo base
        ganancia_realizada = calcular_ganancia_realizada_fecha(ops_plataforma, ticker, fecha)

        # Para Global:
        # - Si fecha="Todos": ventas + valor_cartera - compras (del ticker)
        # - Si fecha específica: realizada + ganancia no realizada de compras del día
        if fecha == "Todos":
            # Comportamiento original para filtro solo por ticker
            total_compras = sum(op.get("precio", 0) * op.get("cantidad", 0)
                               for op in ops_filtradas if op.get("tipo", "").lower() == "compra")
            total_ventas = sum(op.get("precio", 0) * op.get("cantidad", 0)
                              for op in ops_filtradas if op.get("tipo", "").lower() == "venta")

            valor_cartera = 0
            cartera = calcular_cartera(ops_filtradas)
            if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
                try:
                    df_log = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
                    for tk, datos in cartera.items():
                        if datos.get("acciones", 0) > 0:
                            df_ticker = df_log[df_log['Ticker'] == tk].sort_values('Date')
                            if not df_ticker.empty:
                                ultimo_precio = df_ticker.iloc[-1]['Close']
                                if pd.notna(ultimo_precio):
                                    valor_cartera += datos["acciones"] * ultimo_precio
                except:
                    pass

            ganancia_global = (total_ventas + valor_cartera) - total_compras
        else:
            # Para fecha específica: realizada + ganancia no realizada de compras del día
            # Ganancia no realizada = (precio_cierre - precio_compra) * cantidad
            ganancia_no_realizada = 0
            compras_del_dia = [op for op in ops_filtradas if op.get("tipo", "").lower() == "compra"]

            if compras_del_dia and os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
                try:
                    df_log = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
                    for op in compras_del_dia:
                        tk = op.get("ticker_symbol")
                        precio_compra = op.get("precio", 0)
                        cantidad = op.get("cantidad", 0)

                        df_ticker = df_log[df_log['Ticker'] == tk].sort_values('Date')
                        if not df_ticker.empty:
                            precio_cierre = df_ticker.iloc[-1]['Close']
                            if pd.notna(precio_cierre):
                                ganancia_no_realizada += (precio_cierre - precio_compra) * cantidad
                except:
                    pass

            ganancia_global = ganancia_realizada + ganancia_no_realizada

        return ganancia_realizada, ganancia_global

    def actualizar_labels_ticker():
        """Actualiza los labels de ganancia según filtros de ticker y fecha"""
        ticker = filtro_ticker_var.get()
        fecha = filtro_fecha_var.get()
        realizada, global_val = calcular_metricas_filtradas(ticker, fecha)

        if realizada is None:
            lbl_realizada_filtro.config(text="Realizada: -", fg="black")
            lbl_global_filtro.config(text="Global: -", fg="black")
        else:
            color_r = "green" if realizada >= 0 else "red"
            color_g = "green" if global_val >= 0 else "red"
            lbl_realizada_filtro.config(text=f"Realizada: ${realizada:,.2f}", fg=color_r)
            lbl_global_filtro.config(text=f"Global: ${global_val:,.2f}", fg=color_g)

    # Scrollbars
    scrollbar_y = tk.Scrollbar(frame_historial, orient="vertical")
    scrollbar_x = tk.Scrollbar(frame_historial, orient="horizontal")

    # Treeview para historial
    cols_hist = ("Fecha", "Symbol", "Tipo", "Precio", "Cantidad", "Total", "Saldo")
    tree_hist = ttk.Treeview(frame_historial, columns=cols_hist, show="headings",
                              selectmode="extended",
                              yscrollcommand=scrollbar_y.set,
                              xscrollcommand=scrollbar_x.set)

    scrollbar_y.config(command=tree_hist.yview)
    scrollbar_x.config(command=tree_hist.xview)

    anchos = {"Fecha": 100, "Symbol": 80, "Tipo": 70, "Precio": 90, "Cantidad": 70, "Total": 100, "Saldo": 60}
    for col in cols_hist:
        tree_hist.heading(col, text=col)
        tree_hist.column(col, width=anchos.get(col, 80), anchor="center")

    tree_hist.tag_configure("compra", foreground="#008000")
    tree_hist.tag_configure("venta", foreground="#cc0000")
    tree_hist.tag_configure("vendida", foreground="#999999")  # Gris para compras ya vendidas

    def calcular_saldos_compras(operaciones_plataforma):
        """
        Calcula el saldo restante de cada compra después de aplicar ventas.
        Las ventas se aplican a compras realizadas ANTES de la fecha de venta,
        y entre esas, primero a las de menor precio.
        Retorna dict: {indice_operacion: saldo_restante}
        """
        # Agrupar por ticker
        compras_por_ticker = {}
        ventas_por_ticker = {}

        for i, op in enumerate(operaciones_plataforma):
            ticker = op.get("ticker_symbol", "")
            tipo = op.get("tipo", "").lower()
            cantidad = op.get("cantidad", 0)
            precio = op.get("precio", 0)
            fecha = op.get("fecha", "")

            if tipo == "compra":
                if ticker not in compras_por_ticker:
                    compras_por_ticker[ticker] = []
                compras_por_ticker[ticker].append({
                    "indice": i,
                    "precio": precio,
                    "cantidad": cantidad,
                    "fecha": fecha,
                    "saldo": cantidad  # Inicialmente el saldo es la cantidad comprada
                })
            elif tipo == "venta":
                if ticker not in ventas_por_ticker:
                    ventas_por_ticker[ticker] = []
                ventas_por_ticker[ticker].append({
                    "cantidad": cantidad,
                    "fecha": fecha
                })

        # Procesar ventas en orden cronológico
        for ticker, ventas in ventas_por_ticker.items():
            if ticker not in compras_por_ticker:
                continue

            # Ordenar ventas por fecha
            ventas_ordenadas = sorted(ventas, key=lambda x: x["fecha"])
            compras = compras_por_ticker[ticker]

            for venta in ventas_ordenadas:
                fecha_venta = venta["fecha"]
                cantidad_vender = venta["cantidad"]

                # Filtrar compras anteriores a la fecha de venta y ordenar por precio
                compras_elegibles = [c for c in compras if c["fecha"] <= fecha_venta and c["saldo"] > 0]
                compras_por_precio = sorted(compras_elegibles, key=lambda x: x["precio"])

                # Aplicar venta a compras de menor precio primero
                for compra in compras_por_precio:
                    if cantidad_vender <= 0:
                        break
                    reducir = min(compra["saldo"], cantidad_vender)
                    compra["saldo"] -= reducir
                    cantidad_vender -= reducir

        # Crear diccionario de saldos
        saldos = {}
        for ticker, compras in compras_por_ticker.items():
            for compra in compras:
                saldos[compra["indice"]] = compra["saldo"]

        return saldos

    def actualizar_filtros_hist(actualizar_fechas=True):
        """Actualiza las opciones de los combos de filtro para la plataforma seleccionada"""
        ops_plataforma = obtener_operaciones_plataforma()
        tickers = sorted(set(op.get("ticker_symbol", "") for op in ops_plataforma))
        combo_filtro_ticker["values"] = ["Todos"] + tickers

        if actualizar_fechas:
            # Filtrar fechas según el ticker seleccionado
            ticker_seleccionado = filtro_ticker_var.get()
            if ticker_seleccionado == "Todos":
                ops_para_fechas = ops_plataforma
            else:
                ops_para_fechas = [op for op in ops_plataforma if op.get("ticker_symbol") == ticker_seleccionado]

            fechas = sorted(set(op.get("fecha", "") for op in ops_para_fechas), reverse=True)
            combo_filtro_fecha["values"] = ["Todos"] + fechas

            # Si la fecha actual no está en la lista, resetear a "Todos"
            if filtro_fecha_var.get() not in ["Todos"] + fechas:
                filtro_fecha_var.set("Todos")

    def actualizar_historial():
        """Actualiza la vista del historial de la plataforma seleccionada"""
        nonlocal operaciones
        # Recargar datos completos
        datos_hist = cargar_historial_operaciones_completo()
        operaciones = datos_hist.get("operaciones", [])

        actualizar_filtros_hist()

        for item in tree_hist.get_children():
            tree_hist.delete(item)

        filtro_t = filtro_ticker_var.get()
        filtro_f = filtro_fecha_var.get()
        ocultar_vendidas = ocultar_vendidas_var.get()

        # Obtener operaciones de la plataforma seleccionada
        ops_plataforma = obtener_operaciones_plataforma()

        # Calcular saldos de compras (FIFO por precio más bajo)
        saldos = calcular_saldos_compras(ops_plataforma)

        # Ordenar por symbol alfabéticamente y guardar índice original
        ops_con_indice = [(i, op) for i, op in enumerate(ops_plataforma)]
        ops_ordenadas = sorted(ops_con_indice, key=lambda x: x[1].get("ticker_symbol", "").upper())

        for idx_original, op in ops_ordenadas:
            tipo = op.get("tipo", "").lower()

            # Determinar saldo (solo para compras)
            saldo = saldos.get(idx_original, None) if tipo == "compra" else None
            es_vendida = tipo == "compra" and saldo == 0

            # Filtro: ocultar compras ya vendidas y también las ventas
            if ocultar_vendidas and (es_vendida or tipo == "venta"):
                continue

            # Aplicar filtros de ticker y fecha
            if filtro_t != "Todos" and op.get("ticker_symbol", "") != filtro_t:
                continue
            if filtro_f != "Todos" and op.get("fecha", "") != filtro_f:
                continue

            precio = op.get("precio", 0)
            cantidad = op.get("cantidad", 0)
            total = precio * cantidad

            # Determinar texto de saldo
            saldo_texto = str(saldo) if saldo is not None else "-"

            # Determinar tags
            if es_vendida:
                tags = ("vendida",)
            else:
                tags = (tipo,)

            tree_hist.insert("", "end", values=(
                op.get("fecha", ""),
                op.get("ticker_symbol", ""),
                tipo.capitalize(),
                f"${precio:.2f}",
                cantidad,
                f"${total:.2f}",
                saldo_texto
            ), tags=tags)

    def on_filtro_hist_change(*args):
        actualizar_historial()
        actualizar_labels_ticker()

    def on_ticker_change(*args):
        """Cuando cambia el ticker, actualizar lista de fechas y refrescar vista"""
        actualizar_filtros_hist(actualizar_fechas=True)
        actualizar_historial()
        actualizar_labels_ticker()

    combo_filtro_ticker.bind("<<ComboboxSelected>>", on_ticker_change)
    combo_filtro_fecha.bind("<<ComboboxSelected>>", on_filtro_hist_change)
    ocultar_vendidas_var.trace_add("write", on_filtro_hist_change)

    def on_plataforma_change(*args):
        """Actualiza todas las vistas cuando cambia la plataforma seleccionada"""
        # Actualizar etiquetas de moneda y descripción
        plat = plataforma_var.get()
        moneda = config_plataformas.get(plat, {}).get("moneda", "USD")
        desc = config_plataformas.get(plat, {}).get("descripcion", "")
        lbl_moneda.config(text=f"Moneda: {moneda}")
        lbl_desc.config(text=desc)
        # Resetear filtros
        filtro_ticker_var.set("Todos")
        filtro_fecha_var.set("Todos")
        # Actualizar todas las vistas
        actualizar_cartera()
        actualizar_resumen()
        actualizar_historial()
        actualizar_labels_ticker()

    combo_plataforma.bind("<<ComboboxSelected>>", on_plataforma_change)

    def on_modo_change(event=None):
        """Actualiza las vistas cuando cambia el modo"""
        actualizar_cartera()
        actualizar_resumen()
        actualizar_historial()
        actualizar_labels_ticker()

    combo_modo.bind("<<ComboboxSelected>>", on_modo_change)

    actualizar_historial()

    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")
    tree_hist.pack(fill="both", expand=True)

    def agregar_operacion():
        """Abre ventana para agregar nueva operación"""
        ventana_add = tk.Toplevel(ventana_hist)
        ventana_add.title("Registrar Operación")
        ventana_add.geometry("350x330")
        ventana_add.transient(ventana_hist)
        ventana_add.grab_set()

        frame_form = tk.Frame(ventana_add, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        # Fecha
        tk.Label(frame_form, text="Fecha (YYYY-MM-DD):").grid(row=0, column=0, sticky="w", pady=5)
        entry_fecha = tk.Entry(frame_form, width=20)
        entry_fecha.insert(0, datetime.now().strftime("%Y-%m-%d"))
        entry_fecha.grid(row=0, column=1, pady=5)

        # Symbol (auto-mayúsculas)
        tk.Label(frame_form, text="Symbol:").grid(row=1, column=0, sticky="w", pady=5)
        symbol_var = tk.StringVar()
        symbol_var.trace_add("write", lambda *args: symbol_var.set(symbol_var.get().upper()))
        entry_symbol = tk.Entry(frame_form, width=20, textvariable=symbol_var)
        entry_symbol.grid(row=1, column=1, pady=5)

        # Tipo
        tk.Label(frame_form, text="Tipo:").grid(row=2, column=0, sticky="w", pady=5)
        tipo_var = tk.StringVar(value="compra")
        frame_tipo = tk.Frame(frame_form)
        frame_tipo.grid(row=2, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_tipo, text="Compra", variable=tipo_var, value="compra").pack(side="left")
        tk.Radiobutton(frame_tipo, text="Venta", variable=tipo_var, value="venta").pack(side="left")

        # Modo (Paper/Real) - default según plataforma: TYBA=Real, resto=Paper
        tk.Label(frame_form, text="Modo:").grid(row=3, column=0, sticky="w", pady=5)
        plat_actual = plataforma_var.get()
        modo_default = "real" if plat_actual == "TYBA" else "paper"
        modo_inicial = modo_var.get().lower() if modo_var.get() not in ["Todos", ""] else modo_default
        modo_op_var = tk.StringVar(value=modo_inicial)
        frame_modo = tk.Frame(frame_form)
        frame_modo.grid(row=3, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_modo, text="Paper", variable=modo_op_var, value="paper").pack(side="left")
        tk.Radiobutton(frame_modo, text="Real", variable=modo_op_var, value="real").pack(side="left")

        # Precio
        tk.Label(frame_form, text="Precio:").grid(row=4, column=0, sticky="w", pady=5)
        entry_precio = tk.Entry(frame_form, width=20)
        entry_precio.grid(row=4, column=1, pady=5)

        # Cantidad
        tk.Label(frame_form, text="Cantidad:").grid(row=5, column=0, sticky="w", pady=5)
        entry_cantidad = tk.Entry(frame_form, width=20)
        entry_cantidad.grid(row=5, column=1, pady=5)

        def guardar():
            fecha = entry_fecha.get().strip()
            symbol = entry_symbol.get().strip().upper()
            tipo = tipo_var.get()

            if not fecha or not symbol:
                messagebox.showwarning("Campos requeridos", "Completa fecha y symbol", parent=ventana_add)
                return

            # Validar que el ticker exista en la plataforma y modo seleccionados
            plataforma_actual = plataforma_var.get()
            modo_actual = modo_op_var.get().capitalize()  # "paper" -> "Paper"
            tickers_validos = obtener_tickers_plataforma(plataforma_actual, modo_actual)
            if symbol not in tickers_validos:
                messagebox.showerror("Ticker inválido",
                    f"'{symbol}' no es un ticker válido para {plataforma_actual} ({modo_actual}).\n\n"
                    f"Tickers disponibles:\n{', '.join(sorted(tickers_validos)) if tickers_validos else '(ninguno)'}",
                    parent=ventana_add)
                entry_symbol.focus_set()
                entry_symbol.select_range(0, tk.END)
                return

            try:
                precio = float(entry_precio.get().strip().replace(",", "."))
                cantidad = int(entry_cantidad.get().strip())
            except ValueError:
                messagebox.showerror("Error", "Precio y cantidad deben ser numéricos")
                return

            if cantidad <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0", parent=ventana_add)
                return

            # Validar precio vs cantidad (detectar si están invertidos)
            # Obtener precio de cierre más reciente del ticker
            precio_cierre_ref = None
            try:
                if AUTO_UPDATE_LOG_PORTABLE.exists():
                    import pandas as pd
                    df_log = pd.read_csv(AUTO_UPDATE_LOG_PORTABLE)
                    df_ticker = df_log[df_log['Symbol'] == symbol]
                    if not df_ticker.empty:
                        precio_cierre_ref = df_ticker.iloc[-1]['Close']
            except:
                pass

            # Verificar si precio y cantidad podrían estar invertidos
            advertencias = []

            if precio_cierre_ref:
                # Si el precio difiere más de 50% del cierre, advertir
                diferencia_pct = abs(precio - precio_cierre_ref) / precio_cierre_ref * 100
                if diferencia_pct > 50:
                    advertencias.append(
                        f"El precio ingresado (${precio:.2f}) difiere {diferencia_pct:.0f}% "
                        f"del último cierre (${precio_cierre_ref:.2f})."
                    )

            # Si cantidad > precio, probablemente están invertidos
            if cantidad > precio and precio < 100:
                advertencias.append(
                    f"¿Estás seguro? Cantidad ({cantidad}) > Precio (${precio:.2f}).\n"
                    f"Podría ser: Precio=${cantidad:.2f}, Cantidad={int(precio)}"
                )

            # Validar límite de acciones según parámetros (solo para compras)
            if tipo == "compra":
                limite_acciones = 10  # Default
                try:
                    datos_params, _ = cargar_parametros_activos()
                    if datos_params:
                        # Buscar el ticker en cualquier slot para obtener limite_valor
                        for slot_id in ["1", "2", "3", "4", "5", "6"]:
                            params_slot = obtener_parametros_slot(datos_params, slot_id)
                            for p in params_slot:
                                if p.get("ticker_symbol") == symbol:
                                    limite_acciones = int(p.get("limite_valor", 10))
                                    break
                            if limite_acciones != 10:
                                break
                except:
                    pass

                # Obtener acciones actuales en cartera
                ops_plataforma = obtener_operaciones_plataforma()
                cartera = calcular_cartera(ops_plataforma)
                acciones_actuales = cartera.get(symbol, {}).get("acciones", 0)
                total_despues = acciones_actuales + cantidad

                if total_despues > limite_acciones:
                    advertencias.append(
                        f"Excedes el límite de {limite_acciones} acciones para {symbol}.\n"
                        f"Actualmente tienes: {acciones_actuales}\n"
                        f"Después de esta compra: {total_despues}"
                    )

            if advertencias:
                msg = "⚠️ ADVERTENCIA:\n\n" + "\n\n".join(advertencias)
                msg += f"\n\n¿Deseas continuar con:\n{tipo.upper()} {cantidad} {symbol} @ ${precio:.2f}?"
                if not messagebox.askyesno("Verificar datos", msg, parent=ventana_add, icon="warning"):
                    entry_precio.focus_set()
                    entry_precio.select_range(0, tk.END)
                    return

            # Validar que no se venda más de lo que se tiene en la plataforma
            if tipo == "venta":
                ops_plataforma = obtener_operaciones_plataforma()
                cartera = calcular_cartera(ops_plataforma)
                acciones_disponibles = cartera.get(symbol, {}).get("acciones", 0)
                if cantidad > acciones_disponibles:
                    messagebox.showerror("Error",
                        f"No puedes vender {cantidad} acciones de {symbol}.\n"
                        f"Solo tienes {acciones_disponibles} en cartera ({plataforma_var.get()}).")
                    return

            nueva_op = {
                "fecha": fecha,
                "ticker_symbol": symbol,
                "tipo": tipo,
                "precio": precio,
                "cantidad": cantidad,
                "plataforma": plataforma_var.get(),
                "modo": modo_op_var.get()
            }

            operaciones.append(nueva_op)
            guardar_historial_operaciones(operaciones)
            actualizar_historial()
            actualizar_cartera()
            actualizar_resumen()
            messagebox.showinfo("Guardado", f"Operación registrada:\n{tipo.upper()} {cantidad} {symbol} @ ${precio:.2f}", parent=ventana_add)

            # Limpiar campos para registrar otra operación
            entry_symbol.delete(0, tk.END)
            entry_precio.delete(0, tk.END)
            entry_cantidad.delete(0, tk.END)
            entry_symbol.focus_set()

        frame_botones_form = tk.Frame(frame_form)
        frame_botones_form.grid(row=6, column=0, columnspan=2, pady=20)

        tk.Button(frame_botones_form, text="Guardar", command=guardar,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=10)

        tk.Button(frame_botones_form, text="Cerrar", command=ventana_add.destroy,
                  bg="#6c757d", fg="white", font=("Arial", 10)).pack(side="left", padx=10)

    def editar_seleccionado():
        """Edita la operación seleccionada"""
        seleccionados = tree_hist.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona una operación para editar")
            return

        if len(seleccionados) > 1:
            messagebox.showwarning("Múltiple selección", "Selecciona solo una operación para editar")
            return

        # Obtener valores actuales
        item = seleccionados[0]
        valores = tree_hist.item(item, "values")
        fecha_actual = valores[0]
        symbol_actual = valores[1]
        tipo_actual = valores[2].lower()
        precio_actual = float(valores[3].replace("$", ""))
        cantidad_actual = int(valores[4])

        # Buscar índice en operaciones (filtrar por plataforma actual)
        plat_actual = plataforma_var.get()
        indice_editar = None
        for i, op in enumerate(operaciones):
            if (op.get("fecha") == fecha_actual and
                op.get("ticker_symbol") == symbol_actual and
                op.get("tipo") == tipo_actual and
                abs(op.get("precio", 0) - precio_actual) < 0.01 and
                op.get("cantidad") == cantidad_actual and
                op.get("plataforma", "TYBA") == plat_actual):
                indice_editar = i
                break

        if indice_editar is None:
            messagebox.showerror("Error", "No se encontró la operación")
            return

        # Obtener modo actual de la operación (default según plataforma: TYBA=Real, resto=Paper)
        op_plataforma = operaciones[indice_editar].get("plataforma", "TYBA")
        modo_default = "Real" if op_plataforma == "TYBA" else "Paper"
        modo_actual = operaciones[indice_editar].get("modo", modo_default)

        # Ventana de edición
        ventana_edit = tk.Toplevel(ventana_hist)
        ventana_edit.title("Editar Operación")
        ventana_edit.geometry("350x330")
        ventana_edit.transient(ventana_hist)
        ventana_edit.grab_set()

        frame_form = tk.Frame(ventana_edit, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        # Fecha
        tk.Label(frame_form, text="Fecha (YYYY-MM-DD):").grid(row=0, column=0, sticky="w", pady=5)
        entry_fecha = tk.Entry(frame_form, width=20)
        entry_fecha.insert(0, fecha_actual)
        entry_fecha.grid(row=0, column=1, pady=5)

        # Symbol
        tk.Label(frame_form, text="Symbol:").grid(row=1, column=0, sticky="w", pady=5)
        entry_symbol = tk.Entry(frame_form, width=20)
        entry_symbol.insert(0, symbol_actual)
        entry_symbol.grid(row=1, column=1, pady=5)

        # Tipo
        tk.Label(frame_form, text="Tipo:").grid(row=2, column=0, sticky="w", pady=5)
        tipo_var_edit = tk.StringVar(value=tipo_actual)
        frame_tipo = tk.Frame(frame_form)
        frame_tipo.grid(row=2, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_tipo, text="Compra", variable=tipo_var_edit, value="compra").pack(side="left")
        tk.Radiobutton(frame_tipo, text="Venta", variable=tipo_var_edit, value="venta").pack(side="left")

        # Modo (Paper/Real)
        tk.Label(frame_form, text="Modo:").grid(row=3, column=0, sticky="w", pady=5)
        modo_edit_inicial = modo_actual.lower() if modo_actual else "real"
        modo_edit_var = tk.StringVar(value=modo_edit_inicial)
        frame_modo = tk.Frame(frame_form)
        frame_modo.grid(row=3, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_modo, text="Paper", variable=modo_edit_var, value="paper").pack(side="left")
        tk.Radiobutton(frame_modo, text="Real", variable=modo_edit_var, value="real").pack(side="left")

        # Precio
        tk.Label(frame_form, text="Precio:").grid(row=4, column=0, sticky="w", pady=5)
        entry_precio = tk.Entry(frame_form, width=20)
        entry_precio.insert(0, str(precio_actual))
        entry_precio.grid(row=4, column=1, pady=5)

        # Cantidad
        tk.Label(frame_form, text="Cantidad:").grid(row=5, column=0, sticky="w", pady=5)
        entry_cantidad = tk.Entry(frame_form, width=20)
        entry_cantidad.insert(0, str(cantidad_actual))
        entry_cantidad.grid(row=5, column=1, pady=5)

        def guardar_edicion():
            fecha = entry_fecha.get().strip()
            symbol = entry_symbol.get().strip().upper()
            tipo = tipo_var_edit.get()

            if not fecha or not symbol:
                messagebox.showwarning("Campos requeridos", "Completa fecha y symbol")
                return

            try:
                precio = float(entry_precio.get().strip().replace(",", "."))
                cantidad = int(entry_cantidad.get().strip())
            except ValueError:
                messagebox.showerror("Error", "Precio y cantidad deben ser numéricos")
                return

            if cantidad <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            # Actualizar operación (preservar plataforma original)
            plataforma_original = operaciones[indice_editar].get("plataforma", plataforma_var.get())
            operaciones[indice_editar] = {
                "fecha": fecha,
                "ticker_symbol": symbol,
                "tipo": tipo,
                "precio": precio,
                "cantidad": cantidad,
                "plataforma": plataforma_original,
                "modo": modo_edit_var.get()
            }

            guardar_historial_operaciones(operaciones)
            actualizar_historial()
            actualizar_cartera()
            actualizar_resumen()
            messagebox.showinfo("Guardado", f"Operación actualizada:\n{tipo.upper()} {cantidad} {symbol} @ ${precio:.2f}")
            ventana_edit.destroy()

        tk.Button(frame_form, text="Guardar Cambios", command=guardar_edicion,
                  bg="#ffc107", fg="black", font=("Arial", 10, "bold")).grid(row=6, column=0, columnspan=2, pady=20)

    def eliminar_seleccionados():
        """Elimina las operaciones seleccionadas"""
        seleccionados = tree_hist.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona operaciones para eliminar")
            return

        if not messagebox.askyesno("Confirmar", f"¿Eliminar {len(seleccionados)} operación(es)?"):
            return

        # Obtener índices a eliminar (filtrar por plataforma actual)
        plat_actual = plataforma_var.get()
        indices_eliminar = []
        for item in seleccionados:
            valores = tree_hist.item(item, "values")
            fecha = valores[0]
            symbol = valores[1]
            tipo = valores[2].lower()
            precio = float(valores[3].replace("$", ""))
            cantidad = int(valores[4])

            # Buscar en operaciones
            for i, op in enumerate(operaciones):
                if (op.get("fecha") == fecha and
                    op.get("ticker_symbol") == symbol and
                    op.get("tipo") == tipo and
                    abs(op.get("precio", 0) - precio) < 0.01 and
                    op.get("cantidad") == cantidad and
                    op.get("plataforma", "TYBA") == plat_actual):
                    indices_eliminar.append(i)
                    break

        # Eliminar en orden inverso para no afectar índices
        for i in sorted(indices_eliminar, reverse=True):
            operaciones.pop(i)

        guardar_historial_operaciones(operaciones)
        actualizar_historial()
        actualizar_cartera()
        actualizar_resumen()
        messagebox.showinfo("Eliminado", f"Se eliminaron {len(indices_eliminar)} operación(es)")

    def graficar_operaciones():
        """Grafica las operaciones realizadas: cuadrados verdes=compras, triángulos rojos=ventas"""
        ops_plataforma = obtener_operaciones_plataforma()
        if not ops_plataforma:
            messagebox.showinfo("Sin datos", f"No hay operaciones en {plataforma_var.get()} para graficar.")
            return

        # Obtener tickers únicos de la plataforma seleccionada
        tickers_unicos = sorted(set(op.get("ticker_symbol", "") for op in ops_plataforma))

        if not tickers_unicos:
            return

        # Ventana para seleccionar ticker
        ventana_graf = tk.Toplevel(ventana_hist)
        ventana_graf.title("Graficar Operaciones")
        ventana_graf.geometry("800x600")
        ventana_graf.resizable(True, True)
        ventana_graf.minsize(500, 400)

        # Frame para controles
        frame_controles = tk.Frame(ventana_graf)
        frame_controles.pack(pady=5)

        tk.Label(frame_controles, text="Ticker:", font=("Arial", 10)).pack(side="left", padx=(0, 5))
        ticker_var = tk.StringVar(value=tickers_unicos[0] if tickers_unicos else "")
        combo_ticker = ttk.Combobox(frame_controles, textvariable=ticker_var, values=tickers_unicos, state="readonly", width=12)
        combo_ticker.pack(side="left", padx=5)

        tk.Label(frame_controles, text="|", font=("Arial", 10), fg="gray").pack(side="left", padx=10)

        tk.Label(frame_controles, text="Rango:", font=("Arial", 10)).pack(side="left", padx=(0, 5))
        rango_hist_var = tk.StringVar(value="Completo")
        combo_rango_hist = ttk.Combobox(frame_controles, textvariable=rango_hist_var,
                                        values=["Completo", "30 días"], state="readonly", width=10)
        combo_rango_hist.pack(side="left", padx=5)

        # Frame para el gráfico
        frame_grafico = tk.Frame(ventana_graf)
        frame_grafico.pack(fill="both", expand=True, padx=10, pady=10)

        def actualizar_grafico(*args):
            # Limpiar frame
            for widget in frame_grafico.winfo_children():
                widget.destroy()

            ticker_sel = ticker_var.get()
            if not ticker_sel:
                return

            # Filtrar operaciones del ticker (de la plataforma seleccionada)
            ops_ticker = [op for op in ops_plataforma if op.get("ticker_symbol") == ticker_sel]

            if not ops_ticker:
                tk.Label(frame_grafico, text="Sin operaciones para este ticker").pack()
                return

            # Separar compras y ventas
            compras = [(op.get("fecha"), op.get("precio", 0)) for op in ops_ticker if op.get("tipo", "").lower() == "compra"]
            ventas = [(op.get("fecha"), op.get("precio", 0)) for op in ops_ticker if op.get("tipo", "").lower() == "venta"]

            # Cargar precios de cierre del log
            precios_cierre = []
            if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
                try:
                    df_log = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
                    df_ticker = df_log[df_log['Ticker'] == ticker_sel].sort_values('Date')
                    if not df_ticker.empty:
                        precios_cierre = [(row['Date'], row['Close']) for _, row in df_ticker.iterrows()]
                except Exception as e:
                    print(f"[WARN] Error cargando precios: {e}")

            # Filtrar por rango de fechas si está seleccionado "30 días"
            if rango_hist_var.get() == "30 días":
                from datetime import timedelta
                fecha_limite = datetime.now() - timedelta(days=30)
                # Filtrar precios de cierre
                precios_cierre = [(f, p) for f, p in precios_cierre if f >= fecha_limite]
                # Filtrar operaciones
                compras = [(f, p) for f, p in compras if datetime.strptime(f, "%Y-%m-%d") >= fecha_limite]
                ventas = [(f, p) for f, p in ventas if datetime.strptime(f, "%Y-%m-%d") >= fecha_limite]

            # Crear figura
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.subplots_adjust(left=0.06, right=0.98, bottom=0.12, top=0.94)

            # Recopilar todas las fechas y precios
            todas_fechas = []
            todos_precios = []

            # Graficar línea de precios de cierre (azul, sin puntos)
            if precios_cierre:
                fechas_p = [f for f, _ in precios_cierre]
                precios_p = [p for _, p in precios_cierre]
                ax.plot(fechas_p, precios_p, color='blue', linewidth=1.5, label='Precio Cierre', zorder=1)
                todas_fechas.extend(fechas_p)
                todos_precios.extend(precios_p)

            # Graficar compras (cuadrados verdes)
            if compras:
                fechas_c = [datetime.strptime(f, "%Y-%m-%d") for f, _ in compras]
                precios_c = [p for _, p in compras]
                ax.scatter(fechas_c, precios_c, marker='s', s=35, c='green', label='Compras', zorder=5)
                todas_fechas.extend(fechas_c)
                todos_precios.extend(precios_c)

            # Graficar ventas (triángulos rojos)
            if ventas:
                fechas_v = [datetime.strptime(f, "%Y-%m-%d") for f, _ in ventas]
                precios_v = [p for _, p in ventas]
                ax.scatter(fechas_v, precios_v, marker='^', s=35, c='red', label='Ventas', zorder=5)
                todas_fechas.extend(fechas_v)
                todos_precios.extend(precios_v)

            # Ajustar límites del eje X (margen de 7 días)
            if todas_fechas:
                from datetime import timedelta
                fecha_min = min(todas_fechas) - timedelta(days=7)
                fecha_max = max(todas_fechas) + timedelta(days=7)
                ax.set_xlim(fecha_min, fecha_max)

            # Ajustar límites del eje Y (margen del 5%)
            if todos_precios:
                precio_min = min(todos_precios)
                precio_max = max(todos_precios)
                margen = (precio_max - precio_min) * 0.1 if precio_max != precio_min else precio_min * 0.05
                ax.set_ylim(precio_min - margen, precio_max + margen)

            ax.set_title(f"Operaciones - {ticker_sel}")
            ax.set_xlabel("Fecha")
            ax.set_ylabel("Precio ($)")
            ax.legend()
            ax.grid(True, alpha=0.3)

            # Formato de fechas: dd/mm/yy
            import matplotlib.dates as mdates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%y'))
            fig.autofmt_xdate(rotation=45)

            # Mostrar en tkinter
            canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

            plt.close(fig)

        combo_ticker.bind("<<ComboboxSelected>>", actualizar_grafico)
        combo_rango_hist.bind("<<ComboboxSelected>>", actualizar_grafico)
        actualizar_grafico()  # Mostrar gráfico inicial

    tk.Button(frame_botones, text="Registrar Operación", command=agregar_operacion,
              bg="#007bff", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Editar", command=editar_seleccionado,
              bg="#ffc107", fg="black", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Eliminar seleccionadas", command=eliminar_seleccionados,
              bg="#ff6b6b", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Graficar", command=graficar_operaciones,
              bg="#6f42c1", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    def exportar_excel_historial():
        """Exporta el historial de operaciones a Excel (de la plataforma seleccionada)"""
        ops_plataforma = obtener_operaciones_plataforma()
        if not ops_plataforma:
            messagebox.showinfo("Sin datos", f"No hay operaciones en {plataforma_var.get()} para exportar.")
            return

        from tkinter import filedialog
        plat_nombre = plataforma_var.get().replace("-", "_")
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"historial_operaciones_{plat_nombre}.xlsx",
            title="Guardar historial como Excel"
        )
        if not ruta_archivo:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()

            # Hoja 1: Historial de operaciones
            ws = wb.active
            ws.title = "Operaciones"

            encabezados = ["Fecha", "Symbol", "Tipo", "Precio", "Cantidad", "Total"]
            for col_idx, enc in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col_idx, value=enc)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            ops_ordenadas = sorted(ops_plataforma, key=lambda x: x.get("ticker_symbol", "").upper())
            for row_idx, op in enumerate(ops_ordenadas, 2):
                precio = op.get("precio", 0)
                cantidad = op.get("cantidad", 0)
                ws.cell(row=row_idx, column=1, value=op.get("fecha", ""))
                ws.cell(row=row_idx, column=2, value=op.get("ticker_symbol", ""))
                tipo = op.get("tipo", "")
                cell_tipo = ws.cell(row=row_idx, column=3, value=tipo.capitalize())
                if tipo == "compra":
                    cell_tipo.font = Font(color="008000")
                else:
                    cell_tipo.font = Font(color="FF0000")
                ws.cell(row=row_idx, column=4, value=precio)
                ws.cell(row=row_idx, column=5, value=cantidad)
                ws.cell(row=row_idx, column=6, value=round(precio * cantidad, 2))

            # Formato de columnas numéricas
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = '$#,##0.00'
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
                for cell in row:
                    cell.number_format = '$#,##0.00'

            # Ajustar anchos
            for col_idx, enc in enumerate(encabezados, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = max(len(enc) + 4, 12)

            # Hoja 2: Resumen cartera (de la plataforma seleccionada)
            ws2 = wb.create_sheet("Cartera")
            enc_cartera = ["Symbol", "Acciones", "P. Prom. Compra", "Capital Invertido"]
            for col_idx, enc in enumerate(enc_cartera, 1):
                cell = ws2.cell(row=1, column=col_idx, value=enc)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            cartera = calcular_cartera(ops_plataforma)
            row_idx = 2
            for symbol, datos in sorted(cartera.items(), key=lambda x: x[0].upper()):
                if datos["acciones"] > 0:
                    ws2.cell(row=row_idx, column=1, value=symbol)
                    ws2.cell(row=row_idx, column=2, value=datos["acciones"])
                    ws2.cell(row=row_idx, column=3, value=round(datos["precio_promedio_compra"], 2))
                    ws2.cell(row=row_idx, column=4, value=round(datos["capital_invertido"], 2))
                    row_idx += 1

            for col_idx, enc in enumerate(enc_cartera, 1):
                ws2.column_dimensions[chr(64 + col_idx)].width = max(len(enc) + 4, 14)

            wb.save(ruta_archivo)
            messagebox.showinfo("Exportado", f"Historial exportado a:\n{ruta_archivo}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar:\n{str(e)}")

    tk.Button(frame_botones, text="Exportar Excel", command=exportar_excel_historial,
              bg="#17a2b8", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_hist.destroy).pack(side="right", padx=5)


def filtrar_parametros_por_fecha(parametros, fecha_objetivo):
    """
    Filtra parámetros que estén vigentes para una fecha específica.
    Un parámetro está vigente si:
    - fecha_inicio es None O fecha_inicio <= fecha_objetivo
    - fecha_fin es None O fecha_fin >= fecha_objetivo

    Args:
        parametros: Lista de parámetros
        fecha_objetivo: String en formato YYYY-MM-DD o objeto datetime

    Returns:
        Lista de parámetros vigentes para esa fecha
    """
    if not fecha_objetivo:
        return parametros

    # Convertir fecha_objetivo a string si es datetime
    if hasattr(fecha_objetivo, 'strftime'):
        fecha_str = fecha_objetivo.strftime("%Y-%m-%d")
    else:
        fecha_str = str(fecha_objetivo)[:10]  # Tomar solo YYYY-MM-DD

    parametros_vigentes = []
    for param in parametros:
        fecha_inicio = param.get("fecha_inicio")
        fecha_fin = param.get("fecha_fin")

        # Verificar fecha inicio
        if fecha_inicio and fecha_inicio > fecha_str:
            continue  # El parámetro aún no era válido en esa fecha

        # Verificar fecha fin
        if fecha_fin and fecha_fin < fecha_str:
            continue  # El parámetro ya había expirado

        # El parámetro está vigente
        parametros_vigentes.append(param)

    return parametros_vigentes


def calcular_tendencia(df_precios, ticker, dias=10):
    """
    Calcula la tendencia de un ticker usando regresión lineal.

    Args:
        df_precios: DataFrame con columnas Date, Ticker, Close
        ticker: Símbolo del ticker a analizar
        dias: Número de días para el análisis (default 15)

    Returns:
        str: Tendencia en formato "+XX" o "-XX" donde XX es 0-100 en escalas de 10
             Retorna "N/A" si no hay suficientes datos
    """
    try:
        # Verificar que df_precios no sea None
        if df_precios is None:
            return "N/A"

        # Filtrar datos del ticker
        df_ticker = df_precios[df_precios['Ticker'] == ticker].copy()

        if len(df_ticker) < 5:  # Mínimo 5 días para calcular tendencia
            return "N/A"

        # Ordenar por fecha y tomar los últimos N días
        df_ticker = df_ticker.sort_values('Date').tail(dias)

        if len(df_ticker) < 5:
            return "N/A"

        # Preparar datos para regresión
        precios = df_ticker['Close'].values
        x = np.arange(len(precios))

        # Calcular regresión lineal: y = mx + b
        n = len(x)
        sum_x = np.sum(x)
        sum_y = np.sum(precios)
        sum_xy = np.sum(x * precios)
        sum_x2 = np.sum(x ** 2)

        # Pendiente (m)
        pendiente = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)

        # Intercepto (b)
        intercepto = (sum_y - pendiente * sum_x) / n

        # Calcular R² (coeficiente de determinación)
        y_pred = pendiente * x + intercepto
        ss_res = np.sum((precios - y_pred) ** 2)
        ss_tot = np.sum((precios - np.mean(precios)) ** 2)

        if ss_tot == 0:
            r2 = 0
        else:
            r2 = 1 - (ss_res / ss_tot)

        # Determinar dirección
        signo = "+" if pendiente > 0 else "-"

        # Calcular nivel (0-100) basado en R² y magnitud de la pendiente
        # R² indica qué tan consistente es la tendencia (0-1)
        # Normalizar R² a escala 0-100 y redondear a decenas
        nivel = int(round(abs(r2) * 100, -1))  # Redondear a decenas
        nivel = min(100, max(0, nivel))  # Asegurar rango 0-100

        return f"{signo}{nivel}"

    except Exception as e:
        print(f"[WARN] Error calculando tendencia para {ticker}: {e}")
        return "N/A"


def calcular_senales_para_parametros(parametros, df_precios, precios_dict, cartera):
    """Calcula señales para una lista de parámetros (función auxiliar)"""
    LIMITE_TIPO_DEFAULT = "acciones"
    LIMITE_VALOR_DEFAULT = 10.0

    senales = []
    for param in parametros:
        symbol = param.get('ticker_symbol')
        limite_tipo = param.get('limite_tipo', LIMITE_TIPO_DEFAULT)
        limite_valor = param.get('limite_valor', LIMITE_VALOR_DEFAULT)

        info_cartera = cartera.get(symbol, {"acciones": 0, "capital_invertido": 0, "precio_compra_minimo": 0, "precios_fifo": []})
        acciones_en_cartera = info_cartera.get("acciones", 0)
        capital_invertido = info_cartera.get("capital_invertido", 0)
        precio_compra_minimo = info_cartera.get("precio_compra_minimo", 0)
        precios_fifo = info_cartera.get("precios_fifo", [])

        if symbol not in precios_dict:
            senales.append({
                'symbol': symbol,
                'fecha_precio': 'N/A',
                'cierre': 'N/A',
                'precio_compra': 'N/A',
                'cant_compra': '-',
                'opc_compra': 'N/A',
                'precio_venta': 'N/A',
                'cant_venta': '-',
                'opc_venta': 'N/A',
                'acciones_cartera': acciones_en_cartera,
                'precio_compra_minimo': precio_compra_minimo,
                'ganancia_min_pct': param.get('ganancia_min_pct', 0),
                'limite_tipo': limite_tipo,
                'limite_valor': limite_valor,
                'tendencia': 'N/A',
                'estado': 'Sin datos de precio'
            })
            continue

        precio_info = precios_dict[symbol]
        cierre = precio_info['close']
        compra_pct = param.get('compra_pct', 0)
        venta_pct = param.get('venta_pct', 0)
        ganancia_min_pct = param.get('ganancia_min_pct', 0)

        precio_compra = cierre * (1 + compra_pct / 100)

        # Precio de venta basado en el cierre actual
        precio_venta_por_cierre = cierre * (1 + venta_pct / 100)

        # Precio de venta mínimo para garantizar ganancia sobre el precio de compra más bajo
        if precio_compra_minimo > 0 and ganancia_min_pct > 0:
            precio_venta_minimo = precio_compra_minimo * (1 + ganancia_min_pct / 100)
            # El precio de venta debe ser el mayor entre ambos para garantizar la ganancia mínima
            precio_venta = max(precio_venta_por_cierre, precio_venta_minimo)
        else:
            precio_venta = precio_venta_por_cierre

        promedio_minimos = param.get('promedio_minimos', 0)
        promedio_maximos = param.get('promedio_maximos', 0)
        compra_multiple_config = param.get('compra_multiple') or 1
        venta_multiple_config = param.get('venta_multiple') or 1

        usar_compra_multiple = False
        usar_venta_multiple = False

        if df_precios is not None and symbol in df_precios['Ticker'].values:
            try:
                hist_ticker = df_precios[df_precios['Ticker'] == symbol].sort_values('Date')
                if len(hist_ticker) >= 2:
                    precios_cierre = hist_ticker['Close'].values
                    precio_referencia = precios_cierre[0]
                    variacion_diaria_anterior = 0

                    for i in range(1, len(precios_cierre)):
                        precio_anterior = precios_cierre[i - 1]
                        precio_actual_iter = precios_cierre[i]
                        variacion_diaria = ((precio_actual_iter - precio_anterior) / precio_anterior) * 100

                        if variacion_diaria_anterior != 0:
                            if (variacion_diaria_anterior > 0 and variacion_diaria < 0) or \
                               (variacion_diaria_anterior < 0 and variacion_diaria > 0):
                                precio_referencia = precio_anterior

                        variacion_diaria_anterior = variacion_diaria

                    precio_actual = precios_cierre[-1]
                    pct_acumulado = ((precio_actual - precio_referencia) / precio_referencia) * 100

                    if promedio_minimos < 0 and pct_acumulado <= promedio_minimos:
                        usar_compra_multiple = True
                    if promedio_maximos > 0 and pct_acumulado >= promedio_maximos:
                        usar_venta_multiple = True
            except Exception as e:
                print(f"[WARN] Error calculando % acumulado para {symbol}: {e}")

        cant_compra = compra_multiple_config if usar_compra_multiple else 1
        cant_venta = venta_multiple_config if usar_venta_multiple else 1

        if limite_tipo == "acciones":
            limite_acciones = int(limite_valor)
            if acciones_en_cartera >= limite_acciones:
                opc_compra = "N/A (límite)"
            else:
                espacio_disponible = limite_acciones - acciones_en_cartera
                cant_compra = min(cant_compra, espacio_disponible)
                opc_compra = "COMPRAR"
        else:
            limite_monto = float(limite_valor)
            if capital_invertido >= limite_monto:
                opc_compra = "N/A (límite $)"
            else:
                monto_disponible = limite_monto - capital_invertido
                max_acciones_por_monto = int(monto_disponible / precio_compra) if precio_compra > 0 else 0
                if max_acciones_por_monto <= 0:
                    opc_compra = "N/A (límite $)"
                else:
                    cant_compra = min(cant_compra, max_acciones_por_monto)
                    opc_compra = "COMPRAR"

        if acciones_en_cartera <= 0:
            opc_venta = "N/A (sin acciones)"
            cant_venta = 0
        else:
            # Validar cantidad de venta usando FIFO
            cant_venta_deseada = min(cant_venta, acciones_en_cartera)
            if precios_fifo and ganancia_min_pct > 0:
                cant_venta, motivo_venta = calcular_cant_venta_valida_fifo(
                    precios_fifo, precio_venta, cant_venta_deseada, ganancia_min_pct
                )
                if cant_venta > 0:
                    opc_venta = "VENDER"
                else:
                    opc_venta = motivo_venta  # "ESPERAR (pérdida individual)" o "ESPERAR (ganancia insuficiente)"
            else:
                cant_venta = cant_venta_deseada
                opc_venta = "VENDER"

        # Calcular tendencias (corta 10 días, larga 30 días)
        tendencia_corta = calcular_tendencia(df_precios, symbol, dias=10)
        tendencia_larga = calcular_tendencia(df_precios, symbol, dias=30)

        senales.append({
            'symbol': symbol,
            'fecha_precio': precio_info['fecha'].strftime('%Y-%m-%d'),
            'cierre': cierre,
            'precio_compra': precio_compra,
            'cant_compra': cant_compra,
            'opc_compra': opc_compra,
            'precio_venta': precio_venta,
            'cant_venta': cant_venta,
            'opc_venta': opc_venta,
            'acciones_cartera': acciones_en_cartera,
            'precio_compra_minimo': precio_compra_minimo,
            'ganancia_min_pct': ganancia_min_pct,
            'limite_tipo': limite_tipo,
            'limite_valor': limite_valor,
            'tendencia': tendencia_corta,
            'tendencia_larga': tendencia_larga,
            'estado': 'OK'
        })

    return senales


def generar_senales_slot6(df_precios, cartera, plataforma=None, modo=None, fecha_senales=None):
    """Genera senales del Slot 6 (Claude diario) usando analisis contextual.

    El Slot 6 es diferente a los demas: no usa parametros fijos sino que
    analiza el contexto del mercado, indicadores tecnicos y senales de
    otros slots para generar recomendaciones.

    Args:
        df_precios: DataFrame con precios historicos
        cartera: Dict con estado actual de la cartera
        plataforma: Plataforma de inversion
        modo: Modo de operacion (Paper/Real)
        fecha_senales: Fecha para la cual se generan las senales (datetime)

    Returns:
        list: Lista de senales generadas por Claude
    """
    senales_slot6 = []

    try:
        # Cargar decisiones directamente del archivo JSON
        decisiones_file = UBICACION_JSON_PORTABLE / "decisiones_claude.json"

        if not decisiones_file.exists():
            print(f"[WARN] Slot 6: No existe {decisiones_file}")
            print("[INFO] Ejecuta primero: python Trading_Claude.py --analisis-diario")
            return senales_slot6

        # Cargar decisiones
        with open(decisiones_file, 'r', encoding='utf-8') as f:
            decisiones_data = json.load(f)

        # Obtener decisiones - soporta formato nuevo (directo) y antiguo (anidado)
        if not decisiones_data.get('decisiones'):
            print("[WARN] Slot 6: No hay decisiones guardadas")
            return senales_slot6

        # Detectar formato: nuevo (lista directa) vs antiguo (lista anidada con fecha)
        decisiones_list = decisiones_data['decisiones']

        # Formato nuevo: decisiones es lista de {symbol, accion, precio_compra, ...}
        # Formato antiguo: decisiones es lista de {fecha, decisiones_tickers: [...]}
        if decisiones_list and isinstance(decisiones_list[0], dict):
            if 'symbol' in decisiones_list[0] or 'ticker' in decisiones_list[0]:
                # Formato nuevo (directo)
                decisiones_tickers = decisiones_list
                fecha_decisiones = decisiones_data.get('fecha_generacion', 'desconocida')[:10]
            else:
                # Formato antiguo (anidado) - buscar decisiones que coincidan con plataforma/modo
                decisiones_hoy = None
                # Buscar desde la más reciente hacia atrás
                for dec in reversed(decisiones_list):
                    dec_plat = dec.get('plataforma', '')
                    dec_modo = dec.get('modo', '')
                    # Coincidir plataforma y modo (case insensitive)
                    if plataforma and modo:
                        if dec_plat == plataforma and dec_modo.lower() == modo.lower():
                            decisiones_hoy = dec
                            break
                    else:
                        # Si no se especifica plataforma/modo, usar la última
                        decisiones_hoy = dec
                        break

                if not decisiones_hoy:
                    print(f"[WARN] Slot 6: No hay decisiones para {plataforma} {modo}, usando ultima disponible")
                    # Usar la última decisión disponible como fallback
                    if decisiones_list:
                        decisiones_hoy = decisiones_list[-1]
                    else:
                        return senales_slot6

                decisiones_tickers = decisiones_hoy.get('decisiones_tickers', [])
                fecha_decisiones = decisiones_hoy.get('fecha', 'desconocida')
        else:
            print("[WARN] Slot 6: Formato de decisiones no reconocido")
            return senales_slot6

        # Validar fecha_trading: el análisis guardado debe ser para la misma fecha de trading
        fecha_trading_guardada = decisiones_hoy.get('fecha_trading') if decisiones_hoy else None
        fecha_guardada_str = fecha_trading_guardada  # Para mostrar en mensaje

        if fecha_senales and fecha_trading_guardada:
            try:
                fecha_esperada = fecha_senales.date() if hasattr(fecha_senales, 'date') else fecha_senales
                fecha_trading_dt = datetime.strptime(fecha_trading_guardada, "%Y-%m-%d").date()

                if fecha_trading_dt != fecha_esperada:
                    fecha_esperada_fmt = fecha_esperada.strftime("%d-%m-%Y")
                    fecha_guardada_fmt = fecha_trading_dt.strftime("%d-%m-%Y")
                    mensaje = (f"No se muestran precios porque Claude aún no ha realizado su análisis "
                              f"para la fecha de trading {fecha_esperada_fmt}. "
                              f"El último análisis de Claude guardado fue para la fecha de trading {fecha_guardada_fmt}.")
                    print(f"[INFO] Slot 6: {mensaje}")
                    # Retornar señal especial con aviso
                    return [{'estado': 'AVISO', 'mensaje': mensaje, 'symbol': 'AVISO'}]
            except Exception as e:
                print(f"[WARN] Slot 6: Error validando fecha_trading: {e}")
        elif fecha_senales and not fecha_trading_guardada:
            # Si no tiene fecha_trading, es análisis antiguo - verificar con fecha normal
            fecha_decisiones_str = decisiones_hoy.get('fecha') if decisiones_hoy else None
            if fecha_decisiones_str:
                try:
                    fecha_esperada = fecha_senales.date() if hasattr(fecha_senales, 'date') else fecha_senales
                    fecha_dec_dt = datetime.strptime(fecha_decisiones_str[:10], "%Y-%m-%d").date()
                    if fecha_dec_dt != fecha_esperada:
                        fecha_esperada_fmt = fecha_esperada.strftime("%d-%m-%Y")
                        fecha_guardada_fmt = fecha_dec_dt.strftime("%d-%m-%Y")
                        mensaje = (f"No se muestran precios porque Claude aún no ha realizado su análisis "
                                  f"para la fecha de trading {fecha_esperada_fmt}. "
                                  f"El último análisis de Claude guardado fue para la fecha de trading {fecha_guardada_fmt}.")
                        print(f"[INFO] Slot 6: {mensaje}")
                        return [{'estado': 'AVISO', 'mensaje': mensaje, 'symbol': 'AVISO'}]
                except Exception as e:
                    print(f"[WARN] Slot 6: Error validando fecha: {e}")

        print(f"[INFO] Slot 6: Cargando {len(decisiones_tickers)} decisiones ({fecha_decisiones})")

        for decision in decisiones_tickers:
            # Soportar ambos formatos: 'ticker' o 'symbol'
            ticker = decision.get('ticker') or decision.get('symbol', '')
            accion = decision.get('accion', 'esperar').lower()


            # Obtener precio de cierre del ticker
            cierre = None
            if df_precios is not None:
                df_ticker = df_precios[df_precios['Ticker'] == ticker]
                if not df_ticker.empty:
                    cierre = float(df_ticker['Close'].iloc[-1])

            # Calcular tendencias
            tendencia_corta = calcular_tendencia(df_precios, ticker, dias=10)
            tendencia_larga = calcular_tendencia(df_precios, ticker, dias=30)

            # Obtener acciones en cartera - SIEMPRE usar cartera real, no la del archivo
            acciones_cartera = 0
            if cartera and ticker in cartera:
                acciones_cartera = cartera[ticker].get('acciones', 0)

            # Obtener precio de compra mínimo
            precio_compra_minimo = decision.get('precio_compra_minimo')

            # Obtener precios sugeridos (soportar ambos formatos)
            precio_compra = decision.get('precio_compra') or decision.get('precio_compra_sugerido')
            precio_venta = decision.get('precio_venta') or decision.get('precio_venta_sugerido')

            # Cantidades: misma lógica que otros slots
            # cant_compra = 1 si hay precio de compra
            # cant_venta = min(1, acciones_cartera) si hay precio de venta
            cant_compra = 1 if precio_compra else 0
            cant_venta = min(1, acciones_cartera) if precio_venta and acciones_cartera > 0 else 0

            # Opciones de compra/venta: TODO EN MAYÚSCULAS
            opc_compra = 'COMPRAR' if precio_compra else 'N/A'
            opc_venta = 'VENDER' if precio_venta and acciones_cartera > 0 else 'N/A'
            if accion == 'esperar':
                opc_compra = 'ESPERAR' if precio_compra else 'N/A'
                opc_venta = 'ESPERAR' if precio_venta and acciones_cartera > 0 else 'N/A'

            # Ajustar cant_venta si no hay cartera
            if acciones_cartera <= 0:
                opc_venta = 'N/A'
                cant_venta = 0

            senal = {
                'symbol': ticker,
                'cierre': cierre,
                'precio_compra': precio_compra,
                'cant_compra': cant_compra,
                'opc_compra': opc_compra,
                'precio_venta': precio_venta,
                'cant_venta': cant_venta,
                'opc_venta': opc_venta,
                'acciones_cartera': acciones_cartera,
                'precio_compra_minimo': precio_compra_minimo,
                'ganancia_min_pct': 3.0,
                'limite_tipo': 'acciones',
                'limite_valor': 10,
                'tendencia': tendencia_corta if tendencia_corta else '-',
                'tendencia_larga': tendencia_larga if tendencia_larga else '-',
                'estado': 'OK',
                'confianza': decision.get('confianza', 'media'),
                'justificacion': decision.get('justificacion', {}),
                'slot_origen_compra': decision.get('slot_origen_compra', ''),
                'slot_origen_venta': decision.get('slot_origen_venta', ''),
                'slot_nombre': '6.-Claude diario'
            }

            # Agregar plataforma y modo
            if plataforma:
                senal['plataforma'] = plataforma
            if modo:
                senal['modo'] = modo.lower()

            senales_slot6.append(senal)

        print(f"[INFO] Slot 6 genero {len(senales_slot6)} senales")

    except Exception as e:
        print(f"[ERROR] Error generando senales Slot 6: {e}")
        import traceback
        traceback.print_exc()

    return senales_slot6


def generar_senales(plataforma=None, modo=None, mostrar_ventana=True):
    """Genera senales de compra/venta para TODOS los slots de parametros activos.

    Args:
        plataforma: Si se especifica, calcula cartera y filtra tickers para esta plataforma.
                    Si es None, calcula cartera global (comportamiento anterior).
        modo: Modo de operacion (Paper/Real). Por defecto Real.
        mostrar_ventana: Si True, muestra la ventana de señales al final.

    Returns:
        tuple: (senales_por_slot, datos_slots, total_senales) si mostrar_ventana=False, None en caso contrario
    """

    if not verificar_libs_cargadas(["pandas"]):
        if mostrar_ventana:
            messagebox.showwarning("Esperar", "Esperar que se carguen los recursos del sistema.")
        return None if not mostrar_ventana else None

    hoy = datetime.now()
    es_fin_de_semana = hoy.weekday() >= 5
    if es_fin_de_semana and mostrar_ventana:
        dia_semana = "sabado" if hoy.weekday() == 5 else "domingo"
        messagebox.showinfo("Mercado cerrado",
            f"Hoy es {dia_semana}. El mercado esta cerrado.\n\n"
            "Se mostraran las senales basadas en el ultimo dia de trading.\n"
            "(Las senales no se guardaran porque ya estan guardadas)")

    # Usar siempre la ruta portable del log (consistente con sincronizar_desde_github)
    log_file = str(AUTO_UPDATE_LOG_PORTABLE)

    if not os.path.exists(log_file):
        messagebox.showwarning("Sin datos", f"No existe el archivo de log:\n{log_file}\n\nDescarga los precios primero.")
        return

    # Cargar estructura de slots
    datos_slots, error = cargar_parametros_activos()
    if error:
        messagebox.showerror("Error", error)
        return

    # Calcular cartera (filtrada por plataforma y modo si se especifican)
    cartera = calcular_cartera(plataforma=plataforma, modo=modo)

    # Obtener tickers de la plataforma y modo (para filtrar senales)
    tickers_plataforma = obtener_tickers_plataforma(plataforma, modo) if plataforma else None

    try:
        df_precios = pd.read_csv(log_file, parse_dates=['Date'])
    except Exception as e:
        messagebox.showerror("Error", f"Error leyendo archivo de precios:\n{e}")
        return

    df_precios['Date'] = pd.to_datetime(df_precios['Date'])
    ultimos_precios = df_precios.sort_values('Date').groupby('Ticker').last().reset_index()

    precios_dict = {}
    fecha_senales = None
    for _, row in ultimos_precios.iterrows():
        precios_dict[row['Ticker']] = {
            'fecha': row['Date'],
            'close': row['Close'],
            'open': row['Open'],
            'high': row['High'],
            'low': row['Low']
        }
        # Tomar la fecha de cualquier ticker (todas deberían ser iguales para el último día)
        if fecha_senales is None:
            fecha_senales = row['Date']

    # Verificar si hay tickers configurados sin precios en el CSV
    tickers_en_csv = set(precios_dict.keys())
    tickers_requeridos = tickers_plataforma if tickers_plataforma else obtener_tickers_unicos()
    tickers_faltantes = [t for t in tickers_requeridos if t not in tickers_en_csv]

    if tickers_faltantes and mostrar_ventana:
        respuesta = messagebox.askyesno("Tickers sin precios",
            f"Los siguientes tickers no tienen precios en el CSV:\n\n"
            f"{', '.join(tickers_faltantes)}\n\n"
            f"Las señales para estos tickers no se generarán.\n\n"
            f"¿Desea continuar de todos modos?\n"
            f"(Presione 'No' para ir a Sync GitHub y descargar los precios faltantes)")
        if not respuesta:
            return

    # Verificar si debemos guardar las señales (precio de cierre confirmado)
    # - Si es fin de semana → NO guardar (ya se guardaron el viernes)
    # - Si la fecha de los precios NO es hoy → guardar
    # - Si la fecha es hoy Y hora NY >= 16:30 → guardar (mercado cerrado)
    # - Si la fecha es hoy Y hora NY < 16:30 → NO guardar (mercado abierto)
    now_ny = datetime.now(ZoneInfo("America/New_York"))
    hoy_ny = now_ny.date()
    hora_ny = now_ny.hour + now_ny.minute / 60  # Hora decimal (16:30 = 16.5)
    fecha_precios = fecha_senales.date() if fecha_senales else None

    # Guardar señales si:
    # - Es fin de semana (señales para el lunes)
    # - O fecha_precios es anterior a hoy (datos confirmados)
    # - O es hoy pero mercado ya cerró (hora NY >= 16:30)
    if es_fin_de_semana:
        guardar_senales = True  # En fin de semana siempre guardar (son para el lunes)
    else:
        guardar_senales = (fecha_precios != hoy_ny) or (fecha_precios == hoy_ny and hora_ny >= 16.5)

    # Calcular la fecha del siguiente día de trading (las señales son para esa fecha)
    fecha_siguiente_trading, feriados_saltados = siguiente_dia_trading(fecha_senales, retornar_feriados=True)
    fecha_guardar = fecha_siguiente_trading.strftime("%Y-%m-%d") + " 09:30:00"  # Apertura de mercado

    # Generar senales para CADA slot (para TODOS los tickers, el filtrado se hace al mostrar)
    senales_por_slot = {}
    for slot_id in ["1", "2", "3", "4", "5", "6"]:
        parametros = obtener_parametros_slot(datos_slots, slot_id)
        if parametros:
            # Filtrar parametros vigentes para la fecha del siguiente dia de trading
            parametros_vigentes = filtrar_parametros_por_fecha(parametros, fecha_siguiente_trading)

            # NOTA: Ya no filtramos por plataforma aquí - se filtra al mostrar en la ventana

            if parametros_vigentes:
                senales = calcular_senales_para_parametros(parametros_vigentes, df_precios, precios_dict, cartera)

                # Agregar campos plataforma y modo a cada senal
                for s in senales:
                    if plataforma:
                        s['plataforma'] = plataforma
                    if modo:
                        s['modo'] = modo.lower()

                senales_por_slot[slot_id] = senales
                # Solo guardar senales si corresponde (mercado cerrado y no es fin de semana)
                if guardar_senales:
                    nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                    guardar_historial_senales(senales, slot_id, nombre_slot, fecha_guardar, plataforma, modo, fecha_senales)
            else:
                senales_por_slot[slot_id] = []
        else:
            senales_por_slot[slot_id] = []

    # Generar senales del Slot 6 (Claude diario) usando Trading_Claude.py
    senales_slot6 = generar_senales_slot6(df_precios, cartera, plataforma, modo, fecha_siguiente_trading)
    senales_por_slot["6"] = senales_slot6

    # Guardar señales del Slot 6 en el historial (igual que los otros slots)
    if guardar_senales and senales_slot6:
        guardar_historial_senales(senales_slot6, "6", "6.-Claude diario", fecha_guardar, plataforma, modo, fecha_senales)

    # Mostrar ventana con senales de todos los slots o retornar datos
    if mostrar_ventana:
        mostrar_ventana_senales(senales_por_slot, datos_slots, plataforma=plataforma, modo=modo)
        return None
    else:
        total = sum(len(s) for s in senales_por_slot.values())
        # Retornar también las fechas y feriados para uso en mensajes
        return (senales_por_slot, datos_slots, total, fecha_senales, fecha_siguiente_trading, feriados_saltados)


def generar_senales_todas_plataformas(plataforma_mostrar=None, modo_mostrar=None):
    """Genera señales para TODAS las plataformas y modos configurados.

    Args:
        plataforma_mostrar: Plataforma para mostrar en la ventana final
        modo_mostrar: Modo para mostrar en la ventana final
    """
    if not verificar_libs_cargadas(["pandas"]):
        messagebox.showwarning("Esperar", "Esperar que se carguen los recursos del sistema.")
        return

    # Obtener todas las plataformas y modos configurados
    config = cargar_tickers_config()
    plataformas = config.get("plataformas", {})

    resumen = []
    total_global = 0
    fecha_datos = None
    fecha_senal = None
    feriados = []

    for plat_nombre, plat_config in plataformas.items():
        modos = plat_config.get("modos", {})
        for modo_nombre, modo_config in modos.items():
            tickers = modo_config.get("tickers", [])
            if tickers:  # Solo procesar si hay tickers configurados
                resultado = generar_senales(plataforma=plat_nombre, modo=modo_nombre, mostrar_ventana=False)
                if resultado:
                    _, _, total, f_datos, f_senal, f_feriados = resultado
                    # Capturar las fechas y feriados (son iguales para todas las plataformas)
                    if fecha_datos is None and f_datos is not None:
                        fecha_datos = f_datos
                        fecha_senal = f_senal
                        feriados = f_feriados
                    if total > 0:
                        resumen.append(f"  {plat_nombre} ({modo_nombre}): {total} señales")
                        total_global += total

    # Mostrar resumen con fechas
    if resumen:
        fecha_datos_str = fecha_datos.strftime("%d-%m-%Y") if fecha_datos else "N/A"
        fecha_senal_str = fecha_senal.strftime("%d-%m-%Y") if fecha_senal else "N/A"
        # Agregar información de feriados saltados si los hay
        feriados_str = ""
        if feriados:
            feriados_str = "\n\nFeriado USA: " + ", ".join(feriados)
        messagebox.showinfo("Señales Generadas",
            f"Señales generadas en todas las plataformas para el {fecha_senal_str},\n"
            f"en base a los últimos datos del {fecha_datos_str}:\n\n" +
            "\n".join(resumen) +
            f"\n\nTotal: {total_global} señales" + feriados_str)
    else:
        messagebox.showinfo("Sin señales", "No se generaron señales para ninguna plataforma.")

    # Mostrar ventana con la plataforma/modo seleccionados
    if plataforma_mostrar and modo_mostrar:
        generar_senales(plataforma=plataforma_mostrar, modo=modo_mostrar, mostrar_ventana=True)


def regenerar_senales_historicas():
    """Permite regenerar señales para una fecha anterior basándose en datos históricos"""

    # Verificar que las bibliotecas necesarias estén cargadas
    if not verificar_libs_cargadas(["pandas"]):
        messagebox.showwarning("Esperar", "Esperar que se carguen los recursos del sistema.")
        return

    # Verificar que hay un CSV configurado
    csv_file = entry_ruta.get()
    if not csv_file:
        messagebox.showwarning("Sin datos", "Primero selecciona y descarga un CSV de precios")
        return

    # Obtener ruta del log
    log_file = os.path.join(os.path.dirname(csv_file), "auto_update_log.csv")

    if not os.path.exists(log_file):
        messagebox.showwarning("Sin datos", f"No existe el archivo de log:\n{log_file}")
        return

    # Cargar precios del log
    try:
        df_precios = pd.read_csv(log_file, parse_dates=['Date'])
        df_precios['Date'] = pd.to_datetime(df_precios['Date'])
    except Exception as e:
        messagebox.showerror("Error", f"Error leyendo archivo de precios:\n{e}")
        return

    # Obtener fechas disponibles
    fechas_disponibles = sorted(df_precios['Date'].dt.strftime('%Y-%m-%d').unique(), reverse=True)

    if not fechas_disponibles:
        messagebox.showinfo("Sin datos", "No hay fechas disponibles en el log de precios")
        return

    # Crear ventana de selección de fecha
    ventana_fecha = tk.Toplevel(root)
    ventana_fecha.title("Regenerar Señales Históricas")
    ventana_fecha.geometry("450x280")
    ventana_fecha.transient(root)
    ventana_fecha.grab_set()

    tk.Label(ventana_fecha, text="Selecciona la fecha para regenerar señales:",
             font=("Arial", 10)).pack(pady=10)

    # Combobox con fechas disponibles
    fecha_var = tk.StringVar()
    combo_fechas = ttk.Combobox(ventana_fecha, textvariable=fecha_var, values=fechas_disponibles,
                                 state="readonly", width=20)
    combo_fechas.pack(pady=5)
    combo_fechas.current(0)

    tk.Label(ventana_fecha, text="(Se regenerarán señales para TODAS las plataformas/modos)",
             font=("Arial", 9), fg="blue").pack(pady=5)

    tk.Label(ventana_fecha, text="(Las señales se guardarán con la fecha siguiente a la seleccionada,\ndependiendo de la apertura del mercado)",
             font=("Arial", 9), fg="gray").pack(pady=5)

    def procesar_fecha():
        """Regenera señales para la fecha seleccionada, para TODAS las plataformas/modos"""
        fecha_seleccionada = fecha_var.get()
        if not fecha_seleccionada:
            return

        # Obtener todas las plataformas y modos configurados
        config = cargar_tickers_config()
        plataformas_config = config.get("plataformas", {})

        # Construir lista de TODAS las plataformas/modos (incluyendo las sin tickers)
        # Orden: Real primero, luego Paper para cada plataforma
        plat_modos = []
        for plat_nombre in sorted(plataformas_config.keys()):
            plat_config = plataformas_config[plat_nombre]
            modos = plat_config.get("modos", {})
            # Ordenar modos: Real primero, Paper después
            for modo_nombre in ["Real", "Paper"]:
                if modo_nombre in modos:
                    plat_modos.append((plat_nombre, modo_nombre))

        if not plat_modos:
            messagebox.showwarning("Sin configuración", "No hay plataformas configuradas.")
            return

        # Cargar estructura de slots
        datos_slots, error = cargar_parametros_activos()
        if error:
            messagebox.showerror("Error", error)
            return

        # Calcular el siguiente día de trading (las señales son para esa fecha)
        fecha_siguiente_trading = siguiente_dia_trading(datetime.strptime(fecha_seleccionada, "%Y-%m-%d"))

        # Filtrar precios para la fecha seleccionada
        df_fecha = df_precios[df_precios['Date'].dt.strftime('%Y-%m-%d') == fecha_seleccionada]

        if df_fecha.empty:
            messagebox.showwarning("Sin datos", f"No hay datos de precios para {fecha_seleccionada}")
            return

        # Crear diccionario de precios para esa fecha
        precios_dict = {}
        for _, row in df_fecha.iterrows():
            precios_dict[row['Ticker']] = {
                'fecha': row['Date'],
                'close': row['Close'],
                'open': row['Open'],
                'high': row['High'],
                'low': row['Low']
            }

        fecha_generacion = fecha_siguiente_trading.strftime("%Y-%m-%d") + " 09:30:00"
        resumen_resultado = []
        total_global = 0

        # Filtrar df_precios hasta la fecha seleccionada (no usar datos futuros)
        fecha_limite = pd.to_datetime(fecha_seleccionada)
        df_precios_historico = df_precios[df_precios['Date'] <= fecha_limite]

        # Procesar cada plataforma/modo
        for plat_nombre, modo_nombre in plat_modos:
            cartera = calcular_cartera_historica(fecha_siguiente_trading, plat_nombre, modo_nombre)
            total_plat = 0

            # Obtener tickers configurados para esta plataforma/modo
            tickers_plat_modo = obtener_tickers_plataforma(plat_nombre, modo_nombre)
            if not tickers_plat_modo:
                # Agregar al resumen con 0 señales y continuar
                resumen_resultado.append(f"  {plat_nombre} ({modo_nombre}): 0")
                continue

            for slot_id in ["1", "2", "3", "4", "5", "6"]:
                parametros = obtener_parametros_slot(datos_slots, slot_id)
                if not parametros:
                    continue

                parametros_vigentes = filtrar_parametros_por_fecha(parametros, fecha_siguiente_trading)
                if not parametros_vigentes:
                    continue

                # Filtrar parámetros solo para tickers de esta plataforma/modo
                parametros_filtrados = [p for p in parametros_vigentes
                                       if p.get('ticker_symbol') in tickers_plat_modo]
                if not parametros_filtrados:
                    continue

                senales = calcular_senales_para_parametros(parametros_filtrados, df_precios_historico, precios_dict, cartera)

                if senales:
                    for s in senales:
                        s['plataforma'] = plat_nombre
                        s['modo'] = modo_nombre.lower()

                    nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                    # Pasar fecha del cierre usado (fecha_seleccionada es string, convertir a datetime)
                    fecha_cierre_dt = datetime.strptime(fecha_seleccionada, "%Y-%m-%d")
                    guardar_historial_senales(senales, slot_id, nombre_slot, fecha_generacion, plat_nombre, modo_nombre, fecha_cierre_dt)
                    total_plat += len(senales)

            # Siempre agregar al resumen (incluso con 0 señales)
            resumen_resultado.append(f"  {plat_nombre} ({modo_nombre}): {total_plat}")
            total_global += total_plat

        ventana_fecha.destroy()
        fecha_siguiente_str = fecha_siguiente_trading.strftime("%Y-%m-%d")
        messagebox.showinfo("Éxito",
            f"Señales regeneradas para {fecha_siguiente_str}:\n\n"
            + "\n".join(resumen_resultado) +
            f"\n\nTotal: {total_global} señales")

    def procesar_todas_fechas():
        """Regenera señales para TODAS las fechas y TODAS las plataformas/modos"""
        # Obtener todas las plataformas y modos configurados
        config = cargar_tickers_config()
        plataformas_config = config.get("plataformas", {})

        # Construir lista de plataformas/modos con tickers
        plat_modos = []
        for plat_nombre, plat_config in plataformas_config.items():
            modos = plat_config.get("modos", {})
            for modo_nombre, modo_config in modos.items():
                tickers = modo_config.get("tickers", [])
                if tickers:
                    plat_modos.append((plat_nombre, modo_nombre, len(tickers)))

        if not plat_modos:
            messagebox.showwarning("Sin configuración", "No hay plataformas/modos con tickers configurados.")
            return

        resumen_plats = "\n".join([f"  - {p} ({m}): {t} tickers" for p, m, t in plat_modos])
        if not messagebox.askyesno("Confirmar",
            f"¿Regenerar señales para las {len(fechas_disponibles)} fechas disponibles?\n\n"
            f"Plataformas/Modos a procesar:\n{resumen_plats}\n\n"
            "Esto reemplazará TODAS las señales históricas."):
            return

        # Limpiar historial completo
        historial_path = obtener_ruta_senales()
        historial = {"version": "2.0", "senales_por_slot": {"1": [], "2": [], "3": [], "4": [], "5": [], "6": []}}
        with open(historial_path, 'w', encoding='utf-8') as f:
            json.dump(historial, f, indent=2, ensure_ascii=False)

        # Cargar estructura de slots
        datos_slots, error = cargar_parametros_activos()
        if error:
            messagebox.showerror("Error", error)
            return

        resumen_resultado = []
        total_global = 0

        # Procesar cada plataforma/modo
        for plat_nombre, modo_nombre, _ in plat_modos:
            total_plat = 0
            fechas_procesadas = 0

            # Obtener tickers configurados para esta plataforma/modo
            tickers_plat_modo = obtener_tickers_plataforma(plat_nombre, modo_nombre)
            if not tickers_plat_modo:
                # Agregar al resumen con 0 señales y continuar
                resumen_resultado.append(f"  {plat_nombre} ({modo_nombre}): 0 señales")
                continue

            # Procesar cada fecha (de más antigua a más reciente)
            for fecha_str in sorted(fechas_disponibles):
                fecha_siguiente_trading = siguiente_dia_trading(datetime.strptime(fecha_str, "%Y-%m-%d"))
                cartera = calcular_cartera_historica(fecha_siguiente_trading, plat_nombre, modo_nombre)

                df_fecha = df_precios[df_precios['Date'].dt.strftime('%Y-%m-%d') == fecha_str]
                if df_fecha.empty:
                    continue

                precios_dict = {}
                for _, row in df_fecha.iterrows():
                    precios_dict[row['Ticker']] = {
                        'fecha': row['Date'],
                        'close': row['Close'],
                        'open': row['Open'],
                        'high': row['High'],
                        'low': row['Low']
                    }

                fecha_generacion = fecha_siguiente_trading.strftime("%Y-%m-%d") + " 09:30:00"

                # Filtrar df_precios hasta la fecha actual (no usar datos futuros)
                fecha_limite = pd.to_datetime(fecha_str)
                df_precios_historico = df_precios[df_precios['Date'] <= fecha_limite]

                for slot_id in ["1", "2", "3", "4", "5", "6"]:
                    parametros = obtener_parametros_slot(datos_slots, slot_id)
                    if not parametros:
                        continue

                    parametros_vigentes = filtrar_parametros_por_fecha(parametros, fecha_siguiente_trading)
                    if not parametros_vigentes:
                        continue

                    # Filtrar parámetros solo para tickers de esta plataforma/modo
                    parametros_filtrados = [p for p in parametros_vigentes
                                           if p.get('ticker_symbol') in tickers_plat_modo]
                    if not parametros_filtrados:
                        continue

                    senales = calcular_senales_para_parametros(parametros_filtrados, df_precios_historico, precios_dict, cartera)

                    if senales:
                        # Agregar plataforma y modo a cada señal
                        for s in senales:
                            s['plataforma'] = plat_nombre
                            s['modo'] = modo_nombre.lower()

                        nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                        # Pasar fecha del cierre usado (fecha_str es string, convertir a datetime)
                        fecha_cierre_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
                        guardar_historial_senales(senales, slot_id, nombre_slot, fecha_generacion, plat_nombre, modo_nombre, fecha_cierre_dt)
                        total_plat += len(senales)

                fechas_procesadas += 1

            # Siempre agregar al resumen (incluso con 0 señales)
            resumen_resultado.append(f"  {plat_nombre} ({modo_nombre}): {total_plat} señales")
            total_global += total_plat

        ventana_fecha.destroy()
        if resumen_resultado:
            messagebox.showinfo("Completado",
                f"Regeneración completada:\n\n"
                f"Fechas procesadas: {len(fechas_disponibles)}\n\n"
                + "\n".join(resumen_resultado) +
                f"\n\nTotal: {total_global} señales")
        else:
            messagebox.showinfo("Sin señales", "No se generaron señales para ninguna plataforma/modo.")

    frame_botones = tk.Frame(ventana_fecha)
    frame_botones.pack(pady=20)

    tk.Button(frame_botones, text="Regenerar Señales", command=procesar_fecha,
              bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Regenerar TODAS", command=procesar_todas_fechas,
              bg="#6c757d", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cancelar", command=ventana_fecha.destroy).pack(side="left", padx=5)


def mostrar_ventana_senales(senales_por_slot, datos_slots, titulo_extra="", plataforma=None, modo=None):
    """Muestra una ventana con las senales generadas organizadas en pestanas por slot.

    Args:
        senales_por_slot: Dict con senales por slot
        datos_slots: Estructura de slots
        titulo_extra: Texto adicional para el titulo
        plataforma: Plataforma actual (para el selector)
        modo: Modo de operacion (Paper/Real)
    """

    ventana_senales = tk.Toplevel(root)
    ventana_senales.title("Senales de Trading - " + datetime.now().strftime("%Y-%m-%d %H:%M") + titulo_extra)
    ventana_senales.geometry("1250x550")

    fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Frame superior con info
    frame_info = tk.Frame(ventana_senales, pady=5)
    frame_info.pack(fill="x", padx=10)

    total_senales = sum(len(s) for s in senales_por_slot.values())
    lbl_info = tk.Label(frame_info, text=f"Senales generadas: {fecha_generacion}",
             font=("Arial", 10, "bold"), width=50, anchor="w")
    lbl_info.pack(side="left")

    # Selector de plataforma
    tk.Label(frame_info, text="Plataforma:", font=("Arial", 9)).pack(side="left", padx=(15, 2))
    plataforma_senales_var = tk.StringVar(value=plataforma if plataforma else obtener_plataformas()[0])
    combo_plat_senales = ttk.Combobox(frame_info, textvariable=plataforma_senales_var,
                                       values=obtener_plataformas(), state="readonly", width=10)
    combo_plat_senales.pack(side="left")

    # Selector de modo (al lado de plataforma)
    tk.Label(frame_info, text="Modo:", font=("Arial", 9)).pack(side="left", padx=(8, 2))
    modo_senales_var = tk.StringVar(value=modo if modo else "Real")
    combo_modo_senales = ttk.Combobox(frame_info, textvariable=modo_senales_var,
                                       values=["Paper", "Real"], state="readonly", width=6)
    combo_modo_senales.pack(side="left")

    def filtrar_senales_por_plataforma_modo():
        """Filtra las señales según plataforma y modo seleccionados.

        Usa las señales RECIÉN GENERADAS si coinciden con plataforma/modo,
        de lo contrario carga del historial.
        """
        plat = plataforma_senales_var.get()
        modo_sel = modo_senales_var.get()

        # Obtener tickers válidos para esta combinación plataforma+modo
        tickers_validos = set(obtener_tickers_plataforma(plat, modo_sel))
        modo_lower = modo_sel.lower()

        # Usar señales recién generadas si coincide plataforma/modo
        usar_senales_actuales = (plat == plataforma and modo_lower == (modo or '').lower())

        senales_filtradas = {}
        total = 0

        if usar_senales_actuales and senales_por_slot:
            # Usar señales recién generadas
            for slot_id in ["1", "2", "3", "4", "5", "6"]:
                senales_slot = senales_por_slot.get(slot_id, [])
                # Filtrar por tickers válidos, pero preservar señales de AVISO (Slot 6)
                filtradas = [s for s in senales_slot
                            if s.get('symbol', '') in tickers_validos or s.get('estado') == 'AVISO']
                senales_filtradas[slot_id] = filtradas
                # No contar avisos en el total
                total += len([s for s in filtradas if s.get('estado') != 'AVISO'])

            lbl_info.config(text=f"Señales de {plat} ({modo_sel}): {total} - Generadas ahora")
        else:
            # Plataforma/modo diferente: cargar slots 1-5 del historial, regenerar Slot 6
            historial = cargar_historial_senales()

            for slot_id in ["1", "2", "3", "4", "5"]:
                senales_slot = historial.get("senales_por_slot", {}).get(slot_id, [])
                # Filtrar por plataforma, modo y tickers de la plataforma
                filtradas = [s for s in senales_slot
                            if s.get('plataforma') == plat
                            and s.get('modo', 'real').lower() == modo_lower
                            and s.get('symbol', '') in tickers_validos]

                # Si hay señales, tomar solo las de la fecha más reciente
                if filtradas:
                    fecha_mas_reciente = max(s.get('fecha_generacion', '')[:10] for s in filtradas)
                    filtradas = [s for s in filtradas if s.get('fecha_generacion', '')[:10] == fecha_mas_reciente]

                senales_filtradas[slot_id] = filtradas
                total += len(filtradas)

            # Regenerar Slot 6 con la cartera correcta de la plataforma seleccionada
            try:
                cartera_plat = calcular_cartera(plataforma=plat, modo=modo_sel)
                csv_file = entry_ruta.get()
                log_file = os.path.join(os.path.dirname(csv_file), "auto_update_log.csv") if csv_file else None
                df_precios = pd.read_csv(log_file, parse_dates=['Date']) if log_file and os.path.exists(log_file) else None
                # Calcular fecha_trading basándose en la fecha del último cierre del CSV
                fecha_trading_slot6 = None
                if df_precios is not None and not df_precios.empty:
                    fecha_ultimo_cierre = df_precios['Date'].max()
                    fecha_trading_slot6 = siguiente_dia_trading(fecha_ultimo_cierre)
                senales_slot6 = generar_senales_slot6(df_precios, cartera_plat, plat, modo_sel, fecha_trading_slot6)
                # Filtrar por tickers válidos, pero preservar señales de AVISO
                filtradas_s6 = [s for s in senales_slot6
                               if s.get('symbol', '') in tickers_validos or s.get('estado') == 'AVISO']
                senales_filtradas["6"] = filtradas_s6
                # No contar avisos en el total
                total += len([s for s in filtradas_s6 if s.get('estado') != 'AVISO'])
            except Exception as e:
                print(f"[WARN] Error regenerando Slot 6: {e}")
                senales_filtradas["6"] = []

            # Actualizar título con total de señales filtradas
            if total > 0:
                fecha_senales = senales_filtradas.get("1", [{}])[0].get('fecha_generacion', '')[:10] if senales_filtradas.get("1") else ""
                lbl_info.config(text=f"Señales de {plat} ({modo_sel}): {total} - Fecha: {fecha_senales}")
            else:
                lbl_info.config(text=f"Señales de {plat} ({modo_sel}): Sin señales guardadas")

        # Actualizar títulos de pestañas con conteos filtrados
        for i, slot_id in enumerate(["1", "2", "3", "4", "5", "6"]):
            nombre = obtener_nombre_slot(datos_slots, slot_id)
            count = len(senales_filtradas.get(slot_id, []))
            notebook.tab(i, text=f"{nombre} ({count})")

        # Repoblar trees con señales filtradas
        poblar_trees(senales_filtradas)

    def cambiar_plataforma_o_modo(*args):
        """Refresca las señales cuando cambia la plataforma o el modo"""
        # Mantener el modo actual (no cambiarlo automáticamente)
        filtrar_senales_por_plataforma_modo()

    combo_plat_senales.bind("<<ComboboxSelected>>", cambiar_plataforma_o_modo)

    # Checkbox "Ver anteriores" dentro de la ventana
    ver_ant_var = tk.BooleanVar(value=False)
    tk.Checkbutton(frame_info, text="Ver guardadas", variable=ver_ant_var,
                   font=("Arial", 9), command=lambda: toggle_ver_anteriores()).pack(side="left", padx=15)

    # Campo limite de precio
    tk.Label(frame_info, text="Limite %:", font=("Arial", 9)).pack(side="left", padx=(10, 2))
    limite_plataforma_var = tk.StringVar(value="3")
    entry_limite = tk.Entry(frame_info, textvariable=limite_plataforma_var, width=4, font=("Arial", 9), justify="center")
    entry_limite.pack(side="left")
    tk.Label(frame_info, text="%", font=("Arial", 9)).pack(side="left", padx=(0, 5))

    def aplicar_limite():
        if ver_ant_var.get():
            toggle_ver_anteriores()  # Recargar guardadas con nuevo limite
        else:
            filtrar_senales_por_plataforma_modo()  # Recargar actuales filtradas con nuevo limite

    tk.Button(frame_info, text="Aplicar", command=aplicar_limite, font=("Arial", 8),
              bg="#6c757d", fg="white", padx=5).pack(side="left", padx=5)

    # Binding para modo (combo ya definido arriba, junto a plataforma)
    def on_modo_senales_change(event=None):
        """Callback cuando cambia el modo Paper/Real"""
        filtrar_senales_por_plataforma_modo()

    combo_modo_senales.bind("<<ComboboxSelected>>", on_modo_senales_change)

    tk.Label(frame_info, text=f"Total senales: {total_senales}",
             font=("Arial", 10)).pack(side="right")



    # Notebook con pestañas
    notebook = ttk.Notebook(ventana_senales)
    notebook.pack(fill="both", expand=True, padx=10, pady=5)

    columns = ("Symbol", "Cartera", "Cierre últ.", "P.Compra", "Cant.C", "Opc.Compra", "P.Venta", "Cant.V", "Opc.Venta", "Tend.C", "Tend.L")
    anchos = {"Symbol": 70, "Cartera": 60, "Cierre últ.": 85, "P.Compra": 85, "Cant.C": 50,
              "Opc.Compra": 110, "P.Venta": 85, "Cant.V": 50, "Opc.Venta": 120, "Tend.C": 55, "Tend.L": 55}

    trees = {}
    labels_aviso = {}  # Labels para mostrar mensajes de aviso

    def crear_pestaña_slot(slot_id):
        """Crea una pestaña con el treeview vacío para un slot"""
        frame_slot = tk.Frame(notebook)

        # Label de aviso (oculto por defecto)
        lbl_aviso = tk.Label(frame_slot, text="", bg="#FFF3CD", fg="#856404",
                            font=("Arial", 10), wraplength=600, justify="left",
                            padx=10, pady=10)
        labels_aviso[slot_id] = lbl_aviso

        frame_tabla = tk.Frame(frame_slot)
        frame_tabla.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar_y = tk.Scrollbar(frame_tabla, orient="vertical")
        scrollbar_x = tk.Scrollbar(frame_tabla, orient="horizontal")

        tree = ttk.Treeview(frame_tabla, columns=columns, show="headings",
                            yscrollcommand=scrollbar_y.set,
                            xscrollcommand=scrollbar_x.set)

        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=anchos.get(col, 70), anchor="center")

        # Tags para precios ajustados (naranja)
        tree.tag_configure("ajustado", foreground="#FF6600")

        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        trees[slot_id] = tree
        return frame_slot

    def poblar_trees(datos):
        """Llena todos los trees con los datos proporcionados"""
        # Obtener límite de plataforma (None = sin límite)
        valor_limite = limite_plataforma_var.get().strip()
        if valor_limite == "" or valor_limite == "0":
            limite_pct = None  # Sin límite
        else:
            try:
                limite_pct = float(valor_limite) / 100.0
            except ValueError:
                limite_pct = None  # Valor inválido = sin límite

        for slot_id, tree in trees.items():
            tree.delete(*tree.get_children())
            senales = datos.get(slot_id, [])

            # Detectar si hay un mensaje de aviso (Slot 6 sin análisis actualizado)
            if senales and len(senales) == 1 and senales[0].get('estado') == 'AVISO':
                mensaje = senales[0].get('mensaje', 'Análisis no disponible')
                # Mostrar label de aviso
                if slot_id in labels_aviso:
                    labels_aviso[slot_id].config(text=mensaje)
                    labels_aviso[slot_id].pack(fill="x", padx=5, pady=5, before=tree.master)
                # Actualizar texto de pestaña
                idx = int(slot_id) - 1
                nombre = obtener_nombre_slot(datos_slots, slot_id)
                notebook.tab(idx, text=f"{nombre} (0)")
                continue
            else:
                # Ocultar label de aviso si existe
                if slot_id in labels_aviso:
                    labels_aviso[slot_id].pack_forget()

            senales_ordenadas = sorted(senales, key=lambda x: x.get('symbol', '').upper())
            for senal in senales_ordenadas:
                # Detectar formato: señales guardadas usan precio_cierre, señales actuales usan cierre
                es_senal_guardada = 'precio_cierre' in senal and 'cierre' not in senal
                tiene_datos = senal.get('estado') == 'OK' or es_senal_guardada

                if tiene_datos:
                    # Obtener precios del formato correcto
                    cierre = senal.get('precio_cierre') if es_senal_guardada else senal.get('cierre')
                    precio_compra_orig = senal.get('precio_compra_sugerido') if es_senal_guardada else senal.get('precio_compra')
                    precio_venta_orig = senal.get('precio_venta_sugerido') if es_senal_guardada else senal.get('precio_venta')

                    # Validar que tenemos los datos necesarios
                    # Slot 6 puede tener solo precio_compra O precio_venta (no ambos)
                    if cierre is None:
                        continue
                    # Para slots normales, requerir ambos precios; para Slot 6, al menos uno
                    es_slot6 = slot_id == "6"
                    if not es_slot6 and (precio_compra_orig is None or precio_venta_orig is None):
                        continue

                    # Valores por defecto para Slot 6 cuando falta un precio
                    if precio_compra_orig is None:
                        precio_compra_orig = 0
                    if precio_venta_orig is None:
                        precio_venta_orig = 0

                    # Ajustar precios si hay límite activo
                    precio_compra_mostrar = precio_compra_orig
                    precio_venta_mostrar = precio_venta_orig
                    compra_ajustada = False
                    venta_ajustada = False

                    if limite_pct is not None and limite_pct > 0:
                        limite_compra_min = cierre * (1 - limite_pct)
                        limite_venta_max = cierre * (1 + limite_pct)

                        if precio_compra_orig < limite_compra_min:
                            precio_compra_mostrar = limite_compra_min
                            compra_ajustada = True

                        if precio_venta_orig > limite_venta_max:
                            precio_venta_mostrar = limite_venta_max
                            venta_ajustada = True

                    # Formato de precios: agregar * si fue ajustado, "-" si es 0
                    # Para Slot 6: agregar slot de origen (ej: "$250.65 S1")
                    slot_origen_compra = senal.get('slot_origen_compra', '') if es_slot6 else ''
                    slot_origen_venta = senal.get('slot_origen_venta', '') if es_slot6 else ''

                    if precio_compra_mostrar == 0:
                        str_compra = "-"
                    elif compra_ajustada:
                        str_compra = f"*${precio_compra_mostrar:.2f}"
                        if slot_origen_compra:
                            str_compra += f" {slot_origen_compra}"
                    else:
                        str_compra = f"${precio_compra_mostrar:.2f}"
                        if slot_origen_compra:
                            str_compra += f" {slot_origen_compra}"

                    if precio_venta_mostrar == 0:
                        str_venta = "-"
                    elif venta_ajustada:
                        str_venta = f"*${precio_venta_mostrar:.2f}"
                        if slot_origen_venta:
                            str_venta += f" {slot_origen_venta}"
                    else:
                        str_venta = f"${precio_venta_mostrar:.2f}"
                        if slot_origen_venta:
                            str_venta += f" {slot_origen_venta}"

                    # Ajustar opción de venta si el precio ajustado no cumple la ganancia mínima
                    opc_venta_mostrar = senal['opc_venta']
                    precio_compra_min_cartera = senal.get('precio_compra_minimo') or 0
                    ganancia_min_param = senal.get('ganancia_min_pct') or 0
                    if venta_ajustada and precio_compra_min_cartera > 0:
                        # Calcular precio mínimo de venta para cumplir ganancia mínima
                        precio_venta_minimo_req = precio_compra_min_cartera * (1 + ganancia_min_param / 100)
                        if precio_venta_mostrar < precio_venta_minimo_req:
                            opc_venta_mostrar = "ESPERAR"

                    cartera_mostrar = senal['acciones_cartera']
                    tree.insert("", "end", values=(
                        senal['symbol'],
                        cartera_mostrar,
                        f"${cierre:.2f}",
                        str_compra,
                        senal['cant_compra'],
                        senal['opc_compra'],
                        str_venta,
                        senal['cant_venta'],
                        opc_venta_mostrar,
                        senal.get('tendencia', 'N/A'),
                        senal.get('tendencia_larga', 'N/A')
                    ))
                else:
                    cartera_mostrar = senal.get('acciones_cartera', 0)
                    tree.insert("", "end", values=(
                        senal['symbol'],
                        cartera_mostrar,
                        senal.get('cierre', 'N/A'),
                        "-", "-",
                        senal.get('opc_compra', 'N/A'),
                        "-", "-",
                        senal.get('opc_venta', 'N/A'),
                        senal.get('tendencia', 'N/A'),
                        senal.get('tendencia_larga', 'N/A')
                    ))
            # Actualizar texto de pestaña
            idx = int(slot_id) - 1
            nombre = obtener_nombre_slot(datos_slots, slot_id)
            notebook.tab(idx, text=f"{nombre} ({len(senales_ordenadas)})")

    def toggle_ver_anteriores():
        """Alterna entre señales actuales y guardadas (muestra la fecha anterior a la más reciente)"""
        if ver_ant_var.get():
            # Cargar señales guardadas de la fecha ANTERIOR (no la más reciente que es igual a la actual)
            datos_senales = cargar_historial_senales()

            # Recopilar todas las fechas únicas de todos los slots
            todas_fechas = set()
            for slot_id in ["1", "2", "3", "4", "5", "6"]:
                for s in datos_senales.get("senales_por_slot", {}).get(slot_id, []):
                    todas_fechas.add(s.get("fecha_generacion", "")[:10])

            fechas_ordenadas = sorted(todas_fechas, reverse=True)
            if len(fechas_ordenadas) < 2:
                messagebox.showinfo("Sin datos", "No hay señales anteriores guardadas.")
                ver_ant_var.set(False)
                return

            # Usar la segunda fecha más reciente (la anterior a la actual)
            fecha_anterior = fechas_ordenadas[1]

            senales_guardadas = {}
            for slot_id in ["1", "2", "3", "4", "5", "6"]:
                senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
                senales_fecha = [s for s in senales_slot if s.get("fecha_generacion", "")[:10] == fecha_anterior]
                senales_convertidas = []
                for s in senales_fecha:
                    senales_convertidas.append({
                        'symbol': s.get('symbol', ''),
                        'cierre': s.get('precio_cierre', 0),
                        'precio_compra': s.get('precio_compra_sugerido', 0),
                        'cant_compra': s.get('cant_compra', '-'),
                        'opc_compra': s.get('opc_compra', ''),
                        'precio_venta': s.get('precio_venta_sugerido', 0),
                        'cant_venta': s.get('cant_venta', '-'),
                        'opc_venta': s.get('opc_venta', ''),
                        'acciones_cartera': s.get('acciones_cartera', 0),
                        'tendencia': s.get('tendencia', 'N/A'),
                        'tendencia_larga': s.get('tendencia_larga', 'N/A'),
                        'estado': 'OK'
                    })
                senales_guardadas[slot_id] = senales_convertidas
            poblar_trees(senales_guardadas)
            lbl_info.config(text=f"⮜ Señales anteriores (guardadas para: {fecha_anterior})")
        else:
            # Restaurar señales actuales (filtradas)
            filtrar_senales_por_plataforma_modo()

    # Crear pestañas para cada slot
    for slot_id in ["1", "2", "3", "4", "5", "6"]:
        nombre = obtener_nombre_slot(datos_slots, slot_id)
        frame = crear_pestaña_slot(slot_id)
        senales = senales_por_slot.get(slot_id, [])
        notebook.add(frame, text=f"{nombre} ({len(senales)})")

    # Poblar con señales filtradas por plataforma+modo inicial
    filtrar_senales_por_plataforma_modo()

    # Frame de botones
    frame_botones = tk.Frame(ventana_senales, pady=10)
    frame_botones.pack(fill="x", padx=10)

    def exportar_excel():
        """Exporta las señales de todos los slots a Excel"""
        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Señales",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Senales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )

        if not ruta_excel:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            compra_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            venta_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            headers = ["Symbol", "Cartera", "Cierre últ.", "P.Compra", "Cant.C", "Opc.Compra", "P.Venta", "Cant.V", "Opc.Venta"]

            for slot_id in ["1", "2", "3", "4", "5", "6"]:
                senales = senales_por_slot.get(slot_id, [])
                if not senales:
                    continue

                nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                ws = wb.create_sheet(title=nombre_slot[:31])

                ws.cell(row=1, column=1, value=f"Señales - {nombre_slot} ({fecha_generacion})")
                ws.cell(row=1, column=1).font = Font(bold=True)

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                for row_idx, senal in enumerate(senales, 4):
                    ws.cell(row=row_idx, column=1, value=senal['symbol']).border = border
                    ws.cell(row=row_idx, column=2, value=senal.get('acciones_cartera', 0)).border = border

                    if senal.get('estado') == 'OK':
                        cell_cierre = ws.cell(row=row_idx, column=3, value=senal['cierre'])
                        cell_cierre.number_format = '$#,##0.00'
                        cell_cierre.border = border

                        cell_pcompra = ws.cell(row=row_idx, column=4, value=senal['precio_compra'])
                        cell_pcompra.number_format = '$#,##0.00'
                        cell_pcompra.fill = compra_fill
                        cell_pcompra.border = border

                        ws.cell(row=row_idx, column=5, value=senal['cant_compra']).border = border

                        cell_opc_compra = ws.cell(row=row_idx, column=6, value=senal['opc_compra'])
                        cell_opc_compra.border = border
                        if senal['opc_compra'] == "Comprar":
                            cell_opc_compra.fill = compra_fill

                        cell_pventa = ws.cell(row=row_idx, column=7, value=senal['precio_venta'])
                        cell_pventa.number_format = '$#,##0.00'
                        cell_pventa.fill = venta_fill
                        cell_pventa.border = border

                        ws.cell(row=row_idx, column=8, value=senal['cant_venta']).border = border

                        cell_opc_venta = ws.cell(row=row_idx, column=9, value=senal['opc_venta'])
                        cell_opc_venta.border = border
                        if senal['opc_venta'] == "Vender":
                            cell_opc_venta.fill = venta_fill
                    else:
                        ws.cell(row=row_idx, column=3, value=senal.get('cierre', 'N/A')).border = border
                        ws.cell(row=row_idx, column=4, value="-").border = border
                        ws.cell(row=row_idx, column=5, value="-").border = border
                        ws.cell(row=row_idx, column=6, value=senal.get('opc_compra', 'N/A')).border = border
                        ws.cell(row=row_idx, column=7, value="-").border = border
                        ws.cell(row=row_idx, column=8, value="-").border = border
                        ws.cell(row=row_idx, column=9, value=senal.get('opc_venta', 'N/A')).border = border

                for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
                    ws.column_dimensions[col].width = 14

            wb.save(ruta_excel)
            messagebox.showinfo("Exportado", f"Señales exportadas a:\n{ruta_excel}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")

    tk.Button(frame_botones, text="Exportar a Excel", command=exportar_excel,
              bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_senales.destroy).pack(side="right", padx=5)

    # Nota informativa
    frame_nota = tk.Frame(ventana_senales, pady=5)
    frame_nota.pack(fill="x", padx=10)
    tk.Label(frame_nota,
             text="Nota: Cada pestaña muestra las señales de un slot de Parámetros Activos",
             font=("Arial", 9), fg="gray").pack(anchor="w")


def comparar_senales_operaciones():
    """Abre ventana para comparar señales generadas con operaciones reales (con 5 pestañas por slot)"""

    # Verificar que las bibliotecas necesarias estén cargadas
    if not verificar_libs_cargadas(["pandas", "matplotlib"]):
        messagebox.showwarning("Esperar", "Esperar que se carguen los recursos del sistema.")
        return

    ruta_senales = obtener_ruta_senales()
    if ruta_senales is None:
        messagebox.showerror("Error", "No hay ubicacion configurada.\nVerifica que exista la carpeta data/")
        return

    # Cargar datos de slots
    datos_slots, error = cargar_parametros_activos()
    if error:
        messagebox.showerror("Error", f"Error cargando parámetros: {error}")
        return

    datos_senales = cargar_historial_senales()
    operaciones = cargar_historial_operaciones()

    # Verificar que haya al menos algunas señales en algún slot
    hay_senales = any(
        len(datos_senales.get("senales_por_slot", {}).get(slot_id, [])) > 0
        for slot_id in ["1", "2", "3", "4", "5", "6"]
    )
    if not hay_senales:
        messagebox.showinfo("Sin datos", "No hay señales guardadas.\nGenera señales primero con el botón 'Generar Señales'.")
        return

    # Cargar datos de precios del log (portable)
    precios_df = None
    fechas_con_cierre = set()
    if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
        try:
            precios_df = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
            precios_df['Date'] = pd.to_datetime(precios_df['Date']).dt.strftime('%Y-%m-%d')
            fechas_con_cierre = set(precios_df['Date'].unique())
            print(f"[INFO] Precios cargados desde: {AUTO_UPDATE_LOG_PORTABLE}")
        except Exception as e:
            print(f"[WARN] No se pudo cargar log de precios: {e}")

    # Filtrar señales: solo mostrar las que tienen precio de cierre en el log
    if fechas_con_cierre:
        for slot_id in ["1", "2", "3", "4", "5", "6"]:
            senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
            senales_filtradas = [
                s for s in senales_slot
                if s.get("fecha_generacion", "")[:10] in fechas_con_cierre
            ]
            datos_senales["senales_por_slot"][slot_id] = senales_filtradas

    # Crear ventana
    ventana_comp = tk.Toplevel(root)
    ventana_comp.title("Comparación: Señales vs Operaciones Reales")
    ventana_comp.geometry("1450x700")

    # Contar señales totales
    total_senales = sum(
        len(datos_senales.get("senales_por_slot", {}).get(slot_id, []))
        for slot_id in ["1", "2", "3", "4", "5", "6"]
    )

    # Frame superior con info
    frame_info = tk.Frame(ventana_comp, pady=5)
    frame_info.pack(fill="x", padx=10)

    lbl_totales = tk.Label(frame_info, text=f"Total señales: {total_senales}  |  Total operaciones: {len(operaciones)}",
             font=("Arial", 10, "bold"))
    lbl_totales.pack(side="left")

    # Recopilar tickers, fechas y plataformas unicos para filtros
    todos_tickers = set()
    todas_fechas = set()
    todas_plataformas = set()
    for slot_id in ["1", "2", "3", "4", "5", "6"]:
        for sen in datos_senales.get("senales_por_slot", {}).get(slot_id, []):
            todos_tickers.add(sen.get("symbol", ""))
            todas_fechas.add(sen.get("fecha_generacion", "")[:10])
            todas_plataformas.add(sen.get("plataforma", "TYBA"))
    lista_tickers = ["Todos"] + sorted(todos_tickers)
    lista_fechas = ["Todos"] + sorted(todas_fechas, reverse=True)
    lista_plataformas = ["Todos"] + sorted(todas_plataformas)

    # Frame de filtros
    frame_filtros = tk.Frame(ventana_comp, pady=3)
    frame_filtros.pack(fill="x", padx=10)

    tk.Label(frame_filtros, text="Filtrar por:", font=("Arial", 9)).pack(side="left", padx=(0, 5))

    # Filtro Plataforma
    tk.Label(frame_filtros, text="Plataforma:", font=("Arial", 9)).pack(side="left")
    combo_filtro_plat = ttk.Combobox(frame_filtros, values=lista_plataformas, state="readonly", width=10)
    combo_filtro_plat.set("Todos")
    combo_filtro_plat.pack(side="left", padx=(2, 10))

    # Filtro Modo (Paper/Real)
    tk.Label(frame_filtros, text="Modo:", font=("Arial", 9)).pack(side="left")
    combo_filtro_modo = ttk.Combobox(frame_filtros, values=["Todos", "Paper", "Real"], state="readonly", width=7)
    combo_filtro_modo.set("Todos")
    combo_filtro_modo.pack(side="left", padx=(2, 10))

    # Filtro Ticker
    tk.Label(frame_filtros, text="Ticker:", font=("Arial", 9)).pack(side="left")
    combo_filtro_ticker = ttk.Combobox(frame_filtros, values=lista_tickers, state="readonly", width=10)
    combo_filtro_ticker.set("Todos")
    combo_filtro_ticker.pack(side="left", padx=(2, 10))

    # Filtro Fecha
    tk.Label(frame_filtros, text="Fecha:", font=("Arial", 9)).pack(side="left")
    combo_filtro_fecha = ttk.Combobox(frame_filtros, values=lista_fechas, state="readonly", width=12)
    combo_filtro_fecha.set("Todos")
    combo_filtro_fecha.pack(side="left", padx=(2, 10))

    lbl_filtro_count = tk.Label(frame_filtros, text="", font=("Arial", 9), fg="gray")
    lbl_filtro_count.pack(side="left", padx=5)

    # Contenedor para funciones de botones (se llenan después)
    btn_funcs = {'graficar': None, 'exportar': None, 'eliminar': None}

    # Botones a la derecha de los filtros
    tk.Button(frame_filtros, text="Cerrar", command=ventana_comp.destroy).pack(side="right", padx=2)
    tk.Button(frame_filtros, text="Eliminar Sel.",
              command=lambda: btn_funcs['eliminar']() if btn_funcs['eliminar'] else messagebox.showinfo("Espere", "Cargando..."),
              bg="#fd7e14", fg="white", font=("Arial", 8)).pack(side="right", padx=2)
    tk.Button(frame_filtros, text="Exportar Excel",
              command=lambda: btn_funcs['exportar']() if btn_funcs['exportar'] else messagebox.showinfo("Espere", "Cargando..."),
              bg="#28a745", fg="white", font=("Arial", 8)).pack(side="right", padx=2)
    tk.Button(frame_filtros, text="Graficar",
              command=lambda: btn_funcs['graficar']() if btn_funcs['graficar'] else messagebox.showinfo("Espere", "Cargando..."),
              bg="#6f42c1", fg="white", font=("Arial", 8)).pack(side="right", padx=2)

    # Notebook principal con pestañas por slot
    notebook_principal = ttk.Notebook(ventana_comp)
    notebook_principal.pack(fill="both", expand=True, padx=10, pady=5)

    # Diccionario global para mapear items a señales (para eliminación)
    item_to_senal_global = {}
    # Lista global para datos de gráfico
    datos_grafico_global = []
    # Referencias a treeviews para filtrado
    tree_refs = {}

    # Crear pestañas para cada slot
    for slot_id in ["1", "2", "3", "4", "5", "6"]:
        nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
        senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
        cantidad_senales = len(senales_slot)

        # Frame principal del slot
        frame_slot = tk.Frame(notebook_principal)
        texto_tab = f"{nombre_slot} ({cantidad_senales})"
        notebook_principal.add(frame_slot, text=texto_tab)

        if cantidad_senales == 0:
            tk.Label(frame_slot, text="No hay señales en este slot",
                    font=("Arial", 12), fg="gray").pack(expand=True)
            continue

        # Sub-notebook con Señales y Comparación
        sub_notebook = ttk.Notebook(frame_slot)
        sub_notebook.pack(fill="both", expand=True)

        # ===== SUB-PESTAÑA: SEÑALES =====
        frame_senales = tk.Frame(sub_notebook)
        sub_notebook.add(frame_senales, text="Señales")

        scroll_sen_y = tk.Scrollbar(frame_senales, orient="vertical")
        scroll_sen_x = tk.Scrollbar(frame_senales, orient="horizontal")

        cols_sen = ("Fecha", "Symbol", "Cierre fecha", "P.Compra", "Cant.C", "Opc.Compra", "P.Venta", "Cant.V", "Opc.Venta", "Cartera", "Tend.C", "Tend.L")
        tree_senales = ttk.Treeview(frame_senales, columns=cols_sen, show="headings",
                                     selectmode="extended",
                                     yscrollcommand=scroll_sen_y.set, xscrollcommand=scroll_sen_x.set)

        scroll_sen_y.config(command=tree_senales.yview)
        scroll_sen_x.config(command=tree_senales.xview)

        anchos_sen = {"Fecha": 85, "Symbol": 70, "Cierre fecha": 90, "P.Compra": 80, "Cant.C": 55,
                      "Opc.Compra": 85, "P.Venta": 75, "Cant.V": 55, "Opc.Venta": 80, "Cartera": 65, "Tend.C": 55, "Tend.L": 55}
        for col in cols_sen:
            tree_senales.heading(col, text=col)
            tree_senales.column(col, width=anchos_sen.get(col, 80), anchor="center")

        scroll_sen_y.pack(side="right", fill="y")
        scroll_sen_x.pack(side="bottom", fill="x")
        tree_senales.pack(fill="both", expand=True)

        # ===== SUB-PESTAÑA: COMPARACIÓN =====
        frame_comp = tk.Frame(sub_notebook)
        sub_notebook.add(frame_comp, text="Comparación")

        scroll_comp_y = tk.Scrollbar(frame_comp, orient="vertical")
        scroll_comp_x = tk.Scrollbar(frame_comp, orient="horizontal")

        cols_comp = ("Fecha Señal", "Symbol", "Máximo", "Mínimo", "Cierre fecha", "P.Compra", "P.Venta", "Recomendación", "Tend.C", "Tend.L", "Fecha Op.", "Tipo Real", "Precio Real", "Seguida")
        tree_comp = ttk.Treeview(frame_comp, columns=cols_comp, show="headings",
                                  yscrollcommand=scroll_comp_y.set, xscrollcommand=scroll_comp_x.set)

        scroll_comp_y.config(command=tree_comp.yview)
        scroll_comp_x.config(command=tree_comp.xview)

        anchos_comp = {"Fecha Señal": 90, "Symbol": 70, "Máximo": 80, "Mínimo": 80, "Cierre fecha": 90,
                       "P.Compra": 80, "P.Venta": 80, "Recomendación": 95, "Tend.C": 55, "Tend.L": 55, "Fecha Op.": 90,
                       "Tipo Real": 75, "Precio Real": 85, "Seguida": 70}
        for col in cols_comp:
            tree_comp.heading(col, text=col)
            tree_comp.column(col, width=anchos_comp.get(col, 80), anchor="center")

        scroll_comp_y.pack(side="right", fill="y")
        scroll_comp_x.pack(side="bottom", fill="x")
        tree_comp.pack(fill="both", expand=True)

        # Guardar referencias
        tree_refs[slot_id] = {"senales": tree_senales, "comp": tree_comp, "nombre": nombre_slot}

    def poblar_arboles(filtro_plataforma="Todos", filtro_modo="Todos", filtro_ticker="Todos", filtro_fecha="Todos"):
        """Limpia y repuebla todos los treeviews segun los filtros seleccionados"""
        item_to_senal_global.clear()
        datos_grafico_global.clear()
        total_mostradas = 0

        for slot_id, refs in tree_refs.items():
            tree_sen = refs["senales"]
            tree_cmp = refs["comp"]
            nombre_slot = refs["nombre"]

            # Limpiar arboles
            tree_sen.delete(*tree_sen.get_children())
            tree_cmp.delete(*tree_cmp.get_children())

            senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
            senales_ordenadas = sorted(senales_slot, key=lambda x: (x.get("symbol", "").upper(), x.get("fecha_generacion", "")[:10]))

            # Aplicar filtros por plataforma/modo usando los campos de la señal
            # NOTA: Señales sin plataforma se tratan como TYBA, sin modo se tratan como Real
            if filtro_plataforma != "Todos":
                senales_ordenadas = [s for s in senales_ordenadas
                                    if (s.get("plataforma") or "TYBA") == filtro_plataforma]
            if filtro_modo != "Todos":
                modo_lower = filtro_modo.lower()
                # Señales sin modo se tratan como "Real"
                senales_ordenadas = [s for s in senales_ordenadas
                                    if (s.get("modo") or "Real").lower() == modo_lower]

            # Filtrar por tickers válidos de la plataforma/modo
            if filtro_plataforma != "Todos" and filtro_modo != "Todos":
                tickers_validos = obtener_tickers_plataforma(filtro_plataforma, filtro_modo)
                if tickers_validos:
                    senales_ordenadas = [s for s in senales_ordenadas
                                        if s.get("symbol", "") in tickers_validos]
            if filtro_ticker != "Todos":
                senales_ordenadas = [s for s in senales_ordenadas if s.get("symbol", "") == filtro_ticker]
            if filtro_fecha != "Todos":
                senales_ordenadas = [s for s in senales_ordenadas if s.get("fecha_generacion", "")[:10] == filtro_fecha]

            count_slot = len(senales_ordenadas)
            total_mostradas += count_slot

            # Actualizar texto de la pestaña
            idx_tab = int(slot_id) - 1
            notebook_principal.tab(idx_tab, text=f"{nombre_slot} ({count_slot})")

            # Poblar pestaña Señales
            for sen in senales_ordenadas:
                fecha_completa = sen.get("fecha_generacion", "")
                fecha_senal = fecha_completa[:10]
                symbol = sen.get("symbol", "")

                cierre_real = "-"
                if precios_df is not None:
                    precio_row = precios_df[(precios_df['Date'] == fecha_senal) & (precios_df['Ticker'] == symbol)]
                    if not precio_row.empty:
                        cierre_real = f"${precio_row['Close'].iloc[0]:.2f}"

                item_id = tree_sen.insert("", "end", values=(
                    fecha_senal,
                    symbol,
                    cierre_real,
                    f"${(sen.get('precio_compra_sugerido') or 0):.2f}",
                    sen.get("cant_compra", "-"),
                    sen.get("opc_compra", ""),
                    f"${(sen.get('precio_venta_sugerido') or 0):.2f}",
                    sen.get("cant_venta", "-"),
                    sen.get("opc_venta", ""),
                    sen.get("acciones_cartera", 0),
                    sen.get("tendencia", "N/A"),
                    sen.get("tendencia_larga", "N/A")
                ))
                item_to_senal_global[item_id] = {
                    "fecha_generacion": fecha_completa,
                    "symbol": sen.get("symbol", ""),
                    "precio_cierre": sen.get("precio_cierre", 0),
                    "slot_id": slot_id
                }

            # Poblar pestaña Comparación
            for sen in senales_ordenadas:
                fecha_sen = sen.get("fecha_generacion", "")[:10]
                symbol = sen.get("symbol", "")

                precio_max = 0
                precio_min = 0
                precio_cierre = sen.get("precio_cierre", 0)
                datos_disponibles = False

                if precios_df is not None:
                    precio_dia = precios_df[(precios_df['Date'] == fecha_sen) & (precios_df['Ticker'] == symbol)]
                    if not precio_dia.empty:
                        precio_max = precio_dia['High'].values[0]
                        precio_min = precio_dia['Low'].values[0]
                        precio_cierre = precio_dia['Close'].values[0]
                        if pd.notna(precio_max) and pd.notna(precio_min) and pd.notna(precio_cierre):
                            datos_disponibles = True

                if not datos_disponibles:
                    continue

                precio_compra_sug = sen.get("precio_compra_sugerido") or 0
                precio_venta_sug = sen.get("precio_venta_sugerido") or 0

                if sen.get("opc_compra") == "Comprar":
                    recomendacion = "Comprar"
                elif sen.get("opc_venta") == "Vender":
                    recomendacion = "Vender"
                else:
                    recomendacion = "Sin acción"

                op_encontrada = None
                for op in operaciones:
                    if op.get("ticker_symbol") == symbol:
                        fecha_op = op.get("fecha", "")
                        if fecha_op >= fecha_sen:
                            try:
                                from datetime import timedelta
                                fecha_sen_dt = datetime.strptime(fecha_sen, "%Y-%m-%d")
                                fecha_op_dt = datetime.strptime(fecha_op, "%Y-%m-%d")
                                if (fecha_op_dt - fecha_sen_dt).days <= 2:
                                    op_encontrada = op
                                    break
                            except:
                                pass

                if op_encontrada:
                    tipo_real = op_encontrada.get("tipo", "").capitalize()
                    precio_real = op_encontrada.get("precio") or 0
                    fecha_op_str = op_encontrada.get("fecha", "")
                    seguida = "SI" if recomendacion.lower() == tipo_real.lower() else "NO"
                else:
                    tipo_real = "-"
                    precio_real = 0
                    fecha_op_str = "-"
                    seguida = "Pendiente"

                tendencia_sen = sen.get("tendencia", "N/A")
                tendencia_larga_sen = sen.get("tendencia_larga", "N/A")
                tree_cmp.insert("", "end", values=(
                    fecha_sen,
                    symbol,
                    f"${precio_max:.2f}" if precio_max > 0 else "-",
                    f"${precio_min:.2f}" if precio_min > 0 else "-",
                    f"${precio_cierre:.2f}" if precio_cierre > 0 else "-",
                    f"${precio_compra_sug:.2f}" if precio_compra_sug > 0 else "-",
                    f"${precio_venta_sug:.2f}" if precio_venta_sug > 0 else "-",
                    recomendacion,
                    tendencia_sen,
                    tendencia_larga_sen,
                    fecha_op_str,
                    tipo_real,
                    f"${precio_real:.2f}" if precio_real > 0 else "-",
                    seguida
                ))

                datos_grafico_global.append({
                    'fecha': fecha_sen,
                    'symbol': symbol,
                    'maximo': precio_max,
                    'minimo': precio_min,
                    'cierre': precio_cierre,
                    'precio_compra': precio_compra_sug,
                    'precio_venta': precio_venta_sug,
                    'recomendacion': recomendacion,
                    'tendencia': tendencia_sen,
                    'tendencia_larga': tendencia_larga_sen,
                    'slot_id': slot_id,
                    'slot_nombre': nombre_slot
                })

        # Actualizar etiqueta de filtro
        if filtro_plataforma != "Todos" or filtro_modo != "Todos" or filtro_ticker != "Todos" or filtro_fecha != "Todos":
            lbl_filtro_count.config(text=f"(Mostrando {total_mostradas} de {total_senales})")
        else:
            lbl_filtro_count.config(text="")

    def on_filtro_change(event=None):
        """Callback cuando cambia un filtro"""
        poblar_arboles(combo_filtro_plat.get(), combo_filtro_modo.get(), combo_filtro_ticker.get(), combo_filtro_fecha.get())

    combo_filtro_plat.bind("<<ComboboxSelected>>", on_filtro_change)
    combo_filtro_modo.bind("<<ComboboxSelected>>", on_filtro_change)
    combo_filtro_ticker.bind("<<ComboboxSelected>>", on_filtro_change)
    combo_filtro_fecha.bind("<<ComboboxSelected>>", on_filtro_change)

    # Poblar árboles inicialmente (sin filtro)
    try:
        poblar_arboles()
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error en poblar_arboles:\n{error_detail}")
        messagebox.showerror("Error en poblar_arboles", f"{e}\n\nVer consola para detalles.")

    # Definir funciones para los botones

    def exportar_comparacion_excel():
        """Exporta la comparación a Excel con hojas por slot + operaciones"""
        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Comparación",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Comparacion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )

        if not ruta_excel:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            si_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            primera_hoja = True

            # Crear una hoja por cada slot con señales
            for slot_id in ["1", "2", "3", "4", "5", "6"]:
                nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])

                if not senales_slot:
                    continue

                # Crear hoja
                if primera_hoja:
                    ws = wb.active
                    ws.title = f"Slot {nombre_slot}"
                    primera_hoja = False
                else:
                    ws = wb.create_sheet(f"Slot {nombre_slot}")

                headers = ["Fecha", "Symbol", "Cierre", "P.Compra", "Cant.C", "Opc.Compra",
                          "P.Venta", "Cant.V", "Opc.Venta", "Cartera", "Tend.C", "Tend.L", "Slot"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border

                senales_ordenadas = sorted(senales_slot, key=lambda x: (x.get("symbol", "").upper(), x.get("fecha_generacion", "")[:10]))

                for row_idx, sen in enumerate(senales_ordenadas, 2):
                    ws.cell(row=row_idx, column=1, value=sen.get("fecha_generacion", "")[:10]).border = border
                    ws.cell(row=row_idx, column=2, value=sen.get("symbol", "")).border = border
                    ws.cell(row=row_idx, column=3, value=sen.get("precio_cierre", 0)).border = border
                    ws.cell(row=row_idx, column=4, value=sen.get("precio_compra_sugerido", 0)).border = border
                    ws.cell(row=row_idx, column=5, value=sen.get("cant_compra", "-")).border = border
                    ws.cell(row=row_idx, column=6, value=sen.get("opc_compra", "")).border = border
                    ws.cell(row=row_idx, column=7, value=sen.get("precio_venta_sugerido", 0)).border = border
                    ws.cell(row=row_idx, column=8, value=sen.get("cant_venta", "-")).border = border
                    ws.cell(row=row_idx, column=9, value=sen.get("opc_venta", "")).border = border
                    ws.cell(row=row_idx, column=10, value=sen.get("acciones_cartera", 0)).border = border
                    ws.cell(row=row_idx, column=11, value=sen.get("tendencia", "N/A")).border = border
                    ws.cell(row=row_idx, column=12, value=sen.get("tendencia_larga", "N/A")).border = border
                    ws.cell(row=row_idx, column=13, value=nombre_slot).border = border

            # Hoja de Comparación (global con datos de gráfico)
            if primera_hoja:
                ws_comp = wb.active
                ws_comp.title = "Comparación"
            else:
                ws_comp = wb.create_sheet("Comparación")

            headers_comp = ["Fecha Señal", "Symbol", "Slot", "Máximo", "Mínimo", "Cierre",
                           "P.Compra", "P.Venta", "Recomendación", "Tend.C", "Tend.L",
                           "Fecha Op.", "Tipo Real", "Precio Real", "Seguida"]
            for col_idx, header in enumerate(headers_comp, 1):
                cell = ws_comp.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            row_idx = 2
            for dato in datos_grafico_global:
                fecha_sen = dato['fecha']
                symbol = dato['symbol']
                recomendacion = dato['recomendacion']
                slot_id_dato = dato.get('slot_id', '1')
                nombre_slot_dato = obtener_nombre_slot(datos_slots, slot_id_dato)

                op_encontrada = None
                for op in operaciones:
                    if op.get("ticker_symbol") == symbol:
                        fecha_op = op.get("fecha", "")
                        if fecha_op >= fecha_sen:
                            try:
                                fecha_sen_dt = datetime.strptime(fecha_sen, "%Y-%m-%d")
                                fecha_op_dt = datetime.strptime(fecha_op, "%Y-%m-%d")
                                if (fecha_op_dt - fecha_sen_dt).days <= 2:
                                    op_encontrada = op
                                    break
                            except:
                                pass

                if op_encontrada:
                    tipo_real = op_encontrada.get("tipo", "").capitalize()
                    precio_real = op_encontrada.get("precio", 0)
                    fecha_op_str = op_encontrada.get("fecha", "")
                    seguida = "SI" if recomendacion.lower() == tipo_real.lower() else "NO"
                else:
                    tipo_real = "-"
                    precio_real = 0
                    fecha_op_str = "-"
                    seguida = "Pendiente"

                ws_comp.cell(row=row_idx, column=1, value=fecha_sen).border = border
                ws_comp.cell(row=row_idx, column=2, value=symbol).border = border
                ws_comp.cell(row=row_idx, column=3, value=nombre_slot_dato).border = border
                ws_comp.cell(row=row_idx, column=4, value=dato['maximo'] if dato['maximo'] > 0 else "-").border = border
                ws_comp.cell(row=row_idx, column=5, value=dato['minimo'] if dato['minimo'] > 0 else "-").border = border
                ws_comp.cell(row=row_idx, column=6, value=dato['cierre'] if dato['cierre'] > 0 else "-").border = border
                ws_comp.cell(row=row_idx, column=7, value=dato['precio_compra'] if dato['precio_compra'] > 0 else "-").border = border
                ws_comp.cell(row=row_idx, column=8, value=dato['precio_venta'] if dato['precio_venta'] > 0 else "-").border = border
                ws_comp.cell(row=row_idx, column=9, value=recomendacion).border = border
                ws_comp.cell(row=row_idx, column=10, value=dato.get('tendencia', 'N/A')).border = border
                ws_comp.cell(row=row_idx, column=11, value=dato.get('tendencia_larga', 'N/A')).border = border
                ws_comp.cell(row=row_idx, column=12, value=fecha_op_str).border = border
                ws_comp.cell(row=row_idx, column=13, value=tipo_real).border = border
                ws_comp.cell(row=row_idx, column=14, value=precio_real if precio_real > 0 else "-").border = border

                cell_seguida = ws_comp.cell(row=row_idx, column=15, value=seguida)
                cell_seguida.border = border
                if seguida == "SI":
                    cell_seguida.fill = si_fill
                elif seguida == "NO":
                    cell_seguida.fill = no_fill

                row_idx += 1

            # Ajustar anchos
            for ws in wb.worksheets:
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 14

            wb.save(ruta_excel)
            messagebox.showinfo("Exportado", f"Comparación exportada a:\n{ruta_excel}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")

    btn_funcs['exportar'] = exportar_comparacion_excel

    def graficar_datos():
        """Abre ventana con gráfico de precios y señales"""
        if not datos_grafico_global:
            messagebox.showinfo("Sin datos", "No hay datos para graficar")
            return

        # Cargar operaciones reales para mostrar en el gráfico
        operaciones_reales = []
        try:
            ruta_ops = obtener_ruta_operaciones()
            if os.path.exists(ruta_ops):
                with open(ruta_ops, 'r', encoding='utf-8') as f:
                    datos_ops = json.load(f)
                    operaciones_reales = datos_ops.get("operaciones", [])
        except:
            pass

        # Obtener símbolos únicos (ordenados alfabéticamente)
        symbols = sorted(list(set(d['symbol'] for d in datos_grafico_global)))

        # Obtener parámetros únicos disponibles (slot_id -> nombre)
        param_nombres = {}
        for d in datos_grafico_global:
            slot_id = d.get('slot_id', '1')
            slot_nombre = d.get('slot_nombre', slot_id)
            if slot_id not in param_nombres:
                param_nombres[slot_id] = slot_nombre

        # Crear lista ordenada de nombres de parámetros
        param_opciones_ordenadas = sorted(param_nombres.items(), key=lambda x: x[0])
        param_nombres_lista = [nombre for _, nombre in param_opciones_ordenadas]

        # Crear ventana de selección de ticker
        ventana_graf = tk.Toplevel(ventana_comp)
        ventana_graf.title("Graficar Precios y Señales")
        ventana_graf.geometry("900x650")

        # Frame superior para selección
        frame_sel = tk.Frame(ventana_graf, pady=10)
        frame_sel.pack(fill="x", padx=10)

        tk.Label(frame_sel, text="Ticker:", font=("Arial", 10)).pack(side="left", padx=5)

        ticker_var = tk.StringVar(value=symbols[0] if symbols else "")
        combo_ticker = ttk.Combobox(frame_sel, textvariable=ticker_var, values=symbols, state="readonly", width=10)
        combo_ticker.pack(side="left", padx=5)

        tk.Label(frame_sel, text="Parámetro:", font=("Arial", 10)).pack(side="left", padx=(15, 5))

        # Iniciar con el primer parámetro (sin opción "Todos")
        primer_param = param_nombres_lista[0] if param_nombres_lista else ""
        param_var = tk.StringVar(value=primer_param)
        combo_param = ttk.Combobox(frame_sel, textvariable=param_var, values=param_nombres_lista, state="readonly", width=20)
        combo_param.pack(side="left", padx=5)

        # Checkbox para mostrar/ocultar tendencias
        mostrar_tendencias_var = tk.BooleanVar(value=True)
        chk_tendencias = tk.Checkbutton(frame_sel, text="Tendencias", variable=mostrar_tendencias_var,
                                        font=("Arial", 9))
        chk_tendencias.pack(side="left", padx=(15, 5))

        # Checkbox para mostrar/ocultar línea de tendencia lineal
        mostrar_linea_tendencia_var = tk.BooleanVar(value=False)
        chk_linea_tend = tk.Checkbutton(frame_sel, text="Línea Tend.", variable=mostrar_linea_tendencia_var,
                                        font=("Arial", 9))
        chk_linea_tend.pack(side="left", padx=(5, 5))

        # Checkbox para mostrar/ocultar promedio móvil 5 días
        mostrar_pm5_var = tk.BooleanVar(value=False)
        chk_pm5 = tk.Checkbutton(frame_sel, text="PM 5d", variable=mostrar_pm5_var,
                                 font=("Arial", 9))
        chk_pm5.pack(side="left", padx=(5, 5))

        # Checkbox para mostrar/ocultar precios sugeridos
        mostrar_sugeridos_var = tk.BooleanVar(value=True)
        chk_sugeridos = tk.Checkbutton(frame_sel, text="P.Sug.", variable=mostrar_sugeridos_var,
                                       font=("Arial", 9))
        chk_sugeridos.pack(side="left", padx=(5, 5))

        # Checkbox para mostrar/ocultar máximo y mínimo
        mostrar_maxmin_var = tk.BooleanVar(value=True)
        chk_maxmin = tk.Checkbutton(frame_sel, text="Max/Min", variable=mostrar_maxmin_var,
                                    font=("Arial", 9))
        chk_maxmin.pack(side="left", padx=(5, 5))

        # Separador visual
        tk.Label(frame_sel, text="|", font=("Arial", 9), fg="gray").pack(side="left", padx=(10, 5))

        # Combobox para rango de fechas
        tk.Label(frame_sel, text="Rango:", font=("Arial", 9)).pack(side="left")
        rango_var = tk.StringVar(value="Completo")
        combo_rango = ttk.Combobox(frame_sel, textvariable=rango_var,
                                   values=["Completo", "30 días"], state="readonly", width=10)
        combo_rango.pack(side="left", padx=5)

        # Frame para el gráfico
        frame_grafico = tk.Frame(ventana_graf)
        frame_grafico.pack(fill="both", expand=True, padx=10, pady=5)

        # Figura de matplotlib
        fig, ax = plt.subplots(figsize=(10, 5))
        fig.subplots_adjust(left=0.06, right=0.94)  # Reducir espacio izquierdo
        ax2 = ax.twinx()  # Crear eje secundario una sola vez
        canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
        canvas.get_tk_widget().pack(fill="both", expand=True)

        def actualizar_grafico(*args):
            ax.clear()
            ax2.clear()  # Limpiar también el eje secundario
            ticker_sel = ticker_var.get()
            param_sel = param_var.get()

            if not ticker_sel or not param_sel:
                return

            # Filtrar datos del ticker y parámetro seleccionado
            datos_ticker = [d for d in datos_grafico_global
                           if d['symbol'] == ticker_sel and d.get('slot_nombre', d.get('slot_id', '1')) == param_sel]

            if not datos_ticker:
                ax.text(0.5, 0.5, 'Sin datos para este ticker/parámetro', ha='center', va='center', transform=ax.transAxes)
                canvas.draw()
                return

            # Ordenar por fecha y eliminar duplicados (mantener solo el primero por fecha)
            datos_ticker = sorted(datos_ticker, key=lambda x: x['fecha'])
            fechas_vistas = set()
            datos_ticker_unicos = []
            for d in datos_ticker:
                if d['fecha'] not in fechas_vistas:
                    fechas_vistas.add(d['fecha'])
                    datos_ticker_unicos.append(d)
            datos_ticker = datos_ticker_unicos

            # Filtrar por rango de fechas si está seleccionado "30 días"
            if rango_var.get() == "30 días" and datos_ticker:
                from datetime import timedelta
                fecha_limite = datetime.now() - timedelta(days=30)
                datos_ticker = [d for d in datos_ticker
                               if datetime.strptime(d['fecha'], '%Y-%m-%d') >= fecha_limite]
                if not datos_ticker:
                    ax.text(0.5, 0.5, 'Sin datos en los últimos 30 días', ha='center', va='center', transform=ax.transAxes)
                    canvas.draw()
                    return

            # Preparar datos
            fechas = [datetime.strptime(d['fecha'], '%Y-%m-%d') for d in datos_ticker]
            maximos = [d['maximo'] for d in datos_ticker]
            minimos = [d['minimo'] for d in datos_ticker]
            cierres = [d['cierre'] for d in datos_ticker]
            precios_compra = [d['precio_compra'] for d in datos_ticker]
            precios_venta = [d['precio_venta'] for d in datos_ticker]

            # Calcular límites fijos del eje Y (basado en todos los datos)
            todos_precios = maximos + minimos + cierres + precios_compra + precios_venta
            todos_precios = [p for p in todos_precios if p > 0]
            if todos_precios:
                y_min = min(todos_precios)
                y_max = max(todos_precios)
                margen = (y_max - y_min) * 0.05
                y_min_fijo = y_min - margen
                y_max_fijo = y_max + margen
            else:
                y_min_fijo, y_max_fijo = None, None

            # Graficar líneas
            if mostrar_maxmin_var.get():
                if any(m > 0 for m in maximos):
                    ax.plot(fechas, maximos, 'g-', label='Máximo', linewidth=1.5, marker='o', markersize=4)
                if any(m > 0 for m in minimos):
                    ax.plot(fechas, minimos, 'r-', label='Mínimo', linewidth=1.5, marker='o', markersize=4)
            if any(c > 0 for c in cierres):
                ax.plot(fechas, cierres, 'b-', label='Cierre', linewidth=2, marker='s', markersize=5)
            if mostrar_sugeridos_var.get():
                if any(p > 0 for p in precios_compra):
                    ax.plot(fechas, precios_compra, 'g--', label='Precio Compra Sugerido', linewidth=1.5, alpha=0.7)
                if any(p > 0 for p in precios_venta):
                    ax.plot(fechas, precios_venta, 'r--', label='Precio Venta Sugerido', linewidth=1.5, alpha=0.7)

            # Graficar línea de tendencia lineal (regresión) si checkbox activado
            if mostrar_linea_tendencia_var.get() and len(cierres) >= 2:
                # Convertir fechas a timestamps para regresión (línea perfectamente recta)
                x_timestamps = np.array([f.timestamp() for f in fechas])
                y_vals = np.array(cierres)

                # Calcular regresión lineal usando timestamps
                n = len(x_timestamps)
                sum_x = np.sum(x_timestamps)
                sum_y = np.sum(y_vals)
                sum_xy = np.sum(x_timestamps * y_vals)
                sum_x2 = np.sum(x_timestamps ** 2)

                denom = n * sum_x2 - sum_x ** 2
                if denom != 0:
                    pendiente = (n * sum_xy - sum_x * sum_y) / denom
                    intercepto = (sum_y - pendiente * sum_x) / n

                    # Calcular línea de tendencia (solo inicio y fin para línea recta perfecta)
                    x_linea = [fechas[0], fechas[-1]]
                    y_linea = [pendiente * x_timestamps[0] + intercepto,
                               pendiente * x_timestamps[-1] + intercepto]

                    # Dibujar línea de tendencia (púrpura punteada con rayas largas)
                    ax.plot(x_linea, y_linea, color='purple', linestyle='--',
                            linewidth=2, label='Tendencia Lineal', alpha=0.8, dashes=[10, 4])

            # Graficar promedio móvil de 5 días si checkbox activado
            if mostrar_pm5_var.get() and len(cierres) >= 5:
                pm5 = []
                for i in range(len(cierres)):
                    if i < 4:  # No hay suficientes datos anteriores
                        pm5.append(None)
                    else:
                        promedio = sum(cierres[i-4:i+1]) / 5
                        pm5.append(promedio)

                # Filtrar valores válidos
                fechas_pm5 = [f for f, p in zip(fechas, pm5) if p is not None]
                valores_pm5 = [p for p in pm5 if p is not None]

                if valores_pm5:
                    ax.plot(fechas_pm5, valores_pm5, color='black', linestyle='-',
                            linewidth=1.5, label='PM 5d', alpha=0.8)

            # Graficar evolución de los valores de tendencia (eje secundario) solo si checkbox activado
            if mostrar_tendencias_var.get():
                tendencias_cortas = []
                tendencias_largas = []
                for d in datos_ticker:
                    # Convertir tendencia a valor numérico (-100 a +100)
                    tc = d.get('tendencia', 'N/A')
                    tl = d.get('tendencia_larga', 'N/A')
                    try:
                        if tc != 'N/A' and tc:
                            tendencias_cortas.append(int(tc.replace('+', '')))
                        else:
                            tendencias_cortas.append(None)
                    except:
                        tendencias_cortas.append(None)
                    try:
                        if tl != 'N/A' and tl:
                            tendencias_largas.append(int(tl.replace('+', '')))
                        else:
                            tendencias_largas.append(None)
                    except:
                        tendencias_largas.append(None)

                # Configurar eje secundario para tendencias
                ax2.yaxis.set_label_position('right')
                ax2.yaxis.tick_right()
                ax2.set_ylabel('Tendencia', color='gray', rotation=270, labelpad=15)
                ax2.set_ylim(-110, 110)
                ax2.axhline(y=0, color='gray', linestyle=':', linewidth=0.5, alpha=0.5)

                # Graficar tendencias (solo valores válidos)
                fechas_tc = [f for f, t in zip(fechas, tendencias_cortas) if t is not None]
                valores_tc = [t for t in tendencias_cortas if t is not None]
                fechas_tl = [f for f, t in zip(fechas, tendencias_largas) if t is not None]
                valores_tl = [t for t in tendencias_largas if t is not None]

                if valores_tc:
                    ax2.plot(fechas_tc, valores_tc, color='orange', linestyle='-',
                            linewidth=1.2, label='Tend.C (10d)', alpha=0.9)
                if valores_tl:
                    ax2.plot(fechas_tl, valores_tl, color='gray', linestyle='-',
                            linewidth=1.2, label='Tend.L (30d)', alpha=0.9)

                ax2.legend(loc='upper right', fontsize=8)
                ax2.tick_params(axis='y', labelcolor='gray')
            else:
                # Ocultar eje secundario cuando no hay tendencias
                ax2.set_yticks([])
                ax2.set_ylabel('')

            # Marcar operaciones reales (compras/ventas ejecutadas)
            ops_ticker = [op for op in operaciones_reales if op.get('ticker_symbol') == ticker_sel]
            compras_reales_x = []
            compras_reales_y = []
            ventas_reales_x = []
            ventas_reales_y = []

            for op in ops_ticker:
                fecha_str = op.get('fecha', '')
                precio_op = op.get('precio', 0)
                tipo_op = op.get('tipo', '')
                if fecha_str and precio_op > 0:
                    try:
                        fecha_op = datetime.strptime(fecha_str, '%Y-%m-%d')
                        if tipo_op == 'compra':
                            compras_reales_x.append(fecha_op)
                            compras_reales_y.append(precio_op)
                        elif tipo_op == 'venta':
                            ventas_reales_x.append(fecha_op)
                            ventas_reales_y.append(precio_op)
                    except ValueError:
                        pass

            # Graficar operaciones reales con marcadores grandes
            if compras_reales_x:
                ax.scatter(compras_reales_x, compras_reales_y, marker='^', s=150, c='lime',
                          edgecolors='darkgreen', linewidths=2, label='Compra Real', zorder=5)
            if ventas_reales_x:
                ax.scatter(ventas_reales_x, ventas_reales_y, marker='v', s=150, c='salmon',
                          edgecolors='darkred', linewidths=2, label='Venta Real', zorder=5)

            ax.set_title(f'Precios y Señales - {ticker_sel} ({param_sel})', fontsize=12, fontweight='bold')
            ax.set_xlabel('Fecha')
            ax.set_ylabel('Precio ($)')
            if y_min_fijo is not None and y_max_fijo is not None:
                ax.set_ylim(y_min_fijo, y_max_fijo)

            # Ajustar eje X al rango de datos (importante para "30 días")
            if fechas:
                from datetime import timedelta as td
                margen_dias = td(days=1)
                ax.set_xlim(fechas[0] - margen_dias, fechas[-1] + margen_dias)

            ax.legend(loc='upper left', fontsize=8)
            ax.grid(True, alpha=0.3)

            # Formato de fechas - ajustar intervalo según rango
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m'))
            if rango_var.get() == "30 días":
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=2))  # Cada 2 días para 30 días
            else:
                ax.xaxis.set_major_locator(mdates.DayLocator(interval=3))  # Cada 3 días para completo
            plt.setp(ax.xaxis.get_majorticklabels(), fontsize=8, rotation=45)

            canvas.draw()

        # Vincular cambio de ticker, parámetro y checkboxes
        combo_ticker.bind('<<ComboboxSelected>>', actualizar_grafico)
        combo_param.bind('<<ComboboxSelected>>', actualizar_grafico)
        combo_rango.bind('<<ComboboxSelected>>', actualizar_grafico)
        chk_tendencias.config(command=actualizar_grafico)
        chk_linea_tend.config(command=actualizar_grafico)
        chk_pm5.config(command=actualizar_grafico)
        chk_sugeridos.config(command=actualizar_grafico)
        chk_maxmin.config(command=actualizar_grafico)

        # Botón guardar imagen
        def guardar_imagen():
            ruta_img = filedialog.asksaveasfilename(
                title="Guardar Gráfico",
                defaultextension=".png",
                filetypes=[("PNG files", "*.png"), ("PDF files", "*.pdf")],
                initialfile=f"Grafico_{ticker_var.get()}_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
            )
            if ruta_img:
                fig.savefig(ruta_img, dpi=150, bbox_inches='tight')
                messagebox.showinfo("Guardado", f"Gráfico guardado en:\n{ruta_img}")

        tk.Button(frame_sel, text="Guardar Imagen", command=guardar_imagen,
                  bg="#6c757d", fg="white").pack(side="left", padx=5)

        # Frame inferior
        frame_inf = tk.Frame(ventana_graf, pady=5)
        frame_inf.pack(fill="x", padx=10)

        tk.Label(frame_inf, text="▲ = Compra real | ▼ = Venta real | Naranja = Tend.10d | Negro = Tend.30d", font=("Arial", 9), fg="gray").pack(side="left")
        tk.Button(frame_inf, text="Cerrar", command=ventana_graf.destroy).pack(side="right")

        # Graficar el primer ticker
        actualizar_grafico()

    btn_funcs['graficar'] = graficar_datos

    def eliminar_senales_seleccionadas():
        """Elimina las señales seleccionadas (nota: esta función está deshabilitada en la nueva estructura de pestañas)"""
        messagebox.showinfo("Info", "Para eliminar señales, usa 'Limpiar Todo' o regenera las señales.\nLa eliminación individual no está disponible en la vista por slots.")

    btn_funcs['eliminar'] = eliminar_senales_seleccionadas


def seleccionar_csv():
    # Obtener ruta guardada para usar como directorio inicial
    ruta_guardada = cargar_ruta_csv()
    initial_dir = os.path.dirname(ruta_guardada) if ruta_guardada else None

    ruta = filedialog.asksaveasfilename(
        title="Selecciona o crea el archivo CSV",
        defaultextension=".csv",
        filetypes=[("CSV files", "*.csv"), ("Todos los archivos", "*.*")],
        initialdir=initial_dir
    )
    if ruta:
        entry_ruta.delete(0, tk.END)
        entry_ruta.insert(0, ruta)
        # Guardar la ruta para la próxima vez
        guardar_ruta_csv(ruta)

def validar_formato_yfinance(data):
    """
    Valida el formato de datos devuelto por yfinance y detecta cambios.
    Retorna (es_valido, advertencias)
    """
    advertencias = []

    # Detectar MultiIndex en columnas
    if isinstance(data.columns, pd.MultiIndex):
        niveles = data.columns.nlevels
        if niveles > 2:
            advertencias.append(
                f"yfinance devolvió MultiIndex con {niveles} niveles (esperado: 2)")

    # Obtener columnas para verificar
    if isinstance(data.columns, pd.MultiIndex):
        cols = set(data.columns.get_level_values(-1).unique())
    else:
        cols = set(data.columns)

    # Verificar columnas críticas
    cols_criticas = {"Open", "High", "Low", "Close"}
    faltantes = cols_criticas - cols
    if faltantes:
        advertencias.append(f"Faltan columnas críticas: {faltantes}")

    # Verificar índice
    if not isinstance(data.index, pd.DatetimeIndex):
        advertencias.append(f"Índice inesperado: {type(data.index).__name__}")

    es_valido = len(advertencias) == 0
    return es_valido, advertencias


def actualizar_csv():
    """
    Descarga y actualiza datos de precios.
    - Antes de 16:00 NY: muestra advertencia de precios preliminares
    - Después de 16:00 NY: sobrescribe automáticamente datos de hoy
    """
    # Verificar que las bibliotecas necesarias estén cargadas
    if not verificar_libs_cargadas(["yfinance", "pandas"]):
        messagebox.showwarning("Esperar", "Esperar que se carguen los recursos del sistema.")
        return

    csv_file = entry_ruta.get()
    if not csv_file:
        label_status.config(text="Selecciona primero la ruta del CSV", fg="red")
        return

    # Verificar hora de NY
    now_ny = datetime.now(ZoneInfo("America/New_York"))
    hora_ny = now_ny.hour
    es_fin_de_semana = now_ny.weekday() >= 5  # 5=Sábado, 6=Domingo

    # Después de 16:00 = sobrescribir automáticamente (precios de cierre definitivos)
    forzar_actualizacion = (hora_ny >= 16)

    if es_fin_de_semana:
        respuesta = messagebox.askyesno(
            "Fin de semana",
            f"Hoy es {['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo'][now_ny.weekday()]}.\n"
            "El mercado está cerrado los fines de semana.\n\n"
            "¿Deseas descargar los datos del viernes pasado?"
        )
        if not respuesta:
            return
    elif hora_ny < 16:
        respuesta = messagebox.askyesno(
            "Mercado aún abierto",
            f"Hora actual en NY: {now_ny.strftime('%H:%M')}\n\n"
            "El mercado cierra a las 16:00 NY.\n"
            "Los precios descargados ahora son PRELIMINARES\n"
            "(el precio 'Close' será el último precio negociado, no el de cierre).\n\n"
            "¿Deseas descargar los precios preliminares?\n\n"
            "Después de las 16:00, los datos se sobrescribirán\n"
            "automáticamente con los precios de cierre definitivos."
        )
        if not respuesta:
            return

    try:
        print("\n=== INICIO ACTUALIZACIÓN ===")

        print("[1] Descargando datos de Yahoo Finance...")
        data = yf.download(tickers, period="1d", group_by='ticker', auto_adjust=False)
        print("[2] Descarga completada.")

        # Validar formato de yfinance
        es_valido, advertencias = validar_formato_yfinance(data)
        if advertencias:
            msg_aviso = "AVISO: Posible cambio en formato de yfinance:\n\n"
            msg_aviso += "\n".join(f"• {a}" for a in advertencias)
            msg_aviso += "\n\nEl script intentará continuar, pero revisa los datos."
            print(f"[!] {msg_aviso}")
            messagebox.showwarning("Aviso yfinance", msg_aviso)
            if not es_valido:
                label_status.config(text="Error: formato yfinance cambió", fg="red")
                return

        records = []
        for ticker in tickers:
            try:
                if hasattr(data.columns, "levels") and ticker in data.columns.levels[0]:
                    df = data[ticker].copy()
                    df.reset_index(inplace=True)
                    # Manejar MultiIndex de columnas (cambio en yfinance)
                    if isinstance(df.columns, pd.MultiIndex):
                        df.columns = df.columns.get_level_values(0)
                    if 'Adj Close' in df.columns:
                        df.rename(columns={'Adj Close':'Close'}, inplace=True)
                    df['Ticker'] = ticker
                    if not df.empty and 'Close' in df.columns:
                        close_val = df['Close'].iloc[0]
                        if pd.notna(close_val).any() if hasattr(close_val, '__iter__') else pd.notna(close_val):
                            records.append(df[['Date','Ticker','Open','High','Low','Close']])
                else:
                    # Caso de un solo ticker
                    cols_to_check = data.columns.get_level_values(0) if isinstance(data.columns, pd.MultiIndex) else data.columns
                    if 'Open' in cols_to_check and 'Close' in cols_to_check:
                        tmp = data.reset_index().copy()
                        if isinstance(tmp.columns, pd.MultiIndex):
                            tmp.columns = tmp.columns.get_level_values(0)
                        if 'Adj Close' in tmp.columns:
                            tmp.rename(columns={'Adj Close':'Close'}, inplace=True)
                        tmp['Ticker'] = ticker
                        if not tmp.empty:
                            records.append(tmp[['Date','Ticker','Open','High','Low','Close']])
                        break
            except Exception as e:
                print(f"[WARN] Error procesando {ticker}: {e}")

        if not records:
            print("[X] No se encontraron datos.")
            label_status.config(text="No hay datos nuevos disponibles hoy.", fg="blue")
            return

        df_long = pd.concat(records, ignore_index=True)
        df_long = df_long.loc[:, ~df_long.columns.duplicated()]
        df_long['Date'] = pd.to_datetime(df_long['Date']).dt.normalize()

        # ===========================
        # CREAR CSV PRINCIPAL
        # ===========================
        if not os.path.exists(csv_file):
            print("[3] CSV no existe, se creará uno nuevo.")
        else:
            print("[3] CSV ya existe, será sobrescrito con la data descargada.")

        print("[5] Creando CSV con la data descargada...")
        df_long.to_csv(csv_file, index=False, float_format="%.2f")
        print("[6] CSV guardado correctamente.")


        # ===========================
        # ACTUALIZAR LOG AUXILIAR
        # ===========================
        log_file = os.path.join(os.path.dirname(csv_file), "auto_update_log.csv")
        print(f"[7] Actualizando log auxiliar: {log_file}")

        df_long_for_log = df_long.copy()
        df_long_for_log['Date'] = pd.to_datetime(df_long_for_log['Date']).dt.normalize()

        if os.path.exists(log_file):
            print("[8] Leyendo log existente...")
            df_log_existing = pd.read_csv(log_file, parse_dates=['Date'])
            df_log_existing = df_log_existing.loc[:, ~df_log_existing.columns.duplicated()]
            df_log_existing['Date'] = pd.to_datetime(df_log_existing['Date']).dt.normalize()

            # Si forzar_actualizacion, eliminar datos de hoy antes de agregar nuevos
            if forzar_actualizacion:
                fecha_hoy = df_long_for_log['Date'].iloc[0]
                filas_antes = len(df_log_existing)
                df_log_existing = df_log_existing[df_log_existing['Date'] != fecha_hoy]
                filas_eliminadas = filas_antes - len(df_log_existing)
                if filas_eliminadas > 0:
                    print(f"[8.1] Sobrescribiendo: eliminados {filas_eliminadas} registros de {fecha_hoy.strftime('%Y-%m-%d')}")

            existing_keys = set(zip(
                df_log_existing['Date'].dt.strftime('%Y-%m-%d'),
                df_log_existing['Ticker']
            ))

            keys_series = df_long_for_log[['Date','Ticker']].apply(
                lambda r: (r['Date'].strftime('%Y-%m-%d'), r['Ticker']), axis=1
            )

            mask_new = ~keys_series.isin(existing_keys)
            df_log_new = df_long_for_log.loc[mask_new].copy()

            if not df_log_new.empty:
                print(f"[9] Agregando {len(df_log_new)} filas nuevas al log.")
                df_log_to_save = pd.concat([df_log_existing, df_log_new], ignore_index=True)
            else:
                if forzar_actualizacion:
                    # Si se forzó actualización pero no hay filas "nuevas", agregar igualmente
                    print(f"[9] Sobrescribiendo {len(df_long_for_log)} registros de hoy.")
                    df_log_to_save = pd.concat([df_log_existing, df_long_for_log], ignore_index=True)
                else:
                    print("[9] No hay filas nuevas para agregar al log.")
                    df_log_to_save = df_log_existing.copy()

        else:
            print("[8] Log no existe. Creándolo desde cero.")
            df_log_to_save = df_long_for_log.copy()

        print("[10] Guardando log auxiliar...")
        df_log_to_save.to_csv(log_file, index=False, float_format="%.2f")
        print("[11] Log guardado correctamente.")

        # Liberar memoria
        gc.collect()

        # Hora NY
        now_ny = datetime.now(ZoneInfo("America/New_York"))
        fecha_hora_ny = now_ny.strftime("%Y-%m-%d %H:%M")
        label_status.config(
            text=f"CSV actualizado con fecha y hora de Nueva York: {fecha_hora_ny}",
            fg="blue"
        )

        print("=== FIN ACTUALIZACIÓN ===\n")

        mostrar_datos_en_tabla(csv_file)

    except Exception as e:
        print(f"[ERROR GENERAL] {str(e)}")
        label_status.config(text=f"Error: {str(e)}", fg="red")

def mostrar_datos_en_tabla(csv_file):
    # Verificar que pandas esté cargado
    if not verificar_libs_cargadas(["pandas"]):
        return

    df = pd.read_csv(csv_file)

    # Limpiar tabla
    for row in tree.get_children():
        tree.delete(row)

    # Insertar filas
    for _, row in df.iterrows():
        tree.insert(
            "", tk.END,
            values=(
                row['Date'],
                row['Ticker'],
                f"{row['Open']:.2f}",
                f"{row['High']:.2f}",
                f"{row['Low']:.2f}",
                f"{row['Close']:.2f}"
            )
        )

# Crear ventana principal
root = tk.Tk()
root.title("Actualizar precios de acciones")

# Label de progreso de carga (en la parte superior)
label_carga = tk.Label(root, text="Cargando recursos...", fg="blue", font=("Arial", 9))
label_carga.pack(anchor="ne", padx=10, pady=2)

# Iniciar carga de bibliotecas en segundo plano
hilo_carga = threading.Thread(target=cargar_bibliotecas_async, args=(root, label_carga), daemon=True)
hilo_carga.start()

# Frame para selección de archivo
frame1 = tk.Frame(root)
frame1.pack(pady=10, padx=10, fill="x")
tk.Label(frame1, text="Ruta del CSV:").pack(anchor="w")
entry_ruta = tk.Entry(frame1, width=60)
entry_ruta.pack(side="left", padx=(0,5))
tk.Button(frame1, text="Seleccionar CSV", command=seleccionar_csv).pack(side="left")

# Cargar última ruta guardada
ruta_guardada = cargar_ruta_csv()
if ruta_guardada and os.path.exists(ruta_guardada):
    entry_ruta.insert(0, ruta_guardada)

# Frame para editar tickers y plataformas (DOS COLUMNAS)
frame_tickers = tk.Frame(root)
frame_tickers.pack(padx=10, pady=5, fill="x")

# --- COLUMNA IZQUIERDA: Panel de Plataforma ---
frame_plataforma_panel = tk.Frame(frame_tickers)
frame_plataforma_panel.pack(side="left", anchor="n", padx=(0, 20))

# --- Frame superior: Selector de plataforma ---
frame_plataforma_selector = tk.Frame(frame_plataforma_panel)
frame_plataforma_selector.pack(anchor="w", pady=(0, 5))

tk.Label(frame_plataforma_selector, text="Plataforma:").pack(side="left")
plataforma_tickers_var = tk.StringVar()
plataformas_disponibles = obtener_plataformas()
plataforma_tickers_var.set(plataformas_disponibles[0] if plataformas_disponibles else "TYBA")

combo_plataforma_tickers = ttk.Combobox(
    frame_plataforma_selector,
    textvariable=plataforma_tickers_var,
    values=plataformas_disponibles,
    state="readonly",
    width=12
)
combo_plataforma_tickers.pack(side="left", padx=5)

# Selector de Modo (Paper/Real) - default según plataforma
tk.Label(frame_plataforma_selector, text="Modo:").pack(side="left", padx=(15, 0))
modo_inicial = "Real"  # Por defecto Real para todas las plataformas
modo_tickers_var = tk.StringVar(value=modo_inicial)
combo_modo_tickers = ttk.Combobox(
    frame_plataforma_selector,
    textvariable=modo_tickers_var,
    values=["Paper", "Real"],
    state="readonly",
    width=8
)
combo_modo_tickers.pack(side="left", padx=5)

def on_plataforma_change(event=None):
    """Actualiza la lista cuando cambia la plataforma (mantiene el modo actual)"""
    actualizar_listbox_tickers()

# Nota: el bind se hace después de definir actualizar_listbox_tickers


def nueva_plataforma_dialog():
    """Abre dialogo para crear nueva plataforma."""
    dialog = tk.Toplevel(root)
    dialog.title("Nueva Plataforma")
    dialog.geometry("300x180")
    dialog.transient(root)
    dialog.grab_set()

    tk.Label(dialog, text="Nombre (clave):").pack(pady=(10, 0))
    entry_nombre = tk.Entry(dialog, width=20)
    entry_nombre.pack()

    tk.Label(dialog, text="Mercado:").pack(pady=(5, 0))
    entry_mercado = tk.Entry(dialog, width=20)
    entry_mercado.insert(0, "NYSE")
    entry_mercado.pack()

    tk.Label(dialog, text="Moneda:").pack(pady=(5, 0))
    entry_moneda = tk.Entry(dialog, width=20)
    entry_moneda.insert(0, "USD")
    entry_moneda.pack()

    lbl_resultado = tk.Label(dialog, text="")
    lbl_resultado.pack(pady=5)

    def crear():
        nombre = entry_nombre.get().strip().upper()
        mercado = entry_mercado.get().strip().upper() or "NYSE"
        moneda = entry_moneda.get().strip().upper() or "USD"

        if not nombre:
            lbl_resultado.config(text="Ingrese un nombre", fg="red")
            return

        exito, mensaje = agregar_plataforma_tickers(nombre, mercado, moneda)
        if exito:
            # Actualizar combo
            nuevas_plataformas = obtener_plataformas()
            combo_plataforma_tickers.config(values=nuevas_plataformas)
            plataforma_tickers_var.set(nombre)
            actualizar_listbox_tickers()
            dialog.destroy()
            label_status.config(text=mensaje, fg="green")
        else:
            lbl_resultado.config(text=mensaje, fg="red")

    tk.Button(dialog, text="Crear", command=crear, bg="#28a745", fg="white").pack(pady=10)


def eliminar_plataforma_dialog():
    """Elimina la plataforma seleccionada."""
    plat = plataforma_tickers_var.get()
    if not plat:
        return

    # Verificar si tiene operaciones
    try:
        hist = cargar_historial_operaciones_completo()
        ops_plat = [op for op in hist.get("operaciones", []) if op.get("plataforma", "TYBA") == plat]
        if ops_plat:
            messagebox.showwarning("Advertencia",
                f"La plataforma '{plat}' tiene {len(ops_plat)} operaciones registradas.\n"
                "Elimine las operaciones primero.")
            return
    except:
        pass

    if messagebox.askyesno("Confirmar", f"¿Eliminar plataforma '{plat}'?"):
        exito, mensaje = eliminar_plataforma_tickers(plat)
        if exito:
            nuevas_plataformas = obtener_plataformas()
            combo_plataforma_tickers.config(values=nuevas_plataformas)
            plataforma_tickers_var.set(nuevas_plataformas[0] if nuevas_plataformas else "")
            actualizar_listbox_tickers()
            label_status.config(text=mensaje, fg="blue")
        else:
            label_status.config(text=mensaje, fg="red")


tk.Button(frame_plataforma_selector, text="+", command=nueva_plataforma_dialog,
          width=2, bg="#28a745", fg="white").pack(side="left", padx=2)
tk.Button(frame_plataforma_selector, text="-", command=eliminar_plataforma_dialog,
          width=2, bg="#dc3545", fg="white").pack(side="left")

# --- Label de tickers ---
label_tickers_titulo = tk.Label(frame_plataforma_panel, text=f"Tickers de {plataforma_tickers_var.get()}:")
label_tickers_titulo.pack(anchor="w")

# Frame contenedor para listbox y scrollbar
frame_listbox_plat = tk.Frame(frame_plataforma_panel)
frame_listbox_plat.pack(anchor="w")

# Lista de tickers visible (de la plataforma seleccionada)
listbox_tickers = tk.Listbox(frame_listbox_plat, height=10, width=15)
listbox_tickers.pack(side="left", fill="y")

# Funcion para actualizar listbox segun plataforma y modo
def actualizar_listbox_tickers(*args):
    plat = plataforma_tickers_var.get()
    modo = modo_tickers_var.get()

    label_tickers_titulo.config(text=f"Tickers de {plat} ({modo}):")
    listbox_tickers.delete(0, tk.END)

    # Obtener tickers para esta combinación plataforma+modo
    tickers = obtener_tickers_plataforma(plat, modo)

    if tickers:
        for t in tickers:
            listbox_tickers.insert(tk.END, t)
    else:
        # Sin tickers para esta combinación
        listbox_tickers.insert(tk.END, f"({plat} {modo} sin tickers)")

# Cargar tickers iniciales
actualizar_listbox_tickers()

# Bind al cambio de plataforma (cambia modo + actualiza lista) y modo (solo actualiza lista)
combo_plataforma_tickers.bind("<<ComboboxSelected>>", on_plataforma_change)
combo_modo_tickers.bind("<<ComboboxSelected>>", actualizar_listbox_tickers)

# Scrollbar para listbox
scroll_tickers = tk.Scrollbar(frame_listbox_plat, orient="vertical", command=listbox_tickers.yview)
scroll_tickers.pack(side="left", fill="y")
listbox_tickers.config(yscrollcommand=scroll_tickers.set)

# Frame para botones de gestión de tickers de plataforma
frame_ticker_btns = tk.Frame(frame_plataforma_panel)
frame_ticker_btns.pack(anchor="w", pady=5)

def asignar_ticker_plataforma():
    """Asigna un ticker de la lista global a la plataforma seleccionada."""
    global tickers

    # Obtener ticker seleccionado de la lista global
    seleccion_global = listbox_global.curselection()
    if not seleccion_global:
        label_status.config(text="Selecciona un ticker de la Lista General.", fg="orange")
        return

    ticker = listbox_global.get(seleccion_global[0])
    plataforma = plataforma_tickers_var.get()
    modo = modo_tickers_var.get()
    tickers_plat = obtener_tickers_plataforma(plataforma, modo)

    if ticker in tickers_plat:
        label_status.config(text=f"{ticker} ya esta en {plataforma} ({modo}).", fg="orange")
        return

    # Solo asignar (no hacer onboarding, ya existe en la lista global)
    exito, mensaje = agregar_ticker_plataforma(plataforma, ticker, modo)
    if exito:
        tickers = obtener_tickers_unicos()
        actualizar_listbox_tickers()
        label_status.config(text=f"{ticker} asignado a {plataforma} ({modo}).", fg="green")
    else:
        label_status.config(text=mensaje, fg="orange")


def desasignar_ticker_plataforma():
    """Desasigna un ticker de la plataforma (no lo elimina de la lista global)."""
    global tickers
    seleccion = listbox_tickers.curselection()
    if not seleccion:
        label_status.config(text="Selecciona un ticker de la lista de plataforma.", fg="orange")
        return

    idx = seleccion[0]
    t = listbox_tickers.get(idx)
    plataforma = plataforma_tickers_var.get()
    modo = modo_tickers_var.get()

    exito, mensaje = quitar_ticker_plataforma(plataforma, t, modo)

    if exito:
        tickers = obtener_tickers_unicos()
        actualizar_listbox_tickers()
        label_status.config(text=f"{t} desasignado de {plataforma} ({modo}).", fg="blue")
    else:
        label_status.config(text=mensaje, fg="orange")

tk.Button(frame_ticker_btns, text="<< Asignar", command=asignar_ticker_plataforma, width=12).pack(pady=2)
tk.Button(frame_ticker_btns, text="Desasignar", command=desasignar_ticker_plataforma, width=12).pack(pady=2)

# --- COLUMNA DERECHA: Lista General de Tickers ---
frame_global_panel = tk.Frame(frame_tickers)
frame_global_panel.pack(side="left", anchor="n")

tk.Label(frame_global_panel, text="Lista General de Tickers:", font=("Arial", 10, "bold")).pack(anchor="w")

# Frame contenedor para listbox global y scrollbar
frame_listbox_global = tk.Frame(frame_global_panel)
frame_listbox_global.pack(anchor="w")

# Lista global de tickers
listbox_global = tk.Listbox(frame_listbox_global, height=10, width=15)
listbox_global.pack(side="left", fill="y")

# Scrollbar para listbox global
scroll_global = tk.Scrollbar(frame_listbox_global, orient="vertical", command=listbox_global.yview)
scroll_global.pack(side="left", fill="y")
listbox_global.config(yscrollcommand=scroll_global.set)

# Frame para entry y botones de Lista General
frame_global_btns = tk.Frame(frame_global_panel)
frame_global_btns.pack(anchor="w", pady=5)

entry_nuevo_ticker = tk.Entry(frame_global_btns, width=10)
entry_nuevo_ticker.pack(pady=(0, 5))


def actualizar_listbox_global():
    """Actualiza la lista global de tickers."""
    listbox_global.delete(0, tk.END)
    tickers_globales = obtener_tickers_globales()
    for t in tickers_globales:
        listbox_global.insert(tk.END, t)


def agregar_ticker_global_gui():
    """Agrega un ticker a la lista global con onboarding."""
    global tickers

    nuevo = entry_nuevo_ticker.get().strip().upper()
    if not nuevo:
        label_status.config(text="Ingresa un ticker valido.", fg="red")
        return

    # Verificar si ya existe en la lista global
    if ticker_existe_en_global(nuevo):
        label_status.config(text=f"{nuevo} ya existe en la Lista General.", fg="orange")
        return

    # Verificacion rapida con Yahoo Finance
    try:
        df_test = yf.download(nuevo, period="1d", progress=False)
        if df_test.empty:
            raise ValueError("No hay datos para este ticker")
    except Exception:
        label_status.config(text=f"Ticker invalido: {nuevo}", fg="red")
        return

    # Si pasa la verificacion, preguntar si ejecutar onboarding completo
    mensaje_confirmacion = (
        f"El ticker {nuevo} es valido.\n\n"
        f"Deseas ejecutar el proceso completo de onboarding?\n\n"
        f"Este proceso incluye:\n"
        f"1. Descargar datos desde 01-01-2025\n"
        f"2. Agregar al CSV de precios\n"
        f"3. Extraer datos de 12 meses\n"
        f"4. Ejecutar analisis (Completo, 6m, 3m)\n"
        f"5. Calcular parametros Slots 1-5\n\n"
        f"Tiempo estimado: ~5 minutos\n"
        f"La interfaz NO se congelara.\n\n"
        f"Si eliges 'No', solo se agregara el ticker a la lista."
    )

    ejecutar_onboarding = messagebox.askyesno(
        "Onboarding de Nuevo Ticker",
        mensaje_confirmacion
    )

    if ejecutar_onboarding:
        # Ejecutar onboarding en background
        entry_nuevo_ticker.delete(0, tk.END)
        label_status.config(text=f"Iniciando onboarding de {nuevo}...", fg="blue")
        ejecutar_onboarding_global_background(nuevo)
    else:
        # Solo agregar el ticker sin onboarding
        exito, mensaje = agregar_ticker_global(nuevo)
        if exito:
            tickers = obtener_tickers_unicos()
            actualizar_listbox_global()
            entry_nuevo_ticker.delete(0, tk.END)
            label_status.config(text=mensaje, fg="green")
        else:
            label_status.config(text=mensaje, fg="orange")


def ejecutar_onboarding_global_background(ticker):
    """Ejecuta el onboarding para la lista global."""
    global tickers

    def actualizar_progreso(mensaje, porcentaje):
        root.after(0, lambda: label_status.config(
            text=f"[{porcentaje}%] {mensaje}",
            fg="blue"
        ))

    def proceso_onboarding():
        try:
            from onboarding_nuevo_ticker import onboarding_ticker

            resultado = onboarding_ticker(ticker, callback=actualizar_progreso)
            exito = resultado.get('exito', False)
            errores = resultado.get('errores', [])
            pasos = resultado.get('pasos_completados', [])

            if exito:
                def finalizar_exito():
                    global tickers
                    # Agregar a la lista global
                    exito_agregar, msg_agregar = agregar_ticker_global(ticker)
                    if exito_agregar:
                        tickers = obtener_tickers_unicos()
                        actualizar_listbox_global()
                    label_status.config(
                        text=f"Onboarding de {ticker} completado. Parametros calculados para Slots 1-5.",
                        fg="green"
                    )
                    messagebox.showinfo(
                        "Onboarding Completado",
                        f"El ticker {ticker} ha sido configurado exitosamente.\n\n"
                        f"Se han calculado parametros para:\n"
                        f"- Slot 1 y 2 (Base)\n"
                        f"- Slot 3 y 4 (Derivados)\n"
                        f"- Slot 5 (Optimizado)\n\n"
                        f"Ahora puedes asignarlo a plataformas usando '<< Asignar'."
                    )
                root.after(0, finalizar_exito)
            else:
                def finalizar_error():
                    error_msg = ', '.join(errores) if errores else "Error desconocido"
                    label_status.config(
                        text=f"Error en onboarding de {ticker}: {error_msg}",
                        fg="red"
                    )
                    messagebox.showerror(
                        "Error en Onboarding",
                        f"Hubo un error durante el onboarding de {ticker}:\n\n{error_msg}\n\n"
                        f"Pasos completados: {', '.join(pasos) if pasos else 'ninguno'}"
                    )
                root.after(0, finalizar_error)

        except Exception as e:
            def mostrar_error():
                label_status.config(
                    text=f"Error en onboarding: {str(e)}",
                    fg="red"
                )
                messagebox.showerror(
                    "Error en Onboarding",
                    f"Error inesperado:\n\n{str(e)}"
                )
            root.after(0, mostrar_error)

    thread = threading.Thread(target=proceso_onboarding, daemon=True)
    thread.start()


def quitar_ticker_global_gui():
    """Quita un ticker de la lista global (solo si no está en ninguna plataforma)."""
    global tickers

    seleccion = listbox_global.curselection()
    if not seleccion:
        label_status.config(text="Selecciona un ticker de la Lista General.", fg="orange")
        return

    ticker = listbox_global.get(seleccion[0])

    # Intentar quitar (la función verifica si está en alguna plataforma)
    exito, mensaje = quitar_ticker_global(ticker)
    if exito:
        tickers = obtener_tickers_unicos()
        actualizar_listbox_global()
        label_status.config(text=mensaje, fg="blue")
    else:
        # Mostrar advertencia si está en plataformas
        label_status.config(text=mensaje, fg="orange")
        messagebox.showwarning(
            "No se puede quitar",
            f"{mensaje}\n\nPrimero desasigna el ticker de todas las plataformas."
        )


tk.Button(frame_global_btns, text="Agregar Ticker", command=agregar_ticker_global_gui,
          bg="#28a745", fg="white", width=12).pack(pady=2)
tk.Button(frame_global_btns, text="Quitar Ticker", command=quitar_ticker_global_gui,
          bg="#dc3545", fg="white", width=12).pack(pady=2)

# Cargar lista global inicial
actualizar_listbox_global()

# Checkbox para opción automática (activar/desactivar)
auto_var = tk.BooleanVar(value=False)
tk.Checkbutton(root, text="Actualizar automáticamente (activar/desactivar)", variable=auto_var).pack(pady=5)

def auto_actualizar():
    if auto_var.get():
        now_ny = datetime.now(ZoneInfo("America/New_York"))
        # Solo ejecutar de lunes a viernes (0=lunes, 4=viernes)
        if now_ny.weekday() < 5 and now_ny.hour == 16 and now_ny.minute >= 10:
            actualizar_csv()
    root.after(60000, auto_actualizar)  # revisa cada 60 segundos

auto_actualizar()

# Frame para botones principales
frame_botones_principales = tk.Frame(root)
frame_botones_principales.pack(pady=5)

# Botón para actualizar CSV manualmente
tk.Button(frame_botones_principales, text="Actualizar CSV ahora", command=actualizar_csv,
          bg="lightblue", font=("Arial", 10)).pack(side="left", padx=5)

# Boton para generar senales (genera para TODAS las plataformas/modos, muestra la seleccionada)
tk.Button(frame_botones_principales, text="Generar Senales",
          command=lambda: generar_senales_todas_plataformas(
              plataforma_mostrar=plataforma_tickers_var.get(),
              modo_mostrar=modo_tickers_var.get()),
          bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

# Botón para regenerar señales de fechas anteriores
tk.Button(frame_botones_principales, text="Regenerar Históricas", command=regenerar_senales_historicas,
          bg="#6c757d", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

# Botón para historial de operaciones
tk.Button(frame_botones_principales, text="Historial", command=administrar_historial,
          bg="#ffc107", fg="black", font=("Arial", 10)).pack(side="left", padx=5)

# Botón para comparar señales con operaciones reales
tk.Button(frame_botones_principales, text="Comparar Señales", command=comparar_senales_operaciones,
          bg="#17a2b8", fg="white", font=("Arial", 10)).pack(side="left", padx=5)

# Botón para sincronizar desde GitHub
tk.Button(frame_botones_principales, text="Sync GitHub", command=sincronizar_desde_github,
          bg="#6f42c1", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

# Label para mensajes de estado
label_status = tk.Label(root, text="", fg="blue")
label_status.pack(pady=5)

# Frame para tabla
frame_table = tk.Frame(root)
frame_table.pack(padx=10, pady=10, fill="both", expand=True)

columns = ("Date", "Ticker", "Open", "High", "Low", "Close")
tree = ttk.Treeview(frame_table, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, anchor="center", width=80)

# Scrollbars
scroll_y = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
scroll_x = ttk.Scrollbar(frame_table, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
tree.pack(side="left", fill="both", expand=True)
scroll_y.pack(side="right", fill="y")
scroll_x.pack(side="bottom", fill="x")

root.mainloop()
