#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
SCRIPT: Análisis de Inversiones con Optimización Multi-Período
=============================================================================
VERSIÓN: 2.10.0
FECHA DE CREACIÓN: 13/12/2025 10:45:00
ÚLTIMA MODIFICACIÓN: 05/06/2026

MEJORAS EN ESTA VERSIÓN (v2.10.0):
- FIX: Ventana Parámetros Activos ahora se refresca automáticamente al guardar Slot 3/4 o Slot 5
  (las funciones calcular_slots_3_4/calcular_slot_5 usaban variable local datos_slots que no era
  la misma que lee actualizar_tabla_slot; renombrado a datos_calc para que el outer datos_slots
  sea actualizado directamente en guardar_slots/guardar_slot5)

MEJORAS EN VERSIÓN (v2.9.9):
- FIX: Calcular Slots 1/2 ahora usa fechas dinámicas (hoy → hoy+91 días) en vez de hardcodeadas 2026-02-28
- NUEVO: Checkbox "Todos los tickers" — calcula Slot 1/2 usando el análisis más reciente de cada ticker individualmente
- FIX: Calcular Slot 3/4 fallaba con NoneType al no cargar pandas (faltaba _cargar_dependencias_analisis)

MEJORAS EN VERSIÓN (v2.9.0):
- NUEVO: Botón "Calcular Slot 3/4" en ventana Parámetros Activos
- NUEVO: Botón "Calcular Slot 5" en ventana Parámetros Activos
- NUEVO: Función integrada para calcular Slot 3 (largo) y Slot 4 (corto)
- NUEVO: Función integrada para calcular Slot 5 (mejor de 1-4 con ajuste ±30%)
- NUEVO: Optimización de factor individual por ticker (S3: 1.0-1.5, S4: 0.5-1.0)
- NUEVO: Optimización de ajuste individual por ticker (S5: ±30% en compra/venta)
- NUEVO: Comparación automática Slot 1 vs Slot 2 para determinar mejor base
- NUEVO: Simulación de rentabilidad últimos 2 meses (S3/S4) y 30 días (S5)

MEJORAS EN VERSIÓN ANTERIOR (v2.6.2):
- Modo portable para ejecutables (.exe)
- Detección automática de modo ejecutable vs script
- Carpeta data/ relativa al ejecutable para datos portables

MEJORAS EN VERSIÓN ANTERIOR (v2.6.0):
- NUEVO: Checkboxes para objetivos de optimización (Rentabilidad y/o Margen)
- NUEVO: Análisis multi-objetivo en una sola ejecución
- NUEVO: Barra de progreso inteligente con estimación de tiempo
- NUEVO: Historial de tiempos en ~/.analisis_tiempos.json
- NUEVO: Columnas Prom.Min% y Prom.Max% en ventana "Parámetros Activos"
- MEJORADO: Ventana "Administrar JSON" con 31 columnas de estadísticas
- MEJORADO: Valores Prom.Max% y Prom.Min% corregidos (÷100)
- MEJORADO: Anchos de columna dinámicos según título
- MEJORADO: Ordenamiento alfabético de tickers en todas las ventanas

MEJORAS EN VERSIÓN ANTERIOR (v2.5.8):
- NUEVO: Campo ticker_symbol en JSON (ej: "META" extraído de "Datos_META_ENE25_NOV25")
- NUEVO: Función extraer_ticker_symbol() para obtener ticker puro del nombre de archivo
- MEJORADO: Ventana "Administrar JSON" muestra ticker_symbol en lugar de nombre archivo

MEJORAS EN VERSIÓN ANTERIOR (v2.5.7):
- NUEVO: Barra de desplazamiento vertical (scrollbar) para toda la interfaz
- NUEVO: Cuadros de fechas de compras/ventas múltiples restaurados
- MEJORADO: Scroll funciona con rueda del mouse

MEJORAS EN VERSIÓN ANTERIOR (v2.5.6):
- NUEVO: Botón para detener análisis en proceso
- MEJORADO: Al borrar registro del JSON, se actualiza inmediatamente el cuadro de parámetros
- MEJORADO: Registros agrupados por objetivo en el cuadro de parámetros

MEJORAS EN VERSIÓN ANTERIOR (v2.5.5):
- MEJORADO: Guarda nuevo registro si los parámetros son diferentes (mismo ticker/período/objetivo)
- MEJORADO: Solo sobrescribe si ticker/período/objetivo Y parámetros son idénticos

MEJORAS EN VERSIÓN ANTERIOR (v2.5.4):
- NUEVO: Ventana para administrar JSON (ver y eliminar registros)
- NUEVO: Selección múltiple para eliminar varios registros a la vez

MEJORAS EN VERSIÓN ANTERIOR (v2.5.3):
- NUEVO: Refinamiento post-optimización para encontrar el centro del rango óptimo
- NUEVO: Semilla fija (seed=42) para resultados reproducibles
- MEJORADO: Siempre obtiene el mismo resultado óptimo para los mismos datos/parámetros

MEJORAS EN VERSIÓN ANTERIOR (v2.5.2):
- CORREGIDO: Lee correctamente JSON con estructura MIXTA (antigua + nueva)
- CORREGIDO: Botón "Generar DB y Excel" se reactiva al hacer nuevo análisis
- CORREGIDO: Muestra TODOS los análisis guardados (7 análisis en tu caso)
- MEJORADO: Compatible con cualquier combinación de estructuras JSON

AUTOR: Claude (Anthropic)
=============================================================================
"""

import os
import sys
import time
import json
from pathlib import Path
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog

# Imports pesados diferidos (lazy) para apertura rápida
pd = None
np = None
plt = None
FigureCanvasTkAgg = None
FuncFormatter = None
mdates = None
DayLocator = None
Workbook = None
dataframe_to_rows = None
differential_evolution = None
sqlite3 = None

def _cargar_dependencias_analisis():
    """Carga scipy, numpy y pandas cuando se necesitan para el análisis"""
    global pd, np, differential_evolution
    if pd is None:
        import pandas
        pd = pandas
    if np is None:
        import numpy
        np = numpy
    if differential_evolution is None:
        from scipy.optimize import differential_evolution as de
        differential_evolution = de

def _cargar_dependencias_grafico():
    """Carga matplotlib cuando se necesita para graficar"""
    global plt, FigureCanvasTkAgg, FuncFormatter, mdates, DayLocator
    if plt is None:
        import matplotlib.pyplot
        plt = matplotlib.pyplot
    if FigureCanvasTkAgg is None:
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg as fca
        FigureCanvasTkAgg = fca
    if FuncFormatter is None:
        from matplotlib.ticker import FuncFormatter as ff
        FuncFormatter = ff
    if mdates is None:
        import matplotlib.dates
        mdates = matplotlib.dates
    if DayLocator is None:
        from matplotlib.dates import DayLocator as dl
        DayLocator = dl

def _cargar_dependencias_excel():
    """Carga openpyxl cuando se necesita para exportar"""
    global Workbook, dataframe_to_rows
    if Workbook is None:
        from openpyxl import Workbook as wb
        Workbook = wb
    if dataframe_to_rows is None:
        from openpyxl.utils.dataframe import dataframe_to_rows as dtr
        dataframe_to_rows = dtr

def _cargar_sqlite():
    """Carga sqlite3 cuando se necesita"""
    global sqlite3
    if sqlite3 is None:
        import sqlite3 as sq
        sqlite3 = sq

# =====================================================
# DETECCION DE MODO EJECUTABLE vs SCRIPT
# =====================================================
def es_ejecutable():
    """Detecta si el script corre como ejecutable (.exe) compilado con PyInstaller"""
    return getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS')

def obtener_ruta_base():
    """Obtiene la ruta base del ejecutable o del script"""
    if es_ejecutable():
        # Cuando es ejecutable, la ruta es donde está el .exe
        return Path(sys.executable).parent
    else:
        # Cuando es script, la ruta es donde está el .py
        return Path(__file__).parent

def obtener_carpeta_datos():
    """Obtiene la carpeta de datos (data/) - SIEMPRE portable"""
    ruta_base = obtener_ruta_base()
    # Buscar data/ en el directorio actual
    carpeta_data = ruta_base / "data"
    if carpeta_data.exists():
        return carpeta_data
    # Si no existe, buscar en el directorio padre (estructura compartida)
    carpeta_data_padre = ruta_base.parent / "data"
    if carpeta_data_padre.exists():
        return carpeta_data_padre
    # Si no existe en ningun lugar, crear en el directorio actual
    carpeta_data.mkdir(parents=True, exist_ok=True)
    return carpeta_data

# Variable global para modo ejecutable
MODO_EJECUTABLE = es_ejecutable()
CARPETA_DATOS_PORTABLE = obtener_carpeta_datos()

# Funciones para guardar/cargar ultima carpeta de graficos
def guardar_ultima_carpeta_grafico(carpeta):
    """Guarda la ultima carpeta usada para graficos"""
    config_file = CARPETA_DATOS_PORTABLE / "config_grafico.json"
    try:
        config = {"ultima_carpeta_grafico": str(carpeta)}
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(config, f)
    except:
        pass

def cargar_ultima_carpeta_grafico():
    """Carga la ultima carpeta usada para graficos"""
    config_file = CARPETA_DATOS_PORTABLE / "config_grafico.json"
    try:
        if config_file.exists():
            with open(config_file, "r", encoding="utf-8") as f:
                config = json.load(f)
                carpeta = config.get("ultima_carpeta_grafico", "")
                if carpeta and Path(carpeta).exists():
                    return carpeta
    except:
        pass
    return str(CARPETA_DATOS_PORTABLE)

# Valores por defecto para el límite
LIMITE_TIPO = "acciones"
LIMITE_VALOR = 10.0
VENTA_MULTIPLE_ACCIONES = None
COMPRA_MULTIPLE_ACCIONES = None
text_ventas_mult = None
text_compras_mult = None

# Configuracion PORTABLE - siempre usa carpeta data/ relativa al script/exe
CONFIG_FILE = None  # No se usa archivo de configuracion externo
UBICACION_JSON = CARPETA_DATOS_PORTABLE
ARCHIVO_JSON = CARPETA_DATOS_PORTABLE / "Resultado_de_Analisis.json"
ARCHIVO_PARAMETROS_ACTIVOS = CARPETA_DATOS_PORTABLE / "parametros_activos.json"

# Columnas esperadas (exactas)
EXPECTED_COLUMNS = ["Fecha", "Último", "Apertura", "Máximo", "Mínimo", "Vol.", "% var."]

# Mapeo de columnas en inglés a español (para archivos de Investing.com en inglés)
COLUMN_MAPPING_EN_ES = {
    "Date": "Fecha",
    "Price": "Último",
    "Open": "Apertura",
    "High": "Máximo",
    "Low": "Mínimo",
    "Vol.": "Vol.",
    "Change %": "% var."
}

def normalizar_columnas(df):
    """Renombra columnas de inglés a español si es necesario"""
    # Limpiar nombres de columnas
    df.columns = [c.strip() for c in df.columns]

    # Si ya tiene columnas en español, no hacer nada
    if all(col in df.columns for col in EXPECTED_COLUMNS):
        return df

    # Intentar mapear columnas de inglés a español
    columnas_renombrar = {}
    for col_en, col_es in COLUMN_MAPPING_EN_ES.items():
        if col_en in df.columns:
            columnas_renombrar[col_en] = col_es

    if columnas_renombrar:
        df = df.rename(columns=columnas_renombrar)

    return df

# Variable global para almacenar resultados de análisis
resultados_analisis_actuales = {}

# Variables globales para progreso
scipy_evaluaciones = 0
scipy_evaluaciones_max = 0
scipy_inicio_tiempo = None

# Variable global para detener análisis
analisis_detenido = False
error_analisis_mostrado = False  # Evitar mostrar múltiples mensajes de error

# Variable global para objetivo actual durante análisis
OBJETIVO_ACTUAL = None

# Archivo para historial de tiempos de análisis
ARCHIVO_HISTORIAL_TIEMPOS = Path.home() / ".analisis_tiempos.json"

# Variables globales para progreso inteligente
progreso_combinacion_actual = 0
progreso_total_combinaciones = 0
progreso_tiempo_inicio_total = None
progreso_tiempos_combinaciones = []  # Lista de tiempos por combinación en la sesión actual


def obtener_clave_configuracion(num_filas, checks_activos):
    """
    Genera una clave única basada en el rango de filas y checks activos.
    checks_activos es un dict con: {'scipy': bool, 'compra': bool, 'venta': bool,
                                     'ganancia': bool, 'compra_mult': bool, 'venta_mult': bool}
    """
    # Rangos de filas: 0-100, 100-200, 200-300, 300-500, 500+
    if num_filas <= 100:
        rango = "0-100"
    elif num_filas <= 200:
        rango = "100-200"
    elif num_filas <= 300:
        rango = "200-300"
    elif num_filas <= 500:
        rango = "300-500"
    else:
        rango = "500+"

    # Crear string de checks activos
    checks_str = "_".join([k for k, v in sorted(checks_activos.items()) if v])

    return f"{rango}_{checks_str}" if checks_str else f"{rango}_ninguno"


def cargar_historial_tiempos():
    """Carga el historial de tiempos desde el archivo JSON"""
    try:
        if ARCHIVO_HISTORIAL_TIEMPOS.exists():
            with open(ARCHIVO_HISTORIAL_TIEMPOS, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"[WARN] Error cargando historial de tiempos: {e}")
    return {}


def guardar_historial_tiempos(historial):
    """Guarda el historial de tiempos en el archivo JSON"""
    try:
        with open(ARCHIVO_HISTORIAL_TIEMPOS, 'w', encoding='utf-8') as f:
            json.dump(historial, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"[WARN] Error guardando historial de tiempos: {e}")


def registrar_tiempo_combinacion(clave_config, tiempo_segundos):
    """Registra el tiempo de una combinación en el historial"""
    historial = cargar_historial_tiempos()

    if clave_config not in historial:
        historial[clave_config] = {"tiempos": [], "promedio": 0}

    # Mantener solo los últimos 10 tiempos para cada configuración
    historial[clave_config]["tiempos"].append(tiempo_segundos)
    if len(historial[clave_config]["tiempos"]) > 10:
        historial[clave_config]["tiempos"] = historial[clave_config]["tiempos"][-10:]

    # Calcular promedio
    tiempos = historial[clave_config]["tiempos"]
    historial[clave_config]["promedio"] = sum(tiempos) / len(tiempos)

    guardar_historial_tiempos(historial)


def estimar_tiempo_total(clave_config, num_combinaciones):
    """Estima el tiempo total basado en el historial"""
    historial = cargar_historial_tiempos()

    if clave_config in historial and historial[clave_config]["promedio"] > 0:
        tiempo_por_combinacion = historial[clave_config]["promedio"]
        return tiempo_por_combinacion * num_combinaciones, True

    return None, False


def formatear_tiempo(segundos):
    """Formatea segundos a formato legible (mm:ss o hh:mm:ss)"""
    if segundos < 0:
        return "calculando..."

    segundos = int(segundos)
    if segundos < 60:
        return f"{segundos} seg"
    elif segundos < 3600:
        mins = segundos // 60
        segs = segundos % 60
        return f"{mins}m {segs:02d}s"
    else:
        horas = segundos // 3600
        mins = (segundos % 3600) // 60
        segs = segundos % 60
        return f"{horas}h {mins:02d}m {segs:02d}s"


def detener_analisis():
    """Detiene el análisis en proceso"""
    global analisis_detenido
    analisis_detenido = True
    print("[DEBUG] Análisis detenido por el usuario")


def extraer_ticker_symbol(nombre_archivo):
    """
    Extrae el símbolo del ticker de Yahoo Finance desde el nombre del archivo.

    Ejemplos:
        "Datos_META_ENE25_NOV25" -> "META"
        "Datos_AAPL_ENE25_NOV25" -> "AAPL"
        "Datos_BRK-B_ENE25_NOV25" -> "BRK-B"
        "Datos_QQQ_ENE25_NOV25" -> "QQQ"

    Args:
        nombre_archivo: Nombre del archivo sin extensión (ej: "Datos_META_ENE25_NOV25")

    Returns:
        str: Símbolo del ticker (ej: "META") o None si no se puede extraer
    """
    import re

    if not nombre_archivo:
        return None

    # Patrón: Datos_TICKER_MesAño_MesAño
    # Donde TICKER puede contener letras, números y guiones (ej: BRK-B)
    # Y MesAño es 3 letras + 2 dígitos (ej: ENE25, NOV25)
    patron = r'^Datos_([A-Za-z0-9\-]+)_[A-Za-z]{3}\d{2}_[A-Za-z]{3}\d{2}$'

    match = re.match(patron, nombre_archivo)
    if match:
        return match.group(1).upper()

    # Patrón alternativo más flexible: Datos_TICKER_cualquier_cosa
    patron_alternativo = r'^Datos_([A-Za-z0-9\-]+)_'
    match_alt = re.match(patron_alternativo, nombre_archivo)
    if match_alt:
        return match_alt.group(1).upper()

    # Si no hay patrón "Datos_", intentar extraer el primer segmento antes de "_"
    partes = nombre_archivo.split('_')
    if len(partes) >= 2 and partes[0].upper() == "DATOS":
        return partes[1].upper()

    return None


# =========================
# Funciones de configuración JSON
# =========================
def cargar_configuracion():
    """Carga la ubicacion del JSON desde el archivo de configuracion"""
    global UBICACION_JSON, ARCHIVO_JSON, ARCHIVO_PARAMETROS_ACTIVOS

    # Modo PORTABLE: la configuracion ya esta definida (carpeta data/)
    label_json_actual.config(text=f"JSON: {ARCHIVO_JSON} (portable)")


def guardar_configuracion():
    """En modo portable no se guarda configuracion externa"""
    pass  # No se usa archivo de configuracion externo


def seleccionar_ubicacion_json():
    """En modo portable la ubicacion es fija"""
    global UBICACION_JSON, ARCHIVO_JSON, ARCHIVO_PARAMETROS_ACTIVOS

    # Modo portable: ubicacion fija
    messagebox.showinfo("Modo Portable",
        f"Los datos se guardan en:\n{CARPETA_DATOS_PORTABLE}\n\n"
        "Esta ubicacion es fija para mantener portabilidad.")
    return

    carpeta = filedialog.askdirectory(title="Selecciona carpeta para guardar resultados JSON")
    if carpeta:
        UBICACION_JSON = carpeta
        ARCHIVO_JSON = Path(UBICACION_JSON) / "Resultado_de_Analisis.json"
        ARCHIVO_PARAMETROS_ACTIVOS = Path(UBICACION_JSON) / "parametros_activos.json"
        guardar_configuracion()
        label_json_actual.config(text=f"JSON: {ARCHIVO_JSON}")
        messagebox.showinfo("Ubicación guardada", f"Los resultados se guardarán en:\n{ARCHIVO_JSON}")


def verificar_ubicacion_json():
    """Verifica si hay ubicación configurada, si no, pide al usuario"""
    global UBICACION_JSON, ARCHIVO_JSON

    # En modo ejecutable, siempre está configurado
    if MODO_EJECUTABLE:
        return True

    if UBICACION_JSON is None:
        respuesta = messagebox.askyesno(
            "Ubicación JSON no configurada",
            "No has configurado dónde guardar los resultados JSON.\n¿Deseas seleccionar una carpeta ahora?"
        )
        if respuesta:
            seleccionar_ubicacion_json()
            return ARCHIVO_JSON is not None
        return False
    return True


# =========================
# Funciones para Parámetros Activos
# =========================

def fecha_display_to_iso(fecha_display):
    """Convierte fecha DD-MM-YYYY a YYYY-MM-DD para almacenamiento"""
    if not fecha_display:
        return None
    try:
        dt = datetime.strptime(fecha_display.strip(), "%d-%m-%Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return None


def fecha_iso_to_display(fecha_iso):
    """Convierte fecha YYYY-MM-DD a DD-MM-YYYY para mostrar"""
    if not fecha_iso:
        return ""
    try:
        dt = datetime.strptime(fecha_iso.strip(), "%Y-%m-%d")
        return dt.strftime("%d-%m-%Y")
    except ValueError:
        return fecha_iso  # Retornar original si falla


def detectar_traslape_fechas(parametros, ticker, fecha_inicio, fecha_fin, excluir_index=None):
    """
    Detecta si hay traslape de fechas con parámetros existentes del mismo ticker.
    Returns: Lista de diccionarios con info de traslapes, o lista vacía si no hay
    """
    traslapes = []
    if not fecha_inicio and not fecha_fin:
        return traslapes

    def parse_fecha(f):
        if not f:
            return None
        try:
            return datetime.strptime(f, '%Y-%m-%d')
        except:
            return None

    nueva_inicio = parse_fecha(fecha_inicio)
    nueva_fin = parse_fecha(fecha_fin)

    if nueva_inicio and not nueva_fin:
        nueva_fin = datetime(2099, 12, 31)
    if nueva_fin and not nueva_inicio:
        nueva_inicio = datetime(2000, 1, 1)

    for i, p in enumerate(parametros):
        if excluir_index is not None and i == excluir_index:
            continue
        if p.get('ticker_symbol') != ticker:
            continue

        exist_inicio = parse_fecha(p.get('fecha_inicio'))
        exist_fin = parse_fecha(p.get('fecha_fin'))

        if not exist_inicio and not exist_fin:
            continue

        if exist_inicio and not exist_fin:
            exist_fin = datetime(2099, 12, 31)
        if exist_fin and not exist_inicio:
            exist_inicio = datetime(2000, 1, 1)

        if nueva_fin >= exist_inicio and nueva_inicio <= exist_fin:
            traslapes.append({
                'fecha_inicio': p.get('fecha_inicio'),
                'fecha_fin': p.get('fecha_fin'),
                'index': i
            })

    return traslapes


def crear_estructura_slots_vacia():
    """Crea una estructura de slots vacía con valores por defecto"""
    return {
        "version": "2.0",
        "slots": {
            "1": {"nombre": "1", "parametros_activos": []},
            "2": {"nombre": "2", "parametros_activos": []},
            "3": {"nombre": "3", "parametros_activos": []},
            "4": {"nombre": "4", "parametros_activos": []},
            "5": {"nombre": "5", "parametros_activos": []}
        }
    }


def migrar_parametros_v1_a_v2(datos_v1):
    """Migra formato antiguo (lista) a nuevo formato (slots)"""
    parametros_lista = datos_v1.get("parametros_activos", [])
    estructura = crear_estructura_slots_vacia()
    estructura["slots"]["1"]["parametros_activos"] = parametros_lista
    return estructura


def cargar_parametros_activos():
    """Carga los parámetros activos desde el archivo de configuración.
    Retorna estructura con slots (formato v2.0)"""
    if ARCHIVO_PARAMETROS_ACTIVOS is None or not ARCHIVO_PARAMETROS_ACTIVOS.exists():
        return crear_estructura_slots_vacia()

    try:
        with open(ARCHIVO_PARAMETROS_ACTIVOS, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Detectar versión del formato
        if "version" in datos and datos.get("version") == "2.0":
            # Ya es formato nuevo
            return datos
        else:
            # Formato antiguo - migrar a v2
            datos_migrados = migrar_parametros_v1_a_v2(datos)
            # Guardar en formato nuevo
            guardar_parametros_activos(datos_migrados)
            return datos_migrados
    except Exception as e:
        print(f"[ERROR] Error cargando parámetros activos: {e}")
        return crear_estructura_slots_vacia()


def guardar_parametros_activos(datos_slots):
    """Guarda los parámetros activos en el archivo de configuración (formato v2.0 con slots)"""
    if ARCHIVO_PARAMETROS_ACTIVOS is None:
        messagebox.showerror("Error", "No hay ubicación configurada para guardar los parámetros activos.")
        return False

    try:
        # Asegurar que tiene la versión correcta
        if "version" not in datos_slots:
            datos_slots["version"] = "2.0"
        with open(ARCHIVO_PARAMETROS_ACTIVOS, 'w', encoding='utf-8') as f:
            json.dump(datos_slots, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Error guardando parámetros activos:\n{e}")
        return False


def obtener_parametros_slot(datos_slots, slot_id):
    """Obtiene la lista de parámetros de un slot específico"""
    return datos_slots.get("slots", {}).get(slot_id, {}).get("parametros_activos", [])


def obtener_nombre_slot(datos_slots, slot_id):
    """Obtiene el nombre de un slot"""
    return datos_slots.get("slots", {}).get(slot_id, {}).get("nombre", slot_id)


def administrar_parametros_activos():
    """Abre una ventana para gestionar los parámetros activos con 5 pestañas (slots)"""
    if not verificar_ubicacion_json():
        return

    # Cargar estructura completa de parámetros con slots
    datos_slots = cargar_parametros_activos()

    # Crear ventana principal
    ventana_params = tk.Toplevel(ventana)
    ventana_params.title("Parámetros Activos para Señales de Trading")
    ventana_params.geometry("1350x550")

    # Frame superior con instrucciones y botón renombrar
    frame_superior = tk.Frame(ventana_params, pady=5)
    frame_superior.pack(fill="x", padx=10)

    tk.Label(frame_superior,
             text="Configura los parámetros que se usarán para generar señales de compra/venta (5 slots disponibles)",
             font=("Arial", 10), fg="gray").pack(side="left", anchor="w")

    def renombrar_slot_actual():
        """Renombra el slot actualmente seleccionado"""
        tab_actual = notebook.index(notebook.select())
        slot_id = str(tab_actual + 1)
        nombre_actual = obtener_nombre_slot(datos_slots, slot_id)

        nuevo_nombre = simpledialog.askstring("Renombrar Slot",
            f"Nuevo nombre para el slot {slot_id}:",
            initialvalue=nombre_actual,
            parent=ventana_params)

        if nuevo_nombre and nuevo_nombre.strip():
            datos_slots["slots"][slot_id]["nombre"] = nuevo_nombre.strip()
            guardar_parametros_activos(datos_slots)
            actualizar_titulos_pestanas()

    tk.Button(frame_superior, text="Renombrar Slot", command=renombrar_slot_actual,
              bg="#6c757d", fg="white", font=("Arial", 9)).pack(side="right")

    # Notebook para las 5 pestañas
    notebook = ttk.Notebook(ventana_params)
    notebook.pack(fill="both", expand=True, padx=10, pady=5)

    # Diccionario para almacenar los treeviews de cada slot
    trees = {}
    frames_slots = {}
    combos_vigencia = {}  # Comboboxes de vigencia por slot
    periodos_por_slot = {}  # Lista de periodos por slot

    def obtener_periodos_vigencia(slot_id):
        """Obtiene los periodos de vigencia únicos de un slot"""
        parametros = obtener_parametros_slot(datos_slots, slot_id)
        periodos = set()
        for p in parametros:
            fecha_inicio = p.get("fecha_inicio", "")
            fecha_fin = p.get("fecha_fin", "")
            if fecha_inicio and fecha_fin:
                periodos.add((fecha_inicio, fecha_fin))
        # Ordenar por fecha_fin descendente (más reciente primero)
        periodos_ordenados = sorted(list(periodos), key=lambda x: x[1], reverse=True)
        return periodos_ordenados

    def formato_periodo(periodo):
        """Formatea un periodo (fecha_inicio, fecha_fin) para mostrar en combobox"""
        if periodo == "Todos":
            return "Todos"
        fecha_inicio, fecha_fin = periodo
        return f"{fecha_iso_to_display(fecha_inicio)} a {fecha_iso_to_display(fecha_fin)}"

    # Columnas del Treeview (sin Desde/Hasta ya que se filtra por vigencia)
    columns = ("Symbol", "Origen", "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N", "Límite", "Valor Lím.", "Prom.Min%", "Prom.Max%")
    anchos = {"Symbol": 70, "Origen": 85, "Compra%": 60, "Venta%": 60,
              "Gan.Mín%": 65, "Compra N": 60, "Venta N": 60, "Límite": 60, "Valor Lím.": 70,
              "Prom.Min%": 75, "Prom.Max%": 75}

    def crear_pestaña_slot(slot_id):
        """Crea una pestaña para un slot específico"""
        frame_slot = tk.Frame(notebook)
        frames_slots[slot_id] = frame_slot

        # Frame superior con selector de vigencia
        frame_vigencia = tk.Frame(frame_slot)
        frame_vigencia.pack(fill="x", padx=5, pady=(5, 0))
        
        tk.Label(frame_vigencia, text="Vigencia:", font=("Arial", 9)).pack(side="left", padx=(0, 5))
        
        # Obtener periodos de vigencia para este slot
        periodos = obtener_periodos_vigencia(slot_id)
        periodos_por_slot[slot_id] = periodos
        
        # Crear combobox con periodos formateados
        combo_vigencia = ttk.Combobox(frame_vigencia, state="readonly", width=30)
        valores_combo = ["Todos"] + [formato_periodo(p) for p in periodos]
        combo_vigencia["values"] = valores_combo
        # Seleccionar el primer periodo vigente por defecto (índice 1, después de "Todos")
        combo_vigencia.current(1 if len(periodos) > 0 else 0)
        combo_vigencia.pack(side="left")
        combos_vigencia[slot_id] = combo_vigencia
        
        def on_vigencia_change(event=None):
            actualizar_tabla_slot(slot_id)
        
        combo_vigencia.bind("<<ComboboxSelected>>", on_vigencia_change)

        # Frame para el Treeview
        frame_tree = tk.Frame(frame_slot)
        frame_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Scrollbars
        scrollbar_y = tk.Scrollbar(frame_tree, orient="vertical")
        scrollbar_x = tk.Scrollbar(frame_tree, orient="horizontal")

        # Treeview
        tree = ttk.Treeview(frame_tree, columns=columns, show="headings",
                            selectmode="extended",
                            yscrollcommand=scrollbar_y.set,
                            xscrollcommand=scrollbar_x.set)

        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=anchos.get(col, 80), anchor="center")

        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        trees[slot_id] = tree
        return frame_slot

    def actualizar_tabla_slot(slot_id):
        """Actualiza la tabla de un slot específico, filtrando por vigencia seleccionada"""
        tree = trees.get(slot_id)
        if not tree:
            return

        for item in tree.get_children():
            tree.delete(item)

        parametros = obtener_parametros_slot(datos_slots, slot_id)

        # Actualizar combo de vigencia con los períodos actuales (por si cambiaron)
        combo = combos_vigencia.get(slot_id)
        if combo:
            seleccion_actual = combo.get()
            nuevos_periodos = obtener_periodos_vigencia(slot_id)
            periodos_por_slot[slot_id] = nuevos_periodos
            nuevos_valores = ["Todos"] + [formato_periodo(p) for p in nuevos_periodos]
            combo["values"] = nuevos_valores
            # Preservar selección si sigue existiendo, si no usar el primero
            if seleccion_actual in nuevos_valores:
                combo.set(seleccion_actual)
            elif nuevos_valores:
                combo.current(1 if len(nuevos_periodos) > 0 else 0)

        # Filtrar por periodo de vigencia seleccionado
        if combo:
            seleccion = combo.get()
            if seleccion != "Todos":
                periodos = periodos_por_slot.get(slot_id, [])
                for periodo in periodos:
                    if formato_periodo(periodo) == seleccion:
                        fecha_inicio_filtro, fecha_fin_filtro = periodo
                        parametros = [p for p in parametros
                                     if p.get("fecha_inicio") == fecha_inicio_filtro
                                     and p.get("fecha_fin") == fecha_fin_filtro]
                        break
        
        parametros_ordenados = sorted(parametros, key=lambda x: x.get("ticker_symbol", "").upper())

        for param in parametros_ordenados:
            compra_n = param.get("compra_multiple")
            venta_n = param.get("venta_multiple")
            limite_tipo = param.get("limite_tipo", "acciones")
            limite_valor = param.get("limite_valor", 10.0)
            prom_min = param.get("promedio_minimos", 0)
            prom_max = param.get("promedio_maximos", 0)
            tree.insert("", "end", values=(
                param.get("ticker_symbol", ""),
                param.get("origen", ""),
                f"{param.get('compra_pct', 0):.1f}",
                f"{param.get('venta_pct', 0):.1f}",
                f"{param.get('ganancia_min_pct', 0):.1f}",
                compra_n if compra_n else "-",
                venta_n if venta_n else "-",
                limite_tipo.title() if limite_tipo else "Acciones",
                f"{limite_valor:.0f}" if limite_tipo == "acciones" else f"${limite_valor:.0f}",
                f"{prom_min:.2f}%" if prom_min else "-",
                f"{prom_max:.2f}%" if prom_max else "-"
            ))

    def actualizar_titulos_pestanas():
        """Actualiza los títulos de las pestañas con nombre y cantidad de tickers"""
        for slot_id in ["1", "2", "3", "4", "5"]:
            nombre = obtener_nombre_slot(datos_slots, slot_id)
            cantidad = len(obtener_parametros_slot(datos_slots, slot_id))
            titulo = f"{nombre} ({cantidad})"
            notebook.tab(int(slot_id) - 1, text=titulo)

    def obtener_slot_actual():
        """Obtiene el ID del slot actualmente seleccionado"""
        return str(notebook.index(notebook.select()) + 1)

    def obtener_parametros_slot_actual():
        """Obtiene la lista de parámetros del slot actual"""
        slot_id = obtener_slot_actual()
        return datos_slots["slots"][slot_id]["parametros_activos"]

    # Crear las 5 pestañas
    for slot_id in ["1", "2", "3", "4", "5"]:
        frame = crear_pestaña_slot(slot_id)
        nombre = obtener_nombre_slot(datos_slots, slot_id)
        cantidad = len(obtener_parametros_slot(datos_slots, slot_id))
        notebook.add(frame, text=f"{nombre} ({cantidad})")
        actualizar_tabla_slot(slot_id)

    # Frame inferior con botones
    frame_botones = tk.Frame(ventana_params, pady=10)
    frame_botones.pack(fill="x", padx=10)

    def agregar_desde_json():
        """Abre ventana para seleccionar parámetros del JSON calculado"""
        slot_id = obtener_slot_actual()
        parametros = obtener_parametros_slot_actual()

        datos_json = cargar_resultados_json()
        if not datos_json:
            messagebox.showinfo("Sin datos", "No hay parámetros calculados en el JSON")
            return

        ventana_seleccion = tk.Toplevel(ventana_params)
        ventana_seleccion.title(f"Seleccionar desde JSON - Slot {obtener_nombre_slot(datos_slots, slot_id)}")
        ventana_seleccion.geometry("900x400")
        ventana_seleccion.transient(ventana_params)
        ventana_seleccion.grab_set()

        tk.Label(ventana_seleccion, text="Selecciona los parámetros a agregar:",
                 font=("Arial", 10)).pack(pady=5)

        frame_lista = tk.Frame(ventana_seleccion)
        frame_lista.pack(fill="both", expand=True, padx=10, pady=5)

        scrollbar = tk.Scrollbar(frame_lista)
        scrollbar.pack(side="right", fill="y")

        cols_sel = ("Symbol", "Período", "Objetivo", "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N")
        tree_sel = ttk.Treeview(frame_lista, columns=cols_sel, show="headings",
                                selectmode="extended", yscrollcommand=scrollbar.set)
        scrollbar.config(command=tree_sel.yview)

        for col in cols_sel:
            tree_sel.heading(col, text=col)
            tree_sel.column(col, width=90, anchor="center")

        item_datos = {}

        for ticker, contenido_ticker in datos_json.items():
            ticker_symbol = contenido_ticker.get("_ticker_symbol") or extraer_ticker_symbol(ticker) or ticker

            for periodo, contenido_periodo in contenido_ticker.items():
                if periodo in ["ticker", "fecha_guardado", "periodos", "_ticker_symbol"]:
                    continue

                if isinstance(contenido_periodo, dict):
                    for objetivo, datos in contenido_periodo.items():
                        if isinstance(datos, dict) and "parametros_optimos" in datos:
                            params = datos.get("parametros_optimos", {})
                            compra_mult = params.get("compra_multiple")
                            venta_mult = params.get("venta_multiple")

                            item_id = tree_sel.insert("", "end", values=(
                                ticker_symbol,
                                periodo.replace("_", " ").title(),
                                objetivo.replace("_", " ").title(),
                                f"{params.get('compra_pct', 0):.1f}",
                                f"{params.get('venta_pct', 0):.1f}",
                                f"{params.get('ganancia_minima_pct', 0):.1f}",
                                compra_mult if compra_mult else "-",
                                venta_mult if venta_mult else "-"
                            ))

                            item_datos[item_id] = {
                                "ticker_symbol": ticker_symbol,
                                "origen": f"calculado ({periodo}/{objetivo})",
                                "compra_pct": params.get("compra_pct", 0),
                                "venta_pct": params.get("venta_pct", 0),
                                "ganancia_min_pct": params.get("ganancia_minima_pct", 0),
                                "compra_multiple": compra_mult,
                                "venta_multiple": venta_mult,
                                "limite_tipo": params.get("limite_tipo", "acciones"),
                                "limite_valor": params.get("limite_valor", 10.0),
                                "promedio_maximos": params.get("promedio_maximos", 0),
                                "promedio_minimos": params.get("promedio_minimos", 0)
                            }

        tree_sel.pack(fill="both", expand=True)

        def agregar_seleccionados():
            seleccionados = tree_sel.selection()
            if not seleccionados:
                messagebox.showwarning("Sin selección", "Selecciona al menos un registro")
                return

            agregados = 0
            for item_id in seleccionados:
                if item_id in item_datos:
                    nuevo_param = item_datos[item_id].copy()
                    existe = any(p.get("ticker_symbol") == nuevo_param["ticker_symbol"] for p in parametros)
                    if existe:
                        resp = messagebox.askyesno("Ticker existente",
                            f"Ya existe {nuevo_param['ticker_symbol']} en este slot.\n¿Deseas reemplazarlo?")
                        if resp:
                            parametros[:] = [p for p in parametros if p.get("ticker_symbol") != nuevo_param["ticker_symbol"]]
                        else:
                            continue
                    parametros.append(nuevo_param)
                    agregados += 1

            if agregados > 0:
                guardar_parametros_activos(datos_slots)
                actualizar_tabla_slot(slot_id)
                actualizar_titulos_pestanas()
                messagebox.showinfo("Agregados", f"Se agregaron {agregados} parámetro(s) al slot {obtener_nombre_slot(datos_slots, slot_id)}")
            ventana_seleccion.destroy()

        tk.Button(ventana_seleccion, text="Agregar seleccionados", command=agregar_seleccionados,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

    def agregar_personalizado():
        """Abre ventana para agregar parámetros personalizados al slot actual"""
        slot_id = obtener_slot_actual()
        parametros = obtener_parametros_slot_actual()

        ventana_custom = tk.Toplevel(ventana_params)
        ventana_custom.title(f"Agregar Personalizado - Slot {obtener_nombre_slot(datos_slots, slot_id)}")
        ventana_custom.geometry("450x650")
        ventana_custom.transient(ventana_params)
        ventana_custom.grab_set()

        frame_form = tk.Frame(ventana_custom, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        tk.Label(frame_form, text="Symbol (ej: META, AAPL):", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=5)
        entry_symbol = tk.Entry(frame_form, width=20)
        entry_symbol.grid(row=0, column=1, pady=5)

        tk.Label(frame_form, text="Compra %:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=5)
        entry_compra = tk.Entry(frame_form, width=20)
        entry_compra.grid(row=1, column=1, pady=5)

        tk.Label(frame_form, text="Venta %:", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=5)
        entry_venta = tk.Entry(frame_form, width=20)
        entry_venta.grid(row=2, column=1, pady=5)

        tk.Label(frame_form, text="Ganancia Mín %:", font=("Arial", 10)).grid(row=3, column=0, sticky="w", pady=5)
        entry_ganancia = tk.Entry(frame_form, width=20)
        entry_ganancia.grid(row=3, column=1, pady=5)

        tk.Label(frame_form, text="Compra N acciones (opcional):", font=("Arial", 10)).grid(row=4, column=0, sticky="w", pady=5)
        entry_compra_n = tk.Entry(frame_form, width=20)
        entry_compra_n.grid(row=4, column=1, pady=5)

        tk.Label(frame_form, text="Venta N acciones (opcional):", font=("Arial", 10)).grid(row=5, column=0, sticky="w", pady=5)
        entry_venta_n = tk.Entry(frame_form, width=20)
        entry_venta_n.grid(row=5, column=1, pady=5)

        tk.Label(frame_form, text="Tipo de límite:", font=("Arial", 10)).grid(row=6, column=0, sticky="w", pady=5)
        limite_tipo_var = tk.StringVar(value="acciones")
        frame_limite_tipo = tk.Frame(frame_form)
        frame_limite_tipo.grid(row=6, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_limite_tipo, text="Acciones", variable=limite_tipo_var, value="acciones").pack(side="left")
        tk.Radiobutton(frame_limite_tipo, text="Monto $", variable=limite_tipo_var, value="monto").pack(side="left")

        tk.Label(frame_form, text="Valor límite:", font=("Arial", 10)).grid(row=7, column=0, sticky="w", pady=5)
        entry_limite_valor = tk.Entry(frame_form, width=20)
        entry_limite_valor.insert(0, "10")
        entry_limite_valor.grid(row=7, column=1, pady=5)

        tk.Label(frame_form, text="─── Condiciones Múltiples ───", font=("Arial", 9, "italic"), fg="gray").grid(row=8, column=0, columnspan=2, pady=(10,5))

        tk.Label(frame_form, text="Prom. % acum mínimos (-):", font=("Arial", 10)).grid(row=9, column=0, sticky="w", pady=5)
        entry_prom_min = tk.Entry(frame_form, width=20)
        entry_prom_min.insert(0, "0")
        entry_prom_min.grid(row=9, column=1, pady=5)

        tk.Label(frame_form, text="Prom. % acum máximos (+):", font=("Arial", 10)).grid(row=10, column=0, sticky="w", pady=5)
        entry_prom_max = tk.Entry(frame_form, width=20)
        entry_prom_max.insert(0, "0")
        entry_prom_max.grid(row=10, column=1, pady=5)

        tk.Label(frame_form, text="─── Período de Vigencia ───", font=("Arial", 9, "italic"), fg="gray").grid(row=11, column=0, columnspan=2, pady=(10,5))

        tk.Label(frame_form, text="Fecha inicio (DD-MM-YYYY):", font=("Arial", 10)).grid(row=12, column=0, sticky="w", pady=5)
        entry_fecha_inicio = tk.Entry(frame_form, width=20)
        entry_fecha_inicio.grid(row=12, column=1, pady=5)

        tk.Label(frame_form, text="Fecha fin (DD-MM-YYYY):", font=("Arial", 10)).grid(row=13, column=0, sticky="w", pady=5)
        entry_fecha_fin = tk.Entry(frame_form, width=20)
        entry_fecha_fin.grid(row=13, column=1, pady=5)

        tk.Label(frame_form, text="(dejar vacío = vigente indefinidamente)", font=("Arial", 8), fg="gray").grid(row=14, column=0, columnspan=2, pady=2)

        def guardar_personalizado():
            symbol = entry_symbol.get().strip().upper()
            if not symbol:
                messagebox.showwarning("Campo requerido", "Ingresa el símbolo del ticker")
                return

            try:
                compra_pct = float(entry_compra.get().strip().replace(",", "."))
                venta_pct = float(entry_venta.get().strip().replace(",", "."))
                ganancia_pct = float(entry_ganancia.get().strip().replace(",", "."))
            except ValueError:
                messagebox.showerror("Error", "Los valores de porcentaje deben ser numéricos")
                return

            compra_n = entry_compra_n.get().strip()
            venta_n = entry_venta_n.get().strip()
            tipo_limite = limite_tipo_var.get()
            try:
                valor_limite = float(entry_limite_valor.get().strip().replace(",", "."))
            except ValueError:
                valor_limite = 10.0
            try:
                prom_min = float(entry_prom_min.get().strip().replace(",", "."))
            except ValueError:
                prom_min = 0.0
            try:
                prom_max = float(entry_prom_max.get().strip().replace(",", "."))
            except ValueError:
                prom_max = 0.0

            # Obtener fechas de vigencia (formato DD-MM-YYYY)
            fecha_inicio_input = entry_fecha_inicio.get().strip()
            fecha_fin_input = entry_fecha_fin.get().strip()

            # Convertir y validar fechas
            fecha_inicio = None
            fecha_fin = None
            if fecha_inicio_input:
                fecha_inicio = fecha_display_to_iso(fecha_inicio_input)
                if fecha_inicio is None:
                    messagebox.showerror("Error", "Fecha inicio debe tener formato DD-MM-YYYY")
                    return
            if fecha_fin_input:
                fecha_fin = fecha_display_to_iso(fecha_fin_input)
                if fecha_fin is None:
                    messagebox.showerror("Error", "Fecha fin debe tener formato DD-MM-YYYY")
                    return

            nuevo_param = {
                "ticker_symbol": symbol,
                "origen": "personalizado",
                "compra_pct": compra_pct,
                "venta_pct": venta_pct,
                "ganancia_min_pct": ganancia_pct,
                "compra_multiple": int(compra_n) if compra_n else None,
                "venta_multiple": int(venta_n) if venta_n else None,
                "limite_tipo": tipo_limite,
                "limite_valor": valor_limite,
                "promedio_minimos": prom_min,
                "promedio_maximos": prom_max,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin
            }

            # Verificar traslape de fechas con parámetros existentes
            traslapes = detectar_traslape_fechas(parametros, symbol, fecha_inicio, fecha_fin)
            if traslapes:
                # Verificar si es el mismo periodo exacto (reemplazo) o traslape parcial
                es_mismo_periodo = any(
                    t['fecha_inicio'] == fecha_inicio and t['fecha_fin'] == fecha_fin
                    for t in traslapes
                )
                if es_mismo_periodo:
                    resp = messagebox.askyesno("Ticker existente",
                        f"Ya existe {symbol} con el mismo período.\n¿Deseas reemplazarlo?")
                    if resp:
                        parametros[:] = [p for p in parametros if not (
                            p.get("ticker_symbol") == symbol and
                            p.get("fecha_inicio") == fecha_inicio and
                            p.get("fecha_fin") == fecha_fin
                        )]
                    else:
                        return
                else:
                    # Hay traslape parcial - mostrar advertencia
                    periodos_traslape = "\n".join([
                        f"  - {fecha_iso_to_display(t['fecha_inicio'])} a {fecha_iso_to_display(t['fecha_fin'])}"
                        for t in traslapes
                    ])
                    resp = messagebox.askyesno("ADVERTENCIA: Traslape de fechas",
                        f"El período ingresado para {symbol} se traslapa con:\n{periodos_traslape}\n\n"
                        f"Esto puede causar conflictos al regenerar señales históricas.\n\n"
                        f"¿Deseas continuar de todos modos?",
                        icon='warning')
                    if not resp:
                        return

            parametros.append(nuevo_param)

            guardar_parametros_activos(datos_slots)
            actualizar_tabla_slot(slot_id)
            actualizar_titulos_pestanas()
            messagebox.showinfo("Guardado", f"{symbol} guardado en slot {obtener_nombre_slot(datos_slots, slot_id)}")
            ventana_custom.destroy()

        tk.Button(frame_form, text="Guardar", command=guardar_personalizado,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).grid(row=15, column=0, columnspan=2, pady=20)

    def eliminar_seleccionados():
        """Elimina los parámetros seleccionados del slot actual"""
        slot_id = obtener_slot_actual()
        parametros = obtener_parametros_slot_actual()
        tree = trees.get(slot_id)

        seleccionados = tree.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona al menos un registro para eliminar")
            return

        if not messagebox.askyesno("Confirmar", f"¿Eliminar {len(seleccionados)} parámetro(s)?"):
            return

        symbols_eliminar = []
        for item_id in seleccionados:
            valores = tree.item(item_id, "values")
            symbols_eliminar.append(valores[0])

        parametros[:] = [p for p in parametros if p.get("ticker_symbol") not in symbols_eliminar]
        guardar_parametros_activos(datos_slots)
        actualizar_tabla_slot(slot_id)
        actualizar_titulos_pestanas()
        messagebox.showinfo("Eliminados", f"Se eliminaron {len(symbols_eliminar)} parámetro(s)")

    def editar_parametro():
        """Edita el parámetro seleccionado del slot actual"""
        slot_id = obtener_slot_actual()
        parametros = obtener_parametros_slot_actual()
        tree = trees.get(slot_id)

        seleccionados = tree.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona un parámetro para editar")
            return

        if len(seleccionados) > 1:
            messagebox.showwarning("Selección múltiple", "Selecciona solo un parámetro para editar")
            return

        item_id = seleccionados[0]
        valores = tree.item(item_id, "values")
        ticker_editar = valores[0]

        param_editar = None
        param_index = None
        for i, p in enumerate(parametros):
            if p.get("ticker_symbol") == ticker_editar:
                param_editar = p
                param_index = i
                break

        if param_editar is None:
            messagebox.showerror("Error", "No se encontró el parámetro")
            return

        ventana_editar = tk.Toplevel(ventana_params)
        ventana_editar.title(f"Editar {ticker_editar} - Slot {obtener_nombre_slot(datos_slots, slot_id)}")
        ventana_editar.geometry("400x680")
        ventana_editar.transient(ventana_params)
        ventana_editar.grab_set()

        tk.Label(ventana_editar, text="Editar Parámetro", font=("Arial", 12, "bold")).pack(pady=10)

        frame_form = tk.Frame(ventana_editar, padx=20, pady=10)
        frame_form.pack(fill="both", expand=True)

        tk.Label(frame_form, text="Ticker:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        entry_ticker = tk.Entry(frame_form, width=15, font=("Arial", 10, "bold"))
        entry_ticker.insert(0, ticker_editar)
        entry_ticker.grid(row=0, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Compra %:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=5)
        entry_compra = tk.Entry(frame_form, width=15)
        entry_compra.insert(0, str(param_editar.get("compra_pct", 0)))
        entry_compra.grid(row=1, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Venta %:", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=5)
        entry_venta = tk.Entry(frame_form, width=15)
        entry_venta.insert(0, str(param_editar.get("venta_pct", 0)))
        entry_venta.grid(row=2, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Ganancia mín %:", font=("Arial", 10)).grid(row=3, column=0, sticky="w", pady=5)
        entry_ganancia = tk.Entry(frame_form, width=15)
        entry_ganancia.insert(0, str(param_editar.get("ganancia_min_pct", 0)))
        entry_ganancia.grid(row=3, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Compra múltiple:", font=("Arial", 10)).grid(row=4, column=0, sticky="w", pady=5)
        entry_compra_mult = tk.Entry(frame_form, width=15)
        compra_mult_val = param_editar.get("compra_multiple")
        entry_compra_mult.insert(0, str(compra_mult_val) if compra_mult_val else "")
        entry_compra_mult.grid(row=4, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Venta múltiple:", font=("Arial", 10)).grid(row=5, column=0, sticky="w", pady=5)
        entry_venta_mult = tk.Entry(frame_form, width=15)
        venta_mult_val = param_editar.get("venta_multiple")
        entry_venta_mult.insert(0, str(venta_mult_val) if venta_mult_val else "")
        entry_venta_mult.grid(row=5, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Tipo de límite:", font=("Arial", 10)).grid(row=6, column=0, sticky="w", pady=5)
        limite_tipo_var = tk.StringVar(value=param_editar.get("limite_tipo", "acciones"))
        frame_limite_tipo = tk.Frame(frame_form)
        frame_limite_tipo.grid(row=6, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_limite_tipo, text="Acciones", variable=limite_tipo_var, value="acciones").pack(side="left")
        tk.Radiobutton(frame_limite_tipo, text="Monto $", variable=limite_tipo_var, value="monto").pack(side="left")

        tk.Label(frame_form, text="Valor límite:", font=("Arial", 10)).grid(row=7, column=0, sticky="w", pady=5)
        entry_limite_valor = tk.Entry(frame_form, width=15)
        entry_limite_valor.insert(0, str(param_editar.get("limite_valor", 10.0)))
        entry_limite_valor.grid(row=7, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="─── Condiciones Múltiples ───", font=("Arial", 9, "italic"), fg="gray").grid(row=8, column=0, columnspan=2, pady=(10,5))

        tk.Label(frame_form, text="Prom. % mínimos (-):", font=("Arial", 10)).grid(row=9, column=0, sticky="w", pady=5)
        entry_prom_min = tk.Entry(frame_form, width=15)
        entry_prom_min.insert(0, str(param_editar.get("promedio_minimos", 0)))
        entry_prom_min.grid(row=9, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Prom. % máximos (+):", font=("Arial", 10)).grid(row=10, column=0, sticky="w", pady=5)
        entry_prom_max = tk.Entry(frame_form, width=15)
        entry_prom_max.insert(0, str(param_editar.get("promedio_maximos", 0)))
        entry_prom_max.grid(row=10, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="─── Período de Vigencia ───", font=("Arial", 9, "italic"), fg="gray").grid(row=11, column=0, columnspan=2, pady=(10,5))

        tk.Label(frame_form, text="Fecha inicio:", font=("Arial", 10)).grid(row=12, column=0, sticky="w", pady=5)
        entry_fecha_inicio = tk.Entry(frame_form, width=15)
        fecha_inicio_val = param_editar.get("fecha_inicio")
        entry_fecha_inicio.insert(0, fecha_iso_to_display(fecha_inicio_val) if fecha_inicio_val else "")
        entry_fecha_inicio.grid(row=12, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Fecha fin:", font=("Arial", 10)).grid(row=13, column=0, sticky="w", pady=5)
        entry_fecha_fin = tk.Entry(frame_form, width=15)
        fecha_fin_val = param_editar.get("fecha_fin")
        entry_fecha_fin.insert(0, fecha_iso_to_display(fecha_fin_val) if fecha_fin_val else "")
        entry_fecha_fin.grid(row=13, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Formato: DD-MM-YYYY", font=("Arial", 8), fg="gray").grid(row=14, column=0, columnspan=2, pady=2)

        def guardar_cambios():
            try:
                nuevo_ticker = entry_ticker.get().strip().upper()
                if not nuevo_ticker:
                    messagebox.showerror("Error", "El ticker no puede estar vacío")
                    return

                if nuevo_ticker != ticker_editar:
                    existe = any(p.get("ticker_symbol") == nuevo_ticker for p in parametros)
                    if existe:
                        messagebox.showerror("Error", f"Ya existe {nuevo_ticker} en este slot")
                        return

                # Validar y convertir fechas de DD-MM-YYYY a YYYY-MM-DD
                fecha_inicio_display = entry_fecha_inicio.get().strip()
                fecha_fin_display = entry_fecha_fin.get().strip()
                fecha_inicio = None
                fecha_fin = None
                if fecha_inicio_display:
                    fecha_inicio = fecha_display_to_iso(fecha_inicio_display)
                    if fecha_inicio is None:
                        messagebox.showerror("Error", "Fecha inicio debe tener formato DD-MM-YYYY")
                        return
                if fecha_fin_display:
                    fecha_fin = fecha_display_to_iso(fecha_fin_display)
                    if fecha_fin is None:
                        messagebox.showerror("Error", "Fecha fin debe tener formato DD-MM-YYYY")
                        return

                nuevo_param = {
                    "ticker_symbol": nuevo_ticker,
                    "origen": param_editar.get("origen", "editado"),
                    "compra_pct": float(entry_compra.get()),
                    "venta_pct": float(entry_venta.get()),
                    "ganancia_min_pct": float(entry_ganancia.get()),
                    "compra_multiple": int(entry_compra_mult.get()) if entry_compra_mult.get().strip() else None,
                    "venta_multiple": int(entry_venta_mult.get()) if entry_venta_mult.get().strip() else None,
                    "limite_tipo": limite_tipo_var.get(),
                    "limite_valor": float(entry_limite_valor.get()),
                    "promedio_minimos": float(entry_prom_min.get()) if entry_prom_min.get().strip() else 0,
                    "promedio_maximos": float(entry_prom_max.get()) if entry_prom_max.get().strip() else 0,
                    "fecha_inicio": fecha_inicio if fecha_inicio else None,
                    "fecha_fin": fecha_fin if fecha_fin else None
                }



                # Verificar traslape de fechas (excluyendo el parámetro actual)
                traslapes = detectar_traslape_fechas(parametros, nuevo_ticker, fecha_inicio, fecha_fin, excluir_index=param_index)
                if traslapes:
                    periodos_traslape = "\n".join([
                        f"  - {fecha_iso_to_display(t['fecha_inicio'])} a {fecha_iso_to_display(t['fecha_fin'])}"
                        for t in traslapes
                    ])
                    resp = messagebox.askyesno("ADVERTENCIA: Traslape de fechas",
                        f"El período ingresado para {nuevo_ticker} se traslapa con:\n{periodos_traslape}\n\n"
                        f"Esto puede causar conflictos al regenerar señales históricas.\n\n"
                        f"¿Deseas continuar de todos modos?",
                        icon='warning')
                    if not resp:
                        return

                parametros[param_index] = nuevo_param
                guardar_parametros_activos(datos_slots)
                actualizar_tabla_slot(slot_id)
                actualizar_titulos_pestanas()
                ventana_editar.destroy()
                messagebox.showinfo("Guardado", f"Parámetros de {nuevo_ticker} actualizados")

            except ValueError as e:
                messagebox.showerror("Error", f"Valores inválidos: {e}")

        tk.Button(frame_form, text="Guardar cambios", command=guardar_cambios,
                  bg="#ffc107", fg="black", font=("Arial", 10, "bold")).grid(row=15, column=0, columnspan=2, pady=20)

    def exportar_activos_excel():
        """Exporta los parámetros de todos los slots a Excel"""
        tiene_datos = any(obtener_parametros_slot(datos_slots, s) for s in ["1", "2", "3", "4", "5"])
        if not tiene_datos:
            messagebox.showwarning("Sin datos", "No hay parámetros activos para exportar")
            return

        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Parametros_Activos.xlsx"
        )

        if not ruta_excel:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            headers = ["Symbol", "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N", "Límite", "Valor Lím", "Prom.Mín%", "Prom.Máx%"]

            for slot_id in ["1", "2", "3", "4", "5"]:
                parametros = obtener_parametros_slot(datos_slots, slot_id)
                if not parametros:
                    continue

                nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                ws = wb.create_sheet(title=nombre_slot[:31])

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                for row_idx, param in enumerate(parametros, 2):
                    ws.cell(row=row_idx, column=1, value=param.get("ticker_symbol", "")).border = border
                    ws.cell(row=row_idx, column=2, value=param.get("compra_pct", 0)).border = border
                    ws.cell(row=row_idx, column=3, value=param.get("venta_pct", 0)).border = border
                    ws.cell(row=row_idx, column=4, value=param.get("ganancia_min_pct", 0)).border = border
                    ws.cell(row=row_idx, column=5, value=param.get("compra_multiple") or "").border = border
                    ws.cell(row=row_idx, column=6, value=param.get("venta_multiple") or "").border = border
                    ws.cell(row=row_idx, column=7, value=param.get("limite_tipo", "acciones")).border = border
                    ws.cell(row=row_idx, column=8, value=param.get("limite_valor", 10)).border = border
                    ws.cell(row=row_idx, column=9, value=param.get("promedio_minimos") or "").border = border
                    ws.cell(row=row_idx, column=10, value=param.get("promedio_maximos") or "").border = border

                col_widths = {"A": 10, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10, "I": 12, "J": 12}
                for col, width in col_widths.items():
                    ws.column_dimensions[col].width = width

            wb.save(ruta_excel)
            messagebox.showinfo("Exportado", f"Parámetros exportados a:\n{ruta_excel}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")

    # Botones
    tk.Button(frame_botones, text="Agregar desde JSON", command=agregar_desde_json,
              bg="#007bff", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Agregar personalizado", command=agregar_personalizado,
              bg="#17a2b8", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)

    def calcular_ponderado_slot():
        """Calcula parámetros ponderados desde análisis de 12 meses (solo Slot 1 o 2)"""
        ventana_pond = tk.Toplevel(ventana_params)
        ventana_pond.title("Calcular Parámetros Ponderados")
        ventana_pond.geometry("450x380")
        ventana_pond.transient(ventana_params)
        ventana_pond.grab_set()

        tk.Label(ventana_pond, text="Factores de Ponderación",
                 font=("Arial", 12, "bold")).pack(pady=10)

        tk.Label(ventana_pond, text="Los parámetros se calculan promediando Rentabilidad y Margen Prom\n"
                 "de cada período, ponderados por estos factores:",
                 font=("Arial", 9), fg="gray").pack(pady=5)

        frame_factores = tk.Frame(ventana_pond)
        frame_factores.pack(pady=10)

        tk.Label(frame_factores, text="Completo (12 meses):", font=("Arial", 10)).grid(row=0, column=0, sticky="w", padx=5, pady=3)
        entry_completo = tk.Entry(frame_factores, width=10)
        entry_completo.insert(0, "0.5")
        entry_completo.grid(row=0, column=1, padx=5, pady=3)

        tk.Label(frame_factores, text="Últimos 6 meses:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", padx=5, pady=3)
        entry_6m = tk.Entry(frame_factores, width=10)
        entry_6m.insert(0, "0.3")
        entry_6m.grid(row=1, column=1, padx=5, pady=3)

        tk.Label(frame_factores, text="Últimos 3 meses:", font=("Arial", 10)).grid(row=2, column=0, sticky="w", padx=5, pady=3)
        entry_3m = tk.Entry(frame_factores, width=10)
        entry_3m.insert(0, "0.2")
        entry_3m.grid(row=2, column=1, padx=5, pady=3)

        tk.Label(ventana_pond, text="Slot destino (solo 1 o 2):", font=("Arial", 10)).pack(pady=(15, 5))
        combo_slot = ttk.Combobox(ventana_pond, values=["1", "2"], state="readonly", width=10)
        combo_slot.current(0)

        def actualizar_factores(event=None):
            slot = combo_slot.get()
            entry_completo.delete(0, tk.END)
            entry_6m.delete(0, tk.END)
            entry_3m.delete(0, tk.END)
            if slot == "1":
                entry_completo.insert(0, "0.5")
                entry_6m.insert(0, "0.3")
                entry_3m.insert(0, "0.2")
            else:  # Slot 2
                entry_completo.insert(0, "0.4")
                entry_6m.insert(0, "0.3")
                entry_3m.insert(0, "0.3")

        combo_slot.bind("<<ComboboxSelected>>", actualizar_factores)
        combo_slot.pack()

        tk.Label(ventana_pond, text="Período de análisis:", font=("Arial", 10)).pack(pady=(15, 5))
        # Leer períodos disponibles del JSON de análisis (dinámico)
        periodos_disponibles = []
        try:
            with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
                analisis_check = json.load(f)
            for key in analisis_check.keys():
                partes = key.split('_')
                if len(partes) >= 4:  # Datos_TICKER_PART1_PART2
                    periodo = '_'.join(partes[2:])
                    if periodo not in periodos_disponibles:
                        periodos_disponibles.append(periodo)
            periodos_disponibles = sorted(periodos_disponibles)
        except Exception:
            periodos_disponibles = ["FEB25_FEB26"]
        combo_periodo = ttk.Combobox(ventana_pond, values=periodos_disponibles, state="readonly", width=15)
        combo_periodo.current(len(periodos_disponibles) - 1)  # Seleccionar el más reciente
        combo_periodo.pack()

        var_todos_tickers = tk.BooleanVar(value=False)

        def toggle_combo(*args):
            combo_periodo.config(state="disabled" if var_todos_tickers.get() else "readonly")

        tk.Checkbutton(ventana_pond, text="Todos los tickers (análisis más reciente de cada uno)",
                       variable=var_todos_tickers, command=toggle_combo,
                       font=("Arial", 9)).pack(pady=(4, 0))

        def _periodo_sort_key(p):
            """Devuelve (año, mes) de la fecha de fin del período (ej. FEB26 → (2026,2))"""
            meses = {'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
                     'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12}
            fin = p.split('_')[-1]
            mes = meses.get(fin[:3].upper(), 0)
            try:
                anio = int('20' + fin[3:]) if len(fin[3:]) == 2 else int(fin[3:])
            except ValueError:
                anio = 0
            return (anio, mes)

        def _calcular_param_ticker(analisis, json_key, ticker, periodo_analisis, factores,
                                   fecha_inicio_vig, fecha_fin_vig):
            """Calcula parámetros ponderados para un ticker a partir de su entrada en el JSON."""
            datos_ticker = analisis[json_key]
            periodos_datos = {}
            for periodo_key, periodo_val in datos_ticker.items():
                if periodo_key in ['ticker', 'fecha_guardado', '_ticker_symbol', 'periodos']:
                    continue
                if isinstance(periodo_val, dict):
                    for objetivo_key, objetivo_val in periodo_val.items():
                        if isinstance(objetivo_val, dict) and 'parametros_optimos' in objetivo_val:
                            periodos_datos[(periodo_key, objetivo_key)] = objetivo_val['parametros_optimos']
            if not periodos_datos:
                return None
            campos = ['compra_pct', 'venta_pct', 'ganancia_minima_pct', 'promedio_minimos', 'promedio_maximos']
            resultados = {c: {'rentabilidad': 0, 'margen_prom': 0} for c in campos}
            for (periodo, objetivo), p in periodos_datos.items():
                factor = factores.get(periodo, 0)
                obj_key = 'rentabilidad' if objetivo == 'rentabilidad' else 'margen_prom'
                for campo in campos:
                    if campo in p:
                        resultados[campo][obj_key] += p[campo] * factor
            slot_params = {c: round((resultados[c]['rentabilidad'] + resultados[c]['margen_prom']) / 2, 1) for c in campos}
            po = ['completo', 'ultimos_6_meses', 'ultimos_3_meses']
            compra_mult_final = round((round(sum(periodos_datos.get((p, 'rentabilidad'), {}).get('compra_multiple') or 0 for p in po) / 3) +
                                       round(sum(periodos_datos.get((p, 'margen_prom'),  {}).get('compra_multiple') or 0 for p in po) / 3)) / 2)
            venta_mult_final  = round((round(sum(periodos_datos.get((p, 'rentabilidad'), {}).get('venta_multiple')  or 0 for p in po) / 3) +
                                       round(sum(periodos_datos.get((p, 'margen_prom'),  {}).get('venta_multiple')  or 0 for p in po) / 3)) / 2)
            return {
                'ticker_symbol': ticker,
                'origen': f'ponderado_{periodo_analisis}',
                'compra_pct': slot_params['compra_pct'],
                'venta_pct': slot_params['venta_pct'],
                'ganancia_min_pct': slot_params['ganancia_minima_pct'],
                'compra_multiple': compra_mult_final if compra_mult_final > 0 else None,
                'venta_multiple': venta_mult_final if venta_mult_final > 0 else None,
                'limite_tipo': 'acciones',
                'limite_valor': 10.0,
                'promedio_minimos': slot_params['promedio_minimos'],
                'promedio_maximos': slot_params['promedio_maximos'],
                'fecha_inicio': fecha_inicio_vig,
                'fecha_fin': fecha_fin_vig
            }

        def ejecutar_calculo():
            try:
                f_completo = float(entry_completo.get())
                f_6m = float(entry_6m.get())
                f_3m = float(entry_3m.get())

                suma = f_completo + f_6m + f_3m
                if abs(suma - 1.0) > 0.01:
                    messagebox.showwarning("Advertencia", f"Los factores suman {suma:.2f}, deberían sumar 1.0")

                slot_destino = combo_slot.get()

                with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
                    analisis = json.load(f)

                factores = {'completo': f_completo, 'ultimos_6_meses': f_6m, 'ultimos_3_meses': f_3m}
                fecha_inicio_vig = datetime.now().strftime('%Y-%m-%d')
                fecha_fin_vig = (datetime.now() + timedelta(days=91)).strftime('%Y-%m-%d')

                nuevos_params = []

                if var_todos_tickers.get():
                    # Para cada ticker, encontrar su análisis MÁS RECIENTE (por fecha de fin del período)
                    # y usar solo ese — sin mezclar análisis de diferentes momentos
                    ticker_mejor = {}  # ticker → (sort_key, json_key, periodo_nombre)
                    for key in analisis.keys():
                        partes = key.split('_')
                        if len(partes) < 4:
                            continue
                        ticker = partes[1]
                        periodo = '_'.join(partes[2:])
                        sk = _periodo_sort_key(periodo)
                        if ticker not in ticker_mejor or sk > ticker_mejor[ticker][0]:
                            ticker_mejor[ticker] = (sk, key, periodo)

                    for ticker, (_, best_key, best_periodo) in ticker_mejor.items():
                        param = _calcular_param_ticker(analisis, best_key, ticker, best_periodo,
                                                       factores, fecha_inicio_vig, fecha_fin_vig)
                        if param:
                            nuevos_params.append(param)
                else:
                    # Un período específico seleccionado por el usuario
                    periodo_analisis = combo_periodo.get()
                    for key in analisis.keys():
                        partes = key.split('_')
                        if len(partes) < 4 or '_'.join(partes[2:]) != periodo_analisis:
                            continue
                        ticker = partes[1]
                        param = _calcular_param_ticker(analisis, key, ticker, periodo_analisis,
                                                       factores, fecha_inicio_vig, fecha_fin_vig)
                        if param:
                            nuevos_params.append(param)

                if nuevos_params:
                    datos_slots["slots"][slot_destino]["nombre"] = f"{slot_destino}.-Ponderado-12m"
                    datos_slots["slots"][slot_destino]["parametros_activos"] = nuevos_params
                    guardar_parametros_activos(datos_slots)
                    total_slot = len(nuevos_params)
                    if var_todos_tickers.get():
                        detalle = f"Todos los tickers ({total_slot}), cada uno con su análisis más reciente"
                    else:
                        detalle = f"Período {combo_periodo.get()}: {total_slot} tickers"
                    messagebox.showinfo("Completado",
                        f"Slot {slot_destino} actualizado\n\n"
                        f"{detalle}\n\n"
                        f"Vigencia: {fecha_inicio_vig} → {fecha_fin_vig}")
                    ventana_pond.destroy()
                    actualizar_tabla_slot(slot_destino)
                    actualizar_titulos_pestanas()
                else:
                    messagebox.showwarning("Sin datos", f"No se encontraron análisis para {periodo_analisis}")

            except Exception as e:
                messagebox.showerror("Error", f"Error al calcular: {e}")

        tk.Button(ventana_pond, text="Calcular y Guardar", command=ejecutar_calculo,
                  bg="#28a745", fg="white", font=("Arial", 10), width=20).pack(pady=20)

    def calcular_slots_3_4():
        """Calcula Slot 3 (largo) y Slot 4 (corto) optimizando factor por ticker"""
        from datetime import timedelta

        # Constantes
        FACTOR_MIN_CORTO = 0.5
        FACTOR_MAX_CORTO = 1.0
        FACTOR_MIN_LARGO = 1.0
        FACTOR_MAX_LARGO = 1.5
        PASO_FACTOR = 0.1
        MESES_ANALISIS = 2

        def aplicar_factor(params_base, factor):
            params = params_base.copy()
            params['compra_pct'] = round(params.get('compra_pct', -1.0) * factor, 1)
            params['venta_pct'] = round(params.get('venta_pct', 2.0) * factor, 1)
            gan_base = params.get('ganancia_min_pct', 2.5)
            if factor > 1.0:
                ajuste = (factor - 1.0) * 1.5
                params['ganancia_min_pct'] = round(min(gan_base + ajuste, 3.5), 1)
            else:
                ajuste = (1.0 - factor) * 1.5
                params['ganancia_min_pct'] = round(max(gan_base - ajuste, 1.5), 1)
            return params

        def simular_operaciones(df_ticker, params, limite_acciones=10):
            if df_ticker.empty or params is None:
                return {'rentabilidad': 0, 'operaciones': 0}
            compra_pct = params.get('compra_pct', -1.0)
            venta_pct = params.get('venta_pct', 2.0)
            ganancia_min_pct = params.get('ganancia_min_pct', 3.0)
            compra_mult = params.get('compra_multiple')
            venta_mult = params.get('venta_multiple')
            prom_min = params.get('promedio_minimos', -5.0)
            prom_max = params.get('promedio_maximos', 5.0)
            if abs(prom_min) > 50:
                prom_min = prom_min / 100
            if abs(prom_max) > 50:
                prom_max = prom_max / 100
            cartera = []
            total_compras = 0
            total_ventas = 0
            num_operaciones = 0
            df_ticker = df_ticker.reset_index(drop=True)
            for i, row in df_ticker.iterrows():
                cierre = row['Close']
                if i == 0:
                    primer_cierre = cierre
                    acum_pct = 0
                else:
                    acum_pct = ((cierre - primer_cierre) / primer_cierre) * 100
                precio_compra = cierre * (1 + compra_pct / 100)
                precio_venta = cierre * (1 + venta_pct / 100)
                if len(cartera) < limite_acciones:
                    comprar = False
                    cant_compra = 1
                    if acum_pct <= compra_pct:
                        comprar = True
                        if compra_mult and acum_pct <= prom_min:
                            cant_compra = min(compra_mult, limite_acciones - len(cartera))
                    if comprar:
                        for _ in range(cant_compra):
                            if len(cartera) < limite_acciones:
                                cartera.append(precio_compra)
                                total_compras += precio_compra
                                num_operaciones += 1
                if cartera:
                    precio_compra_fifo = cartera[0]
                    ganancia_actual = ((precio_venta - precio_compra_fifo) / precio_compra_fifo) * 100
                    if ganancia_actual >= ganancia_min_pct and acum_pct >= venta_pct:
                        cant_venta = 1
                        if venta_mult and acum_pct >= prom_max:
                            cant_venta = min(venta_mult, len(cartera))
                        for _ in range(cant_venta):
                            if cartera:
                                cartera.pop(0)
                                total_ventas += precio_venta
                                num_operaciones += 1
            if cartera and not df_ticker.empty:
                valor_cartera = len(cartera) * df_ticker.iloc[-1]['Close']
            else:
                valor_cartera = 0
            if total_compras > 0:
                rentabilidad = ((total_ventas + valor_cartera - total_compras) / total_compras) * 100
            else:
                rentabilidad = 0
            return {'rentabilidad': round(rentabilidad, 2), 'operaciones': num_operaciones}

        def encontrar_mejor_factor(df_ticker, params_base, factor_min, factor_max, paso):
            mejor_factor = 1.0
            mejor_rent = -999
            mejor_params = params_base.copy()
            factor = factor_min
            while factor <= factor_max + 0.001:
                params_test = aplicar_factor(params_base, factor)
                sim = simular_operaciones(df_ticker, params_test)
                if sim['rentabilidad'] > mejor_rent:
                    mejor_rent = sim['rentabilidad']
                    mejor_factor = factor
                    mejor_params = params_test.copy()
                factor = round(factor + paso, 1)
            return mejor_factor, mejor_rent, mejor_params

        # Ventana de progreso
        ventana_calc = tk.Toplevel(ventana_params)
        ventana_calc.title("Calculando Slot 3 y 4")
        ventana_calc.geometry("600x500")
        ventana_calc.transient(ventana_params)

        tk.Label(ventana_calc, text="Cálculo de Slot 3 (largo) y Slot 4 (corto)",
                 font=("Arial", 12, "bold")).pack(pady=10)
        tk.Label(ventana_calc, text=f"Slot 3: factores {FACTOR_MIN_LARGO} a {FACTOR_MAX_LARGO}\n"
                 f"Slot 4: factores {FACTOR_MIN_CORTO} a {FACTOR_MAX_CORTO}",
                 font=("Arial", 9), fg="gray").pack(pady=5)

        # Frame para resultados
        frame_tree = tk.Frame(ventana_calc)
        frame_tree.pack(fill="both", expand=True, padx=10, pady=10)

        cols = ("Ticker", "Base", "Rent Base", "Factor S3", "Rent S3", "Factor S4", "Rent S4")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=15)
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=80, anchor="center")
        tree.pack(fill="both", expand=True)

        lbl_status = tk.Label(ventana_calc, text="Calculando...", font=("Arial", 10))
        lbl_status.pack(pady=5)

        ventana_calc.update()

        try:
            _cargar_dependencias_analisis()
            # Cargar datos
            datos_calc = cargar_parametros_activos()
            slots = {}
            for slot_num in ['1', '2']:
                slots[slot_num] = {}
                params_list = datos_calc['slots'].get(slot_num, {}).get('parametros_activos', [])
                for p in params_list:
                    ticker = p.get('ticker_symbol')
                    if ticker:
                        slots[slot_num][ticker] = p.copy()

            # Cargar precios
            precios_csv = obtener_carpeta_datos() / "auto_update_log.csv"
            df = pd.read_csv(precios_csv)
            df['Date'] = pd.to_datetime(df['Date'])
            fecha_fin = df['Date'].max()
            fecha_inicio = fecha_fin - timedelta(days=MESES_ANALISIS * 30)
            df = df[df['Date'] >= fecha_inicio].copy()
            df = df.sort_values(['Ticker', 'Date'])

            tickers = sorted(set(slots['1'].keys()) | set(slots['2'].keys()))
            resultados = []
            params_slot3 = []
            params_slot4 = []

            for ticker in tickers:
                df_ticker = df[df['Ticker'] == ticker].copy()

                # Comparar Slot 1 vs 2
                params_s1 = slots['1'].get(ticker)
                params_s2 = slots['2'].get(ticker)
                rent_s1 = simular_operaciones(df_ticker, params_s1)['rentabilidad'] if params_s1 else -999
                rent_s2 = simular_operaciones(df_ticker, params_s2)['rentabilidad'] if params_s2 else -999

                if rent_s1 >= rent_s2:
                    mejor = '1'
                    params_base = params_s1
                    rent_base = rent_s1
                else:
                    mejor = '2'
                    params_base = params_s2
                    rent_base = rent_s2

                if not params_base:
                    continue

                # Optimizar Slot 3 (largo)
                factor_s3, rent_s3, params_s3 = encontrar_mejor_factor(
                    df_ticker, params_base, FACTOR_MIN_LARGO, FACTOR_MAX_LARGO, PASO_FACTOR)

                # Optimizar Slot 4 (corto)
                factor_s4, rent_s4, params_s4 = encontrar_mejor_factor(
                    df_ticker, params_base, FACTOR_MIN_CORTO, FACTOR_MAX_CORTO, PASO_FACTOR)

                resultados.append({
                    'Ticker': ticker, 'Mejor': f'S{mejor}', 'Rent_Base': rent_base,
                    'Factor_S3': factor_s3, 'Rent_S3': rent_s3,
                    'Factor_S4': factor_s4, 'Rent_S4': rent_s4
                })

                # Preparar parámetros para guardar
                fecha_hoy = datetime.now().strftime('%Y-%m-%d')
                fecha_fin_str = (datetime.now() + timedelta(days=60)).strftime('%Y-%m-%d')

                p_largo = {
                    'ticker_symbol': ticker, 'origen': f'Slot{mejor}', 'factor_aplicado': factor_s3,
                    'compra_pct': params_s3['compra_pct'], 'venta_pct': params_s3['venta_pct'],
                    'ganancia_min_pct': params_s3['ganancia_min_pct'],
                    'compra_multiple': params_base.get('compra_multiple'),
                    'venta_multiple': params_base.get('venta_multiple'),
                    'limite_tipo': params_base.get('limite_tipo', 'acciones'),
                    'limite_valor': params_base.get('limite_valor', 10.0),
                    'promedio_minimos': params_base.get('promedio_minimos'),
                    'promedio_maximos': params_base.get('promedio_maximos'),
                    'fecha_inicio': fecha_hoy, 'fecha_fin': fecha_fin_str
                }
                params_slot3.append(p_largo)

                p_corto = {
                    'ticker_symbol': ticker, 'origen': f'Slot{mejor}', 'factor_aplicado': factor_s4,
                    'compra_pct': params_s4['compra_pct'], 'venta_pct': params_s4['venta_pct'],
                    'ganancia_min_pct': params_s4['ganancia_min_pct'],
                    'compra_multiple': params_base.get('compra_multiple'),
                    'venta_multiple': params_base.get('venta_multiple'),
                    'limite_tipo': params_base.get('limite_tipo', 'acciones'),
                    'limite_valor': params_base.get('limite_valor', 10.0),
                    'promedio_minimos': params_base.get('promedio_minimos'),
                    'promedio_maximos': params_base.get('promedio_maximos'),
                    'fecha_inicio': fecha_hoy, 'fecha_fin': fecha_fin_str
                }
                params_slot4.append(p_corto)

                # Añadir a la tabla
                tree.insert("", "end", values=(
                    ticker, f'S{mejor}', f'{rent_base:.2f}%',
                    factor_s3, f'{rent_s3:.2f}%',
                    factor_s4, f'{rent_s4:.2f}%'
                ))
                ventana_calc.update()

            # Resumen
            mejoras_s3 = sum(1 for r in resultados if r['Rent_S3'] > r['Rent_Base'])
            mejoras_s4 = sum(1 for r in resultados if r['Rent_S4'] > r['Rent_Base'])
            lbl_status.config(text=f"Slot 3 mejora: {mejoras_s3}/{len(resultados)} | Slot 4 mejora: {mejoras_s4}/{len(resultados)}")

            def guardar_slots():
                mes_actual = datetime.now().strftime('%B').lower()[:3]
                datos_slots['slots']['3'] = {
                    'nombre': f'3.-CLAUDE-largo-{mes_actual}',
                    'parametros_activos': params_slot3
                }
                datos_slots['slots']['4'] = {
                    'nombre': f'4.-CLAUDE-corto-{mes_actual}',
                    'parametros_activos': params_slot4
                }
                guardar_parametros_activos(datos_slots)
                for slot_n in ['3', '4']:
                    actualizar_tabla_slot(slot_n)
                actualizar_titulos_pestanas()
                messagebox.showinfo("Guardado", "Slot 3 y Slot 4 guardados correctamente")
                ventana_calc.destroy()

            tk.Button(ventana_calc, text="Guardar Slot 3 y 4", command=guardar_slots,
                      bg="#28a745", fg="white", font=("Arial", 10), width=20).pack(pady=10)

        except Exception as e:
            lbl_status.config(text=f"Error: {e}")
            messagebox.showerror("Error", f"Error al calcular: {e}")

    def calcular_slot_5():
        """Calcula Slot 5 - Mejor de 1-4 con ajuste ±30%"""
        from datetime import timedelta

        # Configuración
        DIAS_ANALISIS = 30
        DIAS_VIGENCIA = 15
        AJUSTE_MIN = -30
        AJUSTE_MAX = 30
        PASO_AJUSTE = 5

        def aplicar_ajuste(params_base, ajuste_compra, ajuste_venta):
            params = params_base.copy()
            compra_base = params.get('compra_pct', -1.0)
            venta_base = params.get('venta_pct', 2.0)
            factor_compra = 1 + (ajuste_compra / 100)
            factor_venta = 1 + (ajuste_venta / 100)
            params['compra_pct'] = round(compra_base * factor_compra, 2)
            params['venta_pct'] = round(venta_base * factor_venta, 2)
            return params

        def simular_ops(df_ticker, params, limite_acciones=10):
            if df_ticker.empty or params is None:
                return {'rentabilidad': -999, 'operaciones': 0}
            compra_pct = params.get('compra_pct', -1.0)
            venta_pct = params.get('venta_pct', 2.0)
            ganancia_min_pct = params.get('ganancia_min_pct', 3.0)
            compra_mult = params.get('compra_multiple')
            venta_mult = params.get('venta_multiple')
            prom_min = params.get('promedio_minimos', -5.0)
            prom_max = params.get('promedio_maximos', 5.0)
            if abs(prom_min) > 50:
                prom_min = prom_min / 100
            if abs(prom_max) > 50:
                prom_max = prom_max / 100
            cartera = []
            total_compras = 0
            total_ventas = 0
            num_operaciones = 0
            df_ticker = df_ticker.reset_index(drop=True)
            for i, row in df_ticker.iterrows():
                cierre = row['Close']
                if i == 0:
                    primer_cierre = cierre
                    acum_pct = 0
                else:
                    acum_pct = ((cierre - primer_cierre) / primer_cierre) * 100
                precio_compra = cierre * (1 + compra_pct / 100)
                precio_venta = cierre * (1 + venta_pct / 100)
                if len(cartera) < limite_acciones:
                    comprar = False
                    cant_compra = 1
                    if acum_pct <= compra_pct:
                        comprar = True
                        if compra_mult and acum_pct <= prom_min:
                            cant_compra = min(compra_mult, limite_acciones - len(cartera))
                    if comprar:
                        for _ in range(cant_compra):
                            if len(cartera) < limite_acciones:
                                cartera.append(precio_compra)
                                total_compras += precio_compra
                                num_operaciones += 1
                if cartera:
                    precio_compra_fifo = cartera[0]
                    ganancia_actual = ((precio_venta - precio_compra_fifo) / precio_compra_fifo) * 100
                    if ganancia_actual >= ganancia_min_pct and acum_pct >= venta_pct:
                        cant_venta = 1
                        if venta_mult and acum_pct >= prom_max:
                            cant_venta = min(venta_mult, len(cartera))
                        for _ in range(cant_venta):
                            if cartera:
                                cartera.pop(0)
                                total_ventas += precio_venta
                                num_operaciones += 1
            if cartera and not df_ticker.empty:
                valor_cartera = len(cartera) * df_ticker.iloc[-1]['Close']
            else:
                valor_cartera = 0
            if total_compras > 0:
                rentabilidad = ((total_ventas + valor_cartera - total_compras) / total_compras) * 100
            else:
                rentabilidad = 0
            return {'rentabilidad': round(rentabilidad, 2), 'operaciones': num_operaciones}

        def encontrar_mejor_slot_base(df_ticker, slots, ticker):
            mejor_slot = '1'
            mejor_rent = -999
            mejor_params = None
            for slot_num in ['1', '2', '3', '4']:
                params = slots[slot_num].get(ticker)
                if params:
                    sim = simular_ops(df_ticker, params)
                    if sim['rentabilidad'] > mejor_rent:
                        mejor_rent = sim['rentabilidad']
                        mejor_slot = slot_num
                        mejor_params = params.copy()
            return mejor_slot, mejor_rent, mejor_params

        def optimizar_ajuste(df_ticker, params_base):
            mejor_ajuste_c = 0
            mejor_ajuste_v = 0
            mejor_rent = simular_ops(df_ticker, params_base)['rentabilidad']
            mejor_params = params_base.copy()
            for ajuste_c in range(AJUSTE_MIN, AJUSTE_MAX + 1, PASO_AJUSTE):
                for ajuste_v in range(AJUSTE_MIN, AJUSTE_MAX + 1, PASO_AJUSTE):
                    params_test = aplicar_ajuste(params_base, ajuste_c, ajuste_v)
                    sim = simular_ops(df_ticker, params_test)
                    if sim['rentabilidad'] > mejor_rent:
                        mejor_rent = sim['rentabilidad']
                        mejor_ajuste_c = ajuste_c
                        mejor_ajuste_v = ajuste_v
                        mejor_params = params_test.copy()
            return mejor_ajuste_c, mejor_ajuste_v, mejor_rent, mejor_params

        # Ventana de progreso
        ventana_calc = tk.Toplevel(ventana_params)
        ventana_calc.title("Calculando Slot 5")
        ventana_calc.geometry("700x500")
        ventana_calc.transient(ventana_params)

        tk.Label(ventana_calc, text="Cálculo de Slot 5 (Optimizado)",
                 font=("Arial", 12, "bold")).pack(pady=10)
        tk.Label(ventana_calc, text=f"Base: Mejor de Slots 1-4 | Ajuste: ±30% | Data: {DIAS_ANALISIS} días",
                 font=("Arial", 9), fg="gray").pack(pady=5)

        frame_tree = tk.Frame(ventana_calc)
        frame_tree.pack(fill="both", expand=True, padx=10, pady=10)

        cols = ("Ticker", "Base", "Rent Base", "Aj.Compra", "Aj.Venta", "Rent Opt", "Mejora")
        tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=15)
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=90, anchor="center")
        tree.pack(fill="both", expand=True)

        lbl_status = tk.Label(ventana_calc, text="Calculando...", font=("Arial", 10))
        lbl_status.pack(pady=5)

        ventana_calc.update()

        try:
            _cargar_dependencias_analisis()  # Asegurar que pd/np estén disponibles

            # Cargar datos
            datos_calc = cargar_parametros_activos()
            slots = {}
            for slot_num in ['1', '2', '3', '4']:
                slots[slot_num] = {}
                params_list = datos_calc['slots'].get(slot_num, {}).get('parametros_activos', [])
                for p in params_list:
                    ticker = p.get('ticker_symbol')
                    if ticker:
                        slots[slot_num][ticker] = p.copy()

            # Cargar precios
            precios_csv = obtener_carpeta_datos() / "auto_update_log.csv"
            df = pd.read_csv(precios_csv)
            df['Date'] = pd.to_datetime(df['Date'])
            fecha_fin = df['Date'].max()
            fecha_inicio = fecha_fin - timedelta(days=DIAS_ANALISIS)
            df = df[df['Date'] >= fecha_inicio].copy()
            df = df.sort_values(['Ticker', 'Date'])

            all_tickers = set()
            for slot_num in ['1', '2', '3', '4']:
                all_tickers.update(slots[slot_num].keys())
            tickers = sorted(all_tickers)

            resultados = []
            params_slot5 = []

            for ticker in tickers:
                df_ticker = df[df['Ticker'] == ticker].copy()
                if df_ticker.empty:
                    continue

                mejor_slot, rent_base, params_base = encontrar_mejor_slot_base(df_ticker, slots, ticker)
                if not params_base:
                    continue

                ajuste_c, ajuste_v, rent_opt, params_opt = optimizar_ajuste(df_ticker, params_base)
                mejora = rent_opt - rent_base

                resultados.append({
                    'Ticker': ticker, 'Base': f'S{mejor_slot}', 'Rent_Base': rent_base,
                    'Ajuste_C': ajuste_c, 'Ajuste_V': ajuste_v,
                    'Rent_Opt': rent_opt, 'Mejora': mejora
                })

                fecha_hoy = datetime.now().strftime('%Y-%m-%d')
                fecha_fin_vig = (datetime.now() + timedelta(days=DIAS_VIGENCIA)).strftime('%Y-%m-%d')

                p_opt = {
                    'ticker_symbol': ticker,
                    'origen': f'Slot{mejor_slot} hasta ±30%',
                    'slot_base': mejor_slot,
                    'ajuste_compra': ajuste_c,
                    'ajuste_venta': ajuste_v,
                    'compra_pct': params_opt['compra_pct'],
                    'venta_pct': params_opt['venta_pct'],
                    'ganancia_min_pct': params_base.get('ganancia_min_pct', 3.0),
                    'compra_multiple': params_base.get('compra_multiple'),
                    'venta_multiple': params_base.get('venta_multiple'),
                    'limite_tipo': params_base.get('limite_tipo', 'acciones'),
                    'limite_valor': params_base.get('limite_valor', 10.0),
                    'promedio_minimos': params_base.get('promedio_minimos'),
                    'promedio_maximos': params_base.get('promedio_maximos'),
                    'fecha_inicio': fecha_hoy,
                    'fecha_fin': fecha_fin_vig
                }
                params_slot5.append(p_opt)

                mejora_str = f"{mejora:+.2f}%" if mejora != 0 else "="
                tree.insert("", "end", values=(
                    ticker, f'S{mejor_slot}', f'{rent_base:.2f}%',
                    f'{ajuste_c:+d}%', f'{ajuste_v:+d}%',
                    f'{rent_opt:.2f}%', mejora_str
                ))
                ventana_calc.update()

            mejoras = sum(1 for r in resultados if r['Mejora'] > 0)
            lbl_status.config(text=f"Tickers mejorados: {mejoras}/{len(resultados)}")

            def guardar_slot5():
                mes_actual = datetime.now().strftime('%B').lower()[:3]
                dia_actual = datetime.now().strftime('%d')
                datos_slots['slots']['5'] = {
                    'nombre': f'5.-Optimizado-{mes_actual}{dia_actual}',
                    'parametros_activos': params_slot5
                }
                guardar_parametros_activos(datos_slots)
                actualizar_tabla_slot('5')
                actualizar_titulos_pestanas()
                messagebox.showinfo("Guardado", "Slot 5 guardado correctamente")
                ventana_calc.destroy()

            tk.Button(ventana_calc, text="Guardar Slot 5", command=guardar_slot5,
                      bg="#28a745", fg="white", font=("Arial", 10), width=20).pack(pady=10)

        except Exception as e:
            lbl_status.config(text=f"Error: {e}")
            messagebox.showerror("Error", f"Error al calcular: {e}")

    tk.Button(frame_botones, text="Calcular Slots 1/2", command=calcular_ponderado_slot,
              bg="#9b59b6", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Calcular Slot 3/4", command=calcular_slots_3_4,
              bg="#e67e22", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Calcular Slot 5", command=calcular_slot_5,
              bg="#3498db", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Editar", command=editar_parametro,
              bg="#ffc107", fg="black", font=("Arial", 9, "bold")).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Exportar a Excel", command=exportar_activos_excel,
              bg="#28a745", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Eliminar seleccionados", command=eliminar_seleccionados,
              bg="#ff6b6b", fg="white", font=("Arial", 9, "bold")).pack(side="right", padx=5)
    tk.Button(frame_botones, text="Cerrar", command=ventana_params.destroy).pack(side="right", padx=5)


# =========================
# Funciones para JSON de resultados (MODIFICADO - Estructura jerárquica)
# =========================
def cargar_resultados_json():
    """Carga todos los resultados guardados en el JSON"""
    if ARCHIVO_JSON is None or not ARCHIVO_JSON.exists():
        return {}

    with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def parametros_son_iguales(params_nuevos, params_existentes, tolerancia=0.01):
    """Compara si dos conjuntos de parámetros son iguales (con tolerancia para decimales)"""
    claves_comparar = ["compra_pct", "venta_pct", "ganancia_minima_pct", "suave_pct",
                       "limite_tipo", "limite_valor", "compra_multiple", "venta_multiple"]

    for clave in claves_comparar:
        val_nuevo = params_nuevos.get(clave)
        val_existente = params_existentes.get(clave)

        # Si ambos son None o iguales, continuar
        if val_nuevo == val_existente:
            continue

        # Si uno es None y otro no, son diferentes
        if val_nuevo is None or val_existente is None:
            return False

        # Para valores numéricos, comparar con tolerancia
        if isinstance(val_nuevo, (int, float)) and isinstance(val_existente, (int, float)):
            if abs(val_nuevo - val_existente) > tolerancia:
                return False
        else:
            # Para strings u otros tipos, comparación exacta
            if val_nuevo != val_existente:
                return False

    return True


def guardar_resultados_en_json():
    """Guarda los resultados actuales en el JSON (botón verde) - ESTRUCTURA JERÁRQUICA"""
    global resultados_analisis_actuales, ARCHIVO_JSON

    if not resultados_analisis_actuales:
        messagebox.showwarning("Sin resultados", "No hay resultados de análisis para guardar.")
        return

    if not verificar_ubicacion_json():
        return

    # Verificar que ARCHIVO_JSON esté configurado
    if ARCHIVO_JSON is None:
        messagebox.showerror("Error", "La ruta del archivo JSON no está configurada.")
        return

    try:
        # Cargar JSON existente
        if ARCHIVO_JSON.exists():
            with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
                datos_json = json.load(f)
        else:
            datos_json = {}

        # Obtener ticker del archivo actual
        ticker = resultados_analisis_actuales.get("ticker", "UNKNOWN")

        # Extraer ticker_symbol (ej: "META" de "Datos_META_ENE25_NOV25")
        ticker_symbol = extraer_ticker_symbol(ticker)
        print(f"[DEBUG] Ticker: {ticker} -> ticker_symbol: {ticker_symbol}")

        # Estructura jerárquica ticker -> período -> objetivo
        if ticker not in datos_json:
            datos_json[ticker] = {}

        # Guardar ticker_symbol a nivel del ticker principal
        if ticker_symbol:
            datos_json[ticker]["_ticker_symbol"] = ticker_symbol

        # Verificar que hay periodos para guardar
        periodos = resultados_analisis_actuales.get("periodos", {})
        if not periodos:
            messagebox.showwarning("Sin períodos", "No hay datos de períodos para guardar.")
            return

        registros_nuevos = 0
        registros_actualizados = 0

        for clave_periodo, datos in periodos.items():
            # La clave tiene formato "periodo_objetivo" (ej: "completo_rentabilidad")
            # Extraer período y objetivo
            objetivo_base = datos.get("objetivo", "rentabilidad")

            # Extraer solo el nombre del período (sin el objetivo)
            if "_rentabilidad" in clave_periodo:
                nombre_periodo = clave_periodo.replace("_rentabilidad", "")
            elif "_margen_prom" in clave_periodo:
                nombre_periodo = clave_periodo.replace("_margen_prom", "")
            else:
                nombre_periodo = clave_periodo

            if nombre_periodo not in datos_json[ticker]:
                datos_json[ticker][nombre_periodo] = {}

            # Crear el nuevo registro
            nuevo_registro = {
                "ticker_symbol": ticker_symbol,  # Símbolo para Yahoo Finance (ej: "META")
                "fecha_guardado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "fecha_inicial": datos.get("fecha_inicial", ""),
                "fecha_final": datos.get("fecha_final", ""),
                "parametros_optimos": {
                    "compra_pct": datos.get("compra_pct", 0),
                    "venta_pct": datos.get("venta_pct", 0),
                    "ganancia_minima_pct": datos.get("ganancia_min", 0),
                    "suave_pct": datos.get("suave_pct", 0),
                    "limite_tipo": datos.get("limite_tipo", "acciones"),
                    "limite_valor": datos.get("limite_valor", 10),
                    "compra_multiple": datos.get("compra_mult"),
                    "venta_multiple": datos.get("venta_mult"),
                    # Condiciones para compra/venta múltiple
                    "promedio_maximos": datos.get("promedio_maximos", 0),
                    "promedio_minimos": datos.get("promedio_minimos", 0)
                },
                "metricas": {
                    "rentabilidad_max": datos.get("rentabilidad_max", 0),
                    "margen_promedio": datos.get("margen_promedio", 0),
                    "rentab_promedio": datos.get("rentab_promedio", 0),
                    "max_margen": datos.get("max_margen", 0),
                    "max_aporte": datos.get("max_aporte", 0)
                },
                "estadisticas_var": {
                    "max_var": datos.get("max_var", 0),
                    "min_var": datos.get("min_var", 0),
                    "fecha_max_var": datos.get("fecha_max_var", ""),
                    "fecha_min_var": datos.get("fecha_min_var", ""),
                    "dif_var": datos.get("dif_var", 0),
                    "max_prom_var": datos.get("max_prom_var", 0),
                    "min_prom_var": datos.get("min_prom_var", 0),
                    "dif_prom_var": datos.get("dif_prom_var", 0)
                },
                "estadisticas_operaciones": {
                    "opc_compra": datos.get("opc_compra", 0),
                    "acciones_compradas": datos.get("acciones_compradas", 0),
                    "opc_venta": datos.get("opc_venta", 0),
                    "acciones_vendidas": datos.get("acciones_vendidas", 0),
                    "max_acc_cartera": datos.get("max_acc_cartera", 0),
                    "fecha_max_rentab": datos.get("fecha_max_rentab", "")
                }
            }

            # Buscar si ya existe un registro con los mismos parámetros
            objetivo_encontrado = None
            for objetivo_key, registro_existente in datos_json[ticker][nombre_periodo].items():
                if objetivo_key.startswith(objetivo_base):
                    if isinstance(registro_existente, dict) and "parametros_optimos" in registro_existente:
                        if parametros_son_iguales(nuevo_registro["parametros_optimos"],
                                                   registro_existente["parametros_optimos"]):
                            objetivo_encontrado = objetivo_key
                            break

            if objetivo_encontrado:
                # Actualizar registro existente (mismos parámetros)
                datos_json[ticker][nombre_periodo][objetivo_encontrado] = nuevo_registro
                registros_actualizados += 1
                print(f"[DEBUG] Actualizado: {ticker}/{nombre_periodo}/{objetivo_encontrado}")
            else:
                # Crear nuevo registro (parámetros diferentes)
                # Buscar un nombre único para el objetivo
                objetivo_final = objetivo_base
                contador = 2
                while objetivo_final in datos_json[ticker][nombre_periodo]:
                    objetivo_final = f"{objetivo_base}_{contador}"
                    contador += 1

                datos_json[ticker][nombre_periodo][objetivo_final] = nuevo_registro
                registros_nuevos += 1
                print(f"[DEBUG] Nuevo registro: {ticker}/{nombre_periodo}/{objetivo_final}")

        # Escribir JSON
        with open(ARCHIVO_JSON, 'w', encoding='utf-8') as f:
            json.dump(datos_json, f, indent=2, ensure_ascii=False)

        print(f"[DEBUG] JSON guardado exitosamente en: {ARCHIVO_JSON}")

        mensaje = f"Resultados guardados para {ticker}\n\n"
        if registros_nuevos > 0:
            mensaje += f"• {registros_nuevos} registro(s) nuevo(s)\n"
        if registros_actualizados > 0:
            mensaje += f"• {registros_actualizados} registro(s) actualizado(s)\n"
        mensaje += f"\nArchivo: {ARCHIVO_JSON}"

        messagebox.showinfo("Guardado exitoso", mensaje)
        btn_guardar_json.config(state="disabled")

    except Exception as e:
        messagebox.showerror("Error al guardar", f"Error al guardar en JSON:\n{str(e)}")


def mostrar_info_json_ticker(ticker, csv_path=None):
    """Muestra información del ticker - línea horizontal + tabla consolidada del JSON"""
    global historial_analisis_por_ticker

    # Limpiar frame de info
    for widget in frame_info_json.winfo_children():
        widget.destroy()

    # Frame horizontal para info básica en UNA LÍNEA
    frame_info_horizontal = tk.Frame(frame_info_json)
    frame_info_horizontal.pack(fill="x", pady=2)

    if csv_path and os.path.exists(csv_path):
        fecha_creacion = datetime.fromtimestamp(os.path.getctime(csv_path))
        fecha_creacion_str = fecha_creacion.strftime("%d/%m/%Y %H:%M:%S")

        # Extraer ticker real (siglas de la acción) del nombre del archivo
        nombre_archivo = os.path.splitext(os.path.basename(csv_path))[0]

        partes = nombre_archivo.split('_')
        ticker_real = ticker  # Por defecto usar el nombre completo
        if len(partes) >= 2:
            for parte in partes:
                if parte.isupper() and 1 <= len(parte) <= 5:
                    ticker_real = parte
                    break

        # LÍNEA HORIZONTAL con colores alternados (azul, negro, azul, negro, azul)
        tk.Label(frame_info_horizontal, text=f"Ticker: {ticker_real}",
                 font=("Arial", 9, "bold"), fg="darkblue").pack(side="left")
        tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")

        tk.Label(frame_info_horizontal, text=f"Archivo: {os.path.basename(csv_path)}",
                 font=("Arial", 8), fg="black").pack(side="left")
        tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")

        tk.Label(frame_info_horizontal, text=f"CSV: {fecha_creacion_str}",
                 font=("Arial", 8), fg="blue").pack(side="left")

        # Verificar si existen archivos Excel y DB
        folder = os.path.dirname(csv_path)
        base_name = nombre_archivo

        excel_path = os.path.join(folder, f"{base_name}_analizado.xlsx")
        db_path = os.path.join(folder, f"{base_name}_analizado.db")

        if os.path.exists(excel_path):
            fecha_excel = datetime.fromtimestamp(os.path.getmtime(excel_path))
            fecha_excel_str = fecha_excel.strftime("%d/%m/%Y %H:%M:%S")
            tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")
            tk.Label(frame_info_horizontal, text=f"Excel: {fecha_excel_str}",
                     font=("Arial", 8), fg="black").pack(side="left")

        if os.path.exists(db_path):
            fecha_db = datetime.fromtimestamp(os.path.getmtime(db_path))
            fecha_db_str = fecha_db.strftime("%d/%m/%Y %H:%M:%S")
            tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")
            tk.Label(frame_info_horizontal, text=f"DB: {fecha_db_str}",
                     font=("Arial", 8), fg="blue").pack(side="left")

    # NUEVO: Cargar y mostrar tabla consolidada con datos del JSON
    if ARCHIVO_JSON is None or not ARCHIVO_JSON.exists():
        return

    datos_json = cargar_resultados_json()

    if ticker not in datos_json:
        return

    info = datos_json[ticker]

    # Limpiar historial anterior de este ticker
    if ticker not in historial_analisis_por_ticker:
        historial_analisis_por_ticker[ticker] = []
    else:
        historial_analisis_por_ticker[ticker] = []

    # COMPATIBILIDAD MEJORADA: Maneja estructura antigua, nueva Y MIXTA

    # 1. Primero procesar estructura ANTIGUA si existe (dentro de "periodos")
    if "periodos" in info and isinstance(info["periodos"], dict):
        for nombre_periodo, datos_periodo in info["periodos"].items():
            if isinstance(datos_periodo,
                          dict) and "parametros_optimos" in datos_periodo and "metricas" in datos_periodo:
                params = datos_periodo["parametros_optimos"]
                metricas = datos_periodo["metricas"]

                historial_analisis_por_ticker[ticker].append({
                    "periodo": nombre_periodo.replace('_', ' ').title(),
                    "objetivo": "Rentabilidad",
                    "compra_pct": params.get('compra_pct', 0),
                    "venta_pct": params.get('venta_pct', 0),
                    "ganancia_min": params.get('ganancia_minima_pct', 0),
                    "suave_pct": params.get('suave_pct', 0),
                    "compra_mult": params.get('compra_multiple'),
                    "venta_mult": params.get('venta_multiple'),
                    "rentabilidad_max": metricas.get('rentabilidad_max', 0),
                    "margen_promedio": metricas.get('margen_promedio', 0)
                })

    # 2. Luego procesar estructura NUEVA (fuera de "periodos")
    for nombre_periodo, contenido in info.items():
        # Saltar claves de estructura antigua
        if nombre_periodo in ["ticker", "fecha_guardado", "periodos"]:
            continue

        # Procesar estructura nueva: periodo -> objetivo -> datos
        if isinstance(contenido, dict):
            for objetivo, datos in contenido.items():
                if isinstance(datos, dict) and "parametros_optimos" in datos and "metricas" in datos:
                    params = datos["parametros_optimos"]
                    metricas = datos["metricas"]

                    historial_analisis_por_ticker[ticker].append({
                        "periodo": nombre_periodo.replace('_', ' ').title(),
                        "objetivo": objetivo.replace('_', ' ').title(),
                        "compra_pct": params.get('compra_pct', 0),
                        "venta_pct": params.get('venta_pct', 0),
                        "ganancia_min": params.get('ganancia_minima_pct', 0),
                        "suave_pct": params.get('suave_pct', 0),
                        "compra_mult": params.get('compra_multiple'),
                        "venta_mult": params.get('venta_multiple'),
                        "rentabilidad_max": metricas.get('rentabilidad_max', 0),
                        "margen_promedio": metricas.get('margen_promedio', 0)
                    })

    # Ordenar por período y luego por objetivo
    orden_periodos = {"Completo": 1, "6 Meses": 2, "3 Meses": 3}
    historial_analisis_por_ticker[ticker].sort(
        key=lambda x: (orden_periodos.get(x['periodo'], 99), x['objetivo'])
    )

    # Mostrar tabla consolidada en frame_stats
    mostrar_tabla_consolidada_desde_json(ticker)


def administrar_json():
    """Abre una ventana para ver y eliminar registros del JSON"""
    _cargar_dependencias_analisis()

    if ARCHIVO_JSON is None or not ARCHIVO_JSON.exists():
        messagebox.showinfo("Sin datos", "No hay archivo JSON configurado o no existe")
        return

    datos_json = cargar_resultados_json()
    if not datos_json:
        messagebox.showinfo("Sin datos", "El archivo JSON está vacío")
        return

    # Obtener ticker actual del CSV seleccionado (para actualizar tabla después de eliminar)
    ticker_actual = None
    try:
        ruta_csv = entry_ruta.get().strip().strip('"')
        if ruta_csv and os.path.exists(ruta_csv):
            ticker_actual = os.path.splitext(os.path.basename(ruta_csv))[0]
    except:
        pass

    # Crear ventana
    ventana_admin = tk.Toplevel(ventana)
    ventana_admin.title("Administrar registros JSON")
    ventana_admin.geometry("1600x550")
    ventana_admin.transient(ventana)
    ventana_admin.grab_set()

    # Frame superior con instrucciones
    frame_instrucciones = tk.Frame(ventana_admin, pady=5)
    frame_instrucciones.pack(fill="x", padx=10)
    tk.Label(frame_instrucciones,
             text="Selecciona los registros que deseas eliminar (puedes seleccionar múltiples con Ctrl+clic)",
             font=("Arial", 9), fg="gray").pack(anchor="w")

    # Frame de filtros
    frame_filtros = tk.Frame(ventana_admin, pady=5)
    frame_filtros.pack(fill="x", padx=10)

    tk.Label(frame_filtros, text="Filtrar por Período de Datos:").pack(side="left", padx=(0, 5))

    # Extraer períodos de datos únicos del JSON
    periodos_datos_set = set()
    for ticker_key in datos_json.keys():
        partes = ticker_key.split('_')
        if len(partes) >= 4:
            periodo_datos = '_'.join(partes[2:])
            periodos_datos_set.add(periodo_datos)
        elif len(partes) >= 2:
            periodo_datos = '_'.join(partes[1:])
            periodos_datos_set.add(periodo_datos)

    periodos_datos_lista = ["Todos"] + sorted(periodos_datos_set)
    combo_periodo_datos = ttk.Combobox(frame_filtros, values=periodos_datos_lista, state="readonly", width=20)
    combo_periodo_datos.set("Todos")
    combo_periodo_datos.pack(side="left", padx=(0, 10))

    # Lista para almacenar todos los items (para filtrado)
    todos_los_items = []

    # Frame para el Treeview
    frame_tree = tk.Frame(ventana_admin)
    frame_tree.pack(fill="both", expand=True, padx=10, pady=5)

    # Scrollbars
    scrollbar_y = tk.Scrollbar(frame_tree, orient="vertical")
    scrollbar_x = tk.Scrollbar(frame_tree, orient="horizontal")

    # Treeview con selección múltiple
    # Columnas completas incluyendo todas las estadísticas
    # ORDEN: Básicos, Parámetros, Métricas (incluyendo Prom.Max/Min%), Estadísticas var, Operaciones, Fecha
    columns = (
        "Symbol", "Datos", "Período", "Objetivo",
        # Parámetros óptimos
        "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N", "Límite", "Valor Lím.",
        # Métricas (Prom.Max% y Prom.Min% movidos aquí, después de Margen.Prom)
        "Rentab.Máx", "Margen.Prom", "Prom.Max%", "Prom.Min%", "Rentab.Prom", "Max.Margen", "Max.Aporte",
        # Estadísticas % variación
        "Max.Var%", "Min.Var%", "Fecha Max.Var", "Fecha Min.Var", "Dif.Var%", "Prom.Subida%", "Prom.Bajada%", "Dif.Prom%",
        # Estadísticas operaciones
        "Opc.Compra", "Acc.Compradas", "Opc.Venta", "Acc.Vendidas", "Max.Acc.Cart",
        # Fecha guardado
        "Fecha"
    )
    tree = ttk.Treeview(frame_tree, columns=columns, show="headings",
                        selectmode="extended",
                        yscrollcommand=scrollbar_y.set,
                        xscrollcommand=scrollbar_x.set)

    scrollbar_y.config(command=tree.yview)
    scrollbar_x.config(command=tree.xview)

    # Configurar columnas con anchos basados en el título (caracteres * 8 + margen)
    for col in columns:
        tree.heading(col, text=col)
        # Ancho basado en longitud del título
        ancho = max(len(col) * 8 + 10, 50)  # Mínimo 50px
        # Columnas de fecha más anchas
        if "Fecha" in col and col != "Fecha":
            ancho = max(ancho, 90)
        elif col == "Fecha":
            ancho = 130
        tree.column(col, width=ancho, anchor="center")

    # Diccionario para mapear items del tree a rutas en el JSON
    item_to_path = {}

    # Llenar el Treeview con datos (ordenados alfabéticamente por ticker_symbol)
    # Primero extraer y ordenar los tickers
    tickers_ordenados = sorted(datos_json.items(),
                                key=lambda x: (x[1].get("_ticker_symbol") or extraer_ticker_symbol(x[0]) or x[0]).upper())

    for ticker, contenido_ticker in tickers_ordenados:
        # Obtener ticker_symbol: desde el JSON o extraerlo del nombre
        ticker_symbol = contenido_ticker.get("_ticker_symbol") or extraer_ticker_symbol(ticker) or ticker

        # Extraer período de datos del nombre del ticker
        partes_ticker = ticker.split('_')
        if len(partes_ticker) >= 4:
            periodo_datos_valor = '_'.join(partes_ticker[2:])
        elif len(partes_ticker) >= 2:
            periodo_datos_valor = '_'.join(partes_ticker[1:])
        else:
            periodo_datos_valor = "-"

        # Manejar estructura antigua (con "periodos")
        if "periodos" in contenido_ticker and isinstance(contenido_ticker["periodos"], dict):
            for periodo, datos_periodo in contenido_ticker["periodos"].items():
                if isinstance(datos_periodo, dict) and "parametros_optimos" in datos_periodo:
                    params = datos_periodo.get("parametros_optimos", {})
                    metricas = datos_periodo.get("metricas", {})
                    stats_var = datos_periodo.get("estadisticas_var", {})
                    stats_ops = datos_periodo.get("estadisticas_operaciones", {})
                    fecha = datos_periodo.get("fecha_guardado", "")
                    compra_mult = params.get("compra_multiple")
                    venta_mult = params.get("venta_multiple")

                    limite_tipo = params.get("limite_tipo", "acciones")
                    limite_valor = params.get("limite_valor", 10.0)

                    valores_fila = (
                        ticker_symbol,
                        periodo_datos_valor,
                        periodo.replace("_", " ").title(),
                        "Rentabilidad",
                        f"{params.get('compra_pct', 0):.2f}",
                        f"{params.get('venta_pct', 0):.2f}",
                        f"{params.get('ganancia_minima_pct', 0):.2f}",
                        compra_mult if compra_mult else "-",
                        venta_mult if venta_mult else "-",
                        limite_tipo.title(),
                        f"{limite_valor:.0f}" if limite_tipo == "acciones" else f"${limite_valor:.0f}",
                        f"{metricas.get('rentabilidad_max', 0):.2f}%",
                        f"{metricas.get('margen_promedio', 0):.2f}",
                        f"{params.get('promedio_maximos', 0):.2f}%",
                        f"{params.get('promedio_minimos', 0):.2f}%",
                        f"{metricas.get('rentab_promedio', 0):.2f}%",
                        f"{metricas.get('max_margen', 0):.2f}",
                        f"{metricas.get('max_aporte', 0):.0f}",
                        f"{stats_var.get('max_var', 0):.2f}%",
                        f"{stats_var.get('min_var', 0):.2f}%",
                        stats_var.get('fecha_max_var', '-'),
                        stats_var.get('fecha_min_var', '-'),
                        f"{stats_var.get('dif_var', 0):.2f}%",
                        f"{stats_var.get('max_prom_var', 0):.2f}%",
                        f"{stats_var.get('min_prom_var', 0):.2f}%",
                        f"{stats_var.get('dif_prom_var', 0):.2f}%",
                        stats_ops.get('opc_compra', 0),
                        stats_ops.get('acciones_compradas', 0),
                        stats_ops.get('opc_venta', 0),
                        stats_ops.get('acciones_vendidas', 0),
                        stats_ops.get('max_acc_cartera', 0),
                        fecha
                    )
                    path_info = {"ticker": ticker, "path": ["periodos", periodo], "tipo": "antiguo"}
                    todos_los_items.append({"valores": valores_fila, "path_info": path_info})
                    item_id = tree.insert("", "end", values=valores_fila)
                    item_to_path[item_id] = path_info

        # Manejar estructura nueva (período -> objetivo -> datos)
        for periodo, contenido_periodo in contenido_ticker.items():
            if periodo in ["ticker", "fecha_guardado", "periodos", "_ticker_symbol"]:
                continue

            if isinstance(contenido_periodo, dict):
                for objetivo, datos in contenido_periodo.items():
                    if isinstance(datos, dict) and "parametros_optimos" in datos:
                        params = datos.get("parametros_optimos", {})
                        metricas = datos.get("metricas", {})
                        stats_var = datos.get("estadisticas_var", {})
                        stats_ops = datos.get("estadisticas_operaciones", {})
                        fecha = datos.get("fecha_guardado", "")
                        # También intentar obtener ticker_symbol del registro individual
                        symbol_mostrar = datos.get("ticker_symbol") or ticker_symbol
                        compra_mult = params.get("compra_multiple")
                        venta_mult = params.get("venta_multiple")
                        limite_tipo = params.get("limite_tipo", "acciones")
                        limite_valor = params.get("limite_valor", 10.0)

                        valores_fila = (
                            symbol_mostrar,
                            periodo_datos_valor,
                            periodo.replace("_", " ").title(),
                            objetivo.replace("_", " ").title(),
                            # Parámetros óptimos
                            f"{params.get('compra_pct', 0):.2f}",
                            f"{params.get('venta_pct', 0):.2f}",
                            f"{params.get('ganancia_minima_pct', 0):.2f}",
                            compra_mult if compra_mult else "-",
                            venta_mult if venta_mult else "-",
                            limite_tipo.title(),
                            f"{limite_valor:.0f}" if limite_tipo == "acciones" else f"${limite_valor:.0f}",
                            # Métricas (Prom.Max% y Prom.Min% después de Margen.Prom)
                            f"{metricas.get('rentabilidad_max', 0):.2f}%",
                            f"{metricas.get('margen_promedio', 0):.2f}",
                            f"{params.get('promedio_maximos', 0):.2f}%",
                            f"{params.get('promedio_minimos', 0):.2f}%",
                            f"{metricas.get('rentab_promedio', 0):.2f}%",
                            f"{metricas.get('max_margen', 0):.2f}",
                            f"{metricas.get('max_aporte', 0):.0f}",
                            # Estadísticas % variación (con símbolos %)
                            f"{stats_var.get('max_var', 0):.2f}%",
                            f"{stats_var.get('min_var', 0):.2f}%",
                            stats_var.get('fecha_max_var', '-'),
                            stats_var.get('fecha_min_var', '-'),
                            f"{stats_var.get('dif_var', 0):.2f}%",
                            f"{stats_var.get('max_prom_var', 0):.2f}%",
                            f"{stats_var.get('min_prom_var', 0):.2f}%",
                            f"{stats_var.get('dif_prom_var', 0):.2f}%",
                            # Estadísticas operaciones
                            stats_ops.get('opc_compra', 0),
                            stats_ops.get('acciones_compradas', 0),
                            stats_ops.get('opc_venta', 0),
                            stats_ops.get('acciones_vendidas', 0),
                            stats_ops.get('max_acc_cartera', 0),
                            # Fecha
                            fecha
                        )
                        path_info = {"ticker": ticker, "path": [periodo, objetivo], "tipo": "nuevo"}
                        todos_los_items.append({"valores": valores_fila, "path_info": path_info})
                        item_id = tree.insert("", "end", values=valores_fila)
                        item_to_path[item_id] = path_info

    # Empaquetar Treeview y scrollbars
    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    # Frame inferior con botones
    frame_botones = tk.Frame(ventana_admin, pady=10)
    frame_botones.pack(fill="x", padx=10)

    # Label para mostrar cantidad seleccionada
    label_seleccion = tk.Label(frame_botones, text="0 registros seleccionados", font=("Arial", 9))
    label_seleccion.pack(side="left")

    def actualizar_contador(event=None):
        cantidad = len(tree.selection())
        label_seleccion.config(text=f"{cantidad} registro(s) seleccionado(s)")

    tree.bind("<<TreeviewSelect>>", actualizar_contador)

    def eliminar_seleccionados():
        seleccionados = tree.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "No has seleccionado ningún registro")
            return

        # Confirmar eliminación
        cantidad = len(seleccionados)
        if not messagebox.askyesno("Confirmar eliminación",
                                    f"¿Estás seguro de eliminar {cantidad} registro(s)?\n\nEsta acción no se puede deshacer."):
            return

        # Cargar JSON actual
        with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Eliminar cada registro seleccionado
        eliminados = 0
        for item_id in seleccionados:
            if item_id in item_to_path:
                info = item_to_path[item_id]
                ticker = info["ticker"]
                path = info["path"]
                tipo = info["tipo"]

                try:
                    if tipo == "antiguo":
                        # Estructura: ticker -> periodos -> periodo
                        if ticker in datos and "periodos" in datos[ticker]:
                            if path[1] in datos[ticker]["periodos"]:
                                del datos[ticker]["periodos"][path[1]]
                                eliminados += 1
                                # Si periodos queda vacío, eliminarlo
                                if not datos[ticker]["periodos"]:
                                    del datos[ticker]["periodos"]
                    else:
                        # Estructura nueva: ticker -> periodo -> objetivo
                        if ticker in datos and path[0] in datos[ticker]:
                            if path[1] in datos[ticker][path[0]]:
                                del datos[ticker][path[0]][path[1]]
                                eliminados += 1
                                # Si periodo queda vacío, eliminarlo
                                if not datos[ticker][path[0]]:
                                    del datos[ticker][path[0]]

                    # Si ticker queda vacío (solo con claves vacías), eliminarlo
                    if ticker in datos:
                        claves_restantes = [k for k in datos[ticker].keys()
                                           if k not in ["ticker", "fecha_guardado"] and datos[ticker][k]]
                        if not claves_restantes:
                            del datos[ticker]

                except Exception as e:
                    print(f"Error eliminando {item_id}: {e}")

        # Guardar JSON actualizado
        with open(ARCHIVO_JSON, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)

        messagebox.showinfo("Eliminación completada", f"Se eliminaron {eliminados} registro(s)")

        # Actualizar el cuadro de parámetros en la interfaz principal
        if ticker_actual:
            # Recargar el historial del ticker y actualizar la tabla
            ruta_csv = entry_ruta.get().strip().strip('"')
            mostrar_info_json_ticker(ticker_actual, ruta_csv)

        # Cerrar y reabrir para refrescar
        ventana_admin.destroy()
        administrar_json()

    def seleccionar_todos():
        for item in tree.get_children():
            tree.selection_add(item)
        actualizar_contador()

    def deseleccionar_todos():
        tree.selection_remove(tree.get_children())
        actualizar_contador()

    def filtrar_por_periodo(event=None):
        """Filtra los items del treeview por período de datos"""
        periodo_seleccionado = combo_periodo_datos.get()

        # Limpiar el treeview
        for item in tree.get_children():
            tree.delete(item)

        # Limpiar el mapeo de paths
        item_to_path.clear()

        # Re-insertar items según el filtro
        for item_data in todos_los_items:
            valores = item_data["valores"]
            path_info = item_data["path_info"]

            # La columna "Datos" está en la posición 1 (índice 1)
            periodo_datos_item = valores[1] if len(valores) > 1 else ""

            if periodo_seleccionado == "Todos" or periodo_datos_item == periodo_seleccionado:
                item_id = tree.insert("", "end", values=valores)
                item_to_path[item_id] = path_info

        actualizar_contador()

    # Vincular el combobox con la función de filtro
    combo_periodo_datos.bind("<<ComboboxSelected>>", filtrar_por_periodo)

    def exportar_a_excel():
        """Exporta los datos del JSON a un archivo Excel"""
        # Obtener todos los items del treeview
        items = tree.get_children()
        if not items:
            messagebox.showwarning("Sin datos", "No hay datos para exportar")
            return

        # Preguntar dónde guardar
        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfile="Parametros_Optimos.xlsx"
        )

        if not ruta_excel:
            return

        try:
            # Crear DataFrame con los datos (todas las columnas en nuevo orden)
            datos_export = []
            for item in items:
                v = tree.item(item, "values")
                datos_export.append({
                    "Symbol": v[0],
                    "Datos": v[1],
                    "Período": v[2],
                    "Objetivo": v[3],
                    # Parámetros óptimos
                    "Compra%": v[4],
                    "Venta%": v[5],
                    "Gan.Mín%": v[6],
                    "Compra N": v[7] if v[7] != "-" else "",
                    "Venta N": v[8] if v[8] != "-" else "",
                    "Límite": v[9],
                    "Valor Lím.": v[10],
                    # Métricas (Prom.Max% y Prom.Min% después de Margen.Prom)
                    "Rentab.Máx": v[11],
                    "Margen.Prom": v[12],
                    "Prom.Max%": v[13],
                    "Prom.Min%": v[14],
                    "Rentab.Prom": v[15],
                    "Max.Margen": v[16],
                    "Max.Aporte": v[17],
                    # Estadísticas % variación
                    "Max.Var%": v[18],
                    "Min.Var%": v[19],
                    "Fecha Max.Var": v[20],
                    "Fecha Min.Var": v[21],
                    "Dif.Var%": v[22],
                    "Prom.Subida%": v[23],
                    "Prom.Bajada%": v[24],
                    "Dif.Prom%": v[25],
                    # Estadísticas operaciones
                    "Opc.Compra": v[26],
                    "Acc.Compradas": v[27],
                    "Opc.Venta": v[28],
                    "Acc.Vendidas": v[29],
                    "Max.Acc.Cart": v[30],
                    # Fecha
                    "Fecha Guardado": v[31]
                })

            df_export = pd.DataFrame(datos_export)

            # Exportar a Excel con formato
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Parámetros Óptimos"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Escribir encabezados
            columnas = list(df_export.columns)
            for col_idx, col_name in enumerate(columnas, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Escribir datos
            for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            # Ajustar anchos de columna automáticamente
            from openpyxl.utils import get_column_letter
            for col_idx, col_name in enumerate(columnas, 1):
                col_letter = get_column_letter(col_idx)
                # Ancho basado en nombre de columna + margen
                ancho = max(len(str(col_name)) + 2, 10)
                # Columnas de fecha más anchas
                if "Fecha" in col_name:
                    ancho = max(ancho, 14)
                ws.column_dimensions[col_letter].width = ancho

            wb.save(ruta_excel)
            messagebox.showinfo("Exportación exitosa",
                               f"Datos exportados correctamente.\n\nArchivo: {ruta_excel}\nRegistros: {len(datos_export)}")

        except PermissionError:
            messagebox.showerror("Error", "El archivo está abierto. Ciérralo e intenta de nuevo.")
        except Exception as e:
            messagebox.showerror("Error al exportar", f"Error: {str(e)}")

    # Botones
    tk.Button(frame_botones, text="Seleccionar todos", command=seleccionar_todos).pack(side="left", padx=(20, 5))
    tk.Button(frame_botones, text="Deseleccionar todos", command=deseleccionar_todos).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Exportar a Excel", command=exportar_a_excel,
              bg="#28a745", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=10)

    tk.Button(frame_botones, text="Eliminar seleccionados", command=eliminar_seleccionados,
              bg="#ff6b6b", fg="white", font=("Arial", 10, "bold")).pack(side="right", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_admin.destroy).pack(side="right", padx=5)


def mostrar_tabla_consolidada_desde_json(ticker):
    """Muestra SOLO la tabla consolidada con datos del JSON (sin pestañas)"""
    # Limpiar frame de estadísticas
    for widget in ventana.frame_stats.winfo_children():
        widget.destroy()

    # Extraer ticker real (siglas) para mostrar
    if ticker:
        partes = ticker.split('_')
        ticker_display = ticker
        if len(partes) >= 2:
            for parte in partes:
                if parte.isupper() and 1 <= len(parte) <= 5:
                    ticker_display = parte
                    break
    else:
        ticker_display = "Actual"

    # Frame con tabla consolidada
    frame_consolidado = tk.Frame(ventana.frame_stats, relief="ridge", borderwidth=2, bg="lightyellow", padx=10, pady=10)
    frame_consolidado.pack(fill="x", pady=(10, 0))

    tk.Label(frame_consolidado, text=f"📊 PARÁMETROS ÓPTIMOS GUARDADOS - {ticker_display}",
             font=("Arial", 11, "bold"), bg="lightyellow", fg="darkgreen").pack(anchor="w")

    # Crear tabla con parámetros
    frame_tabla_params = tk.Frame(frame_consolidado, bg="lightyellow")
    frame_tabla_params.pack(fill="x", pady=(5, 0))

    # Headers
    headers = ["#", "Período", "Objetivo", "Compra %", "Venta %", "Gan Mín %", "Suave %", "Comp", "Venta", "Rentab Máx",
               "Margen Prom"]
    for col, header in enumerate(headers):
        ancho = 5 if col == 0 else 11
        tk.Label(frame_tabla_params, text=header, font=("Arial", 8, "bold"),
                 bg="lightblue", relief="solid", borderwidth=1, width=ancho).grid(row=0, column=col, sticky="ew",
                                                                                  padx=1, pady=1)

    # Datos del ticker actual
    analisis_ticker_actual = historial_analisis_por_ticker.get(ticker, [])

    if not analisis_ticker_actual:
        tk.Label(frame_consolidado, text="No hay análisis guardados en JSON",
                 font=("Arial", 9), bg="lightyellow", fg="gray").pack(pady=10)
        return

    # Colores por objetivo (base del objetivo, sin el número)
    colores_objetivo = {
        "rentabilidad": "#e8f5e9",  # Verde claro
        "margen": "#e3f2fd",         # Azul claro
    }
    color_default = "#fff3e0"  # Naranja claro para otros

    periodo_anterior = None
    fila_actual = 0

    for idx, analisis in enumerate(analisis_ticker_actual, start=1):
        fila_actual += 1

        # Determinar color basado en el objetivo (sin número)
        objetivo_base = analisis['objetivo'].lower().split()[0]
        bg_color = colores_objetivo.get(objetivo_base, color_default)

        # Si cambia el período, agregar línea separadora
        if periodo_anterior is not None and analisis['periodo'] != periodo_anterior:
            # Agregar fila separadora
            for col in range(11):
                ancho = 5 if col == 0 else 11
                tk.Label(frame_tabla_params, text="", font=("Arial", 2),
                         bg="#999999", relief="flat", width=ancho, height=1).grid(
                    row=fila_actual, column=col, sticky="ew", padx=1, pady=0)
            fila_actual += 1

        periodo_anterior = analisis['periodo']

        valores = [
            str(idx),
            analisis['periodo'],
            analisis['objetivo'],
            f"{analisis['compra_pct']:.1f}",
            f"{analisis['venta_pct']:.1f}",
            f"{analisis['ganancia_min']:.1f}",
            f"{analisis['suave_pct']:.1f}",
            str(analisis['compra_mult']) if analisis['compra_mult'] else "-",
            str(analisis['venta_mult']) if analisis['venta_mult'] else "-",
            f"{analisis['rentabilidad_max']:.2f}%",
            f"{analisis['margen_promedio']:.2f}"
        ]

        for col, valor in enumerate(valores):
            ancho = 5 if col == 0 else 11
            tk.Label(frame_tabla_params, text=valor, font=("Arial", 7),
                     bg=bg_color, relief="solid", borderwidth=1, width=ancho).grid(
                row=fila_actual, column=col, sticky="ew", padx=1, pady=1)


# =========================
# Funciones auxiliares
# =========================
def parse_percent_to_decimal(x):
    """Convierte valores de porcentaje a decimal"""
    try:
        if pd.isna(x):
            return float("nan")
        s = str(x).strip().replace(",", ".")
        if s.endswith("%"):
            s = s[:-1].strip()
            if s == "":
                return float("nan")
            return float(s) / 100.0
        try:
            f = float(s)
            if abs(f) <= 1:
                return f
            else:
                return f / 100.0
        except:
            return float("nan")
    except Exception:
        return float("nan")


def to_float_safe(x):
    """Convierte cadenas numéricas con coma/punto a float, devuelve NaN si falla."""
    try:
        if pd.isna(x):
            return float("nan")
        s = str(x).strip().replace('"', '').replace(",", ".")
        if s == "":
            return float("nan")
        return float(s)
    except:
        return float("nan")


def create_sqlite_from_df(folder, name, df):
    """Crea una base sqlite con la tabla 'precios' a partir del DataFrame."""
    db = os.path.join(folder, name)
    conn = sqlite3.connect(db)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS precios")
    cur.execute("""
        CREATE TABLE precios (
            Fecha TEXT,
            Ultimo REAL,
            Apertura REAL,
            Maximo REAL,
            Minimo REAL,
            Vol REAL,
            Var REAL
        )
    """)
    rows = []
    for _, r in df.iterrows():
        fecha = r.get("Fecha", "")
        try:
            ultimo = float(r["Último"]) if pd.notna(r["Último"]) and r["Último"] != "" else None
        except:
            ultimo = to_float_safe(r["Último"])
        try:
            apertura = float(r["Apertura"]) if pd.notna(r["Apertura"]) and r["Apertura"] != "" else None
        except:
            apertura = to_float_safe(r["Apertura"])
        try:
            maximo = float(r["Máximo"]) if pd.notna(r["Máximo"]) and r["Máximo"] != "" else None
        except:
            maximo = to_float_safe(r["Máximo"])
        try:
            minimo = float(r["Mínimo"]) if pd.notna(r["Mínimo"]) and r["Mínimo"] != "" else None
        except:
            minimo = to_float_safe(r["Mínimo"])
        try:
            vol = float(r["Vol."]) if pd.notna(r["Vol."]) and r["Vol."] != "" else None
        except:
            vol = to_float_safe(r["Vol."])

        var_val = r.get("% var.", None)
        if pd.isna(var_val) or var_val is None or str(var_val).strip() == "":
            var_num = None
        else:
            var_num = parse_percent_to_decimal(var_val)
            if pd.isna(var_num):
                var_num = None

        rows.append((fecha, ultimo, apertura, maximo, minimo, vol, var_num))

    cur.executemany("INSERT INTO precios VALUES (?,?,?,?,?,?,?)", rows)
    conn.commit()
    conn.close()
    return db


def filtrar_ultimos_dias(csv_path, dias):
    """Lee el CSV y devuelve un DataFrame con solo los últimos N días"""
    global error_analisis_mostrado

    try:
        # Intentar primero con utf-8-sig para manejar BOM, luego latin-1
        try:
            df = pd.read_csv(csv_path, sep=";", engine='python', dtype=str, encoding='utf-8-sig')
        except:
            df = pd.read_csv(csv_path, sep=";", engine='python', dtype=str, encoding='latin-1')
    except Exception as e:
        error_analisis_mostrado = True
        messagebox.showerror("Error al leer CSV", str(e) + "\n\nEl análisis se detendrá.")
        return None

    df = normalizar_columnas(df)  # Convertir columnas inglés a español si es necesario

    # Procesar fechas
    def parse_mixed_dates(date_str):
        for fmt in ("%d/%m/%Y", "%m/%d/%Y"):
            try:
                return pd.to_datetime(date_str, format=fmt)
            except:
                continue
        return pd.NaT

    df['Fecha'] = df['Fecha'].apply(parse_mixed_dates)
    df = df.dropna(subset=['Fecha'])
    df = df.sort_values('Fecha').reset_index(drop=True)

    # Obtener fecha más reciente y calcular fecha de corte
    fecha_max = df['Fecha'].max()
    fecha_corte = fecha_max - timedelta(days=dias)

    # Filtrar
    df_filtrado = df[df['Fecha'] >= fecha_corte].copy()

    print(f"  -> Periodo: {fecha_corte.strftime('%d/%m/%Y')} a {fecha_max.strftime('%d/%m/%Y')}")
    print(f"  -> Registros: {len(df_filtrado)} de {len(df)}")

    return df_filtrado


# =========================
# Funcion auxiliar para generar grafico en figura
# =========================
def generar_grafico_en_figura(fig, df, titulo):
    """Genera el grafico con 4 ejes en la figura dada.
    Retorna los ejes para referencia.

    CONFIGURACION DE GRAFICO (documentada 19-Mar-2026):
    ===================================================
    Posicion del grafico: pos = [0.05, 0.12, 0.82, 0.78]  # [left, bottom, width, height]

    Ejes Y (lado derecho):
    - ax_margen: outward=0 (pegado al grafico)
    - ax2 (Rentabilidad): outward=45
    - ax3 (Acciones): outward=85

    Escalado de ejes:
    - Margen: ymin - rango*0.3, ymax + rango*0.1 (curva arriba)
    - Rentabilidad: ymin - rango*0.1, ymax + rango*0.5 (curva abajo, separada de margen)
    - Precio: 0 a ymax*1.05
    - Acciones: -rango*0.5 a ymax*1.1

    Fechas eje X (intervalo dinamico):
    - <=30 dias: cada 2 dias
    - <=90 dias: cada 5 dias
    - <=180 dias: cada 10 dias
    - >180 dias: cada 15 dias
    """
    fig.clear()

    # Normalizar nombres de columnas (manejar tildes)
    renombres = {
        'Último': 'Ultimo',
        'Máximo': 'Maximo',
        'Mínimo': 'Minimo'
    }
    for col_tilde, col_sin in renombres.items():
        if col_tilde in df.columns and col_sin not in df.columns:
            df = df.rename(columns={col_tilde: col_sin})

    # Convertir Fecha a datetime si es string
    if df['Fecha'].dtype == 'object':
        df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')

    # Convertir Rentabilidad a numero (quitar % si existe)
    if 'Rentabilidad_num' not in df.columns:
        if df['Rentabilidad'].dtype == 'object':
            df['Rentabilidad_num'] = pd.to_numeric(df['Rentabilidad'].astype(str).str.rstrip('%').str.replace(',', '.'), errors='coerce').fillna(0)
        else:
            df['Rentabilidad_num'] = df['Rentabilidad']
    else:
        # Si ya existe pero es string, convertir a numérico
        if df['Rentabilidad_num'].dtype == 'object':
            df['Rentabilidad_num'] = pd.to_numeric(df['Rentabilidad_num'].astype(str).str.rstrip('%').str.replace(',', '.'), errors='coerce').fillna(0)

    # Crear eje principal
    ax1 = fig.add_subplot(111)
    fig.subplots_adjust(left=0.04, right=0.9, top=0.92, bottom=0.15)

    # Eje Y independiente para Margen (color verde)
    ax_margen = ax1.twinx()
    ax_margen.spines['right'].set_position(('outward', 0))

    # Eje Y para Rentabilidad (color rojo)
    ax2 = ax1.twinx()
    ax2.spines['right'].set_position(('outward', 45))

    # Eje Y para Acciones en cartera (color negro)
    ax3 = ax1.twinx()
    ax3.spines['right'].set_position(('outward', 85))

    # Ajustar la posicion de los ejes para que todo quepa
    pos = [0.05, 0.12, 0.82, 0.78]
    ax1.set_position(pos)
    ax_margen.set_position(pos)
    ax2.set_position(pos)
    ax3.set_position(pos)

    # Graficar series
    ax1.plot(df['Fecha'], df['Ultimo'], color='blue', label='Ultimo', linewidth=2)
    ax_margen.plot(df['Fecha'], df['Margen'], color='green', label='Margen', linewidth=2)

    # Forzar conversión a numérico antes de usar (limpiar % y comas)
    if df['Rentabilidad_num'].dtype == 'object' or not pd.api.types.is_numeric_dtype(df['Rentabilidad_num']):
        df['Rentabilidad_num'] = pd.to_numeric(
            df['Rentabilidad_num'].astype(str).str.replace('%', '', regex=False).str.replace(',', '.', regex=False).str.strip(),
            errors='coerce'
        ).fillna(0)
    # Graficar Rentabilidad sin escalar (usa su propio eje Y)
    ax2.plot(df['Fecha'], df['Rentabilidad_num'], color='red',
             label='Rentabilidad (%)', linestyle='--', linewidth=2)

    ax3.plot(df['Fecha'], df['Acciones en cartera'], color='black',
             label='Acciones en cartera', linestyle=':', linewidth=2)

    # Etiquetas y colores para cada eje
    ax1.set_ylabel('Ultimo', color='blue')
    ax1.tick_params(axis='y', labelcolor='blue')

    ax_margen.set_ylabel('Margen', color='green')
    ax_margen.tick_params(axis='y', labelcolor='green')

    ax2.set_ylabel('Rentabilidad (%)', color='red')
    ax2.tick_params(axis='y', labelcolor='red')

    ax3.set_ylabel('Acciones en cartera', color='black')
    ax3.tick_params(axis='y', labelcolor='black')

    # Formatear valores sin decimales
    ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    ax_margen.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    ax3.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))

    # Autoscale con factor y offset para evitar traslapos
    # Margen: expandir hacia abajo para que la curva quede más arriba
    ax_margen.relim()
    ax_margen.autoscale_view()
    ymin_m, ymax_m = ax_margen.get_ylim()
    rango_m = ymax_m - ymin_m
    ax_margen.set_ylim(ymin_m - rango_m * 0.3, ymax_m + rango_m * 0.1)

    # Rentabilidad: expandir hacia arriba para que la curva quede más abajo (no solapar con margen)
    ax2.relim()
    ax2.autoscale_view()
    ymin_r, ymax_r = ax2.get_ylim()
    rango_r = ymax_r - ymin_r
    ax2.set_ylim(ymin_r - rango_r * 0.1, ymax_r + rango_r * 0.5)

    # Precio: inicia en cero
    ax1.relim()
    ax1.autoscale_view()
    _, ymax_p = ax1.get_ylim()
    ax1.set_ylim(0, ymax_p * 1.05)

    # Acciones en cartera: ajustar para que inicie mas arriba
    ax3.relim()
    ax3.autoscale_view()
    ymin_acc, ymax_acc = ax3.get_ylim()
    rango_acc = ymax_acc - ymin_acc
    ax3.set_ylim(-rango_acc * 0.5, ymax_acc * 1.1)

    # Leyendas combinadas
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax_margen.get_legend_handles_labels()
    lines3, labels3 = ax2.get_legend_handles_labels()
    lines4, labels4 = ax3.get_legend_handles_labels()
    ax1.legend(lines1 + lines2 + lines3 + lines4, labels1 + labels2 + labels3 + labels4, loc='upper left')

    # Formatear fechas en eje X para evitar traslape
    # Calcular intervalo dinámico según cantidad de días
    num_dias = len(df)
    if num_dias <= 30:
        intervalo = 2
    elif num_dias <= 90:
        intervalo = 5
    elif num_dias <= 180:
        intervalo = 10
    else:
        intervalo = 15
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m'))
    ax1.xaxis.set_major_locator(DayLocator(interval=intervalo))
    plt.setp(ax1.xaxis.get_majorticklabels(), fontsize=8, rotation=45, ha='right')

    # Titulo
    fig.suptitle(f'Analisis: Precio, Margen, Rentabilidad y Acciones - {titulo}', fontsize=12)

    return ax1, ax_margen, ax2, ax3


# =========================
# Funcion para mostrar grafico de resultados
# =========================
def mostrar_grafico_resultados():
    """Muestra grafico con Precio, Margen, Rentabilidad y Acciones en cartera.
    Si no hay analisis en memoria, permite cargar desde un Excel existente."""
    global resultados_dfs_por_periodo

    _cargar_dependencias_analisis()
    _cargar_dependencias_grafico()

    df = None
    clave_seleccionada = "Desde Excel"
    archivo_excel_auto = None

    # Intentar obtener el archivo Excel correspondiente al CSV seleccionado
    ruta_csv = entry_ruta.get().strip().strip('"')
    if ruta_csv and os.path.exists(ruta_csv):
        # Derivar la ruta del Excel desde el CSV
        # Ejemplo: DATA/AAPL/Datos_AAPL_FEB25_FEB26.csv -> Resultado_Analisis/AAPL/Analisis_AAPL_FEB25_FEB26.xlsx
        nombre_csv = os.path.basename(ruta_csv)  # Datos_AAPL_FEB25_FEB26.csv
        if nombre_csv.startswith("Datos_") and nombre_csv.endswith(".csv"):
            nombre_excel = nombre_csv.replace("Datos_", "Analisis_").replace(".csv", ".xlsx")
            carpeta_ticker = os.path.dirname(ruta_csv)  # DATA/AAPL
            nombre_ticker = os.path.basename(carpeta_ticker)  # AAPL
            carpeta_resultado = os.path.join(os.path.dirname(carpeta_ticker), "Resultado_Analisis", nombre_ticker)
            archivo_excel_auto = os.path.join(carpeta_resultado, nombre_excel)

            if os.path.exists(archivo_excel_auto):
                # Usar el archivo Excel correspondiente automáticamente
                pass
            else:
                archivo_excel_auto = None

    if resultados_dfs_por_periodo:
        # Hay datos en memoria - preguntar si usar memoria o cargar Excel
        claves = list(resultados_dfs_por_periodo.keys())

        # Crear ventana de seleccion
        ventana_sel = tk.Toplevel(ventana)
        ventana_sel.title("Seleccionar Origen de Datos")
        ventana_sel.geometry("350x200")
        ventana_sel.transient(ventana)
        ventana_sel.grab_set()

        tk.Label(ventana_sel, text="Selecciona el origen de datos:",
                 font=("Arial", 10, "bold")).pack(pady=10)

        # Agregar opcion de cargar desde Excel
        opciones = claves + ["-- Cargar desde Excel --"]
        combo_periodo = ttk.Combobox(ventana_sel, values=opciones, state="readonly", width=35)
        combo_periodo.pack(pady=5)
        combo_periodo.current(0)

        resultado_sel = [None]

        def confirmar():
            resultado_sel[0] = combo_periodo.get()
            ventana_sel.destroy()

        tk.Button(ventana_sel, text="Continuar", command=confirmar,
                  bg="#28a745", fg="white", width=15).pack(pady=15)

        ventana_sel.wait_window()

        if resultado_sel[0] is None:
            return

        if resultado_sel[0] == "-- Cargar desde Excel --":
            # Cargar desde archivo Excel
            df = None
        else:
            clave_seleccionada = resultado_sel[0]
            df = resultados_dfs_por_periodo[clave_seleccionada].copy()

    # Si no hay df, cargar desde Excel
    if df is None:
        # Si hay archivo Excel auto-detectado, usarlo directamente
        if archivo_excel_auto:
            archivo = archivo_excel_auto
        else:
            from tkinter import filedialog
            archivo = filedialog.askopenfilename(
                title="Seleccionar archivo Excel con resultados",
                filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
                initialdir=cargar_ultima_carpeta_grafico()
            )
            if not archivo:
                return

            # Guardar la carpeta del archivo seleccionado
            guardar_ultima_carpeta_grafico(Path(archivo).parent)

        try:
            # Leer el Excel - cargar TODAS las hojas
            xl = pd.ExcelFile(archivo)
            hojas = xl.sheet_names

            # Cargar todas las hojas en un diccionario
            todos_dfs = {}
            for hoja in hojas:
                df_hoja = pd.read_excel(archivo, sheet_name=hoja)
                # Normalizar nombres de columnas
                renombres = {'Último': 'Ultimo', 'Máximo': 'Maximo', 'Mínimo': 'Minimo'}
                for col_tilde, col_sin in renombres.items():
                    if col_tilde in df_hoja.columns and col_sin not in df_hoja.columns:
                        df_hoja = df_hoja.rename(columns={col_tilde: col_sin})
                todos_dfs[hoja] = df_hoja

            # Verificar columnas requeridas en al menos una hoja
            columnas_requeridas = ['Fecha', 'Ultimo', 'Margen', 'Rentabilidad', 'Acciones en cartera']
            primera_hoja = list(todos_dfs.keys())[0]
            columnas_faltantes = [c for c in columnas_requeridas if c not in todos_dfs[primera_hoja].columns]
            if columnas_faltantes:
                messagebox.showerror("Error", f"El archivo no tiene las columnas requeridas:\n{', '.join(columnas_faltantes)}")
                return

            clave_seleccionada = primera_hoja

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{str(e)}")
            return
    else:
        # Solo hay un df desde memoria
        todos_dfs = {clave_seleccionada: df}

    # Crear ventana del grafico
    ventana_grafico = tk.Toplevel(ventana)
    ventana_grafico.title("Grafico de Resultados")
    ventana_grafico.geometry("1500x950")

    # Extraer información para el título
    titulo_grafico = ""
    try:
        # Obtener el ticker - prioridad: archivo cargado, archivo_excel_auto, CSV global
        ticker_titulo = ""

        # 1. Intentar desde archivo cargado
        try:
            if archivo:
                partes = os.path.basename(archivo).replace(".xlsx", "").split("_")
                if len(partes) >= 2:
                    ticker_titulo = partes[1]
        except:
            pass

        # 2. Intentar desde archivo_excel_auto
        if not ticker_titulo and archivo_excel_auto:
            partes = os.path.basename(archivo_excel_auto).replace(".xlsx", "").split("_")
            if len(partes) >= 2:
                ticker_titulo = partes[1]

        # 3. Intentar desde CSV global (entry_ruta)
        if not ticker_titulo:
            csv_global = entry_ruta.get().strip().strip('"')
            if csv_global:
                partes = os.path.basename(csv_global).replace(".csv", "").split("_")
                if len(partes) >= 2:
                    ticker_titulo = partes[1]

        # Obtener rango de fechas - buscar en todas las hojas
        fecha_inicio = None
        fecha_fin = None
        for hoja_nombre, df_hoja in todos_dfs.items():
            if 'Fecha' in df_hoja.columns and len(df_hoja) > 0:
                fechas = pd.to_datetime(df_hoja['Fecha'], errors='coerce').dropna()
                if len(fechas) > 0:
                    fecha_inicio = fechas.min().strftime("%d-%m-%Y")
                    fecha_fin = fechas.max().strftime("%d-%m-%Y")
                    break

        if fecha_inicio and fecha_fin:
            if ticker_titulo:
                titulo_grafico = f"Análisis de {ticker_titulo} del {fecha_inicio} al {fecha_fin}"
            else:
                titulo_grafico = f"Análisis del {fecha_inicio} al {fecha_fin}"
    except Exception as e:
        print(f"Error extrayendo título: {e}")

    # Variable para guardar info del grafico actual
    hoja_actual = [clave_seleccionada]
    modo_actual = ["individual"]  # "individual" o "comparar"

    # ========== FRAME DE CONTROLES (layout horizontal) ==========
    frame_controles = tk.Frame(ventana_grafico)
    frame_controles.pack(fill="x", padx=10, pady=5)

    modo_var = tk.StringVar(value="individual")

    # ----- Columna izquierda: Selector de modo (vertical) -----
    frame_modo = tk.Frame(frame_controles)
    frame_modo.pack(side="left", padx=10)

    tk.Label(frame_modo, text="Modo:", font=("Arial", 10, "bold")).pack(anchor="w")

    rb_individual = tk.Radiobutton(frame_modo, text="Vista individual", variable=modo_var,
                                    value="individual")
    rb_individual.pack(anchor="w")

    rb_comparar = tk.Radiobutton(frame_modo, text="Comparar analisis", variable=modo_var,
                                  value="comparar")
    rb_comparar.pack(anchor="w")

    # ----- Columna central: Opciones segun modo -----
    frame_opciones = tk.Frame(frame_controles)
    frame_opciones.pack(side="left", fill="both", expand=True, padx=10)

    # Título del gráfico (visible en ambos modos)
    if titulo_grafico:
        tk.Label(frame_opciones, text=" " * 60 + titulo_grafico,
                 font=("Arial", 11, "bold"), fg="#0066cc").pack(anchor="w")

    # Frame para Vista Individual
    frame_individual = tk.Frame(frame_opciones)

    if len(todos_dfs) > 1:
        tk.Label(frame_individual, text="Seleccionar hoja:", font=("Arial", 10)).pack(anchor="w")
        combo_hojas = ttk.Combobox(frame_individual, values=list(todos_dfs.keys()), state="readonly", width=40)
        combo_hojas.pack(anchor="w", pady=2)
        combo_hojas.current(0)
    else:
        combo_hojas = None
        tk.Label(frame_individual, text=f"Hoja: {clave_seleccionada}", font=("Arial", 10)).pack(anchor="w")

    # Frame para Modo Comparar
    frame_comparar = tk.Frame(frame_opciones)

    # Selector de variable (en linea horizontal)
    frame_variable = tk.Frame(frame_comparar)
    frame_variable.pack(anchor="w", pady=2)
    tk.Label(frame_variable, text="Variable a comparar:", font=("Arial", 10)).pack(side="left")
    variables_comparar = ["Acciones en cartera", "Margen", "Rentabilidad"]
    combo_variable = ttk.Combobox(frame_variable, values=variables_comparar, state="readonly", width=20)
    combo_variable.pack(side="left", padx=5)
    combo_variable.current(0)

    # Organizar hojas por periodo para los checkboxes
    periodos = {}
    for hoja in todos_dfs.keys():
        if " - " in hoja:
            periodo = hoja.split(" - ")[0].strip()
        elif "_" in hoja:
            periodo = hoja.split("_")[0].strip()
        else:
            periodo = hoja
        if periodo not in periodos:
            periodos[periodo] = []
        periodos[periodo].append(hoja)

    # Frame para checkboxes organizados en columnas
    frame_checks = tk.Frame(frame_comparar)
    frame_checks.pack(anchor="w", pady=3)

    check_vars = {}
    col = 0
    for periodo, hojas_periodo in periodos.items():
        for i, hoja in enumerate(hojas_periodo):
            var = tk.IntVar(value=0)
            check_vars[hoja] = var
            cb = tk.Checkbutton(frame_checks, text=hoja, variable=var)
            cb.grid(row=i, column=col, sticky="w", padx=5)
        col += 1

    # ----- Boton actualizar (junto a las opciones) -----
    frame_boton = tk.Frame(frame_controles)
    frame_boton.pack(side="left", padx=15, anchor="n")

    btn_actualizar = tk.Button(frame_boton, text="Actualizar Grafico",
                                bg="#28a745", fg="white", width=16)
    btn_actualizar.pack(pady=5)

    # Funcion para cambiar entre modos
    def cambiar_modo():
        modo = modo_var.get()
        modo_actual[0] = modo
        if modo == "individual":
            frame_comparar.pack_forget()
            frame_individual.pack(fill="x")
            actualizar_grafico_individual()
        else:
            frame_individual.pack_forget()
            frame_comparar.pack(fill="x")

    # Vincular cambio de modo a los radio buttons
    rb_individual.config(command=cambiar_modo)
    rb_comparar.config(command=cambiar_modo)

    # Mostrar frame individual por defecto
    frame_individual.pack(fill="x")

    # ========== CREAR FIGURA Y CANVAS ==========
    fig = plt.figure(figsize=(16, 10))
    canvas = FigureCanvasTkAgg(fig, master=ventana_grafico)
    canvas.get_tk_widget().pack(fill="both", expand=True)

    # ========== FUNCIONES DE GRAFICADO ==========
    def actualizar_grafico_individual(event=None):
        """Actualiza el grafico con la hoja seleccionada (modo individual)"""
        if combo_hojas:
            hoja = combo_hojas.get()
        else:
            hoja = list(todos_dfs.keys())[0]

        hoja_actual[0] = hoja
        df_grafico = todos_dfs[hoja].copy()
        generar_grafico_en_figura(fig, df_grafico, hoja)
        canvas.draw()

    def actualizar_grafico_comparativo():
        """Genera grafico comparativo de una variable entre multiples analisis.

        CONFIGURACION DE GRAFICO COMPARATIVO (documentada 19-Mar-2026):
        ================================================================
        Margenes: left=0.06, right=0.92, top=0.92, bottom=0.12

        Escalado eje Y secundario (ax2):
        - Margen 10% arriba y abajo del rango de datos

        Formato Rentabilidad: 1 decimal (ej: 25.3)
        Otros: entero (ej: 250)
        """
        variable = combo_variable.get()
        hojas_seleccionadas = [h for h, var in check_vars.items() if var.get() == 1]

        if not hojas_seleccionadas:
            messagebox.showwarning("Aviso", "Selecciona al menos un analisis para comparar")
            return

        # Mapeo de variable a columna del dataframe
        col_map = {
            "Acciones en cartera": "Acciones en cartera",
            "Margen": "Margen",
            "Rentabilidad": "Rentabilidad_num"
        }
        col_variable = col_map.get(variable, variable)

        # Limpiar figura
        fig.clear()
        ax1 = fig.add_subplot(111)
        fig.subplots_adjust(left=0.06, right=0.92, top=0.92, bottom=0.12)

        # Colores para las diferentes series
        colores = ['#2ecc71', '#e74c3c', '#9b59b6', '#f39c12', '#1abc9c', '#e91e63']

        # Graficar precio (siempre el mismo, usar primera hoja)
        df_precio = todos_dfs[hojas_seleccionadas[0]].copy()
        # Normalizar columnas
        if 'Último' in df_precio.columns:
            df_precio = df_precio.rename(columns={'Último': 'Ultimo'})
        if df_precio['Fecha'].dtype == 'object':
            df_precio['Fecha'] = pd.to_datetime(df_precio['Fecha'], dayfirst=True, errors='coerce')

        ax1.plot(df_precio['Fecha'], df_precio['Ultimo'], color='blue', label='Precio (Ultimo)', linewidth=2)
        ax1.set_ylabel('Precio (Ultimo)', color='blue')
        ax1.tick_params(axis='y', labelcolor='blue')

        # Crear segundo eje para la variable comparada
        ax2 = ax1.twinx()

        # Graficar cada analisis seleccionado
        # Nombre corto de la variable para leyenda
        nombre_corto_var = {
            "Acciones en cartera": "Acciones",
            "Margen": "Margen",
            "Rentabilidad": "Rentabilidad"
        }.get(variable, variable)

        for i, hoja in enumerate(hojas_seleccionadas):
            df = todos_dfs[hoja].copy()
            # Normalizar columnas
            if 'Último' in df.columns:
                df = df.rename(columns={'Último': 'Ultimo'})
            if df['Fecha'].dtype == 'object':
                df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')

            # Para Rentabilidad, convertir a numero si es necesario
            if col_variable == "Rentabilidad_num":
                if "Rentabilidad_num" not in df.columns:
                    df['Rentabilidad_num'] = pd.to_numeric(
                        df['Rentabilidad'].astype(str).str.replace('%', '', regex=False).str.replace(',', '.', regex=False).str.strip(),
                        errors='coerce'
                    ).fillna(0)
                elif df['Rentabilidad_num'].dtype == 'object' or not pd.api.types.is_numeric_dtype(df['Rentabilidad_num']):
                    df['Rentabilidad_num'] = pd.to_numeric(
                        df['Rentabilidad_num'].astype(str).str.replace('%', '', regex=False).str.replace(',', '.', regex=False).str.strip(),
                        errors='coerce'
                    ).fillna(0)

            # Extraer tipo de analisis para leyenda corta
            if " - " in hoja:
                tipo_analisis = hoja.split(" - ", 1)[1]
            elif "_" in hoja:
                tipo_analisis = hoja.split("_", 1)[1]
            else:
                tipo_analisis = hoja
            # Eliminar "meses_" si existe
            tipo_analisis = tipo_analisis.replace("meses_", "").replace("Meses_", "")
            # Abreviar "Margen" a "Margen prom" si aplica
            if "Margen" in tipo_analisis or "margen" in tipo_analisis:
                tipo_analisis = "Margen prom"
            # Capitalizar primera letra
            tipo_analisis = tipo_analisis.capitalize() if tipo_analisis else tipo_analisis
            leyenda = f"{nombre_corto_var} ({tipo_analisis})"

            color = colores[i % len(colores)]
            # Estilo de linea segun variable
            if variable == "Acciones en cartera":
                estilo_linea = ':'   # puntos
            elif variable == "Rentabilidad":
                estilo_linea = '--'  # guiones
            else:
                estilo_linea = '-'   # linea solida
            ax2.plot(df['Fecha'], df[col_variable], color=color, label=leyenda, linewidth=2, linestyle=estilo_linea)

        ax2.set_ylabel(variable, color='black')
        ax2.tick_params(axis='y', labelcolor='black')

        # Formatear valores
        ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
        # Para Rentabilidad usar formato con decimales
        if variable == "Rentabilidad":
            ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x:.1f}'))
        else:
            ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))

        # Ajustar escala del Precio para que inicie en cero
        ax1.set_ylim(bottom=0)

        # Ajustar escala del eje Y secundario con margen
        ax2.relim()
        ax2.autoscale_view()
        ymin2, ymax2 = ax2.get_ylim()
        rango2 = ymax2 - ymin2
        ax2.set_ylim(ymin2 - rango2 * 0.1, ymax2 + rango2 * 0.1)

        # Leyendas combinadas
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=9)

        # Formatear fechas en eje X para evitar traslape
        # Calcular intervalo dinámico según cantidad de días
        num_dias = len(df_precio)
        if num_dias <= 30:
            intervalo = 2
        elif num_dias <= 90:
            intervalo = 5
        elif num_dias <= 180:
            intervalo = 10
        else:
            intervalo = 15
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m'))
        ax1.xaxis.set_major_locator(DayLocator(interval=intervalo))
        plt.setp(ax1.xaxis.get_majorticklabels(), fontsize=8, rotation=45, ha='right')

        # Titulo
        fig.suptitle(f'Comparacion de {variable} entre analisis', fontsize=12)

        hoja_actual[0] = f"comparacion_{variable}"
        canvas.draw()

    # Funcion para actualizar segun el modo
    def actualizar_segun_modo():
        if modo_var.get() == "individual":
            actualizar_grafico_individual()
        else:
            actualizar_grafico_comparativo()

    # Vincular eventos
    if combo_hojas:
        combo_hojas.bind("<<ComboboxSelected>>", actualizar_grafico_individual)

    btn_actualizar.config(command=actualizar_segun_modo)

    # Mostrar primer grafico
    actualizar_grafico_individual()

    # ========== FRAME DE BOTONES ==========
    frame_botones = tk.Frame(ventana_grafico)
    frame_botones.pack(pady=5)

    def guardar_grafico():
        from tkinter import filedialog
        ruta = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("PDF", "*.pdf"), ("Todos", "*.*")],
            initialfile=f"grafico_{hoja_actual[0]}.png"
        )
        if ruta:
            fig.savefig(ruta, dpi=150, bbox_inches='tight')
            messagebox.showinfo("Guardado", f"Grafico guardado en:\n{ruta}")

    tk.Button(frame_botones, text="Guardar Grafico", command=guardar_grafico,
              bg="#17a2b8", fg="white", width=15).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_grafico.destroy,
              bg="#6c757d", fg="white", width=10).pack(side="left", padx=5)


# =========================
# Interfaz Gráfica
# =========================
ventana = tk.Tk()
ventana.title("Parámetros del análisis")
ventana.geometry("1100x700")  # Tamaño inicial de la ventana

# =========================================================
# Crear Canvas con Scrollbar vertical para toda la interfaz
# =========================================================
canvas_principal = tk.Canvas(ventana)
scrollbar_vertical = tk.Scrollbar(ventana, orient="vertical", command=canvas_principal.yview)
canvas_principal.configure(yscrollcommand=scrollbar_vertical.set)

scrollbar_vertical.pack(side="right", fill="y")
canvas_principal.pack(side="left", fill="both", expand=True)

# Frame principal dentro del canvas (aquí van todos los widgets)
frame_principal = tk.Frame(canvas_principal)
canvas_window = canvas_principal.create_window((0, 0), window=frame_principal, anchor="nw")

# Configurar el scroll para que funcione con la rueda del mouse
def on_mousewheel(event):
    canvas_principal.yview_scroll(int(-1*(event.delta/120)), "units")

canvas_principal.bind_all("<MouseWheel>", on_mousewheel)

# Actualizar el tamaño del canvas cuando cambie el frame
def configurar_scroll(event):
    canvas_principal.configure(scrollregion=canvas_principal.bbox("all"))
    # Ajustar el ancho del canvas al frame
    canvas_principal.itemconfig(canvas_window, width=event.width if event.width > 1000 else 1000)

frame_principal.bind("<Configure>", configurar_scroll)

# Ajustar el ancho del frame cuando cambie el canvas
def on_canvas_configure(event):
    canvas_principal.itemconfig(canvas_window, width=event.width)

canvas_principal.bind("<Configure>", on_canvas_configure)

# Configuración del grid del frame principal
frame_principal.grid_columnconfigure(1, weight=1)
frame_principal.grid_columnconfigure(0, weight=0)
frame_principal.grid_columnconfigure(2, weight=0)
frame_principal.grid_columnconfigure(3, weight=1)

tk.Label(frame_principal, text="Ruta del CSV (TAB):").grid(row=0, column=0, sticky="w")
entry_ruta = tk.Entry(frame_principal, width=55)
entry_ruta.grid(row=0, column=1, sticky="we")


def seleccionar_csv():
    global ticker_actual

    ruta = filedialog.askopenfilename(
        title="Selecciona el archivo CSV (guardado desde Excel, separado por TAB)",
        filetypes=[("CSV files", "*.csv;*.txt"), ("Todos los archivos", "*.*")]
    )
    entry_ruta.delete(0, tk.END)
    entry_ruta.insert(0, ruta)

    # Mostrar info del JSON si existe
    if ruta:
        nombre_archivo = os.path.splitext(os.path.basename(ruta))[0]
        ticker_actual = nombre_archivo  # Guardar ticker actual
        mostrar_info_json_ticker(nombre_archivo, ruta)  # CORREGIDO: Pasar ruta del CSV


tk.Button(frame_principal, text="Seleccionar", command=seleccionar_csv).grid(row=0, column=2, sticky="w", padx=(6, 0))

# Botón para seleccionar ubicación JSON (row=1, pegado a row=0)
frame_json_config = tk.Frame(frame_principal)
frame_json_config.grid(row=1, column=0, columnspan=3, sticky="w", pady=(1, 0))

tk.Button(frame_json_config, text="Configurar ubicación JSON",
          command=seleccionar_ubicacion_json, bg="lightblue").pack(side="left")

tk.Button(frame_json_config, text="Administrar JSON",
          command=administrar_json, bg="#ffcc80").pack(side="left", padx=(10, 0))

tk.Button(frame_json_config, text="Params Activos",
          command=administrar_parametros_activos, bg="#90EE90").pack(side="left", padx=(10, 0))

label_json_actual = tk.Label(frame_json_config, text="JSON: No configurado", fg="gray")
label_json_actual.pack(side="left", padx=(10, 0))

# AHORA sí cargar configuración (después de crear label_json_actual)
cargar_configuracion()

# Frame para mostrar info JSON del ticker seleccionado (DEBAJO en row=2, ancho completo)
frame_info_json = tk.Frame(frame_principal, relief="groove", borderwidth=2, padx=5, pady=3)
frame_info_json.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(1, 0))

# ------------------------------------------------
# CHECKBOXES: Objetivo de optimización (múltiple selección)
# ------------------------------------------------
tk.Label(frame_principal, text="Objetivo optimización:").grid(row=3, column=0, sticky="w")
# Variables para checkboxes de objetivos
objetivo_rentabilidad_var = tk.IntVar(value=1)  # Por defecto marcado
objetivo_margen_var = tk.IntVar(value=0)
frame_objetivo = tk.Frame(frame_principal)
frame_objetivo.grid(row=3, column=1, sticky="w")
tk.Checkbutton(frame_objetivo, text="Rentabilidad máx", variable=objetivo_rentabilidad_var).pack(side="left")
tk.Checkbutton(frame_objetivo, text="Margen promedio máx", variable=objetivo_margen_var).pack(side="left", padx=(10, 0))

# Función helper para obtener objetivos seleccionados
def obtener_objetivos_seleccionados():
    """Retorna lista de objetivos seleccionados"""
    objetivos = []
    if objetivo_rentabilidad_var.get() == 1:
        objetivos.append("rentabilidad")
    if objetivo_margen_var.get() == 1:
        objetivos.append("margen_prom")
    return objetivos

# CHECKBOX: Usar optimización SciPy
usar_scipy_var = tk.IntVar(value=0)
chk_scipy = tk.Checkbutton(frame_principal, text="Usar optimización avanzada (SciPy)", variable=usar_scipy_var)
chk_scipy.grid(row=3, column=2, sticky="w", padx=(10, 0))

# ------------------------------------------------
# CAMPO Compra (%) + CHECKBOX DE OPTIMIZACIÓN
# ------------------------------------------------
tk.Label(frame_principal, text="Compra (%):").grid(row=4, column=0, sticky="w")

frame_compra = tk.Frame(frame_principal)
frame_compra.grid(row=4, column=1, sticky="w")

entry_compra = tk.Entry(frame_compra, width=6)
entry_compra.insert(0, "-1.6")
entry_compra.pack(side="left")

auto_compra_var = tk.IntVar(value=0)
chk_auto = tk.Checkbutton(frame_compra, text="Auto", variable=auto_compra_var)
chk_auto.pack(side="left", padx=(5, 0))

# Frame para botones de análisis
frame_botones_analisis = tk.Frame(frame_principal)
frame_botones_analisis.grid(row=4, column=2, sticky="w", padx=(10, 0))

btn_iniciar_analisis = tk.Button(frame_botones_analisis, text="▶ Iniciar análisis",
                                  command=lambda: iniciar_proceso(), bg="#90EE90")
btn_iniciar_analisis.pack(side="left")

btn_detener_analisis = tk.Button(frame_botones_analisis, text="⏹ Detener",
                                  command=lambda: detener_analisis(), bg="#ff6b6b", fg="white", state="disabled")
btn_detener_analisis.pack(side="left", padx=(5, 0))

# ------------------------------------------------
# CAMPO Venta (%) + CHECKBOX DE OPTIMIZACIÓN
# ------------------------------------------------
tk.Label(frame_principal, text="Venta (%):").grid(row=5, column=0, sticky="w")

frame_venta = tk.Frame(frame_principal)
frame_venta.grid(row=5, column=1, sticky="w")

entry_venta = tk.Entry(frame_venta, width=6)
entry_venta.insert(0, "1.6")
entry_venta.pack(side="left")

auto_venta_var = tk.IntVar(value=0)
chk_auto_venta = tk.Checkbutton(frame_venta, text="Auto", variable=auto_venta_var)
chk_auto_venta.pack(side="left", padx=(5, 0))

# Botón "Generar DB y Excel" al lado de Venta (azul con letras negras)
btn_generar_db_excel = tk.Button(frame_principal, text="Generar DB y Excel", command=lambda: generar_db_excel(),
                                 bg="#1E90FF", fg="black", font=("Arial", 9, "bold"), width=18)
btn_generar_db_excel.grid(row=5, column=2, sticky="w", padx=(10, 0))

# Boton "Ver Grafico" debajo de Generar DB (verde)
btn_ver_grafico = tk.Button(frame_principal, text="Ver Grafico", command=mostrar_grafico_resultados,
                            bg="#28a745", fg="white", font=("Arial", 9, "bold"), width=12)
btn_ver_grafico.grid(row=6, column=2, sticky="w", padx=(10, 0))

# ------------------------------------------------
# CAMPO: Ganancia mínima (%) + CHECKBOX
# ------------------------------------------------
tk.Label(frame_principal, text="Ganancia mínima (%):").grid(row=6, column=0, sticky="w")

frame_ganancia = tk.Frame(frame_principal)
frame_ganancia.grid(row=6, column=1, sticky="w")

entry_ganancia_minima = tk.Entry(frame_ganancia, width=6)
entry_ganancia_minima.insert(0, "0")
entry_ganancia_minima.pack(side="left")

auto_ganancia_var = tk.IntVar(value=0)
chk_auto_ganancia = tk.Checkbutton(frame_ganancia, text="Auto", variable=auto_ganancia_var)
chk_auto_ganancia.pack(side="left", padx=(5, 0))

tk.Label(frame_principal, text="Suave (%):").grid(row=7, column=0, sticky="w")
entry_suave = tk.Entry(frame_principal, width=6)
entry_suave.insert(0, "0.5")
entry_suave.grid(row=7, column=1, sticky="w")

tipo_limite_var = tk.StringVar(value="acciones")
opciones_limite = ["acciones", "aporte"]
selector_limite = tk.OptionMenu(frame_principal, tipo_limite_var, *opciones_limite)
selector_limite.grid(row=8, column=0, sticky="w")

frame_limite = tk.Frame(frame_principal)
frame_limite.grid(row=8, column=1, sticky="w")
entry_limite = tk.Entry(frame_limite, width=10)
entry_limite.insert(0, "10")
entry_limite.pack(side="left")
tk.Label(frame_limite, text="Valor límite").pack(side="left", padx=(5, 0))

# =========================================================
# Frame para Compra múltiple
# =========================================================
frame_compra_multiple = tk.Frame(frame_principal)
frame_compra_multiple.grid(row=9, column=0, columnspan=2, sticky="w", pady=(5, 0))

tk.Label(frame_compra_multiple, text="Compra de N acciones:").pack(side="left")

entry_compra_multiple = tk.Entry(frame_compra_multiple, width=6)
entry_compra_multiple.pack(side="left", padx=(5, 0))
entry_compra_multiple.insert(0, "")

auto_compra_mult_var = tk.IntVar(value=0)
chk_auto_compra_mult = tk.Checkbutton(frame_compra_multiple, text="Auto", variable=auto_compra_mult_var)
chk_auto_compra_mult.pack(side="left", padx=(5, 0))

# =========================================================
# Frame para Venta múltiple
# =========================================================
frame_venta_multiple = tk.Frame(frame_principal)
frame_venta_multiple.grid(row=10, column=0, columnspan=2, sticky="w", pady=(5, 0))

tk.Label(frame_venta_multiple, text="Venta de N acciones:").pack(side="left")

entry_venta_multiple = tk.Entry(frame_venta_multiple, width=6)
entry_venta_multiple.pack(side="left", padx=(5, 0))
entry_venta_multiple.insert(0, "")

auto_venta_mult_var = tk.IntVar(value=0)
chk_auto_venta_mult = tk.Checkbutton(frame_venta_multiple, text="Auto", variable=auto_venta_mult_var)
chk_auto_venta_mult.pack(side="left", padx=(5, 0))

# =========================================================
# Frame para mostrar fechas de compras/ventas múltiples
# =========================================================
frame_fechas_multiples = tk.Frame(frame_principal)
frame_fechas_multiples.grid(row=9, column=2, rowspan=2, sticky="nw", padx=(10, 0))

# Cuadro de fechas de compras múltiples
tk.Label(frame_fechas_multiples, text="Fechas compras múltiples:", font=("Arial", 8)).grid(row=0, column=0, sticky="w")
text_compras_mult = tk.Text(frame_fechas_multiples, width=15, height=4, font=("Arial", 7))
text_compras_mult.grid(row=1, column=0, sticky="w", padx=(0, 10))

# Cuadro de fechas de ventas múltiples
tk.Label(frame_fechas_multiples, text="Fechas ventas múltiples:", font=("Arial", 8)).grid(row=0, column=1, sticky="w")
text_ventas_mult = tk.Text(frame_fechas_multiples, width=15, height=4, font=("Arial", 7))
text_ventas_mult.grid(row=1, column=1, sticky="w")

# =========================================================
# Frame para selección de períodos a analizar
# =========================================================
frame_periodos = tk.Frame(frame_principal, relief="ridge", borderwidth=2, padx=10, pady=5)
frame_periodos.grid(row=11, column=0, columnspan=3, sticky="w", pady=(10, 0))

tk.Label(frame_periodos, text="Analizar períodos:", font=("Arial", 10, "bold")).pack(side="left", padx=(0, 10))

analizar_completo_var = tk.IntVar(value=1)
analizar_6meses_var = tk.IntVar(value=0)
analizar_3meses_var = tk.IntVar(value=0)

tk.Checkbutton(frame_periodos, text="Completo", variable=analizar_completo_var).pack(side="left", padx=5)
tk.Checkbutton(frame_periodos, text="Últimos 6 meses", variable=analizar_6meses_var).pack(side="left", padx=5)
tk.Checkbutton(frame_periodos, text="Últimos 3 meses", variable=analizar_3meses_var).pack(side="left", padx=5)

# Botón verde para guardar en JSON
btn_guardar_json = tk.Button(frame_periodos, text="💾 Guardar resultados en JSON",
                             command=guardar_resultados_en_json, bg="lightgreen",
                             font=("Arial", 10, "bold"), state="disabled")
btn_guardar_json.pack(side="left", padx=(20, 0))

# =========================================================
# Frame de estadísticas
# =========================================================
ventana.frame_stats = tk.Frame(frame_principal, padx=10, pady=2)
ventana.frame_stats.grid(row=12, column=0, columnspan=3, sticky="w")

# =========================================================
# Barra de progreso para optimización
# =========================================================
frame_progreso = tk.Frame(frame_principal, padx=10, pady=2)
frame_progreso.grid(row=13, column=0, columnspan=3, sticky="we")

ventana.progress_bar = ttk.Progressbar(frame_progreso, length=600, mode='determinate')
ventana.label_progreso = tk.Label(frame_progreso, text="", font=("Arial", 10))

# Label para mostrar resultado de optimización
ventana.label_resultado_opt = tk.Label(frame_principal, text="", font=("Arial", 10, "bold"), fg="darkgreen")

ultimo_df = None
ultima_ruta_excel = ""
ultimo_folder = ""
ultimo_base_name = ""

# Diccionario para almacenar DataFrames por período
resultados_dfs_por_periodo = {}

# Variable global para acumular análisis POR TICKER (no mezclar tickers)
historial_analisis_por_ticker = {}
ticker_actual = None


# =========================
# CAMBIO 1 y 2: Funcion generar DB y Excel (boton)
# =========================
def generar_db_excel():
    global resultados_dfs_por_periodo, ultimo_folder, ultimo_base_name

    _cargar_dependencias_analisis()
    _cargar_dependencias_excel()
    _cargar_sqlite()

    if not resultados_dfs_por_periodo:
        messagebox.showerror("Error", "No hay análisis previo. Ejecuta primero 'Iniciar análisis'.")
        return

    # Cambiar estado del botón
    btn_generar_db_excel.config(state="disabled", bg="gray", text="Generando...")
    ventana.update()

    mensajes_resultado = []
    archivos_generados = []
    errores = []

    try:
        # CAMBIO 1: Excel que ACUMULA pestañas (no sobrescribe)
        ruta_excel = os.path.join(ultimo_folder, f"{ultimo_base_name}_analizado.xlsx")
        objetivo = OBJETIVO_ACTUAL

        # Si el archivo existe, cargar y agregar nuevas pestañas
        if os.path.exists(ruta_excel):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(ruta_excel)
            except PermissionError:
                errores.append(f"❌ Excel: El archivo está abierto, ciérralo primero")
                wb = None
        else:
            from openpyxl import Workbook
            wb = Workbook()
            # Eliminar la hoja por defecto si existe
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        if wb is not None:
            for nombre_periodo, df in resultados_dfs_por_periodo.items():
                # Crear nombre de pestaña descriptivo
                nombre_hoja = f"{nombre_periodo}_{objetivo}"[:31]

                # Si la pestaña ya existe, eliminarla para actualizarla
                if nombre_hoja in wb.sheetnames:
                    del wb[nombre_hoja]

                # Crear nueva pestaña
                ws = wb.create_sheet(nombre_hoja)

                # Escribir datos
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(ruta_excel)
            archivos_generados.append(
                f"✓ Excel: {os.path.basename(ruta_excel)} ({len(resultados_dfs_por_periodo)} pestañas)")

    except PermissionError:
        errores.append(f"❌ Excel: El archivo está abierto, ciérralo primero")
    except Exception as e:
        errores.append(f"❌ Excel: {str(e)}")

    try:
        # CAMBIO 2: SQLite unificado con múltiples tablas
        db_path = os.path.join(ultimo_folder, f"{ultimo_base_name}_analizado.db")

        conn = sqlite3.connect(db_path)

        for nombre_periodo, df in resultados_dfs_por_periodo.items():
            # Nombre de tabla: periodo_objetivo
            objetivo = OBJETIVO_ACTUAL
            tabla_nombre = f"{nombre_periodo}_{objetivo}"

            # Eliminar tabla si existe (con comillas)
            conn.execute(f'DROP TABLE IF EXISTS "{tabla_nombre}"')

            # Crear tabla (con comillas para nombres que empiezan con números)
            conn.execute(f"""
                CREATE TABLE "{tabla_nombre}" (
                    Fecha TEXT,
                    Ultimo REAL,
                    Apertura REAL,
                    Maximo REAL,
                    Minimo REAL,
                    Vol REAL,
                    Var TEXT,
                    Acumulado TEXT,
                    Opcion TEXT,
                    Movimiento INTEGER,
                    Acciones INTEGER,
                    PrecioCompra REAL,
                    CapitalBolsa REAL,
                    CapitalAcciones REAL,
                    CapitalTotal REAL,
                    Aporte REAL,
                    AporteAcumulado REAL,
                    Margen REAL,
                    Rentabilidad TEXT
                )
            """)

            # Insertar datos
            rows = []
            for _, r in df.iterrows():
                rows.append((
                    r.get("Fecha", ""),
                    to_float_safe(r.get("Último", 0)),
                    to_float_safe(r.get("Apertura", 0)),
                    to_float_safe(r.get("Máximo", 0)),
                    to_float_safe(r.get("Mínimo", 0)),
                    to_float_safe(r.get("Vol.", 0)),
                    str(r.get("% var.", "")),
                    str(r.get("% acumulado", "")),
                    str(r.get("Opción", "")),
                    int(r.get("Movimiento de acciones", 0)) if pd.notna(r.get("Movimiento de acciones", 0)) else 0,
                    int(r.get("Acciones en cartera", 0)) if pd.notna(r.get("Acciones en cartera", 0)) else 0,
                    to_float_safe(r.get("Precio de compra", 0)),
                    to_float_safe(r.get("Capital en bolsa", 0)),
                    to_float_safe(r.get("Capital en acciones", 0)),
                    to_float_safe(r.get("Capital total", 0)),
                    to_float_safe(r.get("Aporte", 0)),
                    to_float_safe(r.get("Aporte acumulado", 0)),
                    to_float_safe(r.get("Margen", 0)),
                    str(r.get("Rentabilidad", ""))
                ))

            placeholders = ",".join(["?"] * 19)
            conn.executemany(f'INSERT INTO "{tabla_nombre}" VALUES ({placeholders})', rows)

        conn.commit()
        conn.close()

        archivos_generados.append(f"✓ SQLite: {os.path.basename(db_path)} ({len(resultados_dfs_por_periodo)} tablas)")

    except Exception as e:
        errores.append(f"❌ SQLite: {str(e)}")

    # MEJORA: Una sola ventana de diálogo con todos los resultados
    mensaje_final = ""

    if archivos_generados:
        mensaje_final += "ARCHIVOS GENERADOS:\n\n" + "\n".join(archivos_generados)

    if errores:
        if mensaje_final:
            mensaje_final += "\n\n"
        mensaje_final += "ERRORES:\n\n" + "\n".join(errores)

    # Si todo fue exitoso, deshabilitar botón (como el de JSON)
    if not errores and archivos_generados:
        btn_generar_db_excel.config(state="disabled", bg="lightgray", fg="gray", text="Generar DB y Excel")
    else:
        # Si hubo errores, restaurar botón para reintentar
        btn_generar_db_excel.config(state="normal", bg="#1E90FF", fg="black", text="Generar DB y Excel")

    if errores:
        messagebox.showwarning("Generación completada con errores", mensaje_final)
    else:
        messagebox.showinfo("Generación exitosa", mensaje_final)


# =========================
# Función que ejecuta TODO el análisis con un UMBRAL_COMPRA dado
# =========================
def ejecutar_analisis_con_umbral(umbral_compra_decimal, csv_filtrado=None):
    global text_ventas_mult, text_compras_mult, INPUT_FILE, error_analisis_mostrado

    try:
        local_venta = float(entry_venta.get().replace(",", ".")) / 100
        local_suave = float(entry_suave.get().replace(",", ".")) / 100
        ganancia_minima = float(entry_ganancia_minima.get().replace(",", ".")) / 100
    except:
        messagebox.showerror("Error", "Valores numéricos inválidos en Venta / Suave / Ganancia mínima.")
        return None, -999999, -999999

    # Si se proporciona un CSV filtrado, usarlo; si no, cargar el original
    if csv_filtrado is not None:
        df = csv_filtrado.copy()
    else:
        try:
            # Intentar primero con utf-8-sig para manejar BOM, luego latin-1
            try:
                df = pd.read_csv(INPUT_FILE, sep=";", engine='python', dtype=str, encoding='utf-8-sig')
            except:
                df = pd.read_csv(INPUT_FILE, sep=";", engine='python', dtype=str, encoding='latin-1')
            df = normalizar_columnas(df)  # Convertir columnas inglés a español si es necesario
            df['Fecha'] = df['Fecha'].astype(str).str.strip()

            def parse_mixed_dates(date_str):
                for fmt in ("%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        return pd.to_datetime(date_str, format=fmt)
                    except:
                        continue
                return pd.NaT

            df['Fecha'] = df['Fecha'].apply(parse_mixed_dates)
            df = df.dropna(subset=['Fecha'])
            df = df.sort_values('Fecha').reset_index(drop=True)

        except Exception as e:
            error_analisis_mostrado = True
            messagebox.showerror("Error al leer CSV", str(e) + "\n\nEl análisis se detendrá.")
            return None, -999999, -999999

    # Verificar columnas
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        error_analisis_mostrado = True
        messagebox.showerror("Error", f"Columnas faltantes: {missing}\nColumnas encontradas: {list(df.columns)}\n\nEl análisis se detendrá.")
        return None, -999999, -999999

    df = df[EXPECTED_COLUMNS].copy()

    # Procesar fechas
    df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
    df = df.dropna(subset=['Fecha'])
    df = df.sort_values("Fecha").reset_index(drop=True)

    # Guardar fechas antes de convertir a string
    fecha_inicial = df['Fecha'].min()
    fecha_final = df['Fecha'].max()

    df['Fecha'] = df['Fecha'].dt.strftime("%d/%m/%Y")

    # Convertir numéricos
    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(to_float_safe)

    df['% var.'] = df['% var.'].apply(parse_percent_to_decimal)

    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(lambda x: float(x) if pd.notna(x) else 0.0)

    df['% var.'] = df['% var.'].apply(lambda x: float(x) if pd.notna(x) else 0.0)

    # Calcular % acumulado por signos consecutivos
    acum = 0
    prev = 0
    lst = []
    for v in df['% var.']:
        sign = 1 if v > 0 else -1 if v < 0 else 0
        if sign == prev:
            acum += v
        else:
            acum = v
        lst.append(acum)
        prev = sign
    df['% acumulado'] = lst

    acum_decimal = df['% acumulado'].astype(float)

    valores_seleccionados = []
    seq = []
    indices_seq = []

    for idx, v in enumerate(acum_decimal):
        if v > 0:
            seq.append(v)
            indices_seq.append(idx)
        else:
            if len(seq) >= 2:
                valores_seleccionados.append(seq[-1] * 100.0)
            seq = []
            indices_seq = []
    if len(seq) >= 2:
        valores_seleccionados.append(seq[-1] * 100.0)

    promedio_maximos = (sum(valores_seleccionados) / len(valores_seleccionados) / 100.0) if valores_seleccionados else 0.0

    valores_minimos = []
    seq_neg = []

    for idx, v in enumerate(acum_decimal):
        if v < 0:
            seq_neg.append(v)
        else:
            if len(seq_neg) >= 2:
                valores_minimos.append(seq_neg[-1] * 100.0)
            seq_neg = []
    if len(seq_neg) >= 2:
        valores_minimos.append(seq_neg[-1] * 100.0)

    promedio_minimos = (sum(valores_minimos) / len(valores_minimos) / 100.0) if valores_minimos else 0.0

    def determinar_opcion(v, a):
        if v >= local_venta:
            return "Venta"
        if v <= umbral_compra_decimal:
            return "Compra"
        if a >= local_venta and v >= local_suave:
            return "Venta"
        if a <= umbral_compra_decimal and v <= -local_suave:
            return "Compra"
        return "N/A"

    df['Opción'] = df.apply(lambda r: determinar_opcion(r['% var.'], r['% acumulado']), axis=1)

    try:
        if LIMITE_TIPO == "acciones":
            MAX_ACCIONES = int(LIMITE_VALOR)
            MAX_APORTE = float("inf")
        else:
            MAX_ACCIONES = 10
            MAX_APORTE = float(LIMITE_VALOR)
    except:
        MAX_ACCIONES = 10
        MAX_APORTE = float("inf")

    acciones = 0
    capital_bolsa = 0
    aporte_acumulado = 0

    movs, acts, cap_b, cap_acc, cap_tot, aport, aport_acum, precios_compra = [], [], [], [], [], [], [], []

    precios_en_cartera = []

    acum_pct = df['% acumulado'].astype(float) * 100.0
    comprar_multiple = [False] * len(df)

    seq_idxs_neg = []
    all_negative_sequences = []
    for idx, v in enumerate(acum_pct):
        if v < 0:
            seq_idxs_neg.append(idx)
        else:
            if len(seq_idxs_neg) >= 2:
                all_negative_sequences.append(seq_idxs_neg.copy())
            seq_idxs_neg = []
    if len(seq_idxs_neg) >= 2:
        all_negative_sequences.append(seq_idxs_neg.copy())

    if promedio_minimos < 0.0:
        for s in all_negative_sequences:
            for i in s:
                if acum_pct.iloc[i] <= promedio_minimos:
                    comprar_multiple[i] = True

    vender_doble = [False] * len(df)

    seq_idxs = []
    all_positive_sequences = []
    for idx, v in enumerate(acum_pct):
        if v > 0:
            seq_idxs.append(idx)
        else:
            if len(seq_idxs) >= 2:
                all_positive_sequences.append(seq_idxs.copy())
            seq_idxs = []
    if len(seq_idxs) >= 2:
        all_positive_sequences.append(seq_idxs.copy())

    if promedio_maximos > 0.0:
        for s in all_positive_sequences:
            for i in s:
                if acum_pct.iloc[i] >= promedio_maximos:
                    vender_doble[i] = True

    for idx, row in df.iterrows():
        opcion = row["Opción"]
        precio = row["Último"]
        movimiento = 0
        aporte = 0.0
        precio_operacion = 0.0

        if opcion == "Compra":
            n_compra = 1
            if COMPRA_MULTIPLE_ACCIONES is not None and comprar_multiple[idx]:
                n_compra = COMPRA_MULTIPLE_ACCIONES

            acciones_a_comprar = 0
            for _ in range(n_compra):
                puede_comprar = False
                if LIMITE_TIPO == "acciones" and acciones < MAX_ACCIONES:
                    puede_comprar = True
                elif LIMITE_TIPO == "aporte" and (aporte_acumulado + precio) <= MAX_APORTE:
                    puede_comprar = True

                if puede_comprar:
                    acciones_a_comprar += 1
                    if capital_bolsa >= precio:
                        capital_bolsa -= precio
                    else:
                        aporte += precio
                        aporte_acumulado += precio
                        capital_bolsa += precio
                        capital_bolsa -= precio
                    acciones += 1
                    precios_en_cartera.append(precio)
                    precios_en_cartera.sort()
                else:
                    break

            movimiento = acciones_a_comprar
            if movimiento > 0:
                precio_operacion = -precio

        elif opcion == "Venta" and acciones > 0:
            acciones_vendibles = 0
            for precio_compra in precios_en_cartera:
                ganancia_porcentual = (precio - precio_compra) / precio_compra
                if ganancia_porcentual >= ganancia_minima:
                    acciones_vendibles += 1
                else:
                    break

            if acciones_vendibles > 0:
                n_venta = 1
                if VENTA_MULTIPLE_ACCIONES is not None and vender_doble[idx] and acciones >= VENTA_MULTIPLE_ACCIONES:
                    n_venta = VENTA_MULTIPLE_ACCIONES

                n_venta = min(n_venta, acciones_vendibles, acciones)

                capital_bolsa += precio * n_venta
                acciones -= n_venta
                movimiento = -n_venta

                for _ in range(n_venta):
                    if precios_en_cartera:
                        precios_en_cartera.pop(0)

                if movimiento < 0:
                    precio_operacion = precio

        movs.append(movimiento)
        acts.append(acciones)
        cap_b.append(round(capital_bolsa, 2))
        cap_acc.append(round(acciones * precio, 2))
        cap_tot.append(round(capital_bolsa + acciones * precio, 2))
        aport.append(round(aporte, 2))
        aport_acum.append(round(aporte_acumulado, 2))
        precios_compra.append(precio_operacion)

    df["Movimiento de acciones"] = movs
    df["Acciones en cartera"] = acts
    df["Precio de compra"] = precios_compra
    df["Capital en bolsa"] = cap_b
    df["Capital en acciones"] = cap_acc
    df["Capital total"] = cap_tot
    df["Aporte"] = aport
    df["Aporte acumulado"] = aport_acum

    if text_compras_mult is not None:
        text_compras_mult.delete("1.0", tk.END)

        if COMPRA_MULTIPLE_ACCIONES is not None:
            fechas_compra_multiple = df[df['Movimiento de acciones'] == COMPRA_MULTIPLE_ACCIONES]['Fecha'].tolist()
            if fechas_compra_multiple:
                text_compras_mult.insert(tk.END, "\n".join(fechas_compra_multiple))
            else:
                text_compras_mult.insert(tk.END, f"No hay compras de {COMPRA_MULTIPLE_ACCIONES} acciones")

    if text_ventas_mult is not None:
        text_ventas_mult.delete("1.0", tk.END)

        if VENTA_MULTIPLE_ACCIONES is not None:
            fechas_venta_multiple = df[df['Movimiento de acciones'] == -VENTA_MULTIPLE_ACCIONES]['Fecha'].tolist()
            if fechas_venta_multiple:
                text_ventas_mult.insert(tk.END, "\n".join(fechas_venta_multiple))
            else:
                text_ventas_mult.insert(tk.END, f"No hay ventas de {VENTA_MULTIPLE_ACCIONES} acciones")

    df["Margen"] = df["Capital total"] - df["Aporte acumulado"]
    df["Rentabilidad"] = df.apply(
        lambda r: (r["Margen"] / r["Aporte acumulado"] * 100) if r["Aporte acumulado"] > 0 else 0, axis=1)

    rentab_max = df["Rentabilidad"].max()
    margen_prom = df["Margen"].mean()

    df["Rentabilidad"] = df["Rentabilidad"].round(2).astype(str) + "%"
    df["% var."] = (df["% var."] * 100).round(2).astype(str) + "%"
    df["% acumulado"] = (df["% acumulado"] * 100).round(2).astype(str) + "%"

    return df, rentab_max, margen_prom, fecha_inicial.strftime("%d/%m/%Y"), fecha_final.strftime("%d/%m/%Y")


# =========================
# Variables globales para progreso
# =========================
scipy_evaluaciones = 0
scipy_evaluaciones_max = 0
scipy_inicio_tiempo = None


# =========================
# Función para refinar el óptimo (encontrar centro del rango)
# =========================
def refinar_optimo(params_optimos, bounds, csv_filtrado=None, n_muestras=30, umbral_similitud=0.95):
    """
    Muestrea alrededor del óptimo encontrado para hallar el centro del rango
    que produce resultados similares.

    Args:
        params_optimos: Lista con los parámetros óptimos encontrados [compra, venta, ganancia, compra_mult, venta_mult]
        bounds: Límites de cada parámetro [(min, max), ...]
        csv_filtrado: DataFrame filtrado o None para usar el completo
        n_muestras: Número de puntos a muestrear alrededor del óptimo
        umbral_similitud: Porcentaje mínimo del resultado óptimo para considerar similar (0.95 = 95%)

    Returns:
        Lista con los parámetros promediados
    """
    global COMPRA_MULTIPLE_ACCIONES, VENTA_MULTIPLE_ACCIONES, analisis_detenido

    # Si el análisis fue detenido, retornar los parámetros originales
    if analisis_detenido:
        return params_optimos

    # Evaluar el resultado óptimo original
    compra_orig = params_optimos[0]
    COMPRA_MULTIPLE_ACCIONES = int(round(params_optimos[3])) if params_optimos[3] > 1.5 else None
    VENTA_MULTIPLE_ACCIONES = int(round(params_optimos[4])) if params_optimos[4] > 1.5 else None

    entry_compra.delete(0, tk.END)
    entry_compra.insert(0, f"{params_optimos[0]:.1f}")
    entry_venta.delete(0, tk.END)
    entry_venta.insert(0, f"{params_optimos[1]:.1f}")
    entry_ganancia_minima.delete(0, tk.END)
    entry_ganancia_minima.insert(0, f"{params_optimos[2]:.1f}")

    df_orig, rent_orig, margen_orig, _, _ = ejecutar_analisis_con_umbral(compra_orig / 100, csv_filtrado)

    if df_orig is None:
        return params_optimos  # Si falla, retornar los originales

    # Determinar métrica a usar
    usar_margen = (OBJETIVO_ACTUAL == "margen_prom")
    metrica_optima = margen_orig if usar_margen else rent_orig
    umbral_metrica = metrica_optima * umbral_similitud

    # Generar muestras alrededor del óptimo (±10% de cada parámetro)
    np.random.seed(42)  # Semilla fija para reproducibilidad

    params_similares = [list(params_optimos)]  # Incluir el óptimo original

    for _ in range(n_muestras):
        # Verificar si el análisis fue detenido
        if analisis_detenido:
            break

        params_muestra = []
        for i, (p, (b_min, b_max)) in enumerate(zip(params_optimos, bounds)):
            # Calcular rango de variación (±10% del valor o ±10% del rango total)
            rango = max(abs(p) * 0.1, (b_max - b_min) * 0.05)

            # Para parámetros enteros (compra_mult, venta_mult)
            if i >= 3:
                nuevo_val = p + np.random.uniform(-1, 1)
                nuevo_val = max(b_min, min(b_max, nuevo_val))
            else:
                nuevo_val = p + np.random.uniform(-rango, rango)
                nuevo_val = max(b_min, min(b_max, nuevo_val))

            params_muestra.append(nuevo_val)

        # Evaluar esta muestra
        COMPRA_MULTIPLE_ACCIONES = int(round(params_muestra[3])) if params_muestra[3] > 1.5 else None
        VENTA_MULTIPLE_ACCIONES = int(round(params_muestra[4])) if params_muestra[4] > 1.5 else None

        entry_compra.delete(0, tk.END)
        entry_compra.insert(0, f"{params_muestra[0]:.1f}")
        entry_venta.delete(0, tk.END)
        entry_venta.insert(0, f"{params_muestra[1]:.1f}")
        entry_ganancia_minima.delete(0, tk.END)
        entry_ganancia_minima.insert(0, f"{params_muestra[2]:.1f}")

        try:
            df_test, rent_test, margen_test, _, _ = ejecutar_analisis_con_umbral(params_muestra[0] / 100, csv_filtrado)

            if df_test is not None:
                metrica_test = margen_test if usar_margen else rent_test

                # Si el resultado es similar al óptimo, guardar estos parámetros
                if metrica_test >= umbral_metrica:
                    params_similares.append(params_muestra)
        except:
            continue

    # Calcular promedio de todos los parámetros similares
    if len(params_similares) > 1:
        params_promedio = []
        for i in range(5):
            valores = [p[i] for p in params_similares]
            promedio = sum(valores) / len(valores)

            # Redondear parámetros enteros
            if i >= 3:
                promedio = round(promedio)

            params_promedio.append(promedio)

        print(f"  -> Refinamiento: {len(params_similares)} configuraciones similares encontradas")
        print(f"  -> Parámetros promediados: Compra={params_promedio[0]:.2f}%, Venta={params_promedio[1]:.2f}%")

        return params_promedio
    else:
        print(f"  -> Refinamiento: Solo el óptimo original cumple el umbral")
        return params_optimos


# =========================
# Función objetivo para optimización con SciPy
# =========================
def funcion_objetivo_scipy(params, csv_filtrado=None):
    global COMPRA_MULTIPLE_ACCIONES, VENTA_MULTIPLE_ACCIONES
    global scipy_evaluaciones, scipy_evaluaciones_max, scipy_inicio_tiempo
    global analisis_detenido

    # Verificar si el usuario detuvo el análisis - retornar valor alto para terminar rápido
    if analisis_detenido:
        return float('inf')

    scipy_evaluaciones += 1

    compra_pct = params[0]
    venta_pct = params[1]
    ganancia_min = params[2]
    compra_mult = int(round(params[3])) if params[3] > 1.5 else None
    venta_mult = int(round(params[4])) if params[4] > 1.5 else None

    entry_compra.delete(0, tk.END)
    entry_compra.insert(0, f"{compra_pct:.1f}")

    entry_venta.delete(0, tk.END)
    entry_venta.insert(0, f"{venta_pct:.1f}")

    entry_ganancia_minima.delete(0, tk.END)
    entry_ganancia_minima.insert(0, f"{ganancia_min:.1f}")

    if compra_mult is None:
        entry_compra_multiple.delete(0, tk.END)
    else:
        entry_compra_multiple.delete(0, tk.END)
        entry_compra_multiple.insert(0, str(compra_mult))

    if venta_mult is None:
        entry_venta_multiple.delete(0, tk.END)
    else:
        entry_venta_multiple.delete(0, tk.END)
        entry_venta_multiple.insert(0, str(venta_mult))

    COMPRA_MULTIPLE_ACCIONES = compra_mult
    VENTA_MULTIPLE_ACCIONES = venta_mult

    if scipy_evaluaciones % 5 == 0:
        porcentaje = (scipy_evaluaciones / scipy_evaluaciones_max) * 100
        ventana.progress_bar['value'] = porcentaje

        tiempo_transcurrido = time.time() - scipy_inicio_tiempo
        if scipy_evaluaciones > 10:
            tiempo_por_eval = tiempo_transcurrido / scipy_evaluaciones
            evals_restantes = scipy_evaluaciones_max - scipy_evaluaciones
            tiempo_restante = tiempo_por_eval * evals_restantes * 0.95

            mins_restantes = int(tiempo_restante // 60)
            segs_restantes = int(tiempo_restante % 60)

            ventana.label_progreso.config(
                text=f"Progreso: {scipy_evaluaciones}/{scipy_evaluaciones_max} ({porcentaje:.1f}%) - "
                     f"Tiempo estimado restante: {mins_restantes}m {segs_restantes}s"
            )

        ventana.update()
        time.sleep(0.001)

    try:
        df, rent_tmp, margen_tmp, _, _ = ejecutar_analisis_con_umbral(compra_pct / 100, csv_filtrado)

        if df is None:
            return 999999

        usar_margen = (OBJETIVO_ACTUAL == "margen_prom")
        if usar_margen:
            metrica = margen_tmp
        else:
            metrica = rent_tmp

        return -metrica
    except:
        return 999999


# =========================
# Función para optimizar un período específico
# =========================
def optimizar_periodo(nombre_periodo, dias=None):
    """Ejecuta optimización para un período específico"""
    global scipy_evaluaciones, scipy_evaluaciones_max, scipy_inicio_tiempo
    global COMPRA_MULTIPLE_ACCIONES, VENTA_MULTIPLE_ACCIONES

    print(f"\n{'=' * 60}")
    print(f"Optimizando período: {nombre_periodo}")
    print(f"{'=' * 60}")

    # Filtrar datos si es necesario
    if dias is not None:
        csv_filtrado = filtrar_ultimos_dias(INPUT_FILE, dias)
        if csv_filtrado is None:
            # Error al leer el archivo, ya se mostró mensaje
            return None
    else:
        csv_filtrado = None
        print(f"  -> Analizando datos completos")

    # Determinar si hay optimización activa
    usar_scipy = (usar_scipy_var.get() == 1)
    hay_optimizacion = (auto_compra_var.get() == 1 or auto_venta_var.get() == 1 or
                        auto_ganancia_var.get() == 1 or auto_compra_mult_var.get() == 1 or
                        auto_venta_mult_var.get() == 1)

    mejor_df = None
    mejor_compra = None
    mejor_venta = None
    mejor_ganancia = None
    mejor_compra_mult = None
    mejor_venta_mult = None
    fecha_inicial = None
    fecha_final = None

    # ===============================================================
    # OPTIMIZACIÓN CON SCIPY
    # ===============================================================
    if usar_scipy and hay_optimizacion:
        bounds = []

        if auto_compra_var.get() == 1:
            bounds.append((-3.0, 0.0))
        else:
            try:
                val = float(entry_compra.get().replace(",", "."))
                bounds.append((val, val))
            except:
                bounds.append((-1.6, -1.6))

        if auto_venta_var.get() == 1:
            bounds.append((0.0, 3.0))
        else:
            try:
                val = float(entry_venta.get().replace(",", "."))
                bounds.append((val, val))
            except:
                bounds.append((1.6, 1.6))

        if auto_ganancia_var.get() == 1:
            bounds.append((1.5, 3.0))  # Máximo 3% de ganancia mínima
        else:
            try:
                val = float(entry_ganancia_minima.get().replace(",", "."))
                bounds.append((val, val))
            except:
                bounds.append((0.0, 0.0))

        if auto_compra_mult_var.get() == 1:
            bounds.append((0, 5))
        else:
            val_cm = entry_compra_multiple.get().strip()
            if val_cm == "":
                bounds.append((0, 0))
            else:
                try:
                    val = int(val_cm)
                    bounds.append((val, val))
                except:
                    bounds.append((0, 0))

        if auto_venta_mult_var.get() == 1:
            bounds.append((0, 5))
        else:
            val_vm = entry_venta_multiple.get().strip()
            if val_vm == "":
                bounds.append((0, 0))
            else:
                try:
                    val = int(val_vm)
                    bounds.append((val, val))
                except:
                    bounds.append((0, 0))

        ventana.progress_bar.grid(row=0, column=0, columnspan=2, sticky="we", pady=2)
        ventana.label_progreso.grid(row=1, column=0, columnspan=2, sticky="w")

        # Calcular progreso base (porcentaje de combinaciones completadas)
        if progreso_total_combinaciones > 0:
            progreso_base = ((progreso_combinacion_actual - 1) / progreso_total_combinaciones) * 100
            progreso_slice = 100 / progreso_total_combinaciones  # Porcentaje que representa esta combinación
        else:
            progreso_base = 0
            progreso_slice = 100

        ventana.progress_bar['value'] = progreso_base

        periodo_legible = nombre_periodo.replace("_", " ").title().replace("6 Meses", "6M").replace("3 Meses", "3M")
        obj_texto = "Rent" if OBJETIVO_ACTUAL == "rentabilidad" else "Marg"
        ventana.label_progreso.config(
            text=f"Optimizando {progreso_combinacion_actual}/{progreso_total_combinaciones}: {periodo_legible} - {obj_texto}..."
        )

        maxiter = 100
        popsize = 15
        scipy_evaluaciones_max = maxiter * popsize
        scipy_evaluaciones = 0
        scipy_inicio_tiempo = time.time()

        ventana.update()

        # Callback para actualizar progreso y permitir detención
        def callback_progreso(xk, convergence):
            global scipy_evaluaciones
            scipy_evaluaciones += 1

            # Calcular progreso combinado (global + local)
            if scipy_evaluaciones_max > 0:
                progreso_local = (scipy_evaluaciones / scipy_evaluaciones_max) * progreso_slice
            else:
                progreso_local = 0

            progreso_total = progreso_base + progreso_local
            ventana.progress_bar['value'] = min(progreso_total, 100)
            ventana.update()

            return analisis_detenido  # Retornar True detiene la optimización

        resultado = differential_evolution(
            lambda params: funcion_objetivo_scipy(params, csv_filtrado),
            bounds,
            strategy='best1bin',
            maxiter=maxiter,
            popsize=popsize,
            tol=0.01,
            mutation=(0.5, 1),
            recombination=0.7,
            seed=42,  # Semilla fija para resultados reproducibles
            callback=callback_progreso,
            disp=False,
            polish=False,  # Desactivar polish para permitir detención limpia
            init='latinhypercube',
            atol=0,
            updating='immediate',
            workers=1
        )

        ventana.progress_bar.grid_forget()
        ventana.label_progreso.grid_forget()

        # Verificar si el usuario detuvo el análisis
        if analisis_detenido:
            return None

        # Refinar el óptimo encontrado (encontrar centro del rango)
        print(f"\n  -> Refinando parámetros óptimos...")
        ventana.label_progreso.config(text="Refinando parámetros óptimos...")
        ventana.label_progreso.grid(row=1, column=0, columnspan=2, sticky="w")
        ventana.update()

        params_refinados = refinar_optimo(
            params_optimos=list(resultado.x),
            bounds=bounds,
            csv_filtrado=csv_filtrado,
            n_muestras=30,
            umbral_similitud=0.95
        )

        ventana.label_progreso.grid_forget()

        mejor_compra = params_refinados[0]
        mejor_venta = params_refinados[1]
        mejor_ganancia = params_refinados[2]
        mejor_compra_mult = int(round(params_refinados[3])) if params_refinados[3] > 1.5 else None
        mejor_venta_mult = int(round(params_refinados[4])) if params_refinados[4] > 1.5 else None

        entry_compra.delete(0, tk.END)
        entry_compra.insert(0, f"{mejor_compra:.1f}")

        entry_venta.delete(0, tk.END)
        entry_venta.insert(0, f"{mejor_venta:.1f}")

        entry_ganancia_minima.delete(0, tk.END)
        entry_ganancia_minima.insert(0, f"{mejor_ganancia:.1f}")

        if mejor_compra_mult is None:
            entry_compra_multiple.delete(0, tk.END)
            COMPRA_MULTIPLE_ACCIONES = None
        else:
            entry_compra_multiple.delete(0, tk.END)
            entry_compra_multiple.insert(0, str(mejor_compra_mult))
            COMPRA_MULTIPLE_ACCIONES = mejor_compra_mult

        if mejor_venta_mult is None:
            entry_venta_multiple.delete(0, tk.END)
            VENTA_MULTIPLE_ACCIONES = None
        else:
            entry_venta_multiple.delete(0, tk.END)
            entry_venta_multiple.insert(0, str(mejor_venta_mult))
            VENTA_MULTIPLE_ACCIONES = mejor_venta_mult

        mejor_df, _, _, fecha_inicial, fecha_final = ejecutar_analisis_con_umbral(mejor_compra / 100, csv_filtrado)

    # ===============================================================
    # SIN SCIPY (bucles anidados o ejecución directa)
    # ===============================================================
    else:
        # Aquí iría el código de optimización sin SciPy (bucles anidados)
        # Por brevedad, ejecuto directamente con los valores actuales
        try:
            compra_val = float(entry_compra.get().replace(",", ".")) / 100
        except:
            compra_val = -0.016

        mejor_df, _, _, fecha_inicial, fecha_final = ejecutar_analisis_con_umbral(compra_val, csv_filtrado)
        mejor_compra = compra_val * 100
        mejor_venta = float(entry_venta.get().replace(",", "."))
        mejor_ganancia = float(entry_ganancia_minima.get().replace(",", "."))
        mejor_compra_mult = COMPRA_MULTIPLE_ACCIONES
        mejor_venta_mult = VENTA_MULTIPLE_ACCIONES

    if mejor_df is None:
        return None

    # Calcular estadísticas completas del análisis
    def float_col(col_name):
        return mejor_df[col_name].astype(str).str.rstrip('%').str.replace(',', '.').astype(float)

    # Calcular % acumulado y promedios de máximos/mínimos
    acumulado_float = float_col('% acumulado') * 100.0

    # Promedio de máximos (secuencias positivas)
    valores_seleccionados = []
    secuencia = []
    for v in acumulado_float:
        if v > 0:
            secuencia.append(v)
        else:
            if len(secuencia) >= 2:
                valores_seleccionados.append(secuencia[-1])
            secuencia = []
    if len(secuencia) >= 2:
        valores_seleccionados.append(secuencia[-1])
    promedio_maximos = sum(valores_seleccionados) / len(valores_seleccionados) if valores_seleccionados else 0.0

    # Promedio de mínimos (secuencias negativas)
    valores_minimos = []
    secuencia_neg = []
    for v in acumulado_float:
        if v < 0:
            secuencia_neg.append(v)
        else:
            if len(secuencia_neg) >= 2:
                valores_minimos.append(secuencia_neg[-1])
            secuencia_neg = []
    if len(secuencia_neg) >= 2:
        valores_minimos.append(secuencia_neg[-1])
    promedio_minimos = sum(valores_minimos) / len(valores_minimos) if valores_minimos else 0.0

    # Estadísticas de % variación
    max_var = float_col('% var.').max()
    min_var = float_col('% var.').min()
    fecha_max_var = mejor_df.loc[float_col('% var.').idxmax(), 'Fecha']
    fecha_min_var = mejor_df.loc[float_col('% var.').idxmin(), 'Fecha']
    dif_var = max_var - min_var

    prom_var = float_col('% var.')
    subidas = prom_var[prom_var > 0]
    max_prom = subidas.mean() if not subidas.empty else 0
    bajadas = prom_var[prom_var < 0]
    min_prom = bajadas.mean() if not bajadas.empty else 0
    dif_prom = max_prom - min_prom

    # Estadísticas de operaciones
    opc_compra = int((mejor_df["Opción"] == "Compra").sum())
    acciones_compradas = int(mejor_df.loc[mejor_df["Movimiento de acciones"] > 0, "Movimiento de acciones"].sum())
    opc_venta = int((mejor_df["Opción"] == "Venta").sum())
    acciones_vendidas = int(-mejor_df.loc[mejor_df["Movimiento de acciones"] < 0, "Movimiento de acciones"].sum())
    max_acc_cartera = int(mejor_df["Acciones en cartera"].max())

    # Estadísticas financieras
    max_aporte = float(mejor_df["Aporte acumulado"].max())
    max_margen = float(round(mejor_df["Margen"].max(), 2))
    margen_promedio = float(round(mejor_df["Margen"].mean(), 2))
    max_rentab = float(float_col("Rentabilidad").max())
    rentab_promedio = float(float_col("Rentabilidad").mean())
    fecha_max_rentab = mejor_df.loc[float_col("Rentabilidad").idxmax(), "Fecha"]

    # Preparar resultado con todas las estadísticas
    resultado = {
        "df": mejor_df,
        "compra_pct": mejor_compra,
        "venta_pct": mejor_venta,
        "ganancia_min": mejor_ganancia,
        "suave_pct": float(entry_suave.get().replace(",", ".")),
        "limite_tipo": tipo_limite_var.get(),
        "limite_valor": float(entry_limite.get().replace(",", ".")),
        "compra_mult": mejor_compra_mult,
        "venta_mult": mejor_venta_mult,
        "rentabilidad_max": max_rentab,
        "margen_promedio": margen_promedio,
        "fecha_inicial": fecha_inicial,
        "fecha_final": fecha_final,
        # Nuevos campos de estadísticas
        "promedio_maximos": promedio_maximos,
        "promedio_minimos": promedio_minimos,
        "max_var": max_var,
        "min_var": min_var,
        "fecha_max_var": fecha_max_var,
        "fecha_min_var": fecha_min_var,
        "dif_var": dif_var,
        "max_prom_var": max_prom,
        "min_prom_var": min_prom,
        "dif_prom_var": dif_prom,
        "opc_compra": opc_compra,
        "acciones_compradas": acciones_compradas,
        "opc_venta": opc_venta,
        "acciones_vendidas": acciones_vendidas,
        "max_acc_cartera": max_acc_cartera,
        "max_aporte": max_aporte,
        "max_margen": max_margen,
        "rentab_promedio": rentab_promedio,
        "fecha_max_rentab": fecha_max_rentab
    }

    return resultado


# =========================
# Función iniciar_proceso (principal)
# =========================
def iniciar_proceso():
    global ultimo_df, ultima_ruta_excel, ultimo_folder, ultimo_base_name
    global INPUT_FILE, FOLDER, LIMITE_TIPO, LIMITE_VALOR
    global VENTA_MULTIPLE_ACCIONES, COMPRA_MULTIPLE_ACCIONES
    global resultados_analisis_actuales, resultados_dfs_por_periodo
    global analisis_detenido, error_analisis_mostrado

    _cargar_dependencias_analisis()

    # Resetear variables de control
    analisis_detenido = False
    error_analisis_mostrado = False

    # Configurar botones (deshabilitar Iniciar, habilitar Detener)
    btn_iniciar_analisis.config(state="disabled")
    btn_detener_analisis.config(state="normal")
    ventana.update()

    # Limpiar mensaje anterior de optimización
    ventana.label_resultado_opt.config(text="")
    ventana.label_resultado_opt.grid_forget()

    # Limpiar resultados previos
    resultados_analisis_actuales = {}
    resultados_dfs_por_periodo = {}

    INPUT_FILE = entry_ruta.get().strip().strip('"')
    if not os.path.exists(INPUT_FILE):
        messagebox.showerror("Error", f"La ruta del CSV no existe:\n{INPUT_FILE}")
        return

    FOLDER = os.path.dirname(INPUT_FILE)
    base_name = os.path.splitext(os.path.basename(INPUT_FILE))[0]

    try:
        venta_val = float(entry_venta.get().replace(",", ".")) / 100
        suave_val = float(entry_suave.get().replace(",", ".")) / 100
    except:
        messagebox.showerror("Error", "Valores numéricos inválidos.")
        return

    LIMITE_TIPO = tipo_limite_var.get()
    try:
        LIMITE_VALOR = float(entry_limite.get().replace(",", "."))
    except:
        LIMITE_VALOR = 10.0

    # Validar compra múltiple
    valor_compra_multiple = entry_compra_multiple.get().strip()
    if valor_compra_multiple == "":
        COMPRA_MULTIPLE_ACCIONES = None
    else:
        try:
            n_acciones = int(valor_compra_multiple)
            if n_acciones < 2:
                messagebox.showerror("Error", "La cantidad para 'Compra de N acciones' debe ser 2 o más.")
                return
            COMPRA_MULTIPLE_ACCIONES = n_acciones
        except ValueError:
            messagebox.showerror("Error", "Debes ingresar un número entero válido en 'Compra de N acciones'.")
            return

    # Validar venta múltiple
    valor_venta_multiple = entry_venta_multiple.get().strip()
    if valor_venta_multiple == "":
        VENTA_MULTIPLE_ACCIONES = None
    else:
        try:
            n_acciones = int(valor_venta_multiple)
            if n_acciones < 2:
                messagebox.showerror("Error", "La cantidad para 'Venta de N acciones' debe ser 2 o más.")
                return
            VENTA_MULTIPLE_ACCIONES = n_acciones
        except ValueError:
            messagebox.showerror("Error", "Debes ingresar un número entero válido en 'Venta de N acciones'.")
            return

    # Determinar qué períodos analizar
    periodos_a_analizar = []

    if analizar_completo_var.get() == 1:
        periodos_a_analizar.append(("completo", None))

    if analizar_6meses_var.get() == 1:
        periodos_a_analizar.append(("6_meses", 180))

    if analizar_3meses_var.get() == 1:
        periodos_a_analizar.append(("3_meses", 90))

    if not periodos_a_analizar:
        messagebox.showerror("Error", "Selecciona al menos un período para analizar")
        return

    # Determinar qué objetivos analizar
    objetivos_a_analizar = obtener_objetivos_seleccionados()

    if not objetivos_a_analizar:
        messagebox.showerror("Error", "Selecciona al menos un objetivo de optimización")
        return

    # Variable global para que optimizar_periodo sepa qué objetivo usar
    global OBJETIVO_ACTUAL
    global progreso_combinacion_actual, progreso_total_combinaciones
    global progreso_tiempo_inicio_total, progreso_tiempos_combinaciones

    # =====================================================
    # PROGRESO INTELIGENTE: Preparación
    # =====================================================

    # Contar filas del CSV
    try:
        df_temp = pd.read_csv(INPUT_FILE, encoding='latin-1')
        num_filas = len(df_temp)
        del df_temp
    except:
        num_filas = 200  # Valor por defecto

    # Obtener configuración de checks activos
    checks_activos = {
        'scipy': usar_scipy_var.get() == 1,
        'compra': auto_compra_var.get() == 1,
        'venta': auto_venta_var.get() == 1,
        'ganancia': auto_ganancia_var.get() == 1,
        'compra_mult': auto_compra_mult_var.get() == 1,
        'venta_mult': auto_venta_mult_var.get() == 1
    }

    clave_config = obtener_clave_configuracion(num_filas, checks_activos)
    print(f"[DEBUG] Clave configuración: {clave_config} ({num_filas} filas)")

    # Analizar cada combinación de período y objetivo
    resultados_por_periodo = {}
    total_combinaciones = len(periodos_a_analizar) * len(objetivos_a_analizar)
    combinacion_actual = 0

    # Variables de progreso global
    progreso_combinacion_actual = 0
    progreso_total_combinaciones = total_combinaciones
    progreso_tiempo_inicio_total = time.time()
    progreso_tiempos_combinaciones = []

    # Estimar tiempo total si hay historial
    tiempo_estimado_total, hay_historial = estimar_tiempo_total(clave_config, total_combinaciones)

    if hay_historial:
        print(f"[INFO] Tiempo estimado total: {formatear_tiempo(tiempo_estimado_total)}")

    for objetivo in objetivos_a_analizar:
        OBJETIVO_ACTUAL = objetivo
        objetivo_texto = "Rentabilidad" if objetivo == "rentabilidad" else "Margen Prom"

        for nombre_periodo, dias in periodos_a_analizar:
            combinacion_actual += 1
            progreso_combinacion_actual = combinacion_actual

            # Verificar si el usuario detuvo el análisis
            if analisis_detenido:
                print(f"[DEBUG] Análisis detenido antes de procesar {nombre_periodo}/{objetivo}")
                break

            # Calcular tiempo restante estimado
            tiempo_transcurrido = time.time() - progreso_tiempo_inicio_total
            if hay_historial and tiempo_estimado_total:
                tiempo_restante = max(0, tiempo_estimado_total - tiempo_transcurrido)
                texto_tiempo = f" | Restante: ~{formatear_tiempo(tiempo_restante)}"
            elif len(progreso_tiempos_combinaciones) > 0:
                # Estimar basado en combinaciones ya completadas
                promedio_actual = sum(progreso_tiempos_combinaciones) / len(progreso_tiempos_combinaciones)
                combinaciones_restantes = total_combinaciones - combinacion_actual + 1
                tiempo_restante = promedio_actual * combinaciones_restantes
                texto_tiempo = f" | Restante: ~{formatear_tiempo(tiempo_restante)}"
            else:
                texto_tiempo = ""

            periodo_legible = nombre_periodo.replace("_", " ").title().replace("6 Meses", "6M").replace("3 Meses", "3M")
            obj_corto = "Rent" if objetivo == "rentabilidad" else "Marg"

            print(f"[INFO] Analizando {combinacion_actual}/{total_combinaciones}: {periodo_legible} - {obj_corto}{texto_tiempo}")

            # Mostrar progreso en la interfaz
            ventana.label_progreso.config(
                text=f"Analizando {combinacion_actual}/{total_combinaciones}: {periodo_legible} - {objetivo_texto}{texto_tiempo}"
            )
            ventana.label_progreso.grid(row=1, column=0, columnspan=2, sticky="w")

            # Actualizar barra de progreso global (porcentaje de combinaciones)
            progreso_global = ((combinacion_actual - 1) / total_combinaciones) * 100
            ventana.progress_bar['value'] = progreso_global
            ventana.progress_bar.grid(row=0, column=0, columnspan=2, sticky="we", pady=2)
            ventana.update()

            # Iniciar tiempo de esta combinación
            tiempo_inicio_combinacion = time.time()

            resultado = optimizar_periodo(nombre_periodo, dias)

            # Registrar tiempo de esta combinación
            tiempo_combinacion = time.time() - tiempo_inicio_combinacion
            progreso_tiempos_combinaciones.append(tiempo_combinacion)

            if resultado is None:
                if analisis_detenido:
                    # El usuario detuvo el análisis, salir del bucle
                    break
                elif error_analisis_mostrado:
                    # Ya se mostró un error detallado, solo salir sin mostrar otro mensaje
                    break
                else:
                    messagebox.showerror("Error", f"No se pudo optimizar: {nombre_periodo}/{objetivo_texto}\n\nEl análisis se detendrá.")
                    break  # Detener al primer error en lugar de continuar

            # Agregar el objetivo al resultado
            resultado["objetivo"] = objetivo

            # Guardar con clave que incluye período y objetivo
            clave_resultado = f"{nombre_periodo}_{objetivo}"
            resultados_por_periodo[clave_resultado] = resultado
            resultados_dfs_por_periodo[clave_resultado] = resultado["df"]

        if analisis_detenido:
            break

    # Guardar tiempos en historial (promedio de esta sesión)
    if progreso_tiempos_combinaciones and not analisis_detenido:
        tiempo_promedio = sum(progreso_tiempos_combinaciones) / len(progreso_tiempos_combinaciones)
        registrar_tiempo_combinacion(clave_config, tiempo_promedio)
        print(f"[INFO] Tiempo promedio por combinación: {formatear_tiempo(tiempo_promedio)}")

    # Ocultar barra de progreso y actualizar interfaz
    ventana.progress_bar.grid_forget()
    ventana.label_progreso.grid_forget()
    ventana.update()

    if not resultados_por_periodo:
        # Restaurar botones de análisis
        btn_iniciar_analisis.config(state="normal")
        btn_detener_analisis.config(state="disabled")

        if analisis_detenido:
            ventana.label_resultado_opt.config(text=f"⚠ Análisis detenido por el usuario")
            ventana.label_resultado_opt.config(fg="orange")
            ventana.label_resultado_opt.grid(row=14, column=0, columnspan=3, sticky="w", padx=10, pady=5)
        else:
            messagebox.showerror("Error", "No se obtuvieron resultados válidos")
        return

    # Guardar para JSON (con múltiples objetivos)
    resultados_analisis_actuales = {
        "ticker": base_name,
        "objetivos_analizados": objetivos_a_analizar,
        "periodos": resultados_por_periodo
    }

    # Guardar variables globales
    ultimo_folder = FOLDER
    ultimo_base_name = base_name

    # Mostrar resultados en interfaz
    mostrar_resultados_multiples_periodos(resultados_por_periodo)

    # Habilitar botones
    btn_guardar_json.config(state="normal")
    btn_generar_db_excel.config(state="normal", bg="#1E90FF", fg="black")  # REACTIVAR botón DB/Excel

    # Restaurar botones de análisis
    btn_iniciar_analisis.config(state="normal")
    btn_detener_analisis.config(state="disabled")

    # Mostrar mensaje de completado
    if analisis_detenido:
        ventana.label_resultado_opt.config(text=f"⚠ Análisis detenido por el usuario")
        ventana.label_resultado_opt.config(fg="orange")
    else:
        ventana.label_resultado_opt.config(text=f"✓ Análisis completado para {len(periodos_a_analizar)} período(s)")
        ventana.label_resultado_opt.config(fg="darkgreen")
    ventana.label_resultado_opt.grid(row=14, column=0, columnspan=3, sticky="w", padx=10, pady=5)


# =========================
# Función para mostrar estadísticas en la interfaz
# =========================
def mostrar_resultados_multiples_periodos(resultados):
    """Muestra los resultados de todos los períodos en pestañas"""
    global historial_analisis_por_ticker, ticker_actual

    # Inicializar historial para este ticker si no existe
    if ticker_actual not in historial_analisis_por_ticker:
        historial_analisis_por_ticker[ticker_actual] = []

    # Agregar resultados actuales al historial DEL TICKER ACTUAL
    for clave_periodo, datos in resultados.items():
        # Obtener objetivo de cada resultado individual
        objetivo_actual = datos.get("objetivo", "rentabilidad")
        objetivo_texto = "Rentabilidad" if objetivo_actual == "rentabilidad" else "Margen Prom"

        # Extraer solo el nombre del período (sin el objetivo)
        if "_rentabilidad" in clave_periodo:
            nombre_periodo = clave_periodo.replace("_rentabilidad", "")
        elif "_margen_prom" in clave_periodo:
            nombre_periodo = clave_periodo.replace("_margen_prom", "")
        else:
            nombre_periodo = clave_periodo

        # Convertir a formato legible
        periodo_legible = nombre_periodo.replace("_", " ").title()
        # Corregir "Seis Meses" a "6 Meses" y "Tres Meses" a "3 Meses"
        periodo_legible = periodo_legible.replace("Seis Meses", "6 Meses").replace("Tres Meses", "3 Meses")

        historial_analisis_por_ticker[ticker_actual].append({
            "periodo": periodo_legible,
            "objetivo": objetivo_texto,
            "compra_pct": datos['compra_pct'],
            "venta_pct": datos['venta_pct'],
            "ganancia_min": datos['ganancia_min'],
            "suave_pct": datos['suave_pct'],
            "compra_mult": datos['compra_mult'],
            "venta_mult": datos['venta_mult'],
            "rentabilidad_max": datos['rentabilidad_max'],
            "margen_promedio": datos['margen_promedio']
        })

    # Limpiar frame de estadísticas
    for widget in ventana.frame_stats.winfo_children():
        widget.destroy()

    # Crear notebook (pestañas) para cada período
    notebook = ttk.Notebook(ventana.frame_stats)
    notebook.pack(fill="both", expand=True, pady=(0, 10))

    for clave_periodo, datos in resultados.items():
        # Crear frame para este período
        frame_periodo = tk.Frame(notebook)

        # Crear nombre de pestaña legible (ej: "Completo - Rent" o "6 Meses - Margen")
        objetivo_actual = datos.get("objetivo", "rentabilidad")
        obj_corto = "Rent" if objetivo_actual == "rentabilidad" else "Margen"

        if "_rentabilidad" in clave_periodo:
            periodo_base = clave_periodo.replace("_rentabilidad", "")
        elif "_margen_prom" in clave_periodo:
            periodo_base = clave_periodo.replace("_margen_prom", "")
        else:
            periodo_base = clave_periodo

        periodo_texto = periodo_base.replace("_", " ").title()
        periodo_texto = periodo_texto.replace("Seis Meses", "6 Meses").replace("Tres Meses", "3 Meses")
        nombre_pestana = f"{periodo_texto} - {obj_corto}"

        notebook.add(frame_periodo, text=nombre_pestana)

        # Mostrar estadísticas
        mostrar_estadisticas_en_frame(frame_periodo, datos["df"], datos)

    # NUEVO: Frame INFERIOR con tabla consolidada SOLO del ticker actual
    frame_consolidado = tk.Frame(ventana.frame_stats, relief="ridge", borderwidth=2, bg="lightyellow", padx=10, pady=10)
    frame_consolidado.pack(fill="x", pady=(10, 0))

    # Extraer ticker real (siglas) para mostrar
    if ticker_actual:
        partes = ticker_actual.split('_')
        ticker_display = ticker_actual
        if len(partes) >= 2:
            for parte in partes:
                if parte.isupper() and 1 <= len(parte) <= 5:
                    ticker_display = parte
                    break
    else:
        ticker_display = "Actual"

    tk.Label(frame_consolidado, text=f"📊 HISTORIAL DE ANÁLISIS - {ticker_display} (Acumulativo)",
             font=("Arial", 11, "bold"), bg="lightyellow", fg="darkgreen").pack(anchor="w")

    # Crear tabla con parámetros
    frame_tabla_params = tk.Frame(frame_consolidado, bg="lightyellow")
    frame_tabla_params.pack(fill="x", pady=(5, 0))

    # Headers
    headers = ["#", "Período", "Objetivo", "Compra %", "Venta %", "Gan Mín %", "Suave %", "Comp", "Venta", "Rentab Máx",
               "Margen Prom"]
    for col, header in enumerate(headers):
        ancho = 5 if col == 0 else 11
        tk.Label(frame_tabla_params, text=header, font=("Arial", 8, "bold"),
                 bg="lightblue", relief="solid", borderwidth=1, width=ancho).grid(row=0, column=col, sticky="ew",
                                                                                  padx=1, pady=1)

    # Datos SOLO del ticker actual
    analisis_ticker_actual = historial_analisis_por_ticker.get(ticker_actual, [])

    if not analisis_ticker_actual:
        tk.Label(frame_consolidado, text="No hay análisis en el historial",
                 font=("Arial", 9), bg="lightyellow", fg="gray").pack(pady=10)
        return

    # Ordenar por período y luego por objetivo
    orden_periodos = {"Completo": 1, "6 Meses": 2, "3 Meses": 3}
    analisis_ticker_actual_ordenado = sorted(
        analisis_ticker_actual,
        key=lambda x: (orden_periodos.get(x['periodo'], 99), x['objetivo'])
    )

    # Colores por objetivo (base del objetivo, sin el número)
    colores_objetivo = {
        "rentabilidad": "#e8f5e9",  # Verde claro
        "margen": "#e3f2fd",         # Azul claro
    }
    color_default = "#fff3e0"  # Naranja claro para otros

    periodo_anterior = None
    fila_actual = 0

    for idx, analisis in enumerate(analisis_ticker_actual_ordenado, start=1):
        fila_actual += 1

        # Determinar color basado en el objetivo (sin número)
        objetivo_base = analisis['objetivo'].lower().split()[0]
        bg_color = colores_objetivo.get(objetivo_base, color_default)

        # Si cambia el período, agregar línea separadora
        if periodo_anterior is not None and analisis['periodo'] != periodo_anterior:
            # Agregar fila separadora
            for col in range(11):
                ancho = 5 if col == 0 else 11
                tk.Label(frame_tabla_params, text="", font=("Arial", 2),
                         bg="#999999", relief="flat", width=ancho, height=1).grid(
                    row=fila_actual, column=col, sticky="ew", padx=1, pady=0)
            fila_actual += 1

        periodo_anterior = analisis['periodo']

        valores = [
            str(idx),
            analisis['periodo'],
            analisis['objetivo'],
            f"{analisis['compra_pct']:.1f}",
            f"{analisis['venta_pct']:.1f}",
            f"{analisis['ganancia_min']:.1f}",
            f"{analisis['suave_pct']:.1f}",
            str(analisis['compra_mult']) if analisis['compra_mult'] else "-",
            str(analisis['venta_mult']) if analisis['venta_mult'] else "-",
            f"{analisis['rentabilidad_max']:.2f}%",
            f"{analisis['margen_promedio']:.2f}"
        ]

        for col, valor in enumerate(valores):
            ancho = 5 if col == 0 else 11
            tk.Label(frame_tabla_params, text=valor, font=("Arial", 7),
                     bg=bg_color, relief="solid", borderwidth=1, width=ancho).grid(
                row=fila_actual, column=col, sticky="ew", padx=1, pady=1)


def mostrar_estadisticas_en_frame(frame_parent, df, datos_periodo):
    """Muestra estadísticas de un período específico en un frame"""

    frame1 = tk.Frame(frame_parent, padx=15)
    frame1.grid(row=0, column=0, sticky="nw")
    frame2 = tk.Frame(frame_parent, padx=15)
    frame2.grid(row=0, column=1, sticky="nw")
    frame3 = tk.Frame(frame_parent, padx=15)
    frame3.grid(row=0, column=2, sticky="nw")
    frame4 = tk.Frame(frame_parent, padx=15)
    frame4.grid(row=0, column=3, sticky="nw")

    def float_col(col):
        return df[col].astype(str).str.rstrip('%').replace('', '0').astype(float)

    acumulado_float = df['% acumulado'].str.rstrip('%').astype(float)

    valores_seleccionados = []
    secuencia = []
    for v in acumulado_float:
        if v > 0:
            secuencia.append(v)
        else:
            if len(secuencia) >= 2:
                valores_seleccionados.append(secuencia[-1])
            secuencia = []
    if len(secuencia) >= 2:
        valores_seleccionados.append(secuencia[-1])

    promedio_maximos = sum(valores_seleccionados) / len(valores_seleccionados) if valores_seleccionados else 0.0

    valores_minimos = []
    secuencia_neg = []
    for v in acumulado_float:
        if v < 0:
            secuencia_neg.append(v)
        else:
            if len(secuencia_neg) >= 2:
                valores_minimos.append(secuencia_neg[-1])
            secuencia_neg = []
    if len(secuencia_neg) >= 2:
        valores_minimos.append(secuencia_neg[-1])

    promedio_minimos = sum(valores_minimos) / len(valores_minimos) if valores_minimos else 0.0

    max_var = float_col('% var.').max()
    min_var = float_col('% var.').min()
    fecha_max_var = df.loc[float_col('% var.').idxmax(), 'Fecha']
    fecha_min_var = df.loc[float_col('% var.').idxmin(), 'Fecha']
    dif_var = max_var - min_var

    prom_var = float_col('% var.')
    subidas = prom_var[prom_var > 0]
    max_prom = subidas.mean() if not subidas.empty else 0
    bajadas = prom_var[prom_var < 0]
    min_prom = bajadas.mean() if not bajadas.empty else 0
    dif_prom = max_prom - min_prom

    opc_compra = (df["Opción"] == "Compra").sum()
    acciones_compradas = df.loc[df["Movimiento de acciones"] > 0, "Movimiento de acciones"].sum()
    opc_venta = (df["Opción"] == "Venta").sum()
    acciones_vendidas = -df.loc[df["Movimiento de acciones"] < 0, "Movimiento de acciones"].sum()
    max_acc_cartera = df["Acciones en cartera"].max()
    max_aporte = df["Aporte acumulado"].max()
    max_margen = round(df["Margen"].max(), 2)
    margen_promedio = round(df["Margen"].mean(), 2)
    max_rentab = float_col("Rentabilidad").max()
    rentab_promedio = float_col("Rentabilidad").mean()
    fecha_max_rentab = df.loc[float_col("Rentabilidad").idxmax(), "Fecha"]

    tk.Label(frame1, fg="blue", text=f"Max % var : {max_var:.2f}% ({fecha_max_var})", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame1, fg="blue", text=f"Min % var : {min_var:.2f}% ({fecha_min_var})", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame1, fg="blue", text=f"Diferencia : {dif_var:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame1, fg="blue", text="", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame1, fg="blue", text=f"Prom de % var. acum máximos +: {promedio_maximos:.2f}%",
             font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame1, fg="blue", text=f"Prom de % var. acum mínimos -: {promedio_minimos:.2f}%",
             font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame2, fg="red", text=f"Prom % var + :  {max_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame2, fg="red", text=f"Prom % var - : {min_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame2, fg="red", text=f"Diferencia       :  {dif_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame3, fg="black", text=f"Opciones Compra       : {opc_compra}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Acciones Compradas : {int(acciones_compradas)}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame3, fg="black", text=f"Opciones Venta           : {opc_venta}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Acciones Vendidas      : {int(acciones_vendidas)}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame3, fg="black", text=f"Máx acción en cartera : {max_acc_cartera}", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame4, fg="purple", text=f"Aporte acum max  : {max_aporte:,.0f}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Margen max       : {max_margen:,.2f}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Margen promedio  : {margen_promedio:,.2f}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame4, fg="purple", text=f"Rentab. max      : {max_rentab:.2f}% ({fecha_max_rentab})",
             font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Rentab. promedio : {rentab_promedio:.2f}%", font=("Arial", 12)).pack(
        anchor="w")


# -------------------------
# Manejo de cierre
# -------------------------
def on_closing():
    ventana.quit()
    ventana.destroy()


ventana.protocol("WM_DELETE_WINDOW", on_closing)

try:


    ventana.mainloop()
except KeyboardInterrupt:
    pass
