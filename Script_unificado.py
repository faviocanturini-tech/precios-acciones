#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import tkinter as tk
from tkinter import filedialog, messagebox

# Valores por defecto para el límite
LIMITE_TIPO = "acciones"
LIMITE_VALOR = 10.0

# Columnas esperadas (exactas)
EXPECTED_COLUMNS = ["Fecha", "Último", "Apertura", "Máximo", "Mínimo", "Vol.", "% var."]


# =========================
# Funciones auxiliares
# =========================
def create_sqlite_from_df(folder, name, df):
    """Crea una base sqlite con la tabla 'precios' a partir del DataFrame."""
    db = os.path.join(folder, name)
    conn = sqlite3.connect(db)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS precios")
    cur.execute("""
        CREATE TABLE precios (
            Fecha TEXT,
            Ultimo REAL,
            Apertura REAL,
            Maximo REAL,
            Minimo REAL,
            Vol REAL,
            Var REAL
        )
    """)
    rows = []
    for _, r in df.iterrows():
        rows.append((
            r["Fecha"],
            float(r["Último"]) if pd.notna(r["Último"]) else None,
            float(r["Apertura"]) if pd.notna(r["Apertura"]) else None,
            float(r["Máximo"]) if pd.notna(r["Máximo"]) else None,
            float(r["Mínimo"]) if pd.notna(r["Mínimo"]) else None,
            float(r["Vol."]) if pd.notna(r["Vol."]) else None,
            float(r["% var."]) if pd.notna(r["% var."]) else None
        ))
    cur.executemany("INSERT INTO precios VALUES (?,?,?,?,?,?,?)", rows)
    conn.commit()
    conn.close()
    return db


def parse_percent_to_decimal(value):
    """
    Convierte distintos formatos de porcentaje a decimal:
    - "2.05%" -> 0.0205
    - "2.05"  -> 0.0205  (si se interpreta como porcentaje)
    - "0.0205" -> 0.0205 (ya está en forma decimal)
    Regla práctica:
      * Si contiene '%' -> quitar % y dividir por 100
      * Si el valor absoluto <= 1 -> se interpreta como decimal (p.ej. 0.0205)
      * Si el valor absoluto > 1 -> se interpreta como porcentaje y se divide por 100
    Devuelve float o NaN en caso de fallo.
    """
    try:
        if pd.isna(value):
            return float("nan")
        s = str(value).strip().replace('"', '').replace(",", ".")
        if s == "":
            return float("nan")
        if "%" in s:
            s2 = s.replace("%", "").strip()
            return float(s2) / 100.0
        f = float(s)
        if abs(f) <= 1:
            # ya es decimal
            return f
        else:
            # por ejemplo 2.05 -> 2.05% -> dividir por 100
            return f / 100.0
    except Exception:
        return float("nan")


def to_float_safe(x):
    """Convierte cadenas numéricas con coma/punto a float, devuelve NaN si falla."""
    try:
        if pd.isna(x):
            return float("nan")
        s = str(x).strip().replace('"', '').replace(",", ".")
        if s == "":
            return float("nan")
        return float(s)
    except:
        return float("nan")


# =========================
# Interfaz Gráfica
# =========================
ventana = tk.Tk()
ventana.grid_propagate(True)
ventana.title("Parámetros del análisis")
ventana.grid_columnconfigure(1, weight=1)
ventana.grid_columnconfigure(0, weight=0)
ventana.grid_columnconfigure(2, weight=0)
ventana.grid_columnconfigure(3, weight=1)

# ---- Widgets de entrada ----
tk.Label(ventana, text="Ruta del CSV (TAB):").grid(row=0, column=0, sticky="w")
entry_ruta = tk.Entry(ventana, width=55)
entry_ruta.grid(row=0, column=1, sticky="we")


def seleccionar_csv():
    ruta = filedialog.askopenfilename(
        title="Selecciona el archivo CSV (guardado desde Excel, separado por TAB)",
        filetypes=[("CSV files", "*.csv;*.txt"), ("Todos los archivos", "*.*")]
    )
    entry_ruta.delete(0, tk.END)
    entry_ruta.insert(0, ruta)


tk.Button(ventana, text="Seleccionar", command=seleccionar_csv).grid(row=0, column=2, sticky="w", padx=(6, 0))

# --- Compra / Venta / Suave ---
tk.Label(ventana, text="Compra (%):").grid(row=1, column=0, sticky="w")
entry_compra = tk.Entry(ventana, width=6)
entry_compra.insert(0, "-1.6")
entry_compra.grid(row=1, column=1, sticky="w")

tk.Label(ventana, text="Venta (%):").grid(row=2, column=0, sticky="w")
entry_venta = tk.Entry(ventana, width=6)
entry_venta.insert(0, "1.6")
entry_venta.grid(row=2, column=1, sticky="w")

tk.Label(ventana, text="Suave (%):").grid(row=3, column=0, sticky="w")
entry_suave = tk.Entry(ventana, width=6)
entry_suave.insert(0, "0.5")
entry_suave.grid(row=3, column=1, sticky="w")

# Selector límite
tipo_limite_var = tk.StringVar(value="acciones")
opciones_limite = ["acciones", "aporte"]
selector_limite = tk.OptionMenu(ventana, tipo_limite_var, *opciones_limite)
selector_limite.grid(row=4, column=0, sticky="w")

frame_limite = tk.Frame(ventana)
frame_limite.grid(row=4, column=1, sticky="w")
entry_limite = tk.Entry(frame_limite, width=10)
entry_limite.insert(0, "10")
entry_limite.pack(side="left")
tk.Label(frame_limite, text="Valor límite").pack(side="left", padx=(5, 0))

# Frame estadísticas y gráfico
ventana.frame_stats = tk.Frame(ventana, padx=10, pady=10)
ventana.frame_stats.grid(row=8, column=0, columnspan=3, sticky="w")
ventana.frame_grafico = tk.Frame(ventana, padx=5, pady=5)
ventana.frame_grafico.grid(row=9, column=0, columnspan=6, sticky="nsew")
ventana.grid_rowconfigure(9, weight=1)
ventana.grid_columnconfigure(0, weight=1)
frame_grafico = ventana.frame_grafico


# =========================
# Función principal
# =========================
def iniciar_proceso():
    global INPUT_FILE, UMBRAL_VENTA, UMBRAL_COMPRA, UMBRAL_SUAVE, FOLDER
    global LIMITE_TIPO, LIMITE_VALOR

    INPUT_FILE = entry_ruta.get().strip().strip('"')
    if not os.path.exists(INPUT_FILE):
        messagebox.showerror("Error", f"La ruta del CSV no existe:\n{INPUT_FILE}")
        return

    FOLDER = os.path.dirname(INPUT_FILE)
    base_name = os.path.splitext(os.path.basename(INPUT_FILE))[0]

    # Umbrales
    try:
        UMBRAL_VENTA = float(entry_venta.get().replace(",", ".")) / 100
        UMBRAL_COMPRA = float(entry_compra.get().replace(",", ".")) / 100
        UMBRAL_SUAVE = float(entry_suave.get().replace(",", ".")) / 100
    except:
        messagebox.showerror("Error", "Ingresa valores numéricos válidos para Compra/Venta/Suave.")
        return

    try:
        LIMITE_TIPO = tipo_limite_var.get()
        LIMITE_VALOR = float(entry_limite.get().replace(",", "."))
    except:
        messagebox.showerror("Error", "El valor del límite debe ser numérico.")
        return

    # -------------------------
    # -------------------------
    # Leer CSV preparado (TAB) y normalizar fechas mixtas
    # -------------------------
    try:
        df = pd.read_csv(INPUT_FILE, sep=";", engine='python', dtype=str)

        # Normalizar nombres columnas (strip)
        df.columns = [c.strip() for c in df.columns]

        # Limpiar espacios en la columna de fechas
        df['Fecha'] = df['Fecha'].astype(str).str.strip()

        # Función para convertir fechas mixtas
        def parse_mixed_dates(date_str):
            for fmt in ("%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return pd.to_datetime(date_str, format=fmt)
                except:
                    continue
            return pd.NaT

        # Aplicar la conversión
        df['Fecha'] = df['Fecha'].apply(parse_mixed_dates)

        # Eliminar filas donde la fecha no se pudo convertir
        df = df.dropna(subset=['Fecha'])

        # Ordenar por fecha ascendente
        df = df.sort_values('Fecha').reset_index(drop=True)

    except Exception as e:
        messagebox.showerror("Error al leer CSV", f"No se pudo leer el archivo.\n{e}")
        return

    # Verificar columnas esperadas
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    extra = [c for c in df.columns if c not in EXPECTED_COLUMNS]
    if missing:
        messagebox.showerror(
            "Columnas faltantes",
            f"Faltan las siguientes columnas esperadas en el CSV:\n{missing}\n\n"
            f"Columnas encontradas: {list(df.columns)}"
        )
        return

    # Reordenar/seleccionar exactamente las columnas esperadas
    df = df[EXPECTED_COLUMNS].copy()

    # -------------------------
    # Tipo de datos y transformaciones
    # -------------------------
    # Fecha -> datetime (día primero)
    try:
        df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
    except Exception:
        df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')

    if df['Fecha'].isna().any():
        # aviso pero no aborta: convertiré los NaT a cadena vacía para evitar errores
        # normalmente esto indica que el formato de fecha no es dd/mm/yyyy
        if messagebox.askyesno("Fechas no convertidas",
                               "Algunas filas no pudieron convertirse a fecha.\n¿Deseas continuar igual (esas filas quedarán con fecha vacía)?"):
            df['Fecha'] = df['Fecha'].dt.strftime("%d/%m/%Y")
            df['Fecha'] = df['Fecha'].fillna("")
        else:
            return
    else:
        df = df.sort_values("Fecha").reset_index(drop=True)
        df['Fecha'] = df['Fecha'].dt.strftime("%d/%m/%Y")

    # Columnas numéricas
    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(to_float_safe)

    # Columna % var. -> decimal (p.ej. 2.05% -> 0.0205)
    df['% var.'] = df['% var.'].apply(parse_percent_to_decimal)

    # -------------------------
    # Crear base de datos SQLite (opcional, pero la dejo)
    # -------------------------
    try:
        db_path = create_sqlite_from_df(FOLDER, f"{base_name}.db", df)
        print("Base SQLite generada:", db_path)
    except Exception as e:
        # no abortar; solo aviso
        print("Advertencia: no se pudo crear la DB:", e)

    # -------------------------
    # Guardar Excel analizado
    # -------------------------
    ruta_salida_excel = os.path.join(FOLDER, f"{base_name}_analizado.xlsx")
    try:
        df.to_excel(ruta_salida_excel, index=False)
        print("Excel analizado generado:", ruta_salida_excel)
    except PermissionError:
        messagebox.showerror(
            "Error al guardar Excel",
            f"El archivo {os.path.basename(ruta_salida_excel)} está abierto.\n"
            "Por favor, ciérralo e inicia el análisis nuevamente."
        )
        return

    # -------------------------
    # Continuar con procesamiento (adaptado del script original)
    # -------------------------
    def limpiar_num(x):
        try:
            return float(x)
        except:
            return 0.0

    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(lambda x: float(x) if pd.notna(x) else 0.0)

    # % var. está en decimal (p.ej. 0.0205) -> mantenemos como decimal para cálculos
    df['% var.'] = df['% var.'].apply(lambda x: float(x) if pd.notna(x) else 0.0)

    # Calcular % acumulado por signos consecutivos
    acum = 0
    prev = 0
    lst = []
    for v in df['% var.']:
        sign = 1 if v > 0 else -1 if v < 0 else 0
        if sign == prev:
            acum += v
        else:
            acum = v
        lst.append(acum)
        prev = sign
    df['% acumulado'] = lst

    def determinar_opcion(v, a):
        if v >= UMBRAL_VENTA:
            return "Venta"
        if v <= UMBRAL_COMPRA:
            return "Compra"
        if a >= UMBRAL_VENTA and v >= UMBRAL_SUAVE:
            return "Venta"
        if a <= UMBRAL_COMPRA and v <= -UMBRAL_SUAVE:
            return "Compra"
        return "N/A"

    df['Opción'] = df.apply(lambda r: determinar_opcion(r['% var.'], r['% acumulado']), axis=1)

    # Límite tipo
    try:
        if LIMITE_TIPO == "acciones":
            MAX_ACCIONES = int(LIMITE_VALOR)
            MAX_APORTE = float("inf")
        else:
            MAX_ACCIONES = 10
            MAX_APORTE = float(LIMITE_VALOR)
    except:
        MAX_ACCIONES = 10
        MAX_APORTE = float("inf")

    acciones = 0
    capital_bolsa = 0
    aporte_acumulado = 0

    movs, acts, cap_b, cap_acc, cap_tot, aport, aport_acum = [], [], [], [], [], [], []

    for _, row in df.iterrows():
        opcion = row["Opción"]
        precio = row["Último"]
        movimiento = 0
        aporte = 0.0

        if opcion == "Compra":
            puede_comprar = False
            if LIMITE_TIPO == "acciones" and acciones < MAX_ACCIONES:
                puede_comprar = True
            elif LIMITE_TIPO == "aporte" and (aporte_acumulado + precio) <= MAX_APORTE:
                puede_comprar = True

            if puede_comprar:
                if capital_bolsa >= precio:
                    capital_bolsa -= precio
                else:
                    aporte = precio
                    aporte_acumulado += aporte
                    capital_bolsa += aporte
                    capital_bolsa -= precio
                acciones += 1
                movimiento = 1
        elif opcion == "Venta" and acciones > 0:
            capital_bolsa += precio
            acciones -= 1
            movimiento = -1

        movs.append(movimiento)
        acts.append(acciones)
        cap_b.append(round(capital_bolsa, 2))
        cap_acc.append(round(acciones * precio, 2))
        cap_tot.append(round(capital_bolsa + acciones * precio, 2))
        aport.append(round(aporte, 2))
        aport_acum.append(round(aporte_acumulado, 2))

    df["Movimiento de acciones"] = movs
    df["Acciones en cartera"] = acts
    df["Capital en bolsa"] = cap_b
    df["Capital en acciones"] = cap_acc
    df["Capital total"] = cap_tot
    df["Aporte"] = aport
    df["Aporte acumulado"] = aport_acum
    df["Margen"] = df["Capital total"] - df["Aporte acumulado"]
    df["Rentabilidad"] = df.apply(
        lambda r: (r["Margen"] / r["Aporte acumulado"] * 100) if r["Aporte acumulado"] > 0 else 0, axis=1)

    df["Rentabilidad"] = df["Rentabilidad"].round(2).astype(str) + "%"
    df["% var."] = (df["% var."] * 100).round(2).astype(str) + "%"
    df["% acumulado"] = (df["% acumulado"] * 100).round(2).astype(str) + "%"

    # Guardar Excel actualizado
    try:
        df.to_excel(ruta_salida_excel, index=False)
    except PermissionError:
        messagebox.showerror(
            "Error al guardar Excel",
            f"El archivo {os.path.basename(ruta_salida_excel)} está abierto.\n"
            "Por favor, ciérralo e intenta nuevamente."
        )
        return

    # =========================
    # Mostrar estadísticas en la ventana principal
    # =========================
    for widget in ventana.frame_stats.winfo_children():
        widget.destroy()

    frame1 = tk.Frame(ventana.frame_stats, padx=15)
    frame1.grid(row=0, column=0, sticky="nw")
    frame2 = tk.Frame(ventana.frame_stats, padx=15)
    frame2.grid(row=0, column=1, sticky="nw")
    frame3 = tk.Frame(ventana.frame_stats, padx=15)
    frame3.grid(row=0, column=2, sticky="nw")
    frame4 = tk.Frame(ventana.frame_stats, padx=15)
    frame4.grid(row=0, column=3, sticky="nw")

    def float_col(col):
        return df[col].astype(str).str.rstrip('%').replace('', '0').astype(float)

    max_var = float_col('% var.').max()
    min_var = float_col('% var.').min()
    fecha_max_var = df.loc[float_col('% var.').idxmax(), 'Fecha']
    fecha_min_var = df.loc[float_col('% var.').idxmin(), 'Fecha']
    dif_var = max_var - min_var

    prom_var = float_col('% var.')
    subidas = prom_var[prom_var > 0]
    max_prom = subidas.mean() if not subidas.empty else 0
    bajadas = prom_var[prom_var < 0]
    min_prom = bajadas.mean() if not bajadas.empty else 0
    dif_prom = max_prom - min_prom

    opc_compra = (df["Opción"] == "Compra").sum()
    acciones_compradas = df.loc[df["Movimiento de acciones"] > 0, "Movimiento de acciones"].sum()
    opc_venta = (df["Opción"] == "Venta").sum()
    acciones_vendidas = -df.loc[df["Movimiento de acciones"] < 0, "Movimiento de acciones"].sum()
    max_acc_cartera = df["Acciones en cartera"].max()
    max_aporte = df["Aporte acumulado"].max()
    max_margen = round(df["Margen"].max(), 2)
    max_rentab = float_col("Rentabilidad").max()
    fecha_max_rentab = df.loc[float_col("Rentabilidad").idxmax(), "Fecha"]

    tk.Label(frame1, fg="blue", text=f"Max % var : {max_var:.2f}% ({fecha_max_var})", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame1, fg="blue", text=f"Min % var : {min_var:.2f}% ({fecha_min_var})", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame1, fg="blue", text=f"Diferencia : {dif_var:.2f}%", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame2, fg="red", text=f"Prom % var + :  {max_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame2, fg="red", text=f"Prom % var - : {min_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame2, fg="red", text=f"Diferencia       :  {dif_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame3, fg="black", text=f"Opciones Compra       : {opc_compra}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Acciones Compradas : {acciones_compradas}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Opciones Venta           : {opc_venta}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Acciones Vendidas      : {acciones_vendidas}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame3, fg="black", text=f"Máx acción en cartera : {max_acc_cartera}", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame4, fg="purple", text=f"Aporte acum max : {max_aporte:,.0f}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Margen max    : {max_margen:,.2f}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Rentab. max    : {max_rentab:.2f}% ({fecha_max_rentab})",
             font=("Arial", 12)).pack(anchor="w")

    # -------------------------
    # Paso 3: Generar gráfico y guardarlo en el Excel
    # -------------------------
    data = df.copy()
    data['Fecha'] = pd.to_datetime(data['Fecha'], format='%d/%m/%Y')
    # Rentabilidad ya está como "12.34%" -> convertir a float
    data['Rentabilidad'] = data['Rentabilidad'].astype(str).str.rstrip('%').replace('', '0').astype(float)

    fig, ax1 = plt.subplots(figsize=(16, 10))
    plt.subplots_adjust(left=0.04, right=0.9, top=0.92, bottom=0.1)

    # Eje Y independiente para Margen (color verde)
    ax_margen = ax1.twinx()
    ax_margen.spines['right'].set_position(('outward', 0))

    # Eje Y para Rentabilidad (color rojo)
    ax2 = ax1.twinx()
    ax2.spines['right'].set_position(('outward', 40))

    # Eje Y para Acciones en cartera (color negro)
    ax3 = ax1.twinx()
    ax3.spines['right'].set_position(('outward', 80))

    # Ajustar la posición de los ejes para que todo quepa
    pos = [0.00, 0.18, 0.73, 0.75]  # [left, bottom, width, height]
    ax1.set_position(pos)
    ax_margen.set_position(pos)
    ax2.set_position(pos)
    ax3.set_position(pos)

    # ---- Graficar series con colores explícitos (restaurados) ----
    ax1.plot(data['Fecha'], data['Último'], color='blue', label='Último', linewidth=2)
    ax_margen.plot(data['Fecha'], data['Margen'], color='green', label='Margen', linewidth=2)
    # Escalar Rentabilidad para que no se solape (cálculo seguro)
    max_r = data['Rentabilidad'].abs().replace(0, 1).max()
    max_u = data['Último'].abs().replace(0, 1).max()
    factor_renta = max_u / (max_r if max_r != 0 else 1)
    ax2.plot(data['Fecha'], data['Rentabilidad'] * factor_renta, color='red', label='Rentabilidad (escalada)',
             linestyle='--', linewidth=2)
    ax3.plot(data['Fecha'], data['Acciones en cartera'], color='black', label='Acciones en cartera', linestyle=':',
             linewidth=2)

    # Etiquetas y colores para cada eje (coinciden con las líneas)
    ax1.set_ylabel('Último', color='blue')
    ax1.tick_params(axis='y', labelcolor='blue')

    ax_margen.set_ylabel('Margen', color='green')
    ax_margen.tick_params(axis='y', labelcolor='green')

    ax2.set_ylabel('Rentabilidad (%)', color='red')
    ax2.tick_params(axis='y', labelcolor='red')

    ax3.set_ylabel('Acciones en cartera', color='black')
    ax3.tick_params(axis='y', labelcolor='black')

    # Formatear valores sin decimales en los ejes secundarios
    ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    ax_margen.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    ax3.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))

    # Autoscale con factor y offset personalizado para evitar traslapos
    for ax, factor, offset in [
        (ax1, 1.05, 0),  # Último
        (ax_margen, 1.2, 200),  # Margen: 20% más y offset para separarlo
        (ax2, 1.5, 0),  # Rentabilidad
        (ax3, 1.05, 0)  # Acciones en cartera
    ]:
        ax.relim()
        ax.autoscale_view()
        ymin, ymax = ax.get_ylim()
        ax.set_ylim(ymin, ymax * factor + offset)

    # Leyenda combinada (todas las líneas)
    lines_1, labels_1 = ax1.get_legend_handles_labels()
    lines_2, labels_2 = ax_margen.get_legend_handles_labels()
    lines_3, labels_3 = ax2.get_legend_handles_labels()
    lines_4, labels_4 = ax3.get_legend_handles_labels()
    ax1.legend(lines_1 + lines_2 + lines_3 + lines_4,
               labels_1 + labels_2 + labels_3 + labels_4,
               loc='upper left')

    fig.autofmt_xdate()
    plt.title('Análisis de Último, Margen, Rentabilidad y Acciones en cartera')

    # -------------------------
    # Mostrar gráfico en la interfaz
    # -------------------------
    for widget in frame_grafico.winfo_children():
        widget.destroy()

    canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True, pady=10)

    # Reducir ventana al 80% del tamaño de la pantalla después de mostrar el gráfico
    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()

    win_width = int(screen_width * 0.8)
    win_height = int(screen_height * 0.8)

    x = int((screen_width - win_width) / 2)
    y = int((screen_height - win_height) / 2)

    ventana.geometry(f"{win_width}x{win_height}+{x}+{y}")

    # Guardar imagen y cerrar figura
    img_path = os.path.join(FOLDER, f"Grafico_{base_name}.png")
    fig.savefig(img_path, dpi=150, bbox_inches='tight')
    plt.close(fig)

    try:
        wb = load_workbook(ruta_salida_excel)
        ws = wb.create_sheet("Gráfico")
        img = XLImage(img_path)
        img.anchor = "A1"
        ws.add_image(img)
        wb.save(ruta_salida_excel)
        print("Gráfico insertado en la hoja 'Gráfico' del Excel final.")
    except Exception as e:
        print("Advertencia: no se pudo insertar el gráfico en el Excel:", e)


# -------------------------
# Botón iniciar análisis
# -------------------------
tk.Button(ventana, text="Iniciar análisis", command=iniciar_proceso).grid(row=1, column=2, sticky="w", pady=10)

ventana.mainloop()

