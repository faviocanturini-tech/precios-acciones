import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
from zoneinfo import ZoneInfo
import os
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import gc
import json
from pathlib import Path
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates

# Lista de tickers
tickers = ["AAPL","AMZN","AVGO","BRK-B","GLD","META","MSFT","NVDA","PLTR","QQQ","SPY","TSLA"]

# Archivo de configuración (compartido con Analisis_singrafico.py)
CONFIG_FILE = Path.home() / ".analisis_config.json"

# =====================================================
# CONFIGURACIÓN PORTABLE
# =====================================================
def obtener_ruta_base():
    """Obtiene la ruta base del script"""
    return Path(__file__).parent

def obtener_carpeta_datos():
    """Obtiene la carpeta de datos (data/)"""
    ruta_base = obtener_ruta_base()
    carpeta_data = ruta_base / "data"
    if not carpeta_data.exists():
        carpeta_data.mkdir(parents=True, exist_ok=True)
    return carpeta_data

CARPETA_DATOS_PORTABLE = obtener_carpeta_datos()
DATOS_CSV_PORTABLE = CARPETA_DATOS_PORTABLE / "datos_1dia_crudos.csv"
AUTO_UPDATE_LOG_PORTABLE = CARPETA_DATOS_PORTABLE / "auto_update_log.csv"
BACKUPS_FOLDER = CARPETA_DATOS_PORTABLE / "backups"


def crear_backup_datos(motivo="manual"):
    """
    Crea un backup de todos los archivos críticos en data/backups/
    Se debe llamar ANTES de cualquier operación que modifique datos.
    """
    import shutil

    try:
        BACKUPS_FOLDER.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_folder = BACKUPS_FOLDER / f"{timestamp}_{motivo}"
        backup_folder.mkdir(parents=True, exist_ok=True)

        archivos_criticos = [
            "auto_update_log.csv",
            "datos_1dia_crudos.csv",
            "parametros_activos.json",
            "historial_senales.json",
            "Resultado_de_Analisis.json",
            "tickers_descarga.json",
            "historial_operaciones.json"
        ]

        archivos_respaldados = 0
        for archivo in archivos_criticos:
            origen = CARPETA_DATOS_PORTABLE / archivo
            if origen.exists():
                destino = backup_folder / archivo
                shutil.copy2(origen, destino)
                archivos_respaldados += 1
                print(f"[Backup] {archivo} -> {backup_folder.name}/")

        print(f"[Backup] Completado: {archivos_respaldados} archivos en {backup_folder}")

        # Limpiar backups antiguos (mantener solo los últimos 10)
        limpiar_backups_antiguos(max_backups=10)

        return str(backup_folder)

    except Exception as e:
        print(f"[Backup] ERROR: {e}")
        return None


def limpiar_backups_antiguos(max_backups=10):
    """Elimina backups antiguos, manteniendo solo los más recientes"""
    import shutil

    try:
        if not BACKUPS_FOLDER.exists():
            return

        backups = sorted([d for d in BACKUPS_FOLDER.iterdir() if d.is_dir()])

        while len(backups) > max_backups:
            backup_antiguo = backups.pop(0)
            shutil.rmtree(backup_antiguo)
            print(f"[Backup] Eliminado backup antiguo: {backup_antiguo.name}")

    except Exception as e:
        print(f"[Backup] Error limpiando backups antiguos: {e}")


def siguiente_dia_trading(fecha):
    """
    Calcula el siguiente día de trading después de la fecha dada.
    Salta fines de semana y feriados principales de USA.

    Args:
        fecha: datetime o date object

    Returns:
        datetime.date del siguiente día de trading
    """
    from datetime import timedelta

    # Convertir a date si es datetime
    if hasattr(fecha, 'date'):
        fecha = fecha.date()

    # Feriados principales de USA 2025-2026 (mercado cerrado)
    feriados_usa = {
        # 2025
        datetime(2025, 1, 1).date(),   # New Year's Day
        datetime(2025, 1, 20).date(),  # MLK Day
        datetime(2025, 2, 17).date(),  # Presidents Day
        datetime(2025, 4, 18).date(),  # Good Friday
        datetime(2025, 5, 26).date(),  # Memorial Day
        datetime(2025, 6, 19).date(),  # Juneteenth
        datetime(2025, 7, 4).date(),   # Independence Day
        datetime(2025, 9, 1).date(),   # Labor Day
        datetime(2025, 11, 27).date(), # Thanksgiving
        datetime(2025, 12, 25).date(), # Christmas
        # 2026
        datetime(2026, 1, 1).date(),   # New Year's Day
        datetime(2026, 1, 19).date(),  # MLK Day
        datetime(2026, 2, 16).date(),  # Presidents Day
        datetime(2026, 4, 3).date(),   # Good Friday
        datetime(2026, 5, 25).date(),  # Memorial Day
        datetime(2026, 6, 19).date(),  # Juneteenth
        datetime(2026, 7, 3).date(),   # Independence Day (observed)
        datetime(2026, 9, 7).date(),   # Labor Day
        datetime(2026, 11, 26).date(), # Thanksgiving
        datetime(2026, 12, 25).date(), # Christmas
    }

    siguiente = fecha + timedelta(days=1)

    # Avanzar hasta encontrar un día de trading válido
    while siguiente.weekday() >= 5 or siguiente in feriados_usa:  # 5=sábado, 6=domingo
        siguiente += timedelta(days=1)

    return siguiente


# ===== FUNCIONES DE MIGRACIÓN Y SLOTS v2.0 =====

def crear_estructura_slots_vacia():
    """Crea una estructura de slots vacía con valores por defecto"""
    return {
        "version": "2.0",
        "slots": {
            "1": {"nombre": "1", "parametros_activos": []},
            "2": {"nombre": "2", "parametros_activos": []},
            "3": {"nombre": "3", "parametros_activos": []},
            "4": {"nombre": "4", "parametros_activos": []},
            "5": {"nombre": "5", "parametros_activos": []}
        }
    }


def migrar_parametros_v1_a_v2(datos_v1):
    """Migra formato antiguo (lista) a nuevo formato (slots)"""
    parametros_lista = datos_v1.get("parametros_activos", [])
    estructura = crear_estructura_slots_vacia()
    estructura["slots"]["1"]["parametros_activos"] = parametros_lista
    return estructura


def obtener_parametros_slot(datos_slots, slot_id):
    """Obtiene la lista de parámetros de un slot específico"""
    return datos_slots.get("slots", {}).get(slot_id, {}).get("parametros_activos", [])


def obtener_nombre_slot(datos_slots, slot_id):
    """Obtiene el nombre de un slot"""
    return datos_slots.get("slots", {}).get(slot_id, {}).get("nombre", slot_id)


def crear_estructura_senales_vacia():
    """Crea una estructura de señales vacía con slots"""
    return {
        "version": "2.0",
        "senales_por_slot": {
            "1": [],
            "2": [],
            "3": [],
            "4": [],
            "5": []
        }
    }


def cargar_parametros_activos():
    """Carga los parámetros activos con estructura de slots v2.0"""
    # Primero obtener la ubicación del JSON desde la config
    if not CONFIG_FILE.exists():
        return None, "No se encontró configuración. Ejecuta primero Analisis_singrafico.py"

    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
            ubicacion = config.get("ubicacion_json")

        if not ubicacion:
            return None, "No hay ubicación JSON configurada"

        archivo_params = Path(ubicacion) / "parametros_activos.json"

        if not archivo_params.exists():
            return None, f"No existe el archivo:\n{archivo_params}\n\nConfigura los parámetros activos primero."

        with open(archivo_params, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Detectar versión del formato
        if "version" in datos and datos.get("version") == "2.0":
            # Ya es v2.0 - verificar que hay al menos un parámetro en algún slot
            hay_parametros = any(
                len(datos.get("slots", {}).get(slot_id, {}).get("parametros_activos", [])) > 0
                for slot_id in ["1", "2", "3", "4", "5"]
            )
            if not hay_parametros:
                return None, "No hay parámetros activos configurados en ningún slot"
            return datos, None
        else:
            # Formato antiguo v1.0 - migrar
            parametros = datos.get("parametros_activos", [])
            if not parametros:
                return None, "No hay parámetros activos configurados"

            # Migrar a v2.0
            datos_v2 = migrar_parametros_v1_a_v2(datos)

            # Guardar el formato migrado
            with open(archivo_params, 'w', encoding='utf-8') as f:
                json.dump(datos_v2, f, indent=2, ensure_ascii=False)

            return datos_v2, None

    except Exception as e:
        return None, f"Error cargando parámetros: {e}"


def obtener_ruta_historial():
    """Obtiene la ruta del archivo de historial de operaciones"""
    if not CONFIG_FILE.exists():
        return None

    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
            ubicacion = config.get("ubicacion_json")

        if ubicacion:
            return Path(ubicacion) / "historial_operaciones.json"
    except:
        pass
    return None


def cargar_historial_operaciones():
    """Carga el historial de operaciones confirmadas"""
    ruta = obtener_ruta_historial()
    if ruta is None or not ruta.exists():
        return []

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)
            return datos.get("operaciones", [])
    except Exception as e:
        print(f"[ERROR] Error cargando historial: {e}")
        return []


def guardar_historial_operaciones(operaciones):
    """Guarda el historial de operaciones"""
    ruta = obtener_ruta_historial()
    if ruta is None:
        messagebox.showerror("Error", "No hay ubicación configurada para guardar el historial.")
        return False

    try:
        datos = {"operaciones": operaciones}
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Error guardando historial:\n{e}")
        return False


def obtener_ruta_senales():
    """Obtiene la ruta del archivo de historial de señales"""
    if not CONFIG_FILE.exists():
        return None

    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
            ubicacion = config.get("ubicacion_json")

        if ubicacion:
            return Path(ubicacion) / "historial_senales.json"
    except:
        pass
    return None


def guardar_ruta_csv(ruta_csv):
    """Guarda la última ruta del CSV en la configuración"""
    try:
        config = {}
        if CONFIG_FILE.exists():
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)

        config["ultima_ruta_csv"] = ruta_csv

        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"[WARN] No se pudo guardar ruta CSV: {e}")


def cargar_ruta_csv():
    """Carga la última ruta del CSV desde la configuración"""
    if not CONFIG_FILE.exists():
        return None

    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
            return config.get("ultima_ruta_csv")
    except:
        pass
    return None


def sincronizar_desde_github():
    """
    Sincroniza datos desde GitHub siguiendo el flujo normal:
    0. BACKUP automático antes de cualquier cambio
    1. Descarga datos de GitHub
    2. Filtra solo los que NO existen en auto_update_log.csv local
    3. Guarda los nuevos en datos_1dia_crudos.csv
    4. Muestra en tabla
    5. Merge a auto_update_log.csv
    """
    import subprocess
    import io

    # *** BACKUP AUTOMÁTICO ANTES DE SINCRONIZAR ***
    backup_path = crear_backup_datos("antes_sync_github")
    if backup_path:
        print(f"[Sync] Backup creado: {backup_path}")

    # Rutas (portable)
    repo_path = str(obtener_ruta_base())
    log_file = str(AUTO_UPDATE_LOG_PORTABLE)
    csv_file = str(DATOS_CSV_PORTABLE)

    try:
        # 0. Verificar si es un repositorio git
        check_git = subprocess.run(
            ["git", "rev-parse", "--is-inside-work-tree"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=10
        )
        if check_git.returncode != 0:
            messagebox.showinfo(
                "Sync GitHub",
                "Esta funcion no esta disponible en la version portable.\n\n"
                "Para sincronizar datos desde GitHub, usa la carpeta TRADING\n"
                "que contiene el repositorio git."
            )
            return False

        # 1. Leer datos locales del log histórico
        df_local = None
        local_keys = set()
        if os.path.exists(log_file):
            df_local = pd.read_csv(log_file, parse_dates=['Date'])
            df_local = df_local.loc[:, ~df_local.columns.duplicated()]
            df_local['Date'] = pd.to_datetime(df_local['Date']).dt.normalize()
            local_keys = set(zip(
                df_local['Date'].dt.strftime('%Y-%m-%d'),
                df_local['Ticker']
            ))
            print(f"[Sync] Datos locales en log: {len(df_local)} registros")

        # 2. Hacer git fetch para actualizar referencias
        print("[Sync] Conectando a GitHub...")
        result = subprocess.run(
            ["git", "fetch", "origin", "main"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0:
            messagebox.showerror("Error", f"Error en fetch:\n{result.stderr}")
            return False

        # 3. Obtener el archivo desde GitHub
        result = subprocess.run(
            ["git", "show", "origin/main:data/auto_update_log.csv"],
            cwd=repo_path,
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0 or not result.stdout.strip():
            messagebox.showerror("Error", f"No se pudo obtener datos de GitHub:\n{result.stderr}")
            return False

        df_github = pd.read_csv(io.StringIO(result.stdout), parse_dates=['Date'])
        df_github = df_github.loc[:, ~df_github.columns.duplicated()]
        df_github['Date'] = pd.to_datetime(df_github['Date']).dt.normalize()
        print(f"[Sync] Datos en GitHub: {len(df_github)} registros")

        # 4. Filtrar solo registros que NO existen localmente
        github_keys = df_github[['Date', 'Ticker']].apply(
            lambda r: (r['Date'].strftime('%Y-%m-%d'), r['Ticker']), axis=1
        )
        mask_nuevos = ~github_keys.isin(local_keys)
        df_nuevos = df_github.loc[mask_nuevos].copy()

        if df_nuevos.empty:
            # Aunque no hay datos nuevos, mostrar el último día disponible en tabla
            if df_local is not None and not df_local.empty:
                ultima_fecha = df_local['Date'].max()
                df_ultimo_dia = df_local[df_local['Date'] == ultima_fecha].copy()
                # Guardar en csv temporal para mostrar
                temp_csv = str(DATOS_CSV_PORTABLE)
                df_ultimo_dia.to_csv(temp_csv, index=False, float_format="%.2f")
                mostrar_datos_en_tabla(temp_csv)
                messagebox.showinfo("Sincronización",
                    f"Ya tienes los datos más recientes.\n\n"
                    f"Última fecha: {ultima_fecha.strftime('%Y-%m-%d')}\n"
                    f"Registros: {len(df_ultimo_dia)}")
            else:
                messagebox.showinfo("Sincronización", "Ya tienes los datos más recientes.")
            return True

        print(f"[Sync] Registros nuevos encontrados: {len(df_nuevos)}")

        # 5. Guardar nuevos en datos_1dia_crudos.csv (como el flujo manual)
        df_nuevos.to_csv(csv_file, index=False, float_format="%.2f")
        print(f"[Sync] Guardado en {csv_file}")

        # 6. Mostrar en tabla (no cambiar entry_ruta)
        mostrar_datos_en_tabla(csv_file)

        # 7. Merge a auto_update_log.csv (flujo normal)
        if df_local is not None:
            df_combined = pd.concat([df_local, df_nuevos], ignore_index=True)
            df_combined = df_combined.sort_values(['Date', 'Ticker']).reset_index(drop=True)
        else:
            df_combined = df_nuevos

        df_combined.to_csv(log_file, index=False, float_format="%.2f")
        print(f"[Sync] Log actualizado: {len(df_combined)} registros totales")

        # 8. Verificar si hay valores NaN en los datos
        df_con_nan = df_combined[df_combined['Close'].isna()]
        if not df_con_nan.empty:
            # Construir lista de tickers/fechas con NaN
            nan_info = []
            nan_tickers_fechas = []  # Lista de (ticker, fecha) para re-descarga
            for _, row in df_con_nan.iterrows():
                fecha_str = row['Date'].strftime('%Y-%m-%d') if pd.notna(row['Date']) else 'N/A'
                nan_info.append(f"  - {row['Ticker']} ({fecha_str})")
                if pd.notna(row['Date']):
                    nan_tickers_fechas.append((row['Ticker'], row['Date']))
            nan_list = "\n".join(nan_info[:10])  # Mostrar máximo 10
            if len(nan_info) > 10:
                nan_list += f"\n  ... y {len(nan_info) - 10} más"

            # Preguntar si quiere re-descargar automáticamente
            respuesta = messagebox.askyesno("Datos incompletos",
                f"Se detectaron valores vacíos (NaN) en los siguientes tickers:\n\n"
                f"{nan_list}\n\n"
                f"¿Desea re-descargar automáticamente los datos faltantes desde Yahoo Finance?")

            if respuesta and nan_tickers_fechas:
                # Re-descargar datos faltantes
                import yfinance as yf
                corregidos = 0
                errores = []

                for ticker, fecha in nan_tickers_fechas:
                    try:
                        fecha_inicio = fecha.strftime('%Y-%m-%d')
                        fecha_fin = (fecha + pd.Timedelta(days=1)).strftime('%Y-%m-%d')

                        yf_ticker = yf.Ticker(ticker)
                        hist = yf_ticker.history(start=fecha_inicio, end=fecha_fin)

                        if not hist.empty:
                            # Actualizar valores en df_combined
                            mask = (df_combined['Date'] == fecha) & (df_combined['Ticker'] == ticker)
                            df_combined.loc[mask, 'Open'] = hist.iloc[0]['Open']
                            df_combined.loc[mask, 'High'] = hist.iloc[0]['High']
                            df_combined.loc[mask, 'Low'] = hist.iloc[0]['Low']
                            df_combined.loc[mask, 'Close'] = hist.iloc[0]['Close']
                            corregidos += 1
                            print(f"[Sync] Corregido: {ticker} ({fecha_inicio}) - Close: {hist.iloc[0]['Close']:.2f}")
                        else:
                            errores.append(f"{ticker} ({fecha_inicio})")
                    except Exception as e:
                        errores.append(f"{ticker} ({fecha_inicio}): {str(e)[:30]}")

                # Guardar CSV actualizado
                df_combined.to_csv(log_file, index=False, float_format="%.2f")

                # Mostrar resultado
                if corregidos > 0 and not errores:
                    messagebox.showinfo("Re-descarga completada",
                        f"Se corrigieron {corregidos} registros exitosamente.\n\n"
                        f"Total en log: {len(df_combined)}")
                elif corregidos > 0 and errores:
                    messagebox.showwarning("Re-descarga parcial",
                        f"Corregidos: {corregidos}\n"
                        f"Errores: {len(errores)}\n\n"
                        f"No se pudieron obtener:\n" + "\n".join(errores[:5]))
                else:
                    messagebox.showerror("Error en re-descarga",
                        f"No se pudieron obtener los datos.\n\n"
                        f"Prueba con 'Descargar Precios' manualmente.")
            elif not respuesta:
                messagebox.showinfo("Sincronización",
                    f"Sincronización completada con advertencias.\n\n"
                    f"Registros nuevos: {len(df_nuevos)}\n"
                    f"Registros con NaN: {len(nan_info)}\n\n"
                    f"Usa 'Descargar Precios' para corregir manualmente.")
        else:
            messagebox.showinfo("Sincronización",
                f"Sincronización completada.\n\n"
                f"Registros nuevos: {len(df_nuevos)}\n"
                f"Total en log: {len(df_combined)}")

        return True

    except subprocess.TimeoutExpired:
        messagebox.showerror("Error", "Timeout: La sincronización tardó demasiado.")
        return False
    except FileNotFoundError:
        messagebox.showerror("Error", "Git no está instalado o no se encuentra en el PATH.")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error inesperado:\n{e}")
        return False


def cargar_historial_senales():
    """Carga el historial de señales generadas (estructura con slots v2.0)"""
    ruta = obtener_ruta_senales()
    if ruta is None or not ruta.exists():
        return crear_estructura_senales_vacia()

    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Detectar versión del formato
        if "version" in datos and datos.get("version") == "2.0":
            return datos
        else:
            # Formato antiguo - migrar a v2
            senales_antiguas = datos.get("senales", [])
            estructura = crear_estructura_senales_vacia()
            estructura["senales_por_slot"]["1"] = senales_antiguas
            # Guardar migrado
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(estructura, f, indent=2, ensure_ascii=False)
            return estructura
    except Exception as e:
        print(f"[ERROR] Error cargando historial de señales: {e}")
        return crear_estructura_senales_vacia()


def cargar_senales_slot(slot_id):
    """Carga las señales de un slot específico"""
    datos = cargar_historial_senales()
    return datos.get("senales_por_slot", {}).get(slot_id, [])


def guardar_historial_senales(senales_nuevas, slot_id="1", slot_nombre="1", fecha_override=None):
    """Guarda las señales generadas en el historial para un slot específico

    Args:
        senales_nuevas: Lista de señales a guardar
        slot_id: ID del slot
        slot_nombre: Nombre del slot
        fecha_override: Fecha opcional para señales históricas (formato YYYY-MM-DD HH:MM:SS)
    """
    ruta = obtener_ruta_senales()
    if ruta is None:
        print("[WARN] No hay ubicación configurada para guardar señales.")
        return False

    try:
        # Cargar estructura de señales existente
        datos_senales = cargar_historial_senales()

        # Obtener señales existentes del slot
        senales_existentes = datos_senales.get("senales_por_slot", {}).get(slot_id, [])

        # Usar fecha override si se proporciona, sino usar ahora
        if fecha_override:
            fecha_generacion = fecha_override
        else:
            fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        fecha_hoy = fecha_generacion[:10]

        # Para señales históricas, eliminar señales existentes de esa fecha en este slot
        if fecha_override:
            senales_existentes = [sen for sen in senales_existentes
                                  if sen.get("fecha_generacion", "")[:10] != fecha_hoy]

        # Crear conjunto de señales existentes para verificar duplicados
        senales_existentes_keys = set()
        for sen in senales_existentes:
            fecha_sen = sen.get("fecha_generacion", "")[:10]
            symbol_sen = sen.get("symbol", "")
            senales_existentes_keys.add((fecha_sen, symbol_sen))

        senales_agregadas = 0

        for senal in senales_nuevas:
            if senal.get('estado') == 'OK':
                symbol = senal.get('symbol')

                if (fecha_hoy, symbol) in senales_existentes_keys:
                    print(f"[INFO] Señal duplicada ignorada: {symbol} ({fecha_hoy}) en slot {slot_nombre}")
                    continue

                nueva_senal = {
                    "fecha_generacion": fecha_generacion,
                    "symbol": symbol,
                    "precio_cierre": senal.get('cierre'),
                    "precio_compra_sugerido": senal.get('precio_compra'),
                    "cant_compra": senal.get('cant_compra'),
                    "opc_compra": senal.get('opc_compra'),
                    "precio_venta_sugerido": senal.get('precio_venta'),
                    "cant_venta": senal.get('cant_venta'),
                    "opc_venta": senal.get('opc_venta'),
                    "acciones_cartera": senal.get('acciones_cartera'),
                    "limite_tipo": senal.get('limite_tipo', 'acciones'),
                    "limite_valor": senal.get('limite_valor', 10),
                    "slot_nombre": slot_nombre,
                    "tendencia": senal.get('tendencia', 'N/A'),
                    "tendencia_larga": senal.get('tendencia_larga', 'N/A')
                }
                senales_existentes.append(nueva_senal)
                senales_existentes_keys.add((fecha_hoy, symbol))
                senales_agregadas += 1

        # Actualizar las señales del slot en la estructura
        datos_senales["senales_por_slot"][slot_id] = senales_existentes

        # Guardar
        with open(ruta, 'w', encoding='utf-8') as f:
            json.dump(datos_senales, f, indent=2, ensure_ascii=False)

        print(f"[INFO] Slot {slot_nombre}: {senales_agregadas} señales nuevas guardadas")
        return True

    except Exception as e:
        print(f"[ERROR] Error guardando señales: {e}")
        return False


def calcular_tendencia(df_precios, ticker, dias=10):
    """
    Calcula la tendencia de un ticker usando regresión lineal.
    Retorna un string con formato "+XX" o "-XX" donde XX es el nivel de tendencia (0-100).
    El signo indica dirección (+ alcista, - bajista) y el número indica fuerza (R²).
    """
    try:
        df_ticker = df_precios[df_precios['Ticker'] == ticker].copy()
        if len(df_ticker) < 5:
            return "N/A"

        df_ticker = df_ticker.sort_values('Date').tail(dias)
        if len(df_ticker) < 5:
            return "N/A"

        precios = df_ticker['Close'].values
        x = np.arange(len(precios))

        n = len(x)
        sum_x = np.sum(x)
        sum_y = np.sum(precios)
        sum_xy = np.sum(x * precios)
        sum_x2 = np.sum(x ** 2)

        pendiente = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x ** 2)
        intercepto = (sum_y - pendiente * sum_x) / n
        y_pred = pendiente * x + intercepto

        ss_res = np.sum((precios - y_pred) ** 2)
        ss_tot = np.sum((precios - np.mean(precios)) ** 2)
        if ss_tot == 0:
            r2 = 0
        else:
            r2 = 1 - (ss_res / ss_tot)

        signo = "+" if pendiente > 0 else "-"
        nivel = int(round(abs(r2) * 100, -1))
        nivel = min(100, max(0, nivel))

        return f"{signo}{nivel}"
    except Exception as e:
        print(f"[WARN] Error calculando tendencia para {ticker}: {e}")
        return "N/A"


def calcular_cartera():
    """Calcula el estado actual de la cartera basándose en el historial de operaciones"""
    operaciones = cargar_historial_operaciones()
    cartera = {}

    # Primero, construir lista de compras por ticker para rastrear precios individuales
    compras_por_ticker = {}

    for op in operaciones:
        symbol = op.get("ticker_symbol")
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)

        if symbol not in cartera:
            cartera[symbol] = {
                "acciones": 0,
                "total_comprado": 0,
                "total_vendido": 0,
                "precio_promedio_compra": 0,
                "capital_invertido": 0,
                "precio_compra_minimo": 0
            }
            compras_por_ticker[symbol] = []  # Lista de (precio, cantidad_restante)

        if tipo == "compra":
            # Actualizar precio promedio de compra
            total_acciones_previas = cartera[symbol]["acciones"]
            capital_previo = cartera[symbol]["capital_invertido"]
            nuevo_capital = capital_previo + (precio * cantidad)
            nuevas_acciones = total_acciones_previas + cantidad

            cartera[symbol]["acciones"] = nuevas_acciones
            cartera[symbol]["total_comprado"] += cantidad
            cartera[symbol]["capital_invertido"] = nuevo_capital
            if nuevas_acciones > 0:
                cartera[symbol]["precio_promedio_compra"] = nuevo_capital / nuevas_acciones

            # Agregar compra a la lista (ordenada por precio ascendente para FIFO por precio más bajo)
            compras_por_ticker[symbol].append([precio, cantidad])
            compras_por_ticker[symbol].sort(key=lambda x: x[0])  # Ordenar por precio

        elif tipo == "venta":
            cartera[symbol]["acciones"] -= cantidad
            cartera[symbol]["total_vendido"] += cantidad
            # Ajustar capital invertido proporcionalmente
            if cartera[symbol]["total_comprado"] > 0:
                proporcion = cantidad / cartera[symbol]["total_comprado"]
                cartera[symbol]["capital_invertido"] -= cartera[symbol]["capital_invertido"] * proporcion

            # Descontar de las compras (primero las de precio más bajo)
            cantidad_a_descontar = cantidad
            for compra in compras_por_ticker[symbol]:
                if cantidad_a_descontar <= 0:
                    break
                if compra[1] > 0:
                    descontar = min(compra[1], cantidad_a_descontar)
                    compra[1] -= descontar
                    cantidad_a_descontar -= descontar

            # Limpiar compras agotadas
            compras_por_ticker[symbol] = [c for c in compras_por_ticker[symbol] if c[1] > 0]

    # Calcular precio de compra mínimo para cada ticker (de las acciones restantes)
    for symbol in cartera:
        if compras_por_ticker.get(symbol) and cartera[symbol]["acciones"] > 0:
            # El precio mínimo es el primero de la lista ordenada
            cartera[symbol]["precio_compra_minimo"] = compras_por_ticker[symbol][0][0]
        else:
            cartera[symbol]["precio_compra_minimo"] = 0

    return cartera


def calcular_cartera_historica(fecha_limite):
    """
    Calcula el estado de la cartera hasta una fecha específica.
    Útil para regenerar señales históricas con la cartera que existía en esa fecha.

    Args:
        fecha_limite: Fecha límite (str YYYY-MM-DD o date). Las operaciones de esta fecha
                      en adelante NO se incluyen.

    Returns:
        dict: Cartera con acciones y precio_compra_minimo por ticker
    """
    operaciones = cargar_historial_operaciones()
    cartera = {}
    compras_por_ticker = {}

    # Convertir fecha_limite a string si es necesario
    if hasattr(fecha_limite, 'strftime'):
        fecha_limite_str = fecha_limite.strftime("%Y-%m-%d")
    else:
        fecha_limite_str = str(fecha_limite)

    for op in operaciones:
        fecha_op = op.get("fecha", "")
        # Solo procesar operaciones ANTERIORES a la fecha límite
        if fecha_op >= fecha_limite_str:
            continue

        symbol = op.get("ticker_symbol")
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)

        if symbol not in cartera:
            cartera[symbol] = {
                "acciones": 0,
                "capital_invertido": 0,
                "precio_compra_minimo": 0
            }
            compras_por_ticker[symbol] = []

        if tipo == "compra":
            cartera[symbol]["acciones"] += cantidad
            cartera[symbol]["capital_invertido"] += precio * cantidad
            compras_por_ticker[symbol].append([precio, cantidad])
            compras_por_ticker[symbol].sort(key=lambda x: x[0])

        elif tipo == "venta":
            cartera[symbol]["acciones"] -= cantidad
            # Descontar de las compras más baratas primero (FIFO por precio)
            restante = cantidad
            nuevas_compras = []
            for compra in compras_por_ticker[symbol]:
                if restante <= 0:
                    nuevas_compras.append(compra)
                elif compra[1] <= restante:
                    restante -= compra[1]
                else:
                    compra[1] -= restante
                    restante = 0
                    nuevas_compras.append(compra)
            compras_por_ticker[symbol] = nuevas_compras

    # Calcular precio_compra_minimo
    for symbol in cartera:
        if compras_por_ticker.get(symbol) and cartera[symbol]["acciones"] > 0:
            cartera[symbol]["precio_compra_minimo"] = compras_por_ticker[symbol][0][0]
        else:
            cartera[symbol]["precio_compra_minimo"] = 0

    return cartera


def calcular_ganancia_perdida():
    """Calcula el total de ganancia o pérdida efectiva de las operaciones

    Fórmula: Ganancia/Pérdida = (Ventas + Valor actual cartera) - Compras

    Las acciones no vendidas se valoran al último precio de cierre disponible.
    """
    operaciones = cargar_historial_operaciones()
    total_compras = 0
    total_ventas = 0

    for op in operaciones:
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)
        monto = precio * cantidad

        if tipo == "compra":
            total_compras += monto
        elif tipo == "venta":
            total_ventas += monto

    # Calcular valor actual de la cartera (acciones no vendidas)
    cartera = calcular_cartera()
    valor_cartera = 0

    # Obtener últimos precios de cierre
    ultimos_precios = {}
    if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
        try:
            df_log = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
            # Para cada ticker, obtener el último precio de cierre
            for ticker in df_log['Ticker'].unique():
                df_ticker = df_log[df_log['Ticker'] == ticker].sort_values('Date')
                if not df_ticker.empty:
                    ultimos_precios[ticker] = df_ticker.iloc[-1]['Close']
        except Exception as e:
            print(f"[WARN] Error leyendo precios: {e}")

    # Calcular valor de la cartera
    for symbol, datos in cartera.items():
        acciones = datos.get("acciones", 0)
        if acciones > 0 and symbol in ultimos_precios:
            precio = ultimos_precios[symbol]
            # Ignorar precios NaN
            if pd.notna(precio):
                valor_cartera += acciones * precio

    # Ganancia/Pérdida = (Ventas + Valor cartera) - Compras
    ganancia_perdida = (total_ventas + valor_cartera) - total_compras

    # Calcular ganancia realizada (solo de acciones vendidas)
    ganancia_realizada = calcular_ganancia_realizada()

    return {
        "total_compras": total_compras,
        "total_ventas": total_ventas,
        "valor_cartera": valor_cartera,
        "ganancia_perdida": ganancia_perdida,
        "ganancia_realizada": ganancia_realizada
    }


def calcular_ganancia_realizada():
    """Calcula la ganancia/pérdida realizada solo de acciones que se vendieron.

    Usa FIFO por precio más bajo: las ventas se asignan primero a las compras
    de menor precio. La ganancia realizada es la diferencia entre el precio
    de venta y el precio de compra de cada acción vendida.
    """
    operaciones = cargar_historial_operaciones()

    # Ordenar por fecha para procesar en orden cronológico
    operaciones_ordenadas = sorted(operaciones, key=lambda x: x.get("fecha", ""))

    # Diccionario de compras disponibles por ticker
    # Cada entrada es una lista de [precio, cantidad_disponible]
    compras_por_ticker = {}

    ganancia_total = 0

    for op in operaciones_ordenadas:
        symbol = op.get("ticker_symbol")
        tipo = op.get("tipo")
        cantidad = op.get("cantidad", 0)
        precio = op.get("precio", 0)

        if symbol not in compras_por_ticker:
            compras_por_ticker[symbol] = []

        if tipo == "compra":
            # Agregar compra a la lista (ordenada por precio ascendente)
            compras_por_ticker[symbol].append([precio, cantidad])
            compras_por_ticker[symbol].sort(key=lambda x: x[0])

        elif tipo == "venta":
            # Consumir de las compras de menor precio primero
            cantidad_a_vender = cantidad
            precio_venta = precio

            for compra in compras_por_ticker[symbol]:
                if cantidad_a_vender <= 0:
                    break
                if compra[1] > 0:
                    # Cantidad a consumir de esta compra
                    consumir = min(compra[1], cantidad_a_vender)
                    precio_compra = compra[0]

                    # Calcular ganancia de esta porción
                    ganancia_porcion = (precio_venta - precio_compra) * consumir
                    ganancia_total += ganancia_porcion

                    # Reducir cantidad disponible de esta compra
                    compra[1] -= consumir
                    cantidad_a_vender -= consumir

            # Limpiar compras agotadas
            compras_por_ticker[symbol] = [c for c in compras_por_ticker[symbol] if c[1] > 0]

    return ganancia_total


def administrar_historial():
    """Abre ventana para gestionar el historial de operaciones"""
    ruta = obtener_ruta_historial()
    if ruta is None:
        messagebox.showerror("Error", "No hay ubicación configurada.\nEjecuta primero Analisis_singrafico.py")
        return

    operaciones = cargar_historial_operaciones()

    # Crear ventana
    ventana_hist = tk.Toplevel(root)
    ventana_hist.title("Historial de Operaciones")
    ventana_hist.geometry("900x620")

    # Frame superior - Estado de cartera
    frame_cartera = tk.LabelFrame(ventana_hist, text="Estado Actual de Cartera", pady=5, padx=5)
    frame_cartera.pack(fill="x", padx=10, pady=5)

    # Treeview para cartera
    cols_cartera = ("Symbol", "Acciones", "P. Prom. Compra", "Capital Invertido")
    tree_cartera = ttk.Treeview(frame_cartera, columns=cols_cartera, show="headings", height=4)

    for col in cols_cartera:
        tree_cartera.heading(col, text=col)
        tree_cartera.column(col, width=120, anchor="center")

    tree_cartera.pack(fill="x", pady=5)

    def actualizar_cartera():
        """Actualiza la vista de cartera"""
        for item in tree_cartera.get_children():
            tree_cartera.delete(item)

        cartera = calcular_cartera()
        # Ordenar alfabéticamente por symbol
        for symbol, datos in sorted(cartera.items(), key=lambda x: x[0].upper()):
            if datos["acciones"] > 0 or datos["total_comprado"] > 0:
                tree_cartera.insert("", "end", values=(
                    symbol,
                    datos["acciones"],
                    f"${datos['precio_promedio_compra']:.2f}" if datos['precio_promedio_compra'] > 0 else "-",
                    f"${datos['capital_invertido']:.2f}" if datos['capital_invertido'] > 0 else "-"
                ))

    actualizar_cartera()

    # Frame resumen - Ganancia/Pérdida
    frame_resumen = tk.LabelFrame(ventana_hist, text="Resumen de Operaciones", pady=5, padx=5)
    frame_resumen.pack(fill="x", padx=10, pady=5)

    # Labels para mostrar resumen
    frame_resumen_inner = tk.Frame(frame_resumen)
    frame_resumen_inner.pack(fill="x", pady=5)

    lbl_compras = tk.Label(frame_resumen_inner, text="Compras: $0.00", font=("Arial", 9))
    lbl_compras.pack(side="left", padx=10)

    lbl_ventas = tk.Label(frame_resumen_inner, text="Ventas: $0.00", font=("Arial", 9))
    lbl_ventas.pack(side="left", padx=10)

    lbl_cartera = tk.Label(frame_resumen_inner, text="Cartera: $0.00", font=("Arial", 9), fg="#0066cc")
    lbl_cartera.pack(side="left", padx=10)

    lbl_realizada = tk.Label(frame_resumen_inner, text="Realizada: $0.00", font=("Arial", 9, "bold"))
    lbl_realizada.pack(side="left", padx=10)

    lbl_global = tk.Label(frame_resumen_inner, text="Global: $0.00", font=("Arial", 9, "bold"))
    lbl_global.pack(side="left", padx=10)

    def actualizar_resumen():
        """Actualiza el resumen de ganancia/pérdida"""
        resultado = calcular_ganancia_perdida()
        lbl_compras.config(text=f"Compras: ${resultado['total_compras']:,.2f}")
        lbl_ventas.config(text=f"Ventas: ${resultado['total_ventas']:,.2f}")
        lbl_cartera.config(text=f"Cartera: ${resultado['valor_cartera']:,.2f}")

        # Ganancia realizada (solo de acciones vendidas)
        gr = resultado['ganancia_realizada']
        if gr >= 0:
            lbl_realizada.config(text=f"Realizada: ${gr:,.2f}", fg="green")
        else:
            lbl_realizada.config(text=f"Realizada: -${abs(gr):,.2f}", fg="red")

        # Ganancia global (ventas + cartera - compras)
        gp = resultado['ganancia_perdida']
        if gp >= 0:
            lbl_global.config(text=f"Global: ${gp:,.2f}", fg="green")
        else:
            lbl_global.config(text=f"Global: -${abs(gp):,.2f}", fg="red")

    actualizar_resumen()

    # Frame medio - Historial de operaciones
    frame_historial = tk.LabelFrame(ventana_hist, text="Historial de Operaciones", pady=5, padx=5)
    frame_historial.pack(fill="both", expand=True, padx=10, pady=5)

    # Frame de filtros
    frame_filtros_hist = tk.Frame(frame_historial)
    frame_filtros_hist.pack(fill="x", pady=(0, 5))

    tk.Label(frame_filtros_hist, text="Filtrar:", font=("Arial", 9)).pack(side="left", padx=(0, 5))

    tk.Label(frame_filtros_hist, text="Ticker:", font=("Arial", 9)).pack(side="left")
    filtro_ticker_var = tk.StringVar(value="Todos")
    combo_filtro_ticker = ttk.Combobox(frame_filtros_hist, textvariable=filtro_ticker_var,
                                        state="readonly", width=10)
    combo_filtro_ticker.pack(side="left", padx=(2, 10))

    tk.Label(frame_filtros_hist, text="Fecha:", font=("Arial", 9)).pack(side="left")
    filtro_fecha_var = tk.StringVar(value="Todos")
    combo_filtro_fecha = ttk.Combobox(frame_filtros_hist, textvariable=filtro_fecha_var,
                                       state="readonly", width=12)
    combo_filtro_fecha.pack(side="left", padx=2)

    # Scrollbars
    scrollbar_y = tk.Scrollbar(frame_historial, orient="vertical")
    scrollbar_x = tk.Scrollbar(frame_historial, orient="horizontal")

    # Treeview para historial
    cols_hist = ("Fecha", "Symbol", "Tipo", "Precio", "Cantidad", "Total")
    tree_hist = ttk.Treeview(frame_historial, columns=cols_hist, show="headings",
                              selectmode="extended",
                              yscrollcommand=scrollbar_y.set,
                              xscrollcommand=scrollbar_x.set)

    scrollbar_y.config(command=tree_hist.yview)
    scrollbar_x.config(command=tree_hist.xview)

    anchos = {"Fecha": 100, "Symbol": 80, "Tipo": 70, "Precio": 90, "Cantidad": 70, "Total": 100}
    for col in cols_hist:
        tree_hist.heading(col, text=col)
        tree_hist.column(col, width=anchos.get(col, 80), anchor="center")

    tree_hist.tag_configure("compra", foreground="#008000")
    tree_hist.tag_configure("venta", foreground="#cc0000")

    def actualizar_filtros_hist():
        """Actualiza las opciones de los combos de filtro"""
        tickers = sorted(set(op.get("ticker_symbol", "") for op in operaciones))
        fechas = sorted(set(op.get("fecha", "") for op in operaciones), reverse=True)
        combo_filtro_ticker["values"] = ["Todos"] + tickers
        combo_filtro_fecha["values"] = ["Todos"] + fechas

    def actualizar_historial():
        """Actualiza la vista del historial"""
        nonlocal operaciones
        operaciones = cargar_historial_operaciones()

        actualizar_filtros_hist()

        for item in tree_hist.get_children():
            tree_hist.delete(item)

        filtro_t = filtro_ticker_var.get()
        filtro_f = filtro_fecha_var.get()

        # Ordenar por symbol alfabéticamente
        ops_ordenadas = sorted(operaciones, key=lambda x: x.get("ticker_symbol", "").upper())

        for op in ops_ordenadas:
            # Aplicar filtros
            if filtro_t != "Todos" and op.get("ticker_symbol", "") != filtro_t:
                continue
            if filtro_f != "Todos" and op.get("fecha", "") != filtro_f:
                continue

            precio = op.get("precio", 0)
            cantidad = op.get("cantidad", 0)
            total = precio * cantidad
            tipo = op.get("tipo", "")
            tree_hist.insert("", "end", values=(
                op.get("fecha", ""),
                op.get("ticker_symbol", ""),
                tipo.capitalize(),
                f"${precio:.2f}",
                cantidad,
                f"${total:.2f}"
            ), tags=(tipo,))

    def on_filtro_hist_change(*args):
        actualizar_historial()

    combo_filtro_ticker.bind("<<ComboboxSelected>>", on_filtro_hist_change)
    combo_filtro_fecha.bind("<<ComboboxSelected>>", on_filtro_hist_change)

    actualizar_historial()

    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")
    tree_hist.pack(fill="both", expand=True)

    # Frame inferior - Botones
    frame_botones = tk.Frame(ventana_hist, pady=10)
    frame_botones.pack(fill="x", padx=10)

    def agregar_operacion():
        """Abre ventana para agregar nueva operación"""
        ventana_add = tk.Toplevel(ventana_hist)
        ventana_add.title("Registrar Operación")
        ventana_add.geometry("350x300")
        ventana_add.transient(ventana_hist)
        ventana_add.grab_set()

        frame_form = tk.Frame(ventana_add, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        # Fecha
        tk.Label(frame_form, text="Fecha (YYYY-MM-DD):").grid(row=0, column=0, sticky="w", pady=5)
        entry_fecha = tk.Entry(frame_form, width=20)
        entry_fecha.insert(0, datetime.now().strftime("%Y-%m-%d"))
        entry_fecha.grid(row=0, column=1, pady=5)

        # Symbol
        tk.Label(frame_form, text="Symbol:").grid(row=1, column=0, sticky="w", pady=5)
        entry_symbol = tk.Entry(frame_form, width=20)
        entry_symbol.grid(row=1, column=1, pady=5)

        # Tipo
        tk.Label(frame_form, text="Tipo:").grid(row=2, column=0, sticky="w", pady=5)
        tipo_var = tk.StringVar(value="compra")
        frame_tipo = tk.Frame(frame_form)
        frame_tipo.grid(row=2, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_tipo, text="Compra", variable=tipo_var, value="compra").pack(side="left")
        tk.Radiobutton(frame_tipo, text="Venta", variable=tipo_var, value="venta").pack(side="left")

        # Precio
        tk.Label(frame_form, text="Precio:").grid(row=3, column=0, sticky="w", pady=5)
        entry_precio = tk.Entry(frame_form, width=20)
        entry_precio.grid(row=3, column=1, pady=5)

        # Cantidad
        tk.Label(frame_form, text="Cantidad:").grid(row=4, column=0, sticky="w", pady=5)
        entry_cantidad = tk.Entry(frame_form, width=20)
        entry_cantidad.grid(row=4, column=1, pady=5)

        def guardar():
            fecha = entry_fecha.get().strip()
            symbol = entry_symbol.get().strip().upper()
            tipo = tipo_var.get()

            if not fecha or not symbol:
                messagebox.showwarning("Campos requeridos", "Completa fecha y symbol")
                return

            try:
                precio = float(entry_precio.get().strip().replace(",", "."))
                cantidad = int(entry_cantidad.get().strip())
            except ValueError:
                messagebox.showerror("Error", "Precio y cantidad deben ser numéricos")
                return

            if cantidad <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            # Validar que no se venda más de lo que se tiene
            if tipo == "venta":
                cartera = calcular_cartera()
                acciones_disponibles = cartera.get(symbol, {}).get("acciones", 0)
                if cantidad > acciones_disponibles:
                    messagebox.showerror("Error",
                        f"No puedes vender {cantidad} acciones de {symbol}.\n"
                        f"Solo tienes {acciones_disponibles} en cartera.")
                    return

            nueva_op = {
                "fecha": fecha,
                "ticker_symbol": symbol,
                "tipo": tipo,
                "precio": precio,
                "cantidad": cantidad
            }

            operaciones.append(nueva_op)
            guardar_historial_operaciones(operaciones)
            actualizar_historial()
            actualizar_cartera()
            actualizar_resumen()
            messagebox.showinfo("Guardado", f"Operación registrada:\n{tipo.upper()} {cantidad} {symbol} @ ${precio:.2f}")
            ventana_add.destroy()

        tk.Button(frame_form, text="Guardar", command=guardar,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).grid(row=5, column=0, columnspan=2, pady=20)

    def editar_seleccionado():
        """Edita la operación seleccionada"""
        seleccionados = tree_hist.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona una operación para editar")
            return

        if len(seleccionados) > 1:
            messagebox.showwarning("Múltiple selección", "Selecciona solo una operación para editar")
            return

        # Obtener valores actuales
        item = seleccionados[0]
        valores = tree_hist.item(item, "values")
        fecha_actual = valores[0]
        symbol_actual = valores[1]
        tipo_actual = valores[2].lower()
        precio_actual = float(valores[3].replace("$", ""))
        cantidad_actual = int(valores[4])

        # Buscar índice en operaciones
        indice_editar = None
        for i, op in enumerate(operaciones):
            if (op.get("fecha") == fecha_actual and
                op.get("ticker_symbol") == symbol_actual and
                op.get("tipo") == tipo_actual and
                abs(op.get("precio", 0) - precio_actual) < 0.01 and
                op.get("cantidad") == cantidad_actual):
                indice_editar = i
                break

        if indice_editar is None:
            messagebox.showerror("Error", "No se encontró la operación")
            return

        # Ventana de edición
        ventana_edit = tk.Toplevel(ventana_hist)
        ventana_edit.title("Editar Operación")
        ventana_edit.geometry("350x300")
        ventana_edit.transient(ventana_hist)
        ventana_edit.grab_set()

        frame_form = tk.Frame(ventana_edit, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        # Fecha
        tk.Label(frame_form, text="Fecha (YYYY-MM-DD):").grid(row=0, column=0, sticky="w", pady=5)
        entry_fecha = tk.Entry(frame_form, width=20)
        entry_fecha.insert(0, fecha_actual)
        entry_fecha.grid(row=0, column=1, pady=5)

        # Symbol
        tk.Label(frame_form, text="Symbol:").grid(row=1, column=0, sticky="w", pady=5)
        entry_symbol = tk.Entry(frame_form, width=20)
        entry_symbol.insert(0, symbol_actual)
        entry_symbol.grid(row=1, column=1, pady=5)

        # Tipo
        tk.Label(frame_form, text="Tipo:").grid(row=2, column=0, sticky="w", pady=5)
        tipo_var = tk.StringVar(value=tipo_actual)
        frame_tipo = tk.Frame(frame_form)
        frame_tipo.grid(row=2, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_tipo, text="Compra", variable=tipo_var, value="compra").pack(side="left")
        tk.Radiobutton(frame_tipo, text="Venta", variable=tipo_var, value="venta").pack(side="left")

        # Precio
        tk.Label(frame_form, text="Precio:").grid(row=3, column=0, sticky="w", pady=5)
        entry_precio = tk.Entry(frame_form, width=20)
        entry_precio.insert(0, str(precio_actual))
        entry_precio.grid(row=3, column=1, pady=5)

        # Cantidad
        tk.Label(frame_form, text="Cantidad:").grid(row=4, column=0, sticky="w", pady=5)
        entry_cantidad = tk.Entry(frame_form, width=20)
        entry_cantidad.insert(0, str(cantidad_actual))
        entry_cantidad.grid(row=4, column=1, pady=5)

        def guardar_edicion():
            fecha = entry_fecha.get().strip()
            symbol = entry_symbol.get().strip().upper()
            tipo = tipo_var.get()

            if not fecha or not symbol:
                messagebox.showwarning("Campos requeridos", "Completa fecha y symbol")
                return

            try:
                precio = float(entry_precio.get().strip().replace(",", "."))
                cantidad = int(entry_cantidad.get().strip())
            except ValueError:
                messagebox.showerror("Error", "Precio y cantidad deben ser numéricos")
                return

            if cantidad <= 0:
                messagebox.showerror("Error", "La cantidad debe ser mayor a 0")
                return

            # Actualizar operación
            operaciones[indice_editar] = {
                "fecha": fecha,
                "ticker_symbol": symbol,
                "tipo": tipo,
                "precio": precio,
                "cantidad": cantidad
            }

            guardar_historial_operaciones(operaciones)
            actualizar_historial()
            actualizar_cartera()
            actualizar_resumen()
            messagebox.showinfo("Guardado", f"Operación actualizada:\n{tipo.upper()} {cantidad} {symbol} @ ${precio:.2f}")
            ventana_edit.destroy()

        tk.Button(frame_form, text="Guardar Cambios", command=guardar_edicion,
                  bg="#ffc107", fg="black", font=("Arial", 10, "bold")).grid(row=5, column=0, columnspan=2, pady=20)

    def eliminar_seleccionados():
        """Elimina las operaciones seleccionadas"""
        seleccionados = tree_hist.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona operaciones para eliminar")
            return

        if not messagebox.askyesno("Confirmar", f"¿Eliminar {len(seleccionados)} operación(es)?"):
            return

        # Obtener índices a eliminar
        indices_eliminar = []
        for item in seleccionados:
            valores = tree_hist.item(item, "values")
            fecha = valores[0]
            symbol = valores[1]
            tipo = valores[2].lower()
            precio = float(valores[3].replace("$", ""))
            cantidad = int(valores[4])

            # Buscar en operaciones
            for i, op in enumerate(operaciones):
                if (op.get("fecha") == fecha and
                    op.get("ticker_symbol") == symbol and
                    op.get("tipo") == tipo and
                    abs(op.get("precio", 0) - precio) < 0.01 and
                    op.get("cantidad") == cantidad):
                    indices_eliminar.append(i)
                    break

        # Eliminar en orden inverso para no afectar índices
        for i in sorted(indices_eliminar, reverse=True):
            operaciones.pop(i)

        guardar_historial_operaciones(operaciones)
        actualizar_historial()
        actualizar_cartera()
        actualizar_resumen()
        messagebox.showinfo("Eliminado", f"Se eliminaron {len(indices_eliminar)} operación(es)")

    def graficar_operaciones():
        """Grafica las operaciones realizadas: cuadrados verdes=compras, triángulos rojos=ventas"""
        if not operaciones:
            messagebox.showinfo("Sin datos", "No hay operaciones para graficar.")
            return

        # Obtener tickers únicos
        tickers_unicos = sorted(set(op.get("ticker_symbol", "") for op in operaciones))

        if not tickers_unicos:
            return

        # Ventana para seleccionar ticker
        ventana_graf = tk.Toplevel(ventana_hist)
        ventana_graf.title("Graficar Operaciones")
        ventana_graf.geometry("800x600")
        ventana_graf.resizable(True, True)
        ventana_graf.minsize(500, 400)

        tk.Label(ventana_graf, text="Seleccionar Ticker:", font=("Arial", 10)).pack(pady=5)

        ticker_var = tk.StringVar(value=tickers_unicos[0] if tickers_unicos else "")
        combo_ticker = ttk.Combobox(ventana_graf, textvariable=ticker_var, values=tickers_unicos, state="readonly", width=15)
        combo_ticker.pack(pady=5)

        # Frame para el gráfico
        frame_grafico = tk.Frame(ventana_graf)
        frame_grafico.pack(fill="both", expand=True, padx=10, pady=10)

        def actualizar_grafico(*args):
            # Limpiar frame
            for widget in frame_grafico.winfo_children():
                widget.destroy()

            ticker_sel = ticker_var.get()
            if not ticker_sel:
                return

            # Filtrar operaciones del ticker
            ops_ticker = [op for op in operaciones if op.get("ticker_symbol") == ticker_sel]

            if not ops_ticker:
                tk.Label(frame_grafico, text="Sin operaciones para este ticker").pack()
                return

            # Separar compras y ventas
            compras = [(op.get("fecha"), op.get("precio", 0)) for op in ops_ticker if op.get("tipo", "").lower() == "compra"]
            ventas = [(op.get("fecha"), op.get("precio", 0)) for op in ops_ticker if op.get("tipo", "").lower() == "venta"]

            # Cargar precios de cierre del log
            precios_cierre = []
            if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
                try:
                    df_log = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
                    df_ticker = df_log[df_log['Ticker'] == ticker_sel].sort_values('Date')
                    if not df_ticker.empty:
                        precios_cierre = [(row['Date'], row['Close']) for _, row in df_ticker.iterrows()]
                except Exception as e:
                    print(f"[WARN] Error cargando precios: {e}")

            # Crear figura
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.subplots_adjust(left=0.06, right=0.98, bottom=0.12, top=0.94)

            # Recopilar todas las fechas y precios
            todas_fechas = []
            todos_precios = []

            # Graficar línea de precios de cierre (azul, sin puntos)
            if precios_cierre:
                fechas_p = [f for f, _ in precios_cierre]
                precios_p = [p for _, p in precios_cierre]
                ax.plot(fechas_p, precios_p, color='blue', linewidth=1.5, label='Precio Cierre', zorder=1)
                todas_fechas.extend(fechas_p)
                todos_precios.extend(precios_p)

            # Graficar compras (cuadrados verdes)
            if compras:
                fechas_c = [datetime.strptime(f, "%Y-%m-%d") for f, _ in compras]
                precios_c = [p for _, p in compras]
                ax.scatter(fechas_c, precios_c, marker='s', s=35, c='green', label='Compras', zorder=5)
                todas_fechas.extend(fechas_c)
                todos_precios.extend(precios_c)

            # Graficar ventas (triángulos rojos)
            if ventas:
                fechas_v = [datetime.strptime(f, "%Y-%m-%d") for f, _ in ventas]
                precios_v = [p for _, p in ventas]
                ax.scatter(fechas_v, precios_v, marker='^', s=35, c='red', label='Ventas', zorder=5)
                todas_fechas.extend(fechas_v)
                todos_precios.extend(precios_v)

            # Ajustar límites del eje X (margen de 7 días)
            if todas_fechas:
                from datetime import timedelta
                fecha_min = min(todas_fechas) - timedelta(days=7)
                fecha_max = max(todas_fechas) + timedelta(days=7)
                ax.set_xlim(fecha_min, fecha_max)

            # Ajustar límites del eje Y (margen del 5%)
            if todos_precios:
                precio_min = min(todos_precios)
                precio_max = max(todos_precios)
                margen = (precio_max - precio_min) * 0.1 if precio_max != precio_min else precio_min * 0.05
                ax.set_ylim(precio_min - margen, precio_max + margen)

            ax.set_title(f"Operaciones - {ticker_sel}")
            ax.set_xlabel("Fecha")
            ax.set_ylabel("Precio ($)")
            ax.legend()
            ax.grid(True, alpha=0.3)

            # Formato de fechas: dd/mm/yy
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%y'))
            fig.autofmt_xdate(rotation=45)

            # Mostrar en tkinter
            canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)

            plt.close(fig)

        combo_ticker.bind("<<ComboboxSelected>>", actualizar_grafico)
        actualizar_grafico()  # Mostrar gráfico inicial

    tk.Button(frame_botones, text="Registrar Operación", command=agregar_operacion,
              bg="#007bff", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Editar", command=editar_seleccionado,
              bg="#ffc107", fg="black", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Eliminar seleccionadas", command=eliminar_seleccionados,
              bg="#ff6b6b", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Graficar", command=graficar_operaciones,
              bg="#6f42c1", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    def exportar_excel_historial():
        """Exporta el historial de operaciones a Excel"""
        if not operaciones:
            messagebox.showinfo("Sin datos", "No hay operaciones para exportar.")
            return

        from tkinter import filedialog
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="historial_operaciones.xlsx",
            title="Guardar historial como Excel"
        )
        if not ruta_archivo:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()

            # Hoja 1: Historial de operaciones
            ws = wb.active
            ws.title = "Operaciones"

            encabezados = ["Fecha", "Symbol", "Tipo", "Precio", "Cantidad", "Total"]
            for col_idx, enc in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col_idx, value=enc)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            ops_ordenadas = sorted(operaciones, key=lambda x: x.get("ticker_symbol", "").upper())
            for row_idx, op in enumerate(ops_ordenadas, 2):
                precio = op.get("precio", 0)
                cantidad = op.get("cantidad", 0)
                ws.cell(row=row_idx, column=1, value=op.get("fecha", ""))
                ws.cell(row=row_idx, column=2, value=op.get("ticker_symbol", ""))
                tipo = op.get("tipo", "")
                cell_tipo = ws.cell(row=row_idx, column=3, value=tipo.capitalize())
                if tipo == "compra":
                    cell_tipo.font = Font(color="008000")
                else:
                    cell_tipo.font = Font(color="FF0000")
                ws.cell(row=row_idx, column=4, value=precio)
                ws.cell(row=row_idx, column=5, value=cantidad)
                ws.cell(row=row_idx, column=6, value=round(precio * cantidad, 2))

            # Formato de columnas numéricas
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = '$#,##0.00'
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
                for cell in row:
                    cell.number_format = '$#,##0.00'

            # Ajustar anchos
            for col_idx, enc in enumerate(encabezados, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = max(len(enc) + 4, 12)

            # Hoja 2: Resumen cartera
            ws2 = wb.create_sheet("Cartera")
            enc_cartera = ["Symbol", "Acciones", "P. Prom. Compra", "Capital Invertido"]
            for col_idx, enc in enumerate(enc_cartera, 1):
                cell = ws2.cell(row=1, column=col_idx, value=enc)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            cartera = calcular_cartera()
            row_idx = 2
            for symbol, datos in sorted(cartera.items(), key=lambda x: x[0].upper()):
                if datos["acciones"] > 0:
                    ws2.cell(row=row_idx, column=1, value=symbol)
                    ws2.cell(row=row_idx, column=2, value=datos["acciones"])
                    ws2.cell(row=row_idx, column=3, value=round(datos["precio_promedio_compra"], 2))
                    ws2.cell(row=row_idx, column=4, value=round(datos["capital_invertido"], 2))
                    row_idx += 1

            for col_idx, enc in enumerate(enc_cartera, 1):
                ws2.column_dimensions[chr(64 + col_idx)].width = max(len(enc) + 4, 14)

            wb.save(ruta_archivo)
            messagebox.showinfo("Exportado", f"Historial exportado a:\n{ruta_archivo}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar:\n{str(e)}")

    tk.Button(frame_botones, text="Exportar Excel", command=exportar_excel_historial,
              bg="#17a2b8", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_hist.destroy).pack(side="right", padx=5)


def filtrar_parametros_por_fecha(parametros, fecha_objetivo):
    """
    Filtra parámetros que estén vigentes para una fecha específica.
    Un parámetro está vigente si:
    - fecha_inicio es None O fecha_inicio <= fecha_objetivo
    - fecha_fin es None O fecha_fin >= fecha_objetivo

    Args:
        parametros: Lista de parámetros
        fecha_objetivo: String en formato YYYY-MM-DD o objeto datetime

    Returns:
        Lista de parámetros vigentes para esa fecha
    """
    if not fecha_objetivo:
        return parametros

    # Convertir fecha_objetivo a string si es datetime
    if hasattr(fecha_objetivo, 'strftime'):
        fecha_str = fecha_objetivo.strftime("%Y-%m-%d")
    else:
        fecha_str = str(fecha_objetivo)[:10]  # Tomar solo YYYY-MM-DD

    parametros_vigentes = []
    for param in parametros:
        fecha_inicio = param.get("fecha_inicio")
        fecha_fin = param.get("fecha_fin")

        # Verificar fecha inicio
        if fecha_inicio and fecha_inicio > fecha_str:
            continue  # El parámetro aún no era válido en esa fecha

        # Verificar fecha fin
        if fecha_fin and fecha_fin < fecha_str:
            continue  # El parámetro ya había expirado

        # El parámetro está vigente
        parametros_vigentes.append(param)

    return parametros_vigentes


def calcular_senales_para_parametros(parametros, df_precios, precios_dict, cartera):
    """Calcula señales de compra/venta para un conjunto de parámetros"""

    LIMITE_TIPO_DEFAULT = "acciones"
    LIMITE_VALOR_DEFAULT = 10.0

    senales = []
    for param in parametros:
        symbol = param.get('ticker_symbol')

        limite_tipo = param.get('limite_tipo', LIMITE_TIPO_DEFAULT)
        limite_valor = param.get('limite_valor', LIMITE_VALOR_DEFAULT)

        info_cartera = cartera.get(symbol, {"acciones": 0, "capital_invertido": 0, "precio_compra_minimo": 0})
        acciones_en_cartera = info_cartera.get("acciones", 0)
        capital_invertido = info_cartera.get("capital_invertido", 0)
        precio_compra_minimo = info_cartera.get("precio_compra_minimo", 0)

        if symbol not in precios_dict:
            senales.append({
                'symbol': symbol,
                'fecha_precio': 'N/A',
                'cierre': 'N/A',
                'precio_compra': 'N/A',
                'cant_compra': '-',
                'opc_compra': 'N/A',
                'precio_venta': 'N/A',
                'cant_venta': '-',
                'opc_venta': 'N/A',
                'acciones_cartera': acciones_en_cartera,
                'precio_compra_minimo': precio_compra_minimo,
                'ganancia_min_pct': param.get('ganancia_min_pct', 0),
                'limite_tipo': limite_tipo,
                'limite_valor': limite_valor,
                'tendencia': 'N/A',
                'estado': 'Sin datos de precio'
            })
            continue

        precio_info = precios_dict[symbol]
        cierre = precio_info['close']
        compra_pct = param.get('compra_pct', 0)
        venta_pct = param.get('venta_pct', 0)
        ganancia_min_pct = param.get('ganancia_min_pct', 0)

        precio_compra = cierre * (1 + compra_pct / 100)

        # Precio de venta basado en el cierre actual
        precio_venta_por_cierre = cierre * (1 + venta_pct / 100)

        # Precio de venta mínimo para garantizar ganancia sobre el precio de compra más bajo
        if precio_compra_minimo > 0 and ganancia_min_pct > 0:
            precio_venta_minimo = precio_compra_minimo * (1 + ganancia_min_pct / 100)
            # El precio de venta debe ser el mayor entre ambos para garantizar la ganancia mínima
            precio_venta = max(precio_venta_por_cierre, precio_venta_minimo)
        else:
            precio_venta = precio_venta_por_cierre

        promedio_minimos = param.get('promedio_minimos', 0)
        promedio_maximos = param.get('promedio_maximos', 0)
        compra_multiple_config = param.get('compra_multiple') or 1
        venta_multiple_config = param.get('venta_multiple') or 1

        usar_compra_multiple = False
        usar_venta_multiple = False
        pct_acumulado = 0

        if df_precios is not None and symbol in df_precios['Ticker'].values:
            try:
                hist_ticker = df_precios[df_precios['Ticker'] == symbol].sort_values('Date')
                if len(hist_ticker) >= 2:
                    precios_cierre = hist_ticker['Close'].values
                    precio_referencia = precios_cierre[0]
                    variacion_diaria_anterior = 0

                    for i in range(1, len(precios_cierre)):
                        precio_anterior = precios_cierre[i - 1]
                        precio_actual_iter = precios_cierre[i]
                        variacion_diaria = ((precio_actual_iter - precio_anterior) / precio_anterior) * 100

                        if variacion_diaria_anterior != 0:
                            if (variacion_diaria_anterior > 0 and variacion_diaria < 0) or \
                               (variacion_diaria_anterior < 0 and variacion_diaria > 0):
                                precio_referencia = precio_anterior

                        variacion_diaria_anterior = variacion_diaria

                    precio_actual = precios_cierre[-1]
                    pct_acumulado = ((precio_actual - precio_referencia) / precio_referencia) * 100

                    if promedio_minimos < 0 and pct_acumulado <= promedio_minimos:
                        usar_compra_multiple = True
                    if promedio_maximos > 0 and pct_acumulado >= promedio_maximos:
                        usar_venta_multiple = True
            except Exception as e:
                print(f"[WARN] Error calculando % acumulado para {symbol}: {e}")

        cant_compra = compra_multiple_config if usar_compra_multiple else 1
        cant_venta = venta_multiple_config if usar_venta_multiple else 1

        if limite_tipo == "acciones":
            limite_acciones = int(limite_valor)
            if acciones_en_cartera >= limite_acciones:
                opc_compra = "N/A (límite)"
            else:
                espacio_disponible = limite_acciones - acciones_en_cartera
                cant_compra = min(cant_compra, espacio_disponible)
                opc_compra = "Comprar"
        else:
            limite_monto = float(limite_valor)
            if capital_invertido >= limite_monto:
                opc_compra = "N/A (límite $)"
            else:
                monto_disponible = limite_monto - capital_invertido
                max_acciones_por_monto = int(monto_disponible / precio_compra) if precio_compra > 0 else 0
                if max_acciones_por_monto <= 0:
                    opc_compra = "N/A (límite $)"
                else:
                    cant_compra = min(cant_compra, max_acciones_por_monto)
                    opc_compra = "Comprar"

        if acciones_en_cartera <= 0:
            opc_venta = "N/A (sin acciones)"
            cant_venta = 0
        else:
            cant_venta = min(cant_venta, acciones_en_cartera)
            opc_venta = "Vender"

        # Calcular tendencias (corta 10 días, larga 30 días)
        tendencia_corta = calcular_tendencia(df_precios, symbol, dias=10) if df_precios is not None else "N/A"
        tendencia_larga = calcular_tendencia(df_precios, symbol, dias=30) if df_precios is not None else "N/A"

        senales.append({
            'symbol': symbol,
            'fecha_precio': precio_info['fecha'].strftime('%Y-%m-%d'),
            'cierre': cierre,
            'precio_compra': precio_compra,
            'cant_compra': cant_compra,
            'opc_compra': opc_compra,
            'precio_venta': precio_venta,
            'cant_venta': cant_venta,
            'opc_venta': opc_venta,
            'acciones_cartera': acciones_en_cartera,
            'precio_compra_minimo': precio_compra_minimo,
            'ganancia_min_pct': ganancia_min_pct,
            'limite_tipo': limite_tipo,
            'limite_valor': limite_valor,
            'tendencia': tendencia_corta,
            'tendencia_larga': tendencia_larga,
            'estado': 'OK'
        })

    return senales


def generar_senales():
    """Genera señales de compra/venta basadas en parámetros activos (todos los slots)"""

    # Verificar si es fin de semana (mostrar mensaje pero continuar)
    hoy = datetime.now()
    es_fin_de_semana = hoy.weekday() >= 5
    if es_fin_de_semana:
        dia_semana = "sábado" if hoy.weekday() == 5 else "domingo"
        messagebox.showinfo("Mercado cerrado",
            f"Hoy es {dia_semana}. El mercado está cerrado.\n\n"
            "Se mostrarán las señales basadas en el último día de trading.\n"
            "(Las señales no se guardarán porque ya están guardadas)")

    # Usar siempre la ruta portable del log (consistente con sincronizar_desde_github)
    log_file = str(AUTO_UPDATE_LOG_PORTABLE)

    if not os.path.exists(log_file):
        messagebox.showwarning("Sin datos", f"No existe el archivo de log:\n{log_file}\n\nDescarga los precios primero.")
        return

    # Cargar parámetros activos (estructura de slots v2.0)
    datos_slots, error = cargar_parametros_activos()
    if error:
        messagebox.showerror("Error", error)
        return

    # Cargar estado de cartera
    cartera = calcular_cartera()

    # Cargar precios del log
    try:
        df_precios = pd.read_csv(log_file, parse_dates=['Date'])
    except Exception as e:
        messagebox.showerror("Error", f"Error leyendo archivo de precios:\n{e}")
        return

    # Obtener el último precio de cierre para cada ticker
    df_precios['Date'] = pd.to_datetime(df_precios['Date'])
    ultimos_precios = df_precios.sort_values('Date').groupby('Ticker').last().reset_index()

    # Crear diccionario de precios
    precios_dict = {}
    fecha_senales = None
    for _, row in ultimos_precios.iterrows():
        precios_dict[row['Ticker']] = {
            'fecha': row['Date'],
            'close': row['Close'],
            'open': row['Open'],
            'high': row['High'],
            'low': row['Low']
        }
        # Tomar la fecha de cualquier ticker (todas deberían ser iguales para el último día)
        if fecha_senales is None:
            fecha_senales = row['Date']

    # Verificar si debemos guardar las señales (precio de cierre confirmado)
    # - Si es fin de semana → NO guardar (ya se guardaron el viernes)
    # - Si la fecha de los precios NO es hoy → guardar
    # - Si la fecha es hoy Y hora NY >= 16:30 → guardar (mercado cerrado)
    # - Si la fecha es hoy Y hora NY < 16:30 → NO guardar (mercado abierto)
    now_ny = datetime.now(ZoneInfo("America/New_York"))
    hoy_ny = now_ny.date()
    hora_ny = now_ny.hour + now_ny.minute / 60  # Hora decimal (16:30 = 16.5)
    fecha_precios = fecha_senales.date() if fecha_senales else None

    # En fin de semana no guardar (ya se guardaron el viernes)
    if es_fin_de_semana:
        guardar_senales = False
    else:
        guardar_senales = (fecha_precios != hoy_ny) or (fecha_precios == hoy_ny and hora_ny >= 16.5)

    # Calcular la fecha del siguiente día de trading (las señales son para esa fecha)
    fecha_siguiente_trading = siguiente_dia_trading(fecha_senales)
    fecha_guardar = fecha_siguiente_trading.strftime("%Y-%m-%d") + " 09:30:00"  # Apertura de mercado

    # Generar señales para CADA slot
    senales_por_slot = {}
    for slot_id in ["1", "2", "3", "4", "5"]:
        parametros = obtener_parametros_slot(datos_slots, slot_id)
        if parametros:
            # Filtrar parámetros vigentes para la fecha del siguiente día de trading
            parametros_vigentes = filtrar_parametros_por_fecha(parametros, fecha_siguiente_trading)
            if parametros_vigentes:
                senales = calcular_senales_para_parametros(parametros_vigentes, df_precios, precios_dict, cartera)
                senales_por_slot[slot_id] = senales
                # Solo guardar señales si corresponde (mercado cerrado y no es fin de semana)
                if guardar_senales:
                    nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                    guardar_historial_senales(senales, slot_id, nombre_slot, fecha_guardar)
            else:
                senales_por_slot[slot_id] = []
        else:
            senales_por_slot[slot_id] = []

    # Mostrar ventana con señales (ahora con pestañas por slot)
    mostrar_ventana_senales(senales_por_slot, datos_slots)


def regenerar_senales_historicas():
    """Permite regenerar señales para una fecha anterior basándose en datos históricos"""

    # Verificar que hay un CSV configurado
    csv_file = entry_ruta.get()
    if not csv_file:
        messagebox.showwarning("Sin datos", "Primero selecciona y descarga un CSV de precios")
        return

    # Obtener ruta del log
    log_file = os.path.join(os.path.dirname(csv_file), "auto_update_log.csv")

    if not os.path.exists(log_file):
        messagebox.showwarning("Sin datos", f"No existe el archivo de log:\n{log_file}")
        return

    # Cargar precios del log
    try:
        df_precios = pd.read_csv(log_file, parse_dates=['Date'])
        df_precios['Date'] = pd.to_datetime(df_precios['Date'])
    except Exception as e:
        messagebox.showerror("Error", f"Error leyendo archivo de precios:\n{e}")
        return

    # Obtener fechas disponibles
    fechas_disponibles = sorted(df_precios['Date'].dt.strftime('%Y-%m-%d').unique(), reverse=True)

    if not fechas_disponibles:
        messagebox.showinfo("Sin datos", "No hay fechas disponibles en el log de precios")
        return

    # Crear ventana de selección de fecha
    ventana_fecha = tk.Toplevel(root)
    ventana_fecha.title("Regenerar Señales Históricas")
    ventana_fecha.geometry("400x200")
    ventana_fecha.transient(root)
    ventana_fecha.grab_set()

    tk.Label(ventana_fecha, text="Selecciona la fecha para regenerar señales:",
             font=("Arial", 10)).pack(pady=10)

    # Combobox con fechas disponibles
    fecha_var = tk.StringVar()
    combo_fechas = ttk.Combobox(ventana_fecha, textvariable=fecha_var, values=fechas_disponibles,
                                 state="readonly", width=20)
    combo_fechas.pack(pady=5)
    combo_fechas.current(0)

    tk.Label(ventana_fecha, text="(Las señales se guardarán para la siguiente apertura de mercado)",
             font=("Arial", 9), fg="gray").pack(pady=5)

    def procesar_fecha():
        fecha_seleccionada = fecha_var.get()
        if not fecha_seleccionada:
            return

        # Cargar estructura de slots
        datos_slots, error = cargar_parametros_activos()
        if error:
            messagebox.showerror("Error", error)
            return

        # Calcular el siguiente día de trading (las señales son para esa fecha)
        fecha_siguiente_trading = siguiente_dia_trading(datetime.strptime(fecha_seleccionada, "%Y-%m-%d"))

        # Calcular cartera HISTÓRICA (solo operaciones anteriores a la fecha de la señal)
        # Esto refleja la cartera que se tenía cuando se generaron las señales originalmente
        cartera = calcular_cartera_historica(fecha_siguiente_trading)

        # Filtrar precios para la fecha seleccionada
        df_fecha = df_precios[df_precios['Date'].dt.strftime('%Y-%m-%d') == fecha_seleccionada]

        if df_fecha.empty:
            messagebox.showwarning("Sin datos", f"No hay datos de precios para {fecha_seleccionada}")
            return

        # Crear diccionario de precios para esa fecha
        precios_dict = {}
        for _, row in df_fecha.iterrows():
            precios_dict[row['Ticker']] = {
                'fecha': row['Date'],
                'close': row['Close'],
                'open': row['Open'],
                'high': row['High'],
                'low': row['Low']
            }

        # fecha_siguiente_trading ya fue calculada arriba para la cartera histórica
        fecha_generacion = fecha_siguiente_trading.strftime("%Y-%m-%d") + " 09:30:00"  # Apertura de mercado
        total_senales = 0

        # Generar señales para CADA slot, filtrando por fecha del siguiente día de trading
        for slot_id in ["1", "2", "3", "4", "5"]:
            parametros = obtener_parametros_slot(datos_slots, slot_id)
            if not parametros:
                continue

            # Filtrar parámetros vigentes para la fecha del siguiente día de trading
            parametros_vigentes = filtrar_parametros_por_fecha(parametros, fecha_siguiente_trading)
            if not parametros_vigentes:
                continue

            # Filtrar df_precios hasta la fecha seleccionada (no usar datos futuros)
            fecha_limite = pd.to_datetime(fecha_seleccionada)
            df_precios_historico = df_precios[df_precios['Date'] <= fecha_limite]

            # Calcular señales usando solo datos históricos
            senales = calcular_senales_para_parametros(parametros_vigentes, df_precios_historico, precios_dict, cartera)

            if senales:
                # Guardar en el historial del slot
                nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                guardar_historial_senales(senales, slot_id, nombre_slot, fecha_generacion)
                total_senales += len(senales)

        ventana_fecha.destroy()
        fecha_siguiente_str = fecha_siguiente_trading.strftime("%Y-%m-%d")
        if total_senales > 0:
            messagebox.showinfo("Éxito",
                f"Señales regeneradas:\n"
                f"- Cierre usado: {fecha_seleccionada}\n"
                f"- Fecha de señales: {fecha_siguiente_str}\n"
                f"- {total_senales} señales guardadas en todos los slots")
        else:
            messagebox.showinfo("Sin señales",
                f"No se generaron señales para {fecha_siguiente_str}\n"
                "(Verifica que los parámetros estén vigentes para esa fecha)")

    def procesar_todas_fechas():
        """Regenera señales para TODAS las fechas disponibles"""
        if not messagebox.askyesno("Confirmar",
            f"¿Regenerar señales para las {len(fechas_disponibles)} fechas disponibles?\n\n"
            "Esto reemplazará todas las señales históricas existentes."):
            return

        # Limpiar historial existente
        historial_path = obtener_ruta_senales()
        nuevo_historial = {
            "version": "2.0",
            "senales_por_slot": {"1": [], "2": [], "3": [], "4": [], "5": []}
        }
        with open(historial_path, 'w') as f:
            json.dump(nuevo_historial, f, indent=2)

        # Cargar estructura de slots
        datos_slots, error = cargar_parametros_activos()
        if error:
            messagebox.showerror("Error", error)
            return

        total_global = 0
        fechas_procesadas = 0

        # Procesar cada fecha (de más antigua a más reciente)
        for fecha_str in sorted(fechas_disponibles):
            fecha_siguiente_trading = siguiente_dia_trading(datetime.strptime(fecha_str, "%Y-%m-%d"))
            cartera = calcular_cartera_historica(fecha_siguiente_trading)

            df_fecha = df_precios[df_precios['Date'].dt.strftime('%Y-%m-%d') == fecha_str]
            if df_fecha.empty:
                continue

            precios_dict = {}
            for _, row in df_fecha.iterrows():
                precios_dict[row['Ticker']] = {
                    'fecha': row['Date'],
                    'close': row['Close'],
                    'open': row['Open'],
                    'high': row['High'],
                    'low': row['Low']
                }

            fecha_generacion = fecha_siguiente_trading.strftime("%Y-%m-%d") + " 09:30:00"

            # Filtrar df_precios hasta la fecha actual (no usar datos futuros)
            fecha_limite = pd.to_datetime(fecha_str)
            df_precios_historico = df_precios[df_precios['Date'] <= fecha_limite]

            for slot_id in ["1", "2", "3", "4", "5"]:
                parametros = obtener_parametros_slot(datos_slots, slot_id)
                if not parametros:
                    continue

                parametros_vigentes = filtrar_parametros_por_fecha(parametros, fecha_siguiente_trading)
                if not parametros_vigentes:
                    continue

                senales = calcular_senales_para_parametros(parametros_vigentes, df_precios_historico, precios_dict, cartera)

                if senales:
                    nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                    guardar_historial_senales(senales, slot_id, nombre_slot, fecha_generacion)
                    total_global += len(senales)

            fechas_procesadas += 1

        ventana_fecha.destroy()
        messagebox.showinfo("Completado",
            f"Regeneración completada:\n\n"
            f"- Fechas procesadas: {fechas_procesadas}\n"
            f"- Total señales: {total_global}")

    frame_botones = tk.Frame(ventana_fecha)
    frame_botones.pack(pady=20)

    tk.Button(frame_botones, text="Regenerar Señales", command=procesar_fecha,
              bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Regenerar TODAS", command=procesar_todas_fechas,
              bg="#6c757d", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cancelar", command=ventana_fecha.destroy).pack(side="left", padx=5)


def mostrar_ventana_senales(senales_por_slot, datos_slots, titulo_extra=""):
    """Muestra una ventana con las señales generadas (con pestañas por slot)"""

    ventana_senales = tk.Toplevel(root)
    ventana_senales.title("Señales de Trading - " + datetime.now().strftime("%Y-%m-%d %H:%M") + titulo_extra)
    ventana_senales.geometry("1200x550")

    # Frame superior con info
    frame_info = tk.Frame(ventana_senales, pady=5)
    frame_info.pack(fill="x", padx=10)

    fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_tickers = sum(len(senales) for senales in senales_por_slot.values())

    lbl_info = tk.Label(frame_info, text=f"Señales generadas: {fecha_generacion}",
             font=("Arial", 10, "bold"))
    lbl_info.pack(side="left")

    # Checkbox "Ver guardadas" dentro de la ventana
    ver_ant_var = tk.BooleanVar(value=False)
    tk.Checkbutton(frame_info, text="Ver guardadas", variable=ver_ant_var,
                   font=("Arial", 9), command=lambda: toggle_ver_anteriores()).pack(side="left", padx=15)

    # Campo límite de plataforma
    tk.Label(frame_info, text="Límite plataforma:", font=("Arial", 9)).pack(side="left", padx=(20, 2))
    limite_plataforma_var = tk.StringVar(value="3")
    entry_limite = tk.Entry(frame_info, textvariable=limite_plataforma_var, width=4, font=("Arial", 9), justify="center")
    entry_limite.pack(side="left")
    tk.Label(frame_info, text="%", font=("Arial", 9)).pack(side="left", padx=(0, 5))

    def aplicar_limite():
        if ver_ant_var.get():
            toggle_ver_anteriores()  # Recargar guardadas con nuevo límite
        else:
            poblar_trees(senales_por_slot)  # Recargar actuales con nuevo límite

    tk.Button(frame_info, text="Aplicar", command=aplicar_limite, font=("Arial", 8),
              bg="#6c757d", fg="white", padx=5).pack(side="left", padx=5)

    tk.Label(frame_info, text=f"Total tickers: {total_tickers}",
             font=("Arial", 10)).pack(side="right")

    # Notebook para pestañas de slots
    notebook = ttk.Notebook(ventana_senales)
    notebook.pack(fill="both", expand=True, padx=10, pady=5)

    columns = ("Symbol", "Cartera", "Cierre últ.", "P.Compra", "Cant.C", "Opc.Compra", "P.Venta", "Cant.V", "Opc.Venta", "Tend.C", "Tend.L")
    anchos = {"Symbol": 70, "Cartera": 60, "Cierre últ.": 85, "P.Compra": 85, "Cant.C": 50,
              "Opc.Compra": 110, "P.Venta": 85, "Cant.V": 50, "Opc.Venta": 120, "Tend.C": 55, "Tend.L": 55}

    trees = {}

    # Crear pestañas con treeviews vacíos
    for slot_id in ["1", "2", "3", "4", "5"]:
        nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
        senales = senales_por_slot.get(slot_id, [])

        frame_slot = tk.Frame(notebook)
        notebook.add(frame_slot, text=f"{nombre_slot} ({len(senales)})")

        scrollbar_y = tk.Scrollbar(frame_slot, orient="vertical")
        scrollbar_x = tk.Scrollbar(frame_slot, orient="horizontal")

        tree = ttk.Treeview(frame_slot, columns=columns, show="headings",
                            yscrollcommand=scrollbar_y.set,
                            xscrollcommand=scrollbar_x.set)

        scrollbar_y.config(command=tree.yview)
        scrollbar_x.config(command=tree.xview)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=anchos.get(col, 70), anchor="center")

        # Tags para precios ajustados (naranja)
        tree.tag_configure("ajustado", foreground="#FF6600")

        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        trees[slot_id] = tree

    def poblar_trees(datos):
        """Llena todos los trees con los datos proporcionados"""
        # Obtener límite de plataforma (None = sin límite)
        valor_limite = limite_plataforma_var.get().strip()
        if valor_limite == "" or valor_limite == "0":
            limite_pct = None  # Sin límite
        else:
            try:
                limite_pct = float(valor_limite) / 100.0
            except ValueError:
                limite_pct = None  # Valor inválido = sin límite

        for slot_id, tree in trees.items():
            tree.delete(*tree.get_children())
            senales = datos.get(slot_id, [])
            senales_ordenadas = sorted(senales, key=lambda x: x.get('symbol', '').upper())
            for senal in senales_ordenadas:
                if senal.get('estado') == 'OK':
                    cierre = senal['cierre']
                    precio_compra_orig = senal['precio_compra']
                    precio_venta_orig = senal['precio_venta']

                    # Ajustar precios si hay límite activo
                    precio_compra_mostrar = precio_compra_orig
                    precio_venta_mostrar = precio_venta_orig
                    compra_ajustada = False
                    venta_ajustada = False

                    if limite_pct is not None and limite_pct > 0:
                        limite_compra_min = cierre * (1 - limite_pct)
                        limite_venta_max = cierre * (1 + limite_pct)

                        if precio_compra_orig < limite_compra_min:
                            precio_compra_mostrar = limite_compra_min
                            compra_ajustada = True

                        if precio_venta_orig > limite_venta_max:
                            precio_venta_mostrar = limite_venta_max
                            venta_ajustada = True

                    # Formato de precios: agregar * si fue ajustado
                    str_compra = f"*${precio_compra_mostrar:.2f}" if compra_ajustada else f"${precio_compra_mostrar:.2f}"
                    str_venta = f"*${precio_venta_mostrar:.2f}" if venta_ajustada else f"${precio_venta_mostrar:.2f}"

                    # Ajustar opción de venta si el precio ajustado no cumple la ganancia mínima
                    opc_venta_mostrar = senal['opc_venta']
                    precio_compra_min_cartera = senal.get('precio_compra_minimo', 0)
                    ganancia_min_param = senal.get('ganancia_min_pct', 0)
                    if venta_ajustada and precio_compra_min_cartera > 0:
                        # Calcular precio mínimo de venta para cumplir ganancia mínima
                        precio_venta_minimo_req = precio_compra_min_cartera * (1 + ganancia_min_param / 100)
                        if precio_venta_mostrar < precio_venta_minimo_req:
                            opc_venta_mostrar = "ESPERAR"

                    tree.insert("", "end", values=(
                        senal['symbol'],
                        senal['acciones_cartera'],
                        f"${cierre:.2f}",
                        str_compra,
                        senal['cant_compra'],
                        senal['opc_compra'],
                        str_venta,
                        senal['cant_venta'],
                        opc_venta_mostrar,
                        senal.get('tendencia', 'N/A'),
                        senal.get('tendencia_larga', 'N/A')
                    ))
                else:
                    tree.insert("", "end", values=(
                        senal['symbol'],
                        senal.get('acciones_cartera', 0),
                        senal.get('cierre', 'N/A'),
                        "-", "-",
                        senal.get('opc_compra', 'N/A'),
                        "-", "-",
                        senal.get('opc_venta', 'N/A'),
                        senal.get('tendencia', 'N/A'),
                        senal.get('tendencia_larga', 'N/A')
                    ))
            # Actualizar texto de pestaña
            idx = int(slot_id) - 1
            nombre = obtener_nombre_slot(datos_slots, slot_id)
            notebook.tab(idx, text=f"{nombre} ({len(senales_ordenadas)})")

    def toggle_ver_anteriores():
        """Alterna entre señales actuales y guardadas (muestra la fecha anterior a la más reciente)"""
        if ver_ant_var.get():
            # Cargar señales guardadas de la fecha ANTERIOR (no la más reciente que es igual a la actual)
            datos_senales = cargar_historial_senales()

            # Recopilar todas las fechas únicas de todos los slots
            todas_fechas = set()
            for slot_id in ["1", "2", "3", "4", "5"]:
                for s in datos_senales.get("senales_por_slot", {}).get(slot_id, []):
                    todas_fechas.add(s.get("fecha_generacion", "")[:10])

            fechas_ordenadas = sorted(todas_fechas, reverse=True)
            if len(fechas_ordenadas) < 2:
                messagebox.showinfo("Sin datos", "No hay señales anteriores guardadas.")
                ver_ant_var.set(False)
                return

            # Usar la segunda fecha más reciente (la anterior a la actual)
            fecha_anterior = fechas_ordenadas[1]

            senales_guardadas = {}
            for slot_id in ["1", "2", "3", "4", "5"]:
                senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
                senales_fecha = [s for s in senales_slot if s.get("fecha_generacion", "")[:10] == fecha_anterior]
                senales_convertidas = []
                for s in senales_fecha:
                    senales_convertidas.append({
                        'symbol': s.get('symbol', ''),
                        'cierre': s.get('precio_cierre', 0),
                        'precio_compra': s.get('precio_compra_sugerido', 0),
                        'cant_compra': s.get('cant_compra', '-'),
                        'opc_compra': s.get('opc_compra', ''),
                        'precio_venta': s.get('precio_venta_sugerido', 0),
                        'cant_venta': s.get('cant_venta', '-'),
                        'opc_venta': s.get('opc_venta', ''),
                        'acciones_cartera': s.get('acciones_cartera', 0),
                        'tendencia': s.get('tendencia', 'N/A'),
                        'tendencia_larga': s.get('tendencia_larga', 'N/A'),
                        'estado': 'OK'
                    })
                senales_guardadas[slot_id] = senales_convertidas
            poblar_trees(senales_guardadas)
            lbl_info.config(text=f"⮜ Señales anteriores (guardadas para: {fecha_anterior})")
        else:
            poblar_trees(senales_por_slot)
            lbl_info.config(text=f"Señales actuales (recién calculadas)")

    # Poblar con señales actuales
    poblar_trees(senales_por_slot)

    # Frame de botones
    frame_botones = tk.Frame(ventana_senales, pady=10)
    frame_botones.pack(fill="x", padx=10)

    def exportar_excel():
        """Exporta las señales a Excel con una hoja por slot"""
        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Señales",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Senales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )

        if not ruta_excel:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            primera_hoja = True

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            compra_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            venta_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            for slot_id in ["1", "2", "3", "4", "5"]:
                nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
                senales = senales_por_slot.get(slot_id, [])

                if not senales:
                    continue

                if primera_hoja:
                    ws = wb.active
                    ws.title = f"Slot {nombre_slot}"
                    primera_hoja = False
                else:
                    ws = wb.create_sheet(f"Slot {nombre_slot}")

                ws.cell(row=1, column=1, value=f"Señales generadas: {fecha_generacion} - Slot: {nombre_slot}")
                ws.cell(row=1, column=1).font = Font(bold=True)

                headers = ["Symbol", "Cartera", "Cierre", "P.Compra", "Cant.C", "Opc.Compra", "P.Venta", "Cant.V", "Opc.Venta", "Tendencia"]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                for row_idx, senal in enumerate(sorted(senales, key=lambda x: x.get('symbol', '').upper()), 4):
                    ws.cell(row=row_idx, column=1, value=senal['symbol']).border = border
                    ws.cell(row=row_idx, column=2, value=senal['acciones_cartera']).border = border

                    if senal['estado'] == 'OK':
                        ws.cell(row=row_idx, column=3, value=senal['cierre']).border = border
                        ws.cell(row=row_idx, column=4, value=senal['precio_compra']).border = border
                        ws.cell(row=row_idx, column=5, value=senal['cant_compra']).border = border
                        opc_c = ws.cell(row=row_idx, column=6, value=senal['opc_compra'])
                        opc_c.border = border
                        if senal['opc_compra'] == "Comprar":
                            opc_c.fill = compra_fill
                        ws.cell(row=row_idx, column=7, value=senal['precio_venta']).border = border
                        ws.cell(row=row_idx, column=8, value=senal['cant_venta']).border = border
                        opc_v = ws.cell(row=row_idx, column=9, value=senal['opc_venta'])
                        opc_v.border = border
                        if senal['opc_venta'] == "Vender":
                            opc_v.fill = venta_fill
                        ws.cell(row=row_idx, column=10, value=senal.get('tendencia', 'N/A')).border = border
                    else:
                        ws.cell(row=row_idx, column=3, value=senal['cierre']).border = border
                        for c in range(4, 10):
                            ws.cell(row=row_idx, column=c, value="-").border = border
                        ws.cell(row=row_idx, column=10, value=senal.get('tendencia', 'N/A')).border = border

                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 12

            wb.save(ruta_excel)
            messagebox.showinfo("Exportado", f"Señales exportadas a:\n{ruta_excel}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")

    tk.Button(frame_botones, text="Exportar a Excel", command=exportar_excel,
              bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_senales.destroy).pack(side="right", padx=5)

    # Nota informativa
    frame_nota = tk.Frame(ventana_senales, pady=5)
    frame_nota.pack(fill="x", padx=10)
    tk.Label(frame_nota,
             text="Nota: Cada pestaña muestra las señales de un slot de Parámetros Activos",
             font=("Arial", 9), fg="gray").pack(anchor="w")


def comparar_senales_operaciones():
    """Abre ventana para comparar señales generadas con operaciones reales (con 5 pestañas por slot)"""

    ruta_senales = obtener_ruta_senales()
    if ruta_senales is None:
        messagebox.showerror("Error", "No hay ubicación configurada.\nEjecuta primero Analisis_singrafico.py")
        return

    # Cargar datos de slots
    datos_slots, error = cargar_parametros_activos()
    if error:
        messagebox.showerror("Error", f"Error cargando parámetros: {error}")
        return

    datos_senales = cargar_historial_senales()
    operaciones = cargar_historial_operaciones()

    # Verificar que haya al menos algunas señales en algún slot
    hay_senales = any(
        len(datos_senales.get("senales_por_slot", {}).get(slot_id, [])) > 0
        for slot_id in ["1", "2", "3", "4", "5"]
    )
    if not hay_senales:
        messagebox.showinfo("Sin datos", "No hay señales guardadas.\nGenera señales primero con el botón 'Generar Señales'.")
        return

    # Cargar datos de precios del log (portable)
    precios_df = None
    fechas_con_cierre = set()
    if os.path.exists(str(AUTO_UPDATE_LOG_PORTABLE)):
        try:
            precios_df = pd.read_csv(str(AUTO_UPDATE_LOG_PORTABLE), parse_dates=['Date'])
            precios_df['Date'] = pd.to_datetime(precios_df['Date']).dt.strftime('%Y-%m-%d')
            fechas_con_cierre = set(precios_df['Date'].unique())
            print(f"[INFO] Precios cargados desde: {AUTO_UPDATE_LOG_PORTABLE}")
        except Exception as e:
            print(f"[WARN] No se pudo cargar log de precios: {e}")

    # Filtrar señales: solo mostrar las que tienen precio de cierre en el log
    if fechas_con_cierre:
        for slot_id in ["1", "2", "3", "4", "5"]:
            senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
            senales_filtradas = [
                s for s in senales_slot
                if s.get("fecha_generacion", "")[:10] in fechas_con_cierre
            ]
            datos_senales["senales_por_slot"][slot_id] = senales_filtradas

    # Crear ventana
    ventana_comp = tk.Toplevel(root)
    ventana_comp.title("Comparación: Señales vs Operaciones Reales")
    ventana_comp.geometry("1450x700")

    # Contar señales totales
    total_senales = sum(
        len(datos_senales.get("senales_por_slot", {}).get(slot_id, []))
        for slot_id in ["1", "2", "3", "4", "5"]
    )

    # Frame superior con info
    frame_info = tk.Frame(ventana_comp, pady=5)
    frame_info.pack(fill="x", padx=10)

    tk.Label(frame_info, text=f"Total señales: {total_senales}  |  Total operaciones: {len(operaciones)}",
             font=("Arial", 10, "bold")).pack(side="left")

    # Recopilar tickers y fechas únicos para filtros
    todos_tickers = set()
    todas_fechas = set()
    for slot_id in ["1", "2", "3", "4", "5"]:
        for sen in datos_senales.get("senales_por_slot", {}).get(slot_id, []):
            todos_tickers.add(sen.get("symbol", ""))
            todas_fechas.add(sen.get("fecha_generacion", "")[:10])
    lista_tickers = ["Todos"] + sorted(todos_tickers)
    lista_fechas = ["Todos"] + sorted(todas_fechas, reverse=True)

    # Frame de filtros
    frame_filtros = tk.Frame(ventana_comp, pady=3)
    frame_filtros.pack(fill="x", padx=10)

    tk.Label(frame_filtros, text="Filtrar por:", font=("Arial", 9)).pack(side="left", padx=(0, 5))
    tk.Label(frame_filtros, text="Ticker:", font=("Arial", 9)).pack(side="left")
    combo_filtro_ticker = ttk.Combobox(frame_filtros, values=lista_tickers, state="readonly", width=10)
    combo_filtro_ticker.set("Todos")
    combo_filtro_ticker.pack(side="left", padx=(2, 10))

    tk.Label(frame_filtros, text="Fecha:", font=("Arial", 9)).pack(side="left")
    combo_filtro_fecha = ttk.Combobox(frame_filtros, values=lista_fechas, state="readonly", width=12)
    combo_filtro_fecha.set("Todos")
    combo_filtro_fecha.pack(side="left", padx=(2, 10))

    lbl_filtro_count = tk.Label(frame_filtros, text="", font=("Arial", 9), fg="gray")
    lbl_filtro_count.pack(side="left", padx=5)

    # Notebook principal con pestañas por slot
    notebook_principal = ttk.Notebook(ventana_comp)
    notebook_principal.pack(fill="both", expand=True, padx=10, pady=5)

    # Referencias a treeviews para filtrado
    tree_refs = {}

    # Crear pestañas para cada slot
    for slot_id in ["1", "2", "3", "4", "5"]:
        nombre_slot = obtener_nombre_slot(datos_slots, slot_id)
        senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
        cantidad_senales = len(senales_slot)

        # Frame principal del slot
        frame_slot = tk.Frame(notebook_principal)
        texto_tab = f"{nombre_slot} ({cantidad_senales})"
        notebook_principal.add(frame_slot, text=texto_tab)

        if cantidad_senales == 0:
            tk.Label(frame_slot, text="No hay señales en este slot",
                    font=("Arial", 12), fg="gray").pack(expand=True)
            continue

        # Sub-notebook con Señales y Comparación
        sub_notebook = ttk.Notebook(frame_slot)
        sub_notebook.pack(fill="both", expand=True)

        # ===== SUB-PESTAÑA: SEÑALES =====
        frame_senales = tk.Frame(sub_notebook)
        sub_notebook.add(frame_senales, text="Señales")

        scroll_sen_y = tk.Scrollbar(frame_senales, orient="vertical")
        scroll_sen_x = tk.Scrollbar(frame_senales, orient="horizontal")

        cols_sen = ("Fecha", "Symbol", "Cierre fecha", "P.Compra", "Cant.C", "Opc.Compra", "P.Venta", "Cant.V", "Opc.Venta", "Cartera", "Tendencia")
        tree_senales = ttk.Treeview(frame_senales, columns=cols_sen, show="headings",
                                     selectmode="extended",
                                     yscrollcommand=scroll_sen_y.set, xscrollcommand=scroll_sen_x.set)

        scroll_sen_y.config(command=tree_senales.yview)
        scroll_sen_x.config(command=tree_senales.xview)

        anchos_sen = {"Fecha": 85, "Symbol": 70, "Cierre fecha": 90, "P.Compra": 80, "Cant.C": 55,
                      "Opc.Compra": 85, "P.Venta": 75, "Cant.V": 55, "Opc.Venta": 80, "Cartera": 65, "Tendencia": 70}
        for col in cols_sen:
            tree_senales.heading(col, text=col)
            tree_senales.column(col, width=anchos_sen.get(col, 80), anchor="center")

        scroll_sen_y.pack(side="right", fill="y")
        scroll_sen_x.pack(side="bottom", fill="x")
        tree_senales.pack(fill="both", expand=True)

        # ===== SUB-PESTAÑA: COMPARACIÓN =====
        frame_comp = tk.Frame(sub_notebook)
        sub_notebook.add(frame_comp, text="Comparación")

        scroll_comp_y = tk.Scrollbar(frame_comp, orient="vertical")
        scroll_comp_x = tk.Scrollbar(frame_comp, orient="horizontal")

        cols_comp = ("Fecha Señal", "Symbol", "Máximo", "Mínimo", "Cierre", "P.Compra", "P.Venta", "Recomendación", "Tendencia")
        tree_comp = ttk.Treeview(frame_comp, columns=cols_comp, show="headings",
                                  yscrollcommand=scroll_comp_y.set, xscrollcommand=scroll_comp_x.set)

        scroll_comp_y.config(command=tree_comp.yview)
        scroll_comp_x.config(command=tree_comp.xview)

        for col in cols_comp:
            tree_comp.heading(col, text=col)
            tree_comp.column(col, width=90, anchor="center")

        scroll_comp_y.pack(side="right", fill="y")
        scroll_comp_x.pack(side="bottom", fill="x")
        tree_comp.pack(fill="both", expand=True)

        # Guardar referencias
        tree_refs[slot_id] = {"senales": tree_senales, "comp": tree_comp, "nombre": nombre_slot}

    def poblar_arboles(filtro_ticker="Todos", filtro_fecha="Todos"):
        """Limpia y repuebla todos los treeviews según los filtros seleccionados"""
        total_mostradas = 0

        for slot_id, refs in tree_refs.items():
            tree_sen = refs["senales"]
            tree_cmp = refs["comp"]
            nombre_slot = refs["nombre"]

            # Limpiar árboles
            tree_sen.delete(*tree_sen.get_children())
            tree_cmp.delete(*tree_cmp.get_children())

            senales_slot = datos_senales.get("senales_por_slot", {}).get(slot_id, [])
            senales_ordenadas = sorted(senales_slot, key=lambda x: (x.get("symbol", "").upper(), x.get("fecha_generacion", "")[:10]))

            # Aplicar filtros
            if filtro_ticker != "Todos":
                senales_ordenadas = [s for s in senales_ordenadas if s.get("symbol", "") == filtro_ticker]
            if filtro_fecha != "Todos":
                senales_ordenadas = [s for s in senales_ordenadas if s.get("fecha_generacion", "")[:10] == filtro_fecha]

            count_slot = len(senales_ordenadas)
            total_mostradas += count_slot

            # Actualizar texto de la pestaña
            idx_tab = int(slot_id) - 1
            notebook_principal.tab(idx_tab, text=f"{nombre_slot} ({count_slot})")

            # Poblar pestaña Señales
            for sen in senales_ordenadas:
                fecha_completa = sen.get("fecha_generacion", "")
                fecha_senal = fecha_completa[:10]
                symbol = sen.get("symbol", "")

                cierre_real = "-"
                if precios_df is not None:
                    precio_row = precios_df[(precios_df['Date'] == fecha_senal) & (precios_df['Ticker'] == symbol)]
                    if not precio_row.empty:
                        cierre_real = f"${precio_row['Close'].iloc[0]:.2f}"

                tree_sen.insert("", "end", values=(
                    fecha_senal,
                    symbol,
                    cierre_real,
                    f"${sen.get('precio_compra_sugerido', 0):.2f}",
                    sen.get("cant_compra", "-"),
                    sen.get("opc_compra", ""),
                    f"${sen.get('precio_venta_sugerido', 0):.2f}",
                    sen.get("cant_venta", "-"),
                    sen.get("opc_venta", ""),
                    sen.get("acciones_cartera", 0),
                    sen.get("tendencia", "N/A")
                ))

            # Poblar pestaña Comparación
            for sen in senales_ordenadas:
                fecha_sen = sen.get("fecha_generacion", "")[:10]
                symbol = sen.get("symbol", "")

                precio_max = 0
                precio_min = 0
                precio_cierre = sen.get("precio_cierre", 0)

                if precios_df is not None:
                    precio_dia = precios_df[(precios_df['Date'] == fecha_sen) & (precios_df['Ticker'] == symbol)]
                    if not precio_dia.empty:
                        precio_max = precio_dia['High'].values[0]
                        precio_min = precio_dia['Low'].values[0]
                        precio_cierre = precio_dia['Close'].values[0]

                if sen.get("opc_compra") == "Comprar":
                    recomendacion = "Comprar"
                elif sen.get("opc_venta") == "Vender":
                    recomendacion = "Vender"
                else:
                    recomendacion = "Sin acción"

                tree_cmp.insert("", "end", values=(
                    fecha_sen,
                    symbol,
                    f"${precio_max:.2f}" if precio_max > 0 else "-",
                    f"${precio_min:.2f}" if precio_min > 0 else "-",
                    f"${precio_cierre:.2f}" if precio_cierre > 0 else "-",
                    f"${sen.get('precio_compra_sugerido', 0):.2f}",
                    f"${sen.get('precio_venta_sugerido', 0):.2f}",
                    recomendacion,
                    sen.get("tendencia", "N/A")
                ))

        # Actualizar etiqueta de filtro
        if filtro_ticker != "Todos" or filtro_fecha != "Todos":
            lbl_filtro_count.config(text=f"(Mostrando {total_mostradas} de {total_senales})")
        else:
            lbl_filtro_count.config(text="")

    def on_filtro_change(event=None):
        """Callback cuando cambia un filtro"""
        poblar_arboles(combo_filtro_ticker.get(), combo_filtro_fecha.get())

    combo_filtro_ticker.bind("<<ComboboxSelected>>", on_filtro_change)
    combo_filtro_fecha.bind("<<ComboboxSelected>>", on_filtro_change)

    # Poblar árboles inicialmente (sin filtro)
    poblar_arboles()

    # Frame de botones
    frame_botones = tk.Frame(ventana_comp, pady=10)
    frame_botones.pack(fill="x", padx=10)

    def limpiar_historial_senales():
        """Limpia el historial de señales (todos los slots)"""
        if not messagebox.askyesno("Confirmar", "¿Eliminar todo el historial de señales de TODOS los slots?"):
            return

        ruta = obtener_ruta_senales()
        if ruta and ruta.exists():
            try:
                with open(ruta, 'w', encoding='utf-8') as f:
                    json.dump(crear_estructura_senales_vacia(), f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Limpiado", "Historial de señales eliminado.")
                ventana_comp.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Error: {e}")

    tk.Button(frame_botones, text="Limpiar Todo", command=limpiar_historial_senales,
              bg="#dc3545", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_comp.destroy).pack(side="right", padx=5)


def seleccionar_csv():
    # Obtener ruta guardada para usar como directorio inicial
    ruta_guardada = cargar_ruta_csv()
    initial_dir = os.path.dirname(ruta_guardada) if ruta_guardada else None

    ruta = filedialog.asksaveasfilename(
        title="Selecciona o crea el archivo CSV",
        defaultextension=".csv",
        filetypes=[("CSV files", "*.csv"), ("Todos los archivos", "*.*")],
        initialdir=initial_dir
    )
    if ruta:
        entry_ruta.delete(0, tk.END)
        entry_ruta.insert(0, ruta)
        # Guardar la ruta para la próxima vez
        guardar_ruta_csv(ruta)

def actualizar_csv():
    """
    Descarga y actualiza datos de precios.
    - Antes de 16:00 NY: muestra advertencia de precios preliminares
    - Después de 16:00 NY: sobrescribe automáticamente datos de hoy
    """
    csv_file = entry_ruta.get()
    if not csv_file:
        label_status.config(text="Selecciona primero la ruta del CSV", fg="red")
        return

    # Verificar hora de NY
    now_ny = datetime.now(ZoneInfo("America/New_York"))
    hora_ny = now_ny.hour
    es_fin_de_semana = now_ny.weekday() >= 5  # 5=Sábado, 6=Domingo

    # Después de 16:00 = sobrescribir automáticamente (precios de cierre definitivos)
    forzar_actualizacion = (hora_ny >= 16)

    if es_fin_de_semana:
        respuesta = messagebox.askyesno(
            "Fin de semana",
            f"Hoy es {['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo'][now_ny.weekday()]}.\n"
            "El mercado está cerrado los fines de semana.\n\n"
            "¿Deseas descargar los datos del viernes pasado?"
        )
        if not respuesta:
            return
    elif hora_ny < 16:
        respuesta = messagebox.askyesno(
            "Mercado aún abierto",
            f"Hora actual en NY: {now_ny.strftime('%H:%M')}\n\n"
            "El mercado cierra a las 16:00 NY.\n"
            "Los precios descargados ahora son PRELIMINARES\n"
            "(el precio 'Close' será el último precio negociado, no el de cierre).\n\n"
            "¿Deseas descargar los precios preliminares?\n\n"
            "Después de las 16:00, los datos se sobrescribirán\n"
            "automáticamente con los precios de cierre definitivos."
        )
        if not respuesta:
            return

    try:
        print("\n=== INICIO ACTUALIZACIÓN ===")

        print("[1] Descargando datos de Yahoo Finance...")
        data = yf.download(tickers, period="1d", group_by='ticker', auto_adjust=False)
        print("[2] Descarga completada.")

        records = []
        for ticker in tickers:
            if hasattr(data.columns, "levels") and ticker in data.columns.levels[0]:
                df = data[ticker].copy()
                df.reset_index(inplace=True)
                df.rename(columns={'Adj Close':'Close'}, inplace=True)
                df['Ticker'] = ticker
                records.append(df[['Date','Ticker','Open','High','Low','Close']])
            else:
                if 'Open' in data.columns and 'High' in data.columns and 'Low' in data.columns and 'Close' in data.columns:
                    tmp = data.reset_index().copy()
                    tmp.rename(columns={'Adj Close':'Close'}, inplace=True)
                    tmp['Ticker'] = ticker
                    if not tmp.empty:
                        records.append(tmp[['Date','Ticker','Open','High','Low','Close']])
                    break

        if not records:
            print("[X] No se encontraron datos.")
            label_status.config(text="No hay datos nuevos disponibles hoy.", fg="blue")
            return

        df_long = pd.concat(records, ignore_index=True)
        df_long = df_long.loc[:, ~df_long.columns.duplicated()]
        df_long['Date'] = pd.to_datetime(df_long['Date']).dt.normalize()

        # ===========================
        # CREAR CSV PRINCIPAL
        # ===========================
        if not os.path.exists(csv_file):
            print("[3] CSV no existe, se creará uno nuevo.")
        else:
            print("[3] CSV ya existe, será sobrescrito con la data descargada.")

        print("[5] Creando CSV con la data descargada...")
        df_long.to_csv(csv_file, index=False, float_format="%.2f")
        print("[6] CSV guardado correctamente.")


        # ===========================
        # ACTUALIZAR LOG AUXILIAR
        # ===========================
        log_file = os.path.join(os.path.dirname(csv_file), "auto_update_log.csv")
        print(f"[7] Actualizando log auxiliar: {log_file}")

        df_long_for_log = df_long.copy()
        df_long_for_log['Date'] = pd.to_datetime(df_long_for_log['Date']).dt.normalize()

        if os.path.exists(log_file):
            print("[8] Leyendo log existente...")
            df_log_existing = pd.read_csv(log_file, parse_dates=['Date'])
            df_log_existing = df_log_existing.loc[:, ~df_log_existing.columns.duplicated()]
            df_log_existing['Date'] = pd.to_datetime(df_log_existing['Date']).dt.normalize()

            # Si forzar_actualizacion, eliminar datos de hoy antes de agregar nuevos
            if forzar_actualizacion:
                fecha_hoy = df_long_for_log['Date'].iloc[0]
                filas_antes = len(df_log_existing)
                df_log_existing = df_log_existing[df_log_existing['Date'] != fecha_hoy]
                filas_eliminadas = filas_antes - len(df_log_existing)
                if filas_eliminadas > 0:
                    print(f"[8.1] Sobrescribiendo: eliminados {filas_eliminadas} registros de {fecha_hoy.strftime('%Y-%m-%d')}")

            existing_keys = set(zip(
                df_log_existing['Date'].dt.strftime('%Y-%m-%d'),
                df_log_existing['Ticker']
            ))

            keys_series = df_long_for_log[['Date','Ticker']].apply(
                lambda r: (r['Date'].strftime('%Y-%m-%d'), r['Ticker']), axis=1
            )

            mask_new = ~keys_series.isin(existing_keys)
            df_log_new = df_long_for_log.loc[mask_new].copy()

            if not df_log_new.empty:
                print(f"[9] Agregando {len(df_log_new)} filas nuevas al log.")
                df_log_to_save = pd.concat([df_log_existing, df_log_new], ignore_index=True)
            else:
                if forzar_actualizacion:
                    # Si se forzó actualización pero no hay filas "nuevas", agregar igualmente
                    print(f"[9] Sobrescribiendo {len(df_long_for_log)} registros de hoy.")
                    df_log_to_save = pd.concat([df_log_existing, df_long_for_log], ignore_index=True)
                else:
                    print("[9] No hay filas nuevas para agregar al log.")
                    df_log_to_save = df_log_existing.copy()

        else:
            print("[8] Log no existe. Creándolo desde cero.")
            df_log_to_save = df_long_for_log.copy()

        print("[10] Guardando log auxiliar...")
        df_log_to_save.to_csv(log_file, index=False, float_format="%.2f")
        print("[11] Log guardado correctamente.")

        # Liberar memoria
        gc.collect()

        # Hora NY
        now_ny = datetime.now(ZoneInfo("America/New_York"))
        fecha_hora_ny = now_ny.strftime("%Y-%m-%d %H:%M")
        label_status.config(
            text=f"CSV actualizado con fecha y hora de Nueva York: {fecha_hora_ny}",
            fg="blue"
        )

        print("=== FIN ACTUALIZACIÓN ===\n")

        mostrar_datos_en_tabla(csv_file)

    except Exception as e:
        print(f"[ERROR GENERAL] {str(e)}")
        label_status.config(text=f"Error: {str(e)}", fg="red")

def mostrar_datos_en_tabla(csv_file):
    df = pd.read_csv(csv_file)

    # Limpiar tabla
    for row in tree.get_children():
        tree.delete(row)

    # Insertar filas
    for _, row in df.iterrows():
        tree.insert(
            "", tk.END,
            values=(
                row['Date'],
                row['Ticker'],
                f"{row['Open']:.2f}",
                f"{row['High']:.2f}",
                f"{row['Low']:.2f}",
                f"{row['Close']:.2f}"
            )
        )

# Crear ventana principal
root = tk.Tk()
root.title("Actualizar precios de acciones")

# Frame para selección de archivo
frame1 = tk.Frame(root)
frame1.pack(pady=10, padx=10, fill="x")
tk.Label(frame1, text="Ruta del CSV:").pack(anchor="w")
entry_ruta = tk.Entry(frame1, width=60)
entry_ruta.pack(side="left", padx=(0,5))
tk.Button(frame1, text="Seleccionar CSV", command=seleccionar_csv).pack(side="left")

# Cargar última ruta guardada
ruta_guardada = cargar_ruta_csv()
if ruta_guardada and os.path.exists(ruta_guardada):
    entry_ruta.insert(0, ruta_guardada)

# Frame para editar tickers
frame_tickers = tk.Frame(root)
frame_tickers.pack(padx=10, pady=5, fill="x")

tk.Label(frame_tickers, text="Tickers actuales:").pack(anchor="w")

# Lista de tickers visible
listbox_tickers = tk.Listbox(frame_tickers, height=10)
listbox_tickers.pack(side="left", fill="y")
for t in tickers:
    listbox_tickers.insert(tk.END, t)

# Scrollbar para listbox
scroll_tickers = tk.Scrollbar(frame_tickers, orient="vertical", command=listbox_tickers.yview)
scroll_tickers.pack(side="left", fill="y")
listbox_tickers.config(yscrollcommand=scroll_tickers.set)

# Frame para botones de gestión de tickers
frame_ticker_btns = tk.Frame(frame_tickers)
frame_ticker_btns.pack(side="left", padx=10)

entry_nuevo_ticker = tk.Entry(frame_ticker_btns, width=10)
entry_nuevo_ticker.pack(pady=(0,5))

def agregar_ticker():
    nuevo = entry_nuevo_ticker.get().strip().upper()
    if not nuevo:
        label_status.config(text="Ingresa un ticker válido.", fg="red")
        return
    if nuevo in tickers:
        label_status.config(text=f"{nuevo} ya está en la lista.", fg="orange")
        return
    # Verificación rápida con Yahoo Finance
    try:
        df_test = yf.download(nuevo, period="1d", progress=False)
        if df_test.empty:
            raise ValueError("No hay datos para este ticker")
    except Exception:
        label_status.config(text=f"Ticker inválido: {nuevo}", fg="red")
        return

    # Si pasa la verificación, se agrega
    tickers.append(nuevo)
    listbox_tickers.insert(tk.END, nuevo)
    entry_nuevo_ticker.delete(0, tk.END)
    label_status.config(text=f"Ticker agregado: {nuevo}", fg="green")


def quitar_ticker():
    seleccion = listbox_tickers.curselection()
    if seleccion:
        idx = seleccion[0]
        t = listbox_tickers.get(idx)
        tickers.remove(t)
        listbox_tickers.delete(idx)

tk.Button(frame_ticker_btns, text="Agregar Ticker", command=agregar_ticker).pack(pady=2)
tk.Button(frame_ticker_btns, text="Quitar Ticker", command=quitar_ticker).pack(pady=2)



# Checkbox para opción automática (activar/desactivar)
auto_var = tk.BooleanVar(value=False)
tk.Checkbutton(root, text="Actualizar automáticamente (activar/desactivar)", variable=auto_var).pack(pady=5)

def auto_actualizar():
    if auto_var.get():
        now_ny = datetime.now(ZoneInfo("America/New_York"))
        if now_ny.hour == 16 and now_ny.minute >= 10:
            actualizar_csv()
    root.after(60000, auto_actualizar)  # revisa cada 60 segundos

auto_actualizar()

# Frame para botones principales
frame_botones_principales = tk.Frame(root)
frame_botones_principales.pack(pady=5)

# Botón para actualizar CSV manualmente
tk.Button(frame_botones_principales, text="Actualizar CSV ahora", command=actualizar_csv,
          bg="lightblue", font=("Arial", 10)).pack(side="left", padx=5)

# Botón para generar señales
tk.Button(frame_botones_principales, text="Generar Señales", command=generar_senales,
          bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(side="left", padx=5)

# Botón para regenerar señales de fechas anteriores
tk.Button(frame_botones_principales, text="Regenerar Históricas", command=regenerar_senales_historicas,
          bg="#6c757d", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

# Botón para historial de operaciones
tk.Button(frame_botones_principales, text="Historial", command=administrar_historial,
          bg="#ffc107", fg="black", font=("Arial", 10)).pack(side="left", padx=5)

# Botón para comparar señales con operaciones reales
tk.Button(frame_botones_principales, text="Comparar Señales", command=comparar_senales_operaciones,
          bg="#17a2b8", fg="white", font=("Arial", 10)).pack(side="left", padx=5)

# Botón para sincronizar desde GitHub
tk.Button(frame_botones_principales, text="Sync GitHub", command=sincronizar_desde_github,
          bg="#6f42c1", fg="white", font=("Arial", 9)).pack(side="left", padx=5)

# Label para mensajes de estado
label_status = tk.Label(root, text="", fg="blue")
label_status.pack(pady=5)

# Frame para tabla
frame_table = tk.Frame(root)
frame_table.pack(padx=10, pady=10, fill="both", expand=True)

columns = ("Date", "Ticker", "Open", "High", "Low", "Close")
tree = ttk.Treeview(frame_table, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, anchor="center", width=80)

# Scrollbars
scroll_y = ttk.Scrollbar(frame_table, orient="vertical", command=tree.yview)
scroll_x = ttk.Scrollbar(frame_table, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
tree.pack(side="left", fill="both", expand=True)
scroll_y.pack(side="right", fill="y")
scroll_x.pack(side="bottom", fill="x")

root.mainloop()
