#!/usr/bin/env python3
"""
Script headless para análisis y optimización de parámetros de trading.
Replica la lógica de Analisis_de_Acciones.py sin interfaz gráfica.

Autor: Sistema de Análisis de Inversiones
Fecha: 01/03/2026
Versión: 1.0.0

Uso:
    python analizar_ticker_headless.py DATA/AAPL/Datos_AAPL_FEB25_FEB26.csv
    python analizar_ticker_headless.py DATA/AAPL/Datos_AAPL_FEB25_FEB26.csv --limite 10
"""

import pandas as pd
import numpy as np
import sqlite3
import os
import sys
import time
import argparse
from datetime import datetime, timedelta
from pathlib import Path

# Importar scipy para optimización
from scipy.optimize import differential_evolution

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

EXPECTED_COLUMNS = ["Fecha", "Último", "Apertura", "Máximo", "Mínimo", "Vol.", "% var."]

# Parámetros por defecto
DEFAULT_LIMITE_ACCIONES = 10
DEFAULT_SUAVE = 0.5  # 0.5%

# Bounds para optimización (Auto en todos)
BOUNDS_COMPRA = (-3.0, 0.0)      # % de compra
BOUNDS_VENTA = (0.0, 3.0)        # % de venta
BOUNDS_GANANCIA = (1.5, 3.0)    # % ganancia mínima (máx 3%)
BOUNDS_COMPRA_MULT = (0, 5)     # Múltiplo de compra
BOUNDS_VENTA_MULT = (0, 5)      # Múltiplo de venta

# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================

def log(mensaje):
    """Imprime mensaje con timestamp"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {mensaje}")


def to_float_safe(val):
    """Convierte a float de forma segura"""
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        val_str = str(val).replace(",", ".").replace(" ", "").replace("−", "-")
        return float(val_str)
    except:
        return 0.0


def parse_percent_to_decimal(val):
    """Convierte porcentaje string a decimal"""
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val) / 100 if abs(float(val)) > 1 else float(val)
    try:
        val_str = str(val).replace(",", ".").replace("%", "").replace(" ", "").replace("−", "-")
        return float(val_str) / 100
    except:
        return 0.0


def cargar_csv(filepath):
    """Carga y preprocesa el CSV"""
    try:
        df = pd.read_csv(filepath, sep=";", engine='python', dtype=str, encoding='utf-8-sig')
    except:
        df = pd.read_csv(filepath, sep=";", engine='python', dtype=str, encoding='latin-1')

    # Limpiar nombres de columnas
    df.columns = [c.strip() for c in df.columns]

    # Verificar columnas
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Columnas faltantes: {missing}")

    df = df[EXPECTED_COLUMNS].copy()

    # Procesar fechas
    def parse_date(date_str):
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return pd.to_datetime(date_str, format=fmt)
            except:
                continue
        return pd.NaT

    df['Fecha'] = df['Fecha'].apply(parse_date)
    df = df.dropna(subset=['Fecha'])
    df = df.sort_values('Fecha').reset_index(drop=True)

    # Convertir numéricos
    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(to_float_safe)

    df['% var.'] = df['% var.'].apply(parse_percent_to_decimal)

    return df


def filtrar_ultimos_dias(df, dias):
    """Filtra el DataFrame para los últimos N días"""
    fecha_max = df['Fecha'].max()
    fecha_corte = fecha_max - timedelta(days=dias)
    return df[df['Fecha'] >= fecha_corte].copy().reset_index(drop=True)


def calcular_acumulado(df):
    """Calcula el % acumulado por secuencias de mismo signo"""
    acum = 0
    prev_sign = 0
    lst = []

    for v in df['% var.']:
        sign = 1 if v > 0 else -1 if v < 0 else 0
        if sign == prev_sign:
            acum += v
        else:
            acum = v
        lst.append(acum)
        prev_sign = sign

    df['% acumulado'] = lst
    return df


def calcular_promedios(df):
    """Calcula promedio de máximos y mínimos de secuencias"""
    acum_decimal = df['% acumulado'].astype(float)

    # Promedio de máximos (secuencias positivas >= 2 días)
    valores_maximos = []
    seq = []
    for v in acum_decimal:
        if v > 0:
            seq.append(v)
        else:
            if len(seq) >= 2:
                valores_maximos.append(seq[-1] * 100.0)
            seq = []
    if len(seq) >= 2:
        valores_maximos.append(seq[-1] * 100.0)

    promedio_maximos = sum(valores_maximos) / len(valores_maximos) if valores_maximos else 0.0

    # Promedio de mínimos (secuencias negativas >= 2 días)
    valores_minimos = []
    seq_neg = []
    for v in acum_decimal:
        if v < 0:
            seq_neg.append(v)
        else:
            if len(seq_neg) >= 2:
                valores_minimos.append(seq_neg[-1] * 100.0)
            seq_neg = []
    if len(seq_neg) >= 2:
        valores_minimos.append(seq_neg[-1] * 100.0)

    promedio_minimos = sum(valores_minimos) / len(valores_minimos) if valores_minimos else 0.0

    return promedio_maximos, promedio_minimos


# =============================================================================
# SIMULACIÓN DE OPERACIONES
# =============================================================================

def simular_operaciones(df, compra_pct, venta_pct, ganancia_min_pct,
                        compra_mult, venta_mult, limite_acciones, suave_pct=0.5):
    """
    Simula operaciones de compra/venta y calcula rentabilidad.

    Args:
        df: DataFrame con datos del ticker
        compra_pct: Porcentaje umbral de compra (negativo, ej: -1.6)
        venta_pct: Porcentaje umbral de venta (positivo, ej: 1.6)
        ganancia_min_pct: Ganancia mínima requerida para vender (ej: 2.0)
        compra_mult: Número de acciones a comprar cuando % acum <= promedio_min (None para 1)
        venta_mult: Número de acciones a vender cuando % acum >= promedio_max (None para 1)
        limite_acciones: Máximo de acciones a mantener
        suave_pct: Umbral suave para decisiones (default 0.5%)

    Returns:
        DataFrame con simulación, rentabilidad_max, margen_promedio
    """
    df = df.copy()
    df = calcular_acumulado(df)

    promedio_maximos, promedio_minimos = calcular_promedios(df)

    # Convertir a decimales
    umbral_compra = compra_pct / 100
    umbral_venta = venta_pct / 100
    ganancia_minima = ganancia_min_pct / 100
    suave = suave_pct / 100

    # Determinar días con compra/venta múltiple
    acum_pct = df['% acumulado'].astype(float) * 100.0

    comprar_multiple = [False] * len(df)
    if promedio_minimos < 0.0 and compra_mult is not None:
        seq_idxs = []
        for idx, v in enumerate(acum_pct):
            if v < 0:
                seq_idxs.append(idx)
            else:
                if len(seq_idxs) >= 2:
                    for i in seq_idxs:
                        if acum_pct.iloc[i] <= promedio_minimos:
                            comprar_multiple[i] = True
                seq_idxs = []
        if len(seq_idxs) >= 2:
            for i in seq_idxs:
                if acum_pct.iloc[i] <= promedio_minimos:
                    comprar_multiple[i] = True

    vender_multiple = [False] * len(df)
    if promedio_maximos > 0.0 and venta_mult is not None:
        seq_idxs = []
        for idx, v in enumerate(acum_pct):
            if v > 0:
                seq_idxs.append(idx)
            else:
                if len(seq_idxs) >= 2:
                    for i in seq_idxs:
                        if acum_pct.iloc[i] >= promedio_maximos:
                            vender_multiple[i] = True
                seq_idxs = []
        if len(seq_idxs) >= 2:
            for i in seq_idxs:
                if acum_pct.iloc[i] >= promedio_maximos:
                    vender_multiple[i] = True

    # Función para determinar opción
    def determinar_opcion(var_pct, acum):
        if var_pct >= umbral_venta:
            return "Venta"
        if var_pct <= umbral_compra:
            return "Compra"
        if acum >= umbral_venta and var_pct >= suave:
            return "Venta"
        if acum <= umbral_compra and var_pct <= -suave:
            return "Compra"
        return "N/A"

    df['Opción'] = df.apply(lambda r: determinar_opcion(r['% var.'], r['% acumulado']), axis=1)

    # Simular operaciones
    acciones = 0
    capital_bolsa = 0
    aporte_acumulado = 0
    precios_en_cartera = []

    movs, acts, cap_b, cap_acc, cap_tot, aport, aport_acum, precios_compra = [], [], [], [], [], [], [], []

    for idx, row in df.iterrows():
        opcion = row["Opción"]
        precio = row["Último"]
        movimiento = 0
        aporte = 0.0
        precio_operacion = 0.0

        if opcion == "Compra":
            n_compra = compra_mult if (compra_mult is not None and comprar_multiple[idx]) else 1

            acciones_a_comprar = 0
            for _ in range(n_compra):
                if acciones < limite_acciones:
                    acciones_a_comprar += 1
                    if capital_bolsa >= precio:
                        capital_bolsa -= precio
                    else:
                        aporte += precio
                        aporte_acumulado += precio
                        capital_bolsa += precio
                        capital_bolsa -= precio
                    acciones += 1
                    precios_en_cartera.append(precio)
                    precios_en_cartera.sort()
                else:
                    break

            movimiento = acciones_a_comprar
            if movimiento > 0:
                precio_operacion = -precio

        elif opcion == "Venta" and acciones > 0:
            # Contar acciones que cumplen ganancia mínima (FIFO)
            acciones_vendibles = 0
            for precio_compra_item in precios_en_cartera:
                ganancia_porcentual = (precio - precio_compra_item) / precio_compra_item
                if ganancia_porcentual >= ganancia_minima:
                    acciones_vendibles += 1
                else:
                    break

            if acciones_vendibles > 0:
                n_venta = venta_mult if (venta_mult is not None and vender_multiple[idx] and acciones >= venta_mult) else 1
                n_venta = min(n_venta, acciones_vendibles, acciones)

                capital_bolsa += precio * n_venta
                acciones -= n_venta
                movimiento = -n_venta

                for _ in range(n_venta):
                    if precios_en_cartera:
                        precios_en_cartera.pop(0)

                if movimiento < 0:
                    precio_operacion = precio

        movs.append(movimiento)
        acts.append(acciones)
        cap_b.append(round(capital_bolsa, 2))
        cap_acc.append(round(acciones * precio, 2))
        cap_tot.append(round(capital_bolsa + acciones * precio, 2))
        aport.append(round(aporte, 2))
        aport_acum.append(round(aporte_acumulado, 2))
        precios_compra.append(precio_operacion)

    df["Movimiento de acciones"] = movs
    df["Acciones en cartera"] = acts
    df["Precio de compra"] = precios_compra
    df["Capital en bolsa"] = cap_b
    df["Capital en acciones"] = cap_acc
    df["Capital total"] = cap_tot
    df["Aporte"] = aport
    df["Aporte acumulado"] = aport_acum

    df["Margen"] = df["Capital total"] - df["Aporte acumulado"]
    df["Rentabilidad"] = df.apply(
        lambda r: (r["Margen"] / r["Aporte acumulado"] * 100) if r["Aporte acumulado"] > 0 else 0, axis=1)

    rentab_max = df["Rentabilidad"].max()
    margen_prom = df["Margen"].mean()

    return df, rentab_max, margen_prom, promedio_maximos, promedio_minimos


# =============================================================================
# OPTIMIZACIÓN CON SCIPY
# =============================================================================

def funcion_objetivo(params, df, limite_acciones, objetivo="rentabilidad"):
    """Función objetivo para optimización"""
    compra_pct = params[0]
    venta_pct = params[1]
    ganancia_min_pct = params[2]
    compra_mult = int(round(params[3])) if params[3] > 1.5 else None
    venta_mult = int(round(params[4])) if params[4] > 1.5 else None

    try:
        _, rent, margen, _, _ = simular_operaciones(
            df, compra_pct, venta_pct, ganancia_min_pct,
            compra_mult, venta_mult, limite_acciones
        )

        if objetivo == "rentabilidad":
            return -rent  # Negativo porque differential_evolution minimiza
        else:
            return -margen
    except:
        return 999999


def optimizar_parametros(df, limite_acciones, objetivo="rentabilidad", verbose=True):
    """
    Optimiza parámetros usando SciPy differential_evolution.

    Args:
        df: DataFrame con datos
        limite_acciones: Límite de acciones
        objetivo: "rentabilidad" o "margen_prom"
        verbose: Mostrar progreso

    Returns:
        dict con parámetros óptimos y métricas
    """
    bounds = [
        BOUNDS_COMPRA,
        BOUNDS_VENTA,
        BOUNDS_GANANCIA,
        BOUNDS_COMPRA_MULT,
        BOUNDS_VENTA_MULT
    ]

    evaluaciones = [0]
    mejor_valor = [float('inf')]

    def callback(xk, convergence):
        evaluaciones[0] += 1
        if verbose and evaluaciones[0] % 50 == 0:
            print(f"    Evaluaciones: {evaluaciones[0]}, Mejor: {-mejor_valor[0]:.2f}")
        return False

    def objetivo_wrapper(params):
        val = funcion_objetivo(params, df, limite_acciones, objetivo)
        if val < mejor_valor[0]:
            mejor_valor[0] = val
        return val

    resultado = differential_evolution(
        objetivo_wrapper,
        bounds,
        strategy='best1bin',
        maxiter=100,
        popsize=15,
        tol=0.01,
        mutation=(0.5, 1),
        recombination=0.7,
        seed=42,
        callback=callback,
        disp=False,
        polish=False,
        init='latinhypercube',
        workers=1
    )

    # Extraer parámetros óptimos
    compra_pct = round(resultado.x[0], 1)
    venta_pct = round(resultado.x[1], 1)
    ganancia_min_pct = round(resultado.x[2], 1)
    compra_mult = int(round(resultado.x[3])) if resultado.x[3] > 1.5 else None
    venta_mult = int(round(resultado.x[4])) if resultado.x[4] > 1.5 else None

    # Ejecutar simulación final con parámetros óptimos
    df_sim, rent_max, margen_prom, prom_max, prom_min = simular_operaciones(
        df, compra_pct, venta_pct, ganancia_min_pct,
        compra_mult, venta_mult, limite_acciones
    )

    return {
        'compra_pct': compra_pct,
        'venta_pct': venta_pct,
        'ganancia_min_pct': ganancia_min_pct,
        'compra_mult': compra_mult,
        'venta_mult': venta_mult,
        'rentabilidad_max': rent_max,
        'margen_promedio': margen_prom,
        'promedio_maximos': prom_max,
        'promedio_minimos': prom_min,
        'df_simulacion': df_sim
    }


# =============================================================================
# ANÁLISIS COMPLETO
# =============================================================================

def analizar_ticker(filepath, limite_acciones=10, verbose=True):
    """
    Ejecuta análisis completo de un ticker.

    Períodos: Completo, Últimos 6 meses, Últimos 3 meses
    Objetivos: Rentabilidad máxima, Margen promedio máximo

    Returns:
        dict con resultados por período y objetivo
    """
    log(f"Cargando {filepath}...")
    df_completo = cargar_csv(filepath)

    fecha_inicial = df_completo['Fecha'].min().strftime("%d/%m/%Y")
    fecha_final = df_completo['Fecha'].max().strftime("%d/%m/%Y")
    log(f"Datos: {len(df_completo)} registros ({fecha_inicial} - {fecha_final})")

    # Definir períodos
    periodos = {
        'completo': {'df': df_completo, 'dias': None},
        'ultimos_6_meses': {'df': filtrar_ultimos_dias(df_completo, 180), 'dias': 180},
        'ultimos_3_meses': {'df': filtrar_ultimos_dias(df_completo, 90), 'dias': 90}
    }

    objetivos = ['rentabilidad', 'margen_prom']

    resultados = {}
    total_combinaciones = len(periodos) * len(objetivos)
    combinacion_actual = 0

    for nombre_periodo, config in periodos.items():
        df_periodo = config['df']

        for objetivo in objetivos:
            combinacion_actual += 1
            obj_texto = "Rentabilidad" if objetivo == "rentabilidad" else "Margen Prom"
            log(f"[{combinacion_actual}/{total_combinaciones}] {nombre_periodo} - {obj_texto}")
            log(f"  Registros: {len(df_periodo)}")

            resultado = optimizar_parametros(df_periodo, limite_acciones, objetivo, verbose=False)

            key = f"{nombre_periodo}_{objetivo}"
            resultados[key] = {
                'periodo': nombre_periodo,
                'objetivo': objetivo,
                'fecha_inicio': df_periodo['Fecha'].min().strftime("%d/%m/%Y"),
                'fecha_fin': df_periodo['Fecha'].max().strftime("%d/%m/%Y"),
                'registros': len(df_periodo),
                **{k: v for k, v in resultado.items() if k != 'df_simulacion'},
                'df_simulacion': resultado['df_simulacion']
            }

            log(f"  Compra: {resultado['compra_pct']}%, Venta: {resultado['venta_pct']}%, "
                f"Gan.Min: {resultado['ganancia_min_pct']}%")
            log(f"  Rent.Max: {resultado['rentabilidad_max']:.2f}%, Margen: {resultado['margen_promedio']:.2f}")

    return resultados


# =============================================================================
# GUARDAR EN RESULTADO_DE_ANALISIS.JSON
# =============================================================================

def guardar_en_resultado_json(resultados, filepath_csv, ticker_symbol):
    """
    Guarda los resultados en data/Resultado_de_Analisis.json
    con el formato compatible con la GUI.
    """
    import json

    # Ruta al JSON de resultados
    json_path = os.path.join(os.path.dirname(os.path.dirname(filepath_csv)),
                             "data", "Resultado_de_Analisis.json")

    # Alternativa: buscar en la raíz del proyecto
    if not os.path.exists(os.path.dirname(json_path)):
        json_path = "data/Resultado_de_Analisis.json"

    # Cargar JSON existente o crear nuevo
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            datos_json = json.load(f)
    else:
        datos_json = {}

    # Nombre base del archivo CSV (ej: "Datos_AAPL_FEB25_FEB26")
    nombre_base = os.path.splitext(os.path.basename(filepath_csv))[0]

    # Crear estructura para este ticker
    ticker_data = {
        "_ticker_symbol": ticker_symbol
    }

    # Organizar resultados por período y objetivo
    periodos_map = {
        'completo': 'completo',
        'ultimos_6_meses': 'ultimos_6_meses',
        'ultimos_3_meses': 'ultimos_3_meses'
    }

    for periodo_key in periodos_map.keys():
        periodo_data = {}

        for objetivo in ['rentabilidad', 'margen_prom']:
            result_key = f"{periodo_key}_{objetivo}"

            if result_key in resultados:
                r = resultados[result_key]
                df_sim = r['df_simulacion']

                # Calcular estadísticas adicionales
                var_pct = df_sim['% var.'].astype(float) * 100 if df_sim['% var.'].dtype != 'float64' else df_sim['% var.'] * 100

                # Estadísticas de variación
                max_var = var_pct.max()
                min_var = var_pct.min()
                idx_max_var = var_pct.idxmax()
                idx_min_var = var_pct.idxmin()

                # Estadísticas de operaciones
                movimientos = df_sim['Movimiento de acciones']
                compras = movimientos[movimientos > 0]
                ventas = movimientos[movimientos < 0]

                # Rentabilidad como float
                rent_col = df_sim['Rentabilidad']
                if rent_col.dtype == 'object':
                    rent_values = rent_col.str.replace('%', '').astype(float)
                else:
                    rent_values = rent_col

                # Calcular promedios de variación positiva y negativa
                var_positivas = var_pct[var_pct > 0]
                var_negativas = var_pct[var_pct < 0]
                max_prom_var = round(var_positivas.mean(), 2) if len(var_positivas) > 0 else 0
                min_prom_var = round(var_negativas.mean(), 2) if len(var_negativas) > 0 else 0
                dif_prom_var = round(max_prom_var - min_prom_var, 2)

                periodo_data[objetivo] = {
                    "ticker_symbol": ticker_symbol,
                    "fecha_guardado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "fecha_inicial": r['fecha_inicio'],
                    "fecha_final": r['fecha_fin'],
                    "parametros_optimos": {
                        "compra_pct": r['compra_pct'],
                        "venta_pct": r['venta_pct'],
                        "ganancia_minima_pct": r['ganancia_min_pct'],
                        "suave_pct": 0.5,
                        "limite_tipo": "acciones",
                        "limite_valor": 10.0,
                        "compra_multiple": r['compra_mult'],
                        "venta_multiple": r['venta_mult'],
                        "promedio_maximos": round(r['promedio_maximos'] * 100, 2),
                        "promedio_minimos": round(r['promedio_minimos'] * 100, 2)
                    },
                    "metricas": {
                        "rentabilidad_max": round(r['rentabilidad_max'], 2),
                        "margen_promedio": round(r['margen_promedio'], 2),
                        "rentab_promedio": round(rent_values.mean(), 2),
                        "max_margen": round(df_sim['Margen'].max(), 2),
                        "max_aporte": round(df_sim['Aporte acumulado'].max(), 2)
                    },
                    "estadisticas_var": {
                        "max_var": round(max_var, 2),
                        "min_var": round(min_var, 2),
                        "fecha_max_var": df_sim.loc[idx_max_var, 'Fecha'].strftime("%d/%m/%Y") if hasattr(df_sim.loc[idx_max_var, 'Fecha'], 'strftime') else str(df_sim.loc[idx_max_var, 'Fecha']),
                        "fecha_min_var": df_sim.loc[idx_min_var, 'Fecha'].strftime("%d/%m/%Y") if hasattr(df_sim.loc[idx_min_var, 'Fecha'], 'strftime') else str(df_sim.loc[idx_min_var, 'Fecha']),
                        "dif_var": round(max_var - min_var, 2),
                        "max_prom_var": max_prom_var,
                        "min_prom_var": min_prom_var,
                        "dif_prom_var": dif_prom_var
                    },
                    "estadisticas_operaciones": {
                        "opc_compra": int((movimientos > 0).sum()),
                        "acciones_compradas": int(compras.sum()) if len(compras) > 0 else 0,
                        "opc_venta": int((movimientos < 0).sum()),
                        "acciones_vendidas": int(abs(ventas.sum())) if len(ventas) > 0 else 0,
                        "max_acc_cartera": int(df_sim['Acciones en cartera'].max()),
                        "fecha_max_rentab": df_sim.loc[rent_values.idxmax(), 'Fecha'].strftime("%d/%m/%Y") if hasattr(df_sim.loc[rent_values.idxmax(), 'Fecha'], 'strftime') else str(df_sim.loc[rent_values.idxmax(), 'Fecha'])
                    }
                }

        if periodo_data:
            ticker_data[periodo_key] = periodo_data

    # Agregar al JSON
    datos_json[nombre_base] = ticker_data

    # Guardar JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(datos_json, f, indent=2, ensure_ascii=False)

    log(f"Guardado en: {json_path}")
    return json_path


# =============================================================================
# GENERACIÓN DE SALIDAS (DB Y EXCEL)
# =============================================================================

def generar_db_excel(resultados, filepath_csv, carpeta_salida=None):
    """
    Genera archivos DB (SQLite) y Excel con los resultados.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Determinar carpeta de salida
    if carpeta_salida is None:
        carpeta_salida = os.path.dirname(filepath_csv)

    # Nombre base del archivo
    nombre_base = os.path.splitext(os.path.basename(filepath_csv))[0]

    # Archivo DB
    db_path = os.path.join(carpeta_salida, f"{nombre_base}_analizado.db")
    excel_path = os.path.join(carpeta_salida, f"{nombre_base}_analizado.xlsx")

    log(f"Generando DB: {db_path}")

    # Crear conexión SQLite
    conn = sqlite3.connect(db_path)

    # Guardar cada simulación como tabla
    for key, resultado in resultados.items():
        df_sim = resultado['df_simulacion'].copy()

        # Formatear para guardar
        df_sim['Fecha'] = df_sim['Fecha'].dt.strftime("%d/%m/%Y")
        df_sim['% var.'] = (df_sim['% var.'] * 100).round(2).astype(str) + "%"
        df_sim['% acumulado'] = (df_sim['% acumulado'] * 100).round(2).astype(str) + "%"
        df_sim['Rentabilidad'] = df_sim['Rentabilidad'].round(2).astype(str) + "%"

        tabla_nombre = key.replace(" ", "_")
        df_sim.to_sql(tabla_nombre, conn, if_exists='replace', index=False)

    # Guardar resumen de parámetros
    resumen = []
    for key, resultado in resultados.items():
        resumen.append({
            'Periodo': resultado['periodo'],
            'Objetivo': resultado['objetivo'],
            'Fecha_Inicio': resultado['fecha_inicio'],
            'Fecha_Fin': resultado['fecha_fin'],
            'Compra_%': resultado['compra_pct'],
            'Venta_%': resultado['venta_pct'],
            'Ganancia_Min_%': resultado['ganancia_min_pct'],
            'Compra_Multiple': resultado['compra_mult'] or 0,
            'Venta_Multiple': resultado['venta_mult'] or 0,
            'Rentabilidad_Max': round(resultado['rentabilidad_max'], 2),
            'Margen_Promedio': round(resultado['margen_promedio'], 2),
            'Promedio_Maximos': round(resultado['promedio_maximos'] * 100, 2),
            'Promedio_Minimos': round(resultado['promedio_minimos'] * 100, 2)
        })

    df_resumen = pd.DataFrame(resumen)
    df_resumen.to_sql('resumen_parametros', conn, if_exists='replace', index=False)

    conn.close()

    log(f"Generando Excel: {excel_path}")

    # Crear Excel
    wb = openpyxl.Workbook()

    # Hoja de resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # Escribir encabezados
    headers = list(df_resumen.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Escribir datos
    for row_idx, row in df_resumen.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws_resumen.cell(row=row_idx + 2, column=col_idx, value=value)

    # Ajustar anchos
    for col in ws_resumen.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_resumen.column_dimensions[col[0].column_letter].width = max_length + 2

    # Agregar hojas por período (solo rentabilidad para simplificar)
    for key, resultado in resultados.items():
        if 'rentabilidad' in key:  # Solo incluir resultados de rentabilidad
            df_sim = resultado['df_simulacion'].copy()

            # Formatear
            df_sim['Fecha'] = df_sim['Fecha'].dt.strftime("%d/%m/%Y")
            df_sim['% var.'] = (df_sim['% var.'] * 100).round(2).astype(str) + "%"
            df_sim['% acumulado'] = (df_sim['% acumulado'] * 100).round(2).astype(str) + "%"
            df_sim['Rentabilidad'] = df_sim['Rentabilidad'].round(2).astype(str) + "%"

            # Nombre de hoja (máx 31 caracteres)
            nombre_hoja = resultado['periodo'][:20]
            ws = wb.create_sheet(title=nombre_hoja)

            # Escribir datos
            for col_idx, col_name in enumerate(df_sim.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)

            for row_idx, row in df_sim.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx + 2, column=col_idx, value=value)

    wb.save(excel_path)

    log(f"Archivos generados correctamente")

    return db_path, excel_path


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='Análisis headless de ticker')
    parser.add_argument('csv_path', help='Ruta al archivo CSV del ticker')
    parser.add_argument('--limite', type=int, default=10, help='Límite de acciones (default: 10)')
    parser.add_argument('--no-excel', action='store_true', help='No generar Excel')

    args = parser.parse_args()

    if not os.path.exists(args.csv_path):
        print(f"ERROR: No se encontró el archivo {args.csv_path}")
        sys.exit(1)

    print("=" * 60)
    print("ANÁLISIS DE TICKER - MODO HEADLESS")
    print("=" * 60)
    print()

    # Extraer ticker symbol del nombre del archivo (ej: Datos_AAPL_FEB25_FEB26.csv -> AAPL)
    nombre_archivo = os.path.basename(args.csv_path)
    partes = nombre_archivo.replace('.csv', '').split('_')
    ticker_symbol = partes[1] if len(partes) >= 2 else partes[0]
    log(f"Ticker detectado: {ticker_symbol}")

    # Ejecutar análisis
    resultados = analizar_ticker(args.csv_path, args.limite)

    print()

    # Guardar en Resultado_de_Analisis.json
    log("Guardando en Resultado_de_Analisis.json...")
    guardar_en_resultado_json(resultados, args.csv_path, ticker_symbol)

    # Generar DB y Excel
    if not args.no_excel:
        db_path, excel_path = generar_db_excel(resultados, args.csv_path)

    print()
    print("=" * 60)
    print("RESUMEN DE MEJORES PARÁMETROS")
    print("=" * 60)

    # Mostrar resumen
    for key, resultado in resultados.items():
        if 'rentabilidad' in key:
            print(f"\n{resultado['periodo'].upper()} ({resultado['fecha_inicio']} - {resultado['fecha_fin']})")
            print(f"  Compra: {resultado['compra_pct']}%")
            print(f"  Venta: {resultado['venta_pct']}%")
            print(f"  Ganancia mínima: {resultado['ganancia_min_pct']}%")
            print(f"  Compra múltiple: {resultado['compra_mult'] or 'N/A'}")
            print(f"  Venta múltiple: {resultado['venta_mult'] or 'N/A'}")
            print(f"  Rentabilidad máx: {resultado['rentabilidad_max']:.2f}%")
            print(f"  Margen promedio: {resultado['margen_promedio']:.2f}")
            print(f"  Prom. máximos: {resultado['promedio_maximos']:.2f}%")
            print(f"  Prom. mínimos: {resultado['promedio_minimos']:.2f}%")

    print()
    print("=" * 60)
    print("ANÁLISIS COMPLETADO")
    print("=" * 60)


if __name__ == "__main__":
    main()
