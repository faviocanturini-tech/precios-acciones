#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
SCRIPT: Análisis de Inversiones con Optimización Multi-Período
=============================================================================
VERSIÓN: 2.6.0
FECHA DE CREACIÓN: 13/12/2025 10:45:00
ÚLTIMA MODIFICACIÓN: 18/12/2025 19:00:00

MEJORAS EN ESTA VERSIÓN (v2.6.0):
- NUEVO: Checkboxes para objetivos de optimización (Rentabilidad y/o Margen)
- NUEVO: Análisis multi-objetivo en una sola ejecución
- NUEVO: Barra de progreso inteligente con estimación de tiempo
- NUEVO: Historial de tiempos en ~/.analisis_tiempos.json
- NUEVO: Columnas Prom.Min% y Prom.Max% en ventana "Parámetros Activos"
- MEJORADO: Ventana "Administrar JSON" con 31 columnas de estadísticas
- MEJORADO: Valores Prom.Max% y Prom.Min% corregidos (÷100)
- MEJORADO: Anchos de columna dinámicos según título
- MEJORADO: Ordenamiento alfabético de tickers en todas las ventanas

MEJORAS EN VERSIÓN ANTERIOR (v2.5.8):
- NUEVO: Campo ticker_symbol en JSON (ej: "META" extraído de "Datos_META_ENE25_NOV25")
- NUEVO: Función extraer_ticker_symbol() para obtener ticker puro del nombre de archivo
- MEJORADO: Ventana "Administrar JSON" muestra ticker_symbol en lugar de nombre archivo

MEJORAS EN VERSIÓN ANTERIOR (v2.5.7):
- NUEVO: Barra de desplazamiento vertical (scrollbar) para toda la interfaz
- NUEVO: Cuadros de fechas de compras/ventas múltiples restaurados
- MEJORADO: Scroll funciona con rueda del mouse

MEJORAS EN VERSIÓN ANTERIOR (v2.5.6):
- NUEVO: Botón para detener análisis en proceso
- MEJORADO: Al borrar registro del JSON, se actualiza inmediatamente el cuadro de parámetros
- MEJORADO: Registros agrupados por objetivo en el cuadro de parámetros

MEJORAS EN VERSIÓN ANTERIOR (v2.5.5):
- MEJORADO: Guarda nuevo registro si los parámetros son diferentes (mismo ticker/período/objetivo)
- MEJORADO: Solo sobrescribe si ticker/período/objetivo Y parámetros son idénticos

MEJORAS EN VERSIÓN ANTERIOR (v2.5.4):
- NUEVO: Ventana para administrar JSON (ver y eliminar registros)
- NUEVO: Selección múltiple para eliminar varios registros a la vez

MEJORAS EN VERSIÓN ANTERIOR (v2.5.3):
- NUEVO: Refinamiento post-optimización para encontrar el centro del rango óptimo
- NUEVO: Semilla fija (seed=42) para resultados reproducibles
- MEJORADO: Siempre obtiene el mismo resultado óptimo para los mismos datos/parámetros

MEJORAS EN VERSIÓN ANTERIOR (v2.5.2):
- CORREGIDO: Lee correctamente JSON con estructura MIXTA (antigua + nueva)
- CORREGIDO: Botón "Generar DB y Excel" se reactiva al hacer nuevo análisis
- CORREGIDO: Muestra TODOS los análisis guardados (7 análisis en tu caso)
- MEJORADO: Compatible con cualquier combinación de estructuras JSON

AUTOR: Claude (Anthropic)
=============================================================================
"""

import os
import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.optimize import differential_evolution
import numpy as np
import time
import json
from pathlib import Path
from datetime import datetime, timedelta

# Valores por defecto para el límite
LIMITE_TIPO = "acciones"
LIMITE_VALOR = 10.0
VENTA_MULTIPLE_ACCIONES = None
COMPRA_MULTIPLE_ACCIONES = None
text_ventas_mult = None
text_compras_mult = None

# Archivo de configuración para ubicación del JSON
CONFIG_FILE = Path.home() / ".analisis_config.json"
UBICACION_JSON = None
ARCHIVO_JSON = None

# Archivo de configuración para parámetros activos
ARCHIVO_PARAMETROS_ACTIVOS = None  # Se configura junto con UBICACION_JSON

# Columnas esperadas (exactas)
EXPECTED_COLUMNS = ["Fecha", "Último", "Apertura", "Máximo", "Mínimo", "Vol.", "% var."]

# Variable global para almacenar resultados de análisis
resultados_analisis_actuales = {}

# Variables globales para progreso
scipy_evaluaciones = 0
scipy_evaluaciones_max = 0
scipy_inicio_tiempo = None

# Variable global para detener análisis
analisis_detenido = False

# Variable global para objetivo actual durante análisis
OBJETIVO_ACTUAL = None

# Archivo para historial de tiempos de análisis
ARCHIVO_HISTORIAL_TIEMPOS = Path.home() / ".analisis_tiempos.json"

# Variables globales para progreso inteligente
progreso_combinacion_actual = 0
progreso_total_combinaciones = 0
progreso_tiempo_inicio_total = None
progreso_tiempos_combinaciones = []  # Lista de tiempos por combinación en la sesión actual


def obtener_clave_configuracion(num_filas, checks_activos):
    """
    Genera una clave única basada en el rango de filas y checks activos.
    checks_activos es un dict con: {'scipy': bool, 'compra': bool, 'venta': bool,
                                     'ganancia': bool, 'compra_mult': bool, 'venta_mult': bool}
    """
    # Rangos de filas: 0-100, 100-200, 200-300, 300-500, 500+
    if num_filas <= 100:
        rango = "0-100"
    elif num_filas <= 200:
        rango = "100-200"
    elif num_filas <= 300:
        rango = "200-300"
    elif num_filas <= 500:
        rango = "300-500"
    else:
        rango = "500+"

    # Crear string de checks activos
    checks_str = "_".join([k for k, v in sorted(checks_activos.items()) if v])

    return f"{rango}_{checks_str}" if checks_str else f"{rango}_ninguno"


def cargar_historial_tiempos():
    """Carga el historial de tiempos desde el archivo JSON"""
    try:
        if ARCHIVO_HISTORIAL_TIEMPOS.exists():
            with open(ARCHIVO_HISTORIAL_TIEMPOS, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"[WARN] Error cargando historial de tiempos: {e}")
    return {}


def guardar_historial_tiempos(historial):
    """Guarda el historial de tiempos en el archivo JSON"""
    try:
        with open(ARCHIVO_HISTORIAL_TIEMPOS, 'w', encoding='utf-8') as f:
            json.dump(historial, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"[WARN] Error guardando historial de tiempos: {e}")


def registrar_tiempo_combinacion(clave_config, tiempo_segundos):
    """Registra el tiempo de una combinación en el historial"""
    historial = cargar_historial_tiempos()

    if clave_config not in historial:
        historial[clave_config] = {"tiempos": [], "promedio": 0}

    # Mantener solo los últimos 10 tiempos para cada configuración
    historial[clave_config]["tiempos"].append(tiempo_segundos)
    if len(historial[clave_config]["tiempos"]) > 10:
        historial[clave_config]["tiempos"] = historial[clave_config]["tiempos"][-10:]

    # Calcular promedio
    tiempos = historial[clave_config]["tiempos"]
    historial[clave_config]["promedio"] = sum(tiempos) / len(tiempos)

    guardar_historial_tiempos(historial)


def estimar_tiempo_total(clave_config, num_combinaciones):
    """Estima el tiempo total basado en el historial"""
    historial = cargar_historial_tiempos()

    if clave_config in historial and historial[clave_config]["promedio"] > 0:
        tiempo_por_combinacion = historial[clave_config]["promedio"]
        return tiempo_por_combinacion * num_combinaciones, True

    return None, False


def formatear_tiempo(segundos):
    """Formatea segundos a formato legible (mm:ss o hh:mm:ss)"""
    if segundos < 0:
        return "calculando..."

    segundos = int(segundos)
    if segundos < 60:
        return f"{segundos} seg"
    elif segundos < 3600:
        mins = segundos // 60
        segs = segundos % 60
        return f"{mins}m {segs:02d}s"
    else:
        horas = segundos // 3600
        mins = (segundos % 3600) // 60
        segs = segundos % 60
        return f"{horas}h {mins:02d}m {segs:02d}s"


def detener_analisis():
    """Detiene el análisis en proceso"""
    global analisis_detenido
    analisis_detenido = True
    print("[DEBUG] Análisis detenido por el usuario")


def extraer_ticker_symbol(nombre_archivo):
    """
    Extrae el símbolo del ticker de Yahoo Finance desde el nombre del archivo.

    Ejemplos:
        "Datos_META_ENE25_NOV25" → "META"
        "Datos_AAPL_ENE25_NOV25" → "AAPL"
        "Datos_BRK-B_ENE25_NOV25" → "BRK-B"
        "Datos_QQQ_ENE25_NOV25" → "QQQ"

    Args:
        nombre_archivo: Nombre del archivo sin extensión (ej: "Datos_META_ENE25_NOV25")

    Returns:
        str: Símbolo del ticker (ej: "META") o None si no se puede extraer
    """
    import re

    if not nombre_archivo:
        return None

    # Patrón: Datos_TICKER_MesAño_MesAño
    # Donde TICKER puede contener letras, números y guiones (ej: BRK-B)
    # Y MesAño es 3 letras + 2 dígitos (ej: ENE25, NOV25)
    patron = r'^Datos_([A-Za-z0-9\-]+)_[A-Za-z]{3}\d{2}_[A-Za-z]{3}\d{2}$'

    match = re.match(patron, nombre_archivo)
    if match:
        return match.group(1).upper()

    # Patrón alternativo más flexible: Datos_TICKER_cualquier_cosa
    patron_alternativo = r'^Datos_([A-Za-z0-9\-]+)_'
    match_alt = re.match(patron_alternativo, nombre_archivo)
    if match_alt:
        return match_alt.group(1).upper()

    # Si no hay patrón "Datos_", intentar extraer el primer segmento antes de "_"
    partes = nombre_archivo.split('_')
    if len(partes) >= 2 and partes[0].upper() == "DATOS":
        return partes[1].upper()

    return None


# =========================
# Funciones de configuración JSON
# =========================
def cargar_configuracion():
    """Carga la ubicación del JSON desde el archivo de configuración"""
    global UBICACION_JSON, ARCHIVO_JSON, ARCHIVO_PARAMETROS_ACTIVOS

    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
            UBICACION_JSON = config.get("ubicacion_json", None)
            if UBICACION_JSON:
                ARCHIVO_JSON = Path(UBICACION_JSON) / "Resultado_de_Analisis.json"
                ARCHIVO_PARAMETROS_ACTIVOS = Path(UBICACION_JSON) / "parametros_activos.json"
                label_json_actual.config(text=f"JSON: {ARCHIVO_JSON}")


def guardar_configuracion():
    """Guarda la ubicación del JSON en el archivo de configuración"""
    config = {"ubicacion_json": UBICACION_JSON}
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)


def seleccionar_ubicacion_json():
    """Permite al usuario seleccionar dónde guardar el JSON"""
    global UBICACION_JSON, ARCHIVO_JSON, ARCHIVO_PARAMETROS_ACTIVOS

    carpeta = filedialog.askdirectory(title="Selecciona carpeta para guardar resultados JSON")
    if carpeta:
        UBICACION_JSON = carpeta
        ARCHIVO_JSON = Path(UBICACION_JSON) / "Resultado_de_Analisis.json"
        ARCHIVO_PARAMETROS_ACTIVOS = Path(UBICACION_JSON) / "parametros_activos.json"
        guardar_configuracion()
        label_json_actual.config(text=f"JSON: {ARCHIVO_JSON}")
        messagebox.showinfo("Ubicación guardada", f"Los resultados se guardarán en:\n{ARCHIVO_JSON}")


def verificar_ubicacion_json():
    """Verifica si hay ubicación configurada, si no, pide al usuario"""
    global UBICACION_JSON, ARCHIVO_JSON

    if UBICACION_JSON is None:
        respuesta = messagebox.askyesno(
            "Ubicación JSON no configurada",
            "No has configurado dónde guardar los resultados JSON.\n¿Deseas seleccionar una carpeta ahora?"
        )
        if respuesta:
            seleccionar_ubicacion_json()
            return ARCHIVO_JSON is not None
        return False
    return True


# =========================
# Funciones para Parámetros Activos
# =========================
def cargar_parametros_activos():
    """Carga los parámetros activos desde el archivo de configuración"""
    if ARCHIVO_PARAMETROS_ACTIVOS is None or not ARCHIVO_PARAMETROS_ACTIVOS.exists():
        return []

    try:
        with open(ARCHIVO_PARAMETROS_ACTIVOS, 'r', encoding='utf-8') as f:
            datos = json.load(f)
            return datos.get("parametros_activos", [])
    except Exception as e:
        print(f"[ERROR] Error cargando parámetros activos: {e}")
        return []


def guardar_parametros_activos(parametros):
    """Guarda los parámetros activos en el archivo de configuración"""
    if ARCHIVO_PARAMETROS_ACTIVOS is None:
        messagebox.showerror("Error", "No hay ubicación configurada para guardar los parámetros activos.")
        return False

    try:
        datos = {"parametros_activos": parametros}
        with open(ARCHIVO_PARAMETROS_ACTIVOS, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Error guardando parámetros activos:\n{e}")
        return False


def administrar_parametros_activos():
    """Abre una ventana para gestionar los parámetros activos"""
    if not verificar_ubicacion_json():
        return

    # Cargar parámetros activos actuales
    parametros = cargar_parametros_activos()

    # Crear ventana
    ventana_params = tk.Toplevel(ventana)
    ventana_params.title("Parámetros Activos para Señales de Trading")
    ventana_params.geometry("1100x500")
    ventana_params.transient(ventana)
    ventana_params.grab_set()

    # Frame superior con instrucciones
    frame_instrucciones = tk.Frame(ventana_params, pady=5)
    frame_instrucciones.pack(fill="x", padx=10)
    tk.Label(frame_instrucciones,
             text="Configura los parámetros que se usarán para generar señales diarias de compra/venta",
             font=("Arial", 10), fg="gray").pack(anchor="w")

    # Frame para el Treeview
    frame_tree = tk.Frame(ventana_params)
    frame_tree.pack(fill="both", expand=True, padx=10, pady=5)

    # Scrollbars
    scrollbar_y = tk.Scrollbar(frame_tree, orient="vertical")
    scrollbar_x = tk.Scrollbar(frame_tree, orient="horizontal")

    # Treeview
    columns = ("Symbol", "Origen", "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N", "Límite", "Valor Lím.", "Prom.Min%", "Prom.Max%")
    tree_params = ttk.Treeview(frame_tree, columns=columns, show="headings",
                               selectmode="extended",
                               yscrollcommand=scrollbar_y.set,
                               xscrollcommand=scrollbar_x.set)

    scrollbar_y.config(command=tree_params.yview)
    scrollbar_x.config(command=tree_params.xview)

    # Configurar columnas
    anchos = {"Symbol": 80, "Origen": 100, "Compra%": 80, "Venta%": 80,
              "Gan.Mín%": 80, "Compra N": 80, "Venta N": 80, "Límite": 70, "Valor Lím.": 80,
              "Prom.Min%": 85, "Prom.Max%": 85}

    for col in columns:
        tree_params.heading(col, text=col)
        tree_params.column(col, width=anchos.get(col, 80), anchor="center")

    def actualizar_tabla():
        """Actualiza la tabla con los parámetros activos"""
        for item in tree_params.get_children():
            tree_params.delete(item)

        # Ordenar parámetros alfabéticamente por ticker_symbol
        parametros_ordenados = sorted(parametros, key=lambda x: x.get("ticker_symbol", "").upper())

        for param in parametros_ordenados:
            compra_n = param.get("compra_multiple")
            venta_n = param.get("venta_multiple")
            limite_tipo = param.get("limite_tipo", "acciones")
            limite_valor = param.get("limite_valor", 10.0)
            prom_min = param.get("promedio_minimos", 0)
            prom_max = param.get("promedio_maximos", 0)
            tree_params.insert("", "end", values=(
                param.get("ticker_symbol", ""),
                param.get("origen", ""),
                f"{param.get('compra_pct', 0):.1f}",
                f"{param.get('venta_pct', 0):.1f}",
                f"{param.get('ganancia_min_pct', 0):.1f}",
                compra_n if compra_n else "-",
                venta_n if venta_n else "-",
                limite_tipo.title() if limite_tipo else "Acciones",
                f"{limite_valor:.0f}" if limite_tipo == "acciones" else f"${limite_valor:.0f}",
                f"{prom_min:.2f}%" if prom_min else "-",
                f"{prom_max:.2f}%" if prom_max else "-"
            ))

    actualizar_tabla()

    # Empaquetar Treeview y scrollbars
    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")
    tree_params.pack(fill="both", expand=True)

    # Frame inferior con botones
    frame_botones = tk.Frame(ventana_params, pady=10)
    frame_botones.pack(fill="x", padx=10)

    def agregar_desde_json():
        """Abre ventana para seleccionar parámetros del JSON calculado"""
        datos_json = cargar_resultados_json()
        if not datos_json:
            messagebox.showinfo("Sin datos", "No hay parámetros calculados en el JSON")
            return

        # Crear ventana de selección
        ventana_seleccion = tk.Toplevel(ventana_params)
        ventana_seleccion.title("Seleccionar desde JSON")
        ventana_seleccion.geometry("900x400")
        ventana_seleccion.transient(ventana_params)
        ventana_seleccion.grab_set()

        tk.Label(ventana_seleccion, text="Selecciona los parámetros a agregar:",
                 font=("Arial", 10)).pack(pady=5)

        # Frame para lista
        frame_lista = tk.Frame(ventana_seleccion)
        frame_lista.pack(fill="both", expand=True, padx=10, pady=5)

        scrollbar = tk.Scrollbar(frame_lista)
        scrollbar.pack(side="right", fill="y")

        # Treeview para selección
        cols_sel = ("Symbol", "Período", "Objetivo", "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N")
        tree_sel = ttk.Treeview(frame_lista, columns=cols_sel, show="headings",
                                selectmode="extended", yscrollcommand=scrollbar.set)
        scrollbar.config(command=tree_sel.yview)

        for col in cols_sel:
            tree_sel.heading(col, text=col)
            tree_sel.column(col, width=90, anchor="center")

        # Diccionario para mapear items a datos completos
        item_datos = {}

        # Llenar con datos del JSON
        for ticker, contenido_ticker in datos_json.items():
            ticker_symbol = contenido_ticker.get("_ticker_symbol") or extraer_ticker_symbol(ticker) or ticker

            # Estructura nueva
            for periodo, contenido_periodo in contenido_ticker.items():
                if periodo in ["ticker", "fecha_guardado", "periodos", "_ticker_symbol"]:
                    continue

                if isinstance(contenido_periodo, dict):
                    for objetivo, datos in contenido_periodo.items():
                        if isinstance(datos, dict) and "parametros_optimos" in datos:
                            params = datos.get("parametros_optimos", {})
                            compra_mult = params.get("compra_multiple")
                            venta_mult = params.get("venta_multiple")

                            item_id = tree_sel.insert("", "end", values=(
                                ticker_symbol,
                                periodo.replace("_", " ").title(),
                                objetivo.replace("_", " ").title(),
                                f"{params.get('compra_pct', 0):.1f}",
                                f"{params.get('venta_pct', 0):.1f}",
                                f"{params.get('ganancia_minima_pct', 0):.1f}",
                                compra_mult if compra_mult else "-",
                                venta_mult if venta_mult else "-"
                            ))

                            item_datos[item_id] = {
                                "ticker_symbol": ticker_symbol,
                                "origen": f"calculado ({periodo}/{objetivo})",
                                "compra_pct": params.get("compra_pct", 0),
                                "venta_pct": params.get("venta_pct", 0),
                                "ganancia_min_pct": params.get("ganancia_minima_pct", 0),
                                "compra_multiple": compra_mult,
                                "venta_multiple": venta_mult,
                                "limite_tipo": params.get("limite_tipo", "acciones"),
                                "limite_valor": params.get("limite_valor", 10.0),
                                # Condiciones para compra/venta múltiple
                                "promedio_maximos": params.get("promedio_maximos", 0),
                                "promedio_minimos": params.get("promedio_minimos", 0)
                            }

        tree_sel.pack(fill="both", expand=True)

        def agregar_seleccionados():
            seleccionados = tree_sel.selection()
            if not seleccionados:
                messagebox.showwarning("Sin selección", "Selecciona al menos un registro")
                return

            agregados = 0
            for item_id in seleccionados:
                if item_id in item_datos:
                    nuevo_param = item_datos[item_id].copy()
                    # Verificar si ya existe este ticker
                    existe = any(p.get("ticker_symbol") == nuevo_param["ticker_symbol"] for p in parametros)
                    if existe:
                        resp = messagebox.askyesno("Ticker existente",
                            f"Ya existe un parámetro para {nuevo_param['ticker_symbol']}.\n¿Deseas reemplazarlo?")
                        if resp:
                            parametros[:] = [p for p in parametros if p.get("ticker_symbol") != nuevo_param["ticker_symbol"]]
                        else:
                            continue
                    parametros.append(nuevo_param)
                    agregados += 1

            if agregados > 0:
                guardar_parametros_activos(parametros)
                actualizar_tabla()
                messagebox.showinfo("Agregados", f"Se agregaron {agregados} parámetro(s)")
            ventana_seleccion.destroy()

        tk.Button(ventana_seleccion, text="Agregar seleccionados", command=agregar_seleccionados,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).pack(pady=10)

    def agregar_personalizado():
        """Abre ventana para agregar parámetros personalizados"""
        ventana_custom = tk.Toplevel(ventana_params)
        ventana_custom.title("Agregar Parámetros Personalizados")
        ventana_custom.geometry("450x550")
        ventana_custom.transient(ventana_params)
        ventana_custom.grab_set()

        # Frame para formulario
        frame_form = tk.Frame(ventana_custom, padx=20, pady=20)
        frame_form.pack(fill="both", expand=True)

        # Campos
        tk.Label(frame_form, text="Symbol (ej: META, AAPL):", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=5)
        entry_symbol = tk.Entry(frame_form, width=20)
        entry_symbol.grid(row=0, column=1, pady=5)

        tk.Label(frame_form, text="Compra %:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=5)
        entry_compra = tk.Entry(frame_form, width=20)
        entry_compra.grid(row=1, column=1, pady=5)

        tk.Label(frame_form, text="Venta %:", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=5)
        entry_venta = tk.Entry(frame_form, width=20)
        entry_venta.grid(row=2, column=1, pady=5)

        tk.Label(frame_form, text="Ganancia Mín %:", font=("Arial", 10)).grid(row=3, column=0, sticky="w", pady=5)
        entry_ganancia = tk.Entry(frame_form, width=20)
        entry_ganancia.grid(row=3, column=1, pady=5)

        tk.Label(frame_form, text="Compra N acciones (opcional):", font=("Arial", 10)).grid(row=4, column=0, sticky="w", pady=5)
        entry_compra_n = tk.Entry(frame_form, width=20)
        entry_compra_n.grid(row=4, column=1, pady=5)

        tk.Label(frame_form, text="Venta N acciones (opcional):", font=("Arial", 10)).grid(row=5, column=0, sticky="w", pady=5)
        entry_venta_n = tk.Entry(frame_form, width=20)
        entry_venta_n.grid(row=5, column=1, pady=5)

        # Campos de límite
        tk.Label(frame_form, text="Tipo de límite:", font=("Arial", 10)).grid(row=6, column=0, sticky="w", pady=5)
        limite_tipo_var = tk.StringVar(value="acciones")
        frame_limite_tipo = tk.Frame(frame_form)
        frame_limite_tipo.grid(row=6, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_limite_tipo, text="Acciones", variable=limite_tipo_var, value="acciones").pack(side="left")
        tk.Radiobutton(frame_limite_tipo, text="Monto $", variable=limite_tipo_var, value="monto").pack(side="left")

        tk.Label(frame_form, text="Valor límite:", font=("Arial", 10)).grid(row=7, column=0, sticky="w", pady=5)
        entry_limite_valor = tk.Entry(frame_form, width=20)
        entry_limite_valor.insert(0, "10")
        entry_limite_valor.grid(row=7, column=1, pady=5)

        # Campos para condiciones de compra/venta múltiple
        tk.Label(frame_form, text="─── Condiciones Múltiples ───", font=("Arial", 9, "italic"), fg="gray").grid(row=8, column=0, columnspan=2, pady=(10,5))

        tk.Label(frame_form, text="Prom. % acum mínimos (-):", font=("Arial", 10)).grid(row=9, column=0, sticky="w", pady=5)
        entry_prom_min = tk.Entry(frame_form, width=20)
        entry_prom_min.insert(0, "0")
        entry_prom_min.grid(row=9, column=1, pady=5)

        tk.Label(frame_form, text="Prom. % acum máximos (+):", font=("Arial", 10)).grid(row=10, column=0, sticky="w", pady=5)
        entry_prom_max = tk.Entry(frame_form, width=20)
        entry_prom_max.insert(0, "0")
        entry_prom_max.grid(row=10, column=1, pady=5)

        def guardar_personalizado():
            symbol = entry_symbol.get().strip().upper()
            if not symbol:
                messagebox.showwarning("Campo requerido", "Ingresa el símbolo del ticker")
                return

            try:
                compra_pct = float(entry_compra.get().strip().replace(",", "."))
                venta_pct = float(entry_venta.get().strip().replace(",", "."))
                ganancia_pct = float(entry_ganancia.get().strip().replace(",", "."))
            except ValueError:
                messagebox.showerror("Error", "Los valores de porcentaje deben ser numéricos")
                return

            compra_n = entry_compra_n.get().strip()
            venta_n = entry_venta_n.get().strip()

            # Obtener valores de límite
            tipo_limite = limite_tipo_var.get()
            try:
                valor_limite = float(entry_limite_valor.get().strip().replace(",", "."))
            except ValueError:
                valor_limite = 10.0

            # Obtener valores de promedios para múltiples
            try:
                prom_min = float(entry_prom_min.get().strip().replace(",", "."))
            except ValueError:
                prom_min = 0.0
            try:
                prom_max = float(entry_prom_max.get().strip().replace(",", "."))
            except ValueError:
                prom_max = 0.0

            nuevo_param = {
                "ticker_symbol": symbol,
                "origen": "personalizado",
                "compra_pct": compra_pct,
                "venta_pct": venta_pct,
                "ganancia_min_pct": ganancia_pct,
                "compra_multiple": int(compra_n) if compra_n else None,
                "venta_multiple": int(venta_n) if venta_n else None,
                "limite_tipo": tipo_limite,
                "limite_valor": valor_limite,
                "promedio_minimos": prom_min,
                "promedio_maximos": prom_max
            }

            # Verificar si ya existe
            existe = any(p.get("ticker_symbol") == symbol for p in parametros)
            if existe:
                resp = messagebox.askyesno("Ticker existente",
                    f"Ya existe un parámetro para {symbol}.\n¿Deseas reemplazarlo?")
                if resp:
                    parametros[:] = [p for p in parametros if p.get("ticker_symbol") != symbol]
                else:
                    return

            parametros.append(nuevo_param)
            guardar_parametros_activos(parametros)
            actualizar_tabla()
            messagebox.showinfo("Guardado", f"Parámetros para {symbol} guardados correctamente")
            ventana_custom.destroy()

        tk.Button(frame_form, text="Guardar", command=guardar_personalizado,
                  bg="#28a745", fg="white", font=("Arial", 10, "bold")).grid(row=11, column=0, columnspan=2, pady=20)

    def eliminar_seleccionados():
        """Elimina los parámetros seleccionados"""
        seleccionados = tree_params.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona al menos un registro para eliminar")
            return

        if not messagebox.askyesno("Confirmar", f"¿Eliminar {len(seleccionados)} parámetro(s)?"):
            return

        # Obtener los symbols a eliminar
        symbols_eliminar = []
        for item_id in seleccionados:
            valores = tree_params.item(item_id, "values")
            symbols_eliminar.append(valores[0])

        # Filtrar parámetros
        parametros[:] = [p for p in parametros if p.get("ticker_symbol") not in symbols_eliminar]
        guardar_parametros_activos(parametros)
        actualizar_tabla()
        messagebox.showinfo("Eliminados", f"Se eliminaron {len(symbols_eliminar)} parámetro(s)")

    def editar_parametro():
        """Edita el parámetro seleccionado"""
        seleccionados = tree_params.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "Selecciona un parámetro para editar")
            return

        if len(seleccionados) > 1:
            messagebox.showwarning("Selección múltiple", "Selecciona solo un parámetro para editar")
            return

        # Obtener el ticker del item seleccionado
        item_id = seleccionados[0]
        valores = tree_params.item(item_id, "values")
        ticker_editar = valores[0]

        # Buscar el parámetro en la lista
        param_editar = None
        param_index = None
        for i, p in enumerate(parametros):
            if p.get("ticker_symbol") == ticker_editar:
                param_editar = p
                param_index = i
                break

        if param_editar is None:
            messagebox.showerror("Error", "No se encontró el parámetro")
            return

        # Crear ventana de edición
        ventana_editar = tk.Toplevel(ventana_params)
        ventana_editar.title(f"Editar parámetros - {ticker_editar}")
        ventana_editar.geometry("400x580")
        ventana_editar.transient(ventana_params)
        ventana_editar.grab_set()

        tk.Label(ventana_editar, text="Editar Parámetro",
                 font=("Arial", 12, "bold")).pack(pady=10)

        frame_form = tk.Frame(ventana_editar, padx=20, pady=10)
        frame_form.pack(fill="both", expand=True)

        # Campo Ticker (editable)
        tk.Label(frame_form, text="Ticker:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        entry_ticker = tk.Entry(frame_form, width=15, font=("Arial", 10, "bold"))
        entry_ticker.insert(0, ticker_editar)
        entry_ticker.grid(row=0, column=1, sticky="w", pady=5)

        # Campos editables
        tk.Label(frame_form, text="Compra %:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=5)
        entry_compra = tk.Entry(frame_form, width=15)
        entry_compra.insert(0, str(param_editar.get("compra_pct", 0)))
        entry_compra.grid(row=1, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Venta %:", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=5)
        entry_venta = tk.Entry(frame_form, width=15)
        entry_venta.insert(0, str(param_editar.get("venta_pct", 0)))
        entry_venta.grid(row=2, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Ganancia mín %:", font=("Arial", 10)).grid(row=3, column=0, sticky="w", pady=5)
        entry_ganancia = tk.Entry(frame_form, width=15)
        entry_ganancia.insert(0, str(param_editar.get("ganancia_min_pct", 0)))
        entry_ganancia.grid(row=3, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Compra múltiple:", font=("Arial", 10)).grid(row=4, column=0, sticky="w", pady=5)
        entry_compra_mult = tk.Entry(frame_form, width=15)
        compra_mult_val = param_editar.get("compra_multiple")
        entry_compra_mult.insert(0, str(compra_mult_val) if compra_mult_val else "")
        entry_compra_mult.grid(row=4, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Venta múltiple:", font=("Arial", 10)).grid(row=5, column=0, sticky="w", pady=5)
        entry_venta_mult = tk.Entry(frame_form, width=15)
        venta_mult_val = param_editar.get("venta_multiple")
        entry_venta_mult.insert(0, str(venta_mult_val) if venta_mult_val else "")
        entry_venta_mult.grid(row=5, column=1, sticky="w", pady=5)

        # Campos de límite
        tk.Label(frame_form, text="Tipo de límite:", font=("Arial", 10)).grid(row=6, column=0, sticky="w", pady=5)
        limite_tipo_var = tk.StringVar(value=param_editar.get("limite_tipo", "acciones"))
        frame_limite_tipo = tk.Frame(frame_form)
        frame_limite_tipo.grid(row=6, column=1, sticky="w", pady=5)
        tk.Radiobutton(frame_limite_tipo, text="Acciones", variable=limite_tipo_var, value="acciones").pack(side="left")
        tk.Radiobutton(frame_limite_tipo, text="Monto $", variable=limite_tipo_var, value="monto").pack(side="left")

        tk.Label(frame_form, text="Valor límite:", font=("Arial", 10)).grid(row=7, column=0, sticky="w", pady=5)
        entry_limite_valor = tk.Entry(frame_form, width=15)
        entry_limite_valor.insert(0, str(param_editar.get("limite_valor", 10.0)))
        entry_limite_valor.grid(row=7, column=1, sticky="w", pady=5)

        # Campos para condiciones de compra/venta múltiple
        tk.Label(frame_form, text="─── Condiciones Múltiples ───", font=("Arial", 9, "italic"), fg="gray").grid(row=8, column=0, columnspan=2, pady=(10,5))

        tk.Label(frame_form, text="Prom. % mínimos (-):", font=("Arial", 10)).grid(row=9, column=0, sticky="w", pady=5)
        entry_prom_min = tk.Entry(frame_form, width=15)
        entry_prom_min.insert(0, str(param_editar.get("promedio_minimos", 0)))
        entry_prom_min.grid(row=9, column=1, sticky="w", pady=5)

        tk.Label(frame_form, text="Prom. % máximos (+):", font=("Arial", 10)).grid(row=10, column=0, sticky="w", pady=5)
        entry_prom_max = tk.Entry(frame_form, width=15)
        entry_prom_max.insert(0, str(param_editar.get("promedio_maximos", 0)))
        entry_prom_max.grid(row=10, column=1, sticky="w", pady=5)

        def guardar_cambios():
            try:
                nuevo_ticker = entry_ticker.get().strip().upper()
                if not nuevo_ticker:
                    messagebox.showerror("Error", "El ticker no puede estar vacío")
                    return

                # Verificar si el nuevo ticker ya existe (y no es el mismo que estamos editando)
                if nuevo_ticker != ticker_editar:
                    existe = any(p.get("ticker_symbol") == nuevo_ticker for p in parametros)
                    if existe:
                        messagebox.showerror("Error", f"Ya existe un parámetro para {nuevo_ticker}")
                        return

                nuevo_param = {
                    "ticker_symbol": nuevo_ticker,
                    "origen": param_editar.get("origen", "editado"),
                    "compra_pct": float(entry_compra.get()),
                    "venta_pct": float(entry_venta.get()),
                    "ganancia_min_pct": float(entry_ganancia.get()),
                    "compra_multiple": int(entry_compra_mult.get()) if entry_compra_mult.get().strip() else None,
                    "venta_multiple": int(entry_venta_mult.get()) if entry_venta_mult.get().strip() else None,
                    "limite_tipo": limite_tipo_var.get(),
                    "limite_valor": float(entry_limite_valor.get()),
                    "promedio_minimos": float(entry_prom_min.get()) if entry_prom_min.get().strip() else 0,
                    "promedio_maximos": float(entry_prom_max.get()) if entry_prom_max.get().strip() else 0
                }

                parametros[param_index] = nuevo_param
                guardar_parametros_activos(parametros)
                actualizar_tabla()
                ventana_editar.destroy()
                messagebox.showinfo("Guardado", f"Parámetros de {nuevo_ticker} actualizados")

            except ValueError as e:
                messagebox.showerror("Error", f"Valores inválidos: {e}")

        tk.Button(frame_form, text="Guardar cambios", command=guardar_cambios,
                  bg="#ffc107", fg="black", font=("Arial", 10, "bold")).grid(row=11, column=0, columnspan=2, pady=20)

    def exportar_activos_excel():
        """Exporta los parámetros activos a Excel"""
        if not parametros:
            messagebox.showwarning("Sin datos", "No hay parámetros activos para exportar")
            return

        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Parametros_Activos.xlsx"
        )

        if not ruta_excel:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Parámetros Activos"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Encabezados
            headers = ["Symbol", "Origen", "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Datos
            for row_idx, param in enumerate(parametros, 2):
                ws.cell(row=row_idx, column=1, value=param.get("ticker_symbol", "")).border = border
                ws.cell(row=row_idx, column=2, value=param.get("origen", "")).border = border
                ws.cell(row=row_idx, column=3, value=param.get("compra_pct", 0)).border = border
                ws.cell(row=row_idx, column=4, value=param.get("venta_pct", 0)).border = border
                ws.cell(row=row_idx, column=5, value=param.get("ganancia_min_pct", 0)).border = border
                ws.cell(row=row_idx, column=6, value=param.get("compra_multiple") or "").border = border
                ws.cell(row=row_idx, column=7, value=param.get("venta_multiple") or "").border = border

            # Ajustar anchos
            for col in ["A", "B", "C", "D", "E", "F", "G"]:
                ws.column_dimensions[col].width = 12

            wb.save(ruta_excel)
            messagebox.showinfo("Exportado", f"Parámetros exportados a:\n{ruta_excel}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {e}")

    # Botones
    tk.Button(frame_botones, text="Agregar desde JSON", command=agregar_desde_json,
              bg="#007bff", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Agregar personalizado", command=agregar_personalizado,
              bg="#17a2b8", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Editar", command=editar_parametro,
              bg="#ffc107", fg="black", font=("Arial", 9, "bold")).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Exportar a Excel", command=exportar_activos_excel,
              bg="#28a745", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=5)

    tk.Button(frame_botones, text="Eliminar seleccionados", command=eliminar_seleccionados,
              bg="#ff6b6b", fg="white", font=("Arial", 9, "bold")).pack(side="right", padx=5)
    tk.Button(frame_botones, text="Cerrar", command=ventana_params.destroy).pack(side="right", padx=5)


# =========================
# Funciones para JSON de resultados (MODIFICADO - Estructura jerárquica)
# =========================
def cargar_resultados_json():
    """Carga todos los resultados guardados en el JSON"""
    if ARCHIVO_JSON is None or not ARCHIVO_JSON.exists():
        return {}

    with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def parametros_son_iguales(params_nuevos, params_existentes, tolerancia=0.01):
    """Compara si dos conjuntos de parámetros son iguales (con tolerancia para decimales)"""
    claves_comparar = ["compra_pct", "venta_pct", "ganancia_minima_pct", "suave_pct",
                       "limite_tipo", "limite_valor", "compra_multiple", "venta_multiple"]

    for clave in claves_comparar:
        val_nuevo = params_nuevos.get(clave)
        val_existente = params_existentes.get(clave)

        # Si ambos son None o iguales, continuar
        if val_nuevo == val_existente:
            continue

        # Si uno es None y otro no, son diferentes
        if val_nuevo is None or val_existente is None:
            return False

        # Para valores numéricos, comparar con tolerancia
        if isinstance(val_nuevo, (int, float)) and isinstance(val_existente, (int, float)):
            if abs(val_nuevo - val_existente) > tolerancia:
                return False
        else:
            # Para strings u otros tipos, comparación exacta
            if val_nuevo != val_existente:
                return False

    return True


def guardar_resultados_en_json():
    """Guarda los resultados actuales en el JSON (botón verde) - ESTRUCTURA JERÁRQUICA"""
    global resultados_analisis_actuales, ARCHIVO_JSON

    if not resultados_analisis_actuales:
        messagebox.showwarning("Sin resultados", "No hay resultados de análisis para guardar.")
        return

    if not verificar_ubicacion_json():
        return

    # Verificar que ARCHIVO_JSON esté configurado
    if ARCHIVO_JSON is None:
        messagebox.showerror("Error", "La ruta del archivo JSON no está configurada.")
        return

    try:
        # Cargar JSON existente
        if ARCHIVO_JSON.exists():
            with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
                datos_json = json.load(f)
        else:
            datos_json = {}

        # Obtener ticker del archivo actual
        ticker = resultados_analisis_actuales.get("ticker", "UNKNOWN")

        # Extraer ticker_symbol (ej: "META" de "Datos_META_ENE25_NOV25")
        ticker_symbol = extraer_ticker_symbol(ticker)
        print(f"[DEBUG] Ticker: {ticker} → ticker_symbol: {ticker_symbol}")

        # Estructura jerárquica ticker → período → objetivo
        if ticker not in datos_json:
            datos_json[ticker] = {}

        # Guardar ticker_symbol a nivel del ticker principal
        if ticker_symbol:
            datos_json[ticker]["_ticker_symbol"] = ticker_symbol

        # Verificar que hay periodos para guardar
        periodos = resultados_analisis_actuales.get("periodos", {})
        if not periodos:
            messagebox.showwarning("Sin períodos", "No hay datos de períodos para guardar.")
            return

        registros_nuevos = 0
        registros_actualizados = 0

        for clave_periodo, datos in periodos.items():
            # La clave tiene formato "periodo_objetivo" (ej: "completo_rentabilidad")
            # Extraer período y objetivo
            objetivo_base = datos.get("objetivo", "rentabilidad")

            # Extraer solo el nombre del período (sin el objetivo)
            if "_rentabilidad" in clave_periodo:
                nombre_periodo = clave_periodo.replace("_rentabilidad", "")
            elif "_margen_prom" in clave_periodo:
                nombre_periodo = clave_periodo.replace("_margen_prom", "")
            else:
                nombre_periodo = clave_periodo

            if nombre_periodo not in datos_json[ticker]:
                datos_json[ticker][nombre_periodo] = {}

            # Crear el nuevo registro
            nuevo_registro = {
                "ticker_symbol": ticker_symbol,  # Símbolo para Yahoo Finance (ej: "META")
                "fecha_guardado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "fecha_inicial": datos.get("fecha_inicial", ""),
                "fecha_final": datos.get("fecha_final", ""),
                "parametros_optimos": {
                    "compra_pct": datos.get("compra_pct", 0),
                    "venta_pct": datos.get("venta_pct", 0),
                    "ganancia_minima_pct": datos.get("ganancia_min", 0),
                    "suave_pct": datos.get("suave_pct", 0),
                    "limite_tipo": datos.get("limite_tipo", "acciones"),
                    "limite_valor": datos.get("limite_valor", 10),
                    "compra_multiple": datos.get("compra_mult"),
                    "venta_multiple": datos.get("venta_mult"),
                    # Condiciones para compra/venta múltiple
                    "promedio_maximos": datos.get("promedio_maximos", 0),
                    "promedio_minimos": datos.get("promedio_minimos", 0)
                },
                "metricas": {
                    "rentabilidad_max": datos.get("rentabilidad_max", 0),
                    "margen_promedio": datos.get("margen_promedio", 0),
                    "rentab_promedio": datos.get("rentab_promedio", 0),
                    "max_margen": datos.get("max_margen", 0),
                    "max_aporte": datos.get("max_aporte", 0)
                },
                "estadisticas_var": {
                    "max_var": datos.get("max_var", 0),
                    "min_var": datos.get("min_var", 0),
                    "fecha_max_var": datos.get("fecha_max_var", ""),
                    "fecha_min_var": datos.get("fecha_min_var", ""),
                    "dif_var": datos.get("dif_var", 0),
                    "max_prom_var": datos.get("max_prom_var", 0),
                    "min_prom_var": datos.get("min_prom_var", 0),
                    "dif_prom_var": datos.get("dif_prom_var", 0)
                },
                "estadisticas_operaciones": {
                    "opc_compra": datos.get("opc_compra", 0),
                    "acciones_compradas": datos.get("acciones_compradas", 0),
                    "opc_venta": datos.get("opc_venta", 0),
                    "acciones_vendidas": datos.get("acciones_vendidas", 0),
                    "max_acc_cartera": datos.get("max_acc_cartera", 0),
                    "fecha_max_rentab": datos.get("fecha_max_rentab", "")
                }
            }

            # Buscar si ya existe un registro con los mismos parámetros
            objetivo_encontrado = None
            for objetivo_key, registro_existente in datos_json[ticker][nombre_periodo].items():
                if objetivo_key.startswith(objetivo_base):
                    if isinstance(registro_existente, dict) and "parametros_optimos" in registro_existente:
                        if parametros_son_iguales(nuevo_registro["parametros_optimos"],
                                                   registro_existente["parametros_optimos"]):
                            objetivo_encontrado = objetivo_key
                            break

            if objetivo_encontrado:
                # Actualizar registro existente (mismos parámetros)
                datos_json[ticker][nombre_periodo][objetivo_encontrado] = nuevo_registro
                registros_actualizados += 1
                print(f"[DEBUG] Actualizado: {ticker}/{nombre_periodo}/{objetivo_encontrado}")
            else:
                # Crear nuevo registro (parámetros diferentes)
                # Buscar un nombre único para el objetivo
                objetivo_final = objetivo_base
                contador = 2
                while objetivo_final in datos_json[ticker][nombre_periodo]:
                    objetivo_final = f"{objetivo_base}_{contador}"
                    contador += 1

                datos_json[ticker][nombre_periodo][objetivo_final] = nuevo_registro
                registros_nuevos += 1
                print(f"[DEBUG] Nuevo registro: {ticker}/{nombre_periodo}/{objetivo_final}")

        # Escribir JSON
        with open(ARCHIVO_JSON, 'w', encoding='utf-8') as f:
            json.dump(datos_json, f, indent=2, ensure_ascii=False)

        print(f"[DEBUG] JSON guardado exitosamente en: {ARCHIVO_JSON}")

        mensaje = f"Resultados guardados para {ticker}\n\n"
        if registros_nuevos > 0:
            mensaje += f"• {registros_nuevos} registro(s) nuevo(s)\n"
        if registros_actualizados > 0:
            mensaje += f"• {registros_actualizados} registro(s) actualizado(s)\n"
        mensaje += f"\nArchivo: {ARCHIVO_JSON}"

        messagebox.showinfo("Guardado exitoso", mensaje)
        btn_guardar_json.config(state="disabled")

    except Exception as e:
        messagebox.showerror("Error al guardar", f"Error al guardar en JSON:\n{str(e)}")


def mostrar_info_json_ticker(ticker, csv_path=None):
    """Muestra información del ticker - línea horizontal + tabla consolidada del JSON"""
    global historial_analisis_por_ticker

    # Limpiar frame de info
    for widget in frame_info_json.winfo_children():
        widget.destroy()

    # Frame horizontal para info básica en UNA LÍNEA
    frame_info_horizontal = tk.Frame(frame_info_json)
    frame_info_horizontal.pack(fill="x", pady=2)

    if csv_path and os.path.exists(csv_path):
        fecha_creacion = datetime.fromtimestamp(os.path.getctime(csv_path))
        fecha_creacion_str = fecha_creacion.strftime("%d/%m/%Y %H:%M:%S")

        # Extraer ticker real (siglas de la acción) del nombre del archivo
        nombre_archivo = os.path.splitext(os.path.basename(csv_path))[0]

        partes = nombre_archivo.split('_')
        ticker_real = ticker  # Por defecto usar el nombre completo
        if len(partes) >= 2:
            for parte in partes:
                if parte.isupper() and 1 <= len(parte) <= 5:
                    ticker_real = parte
                    break

        # LÍNEA HORIZONTAL con colores alternados (azul, negro, azul, negro, azul)
        tk.Label(frame_info_horizontal, text=f"Ticker: {ticker_real}",
                 font=("Arial", 9, "bold"), fg="darkblue").pack(side="left")
        tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")

        tk.Label(frame_info_horizontal, text=f"Archivo: {os.path.basename(csv_path)}",
                 font=("Arial", 8), fg="black").pack(side="left")
        tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")

        tk.Label(frame_info_horizontal, text=f"CSV: {fecha_creacion_str}",
                 font=("Arial", 8), fg="blue").pack(side="left")

        # Verificar si existen archivos Excel y DB
        folder = os.path.dirname(csv_path)
        base_name = nombre_archivo

        excel_path = os.path.join(folder, f"{base_name}_analizado.xlsx")
        db_path = os.path.join(folder, f"{base_name}_analizado.db")

        if os.path.exists(excel_path):
            fecha_excel = datetime.fromtimestamp(os.path.getmtime(excel_path))
            fecha_excel_str = fecha_excel.strftime("%d/%m/%Y %H:%M:%S")
            tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")
            tk.Label(frame_info_horizontal, text=f"Excel: {fecha_excel_str}",
                     font=("Arial", 8), fg="black").pack(side="left")

        if os.path.exists(db_path):
            fecha_db = datetime.fromtimestamp(os.path.getmtime(db_path))
            fecha_db_str = fecha_db.strftime("%d/%m/%Y %H:%M:%S")
            tk.Label(frame_info_horizontal, text=" | ", font=("Arial", 9)).pack(side="left")
            tk.Label(frame_info_horizontal, text=f"DB: {fecha_db_str}",
                     font=("Arial", 8), fg="blue").pack(side="left")

    # NUEVO: Cargar y mostrar tabla consolidada con datos del JSON
    if ARCHIVO_JSON is None or not ARCHIVO_JSON.exists():
        return

    datos_json = cargar_resultados_json()

    if ticker not in datos_json:
        return

    info = datos_json[ticker]

    # Limpiar historial anterior de este ticker
    if ticker not in historial_analisis_por_ticker:
        historial_analisis_por_ticker[ticker] = []
    else:
        historial_analisis_por_ticker[ticker] = []

    # COMPATIBILIDAD MEJORADA: Maneja estructura antigua, nueva Y MIXTA

    # 1. Primero procesar estructura ANTIGUA si existe (dentro de "periodos")
    if "periodos" in info and isinstance(info["periodos"], dict):
        for nombre_periodo, datos_periodo in info["periodos"].items():
            if isinstance(datos_periodo,
                          dict) and "parametros_optimos" in datos_periodo and "metricas" in datos_periodo:
                params = datos_periodo["parametros_optimos"]
                metricas = datos_periodo["metricas"]

                historial_analisis_por_ticker[ticker].append({
                    "periodo": nombre_periodo.replace('_', ' ').title(),
                    "objetivo": "Rentabilidad",
                    "compra_pct": params.get('compra_pct', 0),
                    "venta_pct": params.get('venta_pct', 0),
                    "ganancia_min": params.get('ganancia_minima_pct', 0),
                    "suave_pct": params.get('suave_pct', 0),
                    "compra_mult": params.get('compra_multiple'),
                    "venta_mult": params.get('venta_multiple'),
                    "rentabilidad_max": metricas.get('rentabilidad_max', 0),
                    "margen_promedio": metricas.get('margen_promedio', 0)
                })

    # 2. Luego procesar estructura NUEVA (fuera de "periodos")
    for nombre_periodo, contenido in info.items():
        # Saltar claves de estructura antigua
        if nombre_periodo in ["ticker", "fecha_guardado", "periodos"]:
            continue

        # Procesar estructura nueva: periodo → objetivo → datos
        if isinstance(contenido, dict):
            for objetivo, datos in contenido.items():
                if isinstance(datos, dict) and "parametros_optimos" in datos and "metricas" in datos:
                    params = datos["parametros_optimos"]
                    metricas = datos["metricas"]

                    historial_analisis_por_ticker[ticker].append({
                        "periodo": nombre_periodo.replace('_', ' ').title(),
                        "objetivo": objetivo.replace('_', ' ').title(),
                        "compra_pct": params.get('compra_pct', 0),
                        "venta_pct": params.get('venta_pct', 0),
                        "ganancia_min": params.get('ganancia_minima_pct', 0),
                        "suave_pct": params.get('suave_pct', 0),
                        "compra_mult": params.get('compra_multiple'),
                        "venta_mult": params.get('venta_multiple'),
                        "rentabilidad_max": metricas.get('rentabilidad_max', 0),
                        "margen_promedio": metricas.get('margen_promedio', 0)
                    })

    # Ordenar por período y luego por objetivo
    orden_periodos = {"Completo": 1, "6 Meses": 2, "3 Meses": 3}
    historial_analisis_por_ticker[ticker].sort(
        key=lambda x: (orden_periodos.get(x['periodo'], 99), x['objetivo'])
    )

    # Mostrar tabla consolidada en frame_stats
    mostrar_tabla_consolidada_desde_json(ticker)


def administrar_json():
    """Abre una ventana para ver y eliminar registros del JSON"""
    if ARCHIVO_JSON is None or not ARCHIVO_JSON.exists():
        messagebox.showinfo("Sin datos", "No hay archivo JSON configurado o no existe")
        return

    datos_json = cargar_resultados_json()
    if not datos_json:
        messagebox.showinfo("Sin datos", "El archivo JSON está vacío")
        return

    # Obtener ticker actual del CSV seleccionado (para actualizar tabla después de eliminar)
    ticker_actual = None
    try:
        ruta_csv = entry_ruta.get().strip().strip('"')
        if ruta_csv and os.path.exists(ruta_csv):
            ticker_actual = os.path.splitext(os.path.basename(ruta_csv))[0]
    except:
        pass

    # Crear ventana
    ventana_admin = tk.Toplevel(ventana)
    ventana_admin.title("Administrar registros JSON")
    ventana_admin.geometry("1600x550")
    ventana_admin.transient(ventana)
    ventana_admin.grab_set()

    # Frame superior con instrucciones
    frame_instrucciones = tk.Frame(ventana_admin, pady=5)
    frame_instrucciones.pack(fill="x", padx=10)
    tk.Label(frame_instrucciones,
             text="Selecciona los registros que deseas eliminar (puedes seleccionar múltiples con Ctrl+clic)",
             font=("Arial", 9), fg="gray").pack(anchor="w")

    # Frame para el Treeview
    frame_tree = tk.Frame(ventana_admin)
    frame_tree.pack(fill="both", expand=True, padx=10, pady=5)

    # Scrollbars
    scrollbar_y = tk.Scrollbar(frame_tree, orient="vertical")
    scrollbar_x = tk.Scrollbar(frame_tree, orient="horizontal")

    # Treeview con selección múltiple
    # Columnas completas incluyendo todas las estadísticas
    # ORDEN: Básicos, Parámetros, Métricas (incluyendo Prom.Max/Min%), Estadísticas var, Operaciones, Fecha
    columns = (
        "Symbol", "Período", "Objetivo",
        # Parámetros óptimos
        "Compra%", "Venta%", "Gan.Mín%", "Compra N", "Venta N", "Límite", "Valor Lím.",
        # Métricas (Prom.Max% y Prom.Min% movidos aquí, después de Margen.Prom)
        "Rentab.Máx", "Margen.Prom", "Prom.Max%", "Prom.Min%", "Rentab.Prom", "Max.Margen", "Max.Aporte",
        # Estadísticas % variación
        "Max.Var%", "Min.Var%", "Fecha Max.Var", "Fecha Min.Var", "Dif.Var%", "Prom.Subida%", "Prom.Bajada%", "Dif.Prom%",
        # Estadísticas operaciones
        "Opc.Compra", "Acc.Compradas", "Opc.Venta", "Acc.Vendidas", "Max.Acc.Cart",
        # Fecha guardado
        "Fecha"
    )
    tree = ttk.Treeview(frame_tree, columns=columns, show="headings",
                        selectmode="extended",
                        yscrollcommand=scrollbar_y.set,
                        xscrollcommand=scrollbar_x.set)

    scrollbar_y.config(command=tree.yview)
    scrollbar_x.config(command=tree.xview)

    # Configurar columnas con anchos basados en el título (caracteres * 8 + margen)
    for col in columns:
        tree.heading(col, text=col)
        # Ancho basado en longitud del título
        ancho = max(len(col) * 8 + 10, 50)  # Mínimo 50px
        # Columnas de fecha más anchas
        if "Fecha" in col and col != "Fecha":
            ancho = max(ancho, 90)
        elif col == "Fecha":
            ancho = 130
        tree.column(col, width=ancho, anchor="center")

    # Diccionario para mapear items del tree a rutas en el JSON
    item_to_path = {}

    # Llenar el Treeview con datos (ordenados alfabéticamente por ticker_symbol)
    # Primero extraer y ordenar los tickers
    tickers_ordenados = sorted(datos_json.items(),
                                key=lambda x: (x[1].get("_ticker_symbol") or extraer_ticker_symbol(x[0]) or x[0]).upper())

    for ticker, contenido_ticker in tickers_ordenados:
        # Obtener ticker_symbol: desde el JSON o extraerlo del nombre
        ticker_symbol = contenido_ticker.get("_ticker_symbol") or extraer_ticker_symbol(ticker) or ticker

        # Manejar estructura antigua (con "periodos")
        if "periodos" in contenido_ticker and isinstance(contenido_ticker["periodos"], dict):
            for periodo, datos_periodo in contenido_ticker["periodos"].items():
                if isinstance(datos_periodo, dict) and "parametros_optimos" in datos_periodo:
                    params = datos_periodo.get("parametros_optimos", {})
                    metricas = datos_periodo.get("metricas", {})
                    stats_var = datos_periodo.get("estadisticas_var", {})
                    stats_ops = datos_periodo.get("estadisticas_operaciones", {})
                    fecha = datos_periodo.get("fecha_guardado", "")
                    compra_mult = params.get("compra_multiple")
                    venta_mult = params.get("venta_multiple")

                    limite_tipo = params.get("limite_tipo", "acciones")
                    limite_valor = params.get("limite_valor", 10.0)

                    item_id = tree.insert("", "end", values=(
                        ticker_symbol,
                        periodo.replace("_", " ").title(),
                        "Rentabilidad",  # Estructura antigua no tenía objetivo explícito
                        # Parámetros óptimos
                        f"{params.get('compra_pct', 0):.2f}",
                        f"{params.get('venta_pct', 0):.2f}",
                        f"{params.get('ganancia_minima_pct', 0):.2f}",
                        compra_mult if compra_mult else "-",
                        venta_mult if venta_mult else "-",
                        limite_tipo.title(),
                        f"{limite_valor:.0f}" if limite_tipo == "acciones" else f"${limite_valor:.0f}",
                        # Métricas (Prom.Max% y Prom.Min% después de Margen.Prom)
                        f"{metricas.get('rentabilidad_max', 0):.2f}%",
                        f"{metricas.get('margen_promedio', 0):.2f}",
                        f"{params.get('promedio_maximos', 0) / 100:.2f}%",
                        f"{params.get('promedio_minimos', 0) / 100:.2f}%",
                        f"{metricas.get('rentab_promedio', 0):.2f}%",
                        f"{metricas.get('max_margen', 0):.2f}",
                        f"{metricas.get('max_aporte', 0):.0f}",
                        # Estadísticas % variación (con símbolos %)
                        f"{stats_var.get('max_var', 0):.2f}%",
                        f"{stats_var.get('min_var', 0):.2f}%",
                        stats_var.get('fecha_max_var', '-'),
                        stats_var.get('fecha_min_var', '-'),
                        f"{stats_var.get('dif_var', 0):.2f}%",
                        f"{stats_var.get('max_prom_var', 0):.2f}%",
                        f"{stats_var.get('min_prom_var', 0):.2f}%",
                        f"{stats_var.get('dif_prom_var', 0):.2f}%",
                        # Estadísticas operaciones
                        stats_ops.get('opc_compra', 0),
                        stats_ops.get('acciones_compradas', 0),
                        stats_ops.get('opc_venta', 0),
                        stats_ops.get('acciones_vendidas', 0),
                        stats_ops.get('max_acc_cartera', 0),
                        # Fecha
                        fecha
                    ))
                    item_to_path[item_id] = {"ticker": ticker, "path": ["periodos", periodo], "tipo": "antiguo"}

        # Manejar estructura nueva (período → objetivo → datos)
        for periodo, contenido_periodo in contenido_ticker.items():
            if periodo in ["ticker", "fecha_guardado", "periodos", "_ticker_symbol"]:
                continue

            if isinstance(contenido_periodo, dict):
                for objetivo, datos in contenido_periodo.items():
                    if isinstance(datos, dict) and "parametros_optimos" in datos:
                        params = datos.get("parametros_optimos", {})
                        metricas = datos.get("metricas", {})
                        stats_var = datos.get("estadisticas_var", {})
                        stats_ops = datos.get("estadisticas_operaciones", {})
                        fecha = datos.get("fecha_guardado", "")
                        # También intentar obtener ticker_symbol del registro individual
                        symbol_mostrar = datos.get("ticker_symbol") or ticker_symbol
                        compra_mult = params.get("compra_multiple")
                        venta_mult = params.get("venta_multiple")
                        limite_tipo = params.get("limite_tipo", "acciones")
                        limite_valor = params.get("limite_valor", 10.0)

                        item_id = tree.insert("", "end", values=(
                            symbol_mostrar,
                            periodo.replace("_", " ").title(),
                            objetivo.replace("_", " ").title(),
                            # Parámetros óptimos
                            f"{params.get('compra_pct', 0):.2f}",
                            f"{params.get('venta_pct', 0):.2f}",
                            f"{params.get('ganancia_minima_pct', 0):.2f}",
                            compra_mult if compra_mult else "-",
                            venta_mult if venta_mult else "-",
                            limite_tipo.title(),
                            f"{limite_valor:.0f}" if limite_tipo == "acciones" else f"${limite_valor:.0f}",
                            # Métricas (Prom.Max% y Prom.Min% después de Margen.Prom)
                            f"{metricas.get('rentabilidad_max', 0):.2f}%",
                            f"{metricas.get('margen_promedio', 0):.2f}",
                            f"{params.get('promedio_maximos', 0) / 100:.2f}%",
                            f"{params.get('promedio_minimos', 0) / 100:.2f}%",
                            f"{metricas.get('rentab_promedio', 0):.2f}%",
                            f"{metricas.get('max_margen', 0):.2f}",
                            f"{metricas.get('max_aporte', 0):.0f}",
                            # Estadísticas % variación (con símbolos %)
                            f"{stats_var.get('max_var', 0):.2f}%",
                            f"{stats_var.get('min_var', 0):.2f}%",
                            stats_var.get('fecha_max_var', '-'),
                            stats_var.get('fecha_min_var', '-'),
                            f"{stats_var.get('dif_var', 0):.2f}%",
                            f"{stats_var.get('max_prom_var', 0):.2f}%",
                            f"{stats_var.get('min_prom_var', 0):.2f}%",
                            f"{stats_var.get('dif_prom_var', 0):.2f}%",
                            # Estadísticas operaciones
                            stats_ops.get('opc_compra', 0),
                            stats_ops.get('acciones_compradas', 0),
                            stats_ops.get('opc_venta', 0),
                            stats_ops.get('acciones_vendidas', 0),
                            stats_ops.get('max_acc_cartera', 0),
                            # Fecha
                            fecha
                        ))
                        item_to_path[item_id] = {"ticker": ticker, "path": [periodo, objetivo], "tipo": "nuevo"}

    # Empaquetar Treeview y scrollbars
    scrollbar_y.pack(side="right", fill="y")
    scrollbar_x.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    # Frame inferior con botones
    frame_botones = tk.Frame(ventana_admin, pady=10)
    frame_botones.pack(fill="x", padx=10)

    # Label para mostrar cantidad seleccionada
    label_seleccion = tk.Label(frame_botones, text="0 registros seleccionados", font=("Arial", 9))
    label_seleccion.pack(side="left")

    def actualizar_contador(event=None):
        cantidad = len(tree.selection())
        label_seleccion.config(text=f"{cantidad} registro(s) seleccionado(s)")

    tree.bind("<<TreeviewSelect>>", actualizar_contador)

    def eliminar_seleccionados():
        seleccionados = tree.selection()
        if not seleccionados:
            messagebox.showwarning("Sin selección", "No has seleccionado ningún registro")
            return

        # Confirmar eliminación
        cantidad = len(seleccionados)
        if not messagebox.askyesno("Confirmar eliminación",
                                    f"¿Estás seguro de eliminar {cantidad} registro(s)?\n\nEsta acción no se puede deshacer."):
            return

        # Cargar JSON actual
        with open(ARCHIVO_JSON, 'r', encoding='utf-8') as f:
            datos = json.load(f)

        # Eliminar cada registro seleccionado
        eliminados = 0
        for item_id in seleccionados:
            if item_id in item_to_path:
                info = item_to_path[item_id]
                ticker = info["ticker"]
                path = info["path"]
                tipo = info["tipo"]

                try:
                    if tipo == "antiguo":
                        # Estructura: ticker → periodos → periodo
                        if ticker in datos and "periodos" in datos[ticker]:
                            if path[1] in datos[ticker]["periodos"]:
                                del datos[ticker]["periodos"][path[1]]
                                eliminados += 1
                                # Si periodos queda vacío, eliminarlo
                                if not datos[ticker]["periodos"]:
                                    del datos[ticker]["periodos"]
                    else:
                        # Estructura nueva: ticker → periodo → objetivo
                        if ticker in datos and path[0] in datos[ticker]:
                            if path[1] in datos[ticker][path[0]]:
                                del datos[ticker][path[0]][path[1]]
                                eliminados += 1
                                # Si periodo queda vacío, eliminarlo
                                if not datos[ticker][path[0]]:
                                    del datos[ticker][path[0]]

                    # Si ticker queda vacío (solo con claves vacías), eliminarlo
                    if ticker in datos:
                        claves_restantes = [k for k in datos[ticker].keys()
                                           if k not in ["ticker", "fecha_guardado"] and datos[ticker][k]]
                        if not claves_restantes:
                            del datos[ticker]

                except Exception as e:
                    print(f"Error eliminando {item_id}: {e}")

        # Guardar JSON actualizado
        with open(ARCHIVO_JSON, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)

        messagebox.showinfo("Eliminación completada", f"Se eliminaron {eliminados} registro(s)")

        # Actualizar el cuadro de parámetros en la interfaz principal
        if ticker_actual:
            # Recargar el historial del ticker y actualizar la tabla
            ruta_csv = entry_ruta.get().strip().strip('"')
            mostrar_info_json_ticker(ticker_actual, ruta_csv)

        # Cerrar y reabrir para refrescar
        ventana_admin.destroy()
        administrar_json()

    def seleccionar_todos():
        for item in tree.get_children():
            tree.selection_add(item)
        actualizar_contador()

    def deseleccionar_todos():
        tree.selection_remove(tree.get_children())
        actualizar_contador()

    def exportar_a_excel():
        """Exporta los datos del JSON a un archivo Excel"""
        # Obtener todos los items del treeview
        items = tree.get_children()
        if not items:
            messagebox.showwarning("Sin datos", "No hay datos para exportar")
            return

        # Preguntar dónde guardar
        ruta_excel = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfile="Parametros_Optimos.xlsx"
        )

        if not ruta_excel:
            return

        try:
            # Crear DataFrame con los datos (todas las columnas en nuevo orden)
            datos_export = []
            for item in items:
                v = tree.item(item, "values")
                datos_export.append({
                    "Symbol": v[0],
                    "Período": v[1],
                    "Objetivo": v[2],
                    # Parámetros óptimos
                    "Compra%": v[3],
                    "Venta%": v[4],
                    "Gan.Mín%": v[5],
                    "Compra N": v[6] if v[6] != "-" else "",
                    "Venta N": v[7] if v[7] != "-" else "",
                    "Límite": v[8],
                    "Valor Lím.": v[9],
                    # Métricas (Prom.Max% y Prom.Min% después de Margen.Prom)
                    "Rentab.Máx": v[10],
                    "Margen.Prom": v[11],
                    "Prom.Max%": v[12],
                    "Prom.Min%": v[13],
                    "Rentab.Prom": v[14],
                    "Max.Margen": v[15],
                    "Max.Aporte": v[16],
                    # Estadísticas % variación
                    "Max.Var%": v[17],
                    "Min.Var%": v[18],
                    "Fecha Max.Var": v[19],
                    "Fecha Min.Var": v[20],
                    "Dif.Var%": v[21],
                    "Prom.Subida%": v[22],
                    "Prom.Bajada%": v[23],
                    "Dif.Prom%": v[24],
                    # Estadísticas operaciones
                    "Opc.Compra": v[25],
                    "Acc.Compradas": v[26],
                    "Opc.Venta": v[27],
                    "Acc.Vendidas": v[28],
                    "Max.Acc.Cart": v[29],
                    # Fecha
                    "Fecha Guardado": v[30]
                })

            df_export = pd.DataFrame(datos_export)

            # Exportar a Excel con formato
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Parámetros Óptimos"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Escribir encabezados
            columnas = list(df_export.columns)
            for col_idx, col_name in enumerate(columnas, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Escribir datos
            for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            # Ajustar anchos de columna automáticamente
            from openpyxl.utils import get_column_letter
            for col_idx, col_name in enumerate(columnas, 1):
                col_letter = get_column_letter(col_idx)
                # Ancho basado en nombre de columna + margen
                ancho = max(len(str(col_name)) + 2, 10)
                # Columnas de fecha más anchas
                if "Fecha" in col_name:
                    ancho = max(ancho, 14)
                ws.column_dimensions[col_letter].width = ancho

            wb.save(ruta_excel)
            messagebox.showinfo("Exportación exitosa",
                               f"Datos exportados correctamente.\n\nArchivo: {ruta_excel}\nRegistros: {len(datos_export)}")

        except PermissionError:
            messagebox.showerror("Error", "El archivo está abierto. Ciérralo e intenta de nuevo.")
        except Exception as e:
            messagebox.showerror("Error al exportar", f"Error: {str(e)}")

    # Botones
    tk.Button(frame_botones, text="Seleccionar todos", command=seleccionar_todos).pack(side="left", padx=(20, 5))
    tk.Button(frame_botones, text="Deseleccionar todos", command=deseleccionar_todos).pack(side="left", padx=5)
    tk.Button(frame_botones, text="Exportar a Excel", command=exportar_a_excel,
              bg="#28a745", fg="white", font=("Arial", 9, "bold")).pack(side="left", padx=10)

    tk.Button(frame_botones, text="Eliminar seleccionados", command=eliminar_seleccionados,
              bg="#ff6b6b", fg="white", font=("Arial", 10, "bold")).pack(side="right", padx=5)

    tk.Button(frame_botones, text="Cerrar", command=ventana_admin.destroy).pack(side="right", padx=5)


def mostrar_tabla_consolidada_desde_json(ticker):
    """Muestra SOLO la tabla consolidada con datos del JSON (sin pestañas)"""
    # Limpiar frame de estadísticas
    for widget in ventana.frame_stats.winfo_children():
        widget.destroy()

    # Extraer ticker real (siglas) para mostrar
    if ticker:
        partes = ticker.split('_')
        ticker_display = ticker
        if len(partes) >= 2:
            for parte in partes:
                if parte.isupper() and 1 <= len(parte) <= 5:
                    ticker_display = parte
                    break
    else:
        ticker_display = "Actual"

    # Frame con tabla consolidada
    frame_consolidado = tk.Frame(ventana.frame_stats, relief="ridge", borderwidth=2, bg="lightyellow", padx=10, pady=10)
    frame_consolidado.pack(fill="x", pady=(10, 0))

    tk.Label(frame_consolidado, text=f"📊 PARÁMETROS ÓPTIMOS GUARDADOS - {ticker_display}",
             font=("Arial", 11, "bold"), bg="lightyellow", fg="darkgreen").pack(anchor="w")

    # Crear tabla con parámetros
    frame_tabla_params = tk.Frame(frame_consolidado, bg="lightyellow")
    frame_tabla_params.pack(fill="x", pady=(5, 0))

    # Headers
    headers = ["#", "Período", "Objetivo", "Compra %", "Venta %", "Gan Mín %", "Suave %", "Comp", "Venta", "Rentab Máx",
               "Margen Prom"]
    for col, header in enumerate(headers):
        ancho = 5 if col == 0 else 11
        tk.Label(frame_tabla_params, text=header, font=("Arial", 8, "bold"),
                 bg="lightblue", relief="solid", borderwidth=1, width=ancho).grid(row=0, column=col, sticky="ew",
                                                                                  padx=1, pady=1)

    # Datos del ticker actual
    analisis_ticker_actual = historial_analisis_por_ticker.get(ticker, [])

    if not analisis_ticker_actual:
        tk.Label(frame_consolidado, text="No hay análisis guardados en JSON",
                 font=("Arial", 9), bg="lightyellow", fg="gray").pack(pady=10)
        return

    # Colores por objetivo (base del objetivo, sin el número)
    colores_objetivo = {
        "rentabilidad": "#e8f5e9",  # Verde claro
        "margen": "#e3f2fd",         # Azul claro
    }
    color_default = "#fff3e0"  # Naranja claro para otros

    periodo_anterior = None
    fila_actual = 0

    for idx, analisis in enumerate(analisis_ticker_actual, start=1):
        fila_actual += 1

        # Determinar color basado en el objetivo (sin número)
        objetivo_base = analisis['objetivo'].lower().split()[0]
        bg_color = colores_objetivo.get(objetivo_base, color_default)

        # Si cambia el período, agregar línea separadora
        if periodo_anterior is not None and analisis['periodo'] != periodo_anterior:
            # Agregar fila separadora
            for col in range(11):
                ancho = 5 if col == 0 else 11
                tk.Label(frame_tabla_params, text="", font=("Arial", 2),
                         bg="#999999", relief="flat", width=ancho, height=1).grid(
                    row=fila_actual, column=col, sticky="ew", padx=1, pady=0)
            fila_actual += 1

        periodo_anterior = analisis['periodo']

        valores = [
            str(idx),
            analisis['periodo'],
            analisis['objetivo'],
            f"{analisis['compra_pct']:.1f}",
            f"{analisis['venta_pct']:.1f}",
            f"{analisis['ganancia_min']:.1f}",
            f"{analisis['suave_pct']:.1f}",
            str(analisis['compra_mult']) if analisis['compra_mult'] else "-",
            str(analisis['venta_mult']) if analisis['venta_mult'] else "-",
            f"{analisis['rentabilidad_max']:.2f}%",
            f"{analisis['margen_promedio']:.2f}"
        ]

        for col, valor in enumerate(valores):
            ancho = 5 if col == 0 else 11
            tk.Label(frame_tabla_params, text=valor, font=("Arial", 7),
                     bg=bg_color, relief="solid", borderwidth=1, width=ancho).grid(
                row=fila_actual, column=col, sticky="ew", padx=1, pady=1)


# =========================
# Funciones auxiliares
# =========================
def parse_percent_to_decimal(x):
    """Convierte valores de porcentaje a decimal"""
    try:
        if pd.isna(x):
            return float("nan")
        s = str(x).strip().replace(",", ".")
        if s.endswith("%"):
            s = s[:-1].strip()
            if s == "":
                return float("nan")
            return float(s) / 100.0
        try:
            f = float(s)
            if abs(f) <= 1:
                return f
            else:
                return f / 100.0
        except:
            return float("nan")
    except Exception:
        return float("nan")


def to_float_safe(x):
    """Convierte cadenas numéricas con coma/punto a float, devuelve NaN si falla."""
    try:
        if pd.isna(x):
            return float("nan")
        s = str(x).strip().replace('"', '').replace(",", ".")
        if s == "":
            return float("nan")
        return float(s)
    except:
        return float("nan")


def create_sqlite_from_df(folder, name, df):
    """Crea una base sqlite con la tabla 'precios' a partir del DataFrame."""
    db = os.path.join(folder, name)
    conn = sqlite3.connect(db)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS precios")
    cur.execute("""
        CREATE TABLE precios (
            Fecha TEXT,
            Ultimo REAL,
            Apertura REAL,
            Maximo REAL,
            Minimo REAL,
            Vol REAL,
            Var REAL
        )
    """)
    rows = []
    for _, r in df.iterrows():
        fecha = r.get("Fecha", "")
        try:
            ultimo = float(r["Último"]) if pd.notna(r["Último"]) and r["Último"] != "" else None
        except:
            ultimo = to_float_safe(r["Último"])
        try:
            apertura = float(r["Apertura"]) if pd.notna(r["Apertura"]) and r["Apertura"] != "" else None
        except:
            apertura = to_float_safe(r["Apertura"])
        try:
            maximo = float(r["Máximo"]) if pd.notna(r["Máximo"]) and r["Máximo"] != "" else None
        except:
            maximo = to_float_safe(r["Máximo"])
        try:
            minimo = float(r["Mínimo"]) if pd.notna(r["Mínimo"]) and r["Mínimo"] != "" else None
        except:
            minimo = to_float_safe(r["Mínimo"])
        try:
            vol = float(r["Vol."]) if pd.notna(r["Vol."]) and r["Vol."] != "" else None
        except:
            vol = to_float_safe(r["Vol."])

        var_val = r.get("% var.", None)
        if pd.isna(var_val) or var_val is None or str(var_val).strip() == "":
            var_num = None
        else:
            var_num = parse_percent_to_decimal(var_val)
            if pd.isna(var_num):
                var_num = None

        rows.append((fecha, ultimo, apertura, maximo, minimo, vol, var_num))

    cur.executemany("INSERT INTO precios VALUES (?,?,?,?,?,?,?)", rows)
    conn.commit()
    conn.close()
    return db


def filtrar_ultimos_dias(csv_path, dias):
    """Lee el CSV y devuelve un DataFrame con solo los últimos N días"""
    df = pd.read_csv(csv_path, sep=";", engine='python', dtype=str)
    df.columns = [c.strip() for c in df.columns]

    # Procesar fechas
    def parse_mixed_dates(date_str):
        for fmt in ("%d/%m/%Y", "%m/%d/%Y"):
            try:
                return pd.to_datetime(date_str, format=fmt)
            except:
                continue
        return pd.NaT

    df['Fecha'] = df['Fecha'].apply(parse_mixed_dates)
    df = df.dropna(subset=['Fecha'])
    df = df.sort_values('Fecha').reset_index(drop=True)

    # Obtener fecha más reciente y calcular fecha de corte
    fecha_max = df['Fecha'].max()
    fecha_corte = fecha_max - timedelta(days=dias)

    # Filtrar
    df_filtrado = df[df['Fecha'] >= fecha_corte].copy()

    print(f"  → Período: {fecha_corte.strftime('%d/%m/%Y')} a {fecha_max.strftime('%d/%m/%Y')}")
    print(f"  → Registros: {len(df_filtrado)} de {len(df)}")

    return df_filtrado


# =========================
# Interfaz Gráfica
# =========================
ventana = tk.Tk()
ventana.title("Parámetros del análisis")
ventana.geometry("1100x700")  # Tamaño inicial de la ventana

# =========================================================
# Crear Canvas con Scrollbar vertical para toda la interfaz
# =========================================================
canvas_principal = tk.Canvas(ventana)
scrollbar_vertical = tk.Scrollbar(ventana, orient="vertical", command=canvas_principal.yview)
canvas_principal.configure(yscrollcommand=scrollbar_vertical.set)

scrollbar_vertical.pack(side="right", fill="y")
canvas_principal.pack(side="left", fill="both", expand=True)

# Frame principal dentro del canvas (aquí van todos los widgets)
frame_principal = tk.Frame(canvas_principal)
canvas_window = canvas_principal.create_window((0, 0), window=frame_principal, anchor="nw")

# Configurar el scroll para que funcione con la rueda del mouse
def on_mousewheel(event):
    canvas_principal.yview_scroll(int(-1*(event.delta/120)), "units")

canvas_principal.bind_all("<MouseWheel>", on_mousewheel)

# Actualizar el tamaño del canvas cuando cambie el frame
def configurar_scroll(event):
    canvas_principal.configure(scrollregion=canvas_principal.bbox("all"))
    # Ajustar el ancho del canvas al frame
    canvas_principal.itemconfig(canvas_window, width=event.width if event.width > 1000 else 1000)

frame_principal.bind("<Configure>", configurar_scroll)

# Ajustar el ancho del frame cuando cambie el canvas
def on_canvas_configure(event):
    canvas_principal.itemconfig(canvas_window, width=event.width)

canvas_principal.bind("<Configure>", on_canvas_configure)

# Configuración del grid del frame principal
frame_principal.grid_columnconfigure(1, weight=1)
frame_principal.grid_columnconfigure(0, weight=0)
frame_principal.grid_columnconfigure(2, weight=0)
frame_principal.grid_columnconfigure(3, weight=1)

tk.Label(frame_principal, text="Ruta del CSV (TAB):").grid(row=0, column=0, sticky="w")
entry_ruta = tk.Entry(frame_principal, width=55)
entry_ruta.grid(row=0, column=1, sticky="we")


def seleccionar_csv():
    global ticker_actual

    ruta = filedialog.askopenfilename(
        title="Selecciona el archivo CSV (guardado desde Excel, separado por TAB)",
        filetypes=[("CSV files", "*.csv;*.txt"), ("Todos los archivos", "*.*")]
    )
    entry_ruta.delete(0, tk.END)
    entry_ruta.insert(0, ruta)

    # Mostrar info del JSON si existe
    if ruta:
        nombre_archivo = os.path.splitext(os.path.basename(ruta))[0]
        ticker_actual = nombre_archivo  # Guardar ticker actual
        mostrar_info_json_ticker(nombre_archivo, ruta)  # CORREGIDO: Pasar ruta del CSV


tk.Button(frame_principal, text="Seleccionar", command=seleccionar_csv).grid(row=0, column=2, sticky="w", padx=(6, 0))

# Botón para seleccionar ubicación JSON (row=1, pegado a row=0)
frame_json_config = tk.Frame(frame_principal)
frame_json_config.grid(row=1, column=0, columnspan=3, sticky="w", pady=(1, 0))

tk.Button(frame_json_config, text="Configurar ubicación JSON",
          command=seleccionar_ubicacion_json, bg="lightblue").pack(side="left")

tk.Button(frame_json_config, text="Administrar JSON",
          command=administrar_json, bg="#ffcc80").pack(side="left", padx=(10, 0))

tk.Button(frame_json_config, text="Params Activos",
          command=administrar_parametros_activos, bg="#90EE90").pack(side="left", padx=(10, 0))

label_json_actual = tk.Label(frame_json_config, text="JSON: No configurado", fg="gray")
label_json_actual.pack(side="left", padx=(10, 0))

# AHORA sí cargar configuración (después de crear label_json_actual)
cargar_configuracion()

# Frame para mostrar info JSON del ticker seleccionado (DEBAJO en row=2, ancho completo)
frame_info_json = tk.Frame(frame_principal, relief="groove", borderwidth=2, padx=5, pady=3)
frame_info_json.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(1, 0))

# ------------------------------------------------
# CHECKBOXES: Objetivo de optimización (múltiple selección)
# ------------------------------------------------
tk.Label(frame_principal, text="Objetivo optimización:").grid(row=3, column=0, sticky="w")
# Variables para checkboxes de objetivos
objetivo_rentabilidad_var = tk.IntVar(value=1)  # Por defecto marcado
objetivo_margen_var = tk.IntVar(value=0)
frame_objetivo = tk.Frame(frame_principal)
frame_objetivo.grid(row=3, column=1, sticky="w")
tk.Checkbutton(frame_objetivo, text="Rentabilidad máx", variable=objetivo_rentabilidad_var).pack(side="left")
tk.Checkbutton(frame_objetivo, text="Margen promedio máx", variable=objetivo_margen_var).pack(side="left", padx=(10, 0))

# Función helper para obtener objetivos seleccionados
def obtener_objetivos_seleccionados():
    """Retorna lista de objetivos seleccionados"""
    objetivos = []
    if objetivo_rentabilidad_var.get() == 1:
        objetivos.append("rentabilidad")
    if objetivo_margen_var.get() == 1:
        objetivos.append("margen_prom")
    return objetivos

# CHECKBOX: Usar optimización SciPy
usar_scipy_var = tk.IntVar(value=0)
chk_scipy = tk.Checkbutton(frame_principal, text="Usar optimización avanzada (SciPy)", variable=usar_scipy_var)
chk_scipy.grid(row=3, column=2, sticky="w", padx=(10, 0))

# ------------------------------------------------
# CAMPO Compra (%) + CHECKBOX DE OPTIMIZACIÓN
# ------------------------------------------------
tk.Label(frame_principal, text="Compra (%):").grid(row=4, column=0, sticky="w")

frame_compra = tk.Frame(frame_principal)
frame_compra.grid(row=4, column=1, sticky="w")

entry_compra = tk.Entry(frame_compra, width=6)
entry_compra.insert(0, "-1.6")
entry_compra.pack(side="left")

auto_compra_var = tk.IntVar(value=0)
chk_auto = tk.Checkbutton(frame_compra, text="Auto", variable=auto_compra_var)
chk_auto.pack(side="left", padx=(5, 0))

# Frame para botones de análisis
frame_botones_analisis = tk.Frame(frame_principal)
frame_botones_analisis.grid(row=4, column=2, sticky="w", padx=(10, 0))

btn_iniciar_analisis = tk.Button(frame_botones_analisis, text="▶ Iniciar análisis",
                                  command=lambda: iniciar_proceso(), bg="#90EE90")
btn_iniciar_analisis.pack(side="left")

btn_detener_analisis = tk.Button(frame_botones_analisis, text="⏹ Detener",
                                  command=lambda: detener_analisis(), bg="#ff6b6b", fg="white", state="disabled")
btn_detener_analisis.pack(side="left", padx=(5, 0))

# ------------------------------------------------
# CAMPO Venta (%) + CHECKBOX DE OPTIMIZACIÓN
# ------------------------------------------------
tk.Label(frame_principal, text="Venta (%):").grid(row=5, column=0, sticky="w")

frame_venta = tk.Frame(frame_principal)
frame_venta.grid(row=5, column=1, sticky="w")

entry_venta = tk.Entry(frame_venta, width=6)
entry_venta.insert(0, "1.6")
entry_venta.pack(side="left")

auto_venta_var = tk.IntVar(value=0)
chk_auto_venta = tk.Checkbutton(frame_venta, text="Auto", variable=auto_venta_var)
chk_auto_venta.pack(side="left", padx=(5, 0))

# Botón "Generar DB y Excel" al lado de Venta (azul con letras negras)
btn_generar_db_excel = tk.Button(frame_principal, text="Generar DB y Excel", command=lambda: generar_db_excel(),
                                 bg="#1E90FF", fg="black", font=("Arial", 9, "bold"), width=18)
btn_generar_db_excel.grid(row=5, column=2, sticky="w", padx=(10, 0))

# ------------------------------------------------
# CAMPO: Ganancia mínima (%) + CHECKBOX
# ------------------------------------------------
tk.Label(frame_principal, text="Ganancia mínima (%):").grid(row=6, column=0, sticky="w")

frame_ganancia = tk.Frame(frame_principal)
frame_ganancia.grid(row=6, column=1, sticky="w")

entry_ganancia_minima = tk.Entry(frame_ganancia, width=6)
entry_ganancia_minima.insert(0, "0")
entry_ganancia_minima.pack(side="left")

auto_ganancia_var = tk.IntVar(value=0)
chk_auto_ganancia = tk.Checkbutton(frame_ganancia, text="Auto", variable=auto_ganancia_var)
chk_auto_ganancia.pack(side="left", padx=(5, 0))

tk.Label(frame_principal, text="Suave (%):").grid(row=7, column=0, sticky="w")
entry_suave = tk.Entry(frame_principal, width=6)
entry_suave.insert(0, "0.5")
entry_suave.grid(row=7, column=1, sticky="w")

tipo_limite_var = tk.StringVar(value="acciones")
opciones_limite = ["acciones", "aporte"]
selector_limite = tk.OptionMenu(frame_principal, tipo_limite_var, *opciones_limite)
selector_limite.grid(row=8, column=0, sticky="w")

frame_limite = tk.Frame(frame_principal)
frame_limite.grid(row=8, column=1, sticky="w")
entry_limite = tk.Entry(frame_limite, width=10)
entry_limite.insert(0, "10")
entry_limite.pack(side="left")
tk.Label(frame_limite, text="Valor límite").pack(side="left", padx=(5, 0))

# =========================================================
# Frame para Compra múltiple
# =========================================================
frame_compra_multiple = tk.Frame(frame_principal)
frame_compra_multiple.grid(row=9, column=0, columnspan=2, sticky="w", pady=(5, 0))

tk.Label(frame_compra_multiple, text="Compra de N acciones:").pack(side="left")

entry_compra_multiple = tk.Entry(frame_compra_multiple, width=6)
entry_compra_multiple.pack(side="left", padx=(5, 0))
entry_compra_multiple.insert(0, "")

auto_compra_mult_var = tk.IntVar(value=0)
chk_auto_compra_mult = tk.Checkbutton(frame_compra_multiple, text="Auto", variable=auto_compra_mult_var)
chk_auto_compra_mult.pack(side="left", padx=(5, 0))

# =========================================================
# Frame para Venta múltiple
# =========================================================
frame_venta_multiple = tk.Frame(frame_principal)
frame_venta_multiple.grid(row=10, column=0, columnspan=2, sticky="w", pady=(5, 0))

tk.Label(frame_venta_multiple, text="Venta de N acciones:").pack(side="left")

entry_venta_multiple = tk.Entry(frame_venta_multiple, width=6)
entry_venta_multiple.pack(side="left", padx=(5, 0))
entry_venta_multiple.insert(0, "")

auto_venta_mult_var = tk.IntVar(value=0)
chk_auto_venta_mult = tk.Checkbutton(frame_venta_multiple, text="Auto", variable=auto_venta_mult_var)
chk_auto_venta_mult.pack(side="left", padx=(5, 0))

# =========================================================
# Frame para mostrar fechas de compras/ventas múltiples
# =========================================================
frame_fechas_multiples = tk.Frame(frame_principal)
frame_fechas_multiples.grid(row=9, column=2, rowspan=2, sticky="nw", padx=(10, 0))

# Cuadro de fechas de compras múltiples
tk.Label(frame_fechas_multiples, text="Fechas compras múltiples:", font=("Arial", 8)).grid(row=0, column=0, sticky="w")
text_compras_mult = tk.Text(frame_fechas_multiples, width=15, height=4, font=("Arial", 7))
text_compras_mult.grid(row=1, column=0, sticky="w", padx=(0, 10))

# Cuadro de fechas de ventas múltiples
tk.Label(frame_fechas_multiples, text="Fechas ventas múltiples:", font=("Arial", 8)).grid(row=0, column=1, sticky="w")
text_ventas_mult = tk.Text(frame_fechas_multiples, width=15, height=4, font=("Arial", 7))
text_ventas_mult.grid(row=1, column=1, sticky="w")

# =========================================================
# Frame para selección de períodos a analizar
# =========================================================
frame_periodos = tk.Frame(frame_principal, relief="ridge", borderwidth=2, padx=10, pady=5)
frame_periodos.grid(row=11, column=0, columnspan=3, sticky="w", pady=(10, 0))

tk.Label(frame_periodos, text="Analizar períodos:", font=("Arial", 10, "bold")).pack(side="left", padx=(0, 10))

analizar_completo_var = tk.IntVar(value=1)
analizar_6meses_var = tk.IntVar(value=0)
analizar_3meses_var = tk.IntVar(value=0)

tk.Checkbutton(frame_periodos, text="Completo", variable=analizar_completo_var).pack(side="left", padx=5)
tk.Checkbutton(frame_periodos, text="Últimos 6 meses", variable=analizar_6meses_var).pack(side="left", padx=5)
tk.Checkbutton(frame_periodos, text="Últimos 3 meses", variable=analizar_3meses_var).pack(side="left", padx=5)

# Botón verde para guardar en JSON
btn_guardar_json = tk.Button(frame_periodos, text="💾 Guardar resultados en JSON",
                             command=guardar_resultados_en_json, bg="lightgreen",
                             font=("Arial", 10, "bold"), state="disabled")
btn_guardar_json.pack(side="left", padx=(20, 0))

# =========================================================
# Frame de estadísticas
# =========================================================
ventana.frame_stats = tk.Frame(frame_principal, padx=10, pady=2)
ventana.frame_stats.grid(row=12, column=0, columnspan=3, sticky="w")

# =========================================================
# Barra de progreso para optimización
# =========================================================
frame_progreso = tk.Frame(frame_principal, padx=10, pady=2)
frame_progreso.grid(row=13, column=0, columnspan=3, sticky="we")

ventana.progress_bar = ttk.Progressbar(frame_progreso, length=600, mode='determinate')
ventana.label_progreso = tk.Label(frame_progreso, text="", font=("Arial", 10))

# Label para mostrar resultado de optimización
ventana.label_resultado_opt = tk.Label(frame_principal, text="", font=("Arial", 10, "bold"), fg="darkgreen")

ultimo_df = None
ultima_ruta_excel = ""
ultimo_folder = ""
ultimo_base_name = ""

# Diccionario para almacenar DataFrames por período
resultados_dfs_por_periodo = {}

# Variable global para acumular análisis POR TICKER (no mezclar tickers)
historial_analisis_por_ticker = {}
ticker_actual = None


# =========================
# CAMBIO 1 y 2: Función generar DB y Excel (botón)
# =========================
def generar_db_excel():
    global resultados_dfs_por_periodo, ultimo_folder, ultimo_base_name

    if not resultados_dfs_por_periodo:
        messagebox.showerror("Error", "No hay análisis previo. Ejecuta primero 'Iniciar análisis'.")
        return

    # Cambiar estado del botón
    btn_generar_db_excel.config(state="disabled", bg="gray", text="Generando...")
    ventana.update()

    mensajes_resultado = []
    archivos_generados = []
    errores = []

    try:
        # CAMBIO 1: Excel que ACUMULA pestañas (no sobrescribe)
        ruta_excel = os.path.join(ultimo_folder, f"{ultimo_base_name}_analizado.xlsx")
        objetivo = OBJETIVO_ACTUAL

        # Si el archivo existe, cargar y agregar nuevas pestañas
        if os.path.exists(ruta_excel):
            from openpyxl import load_workbook
            try:
                wb = load_workbook(ruta_excel)
            except PermissionError:
                errores.append(f"❌ Excel: El archivo está abierto, ciérralo primero")
                wb = None
        else:
            from openpyxl import Workbook
            wb = Workbook()
            # Eliminar la hoja por defecto si existe
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        if wb is not None:
            for nombre_periodo, df in resultados_dfs_por_periodo.items():
                # Crear nombre de pestaña descriptivo
                nombre_hoja = f"{nombre_periodo}_{objetivo}"[:31]

                # Si la pestaña ya existe, eliminarla para actualizarla
                if nombre_hoja in wb.sheetnames:
                    del wb[nombre_hoja]

                # Crear nueva pestaña
                ws = wb.create_sheet(nombre_hoja)

                # Escribir datos
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(ruta_excel)
            archivos_generados.append(
                f"✓ Excel: {os.path.basename(ruta_excel)} ({len(resultados_dfs_por_periodo)} pestañas)")

    except PermissionError:
        errores.append(f"❌ Excel: El archivo está abierto, ciérralo primero")
    except Exception as e:
        errores.append(f"❌ Excel: {str(e)}")

    try:
        # CAMBIO 2: SQLite unificado con múltiples tablas
        db_path = os.path.join(ultimo_folder, f"{ultimo_base_name}_analizado.db")

        conn = sqlite3.connect(db_path)

        for nombre_periodo, df in resultados_dfs_por_periodo.items():
            # Nombre de tabla: periodo_objetivo
            objetivo = OBJETIVO_ACTUAL
            tabla_nombre = f"{nombre_periodo}_{objetivo}"

            # Eliminar tabla si existe (con comillas)
            conn.execute(f'DROP TABLE IF EXISTS "{tabla_nombre}"')

            # Crear tabla (con comillas para nombres que empiezan con números)
            conn.execute(f"""
                CREATE TABLE "{tabla_nombre}" (
                    Fecha TEXT,
                    Ultimo REAL,
                    Apertura REAL,
                    Maximo REAL,
                    Minimo REAL,
                    Vol REAL,
                    Var TEXT,
                    Acumulado TEXT,
                    Opcion TEXT,
                    Movimiento INTEGER,
                    Acciones INTEGER,
                    PrecioCompra REAL,
                    CapitalBolsa REAL,
                    CapitalAcciones REAL,
                    CapitalTotal REAL,
                    Aporte REAL,
                    AporteAcumulado REAL,
                    Margen REAL,
                    Rentabilidad TEXT
                )
            """)

            # Insertar datos
            rows = []
            for _, r in df.iterrows():
                rows.append((
                    r.get("Fecha", ""),
                    to_float_safe(r.get("Último", 0)),
                    to_float_safe(r.get("Apertura", 0)),
                    to_float_safe(r.get("Máximo", 0)),
                    to_float_safe(r.get("Mínimo", 0)),
                    to_float_safe(r.get("Vol.", 0)),
                    str(r.get("% var.", "")),
                    str(r.get("% acumulado", "")),
                    str(r.get("Opción", "")),
                    int(r.get("Movimiento de acciones", 0)) if pd.notna(r.get("Movimiento de acciones", 0)) else 0,
                    int(r.get("Acciones en cartera", 0)) if pd.notna(r.get("Acciones en cartera", 0)) else 0,
                    to_float_safe(r.get("Precio de compra", 0)),
                    to_float_safe(r.get("Capital en bolsa", 0)),
                    to_float_safe(r.get("Capital en acciones", 0)),
                    to_float_safe(r.get("Capital total", 0)),
                    to_float_safe(r.get("Aporte", 0)),
                    to_float_safe(r.get("Aporte acumulado", 0)),
                    to_float_safe(r.get("Margen", 0)),
                    str(r.get("Rentabilidad", ""))
                ))

            placeholders = ",".join(["?"] * 19)
            conn.executemany(f'INSERT INTO "{tabla_nombre}" VALUES ({placeholders})', rows)

        conn.commit()
        conn.close()

        archivos_generados.append(f"✓ SQLite: {os.path.basename(db_path)} ({len(resultados_dfs_por_periodo)} tablas)")

    except Exception as e:
        errores.append(f"❌ SQLite: {str(e)}")

    # MEJORA: Una sola ventana de diálogo con todos los resultados
    mensaje_final = ""

    if archivos_generados:
        mensaje_final += "ARCHIVOS GENERADOS:\n\n" + "\n".join(archivos_generados)

    if errores:
        if mensaje_final:
            mensaje_final += "\n\n"
        mensaje_final += "ERRORES:\n\n" + "\n".join(errores)

    # Si todo fue exitoso, deshabilitar botón (como el de JSON)
    if not errores and archivos_generados:
        btn_generar_db_excel.config(state="disabled", bg="lightgray", fg="gray", text="Generar DB y Excel")
    else:
        # Si hubo errores, restaurar botón para reintentar
        btn_generar_db_excel.config(state="normal", bg="#1E90FF", fg="black", text="Generar DB y Excel")

    if errores:
        messagebox.showwarning("Generación completada con errores", mensaje_final)
    else:
        messagebox.showinfo("Generación exitosa", mensaje_final)


# =========================
# Función que ejecuta TODO el análisis con un UMBRAL_COMPRA dado
# =========================
def ejecutar_analisis_con_umbral(umbral_compra_decimal, csv_filtrado=None):
    global text_ventas_mult, text_compras_mult, INPUT_FILE

    try:
        local_venta = float(entry_venta.get().replace(",", ".")) / 100
        local_suave = float(entry_suave.get().replace(",", ".")) / 100
        ganancia_minima = float(entry_ganancia_minima.get().replace(",", ".")) / 100
    except:
        messagebox.showerror("Error", "Valores numéricos inválidos en Venta / Suave / Ganancia mínima.")
        return None, -999999, -999999

    # Si se proporciona un CSV filtrado, usarlo; si no, cargar el original
    if csv_filtrado is not None:
        df = csv_filtrado.copy()
    else:
        try:
            df = pd.read_csv(INPUT_FILE, sep=";", engine='python', dtype=str)
            df.columns = [c.strip() for c in df.columns]
            df['Fecha'] = df['Fecha'].astype(str).str.strip()

            def parse_mixed_dates(date_str):
                for fmt in ("%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        return pd.to_datetime(date_str, format=fmt)
                    except:
                        continue
                return pd.NaT

            df['Fecha'] = df['Fecha'].apply(parse_mixed_dates)
            df = df.dropna(subset=['Fecha'])
            df = df.sort_values('Fecha').reset_index(drop=True)

        except Exception as e:
            messagebox.showerror("Error al leer CSV", str(e))
            return None, -999999, -999999

    # Verificar columnas
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        return None, -999999, -999999

    df = df[EXPECTED_COLUMNS].copy()

    # Procesar fechas
    df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True, errors='coerce')
    df = df.dropna(subset=['Fecha'])
    df = df.sort_values("Fecha").reset_index(drop=True)

    # Guardar fechas antes de convertir a string
    fecha_inicial = df['Fecha'].min()
    fecha_final = df['Fecha'].max()

    df['Fecha'] = df['Fecha'].dt.strftime("%d/%m/%Y")

    # Convertir numéricos
    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(to_float_safe)

    df['% var.'] = df['% var.'].apply(parse_percent_to_decimal)

    for col in ['Último', 'Apertura', 'Máximo', 'Mínimo', 'Vol.']:
        df[col] = df[col].apply(lambda x: float(x) if pd.notna(x) else 0.0)

    df['% var.'] = df['% var.'].apply(lambda x: float(x) if pd.notna(x) else 0.0)

    # Calcular % acumulado por signos consecutivos
    acum = 0
    prev = 0
    lst = []
    for v in df['% var.']:
        sign = 1 if v > 0 else -1 if v < 0 else 0
        if sign == prev:
            acum += v
        else:
            acum = v
        lst.append(acum)
        prev = sign
    df['% acumulado'] = lst

    acum_decimal = df['% acumulado'].astype(float)

    valores_seleccionados = []
    seq = []
    indices_seq = []

    for idx, v in enumerate(acum_decimal):
        if v > 0:
            seq.append(v)
            indices_seq.append(idx)
        else:
            if len(seq) >= 2:
                valores_seleccionados.append(seq[-1] * 100.0)
            seq = []
            indices_seq = []
    if len(seq) >= 2:
        valores_seleccionados.append(seq[-1] * 100.0)

    promedio_maximos = sum(valores_seleccionados) / len(valores_seleccionados) if valores_seleccionados else 0.0

    valores_minimos = []
    seq_neg = []

    for idx, v in enumerate(acum_decimal):
        if v < 0:
            seq_neg.append(v)
        else:
            if len(seq_neg) >= 2:
                valores_minimos.append(seq_neg[-1] * 100.0)
            seq_neg = []
    if len(seq_neg) >= 2:
        valores_minimos.append(seq_neg[-1] * 100.0)

    promedio_minimos = sum(valores_minimos) / len(valores_minimos) if valores_minimos else 0.0

    def determinar_opcion(v, a):
        if v >= local_venta:
            return "Venta"
        if v <= umbral_compra_decimal:
            return "Compra"
        if a >= local_venta and v >= local_suave:
            return "Venta"
        if a <= umbral_compra_decimal and v <= -local_suave:
            return "Compra"
        return "N/A"

    df['Opción'] = df.apply(lambda r: determinar_opcion(r['% var.'], r['% acumulado']), axis=1)

    try:
        if LIMITE_TIPO == "acciones":
            MAX_ACCIONES = int(LIMITE_VALOR)
            MAX_APORTE = float("inf")
        else:
            MAX_ACCIONES = 10
            MAX_APORTE = float(LIMITE_VALOR)
    except:
        MAX_ACCIONES = 10
        MAX_APORTE = float("inf")

    acciones = 0
    capital_bolsa = 0
    aporte_acumulado = 0

    movs, acts, cap_b, cap_acc, cap_tot, aport, aport_acum, precios_compra = [], [], [], [], [], [], [], []

    precios_en_cartera = []

    acum_pct = df['% acumulado'].astype(float) * 100.0
    comprar_multiple = [False] * len(df)

    seq_idxs_neg = []
    all_negative_sequences = []
    for idx, v in enumerate(acum_pct):
        if v < 0:
            seq_idxs_neg.append(idx)
        else:
            if len(seq_idxs_neg) >= 2:
                all_negative_sequences.append(seq_idxs_neg.copy())
            seq_idxs_neg = []
    if len(seq_idxs_neg) >= 2:
        all_negative_sequences.append(seq_idxs_neg.copy())

    if promedio_minimos < 0.0:
        for s in all_negative_sequences:
            for i in s:
                if acum_pct.iloc[i] <= promedio_minimos:
                    comprar_multiple[i] = True

    vender_doble = [False] * len(df)

    seq_idxs = []
    all_positive_sequences = []
    for idx, v in enumerate(acum_pct):
        if v > 0:
            seq_idxs.append(idx)
        else:
            if len(seq_idxs) >= 2:
                all_positive_sequences.append(seq_idxs.copy())
            seq_idxs = []
    if len(seq_idxs) >= 2:
        all_positive_sequences.append(seq_idxs.copy())

    if promedio_maximos > 0.0:
        for s in all_positive_sequences:
            for i in s:
                if acum_pct.iloc[i] >= promedio_maximos:
                    vender_doble[i] = True

    for idx, row in df.iterrows():
        opcion = row["Opción"]
        precio = row["Último"]
        movimiento = 0
        aporte = 0.0
        precio_operacion = 0.0

        if opcion == "Compra":
            n_compra = 1
            if COMPRA_MULTIPLE_ACCIONES is not None and comprar_multiple[idx]:
                n_compra = COMPRA_MULTIPLE_ACCIONES

            acciones_a_comprar = 0
            for _ in range(n_compra):
                puede_comprar = False
                if LIMITE_TIPO == "acciones" and acciones < MAX_ACCIONES:
                    puede_comprar = True
                elif LIMITE_TIPO == "aporte" and (aporte_acumulado + precio) <= MAX_APORTE:
                    puede_comprar = True

                if puede_comprar:
                    acciones_a_comprar += 1
                    if capital_bolsa >= precio:
                        capital_bolsa -= precio
                    else:
                        aporte += precio
                        aporte_acumulado += precio
                        capital_bolsa += precio
                        capital_bolsa -= precio
                    acciones += 1
                    precios_en_cartera.append(precio)
                    precios_en_cartera.sort()
                else:
                    break

            movimiento = acciones_a_comprar
            if movimiento > 0:
                precio_operacion = -precio

        elif opcion == "Venta" and acciones > 0:
            acciones_vendibles = 0
            for precio_compra in precios_en_cartera:
                ganancia_porcentual = (precio - precio_compra) / precio_compra
                if ganancia_porcentual >= ganancia_minima:
                    acciones_vendibles += 1
                else:
                    break

            if acciones_vendibles > 0:
                n_venta = 1
                if VENTA_MULTIPLE_ACCIONES is not None and vender_doble[idx] and acciones >= VENTA_MULTIPLE_ACCIONES:
                    n_venta = VENTA_MULTIPLE_ACCIONES

                n_venta = min(n_venta, acciones_vendibles, acciones)

                capital_bolsa += precio * n_venta
                acciones -= n_venta
                movimiento = -n_venta

                for _ in range(n_venta):
                    if precios_en_cartera:
                        precios_en_cartera.pop(0)

                if movimiento < 0:
                    precio_operacion = precio

        movs.append(movimiento)
        acts.append(acciones)
        cap_b.append(round(capital_bolsa, 2))
        cap_acc.append(round(acciones * precio, 2))
        cap_tot.append(round(capital_bolsa + acciones * precio, 2))
        aport.append(round(aporte, 2))
        aport_acum.append(round(aporte_acumulado, 2))
        precios_compra.append(precio_operacion)

    df["Movimiento de acciones"] = movs
    df["Acciones en cartera"] = acts
    df["Precio de compra"] = precios_compra
    df["Capital en bolsa"] = cap_b
    df["Capital en acciones"] = cap_acc
    df["Capital total"] = cap_tot
    df["Aporte"] = aport
    df["Aporte acumulado"] = aport_acum

    if text_compras_mult is not None:
        text_compras_mult.delete("1.0", tk.END)

        if COMPRA_MULTIPLE_ACCIONES is not None:
            fechas_compra_multiple = df[df['Movimiento de acciones'] == COMPRA_MULTIPLE_ACCIONES]['Fecha'].tolist()
            if fechas_compra_multiple:
                text_compras_mult.insert(tk.END, "\n".join(fechas_compra_multiple))
            else:
                text_compras_mult.insert(tk.END, f"No hay compras de {COMPRA_MULTIPLE_ACCIONES} acciones")

    if text_ventas_mult is not None:
        text_ventas_mult.delete("1.0", tk.END)

        if VENTA_MULTIPLE_ACCIONES is not None:
            fechas_venta_multiple = df[df['Movimiento de acciones'] == -VENTA_MULTIPLE_ACCIONES]['Fecha'].tolist()
            if fechas_venta_multiple:
                text_ventas_mult.insert(tk.END, "\n".join(fechas_venta_multiple))
            else:
                text_ventas_mult.insert(tk.END, f"No hay ventas de {VENTA_MULTIPLE_ACCIONES} acciones")

    df["Margen"] = df["Capital total"] - df["Aporte acumulado"]
    df["Rentabilidad"] = df.apply(
        lambda r: (r["Margen"] / r["Aporte acumulado"] * 100) if r["Aporte acumulado"] > 0 else 0, axis=1)

    rentab_max = df["Rentabilidad"].max()
    margen_prom = df["Margen"].mean()

    df["Rentabilidad"] = df["Rentabilidad"].round(2).astype(str) + "%"
    df["% var."] = (df["% var."] * 100).round(2).astype(str) + "%"
    df["% acumulado"] = (df["% acumulado"] * 100).round(2).astype(str) + "%"

    return df, rentab_max, margen_prom, fecha_inicial.strftime("%d/%m/%Y"), fecha_final.strftime("%d/%m/%Y")


# =========================
# Variables globales para progreso
# =========================
scipy_evaluaciones = 0
scipy_evaluaciones_max = 0
scipy_inicio_tiempo = None


# =========================
# Función para refinar el óptimo (encontrar centro del rango)
# =========================
def refinar_optimo(params_optimos, bounds, csv_filtrado=None, n_muestras=30, umbral_similitud=0.95):
    """
    Muestrea alrededor del óptimo encontrado para hallar el centro del rango
    que produce resultados similares.

    Args:
        params_optimos: Lista con los parámetros óptimos encontrados [compra, venta, ganancia, compra_mult, venta_mult]
        bounds: Límites de cada parámetro [(min, max), ...]
        csv_filtrado: DataFrame filtrado o None para usar el completo
        n_muestras: Número de puntos a muestrear alrededor del óptimo
        umbral_similitud: Porcentaje mínimo del resultado óptimo para considerar similar (0.95 = 95%)

    Returns:
        Lista con los parámetros promediados
    """
    global COMPRA_MULTIPLE_ACCIONES, VENTA_MULTIPLE_ACCIONES, analisis_detenido

    # Si el análisis fue detenido, retornar los parámetros originales
    if analisis_detenido:
        return params_optimos

    # Evaluar el resultado óptimo original
    compra_orig = params_optimos[0]
    COMPRA_MULTIPLE_ACCIONES = int(round(params_optimos[3])) if params_optimos[3] > 1.5 else None
    VENTA_MULTIPLE_ACCIONES = int(round(params_optimos[4])) if params_optimos[4] > 1.5 else None

    entry_compra.delete(0, tk.END)
    entry_compra.insert(0, f"{params_optimos[0]:.1f}")
    entry_venta.delete(0, tk.END)
    entry_venta.insert(0, f"{params_optimos[1]:.1f}")
    entry_ganancia_minima.delete(0, tk.END)
    entry_ganancia_minima.insert(0, f"{params_optimos[2]:.1f}")

    df_orig, rent_orig, margen_orig, _, _ = ejecutar_analisis_con_umbral(compra_orig / 100, csv_filtrado)

    if df_orig is None:
        return params_optimos  # Si falla, retornar los originales

    # Determinar métrica a usar
    usar_margen = (OBJETIVO_ACTUAL == "margen_prom")
    metrica_optima = margen_orig if usar_margen else rent_orig
    umbral_metrica = metrica_optima * umbral_similitud

    # Generar muestras alrededor del óptimo (±10% de cada parámetro)
    np.random.seed(42)  # Semilla fija para reproducibilidad

    params_similares = [list(params_optimos)]  # Incluir el óptimo original

    for _ in range(n_muestras):
        # Verificar si el análisis fue detenido
        if analisis_detenido:
            break

        params_muestra = []
        for i, (p, (b_min, b_max)) in enumerate(zip(params_optimos, bounds)):
            # Calcular rango de variación (±10% del valor o ±10% del rango total)
            rango = max(abs(p) * 0.1, (b_max - b_min) * 0.05)

            # Para parámetros enteros (compra_mult, venta_mult)
            if i >= 3:
                nuevo_val = p + np.random.uniform(-1, 1)
                nuevo_val = max(b_min, min(b_max, nuevo_val))
            else:
                nuevo_val = p + np.random.uniform(-rango, rango)
                nuevo_val = max(b_min, min(b_max, nuevo_val))

            params_muestra.append(nuevo_val)

        # Evaluar esta muestra
        COMPRA_MULTIPLE_ACCIONES = int(round(params_muestra[3])) if params_muestra[3] > 1.5 else None
        VENTA_MULTIPLE_ACCIONES = int(round(params_muestra[4])) if params_muestra[4] > 1.5 else None

        entry_compra.delete(0, tk.END)
        entry_compra.insert(0, f"{params_muestra[0]:.1f}")
        entry_venta.delete(0, tk.END)
        entry_venta.insert(0, f"{params_muestra[1]:.1f}")
        entry_ganancia_minima.delete(0, tk.END)
        entry_ganancia_minima.insert(0, f"{params_muestra[2]:.1f}")

        try:
            df_test, rent_test, margen_test, _, _ = ejecutar_analisis_con_umbral(params_muestra[0] / 100, csv_filtrado)

            if df_test is not None:
                metrica_test = margen_test if usar_margen else rent_test

                # Si el resultado es similar al óptimo, guardar estos parámetros
                if metrica_test >= umbral_metrica:
                    params_similares.append(params_muestra)
        except:
            continue

    # Calcular promedio de todos los parámetros similares
    if len(params_similares) > 1:
        params_promedio = []
        for i in range(5):
            valores = [p[i] for p in params_similares]
            promedio = sum(valores) / len(valores)

            # Redondear parámetros enteros
            if i >= 3:
                promedio = round(promedio)

            params_promedio.append(promedio)

        print(f"  → Refinamiento: {len(params_similares)} configuraciones similares encontradas")
        print(f"  → Parámetros promediados: Compra={params_promedio[0]:.2f}%, Venta={params_promedio[1]:.2f}%")

        return params_promedio
    else:
        print(f"  → Refinamiento: Solo el óptimo original cumple el umbral")
        return params_optimos


# =========================
# Función objetivo para optimización con SciPy
# =========================
def funcion_objetivo_scipy(params, csv_filtrado=None):
    global COMPRA_MULTIPLE_ACCIONES, VENTA_MULTIPLE_ACCIONES
    global scipy_evaluaciones, scipy_evaluaciones_max, scipy_inicio_tiempo
    global analisis_detenido

    # Verificar si el usuario detuvo el análisis - retornar valor alto para terminar rápido
    if analisis_detenido:
        return float('inf')

    scipy_evaluaciones += 1

    compra_pct = params[0]
    venta_pct = params[1]
    ganancia_min = params[2]
    compra_mult = int(round(params[3])) if params[3] > 1.5 else None
    venta_mult = int(round(params[4])) if params[4] > 1.5 else None

    entry_compra.delete(0, tk.END)
    entry_compra.insert(0, f"{compra_pct:.1f}")

    entry_venta.delete(0, tk.END)
    entry_venta.insert(0, f"{venta_pct:.1f}")

    entry_ganancia_minima.delete(0, tk.END)
    entry_ganancia_minima.insert(0, f"{ganancia_min:.1f}")

    if compra_mult is None:
        entry_compra_multiple.delete(0, tk.END)
    else:
        entry_compra_multiple.delete(0, tk.END)
        entry_compra_multiple.insert(0, str(compra_mult))

    if venta_mult is None:
        entry_venta_multiple.delete(0, tk.END)
    else:
        entry_venta_multiple.delete(0, tk.END)
        entry_venta_multiple.insert(0, str(venta_mult))

    COMPRA_MULTIPLE_ACCIONES = compra_mult
    VENTA_MULTIPLE_ACCIONES = venta_mult

    if scipy_evaluaciones % 5 == 0:
        porcentaje = (scipy_evaluaciones / scipy_evaluaciones_max) * 100
        ventana.progress_bar['value'] = porcentaje

        tiempo_transcurrido = time.time() - scipy_inicio_tiempo
        if scipy_evaluaciones > 10:
            tiempo_por_eval = tiempo_transcurrido / scipy_evaluaciones
            evals_restantes = scipy_evaluaciones_max - scipy_evaluaciones
            tiempo_restante = tiempo_por_eval * evals_restantes * 0.95

            mins_restantes = int(tiempo_restante // 60)
            segs_restantes = int(tiempo_restante % 60)

            ventana.label_progreso.config(
                text=f"Progreso: {scipy_evaluaciones}/{scipy_evaluaciones_max} ({porcentaje:.1f}%) - "
                     f"Tiempo estimado restante: {mins_restantes}m {segs_restantes}s"
            )

        ventana.update()
        time.sleep(0.001)

    try:
        df, rent_tmp, margen_tmp, _, _ = ejecutar_analisis_con_umbral(compra_pct / 100, csv_filtrado)

        if df is None:
            return 999999

        usar_margen = (OBJETIVO_ACTUAL == "margen_prom")
        if usar_margen:
            metrica = margen_tmp
        else:
            metrica = rent_tmp

        return -metrica
    except:
        return 999999


# =========================
# Función para optimizar un período específico
# =========================
def optimizar_periodo(nombre_periodo, dias=None):
    """Ejecuta optimización para un período específico"""
    global scipy_evaluaciones, scipy_evaluaciones_max, scipy_inicio_tiempo
    global COMPRA_MULTIPLE_ACCIONES, VENTA_MULTIPLE_ACCIONES

    print(f"\n{'=' * 60}")
    print(f"Optimizando período: {nombre_periodo}")
    print(f"{'=' * 60}")

    # Filtrar datos si es necesario
    if dias is not None:
        csv_filtrado = filtrar_ultimos_dias(INPUT_FILE, dias)
    else:
        csv_filtrado = None
        print(f"  → Analizando datos completos")

    # Determinar si hay optimización activa
    usar_scipy = (usar_scipy_var.get() == 1)
    hay_optimizacion = (auto_compra_var.get() == 1 or auto_venta_var.get() == 1 or
                        auto_ganancia_var.get() == 1 or auto_compra_mult_var.get() == 1 or
                        auto_venta_mult_var.get() == 1)

    mejor_df = None
    mejor_compra = None
    mejor_venta = None
    mejor_ganancia = None
    mejor_compra_mult = None
    mejor_venta_mult = None
    fecha_inicial = None
    fecha_final = None

    # ===============================================================
    # OPTIMIZACIÓN CON SCIPY
    # ===============================================================
    if usar_scipy and hay_optimizacion:
        bounds = []

        if auto_compra_var.get() == 1:
            bounds.append((-3.0, 0.0))
        else:
            try:
                val = float(entry_compra.get().replace(",", "."))
                bounds.append((val, val))
            except:
                bounds.append((-1.6, -1.6))

        if auto_venta_var.get() == 1:
            bounds.append((0.0, 3.0))
        else:
            try:
                val = float(entry_venta.get().replace(",", "."))
                bounds.append((val, val))
            except:
                bounds.append((1.6, 1.6))

        if auto_ganancia_var.get() == 1:
            bounds.append((1.5, 5.0))
        else:
            try:
                val = float(entry_ganancia_minima.get().replace(",", "."))
                bounds.append((val, val))
            except:
                bounds.append((0.0, 0.0))

        if auto_compra_mult_var.get() == 1:
            bounds.append((0, 5))
        else:
            val_cm = entry_compra_multiple.get().strip()
            if val_cm == "":
                bounds.append((0, 0))
            else:
                try:
                    val = int(val_cm)
                    bounds.append((val, val))
                except:
                    bounds.append((0, 0))

        if auto_venta_mult_var.get() == 1:
            bounds.append((0, 5))
        else:
            val_vm = entry_venta_multiple.get().strip()
            if val_vm == "":
                bounds.append((0, 0))
            else:
                try:
                    val = int(val_vm)
                    bounds.append((val, val))
                except:
                    bounds.append((0, 0))

        ventana.progress_bar.grid(row=0, column=0, columnspan=2, sticky="we", pady=2)
        ventana.label_progreso.grid(row=1, column=0, columnspan=2, sticky="w")

        # Calcular progreso base (porcentaje de combinaciones completadas)
        if progreso_total_combinaciones > 0:
            progreso_base = ((progreso_combinacion_actual - 1) / progreso_total_combinaciones) * 100
            progreso_slice = 100 / progreso_total_combinaciones  # Porcentaje que representa esta combinación
        else:
            progreso_base = 0
            progreso_slice = 100

        ventana.progress_bar['value'] = progreso_base

        periodo_legible = nombre_periodo.replace("_", " ").title().replace("6 Meses", "6M").replace("3 Meses", "3M")
        obj_texto = "Rent" if OBJETIVO_ACTUAL == "rentabilidad" else "Marg"
        ventana.label_progreso.config(
            text=f"Optimizando {progreso_combinacion_actual}/{progreso_total_combinaciones}: {periodo_legible} - {obj_texto}..."
        )

        maxiter = 100
        popsize = 15
        scipy_evaluaciones_max = maxiter * popsize
        scipy_evaluaciones = 0
        scipy_inicio_tiempo = time.time()

        ventana.update()

        # Callback para actualizar progreso y permitir detención
        def callback_progreso(xk, convergence):
            global scipy_evaluaciones
            scipy_evaluaciones += 1

            # Calcular progreso combinado (global + local)
            if scipy_evaluaciones_max > 0:
                progreso_local = (scipy_evaluaciones / scipy_evaluaciones_max) * progreso_slice
            else:
                progreso_local = 0

            progreso_total = progreso_base + progreso_local
            ventana.progress_bar['value'] = min(progreso_total, 100)
            ventana.update()

            return analisis_detenido  # Retornar True detiene la optimización

        resultado = differential_evolution(
            lambda params: funcion_objetivo_scipy(params, csv_filtrado),
            bounds,
            strategy='best1bin',
            maxiter=maxiter,
            popsize=popsize,
            tol=0.01,
            mutation=(0.5, 1),
            recombination=0.7,
            seed=42,  # Semilla fija para resultados reproducibles
            callback=callback_progreso,
            disp=False,
            polish=False,  # Desactivar polish para permitir detención limpia
            init='latinhypercube',
            atol=0,
            updating='immediate',
            workers=1
        )

        ventana.progress_bar.grid_forget()
        ventana.label_progreso.grid_forget()

        # Verificar si el usuario detuvo el análisis
        if analisis_detenido:
            return None

        # Refinar el óptimo encontrado (encontrar centro del rango)
        print(f"\n  → Refinando parámetros óptimos...")
        ventana.label_progreso.config(text="Refinando parámetros óptimos...")
        ventana.label_progreso.grid(row=1, column=0, columnspan=2, sticky="w")
        ventana.update()

        params_refinados = refinar_optimo(
            params_optimos=list(resultado.x),
            bounds=bounds,
            csv_filtrado=csv_filtrado,
            n_muestras=30,
            umbral_similitud=0.95
        )

        ventana.label_progreso.grid_forget()

        mejor_compra = params_refinados[0]
        mejor_venta = params_refinados[1]
        mejor_ganancia = params_refinados[2]
        mejor_compra_mult = int(round(params_refinados[3])) if params_refinados[3] > 1.5 else None
        mejor_venta_mult = int(round(params_refinados[4])) if params_refinados[4] > 1.5 else None

        entry_compra.delete(0, tk.END)
        entry_compra.insert(0, f"{mejor_compra:.1f}")

        entry_venta.delete(0, tk.END)
        entry_venta.insert(0, f"{mejor_venta:.1f}")

        entry_ganancia_minima.delete(0, tk.END)
        entry_ganancia_minima.insert(0, f"{mejor_ganancia:.1f}")

        if mejor_compra_mult is None:
            entry_compra_multiple.delete(0, tk.END)
            COMPRA_MULTIPLE_ACCIONES = None
        else:
            entry_compra_multiple.delete(0, tk.END)
            entry_compra_multiple.insert(0, str(mejor_compra_mult))
            COMPRA_MULTIPLE_ACCIONES = mejor_compra_mult

        if mejor_venta_mult is None:
            entry_venta_multiple.delete(0, tk.END)
            VENTA_MULTIPLE_ACCIONES = None
        else:
            entry_venta_multiple.delete(0, tk.END)
            entry_venta_multiple.insert(0, str(mejor_venta_mult))
            VENTA_MULTIPLE_ACCIONES = mejor_venta_mult

        mejor_df, _, _, fecha_inicial, fecha_final = ejecutar_analisis_con_umbral(mejor_compra / 100, csv_filtrado)

    # ===============================================================
    # SIN SCIPY (bucles anidados o ejecución directa)
    # ===============================================================
    else:
        # Aquí iría el código de optimización sin SciPy (bucles anidados)
        # Por brevedad, ejecuto directamente con los valores actuales
        try:
            compra_val = float(entry_compra.get().replace(",", ".")) / 100
        except:
            compra_val = -0.016

        mejor_df, _, _, fecha_inicial, fecha_final = ejecutar_analisis_con_umbral(compra_val, csv_filtrado)
        mejor_compra = compra_val * 100
        mejor_venta = float(entry_venta.get().replace(",", "."))
        mejor_ganancia = float(entry_ganancia_minima.get().replace(",", "."))
        mejor_compra_mult = COMPRA_MULTIPLE_ACCIONES
        mejor_venta_mult = VENTA_MULTIPLE_ACCIONES

    if mejor_df is None:
        return None

    # Calcular estadísticas completas del análisis
    def float_col(col_name):
        return mejor_df[col_name].astype(str).str.rstrip('%').str.replace(',', '.').astype(float)

    # Calcular % acumulado y promedios de máximos/mínimos
    acumulado_float = float_col('% acumulado') * 100.0

    # Promedio de máximos (secuencias positivas)
    valores_seleccionados = []
    secuencia = []
    for v in acumulado_float:
        if v > 0:
            secuencia.append(v)
        else:
            if len(secuencia) >= 2:
                valores_seleccionados.append(secuencia[-1])
            secuencia = []
    if len(secuencia) >= 2:
        valores_seleccionados.append(secuencia[-1])
    promedio_maximos = sum(valores_seleccionados) / len(valores_seleccionados) if valores_seleccionados else 0.0

    # Promedio de mínimos (secuencias negativas)
    valores_minimos = []
    secuencia_neg = []
    for v in acumulado_float:
        if v < 0:
            secuencia_neg.append(v)
        else:
            if len(secuencia_neg) >= 2:
                valores_minimos.append(secuencia_neg[-1])
            secuencia_neg = []
    if len(secuencia_neg) >= 2:
        valores_minimos.append(secuencia_neg[-1])
    promedio_minimos = sum(valores_minimos) / len(valores_minimos) if valores_minimos else 0.0

    # Estadísticas de % variación
    max_var = float_col('% var.').max()
    min_var = float_col('% var.').min()
    fecha_max_var = mejor_df.loc[float_col('% var.').idxmax(), 'Fecha']
    fecha_min_var = mejor_df.loc[float_col('% var.').idxmin(), 'Fecha']
    dif_var = max_var - min_var

    prom_var = float_col('% var.')
    subidas = prom_var[prom_var > 0]
    max_prom = subidas.mean() if not subidas.empty else 0
    bajadas = prom_var[prom_var < 0]
    min_prom = bajadas.mean() if not bajadas.empty else 0
    dif_prom = max_prom - min_prom

    # Estadísticas de operaciones
    opc_compra = int((mejor_df["Opción"] == "Compra").sum())
    acciones_compradas = int(mejor_df.loc[mejor_df["Movimiento de acciones"] > 0, "Movimiento de acciones"].sum())
    opc_venta = int((mejor_df["Opción"] == "Venta").sum())
    acciones_vendidas = int(-mejor_df.loc[mejor_df["Movimiento de acciones"] < 0, "Movimiento de acciones"].sum())
    max_acc_cartera = int(mejor_df["Acciones en cartera"].max())

    # Estadísticas financieras
    max_aporte = float(mejor_df["Aporte acumulado"].max())
    max_margen = float(round(mejor_df["Margen"].max(), 2))
    margen_promedio = float(round(mejor_df["Margen"].mean(), 2))
    max_rentab = float(float_col("Rentabilidad").max())
    rentab_promedio = float(float_col("Rentabilidad").mean())
    fecha_max_rentab = mejor_df.loc[float_col("Rentabilidad").idxmax(), "Fecha"]

    # Preparar resultado con todas las estadísticas
    resultado = {
        "df": mejor_df,
        "compra_pct": mejor_compra,
        "venta_pct": mejor_venta,
        "ganancia_min": mejor_ganancia,
        "suave_pct": float(entry_suave.get().replace(",", ".")),
        "limite_tipo": tipo_limite_var.get(),
        "limite_valor": float(entry_limite.get().replace(",", ".")),
        "compra_mult": mejor_compra_mult,
        "venta_mult": mejor_venta_mult,
        "rentabilidad_max": max_rentab,
        "margen_promedio": margen_promedio,
        "fecha_inicial": fecha_inicial,
        "fecha_final": fecha_final,
        # Nuevos campos de estadísticas
        "promedio_maximos": promedio_maximos,
        "promedio_minimos": promedio_minimos,
        "max_var": max_var,
        "min_var": min_var,
        "fecha_max_var": fecha_max_var,
        "fecha_min_var": fecha_min_var,
        "dif_var": dif_var,
        "max_prom_var": max_prom,
        "min_prom_var": min_prom,
        "dif_prom_var": dif_prom,
        "opc_compra": opc_compra,
        "acciones_compradas": acciones_compradas,
        "opc_venta": opc_venta,
        "acciones_vendidas": acciones_vendidas,
        "max_acc_cartera": max_acc_cartera,
        "max_aporte": max_aporte,
        "max_margen": max_margen,
        "rentab_promedio": rentab_promedio,
        "fecha_max_rentab": fecha_max_rentab
    }

    return resultado


# =========================
# Función iniciar_proceso (principal)
# =========================
def iniciar_proceso():
    global ultimo_df, ultima_ruta_excel, ultimo_folder, ultimo_base_name
    global INPUT_FILE, FOLDER, LIMITE_TIPO, LIMITE_VALOR
    global VENTA_MULTIPLE_ACCIONES, COMPRA_MULTIPLE_ACCIONES
    global resultados_analisis_actuales, resultados_dfs_por_periodo
    global analisis_detenido

    # Resetear variable de detención
    analisis_detenido = False

    # Configurar botones (deshabilitar Iniciar, habilitar Detener)
    btn_iniciar_analisis.config(state="disabled")
    btn_detener_analisis.config(state="normal")
    ventana.update()

    # Limpiar mensaje anterior de optimización
    ventana.label_resultado_opt.config(text="")
    ventana.label_resultado_opt.grid_forget()

    # Limpiar resultados previos
    resultados_analisis_actuales = {}
    resultados_dfs_por_periodo = {}

    INPUT_FILE = entry_ruta.get().strip().strip('"')
    if not os.path.exists(INPUT_FILE):
        messagebox.showerror("Error", f"La ruta del CSV no existe:\n{INPUT_FILE}")
        return

    FOLDER = os.path.dirname(INPUT_FILE)
    base_name = os.path.splitext(os.path.basename(INPUT_FILE))[0]

    try:
        venta_val = float(entry_venta.get().replace(",", ".")) / 100
        suave_val = float(entry_suave.get().replace(",", ".")) / 100
    except:
        messagebox.showerror("Error", "Valores numéricos inválidos.")
        return

    LIMITE_TIPO = tipo_limite_var.get()
    try:
        LIMITE_VALOR = float(entry_limite.get().replace(",", "."))
    except:
        LIMITE_VALOR = 10.0

    # Validar compra múltiple
    valor_compra_multiple = entry_compra_multiple.get().strip()
    if valor_compra_multiple == "":
        COMPRA_MULTIPLE_ACCIONES = None
    else:
        try:
            n_acciones = int(valor_compra_multiple)
            if n_acciones < 2:
                messagebox.showerror("Error", "La cantidad para 'Compra de N acciones' debe ser 2 o más.")
                return
            COMPRA_MULTIPLE_ACCIONES = n_acciones
        except ValueError:
            messagebox.showerror("Error", "Debes ingresar un número entero válido en 'Compra de N acciones'.")
            return

    # Validar venta múltiple
    valor_venta_multiple = entry_venta_multiple.get().strip()
    if valor_venta_multiple == "":
        VENTA_MULTIPLE_ACCIONES = None
    else:
        try:
            n_acciones = int(valor_venta_multiple)
            if n_acciones < 2:
                messagebox.showerror("Error", "La cantidad para 'Venta de N acciones' debe ser 2 o más.")
                return
            VENTA_MULTIPLE_ACCIONES = n_acciones
        except ValueError:
            messagebox.showerror("Error", "Debes ingresar un número entero válido en 'Venta de N acciones'.")
            return

    # Determinar qué períodos analizar
    periodos_a_analizar = []

    if analizar_completo_var.get() == 1:
        periodos_a_analizar.append(("completo", None))

    if analizar_6meses_var.get() == 1:
        periodos_a_analizar.append(("6_meses", 180))

    if analizar_3meses_var.get() == 1:
        periodos_a_analizar.append(("3_meses", 90))

    if not periodos_a_analizar:
        messagebox.showerror("Error", "Selecciona al menos un período para analizar")
        return

    # Determinar qué objetivos analizar
    objetivos_a_analizar = obtener_objetivos_seleccionados()

    if not objetivos_a_analizar:
        messagebox.showerror("Error", "Selecciona al menos un objetivo de optimización")
        return

    # Variable global para que optimizar_periodo sepa qué objetivo usar
    global OBJETIVO_ACTUAL
    global progreso_combinacion_actual, progreso_total_combinaciones
    global progreso_tiempo_inicio_total, progreso_tiempos_combinaciones

    # =====================================================
    # PROGRESO INTELIGENTE: Preparación
    # =====================================================

    # Contar filas del CSV
    try:
        df_temp = pd.read_csv(INPUT_FILE)
        num_filas = len(df_temp)
        del df_temp
    except:
        num_filas = 200  # Valor por defecto

    # Obtener configuración de checks activos
    checks_activos = {
        'scipy': usar_scipy_var.get() == 1,
        'compra': auto_compra_var.get() == 1,
        'venta': auto_venta_var.get() == 1,
        'ganancia': auto_ganancia_var.get() == 1,
        'compra_mult': auto_compra_mult_var.get() == 1,
        'venta_mult': auto_venta_mult_var.get() == 1
    }

    clave_config = obtener_clave_configuracion(num_filas, checks_activos)
    print(f"[DEBUG] Clave configuración: {clave_config} ({num_filas} filas)")

    # Analizar cada combinación de período y objetivo
    resultados_por_periodo = {}
    total_combinaciones = len(periodos_a_analizar) * len(objetivos_a_analizar)
    combinacion_actual = 0

    # Variables de progreso global
    progreso_combinacion_actual = 0
    progreso_total_combinaciones = total_combinaciones
    progreso_tiempo_inicio_total = time.time()
    progreso_tiempos_combinaciones = []

    # Estimar tiempo total si hay historial
    tiempo_estimado_total, hay_historial = estimar_tiempo_total(clave_config, total_combinaciones)

    if hay_historial:
        print(f"[INFO] Tiempo estimado total: {formatear_tiempo(tiempo_estimado_total)}")

    for objetivo in objetivos_a_analizar:
        OBJETIVO_ACTUAL = objetivo
        objetivo_texto = "Rentabilidad" if objetivo == "rentabilidad" else "Margen Prom"

        for nombre_periodo, dias in periodos_a_analizar:
            combinacion_actual += 1
            progreso_combinacion_actual = combinacion_actual

            # Verificar si el usuario detuvo el análisis
            if analisis_detenido:
                print(f"[DEBUG] Análisis detenido antes de procesar {nombre_periodo}/{objetivo}")
                break

            # Calcular tiempo restante estimado
            tiempo_transcurrido = time.time() - progreso_tiempo_inicio_total
            if hay_historial and tiempo_estimado_total:
                tiempo_restante = max(0, tiempo_estimado_total - tiempo_transcurrido)
                texto_tiempo = f" | Restante: ~{formatear_tiempo(tiempo_restante)}"
            elif len(progreso_tiempos_combinaciones) > 0:
                # Estimar basado en combinaciones ya completadas
                promedio_actual = sum(progreso_tiempos_combinaciones) / len(progreso_tiempos_combinaciones)
                combinaciones_restantes = total_combinaciones - combinacion_actual + 1
                tiempo_restante = promedio_actual * combinaciones_restantes
                texto_tiempo = f" | Restante: ~{formatear_tiempo(tiempo_restante)}"
            else:
                texto_tiempo = ""

            periodo_legible = nombre_periodo.replace("_", " ").title().replace("6 Meses", "6M").replace("3 Meses", "3M")
            obj_corto = "Rent" if objetivo == "rentabilidad" else "Marg"

            print(f"[INFO] Analizando {combinacion_actual}/{total_combinaciones}: {periodo_legible} - {obj_corto}{texto_tiempo}")

            # Mostrar progreso en la interfaz
            ventana.label_progreso.config(
                text=f"Analizando {combinacion_actual}/{total_combinaciones}: {periodo_legible} - {objetivo_texto}{texto_tiempo}"
            )
            ventana.label_progreso.grid(row=1, column=0, columnspan=2, sticky="w")

            # Actualizar barra de progreso global (porcentaje de combinaciones)
            progreso_global = ((combinacion_actual - 1) / total_combinaciones) * 100
            ventana.progress_bar['value'] = progreso_global
            ventana.progress_bar.grid(row=0, column=0, columnspan=2, sticky="we", pady=2)
            ventana.update()

            # Iniciar tiempo de esta combinación
            tiempo_inicio_combinacion = time.time()

            resultado = optimizar_periodo(nombre_periodo, dias)

            # Registrar tiempo de esta combinación
            tiempo_combinacion = time.time() - tiempo_inicio_combinacion
            progreso_tiempos_combinaciones.append(tiempo_combinacion)

            if resultado is None:
                if analisis_detenido:
                    # El usuario detuvo el análisis, salir del bucle
                    break
                else:
                    messagebox.showerror("Error", f"No se pudo optimizar: {nombre_periodo}/{objetivo_texto}")
                    continue

            # Agregar el objetivo al resultado
            resultado["objetivo"] = objetivo

            # Guardar con clave que incluye período y objetivo
            clave_resultado = f"{nombre_periodo}_{objetivo}"
            resultados_por_periodo[clave_resultado] = resultado
            resultados_dfs_por_periodo[clave_resultado] = resultado["df"]

        if analisis_detenido:
            break

    # Guardar tiempos en historial (promedio de esta sesión)
    if progreso_tiempos_combinaciones and not analisis_detenido:
        tiempo_promedio = sum(progreso_tiempos_combinaciones) / len(progreso_tiempos_combinaciones)
        registrar_tiempo_combinacion(clave_config, tiempo_promedio)
        print(f"[INFO] Tiempo promedio por combinación: {formatear_tiempo(tiempo_promedio)}")

    # Ocultar barra de progreso y actualizar interfaz
    ventana.progress_bar.grid_forget()
    ventana.label_progreso.grid_forget()
    ventana.update()

    if not resultados_por_periodo:
        # Restaurar botones de análisis
        btn_iniciar_analisis.config(state="normal")
        btn_detener_analisis.config(state="disabled")

        if analisis_detenido:
            ventana.label_resultado_opt.config(text=f"⚠ Análisis detenido por el usuario")
            ventana.label_resultado_opt.config(fg="orange")
            ventana.label_resultado_opt.grid(row=14, column=0, columnspan=3, sticky="w", padx=10, pady=5)
        else:
            messagebox.showerror("Error", "No se obtuvieron resultados válidos")
        return

    # Guardar para JSON (con múltiples objetivos)
    resultados_analisis_actuales = {
        "ticker": base_name,
        "objetivos_analizados": objetivos_a_analizar,
        "periodos": resultados_por_periodo
    }

    # Guardar variables globales
    ultimo_folder = FOLDER
    ultimo_base_name = base_name

    # Mostrar resultados en interfaz
    mostrar_resultados_multiples_periodos(resultados_por_periodo)

    # Habilitar botones
    btn_guardar_json.config(state="normal")
    btn_generar_db_excel.config(state="normal", bg="#1E90FF", fg="black")  # REACTIVAR botón DB/Excel

    # Restaurar botones de análisis
    btn_iniciar_analisis.config(state="normal")
    btn_detener_analisis.config(state="disabled")

    # Mostrar mensaje de completado
    if analisis_detenido:
        ventana.label_resultado_opt.config(text=f"⚠ Análisis detenido por el usuario")
        ventana.label_resultado_opt.config(fg="orange")
    else:
        ventana.label_resultado_opt.config(text=f"✓ Análisis completado para {len(periodos_a_analizar)} período(s)")
        ventana.label_resultado_opt.config(fg="darkgreen")
    ventana.label_resultado_opt.grid(row=14, column=0, columnspan=3, sticky="w", padx=10, pady=5)


# =========================
# Función para mostrar estadísticas en la interfaz
# =========================
def mostrar_resultados_multiples_periodos(resultados):
    """Muestra los resultados de todos los períodos en pestañas"""
    global historial_analisis_por_ticker, ticker_actual

    # Inicializar historial para este ticker si no existe
    if ticker_actual not in historial_analisis_por_ticker:
        historial_analisis_por_ticker[ticker_actual] = []

    # Agregar resultados actuales al historial DEL TICKER ACTUAL
    for clave_periodo, datos in resultados.items():
        # Obtener objetivo de cada resultado individual
        objetivo_actual = datos.get("objetivo", "rentabilidad")
        objetivo_texto = "Rentabilidad" if objetivo_actual == "rentabilidad" else "Margen Prom"

        # Extraer solo el nombre del período (sin el objetivo)
        if "_rentabilidad" in clave_periodo:
            nombre_periodo = clave_periodo.replace("_rentabilidad", "")
        elif "_margen_prom" in clave_periodo:
            nombre_periodo = clave_periodo.replace("_margen_prom", "")
        else:
            nombre_periodo = clave_periodo

        # Convertir a formato legible
        periodo_legible = nombre_periodo.replace("_", " ").title()
        # Corregir "Seis Meses" a "6 Meses" y "Tres Meses" a "3 Meses"
        periodo_legible = periodo_legible.replace("Seis Meses", "6 Meses").replace("Tres Meses", "3 Meses")

        historial_analisis_por_ticker[ticker_actual].append({
            "periodo": periodo_legible,
            "objetivo": objetivo_texto,
            "compra_pct": datos['compra_pct'],
            "venta_pct": datos['venta_pct'],
            "ganancia_min": datos['ganancia_min'],
            "suave_pct": datos['suave_pct'],
            "compra_mult": datos['compra_mult'],
            "venta_mult": datos['venta_mult'],
            "rentabilidad_max": datos['rentabilidad_max'],
            "margen_promedio": datos['margen_promedio']
        })

    # Limpiar frame de estadísticas
    for widget in ventana.frame_stats.winfo_children():
        widget.destroy()

    # Crear notebook (pestañas) para cada período
    notebook = ttk.Notebook(ventana.frame_stats)
    notebook.pack(fill="both", expand=True, pady=(0, 10))

    for clave_periodo, datos in resultados.items():
        # Crear frame para este período
        frame_periodo = tk.Frame(notebook)

        # Crear nombre de pestaña legible (ej: "Completo - Rent" o "6 Meses - Margen")
        objetivo_actual = datos.get("objetivo", "rentabilidad")
        obj_corto = "Rent" if objetivo_actual == "rentabilidad" else "Margen"

        if "_rentabilidad" in clave_periodo:
            periodo_base = clave_periodo.replace("_rentabilidad", "")
        elif "_margen_prom" in clave_periodo:
            periodo_base = clave_periodo.replace("_margen_prom", "")
        else:
            periodo_base = clave_periodo

        periodo_texto = periodo_base.replace("_", " ").title()
        periodo_texto = periodo_texto.replace("Seis Meses", "6 Meses").replace("Tres Meses", "3 Meses")
        nombre_pestana = f"{periodo_texto} - {obj_corto}"

        notebook.add(frame_periodo, text=nombre_pestana)

        # Mostrar estadísticas
        mostrar_estadisticas_en_frame(frame_periodo, datos["df"], datos)

    # NUEVO: Frame INFERIOR con tabla consolidada SOLO del ticker actual
    frame_consolidado = tk.Frame(ventana.frame_stats, relief="ridge", borderwidth=2, bg="lightyellow", padx=10, pady=10)
    frame_consolidado.pack(fill="x", pady=(10, 0))

    # Extraer ticker real (siglas) para mostrar
    if ticker_actual:
        partes = ticker_actual.split('_')
        ticker_display = ticker_actual
        if len(partes) >= 2:
            for parte in partes:
                if parte.isupper() and 1 <= len(parte) <= 5:
                    ticker_display = parte
                    break
    else:
        ticker_display = "Actual"

    tk.Label(frame_consolidado, text=f"📊 HISTORIAL DE ANÁLISIS - {ticker_display} (Acumulativo)",
             font=("Arial", 11, "bold"), bg="lightyellow", fg="darkgreen").pack(anchor="w")

    # Crear tabla con parámetros
    frame_tabla_params = tk.Frame(frame_consolidado, bg="lightyellow")
    frame_tabla_params.pack(fill="x", pady=(5, 0))

    # Headers
    headers = ["#", "Período", "Objetivo", "Compra %", "Venta %", "Gan Mín %", "Suave %", "Comp", "Venta", "Rentab Máx",
               "Margen Prom"]
    for col, header in enumerate(headers):
        ancho = 5 if col == 0 else 11
        tk.Label(frame_tabla_params, text=header, font=("Arial", 8, "bold"),
                 bg="lightblue", relief="solid", borderwidth=1, width=ancho).grid(row=0, column=col, sticky="ew",
                                                                                  padx=1, pady=1)

    # Datos SOLO del ticker actual
    analisis_ticker_actual = historial_analisis_por_ticker.get(ticker_actual, [])

    if not analisis_ticker_actual:
        tk.Label(frame_consolidado, text="No hay análisis en el historial",
                 font=("Arial", 9), bg="lightyellow", fg="gray").pack(pady=10)
        return

    # Ordenar por período y luego por objetivo
    orden_periodos = {"Completo": 1, "6 Meses": 2, "3 Meses": 3}
    analisis_ticker_actual_ordenado = sorted(
        analisis_ticker_actual,
        key=lambda x: (orden_periodos.get(x['periodo'], 99), x['objetivo'])
    )

    # Colores por objetivo (base del objetivo, sin el número)
    colores_objetivo = {
        "rentabilidad": "#e8f5e9",  # Verde claro
        "margen": "#e3f2fd",         # Azul claro
    }
    color_default = "#fff3e0"  # Naranja claro para otros

    periodo_anterior = None
    fila_actual = 0

    for idx, analisis in enumerate(analisis_ticker_actual_ordenado, start=1):
        fila_actual += 1

        # Determinar color basado en el objetivo (sin número)
        objetivo_base = analisis['objetivo'].lower().split()[0]
        bg_color = colores_objetivo.get(objetivo_base, color_default)

        # Si cambia el período, agregar línea separadora
        if periodo_anterior is not None and analisis['periodo'] != periodo_anterior:
            # Agregar fila separadora
            for col in range(11):
                ancho = 5 if col == 0 else 11
                tk.Label(frame_tabla_params, text="", font=("Arial", 2),
                         bg="#999999", relief="flat", width=ancho, height=1).grid(
                    row=fila_actual, column=col, sticky="ew", padx=1, pady=0)
            fila_actual += 1

        periodo_anterior = analisis['periodo']

        valores = [
            str(idx),
            analisis['periodo'],
            analisis['objetivo'],
            f"{analisis['compra_pct']:.1f}",
            f"{analisis['venta_pct']:.1f}",
            f"{analisis['ganancia_min']:.1f}",
            f"{analisis['suave_pct']:.1f}",
            str(analisis['compra_mult']) if analisis['compra_mult'] else "-",
            str(analisis['venta_mult']) if analisis['venta_mult'] else "-",
            f"{analisis['rentabilidad_max']:.2f}%",
            f"{analisis['margen_promedio']:.2f}"
        ]

        for col, valor in enumerate(valores):
            ancho = 5 if col == 0 else 11
            tk.Label(frame_tabla_params, text=valor, font=("Arial", 7),
                     bg=bg_color, relief="solid", borderwidth=1, width=ancho).grid(
                row=fila_actual, column=col, sticky="ew", padx=1, pady=1)


def mostrar_estadisticas_en_frame(frame_parent, df, datos_periodo):
    """Muestra estadísticas de un período específico en un frame"""

    frame1 = tk.Frame(frame_parent, padx=15)
    frame1.grid(row=0, column=0, sticky="nw")
    frame2 = tk.Frame(frame_parent, padx=15)
    frame2.grid(row=0, column=1, sticky="nw")
    frame3 = tk.Frame(frame_parent, padx=15)
    frame3.grid(row=0, column=2, sticky="nw")
    frame4 = tk.Frame(frame_parent, padx=15)
    frame4.grid(row=0, column=3, sticky="nw")

    def float_col(col):
        return df[col].astype(str).str.rstrip('%').replace('', '0').astype(float)

    acumulado_float = df['% acumulado'].str.rstrip('%').astype(float)

    valores_seleccionados = []
    secuencia = []
    for v in acumulado_float:
        if v > 0:
            secuencia.append(v)
        else:
            if len(secuencia) >= 2:
                valores_seleccionados.append(secuencia[-1])
            secuencia = []
    if len(secuencia) >= 2:
        valores_seleccionados.append(secuencia[-1])

    promedio_maximos = sum(valores_seleccionados) / len(valores_seleccionados) if valores_seleccionados else 0.0

    valores_minimos = []
    secuencia_neg = []
    for v in acumulado_float:
        if v < 0:
            secuencia_neg.append(v)
        else:
            if len(secuencia_neg) >= 2:
                valores_minimos.append(secuencia_neg[-1])
            secuencia_neg = []
    if len(secuencia_neg) >= 2:
        valores_minimos.append(secuencia_neg[-1])

    promedio_minimos = sum(valores_minimos) / len(valores_minimos) if valores_minimos else 0.0

    max_var = float_col('% var.').max()
    min_var = float_col('% var.').min()
    fecha_max_var = df.loc[float_col('% var.').idxmax(), 'Fecha']
    fecha_min_var = df.loc[float_col('% var.').idxmin(), 'Fecha']
    dif_var = max_var - min_var

    prom_var = float_col('% var.')
    subidas = prom_var[prom_var > 0]
    max_prom = subidas.mean() if not subidas.empty else 0
    bajadas = prom_var[prom_var < 0]
    min_prom = bajadas.mean() if not bajadas.empty else 0
    dif_prom = max_prom - min_prom

    opc_compra = (df["Opción"] == "Compra").sum()
    acciones_compradas = df.loc[df["Movimiento de acciones"] > 0, "Movimiento de acciones"].sum()
    opc_venta = (df["Opción"] == "Venta").sum()
    acciones_vendidas = -df.loc[df["Movimiento de acciones"] < 0, "Movimiento de acciones"].sum()
    max_acc_cartera = df["Acciones en cartera"].max()
    max_aporte = df["Aporte acumulado"].max()
    max_margen = round(df["Margen"].max(), 2)
    margen_promedio = round(df["Margen"].mean(), 2)
    max_rentab = float_col("Rentabilidad").max()
    rentab_promedio = float_col("Rentabilidad").mean()
    fecha_max_rentab = df.loc[float_col("Rentabilidad").idxmax(), "Fecha"]

    tk.Label(frame1, fg="blue", text=f"Max % var : {max_var:.2f}% ({fecha_max_var})", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame1, fg="blue", text=f"Min % var : {min_var:.2f}% ({fecha_min_var})", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame1, fg="blue", text=f"Diferencia : {dif_var:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame1, fg="blue", text="", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame1, fg="blue", text=f"Prom de % var. acum máximos +: {promedio_maximos:.2f}%",
             font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame1, fg="blue", text=f"Prom de % var. acum mínimos -: {promedio_minimos:.2f}%",
             font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame2, fg="red", text=f"Prom % var + :  {max_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame2, fg="red", text=f"Prom % var - : {min_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame2, fg="red", text=f"Diferencia       :  {dif_prom:.2f}%", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame3, fg="black", text=f"Opciones Compra       : {opc_compra}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Acciones Compradas : {int(acciones_compradas)}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame3, fg="black", text=f"Opciones Venta           : {opc_venta}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame3, fg="black", text=f"Acciones Vendidas      : {int(acciones_vendidas)}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame3, fg="black", text=f"Máx acción en cartera : {max_acc_cartera}", font=("Arial", 12)).pack(anchor="w")

    tk.Label(frame4, fg="purple", text=f"Aporte acum max  : {max_aporte:,.0f}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Margen max       : {max_margen:,.2f}", font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Margen promedio  : {margen_promedio:,.2f}", font=("Arial", 12)).pack(
        anchor="w")
    tk.Label(frame4, fg="purple", text=f"Rentab. max      : {max_rentab:.2f}% ({fecha_max_rentab})",
             font=("Arial", 12)).pack(anchor="w")
    tk.Label(frame4, fg="purple", text=f"Rentab. promedio : {rentab_promedio:.2f}%", font=("Arial", 12)).pack(
        anchor="w")


# -------------------------
# Manejo de cierre
# -------------------------
def on_closing():
    ventana.quit()
    ventana.destroy()


ventana.protocol("WM_DELETE_WINDOW", on_closing)

try:


    ventana.mainloop()
except KeyboardInterrupt:
    pass
